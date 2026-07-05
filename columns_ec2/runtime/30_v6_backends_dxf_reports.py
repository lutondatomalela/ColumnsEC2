# -*- coding: utf-8 -*-
# Auto-split from ColumnsEC2 v0.9 RC8.
# This module is executed in the shared runtime namespace by columns_ec2.runtime.loader.
# Keep execution order defined in columns_ec2/runtime/manifest.py.

APP_VERSION = "v6.7"


@dataclass
class DesignDecisionV56:
    status: str
    severity: str
    blocking: List[str]
    warnings: List[str]
    info: List[str]

    @property
    def reason(self) -> str:
        if self.blocking:
            return "; ".join(self.blocking)
        return ""

    @property
    def notes(self) -> str:
        return "; ".join(self.warnings + self.info)


def _v56_is_structuralcodes_backend_value(value) -> bool:
    try:
        return _is_structuralcodes_backend_v52(value)
    except Exception:
        return "structuralcodes" in str(value or "").lower()


def _v56_layout_signature(layout) -> Tuple:
    """Assinatura completa para cache de capacidade. Corrige o erro de layouts mistos
    com mesma área/envolvente mas diâmetros de face distintos.
    """
    if isinstance(layout, MixedLayout):
        return (
            "mixed",
            round(layout.b_mm, 1), round(layout.h_mm, 1), round(layout.cover_mm, 1),
            round(layout.phi_corner_mm, 3), round(layout.phi_face_mm, 3),
            int(layout.n_face_y_extra), int(layout.n_face_z_extra), round(layout.phi_st_mm, 3),
        )
    return (
        "uniform",
        round(layout.b_mm, 1), round(layout.h_mm, 1), round(layout.cover_mm, 1),
        round(getattr(layout, "phi_long_mm", 0.0), 3),
        int(getattr(layout, "n_bars_y", 0)), int(getattr(layout, "n_bars_z", 0)),
        round(getattr(layout, "phi_st_mm", 0.0), 3),
    )


def _v56_layout_description(layout) -> str:
    if isinstance(layout, MixedLayout):
        n_face = max(0, int(layout.n_total) - 4)
        if n_face <= 0:
            return f"4Ø{int(layout.phi_corner_mm)} (cantos)"
        face_parts = []
        if int(layout.n_face_y_extra) > 0:
            face_parts.append(f"{2*int(layout.n_face_y_extra)}Ø{int(layout.phi_face_mm)} nas faces b")
        if int(layout.n_face_z_extra) > 0:
            face_parts.append(f"{2*int(layout.n_face_z_extra)}Ø{int(layout.phi_face_mm)} nas faces h")
        return f"4Ø{int(layout.phi_corner_mm)} (cantos) + " + " + ".join(face_parts)
    n = int(getattr(layout, "n_total", 0))
    phi = int(getattr(layout, "phi_long_mm", 0))
    if n == 4:
        return f"4Ø{phi} (cantos)"
    return f"{n}Ø{phi} distribuídos no perímetro"


def _v56_bar_points(layout) -> List[Tuple[float, float, float]]:
    """Pontos reais dos varões: (y,z,phi), incluindo layouts mistos."""
    return _layout_bar_points_v45(layout) if "_layout_bar_points_v45" in globals() else []


def _v56_clear_spacing_from_layout(layout, agg_mm: float = 20.0) -> Tuple[bool, float, float]:
    req = max(20.0, float(getattr(layout, "phi_long_mm", 0.0)), agg_mm + 5.0)
    pts = _v56_bar_points(layout)
    if not pts:
        return False, 0.0, req
    # Verifica espaçamentos adjacentes nas quatro faces. É mais adequado para layouts perimetrais.
    tol = 1e-6
    ys = sorted(set(round(p[0], 6) for p in pts))
    zs = sorted(set(round(p[1], 6) for p in pts))
    y_min, y_max = min(ys), max(ys)
    z_min, z_max = min(zs), max(zs)
    min_clear = 1e9

    def _check_line(items, along="y"):
        nonlocal min_clear
        if len(items) < 2:
            return
        items = sorted(items, key=lambda x: x[0] if along == "y" else x[1])
        for a, b in zip(items[:-1], items[1:]):
            dist = abs((b[0] - a[0]) if along == "y" else (b[1] - a[1]))
            clear = dist - 0.5 * (a[2] + b[2])
            min_clear = min(min_clear, clear)

    top = [p for p in pts if abs(round(p[1], 6) - z_max) < tol]
    bot = [p for p in pts if abs(round(p[1], 6) - z_min) < tol]
    left = [p for p in pts if abs(round(p[0], 6) - y_min) < tol]
    right = [p for p in pts if abs(round(p[0], 6) - y_max) < tol]
    _check_line(top, "y")
    _check_line(bot, "y")
    _check_line(left, "z")
    _check_line(right, "z")
    if min_clear == 1e9:
        min_clear = 999.0
    return min_clear >= req, min_clear, req


def _v56_constructive_layouts(self, b_mm, h_mm, is_circular=False) -> List:
    """Gera layouts construtivos perimetrais explícitos.
    Para secções rectangulares, privilegia 4 varões nos cantos e varões adicionais nas faces.
    """
    if is_circular:
        return list(_old_build_candidate_layouts_v45_base(self, b_mm, h_mm, is_circular=True)) if "_old_build_candidate_layouts_v45_base" in globals() else list(_old_build_candidate_layouts_v45_base(self, b_mm, h_mm, True))
    max_y, max_z = self.max_bars_per_face(b_mm, h_mm, is_circular=False)
    max_y = max(2, int(max_y))
    max_z = max(2, int(max_z))
    corner_diams = [10.0, 12.0, 16.0, 20.0, 25.0, 32.0]
    face_diams = [10.0, 12.0, 16.0, 20.0, 25.0]
    layouts = []
    seen = set()

    def _add(ly):
        if not ly.clear_spacing_ok():
            return
        ok_spacing, _clear, _req = _v56_clear_spacing_from_layout(ly)
        if not ok_spacing:
            return
        sig = _v56_layout_signature(ly)
        if sig in seen:
            return
        seen.add(sig)
        layouts.append(ly)

    # 4 varões nos cantos.
    for pc in corner_diams:
        _add(MixedLayout(pc, pc, 0, 0, b_mm, h_mm, self.cover_mm, self.choose_stirrup(pc)))

    # Extras nas faces. ey/ez são varões adicionais por face horizontal/vertical.
    for pc in corner_diams:
        for pf in face_diams:
            if pf > pc:
                continue
            phi_st = self.choose_stirrup(pc)
            for ey in range(0, max_y - 1):  # até max_y-2 extras
                for ez in range(0, max_z - 1):
                    if ey == 0 and ez == 0:
                        continue
                    ly = MixedLayout(pc, pf, ey, ez, b_mm, h_mm, self.cover_mm, phi_st)
                    if ly.n_bars_y > max_y or ly.n_bars_z > max_z:
                        continue
                    # Evitar layouts sem sentido: muitos varões pequenos com canto pequeno.
                    if ly.n_total > 20:
                        continue
                    _add(ly)

    # Manter também alguns layouts uniformes existentes como alternativa, mas no fim da ordem prática.
    try:
        base = list(_old_build_candidate_layouts_v45_base(self, b_mm, h_mm, is_circular=False))
        for ly in base:
            if int(getattr(ly, "n_total", 0)) <= 20:
                _add(ly)
    except Exception:
        pass

    layouts.sort(key=lambda ly: _v56_layout_score(ly, 0.0, b_mm, h_mm))
    return layouts


def _v56_layout_score(layout, as_target: float, b_mm: float, h_mm: float) -> Tuple:
    asprov = float(getattr(layout, "as_prov_mm2", 0.0))
    n = int(getattr(layout, "n_total", 999))
    phi_long = float(getattr(layout, "phi_long_mm", 999.0))
    excess = max(0.0, asprov - float(as_target or 0.0))
    deficit = max(0.0, float(as_target or 0.0) - asprov)
    # Penalizações construtivas: muitos varões, diâmetros mistos, distribuição desequilibrada.
    mixed_penalty = 3 if isinstance(layout, MixedLayout) and int(layout.n_total) > 4 and layout.phi_corner_mm != layout.phi_face_mm else 0
    imbalance = abs(int(getattr(layout, "n_bars_y", 2)) - int(getattr(layout, "n_bars_z", 2)))
    # Incentivo para mais varões na face mais longa.
    geom_penalty = 0
    if b_mm > h_mm and getattr(layout, "n_bars_y", 2) < getattr(layout, "n_bars_z", 2):
        geom_penalty = 5
    if h_mm > b_mm and getattr(layout, "n_bars_z", 2) < getattr(layout, "n_bars_y", 2):
        geom_penalty = 5
    return (deficit > 0.0, excess / 100.0, mixed_penalty, geom_penalty, n, phi_long, imbalance)


def _capacity_for_layout_v56(self, layout, n_ed_kN: float, fcd: float, fyd: float, Es: float):
    """Capacidade com chave de cache completa; fundamental para layouts mistos."""
    mode = getattr(self, "requested_calc_mode", self.calc_mode)
    sig = _v56_layout_signature(layout)
    key = ("v56", mode, sig, round(float(n_ed_kN), 1), round(float(fcd), 4), round(float(fyd), 4), round(float(Es), 1))
    if key in self._capacity_cache:
        return self._capacity_cache[key]
    # Usa malha um pouco mais fina em dimensionamento; mantém desempenho aceitável.
    if str(mode).lower() in ["rigoroso", "dimensionamento"]:
        n_ang = 37
        n_c = 110
    else:
        n_ang = 25
        n_c = 70
    angles = [i * math.pi / (2 * max(n_ang - 1, 1)) for i in range(n_ang)]
    capacities = []
    c_max = 3.0 * max(layout.b_mm, layout.h_mm)
    for ang in angles:
        best = None
        # bissecção se possível; fallback em grelha.
        def n_at(c):
            return self.section_response(layout, n_ed_kN, ang, c, fcd, fyd, Es)
        lo, hi = 2.0, c_max
        n_lo, _, _ = n_at(lo)
        n_hi, _, _ = n_at(hi)
        if (n_lo - n_ed_kN) * (n_hi - n_ed_kN) <= 0:
            for _ in range(42):
                mid = 0.5 * (lo + hi)
                N, My, Mz = n_at(mid)
                diff = N - n_ed_kN
                if best is None or abs(diff) < best[0]:
                    best = (abs(diff), My, Mz)
                if abs(diff) < 1e-3:
                    break
                if (n_lo - n_ed_kN) * diff <= 0:
                    hi = mid
                    n_hi = N
                else:
                    lo = mid
                    n_lo = N
        else:
            for i in range(n_c):
                c_mm = 2.0 + i * (c_max - 2.0) / max(n_c - 1, 1)
                N, My, Mz = n_at(c_mm)
                diff = abs(N - n_ed_kN)
                if best is None or diff < best[0]:
                    best = (diff, My, Mz)
        if best is not None:
            capacities.append((best[1], best[2]))
    self._capacity_cache[key] = capacities
    return capacities


ColumnDesigner.capacity_for_layout = _capacity_for_layout_v56
ColumnDesigner.build_candidate_layouts = _v56_constructive_layouts


def _v56_axial_capacity_kN(layout, ac_mm2: float, fcd: float, fyd: float) -> float:
    # Capacidade axial simplificada para filtro; a verificação final continua pela superfície.
    asprov = float(getattr(layout, "as_prov_mm2", 0.0))
    return max(0.0, (0.85 * ac_mm2 * fcd + asprov * fyd) / 1000.0)


def _v56_detailing_check(layout, b_mm: float, h_mm: float, cover_mm: float, as_req_mm2: float, as_max_mm2: float) -> Dict:
    blocking = []
    warnings = []
    info = []
    phi_max = float(getattr(layout, "phi_long_mm", 0.0))
    phi_st = float(getattr(layout, "phi_st_mm", 0.0))
    n_total = int(getattr(layout, "n_total", 0))
    asprov = float(getattr(layout, "as_prov_mm2", 0.0))
    if n_total < 4:
        blocking.append("número mínimo de 4 varões longitudinais não cumprido")
    if phi_max < 10.0:
        blocking.append("diâmetro longitudinal inferior a Ø10")
    if phi_st < max(6.0, phi_max / 4.0):
        blocking.append("diâmetro dos estribos inferior a max(6 mm; Øl,max/4)")
    if asprov < as_req_mm2 - 1e-6:
        blocking.append("As,prov inferior a As,req")
    if asprov > as_max_mm2 + 1e-6:
        blocking.append("As,prov superior a As,max")
    ok_spacing, min_clear, req_clear = _v56_clear_spacing_from_layout(layout)
    if not ok_spacing:
        blocking.append(f"espaçamento livre insuficiente entre varões ({min_clear:.0f} < {req_clear:.0f} mm)")
    smax_ties = min(15.0 * phi_max if phi_max else 999.0, min(b_mm, h_mm), 300.0)
    # O espaçamento adoptado é escolhido depois; aqui fica o limite normativo.
    if isinstance(layout, MixedLayout) and layout.phi_corner_mm != layout.phi_face_mm:
        info.append("layout com dois diâmetros; confirmar legenda e montagem em obra")
    if int(getattr(layout, "n_bars_y", 2)) > 2 or int(getattr(layout, "n_bars_z", 2)) > 2:
        warnings.append("prever cintas/grampos intermédios para travar todos os varões comprimidos")
    if as_req_mm2 > 0 and asprov / as_req_mm2 > 1.60:
        info.append("As,prov significativamente superior a As,req; avaliar optimização")
    status = "Não conforme" if blocking else ("Verificar" if warnings else "OK")
    return {
        "detailing_status": status,
        "detailing_blocking_issues": "; ".join(blocking) if blocking else "-",
        "detailing_warnings": "; ".join(warnings) if warnings else "-",
        "detailing_info": "; ".join(info) if info else "-",
        "detailing_issues": "; ".join(blocking + warnings + info) if (blocking or warnings or info) else "-",
        "detailing_smax_ties_mm": smax_ties,
        "detailing_min_clear_mm": min_clear,
        "detailing_required_clear_mm": req_clear,
    }


def _v56_torsion_check(t_ed_kNm, b_mm, h_mm, cover_mm, fck, fcd, fyd, my_ed=0.0, mz_ed=0.0) -> Dict:
    tor = torsion_check_ec2_v4(t_ed_kNm, b_mm, h_mm, cover_mm, fck, fcd, fyd)
    trd = _finite(tor.get("TRdmax_kNm"), 0.0)
    ted = abs(_finite(t_ed_kNm, 0.0))
    ratio_trd = ted / trd if trd > 1e-9 else 0.0
    ratio_bend = ted / max(abs(_finite(my_ed)), abs(_finite(mz_ed)), 1.0)
    if ted <= 0.10:
        status = "Torção desprezável — TEd≈0"
        asw_s = 0.0
        asl = 0.0
    elif trd > 0 and ratio_trd <= 0.05:
        status = "Torção desprezável — não condicionante"
        asw_s = 0.0
        asl = 0.0
    elif "Não conforme" in str(tor.get("torsion_status", "")):
        status = "Não conforme: TEd > TRd,max"
        asw_s = tor.get("Asw_s_t_req_mm2_per_m")
        asl = tor.get("Asl_t_req_mm2")
    elif ratio_bend <= 0.05:
        status = "Torção pequena face à flexão — confirmar se pode ser desprezada"
        asw_s = 0.0
        asl = 0.0
    else:
        status = "Requer verificação/dimensionamento à torção"
        asw_s = tor.get("Asw_s_t_req_mm2_per_m")
        asl = tor.get("Asl_t_req_mm2")
    tor.update({
        "torsion_status": status,
        "torsion_utilization_TRdmax": ratio_trd,
        "torsion_ratio": ratio_bend,
        "Asw_s_t_req_mm2_per_m": asw_s,
        "Asl_t_req_mm2": asl,
    })
    return tor


def _v56_biaxial_capacity_check(self, layout, n_ed_kN, my_ed_kNm, mz_ed_kNm, ac_mm2, fcd, fyd, Es) -> Tuple[bool, Optional[float], Optional[float], Optional[float], str]:
    nrdmax = _v56_axial_capacity_kN(layout, ac_mm2, fcd, fyd)
    if abs(n_ed_kN) > nrdmax + 1e-6:
        return False, None, None, None, f"NEd superior à capacidade axial simplificada ({abs(n_ed_kN):.1f} > {nrdmax:.1f} kN)"
    capacities = self.capacity_for_layout(layout, abs(n_ed_kN), fcd, fyd, Es)
    ok, util, my_cap, mz_cap = self.biaxial_ok(my_ed_kNm, mz_ed_kNm, capacities)
    if ok:
        return True, util, my_cap, mz_cap, ""
    return False, util, my_cap, mz_cap, "interação N-My-Mz não verificada"


_old_design_one_v55_for_v56 = ColumnDesigner.design_one


def _design_one_v56(self, row: pd.Series, prebuilt_candidates=None):
    # Backends structuralcodes permanecem no fluxo strict existente.
    backend = getattr(self, "code_backend", globals().get("ACTIVE_CODE_BACKEND_V48", BACKEND_EC2_PT_2010 if "BACKEND_EC2_PT_2010" in globals() else ""))
    if _v56_is_structuralcodes_backend_value(backend):
        return _old_design_one_v55_for_v56(self, row, prebuilt_candidates=prebuilt_candidates)

    material = str(row.get("material", "") or DEFAULT_CONCRETE_CLASS)
    if material.strip().lower() in ["", "nan", "none"]:
        material = DEFAULT_CONCRETE_CLASS
    fck = parse_concrete_strength(material)
    cp = concrete_props(fck, gamma_c=self.gamma_c)
    sp = steel_props(self.fyk, gamma_s=self.gamma_s)
    fyd = sp["fyd"]
    Es = sp["Es"]

    b_mm = cm_to_mm(row.get("hy", 0.0))
    h_mm = cm_to_mm(row.get("hz", 0.0))
    ac_mm2 = safe_float(row.get("ax", float("nan"))) * 100.0
    iy_mm4 = safe_float(row.get("iy", float("nan"))) * 10000.0
    iz_mm4 = safe_float(row.get("iz", float("nan"))) * 10000.0
    if b_mm <= 0 and math.isfinite(ac_mm2) and ac_mm2 > 0:
        b_mm = math.sqrt(ac_mm2)
    if h_mm <= 0:
        h_mm = b_mm
    if not math.isfinite(ac_mm2) or ac_mm2 <= 0:
        ac_mm2 = b_mm * h_mm
    if not math.isfinite(iy_mm4) or iy_mm4 <= 0:
        iy_mm4 = b_mm * h_mm ** 3 / 12.0
    if not math.isfinite(iz_mm4) or iz_mm4 <= 0:
        iz_mm4 = h_mm * b_mm ** 3 / 12.0

    case_value = row.get("case", "")
    combination_number = extract_combination_number(case_value)
    limit_state = classify_limit_state_from_case(case_value)
    n_nodes = int(safe_float(row.get("n_nodes_found", 0), 0))
    n_ed_kN = max(abs(safe_float(row.get("fx_i", 0.0), 0.0)), abs(safe_float(row.get("fx_j", 0.0), 0.0)))
    vy_ed_kN = max(abs(safe_float(row.get("fy_i", row.get("fy", 0.0)), 0.0)), abs(safe_float(row.get("fy_j", row.get("fy", 0.0)), 0.0)))
    vz_ed_kN = max(abs(safe_float(row.get("fz_i", row.get("fz", 0.0)), 0.0)), abs(safe_float(row.get("fz_j", row.get("fz", 0.0)), 0.0)))
    mx_ed_kNm = max(abs(safe_float(row.get("mx_i", row.get("mx", 0.0)), 0.0)), abs(safe_float(row.get("mx_j", row.get("mx", 0.0)), 0.0)))
    my1_kNm = safe_float(row.get("my_i", 0.0), 0.0)
    my2_kNm = safe_float(row.get("my_j", 0.0), 0.0)
    mz1_kNm = safe_float(row.get("mz_i", 0.0), 0.0)
    mz2_kNm = safe_float(row.get("mz_j", 0.0), 0.0)

    base = {
        "member": row.get("member", ""), "case": case_value, "combination_number": combination_number, "limit_state": limit_state,
        "name": row.get("name", ""), "node_i": row.get("node_i", ""), "node_j": row.get("node_j", ""),
        "member_case_i": row.get("member_case_i", ""), "member_case_j": row.get("member_case_j", ""),
        "material": material, "fck_MPa": fck, "b_cm": b_mm / 10.0 if b_mm else None, "h_cm": h_mm / 10.0 if h_mm else None,
        "length_m": safe_float(row.get("length", 0.0), 0.0), "n_nodes_found": n_nodes,
        "n_ed_kN": n_ed_kN, "vy_ed_kN": vy_ed_kN, "vz_ed_kN": vz_ed_kN, "mx_ed_kNm": mx_ed_kNm,
        "my_i_kNm": my1_kNm, "my_j_kNm": my2_kNm, "mz_i_kNm": mz1_kNm, "mz_j_kNm": mz2_kNm,
        "code_backend": BACKEND_EC2_PT_2010 if "BACKEND_EC2_PT_2010" in globals() else "NP EN 1992-1-1:2010 PT",
        "cover_mm": self.cover_mm,
        "engine_v56": "default PT revisto: catálogo construtivo + N-My-Mz discreto",
    }
    if n_nodes < 2:
        reason = "Falha de dados: member/case sem os dois nós necessários"
        out = {**base, "status": "Falha", "failure_reason": reason, "failure_type": classify_failure_reason(reason), "failure_severity": "Bloqueante", "design_decision": "Corrigir dados de entrada", "review_priority": "Alta", "failure_action": "Confirmar os dois nós por member/case.", "shortlist_text": ""}
        return out

    ly_mm = m_to_mm(row.get("length", 0.0)) * self.l0y_factor
    lz_mm = m_to_mm(row.get("length", 0.0)) * self.l0z_factor
    as_min_mm2 = self.min_longitudinal_as(n_ed_kN, ac_mm2, fyd)
    as_max_mm2 = self.max_longitudinal_as(ac_mm2)
    as_seed_mm2 = max(as_min_mm2, 0.002 * ac_mm2)

    # Imperfeições práticas.
    L_m = max(safe_float(row.get("length", 0.0), 0.0), 1e-6)
    theta0 = 1.0 / 200.0
    alpha_h = max(2.0 / 3.0, min(1.0, 2.0 / math.sqrt(max(L_m, 1e-9))))
    theta_i = theta0 * alpha_h
    m_imp_y_kNm = n_ed_kN * (theta_i * ly_mm / 2.0) / 1000.0
    m_imp_z_kNm = n_ed_kN * (theta_i * lz_mm / 2.0) / 1000.0

    my_ed_raw, lambda_y, invr_y, kphi_y, _soy, my0eq_kNm, my2nd_kNm = self.second_order_nominal_curvature(n_ed_kN, my1_kNm, my2_kNm, ly_mm, ac_mm2, iy_mm4, h_mm, as_seed_mm2, cp["fcd"], fck, fyd, Es)
    mz_ed_raw, lambda_z, invr_z, kphi_z, _soz, mz0eq_kNm, mz2nd_kNm = self.second_order_nominal_curvature(n_ed_kN, mz1_kNm, mz2_kNm, lz_mm, ac_mm2, iz_mm4, b_mm, as_seed_mm2, cp["fcd"], fck, fyd, Es)
    n_red = max((n_ed_kN * 1e3) / max(ac_mm2 * cp["fcd"], 1e-9), 1e-6)
    omega_seed = (as_seed_mm2 * fyd) / max(ac_mm2 * cp["fcd"], 1e-9)
    lambda_lim_y, rm_y, A_lam, B_lam, C_y = lambda_lim_ec2_practical(my1_kNm, my2_kNm, n_red, omega_seed, self.phi_eff)
    lambda_lim_z, rm_z, _A, _B, C_z = lambda_lim_ec2_practical(mz1_kNm, mz2_kNm, n_red, omega_seed, self.phi_eff)
    needs_2y = lambda_y > lambda_lim_y
    needs_2z = lambda_z > lambda_lim_z
    if not needs_2y:
        my2nd_kNm = 0.0
    if not needs_2z:
        mz2nd_kNm = 0.0
    my0eq_kNm = max(my0eq_kNm, m_imp_y_kNm)
    mz0eq_kNm = max(mz0eq_kNm, m_imp_z_kNm)
    my_ed_kNm = my0eq_kNm + my2nd_kNm
    mz_ed_kNm = mz0eq_kNm + mz2nd_kNm
    as_req_mm2 = self.approx_required_as(n_ed_kN, my_ed_kNm, mz_ed_kNm, b_mm, h_mm, fyd, as_min_mm2)

    m01y, m02y, rmy_signed, curv_y = order_end_moments_ec2(my1_kNm, my2_kNm)
    m01z, m02z, rmz_signed, curv_z = order_end_moments_ec2(mz1_kNm, mz2_kNm)
    alpha_bi, nu_bi = biaxial_alpha_ec2_practical(n_ed_kN, ac_mm2, cp["fcd"], as_seed_mm2, fyd)
    self._biaxial_alpha = alpha_bi

    base.update({
        "my_ed_kNm": my_ed_kNm, "mz_ed_kNm": mz_ed_kNm,
        "lambda_y": lambda_y, "lambda_z": lambda_z, "lambda_lim_y": lambda_lim_y, "lambda_lim_z": lambda_lim_z,
        "lambda_check_y": "Considerar 2.ª ordem" if needs_2y else "Dispensa 2.ª ordem",
        "lambda_check_z": "Considerar 2.ª ordem" if needs_2z else "Dispensa 2.ª ordem",
        "m0e_y_kNm": my0eq_kNm, "m0e_z_kNm": mz0eq_kNm, "m2_y_kNm": my2nd_kNm, "m2_z_kNm": mz2nd_kNm,
        "m_imp_y_kNm": m_imp_y_kNm, "m_imp_z_kNm": m_imp_z_kNm,
        "second_order_y": "Sim" if needs_2y else "Dispensada", "second_order_z": "Sim" if needs_2z else "Dispensada",
        "as_min_mm2": as_min_mm2, "as_req_mm2": as_req_mm2, "as_max_mm2": as_max_mm2,
        "m01_y_ec2_kNm": m01y, "m02_y_ec2_kNm": m02y, "rm_y_signed": rmy_signed, "curvature_y": curv_y,
        "m01_z_ec2_kNm": m01z, "m02_z_ec2_kNm": m02z, "rm_z_signed": rmz_signed, "curvature_z": curv_z,
        "biaxial_method": "EC2 5.8.9 simplificado + superfície discreta N-My-Mz; catálogo construtivo v5.6",
        "biaxial_alpha": alpha_bi, "biaxial_n_ratio": nu_bi,
    })

    is_circular = self.infer_is_circular(row, b_mm, h_mm)
    candidates = list(prebuilt_candidates) if prebuilt_candidates is not None else self.build_candidate_layouts(b_mm, h_mm, is_circular=is_circular)
    candidates = [ly for ly in candidates if float(getattr(ly, "as_prov_mm2", 0.0)) >= as_req_mm2 - 1e-6 and float(getattr(ly, "as_prov_mm2", 0.0)) <= as_max_mm2 + 1e-6]
    candidates.sort(key=lambda ly: _v56_layout_score(ly, as_req_mm2, b_mm, h_mm))

    if not candidates:
        reason = "Falha por insuficiência/pormenorização: nenhum layout construtivo admissível cumpre As,req e As,max"
        out = {**base, "status": "Falha", "failure_reason": reason, "failure_type": "armadura_insuficiente", "failure_severity": "Bloqueante", "design_decision": "Sem solução no catálogo construtivo v5.6", "review_priority": "Alta", "failure_action": "Aumentar secção, rever esforços ou permitir solução especial de armadura.", "shortlist_text": ""}
        return out

    chosen = None
    best_ok = None
    best_fail = None
    shortlist_rows = []
    # v6.3: em Dimensionamento/Rigoroso avalia todo o catálogo admissível e
    # escolhe a solução com menor utilização N-My-Mz, em vez da primeira que verifica.
    # v6.7: limitar o número de layouts testados para evitar tempos excessivos.
    # A lista já está ordenada por viabilidade construtiva e proximidade a As,req.
    _strategy_for_speed = str(getattr(self, "design_strategy", globals().get("ACTIVE_REBAR_STRATEGY_V64", "equilibrada"))).lower()
    _mode_for_speed = str(self.calc_mode).lower()
    if _mode_for_speed == "pre_dimensionamento":
        max_tests = min(35, len(candidates))
    elif _strategy_for_speed.startswith("econ"):
        max_tests = min(80, len(candidates))
    elif _strategy_for_speed.startswith("rob") or _mode_for_speed == "rigoroso":
        max_tests = min(180, len(candidates))
    else:
        max_tests = min(120, len(candidates))
    for ly in candidates[:max_tests]:
        ok_det, min_clear, req_clear = _v56_clear_spacing_from_layout(ly)
        det = _v56_detailing_check(ly, b_mm, h_mm, self.cover_mm, as_req_mm2, as_max_mm2)
        ok_bi, util, my_cap, mz_cap, bi_reason = _v56_biaxial_capacity_check(self, ly, n_ed_kN, my_ed_kNm, mz_ed_kNm, ac_mm2, cp["fcd"], fyd, Es)
        desc = _v56_layout_description(ly)
        shortlist_rows.append({
            "solucao": desc,
            "as_prov_mm2": float(getattr(ly, "as_prov_mm2", 0.0)),
            "utilizacao": "" if util is None else f"{util:.3f}",
            "status_short": "OK" if (ok_bi and det["detailing_status"] != "Não conforme") else "Rejeitada",
            "failure_short": "" if ok_bi else bi_reason,
        })
        item = (999.0 if util is None else float(util), float(getattr(ly, "as_prov_mm2", 0.0)), ly, my_cap, mz_cap, det, bi_reason)
        if best_fail is None or item[0] < best_fail[0]:
            best_fail = item
        if str(self.calc_mode).lower() == "pre_dimensionamento":
            # Em pré-dimensionamento, basta a primeira solução construtiva com área suficiente.
            chosen = (ly, util if util is not None else (as_req_mm2 / max(float(getattr(ly, "as_prov_mm2", 1.0)), 1e-9)), my_cap, mz_cap, det, True, "pré-dimensionamento")
            break
        if ok_bi and det["detailing_status"] != "Não conforme":
            # v6.3: não parar na primeira solução OK. Guardar a solução com menor utilização;
            # em empate, preferir menor As e, depois, solução construtivamente mais simples.
            _eta = 999.0 if util is None else float(util)
            _as = float(getattr(ly, "as_prov_mm2", 0.0))
            _score = _v56_layout_score(ly, as_req_mm2, b_mm, h_mm)
            _strategy = str(getattr(self, "design_strategy", globals().get("ACTIVE_REBAR_STRATEGY_V64", "equilibrada"))).lower()
            if _strategy.startswith("econ"):
                _key = (_as, _eta, _score)
            elif _strategy.startswith("rob"):
                _key = (_eta, _as, _score)
            else:
                # Estratégia equilibrada: procurar uma utilização-alvo, evitando soluções no limite
                # e evitando também excesso de armadura sem necessidade.
                _target = float(globals().get("REBAR_TARGET_ETA_V64", 0.80))
                _eta_min = float(globals().get("REBAR_ETA_MIN_V68", 0.70))
                _eta_max = float(globals().get("REBAR_ETA_MAX_V68", 0.90))
                _as_excess_max = float(globals().get("REBAR_AS_EXCESS_MAX_V68", 0.40))
                _inside = 0 if _eta_min <= _eta <= _eta_max else 1
                _excess_ratio = max(0.0, _as / max(as_req_mm2, 1e-9) - 1.0)
                _excess_penalty = 0 if _excess_ratio <= _as_excess_max else 1
                _key = (_inside, _excess_penalty, abs(_eta - _target), _as, _score)
            ok_item = (_key, _eta, _as, _score, ly, my_cap, mz_cap, det)
            if best_ok is None or ok_item[0] < best_ok[0]:
                best_ok = ok_item

    if chosen is None and best_ok is not None:
        _key, _util, _as, _score, _ly, _my_cap, _mz_cap, _det = best_ok
        chosen = (_ly, _util, _my_cap, _mz_cap, _det, True, "")

    if chosen is None:
        util, _as, ly, my_cap, mz_cap, det, bi_reason = best_fail if best_fail is not None else (None, None, candidates[0], None, None, {}, "sem solução")
        chosen = (ly, util, my_cap, mz_cap, det, False, bi_reason or "interação N-My-Mz não verificada")

    ly, util, my_cap, mz_cap, det, bi_ok, bi_reason = chosen
    smax = self.tie_spacing_max(b_mm, h_mm, float(getattr(ly, "phi_long_mm", 10.0)))
    sprov = self.choose_spacing(smax)
    # sprov deve respeitar limite de pormenorização calculado pela geometria.
    if det.get("detailing_smax_ties_mm"):
        sprov = min(sprov, self.choose_spacing(_finite(det.get("detailing_smax_ties_mm"), smax)))

    sh_y = shear_check_ec2_v4(vy_ed_kN, n_ed_kN, h_mm, b_mm, 0.8 * b_mm, float(getattr(ly, "as_prov_mm2", 0.0)), fck, cp["fcd"], fyd, self.gamma_c)
    sh_z = shear_check_ec2_v4(vz_ed_kN, n_ed_kN, b_mm, h_mm, 0.8 * h_mm, float(getattr(ly, "as_prov_mm2", 0.0)), fck, cp["fcd"], fyd, self.gamma_c)
    tor = _v56_torsion_check(mx_ed_kNm, b_mm, h_mm, self.cover_mm, fck, cp["fcd"], fyd, my_ed_kNm, mz_ed_kNm)
    els = elastic_service_check_v4(n_ed_kN, my_ed_kNm, mz_ed_kNm, b_mm, h_mm, iy_mm4, iz_mm4, float(getattr(ly, "as_prov_mm2", 0.0)), fck, self.fyk, cp["Ecm"], cp["fctm"])
    if limit_state != "ELS":
        els["service_status"] = "Informativo — caso não identificado como ELS"

    blockers = []
    warnings = []
    info = []
    if not bi_ok and str(self.calc_mode).lower() != "pre_dimensionamento":
        blockers.append("resistência biaxial não verificada: " + str(bi_reason))
    if det.get("detailing_status") == "Não conforme":
        blockers.append("pormenorização EC2 não conforme: " + str(det.get("detailing_blocking_issues", "")))
    elif det.get("detailing_status") == "Verificar":
        warnings.append("pormenorização a confirmar: " + str(det.get("detailing_warnings", "")))
    if "Não conforme" in str(sh_y.get("status", "")) or "Não conforme" in str(sh_z.get("status", "")):
        blockers.append("esforço transverso: VEd > VRd,max")
    elif "Requer" in str(sh_y.get("status", "")) or "Requer" in str(sh_z.get("status", "")):
        warnings.append("dimensionar armadura transversal por esforço transverso")
    if "Não conforme" in str(tor.get("torsion_status", "")):
        blockers.append("torção: TEd > TRd,max")
    elif "Requer" in str(tor.get("torsion_status", "")):
        warnings.append("verificar/dimensionar armadura de torção")
    if det.get("detailing_info") and det.get("detailing_info") != "-":
        info.append(str(det.get("detailing_info")))
    if str(self.calc_mode).lower() == "pre_dimensionamento":
        info.append("pré-dimensionamento: verificar em modo Dimensionamento antes de adoptar")

    if blockers:
        status = "Falha"
        severity = "Bloqueante"
    elif warnings:
        status = "Aviso"
        severity = "Aviso"
    elif str(self.calc_mode).lower() == "pre_dimensionamento":
        status = "Pré-dimensionado"
        severity = "Informativo"
    else:
        status = "OK"
        severity = "OK"

    desc = _v56_layout_description(ly)
    solucao = f"{desc} + estribos Ø{int(float(getattr(ly, 'phi_st_mm', 8.0)))}//{sprov/10:.1f} cm"
    out = {
        **base,
        "phi_long_mm": float(getattr(ly, "phi_long_mm", 0.0)),
        "phi_corner_mm": float(getattr(ly, "phi_corner_mm", getattr(ly, "phi_long_mm", 0.0))),
        "phi_face_mm": float(getattr(ly, "phi_face_mm", getattr(ly, "phi_long_mm", 0.0))),
        "n_face_y_extra": int(getattr(ly, "n_face_y_extra", max(0, int(getattr(ly, "n_bars_y", 2)) - 2))),
        "n_face_z_extra": int(getattr(ly, "n_face_z_extra", max(0, int(getattr(ly, "n_bars_z", 2)) - 2))),
        "n_total": int(getattr(ly, "n_total", 0)),
        "n_bars_y": int(getattr(ly, "n_bars_y", 0)),
        "n_bars_z": int(getattr(ly, "n_bars_z", 0)),
        "as_prov_mm2": float(getattr(ly, "as_prov_mm2", 0.0)),
        "phi_st_mm": float(getattr(ly, "phi_st_mm", 8.0)),
        "s_st_mm": sprov,
        "s_st_max_mm": smax,
        "mrd_y_kNm": my_cap,
        "mrd_z_kNm": mz_cap,
        "utilizacao": util,
        "status": status,
        "solucao": solucao,
        "layout_type": "construtivo_misto" if isinstance(ly, MixedLayout) and ly.phi_corner_mm != ly.phi_face_mm else "construtivo_uniforme",
        "layout_description": desc,
        "layout_policy_v56": "varões de canto explícitos + varões de face distribuídos simetricamente",
        "v_ed_y_kN": vy_ed_kN, "v_ed_z_kN": vz_ed_kN,
        "v_rd_c_y_kN": sh_y.get("VRdc_kN"), "v_rd_max_y_kN": sh_y.get("VRdmax_kN"), "asw_s_y_req_mm2_per_m": sh_y.get("Asw_s_req_mm2_per_m"), "shear_status_y": sh_y.get("status"),
        "v_rd_c_z_kN": sh_z.get("VRdc_kN"), "v_rd_max_z_kN": sh_z.get("VRdmax_kN"), "asw_s_z_req_mm2_per_m": sh_z.get("Asw_s_req_mm2_per_m"), "shear_status_z": sh_z.get("status"),
        "shear_status": "Verificar" if any("Requer" in str(x.get("status", "")) for x in [sh_y, sh_z]) else ("Não conforme" if any("Não conforme" in str(x.get("status", "")) for x in [sh_y, sh_z]) else "OK"),
        "t_rd_max_kNm": tor.get("TRdmax_kNm"), "asw_s_t_req_mm2_per_m": tor.get("Asw_s_t_req_mm2_per_m"), "asl_t_req_mm2": tor.get("Asl_t_req_mm2"),
        "torsion_status": tor.get("torsion_status"), "torsion_utilization_TRdmax": tor.get("torsion_utilization_TRdmax"), "torsion_ratio": tor.get("torsion_ratio"),
        **det,
        **els,
        "failure_reason": "; ".join(blockers),
        "failure_type": classify_failure_reason("; ".join(blockers)),
        "failure_severity": severity,
        "failure_warnings": "; ".join(warnings) if warnings else "-",
        "failure_info": "; ".join(info) if info else "-",
        "design_decision": "Não adoptar sem revisão" if blockers else ("Adoptável com revisão indicada" if warnings else "OK"),
        "review_priority": "Alta" if blockers else ("Média" if warnings else "Normal"),
        "failure_action": "; ".join(blockers + warnings) if (blockers or warnings) else "-",
        "failure_summary": f"{severity} | {('; '.join(blockers + warnings)) if (blockers or warnings) else 'OK'}",
        "recommendations": "; ".join(dict.fromkeys([x for x in blockers + warnings + info if x])),
        "shortlist_text": serialize_shortlist(shortlist_rows[:50]),
        "biaxial_catalogue_v56": f"testadas {min(len(candidates), max_tests)} soluções construtivas admissíveis",
        "surface_method": "catálogo construtivo + superfície discreta N-My-Mz com cache v5.6",
    }
    return out


ColumnDesigner.design_one = _design_one_v56


def _v56_bar_points_from_result(r) -> List[Tuple[float, float, float]]:
    b = _finite(r.get("b_cm")) * 10.0
    h = _finite(r.get("h_cm")) * 10.0
    cover = _finite(r.get("cover_mm"), 35.0)
    phi_st = _finite(r.get("phi_st_mm"), 8.0)
    pc = _finite(r.get("phi_corner_mm", r.get("phi_long_mm", 0.0)), 0.0)
    pf = _finite(r.get("phi_face_mm", r.get("phi_long_mm", 0.0)), 0.0)
    ey = int(_finite(r.get("n_face_y_extra", 0), 0))
    ez = int(_finite(r.get("n_face_z_extra", 0), 0))
    if b <= 0 or h <= 0 or pc <= 0:
        return []
    ly = MixedLayout(pc, pf if pf > 0 else pc, ey, ez, b, h, cover, phi_st)
    return _v56_bar_points(ly)


def _bar_points_for_result_v56(r):
    pts = _v56_bar_points_from_result(r)
    if pts:
        # Compatibilidade com rotinas antigas que esperavam (y,z); DXF v56 usa phi.
        return [(y, z) for y, z, _phi in pts]
    return _bar_points_for_result(r) if "_bar_points_for_result" in globals() else []

# Não substituir _bar_points_for_result global por causa de recursão; apenas o DXF v56 usa a função de 3 valores.


def write_columns_dxf_v56(path: str, df: pd.DataFrame):
    parts = [
        "0\nSECTION\n2\nHEADER\n0\nENDSEC\n",
        "0\nSECTION\n2\nTABLES\n0\nENDSEC\n",
        "0\nSECTION\n2\nENTITIES\n",
    ]
    if df is None or df.empty:
        parts.append(_dxf_text(0, 0, "Sem resultados", 50, "COLUMN_TEXT"))
    else:
        # Secções-tipo: agrupar por secção/material/armadura.
        work = df.copy()
        for col in ["b_cm", "h_cm", "material", "layout_description", "solucao", "status"]:
            if col not in work.columns:
                work[col] = ""
        work["_type_key"] = work.apply(lambda r: f"{_finite(r.get('b_cm')):.0f}x{_finite(r.get('h_cm')):.0f}|{r.get('material','')}|{r.get('layout_description', r.get('solucao',''))}", axis=1)
        types = work.groupby("_type_key", dropna=False).first().reset_index()
        x0, y0 = 0.0, 0.0
        dx, dy = 1700.0, -1350.0
        for idx, (_, r) in enumerate(types.head(80).iterrows()):
            col = idx % 4
            row = idx // 4
            ox = x0 + col * dx
            oy = y0 + row * dy
            b = _finite(r.get("b_cm")) * 10.0
            h = _finite(r.get("h_cm")) * 10.0
            if b <= 0 or h <= 0:
                continue
            left, right = ox - b/2, ox + b/2
            bot, top = oy - h/2, oy + h/2
            parts += [
                _dxf_line(left, bot, right, bot, "COLUMN_CONCRETE"),
                _dxf_line(right, bot, right, top, "COLUMN_CONCRETE"),
                _dxf_line(right, top, left, top, "COLUMN_CONCRETE"),
                _dxf_line(left, top, left, bot, "COLUMN_CONCRETE"),
            ]
            c = _finite(r.get("cover_mm"), 35.0)
            l2, r2, b2, t2 = left + c, right - c, bot + c, top - c
            parts += [
                _dxf_line(l2, b2, r2, b2, "COLUMN_STIRRUP"),
                _dxf_line(r2, b2, r2, t2, "COLUMN_STIRRUP"),
                _dxf_line(r2, t2, l2, t2, "COLUMN_STIRRUP"),
                _dxf_line(l2, t2, l2, b2, "COLUMN_STIRRUP"),
            ]
            for y, z, phi in _v56_bar_points_from_result(r):
                parts.append(_dxf_circle(ox + y, oy + z, max(phi/2.0, 3.0), "COLUMN_REBAR"))
            parts.append(_dxf_text(ox - b/2, oy - h/2 - 90, f"Tipo {idx+1}: {_finite(r.get('b_cm')):.0f}x{_finite(r.get('h_cm')):.0f} cm", 28, "COLUMN_TEXT"))
            parts.append(_dxf_text(ox - b/2, oy - h/2 - 135, str(r.get("solucao", "")), 24, "COLUMN_TEXT"))
        # Quadro-resumo.
        tx, ty = 0.0, y0 + ((len(types.head(80)) // 4) + 2) * dy
        headers = ["PRUMADA", "MEMBER", "SECÇÃO", "MATERIAL", "ARMADURA", "ESTRIBOS", "ESTADO"]
        widths = [300, 260, 300, 300, 700, 350, 260]
        x = tx
        for htxt, w in zip(headers, widths):
            parts.append(_dxf_text(x + 5, ty, htxt, 28, "COLUMN_TABLE")); x += w
        for i, (_, r) in enumerate(work.head(160).iterrows(), start=1):
            y = ty - i * 80
            vals = [
                r.get("name", ""), r.get("member", ""),
                f"{_finite(r.get('b_cm')):.0f}x{_finite(r.get('h_cm')):.0f} cm",
                r.get("material", ""), r.get("layout_description", r.get("solucao", "")),
                f"Ø{int(_finite(r.get('phi_st_mm'), 0))}//{_finite(r.get('s_st_mm'))/10:.1f}cm",
                r.get("status", ""),
            ]
            x = tx
            for val, w in zip(vals, widths):
                parts.append(_dxf_text(x + 5, y, val, 22, "COLUMN_TABLE")); x += w
    parts.append("0\nENDSEC\n0\nEOF\n")
    Path(path).write_text("".join(parts), encoding="utf-8")


def _export_dxf_v56(self):
    src = self.df_summary if self.df_summary is not None and not self.df_summary.empty else self.df_results
    if src is None or src.empty:
        messagebox.showwarning("Aviso", "Não há resultados para exportar em DXF.")
        return
    path = filedialog.asksaveasfilename(title="Exportar secções-tipo e quadro de pilares", defaultextension=".dxf", filetypes=[("DXF", "*.dxf")])
    if not path:
        return
    try:
        self.status_var.set("A exportar DXF v5.6...")
        self.progress_var.set(10.0)
        write_columns_dxf_v56(path, src)
        self.progress_var.set(100.0)
        self.status_var.set(f"DXF exportado para: {path}")
    except Exception as err:
        messagebox.showerror("Erro", f"Não foi possível exportar DXF.\n\n{err}")

try:
    ColumnsEC2App.export_dxf = _export_dxf_v56
except Exception:
    pass


def _metadata_df_v56(self) -> pd.DataFrame:
    try:
        df = _metadata_df_v55(self).copy()
    except Exception:
        df = pd.DataFrame(columns=["Campo", "Valor"])
    extra = pd.DataFrame([
        ["Versão de cálculo", APP_VERSION],
        ["Motor default v5.6", "catálogo construtivo explícito + verificação N-My-Mz por layouts admissíveis"],
        ["Cache de capacidade", "chave completa incluindo diâmetros de canto/face e distribuição"],
        ["Pormenorização", "separação entre bloqueante, aviso e informativo"],
        ["Torção", "limiar de relevância antes de exigir armadura de torção"],
    ], columns=["Campo", "Valor"])
    return pd.concat([df, extra], ignore_index=True)

try:
    ColumnsEC2App._metadata_df = _metadata_df_v56
except Exception:
    pass


def _build_normative_notes_v56(self) -> pd.DataFrame:
    try:
        base = _build_normative_notes_v54(self).copy() if "_build_normative_notes_v54" in globals() else self.build_normative_notes()
    except Exception:
        base = pd.DataFrame(columns=["Tema", "Referência", "Nota"])
    add = pd.DataFrame([
        ("v5.6", "Layouts de pilares", "As soluções são geradas como varões de canto + varões de face, com coordenadas explícitas."),
        ("v5.6", "Biaxial", "A falha biaxial só é declarada após pesquisa no catálogo construtivo admissível."),
        ("v5.6", "Pormenorização", "Avisos construtivos não são automaticamente falhas bloqueantes."),
        ("v5.6", "Torção", "TEd pequeno face a TRd,max/flexão é classificado como não condicionante ou aviso, não como falha."),
    ], columns=base.columns if len(base.columns) == 3 else ["Tema", "Referência", "Nota"])
    return pd.concat([base, add], ignore_index=True)

try:
    ColumnsEC2App.build_normative_notes = _build_normative_notes_v56
except Exception:
    pass

# Garantir que o filtro permite Aviso.
def _patch_status_filter_values_v56(app):
    try:
        def walk(w):
            try:
                if isinstance(w, ttk.Combobox) and str(w.cget("textvariable")) == str(app.var_filter_status):
                    w.configure(values=["Todos", "OK", "Aviso", "Falha", "Pré-dimensionado"])
            except Exception:
                pass
            for ch in w.winfo_children():
                walk(ch)
        walk(app)
    except Exception:
        pass

_old_init_v56 = ColumnsEC2App.__init__
def _init_v56(self, *args, **kwargs):
    _old_init_v56(self, *args, **kwargs)
    try:
        _patch_status_filter_values_v56(self)
    except Exception:
        pass
ColumnsEC2App.__init__ = _init_v56


# ============================================================
# ColumnsEC2 v5.7 — reparação da rota structuralcodes
# - corrige o dispatch dos backends EC2 2004 / EC2 2023 / fib MC2010
#   depois da revisão v5.6 do motor PT;
# - força criação explícita dos materiais/geom. com design_code;
# - corrige chamadas EC2 2004 de esforço transverso segundo a API actual
#   do structuralcodes;
# - mantém o motor NP EN 1992-1-1:2010 PT como default.
# ============================================================
APP_VERSION = "v5.7"


def _is_structuralcodes_backend_v52(value=None) -> bool:
    """Função de compatibilidade: alguns patches v5.6 chamavam este nome,
    mas nem sempre ele existia no ficheiro final. Mantém a decisão explícita.
    """
    try:
        return _backend_selected_v52(value) in SC_BACKENDS_V52
    except Exception:
        return "structuralcodes" in str(value or "").lower()

globals()["_is_structuralcodes_backend_v52"] = _is_structuralcodes_backend_v52


# Guardar importador anterior para diagnóstico, mas substituir por uma versão
# mais explícita e compatível com structuralcodes 0.7.x.
_old_sc_import_backend_v52_for_v57 = _sc_import_backend_v52

def _sc_import_backend_v52(backend=None):
    b = _backend_selected_v52(backend)
    key = _sc_backend_key_v52(b)
    if key == "pt_2010":
        return None, "Backend interno PT não usa structuralcodes."
    try:
        import importlib
        from structuralcodes import set_design_code
        from structuralcodes.geometry import SurfaceGeometry, add_reinforcement
        from structuralcodes.materials.concrete import create_concrete
        from structuralcodes.materials.reinforcement import create_reinforcement
        from structuralcodes.sections import BeamSection
        try:
            from shapely import Polygon
        except Exception:
            from shapely.geometry import Polygon

        # Abreviações oficialmente aceites pela API actual.
        code_name = {"ec2_2004": "ec2_2004", "ec2_2023": "ec2_2023", "mc2010": "mc2010"}[key]
        set_design_code(code_name)
        modname = {
            "ec2_2004": "structuralcodes.codes.ec2_2004",
            "ec2_2023": "structuralcodes.codes.ec2_2023",
            "mc2010": "structuralcodes.codes.mc2010",
        }[key]
        return {
            "backend": b,
            "key": key,
            "design_code": code_name,
            "module": importlib.import_module(modname),
            "Polygon": Polygon,
            "SurfaceGeometry": SurfaceGeometry,
            "add_reinforcement": add_reinforcement,
            "create_concrete": create_concrete,
            "create_reinforcement": create_reinforcement,
            "BeamSection": BeamSection,
        }, None
    except Exception as err:
        return None, err

globals()["_sc_import_backend_v52"] = _sc_import_backend_v52


# Materiais structuralcodes mais tolerantes às assinaturas actuais, mas sem fallback
# para fórmulas internas. Se uma propriedade essencial não existir, o erro é reportado.
def _sc_materials_v52(material: str, fyk=500.0, backend=None):
    sc, err = _sc_import_backend_v52(backend)
    if sc is None:
        raise RuntimeError(f"structuralcodes indisponível: {err}")
    mod = sc["module"]
    key = sc["key"]
    fck = parse_concrete_strength(material)
    gamma_c = 1.5
    gamma_s = 1.15
    if key == "ec2_2023":
        eta_cc_val = _safe_sc_call_v52(mod, ["eta_cc"], fck=fck, f_ck=fck)[0] if hasattr(mod, "eta_cc") else 1.0
        k_tc_val = _safe_sc_call_v52(mod, ["k_tc"], t_ref=28, t0=28, strength_dev_class="CN")[0] if hasattr(mod, "k_tc") else 1.0
        fcd_val, fcd_src = _safe_sc_call_v52(mod, ["fcd"], fck=fck, eta_cc=eta_cc_val, k_tc=k_tc_val, gamma_c=gamma_c)
        fctm_val, fctm_src = _safe_sc_call_v52(mod, ["fctm"], fck=fck)
        E_val, E_src = _safe_sc_call_v52(mod, ["Ecm"], fcm=fck + 8.0)
        fyd_val, fyd_src = _safe_sc_call_v52(mod, ["fyd"], fyk=float(fyk), gamma_s=gamma_s)
        Es_val, Es_src = _safe_sc_call_v52(mod, ["Es"])
    elif key == "ec2_2004":
        fcd_val, fcd_src = _safe_sc_call_v52(mod, ["fcd"], fck=fck, alpha_cc=1.0, gamma_c=gamma_c)
        fctm_val, fctm_src = _safe_sc_call_v52(mod, ["fctm"], fck=fck)
        E_val, E_src = _safe_sc_call_v52(mod, ["Ecm"], fcm=fck + 8.0)
        fyd_val, fyd_src = _safe_sc_call_v52(mod, ["fyd"], fyk=float(fyk), gamma_s=gamma_s)
        # EC2 2004 no pacote pode não expor Es(); neste backend o aço é criado
        # com Es=210000 MPa, ficando isto declarado no relatório.
        Es_val, Es_src = 210000.0, "input structuralcodes.create_reinforcement(Es=210000)"
    else:  # mc2010
        fcd_val, fcd_src = _safe_sc_call_v52(mod, ["fcd"], fck=fck, alpha_cc=1.0, gamma_c=gamma_c)
        fctm_val, fctm_src = _safe_sc_call_v52(mod, ["fctm"], fck=fck)
        E_val, E_src = _safe_sc_call_v52(mod, ["Eci"], fcm=fck + 8.0)
        fyd_val, fyd_src = _safe_sc_call_v52(mod, ["fyd"], fyk=float(fyk), gamma_s=gamma_s)
        Es_val, Es_src = 210000.0, "input structuralcodes.create_reinforcement(Es=210000)"
    return {
        "fck": float(fck),
        "fcm": float(fck) + 8.0,
        "fcd": float(fcd_val),
        "fctm": float(fctm_val),
        "Ecm": float(E_val),
        "fyd": float(fyd_val),
        "Es": float(Es_val),
        "backend": sc["backend"],
        "sources": f"fcd:{fcd_src}; fctm:{fctm_src}; E:{E_src}; fyd:{fyd_src}; Es:{Es_src}",
    }

globals()["_sc_materials_v52"] = _sc_materials_v52


# Recriar secção structuralcodes com design_code explícito. Isto evita que o
# backend global fique no código errado depois de alternar EC2 2004/2023/MC2010.
def _sc_section_from_layout_v52(layout, material: str, fyk: float, backend=None):
    sc, err = _sc_import_backend_v52(backend)
    if sc is None:
        raise RuntimeError(f"structuralcodes indisponível: {err}")
    design_code = sc.get("design_code")
    try:
        concrete = sc["create_concrete"](
            fck=float(parse_concrete_strength(material)),
            gamma_c=1.5,
            design_code=design_code,
        )
    except Exception as e:
        raise RuntimeError(f"create_concrete falhou ({design_code}): {e}")
    try:
        reinforcement = sc["create_reinforcement"](
            fyk=float(fyk),
            Es=210000.0,
            ftk=max(float(fyk) * 1.08, float(fyk) + 1.0),
            epsuk=0.05,
            gamma_s=1.15,
            design_code=design_code,
        )
    except Exception as e:
        raise RuntimeError(f"create_reinforcement falhou ({design_code}): {e}")
    b = float(getattr(layout, "b_mm", 0.0))
    h = float(getattr(layout, "h_mm", 0.0))
    if b <= 0 or h <= 0:
        raise RuntimeError("Geometria inválida para criação da secção structuralcodes.")
    poly = sc["Polygon"]([(-b / 2, -h / 2), (b / 2, -h / 2), (b / 2, h / 2), (-b / 2, h / 2)])
    geometry = sc["SurfaceGeometry"](poly=poly, material=concrete, concrete=True)
    pts = _layout_points_v52(layout)
    if not pts:
        pts = _v56_bar_points(layout) if "_v56_bar_points" in globals() else []
    for y, z, phi in pts:
        geometry = sc["add_reinforcement"](geometry, (float(y), float(z)), float(phi), reinforcement)
    return sc["BeamSection"](geometry=geometry)

globals()["_sc_section_from_layout_v52"] = _sc_section_from_layout_v52


# Esforço transverso structuralcodes: corrigir EC2 2004 para a assinatura actual.
def _sc_shear_check_v52(out, row, layout, mats, backend=None):
    sc, err = _sc_import_backend_v52(backend)
    if sc is None:
        out["shear_status_y"] = out["shear_status_z"] = f"Aviso: structuralcodes indisponível ({err})"
        return out
    mod = sc["module"]
    key = sc["key"]
    fck = float(mats["fck"])
    fcd = float(mats.get("fcd", 0.0))
    fyd = float(mats.get("fyd", 500.0 / 1.15))
    Es = float(mats.get("Es", 210000.0))
    As = float(out.get("as_prov_mm2") or getattr(layout, "as_prov_mm2", 0.0) or 0.0)
    b = float(getattr(layout, "b_mm", 0.0))
    h = float(getattr(layout, "h_mm", 0.0))
    z_y = 0.8 * h
    z_z = 0.8 * b
    Ac = max(b * h, 1.0)
    NEd = abs(float(out.get("n_ed_kN", safe_float(row.get("fx", 0.0), 0.0)))) * 1000.0
    fy = max(abs(safe_float(row.get("fy_i", row.get("fy", 0.0)), 0.0)), abs(safe_float(row.get("fy_j", row.get("fy", 0.0)), 0.0)))
    fz = max(abs(safe_float(row.get("fz_i", row.get("fz", 0.0)), 0.0)), abs(safe_float(row.get("fz_j", row.get("fz", 0.0)), 0.0)))
    if key == "ec2_2023":
        out["shear_status_y"] = out["shear_status_z"] = "Aviso: esforço transverso não exposto no módulo structuralcodes EC2:2023"
        out["shear_backend"] = sc["backend"]
        return out
    try:
        if key == "ec2_2004":
            vrdc_y, src1 = _safe_sc_call_v52(mod, ["VRdc"], fck=fck, d=z_y, Asl=As, bw=b, NEd=NEd, Ac=Ac, fcd=fcd, gamma_c=1.5)
            vrdmax_y, src2 = _safe_sc_call_v52(mod, ["VRdmax"], bw=b, z=z_y, fck=fck, theta=45.0, NEd=NEd, Ac=Ac, fcd=fcd, alpha=90.0)
            vrdc_z, src3 = _safe_sc_call_v52(mod, ["VRdc"], fck=fck, d=z_z, Asl=As, bw=h, NEd=NEd, Ac=Ac, fcd=fcd, gamma_c=1.5)
            vrdmax_z, src4 = _safe_sc_call_v52(mod, ["VRdmax"], bw=h, z=z_z, fck=fck, theta=45.0, NEd=NEd, Ac=Ac, fcd=fcd, alpha=90.0)
        else:
            loads_y = {"ned": NEd, "med": abs(float(out.get("my_ed_kNm", 0.0))) * 1e6, "ved": fy * 1e3}
            loads_z = {"ned": NEd, "med": abs(float(out.get("mz_ed_kNm", 0.0))) * 1e6, "ved": fz * 1e3}
            vrdc_y, src1 = _safe_sc_call_v52(mod, ["v_rdc_approx1", "v_rdc"], approx_lvl=1, fck=fck, z=z_y, bw=b, dg=16.0, E_s=Es, As=As, loads=loads_y, gamma_c=1.5)
            vrdmax_y, src2 = _safe_sc_call_v52(mod, ["v_rd_max_approx1", "v_rd_max_approx2"], fck=fck, f_ck=fck, bw=b, theta=45.0, z=z_y, E_s=Es, As=As, loads=loads_y, gamma_c=1.5)
            vrdc_z, src3 = _safe_sc_call_v52(mod, ["v_rdc_approx1", "v_rdc"], approx_lvl=1, fck=fck, z=z_z, bw=h, dg=16.0, E_s=Es, As=As, loads=loads_z, gamma_c=1.5)
            vrdmax_z, src4 = _safe_sc_call_v52(mod, ["v_rd_max_approx1", "v_rd_max_approx2"], fck=fck, f_ck=fck, bw=h, theta=45.0, z=z_z, E_s=Es, As=As, loads=loads_z, gamma_c=1.5)
        def tokN(v):
            v = float(v)
            return v / 1000.0 if abs(v) > 1e4 else v
        out["v_rd_c_y_kN"] = tokN(vrdc_y)
        out["v_rd_c_z_kN"] = tokN(vrdc_z)
        out["v_rd_max_y_kN"] = tokN(vrdmax_y)
        out["v_rd_max_z_kN"] = tokN(vrdmax_z)
        out["shear_backend"] = sc["backend"]
        out["shear_sources"] = f"Y: {src1}/{src2}; Z: {src3}/{src4}"
        out["shear_status_y"] = "OK" if fy <= out["v_rd_max_y_kN"] else "Falha: VEd,y > VRd,max"
        out["shear_status_z"] = "OK" if fz <= out["v_rd_max_z_kN"] else "Falha: VEd,z > VRd,max"
    except Exception as e:
        out["shear_status_y"] = out["shear_status_z"] = f"Aviso: esforço transverso não calculado por structuralcodes ({e})"
        out["shear_backend"] = sc["backend"]
    return out

globals()["_sc_shear_check_v52"] = _sc_shear_check_v52


# Dispatch final: garante que qualquer backend structuralcodes chama directamente
# a rota strict, sem passar pelo novo motor v5.6 do backend PT.
_old_design_one_v56_for_v57 = ColumnDesigner.design_one

def _design_one_v57(self, row: pd.Series, prebuilt_candidates=None):
    backend = _backend_selected_v52(getattr(self, "code_backend", globals().get("ACTIVE_CODE_BACKEND_V48", BACKEND_EC2_PT_2010)))
    if _sc_backend_active_v52(backend):
        # Não usar fórmulas internas de dimensionamento PT. A geração geométrica de
        # candidatos é apenas input; a verificação resistente é feita via structuralcodes.
        return _strict_sc_design_one_v52(self, row, prebuilt_candidates=prebuilt_candidates)
    return _old_design_one_v56_for_v57(self, row, prebuilt_candidates=prebuilt_candidates)

ColumnDesigner.design_one = _design_one_v57


# Diagnóstico textual mais claro quando o pacote não estiver instalado.
_old_validate_inputs_v56_for_v57 = ColumnsEC2App.validate_inputs

def _validate_inputs_v57(self):
    err = _old_validate_inputs_v56_for_v57(self)
    if err:
        return err
    backend = _backend_selected_v52(getattr(self, "var_code_backend", tk.StringVar(value=BACKEND_EC2_PT_2010)).get())
    if _sc_backend_active_v52(backend):
        sc, sc_err = _sc_import_backend_v52(backend)
        if sc is None:
            return (
                f"O backend {backend} não está operacional porque o pacote structuralcodes não foi carregado.\n\n"
                "Instale/actualize com:\n"
                "python -m pip install --upgrade structuralcodes shapely numpy\n\n"
                "Neste modo não há fallback para fórmulas internas.\n\n"
                f"Erro original: {sc_err}"
            )
    return None

ColumnsEC2App.validate_inputs = _validate_inputs_v57


# Actualizar notas/metadados para v5.7.
_old_metadata_df_v56_for_v57 = ColumnsEC2App._metadata_df

def _metadata_df_v57(self) -> pd.DataFrame:
    try:
        df = _old_metadata_df_v56_for_v57(self).copy()
    except Exception:
        df = pd.DataFrame(columns=["Campo", "Valor"])
    extra = pd.DataFrame([
        ["Versão", APP_VERSION],
        ["Correcção v5.7", "reparada a rota structuralcodes para EC2 2004, EC2 2023 e fib MC2010"],
        ["Backends", "default PT mantém motor interno; structuralcodes mantém modo strict sem fallback"],
    ], columns=["Campo", "Valor"])
    return pd.concat([df, extra], ignore_index=True)

ColumnsEC2App._metadata_df = _metadata_df_v57


_old_build_normative_notes_v56_for_v57 = ColumnsEC2App.build_normative_notes

def _build_normative_notes_v57(self) -> pd.DataFrame:
    try:
        notes = _old_build_normative_notes_v56_for_v57(self).copy()
    except Exception:
        notes = pd.DataFrame(columns=["Tema", "Referência", "Nota"])
    extra = pd.DataFrame([
        ("v5.7", "structuralcodes", "Reparado o dispatch dos backends EC2 2004, EC2 2023 e fib MC2010 após a revisão v5.6 do motor PT."),
        ("v5.7", "EC2 2004 structuralcodes", "Corrigidas as chamadas VRdc/VRdmax para a assinatura actual do pacote."),
        ("v5.7", "Materiais structuralcodes", "Materiais e secções passam a ser criados com design_code explícito para evitar mistura de backends."),
    ], columns=["Tema", "Referência", "Nota"])
    return pd.concat([notes, extra], ignore_index=True).drop_duplicates()

ColumnsEC2App.build_normative_notes = _build_normative_notes_v57


# v5.7.1 — aceleração dos backends structuralcodes
APP_VERSION = "v5.7.1"


def _sc_nmm_capacities_v52(layout, n_ed_kN: float, material: str, fyk: float, backend=None):
    """Capacidade N-My-Mz via structuralcodes com cache e discretização controlada.

    Continua strict: a resistência vem de structuralcodes.sections; não há
    substituição por fórmulas internas. A alteração é apenas de desempenho.
    """
    if not hasattr(_sc_nmm_capacities_v52, "_cache"):
        _sc_nmm_capacities_v52._cache = {}
    try:
        sig = _v56_layout_signature(layout) if "_v56_layout_signature" in globals() else (
            round(float(getattr(layout, "b_mm", 0.0)), 1), round(float(getattr(layout, "h_mm", 0.0)), 1),
            round(float(getattr(layout, "phi_long_mm", 0.0)), 2), int(getattr(layout, "n_total", 0)),
        )
    except Exception:
        sig = str(layout)
    bkey = _backend_selected_v52(backend)
    ckey = (bkey, str(material), round(float(fyk), 1), sig, round(float(n_ed_kN), 1))
    if ckey in _sc_nmm_capacities_v52._cache:
        return _sc_nmm_capacities_v52._cache[ckey]

    section = _sc_section_from_layout_v52(layout, material, fyk, backend=backend)
    calc = getattr(section, "section_calculator", None)
    if calc is None:
        raise RuntimeError("BeamSection não expõe section_calculator.")
    method = None
    method_name = ""
    for name in ["calculate_nmm_interaction_domain", "nmm_interaction_domain", "calculate_nm_interaction_domain"]:
        if hasattr(calc, name):
            method = getattr(calc, name)
            method_name = name
            break
    if method is None:
        raise RuntimeError("SectionCalculator não expõe função de domínio N-My-Mz.")

    # Parâmetros leves, mas ainda estáveis para triagem de layouts.
    attempts = [
        {"num_theta": 18, "num_3": 10, "num_4": 6},
        {"num_theta": 24, "num_3": 12, "num_4": 8},
        {"num_theta": 18},
        {},
    ]
    result = None
    last = None
    for kw in attempts:
        try:
            result = _call_sc_func_v52(method, **kw)
            break
        except Exception as err:
            last = err
    if result is None:
        raise RuntimeError(f"{method_name} falhou: {last}")
    n, my, mz = _extract_nmm_arrays_v52(result)
    if n is None:
        raise RuntimeError("Não foi possível extrair arrays N, My, Mz do resultado do structuralcodes.")
    import numpy as np
    n = np.asarray(n, dtype=float)
    my = np.asarray(my, dtype=float)
    mz = np.asarray(mz, dtype=float)
    mask = np.isfinite(n) & np.isfinite(my) & np.isfinite(mz)
    n = n[mask]
    my = my[mask]
    mz = mz[mask]
    if len(n) < 3:
        raise RuntimeError("Domínio N-My-Mz devolveu pontos insuficientes.")
    target = abs(float(n_ed_kN) * 1e3)
    order = np.argsort(np.abs(np.abs(n) - target))[:max(12, min(50, len(n)))]
    caps = []
    for i in order:
        myk = abs(float(my[i])) / 1e6
        mzk = abs(float(mz[i])) / 1e6
        if myk > 1e-9 and mzk > 1e-9:
            caps.append((myk, mzk))
    if not caps:
        raise RuntimeError("Domínio N-My-Mz sem pontos resistentes úteis para o nível de NEd.")
    value = (caps, f"structuralcodes.sections.{method_name}")
    _sc_nmm_capacities_v52._cache[ckey] = value
    return value

globals()["_sc_nmm_capacities_v52"] = _sc_nmm_capacities_v52


# Limitar a pesquisa de layouts nos backends structuralcodes para evitar tempos excessivos.
_old_strict_sc_design_one_v52_for_v571 = _strict_sc_design_one_v52

def _strict_sc_design_one_v52(self, row: pd.Series, prebuilt_candidates=None):
    if prebuilt_candidates is not None:
        # Ordenação prática antes da rota strict original.
        try:
            material = str(row.get("material", "") or "").strip()
            fck = parse_concrete_strength(material) if material else 30.0
            mats = _sc_materials_v52(material or "C30/37", fyk=float(getattr(self, "fyk", 500.0)), backend=getattr(self, "code_backend", None))
            b_mm = cm_to_mm(row.get("hy", 0.0)); h_mm = cm_to_mm(row.get("hz", 0.0))
            ac_mm2 = safe_float(row.get("ax", float('nan'))) * 100.0
            if b_mm <= 0 and ac_mm2 > 0: b_mm = math.sqrt(ac_mm2)
            if h_mm <= 0: h_mm = b_mm
            if not math.isfinite(ac_mm2) or ac_mm2 <= 0: ac_mm2 = b_mm * h_mm
            n_ed = max(abs(safe_float(row.get("fx_i", row.get("fx", 0.0)), 0.0)), abs(safe_float(row.get("fx_j", row.get("fx", 0.0)), 0.0)))
            as_min = max(0.10 * n_ed * 1e3 / max(float(mats.get("fyd", 1.0)), 1e-9), 0.002 * ac_mm2)
            as_max = 0.04 * ac_mm2
            pcs = [c for c in prebuilt_candidates if c.as_prov_mm2 >= as_min and c.as_prov_mm2 <= as_max]
            pcs = sorted(pcs, key=lambda c: (_v56_layout_score(c, as_min, b_mm, h_mm) if "_v56_layout_score" in globals() else (c.as_prov_mm2, c.n_total)))[:18]
            prebuilt_candidates = pcs
        except Exception:
            try:
                prebuilt_candidates = list(prebuilt_candidates)[:18]
            except Exception:
                pass
    return _old_strict_sc_design_one_v52_for_v571(self, row, prebuilt_candidates=prebuilt_candidates)

globals()["_strict_sc_design_one_v52"] = _strict_sc_design_one_v52

# ============================================================
# ColumnsEC2 v5.8 — correcção de activação dos backends structuralcodes
# A versão v5.7 tinha o bloco __main__ antes dos patches v5.7/v5.7.1;
# quando executada como script, a GUI arrancava antes de estes patches ficarem activos.
# Nesta versão, o arranque fica no fim do ficheiro.
# ============================================================
APP_VERSION = "v5.8"


# ============================================================
# ColumnsEC2 v5.9 — relatórios limpos por backend, notas normativas
# dinâmicas, parâmetros de fluência automáticos e correcção iterativa
# controlada por propostas.
# ============================================================
APP_VERSION = "v5.9"
APP_XLSX_DESCRIPTION = (
    "Workbook profissional de dimensionamento/verificação de pilares de betão armado. "
    "O motor NP EN 1992-1-1:2010 PT executa ELU, 2.ª ordem, N-My-Mz, V, T, ELS e pormenorização "
    "pelo motor interno do ColumnsEC2. Os modos Eurocode 2:2004, Eurocode 2:2023 e fib Model Code 2010 "
    "usam apenas o pacote structuralcodes nas verificações disponibilizadas pela respectiva API, sem fallback "
    "normativo por fórmulas internas."
)

# Rótulos comerciais/limpos apresentados na GUI e relatórios.
BACKEND_DISPLAY_PT_V59 = "NP EN 1992-1-1:2010 + AC:2012 + A1:2019 (Portugal)"
BACKEND_DISPLAY_EC2004_V59 = "Eurocode 2:2004"
BACKEND_DISPLAY_EC2023_V59 = "Eurocode 2:2023"
BACKEND_DISPLAY_MC2010_V59 = "fib Model Code 2010"
BACKEND_DISPLAY_CHOICES_V59 = [
    BACKEND_DISPLAY_PT_V59,
    BACKEND_DISPLAY_EC2004_V59,
    BACKEND_DISPLAY_EC2023_V59,
    BACKEND_DISPLAY_MC2010_V59,
]

def _backend_selected_v52(value=None) -> str:
    """Aceita rótulos limpos da GUI e devolve o identificador interno."""
    s0 = str(value or globals().get("ACTIVE_CODE_BACKEND_V48", BACKEND_EC2_PT_2010)).strip()
    s = s0.lower()
    if "fib" in s and ("2010" in s or "model" in s):
        return BACKEND_SC_MC2010
    if "2004" in s and ("eurocode" in s or "ec2" in s or "structuralcodes" in s):
        return BACKEND_SC_EC2_2004
    if "2023" in s and ("eurocode" in s or "ec2" in s or "structuralcodes" in s):
        return BACKEND_SC_EC2_2023
    if "portugal" in s or "np en 1992" in s or "default" in s:
        return BACKEND_EC2_PT_2010
    if s0 in [BACKEND_SC_EC2_2004, BACKEND_SC_EC2_2023, BACKEND_SC_MC2010, BACKEND_EC2_PT_2010]:
        return s0
    return BACKEND_EC2_PT_2010

globals()["_backend_selected_v48"] = _backend_selected_v52

def _backend_display_v59(value=None) -> str:
    b = _backend_selected_v52(value)
    if b == BACKEND_SC_EC2_2004:
        return BACKEND_DISPLAY_EC2004_V59
    if b == BACKEND_SC_EC2_2023:
        return BACKEND_DISPLAY_EC2023_V59
    if b == BACKEND_SC_MC2010:
        return BACKEND_DISPLAY_MC2010_V59
    return BACKEND_DISPLAY_PT_V59

def _backend_reference_v59(app=None) -> str:
    b = _v59_get_backend(app)
    if b == BACKEND_SC_EC2_2004:
        return "EN 1992-1-1:2004 / Eurocode 2:2004"
    if b == BACKEND_SC_EC2_2023:
        return "Eurocode 2:2023"
    if b == BACKEND_SC_MC2010:
        return "fib Model Code for Concrete Structures 2010"
    return "NP EN 1992-1-1:2010 + AC:2012 + A1:2019, com Anexo Nacional português"

def _v59_get_backend(app=None):
    try:
        return _backend_selected_v52(getattr(app, "var_code_backend", tk.StringVar(value=BACKEND_DISPLAY_PT_V59)).get())
    except Exception:
        return _backend_selected_v52(globals().get("ACTIVE_CODE_BACKEND_V48", BACKEND_EC2_PT_2010))

def _v59_material_classes(app=None) -> str:
    vals = []
    for attr in ["df_pair", "df_clean"]:
        try:
            df = getattr(app, attr, pd.DataFrame())
            if df is not None and not df.empty and "material" in df.columns:
                vals.extend([str(v).strip() for v in df["material"].dropna().tolist()])
        except Exception:
            pass
    clean = []
    for v in vals:
        m = re.search(r"C\s*(\d{2,3})\s*/\s*(\d{2,3})", str(v), re.I)
        if m:
            clean.append(f"C{m.group(1)}/{m.group(2)}")
    clean = sorted(set(clean), key=lambda x: [int(n) for n in re.findall(r"\d+", x)[:1]] or [999])
    return ", ".join(clean) if clean else "não identificado na tabela"

# Aço: aceitar "A400", "A500", "A500 (500 MPa)" sem quebrar os cálculos existentes.
_safe_float_before_v59 = safe_float
def safe_float(value, default=float("nan")):
    try:
        s = str(value).strip().upper()
        m = re.search(r"\bA\s*(400|500)\b", s)
        if m:
            return float(m.group(1))
        m = re.search(r"\b(400|500)\s*MPA\b", s)
        if m:
            return float(m.group(1))
    except Exception:
        pass
    return _safe_float_before_v59(value, default)

globals()["safe_float"] = safe_float

def _v59_steel_class_label(app=None) -> str:
    try:
        fyk = safe_float(getattr(app, "var_fyk", tk.StringVar(value="A500")).get(), 500.0)
    except Exception:
        fyk = 500.0
    return f"A{int(round(fyk))}"

def _v59_estimate_h0_for_app(app=None) -> float:
    try:
        return float(_v54_estimate_h0_for_app(app))
    except Exception:
        try:
            return float(_v53_estimate_h0_from_df(getattr(app, "df_pair", pd.DataFrame())))
        except Exception:
            return 200.0

def _v59_compute_phi_eff_for_app(app=None):
    """φef automático para todos os backends.

    Nos backends structuralcodes, tenta-se primeiro a API do pacote. Se a API não expuser
    fluência, o valor é mantido apenas como parâmetro informativo; não é usado para aplicar
    fórmulas internas de 2.ª ordem nesses modos strict.
    """
    backend = _v59_get_backend(app)
    RH = safe_float(getattr(app, "var_creep_RH", tk.StringVar(value="70")).get(), 70.0)
    t0 = safe_float(getattr(app, "var_creep_t0", tk.StringVar(value="28")).get(), 28.0)
    h0 = _v59_estimate_h0_for_app(app)
    fck = 30.0
    try:
        mats = _v59_material_classes(app)
        m = re.search(r"C\s*(\d+)\s*/", mats)
        if m:
            fck = float(m.group(1))
    except Exception:
        pass

    if _sc_backend_active_v52(backend):
        try:
            val, src = _v53_try_structuralcodes_phi(backend, fck, RH, t0, h0)
            if val is not None and math.isfinite(float(val)) and float(val) > 0:
                return float(val), f"{src}; RH={RH:.0f}%; t0={t0:.0f} d; h0={h0:.0f} mm"
        except Exception as e:
            src = str(e)
        # Valor auxiliar apenas para preencher parâmetros e transparência; não afecta o cálculo strict.
        if backend in [BACKEND_SC_EC2_2004, BACKEND_SC_EC2_2023]:
            try:
                val = _v53_ec2_creep_phi(fck, RH=RH, t0=t0, h0_mm=h0)
                return float(val), (
                    f"φef estimado por expressão EC2 auxiliar porque a API de fluência não foi exposta; "
                    f"não usado nas verificações strict; RH={RH:.0f}%; t0={t0:.0f} d; h0={h0:.0f} mm"
                )
            except Exception:
                pass
        return 0.0, f"φef não disponível na API structuralcodes local para {_backend_display_v59(backend)}; não aplicado no backend strict; RH={RH:.0f}%; t0={t0:.0f} d; h0={h0:.0f} mm"

    try:
        val = _v53_ec2_creep_phi(fck, RH=RH, t0=t0, h0_mm=h0)
        return float(val), f"NP EN 1992-1-1: φef automático; RH={RH:.0f}%; t0={t0:.0f} d; h0={h0:.0f} mm; fck={fck:.0f} MPa"
    except Exception as e:
        return 2.0, f"φef automático não calculado ({e}); valor de segurança adoptado = 2.0"

def _validate_inputs_v59(self):
    err = _old_validate_inputs_v52(self) if "_old_validate_inputs_v52" in globals() else None
    if err:
        return err
    try:
        backend = _v59_get_backend(self)
        globals()["ACTIVE_CODE_BACKEND_V48"] = backend
        # Verificação structuralcodes se aplicável.
        if _sc_backend_active_v52(backend):
            sc, sc_err = _sc_import_backend_v52(backend)
            if sc is None:
                return (
                    f"O modo {_backend_display_v59(backend)} requer structuralcodes e shapely.\n\n"
                    "Instale/actualize com:\npython -m pip install --upgrade structuralcodes shapely numpy\n\n"
                    "Neste modo não há fallback para fórmulas internas.\n\n"
                    f"Erro original: {sc_err}"
                )
        phi, src = _v59_compute_phi_eff_for_app(self)
        if hasattr(self, "var_phi_eff"):
            self.var_phi_eff.set(f"{phi:.3f}" if phi and phi > 0 else "0.000")
        self._phi_eff_source_v53 = src
        self._h0_auto_v54 = _v59_estimate_h0_for_app(self)
    except Exception as e:
        self._phi_eff_source_v53 = f"φef automático não calculado: {e}"
        try:
            self.var_phi_eff.set("0.000")
        except Exception:
            pass
    return None

ColumnsEC2App.validate_inputs = _validate_inputs_v59

def _v59_hide_phi_controls(app):
    """Remove controlos editáveis de φef/h0, mas mantém RH e t0."""
    try:
        def walk(w):
            for c in w.winfo_children():
                yield c
                yield from walk(c)
        # renomear label frame de fluência
        for w in list(walk(app)):
            try:
                if isinstance(w, ttk.LabelFrame):
                    txt = str(w.cget("text"))
                    if "φef" in txt or "phi" in txt.lower():
                        w.configure(text="Fluência")
            except Exception:
                pass
        # destruir widgets explícitos de φef/h0
        for w in list(walk(app)):
            try:
                txt = str(w.cget("text")) if hasattr(w, "cget") else ""
                if ("φef" in txt or "h0" in txt or "h₀" in txt or "automaticamente" in txt.lower()):
                    w.destroy()
            except Exception:
                pass
        # esconder entries associados a var_phi_eff e var_creep_h0
        for w in list(walk(app)):
            try:
                if isinstance(w, (ttk.Entry, tk.Entry)) and str(w.cget("textvariable")) in [str(getattr(app, "var_phi_eff", "")), str(getattr(app, "var_creep_h0", ""))]:
                    w.destroy()
            except Exception:
                pass
    except Exception:
        pass

def _build_sidebar_v59(self, parent):
    # variáveis antes da UI
    if not hasattr(self, "var_code_backend"):
        self.var_code_backend = tk.StringVar(value=BACKEND_DISPLAY_PT_V59)
    if not hasattr(self, "var_creep_RH"):
        self.var_creep_RH = tk.StringVar(value="70")
    if not hasattr(self, "var_creep_t0"):
        self.var_creep_t0 = tk.StringVar(value="28")
    if not hasattr(self, "var_creep_h0"):
        self.var_creep_h0 = tk.StringVar(value="0")
    if not hasattr(self, "var_phi_eff_auto"):
        self.var_phi_eff_auto = tk.BooleanVar(value=True)
    # Construir UI antiga
    _old_build_sidebar_v53_base(self, parent) if "_old_build_sidebar_v53_base" in globals() else _old_build_sidebar_v31(self, parent)

    # backend com rótulos limpos
    try:
        self.var_code_backend.set(_backend_display_v59(getattr(self, "var_code_backend", tk.StringVar(value=BACKEND_DISPLAY_PT_V59)).get()))
    except Exception:
        pass

    def walk(w):
        for c in w.winfo_children():
            yield c
            yield from walk(c)

    # Labels/combobox: aço como classe
    for w in list(walk(parent)):
        try:
            txt = str(w.cget("text"))
            if txt == "Aço fyk [MPa]" or "Aço fyk" in txt:
                w.configure(text="Classe de Aço")
        except Exception:
            pass
        try:
            if isinstance(w, ttk.Combobox) and str(w.cget("textvariable")) == str(self.var_fyk):
                w.configure(values=["A400", "A500"], state="readonly", width=14)
                if str(self.var_fyk.get()).strip() in ["400", "A400"]:
                    self.var_fyk.set("A400")
                else:
                    self.var_fyk.set("A500")
        except Exception:
            pass
        try:
            if isinstance(w, ttk.Combobox) and str(w.cget("textvariable")) == str(self.var_code_backend):
                w.configure(values=BACKEND_DISPLAY_CHOICES_V59, state="readonly", width=34)
        except Exception:
            pass
        try:
            if isinstance(w, ttk.Combobox) and str(w.cget("textvariable")) == str(self.var_calc_mode):
                w.configure(values=["Pré-dimensionamento", "Dimensionamento"], state="readonly")
        except Exception:
            pass

    # Remove painel permanente de correcção interactiva.
    for w in list(walk(parent)):
        try:
            if isinstance(w, ttk.LabelFrame) and "Correc" in str(w.cget("text")):
                w.destroy()
        except Exception:
            pass

    # criar/actualizar painel de fluência limpo se não existir
    has_creep = False
    for w in walk(parent):
        try:
            if isinstance(w, ttk.LabelFrame) and str(w.cget("text")) == "Fluência":
                has_creep = True
        except Exception:
            pass
    if not has_creep:
        creep = ttk.LabelFrame(parent, text="Fluência")
        creep.pack(fill="x", pady=(0, 8))
        ttk.Label(creep, text="Humidade relativa RH [%]").grid(row=0, column=0, sticky="w", padx=6, pady=2)
        ttk.Entry(creep, textvariable=self.var_creep_RH, width=8).grid(row=0, column=1, sticky="ew", padx=6, pady=2)
        ttk.Label(creep, text="Idade do betão t0 [dias]").grid(row=1, column=0, sticky="w", padx=6, pady=2)
        ttk.Entry(creep, textvariable=self.var_creep_t0, width=8).grid(row=1, column=1, sticky="ew", padx=6, pady=2)
        ttk.Label(creep, text="h0 e φef são calculados automaticamente.", style="Subtle.TLabel").grid(row=2, column=0, columnspan=2, sticky="w", padx=6, pady=(0,4))
        creep.columnconfigure(1, weight=1)

    _v59_hide_phi_controls(self)

    # Notas rápidas dinâmicas limpas
    try:
        if not hasattr(self, "quick_notes_var"):
            self.quick_notes_var = tk.StringVar()
        self.var_code_backend.trace_add("write", lambda *_: _v59_update_quick_notes(self))
        self.var_calc_mode.trace_add("write", lambda *_: _v59_update_quick_notes(self))
        _v59_update_quick_notes(self)
    except Exception:
        pass

ColumnsEC2App._build_sidebar = _build_sidebar_v59

_old_build_ui_v59_base = ColumnsEC2App._build_ui
def _build_ui_v59(self):
    _old_build_ui_v59_base(self)
    try:
        self.title(f"{APP_NAME} {APP_VERSION}")
        # corrigir combobox do backend depois de todos os patches antigos alterarem valores
        def walk(w):
            yield w
            for c in w.winfo_children():
                yield from walk(c)
        for w in walk(self):
            try:
                if isinstance(w, ttk.Combobox) and str(w.cget("textvariable")) == str(self.var_code_backend):
                    w.configure(values=BACKEND_DISPLAY_CHOICES_V59)
                    self.var_code_backend.set(_backend_display_v59(self.var_code_backend.get()))
                if isinstance(w, ttk.Combobox) and str(w.cget("textvariable")) == str(self.var_fyk):
                    w.configure(values=["A400", "A500"])
                    self.var_fyk.set("A400" if "400" in str(self.var_fyk.get()) else "A500")
            except Exception:
                pass
        _v59_hide_phi_controls(self)
    except Exception:
        pass

ColumnsEC2App._build_ui = _build_ui_v59

def _v59_update_quick_notes(app):
    try:
        b = _v59_get_backend(app)
        mode = _v53_mode_to_label(getattr(app, "var_calc_mode", tk.StringVar(value="Dimensionamento")).get())
        if b == BACKEND_EC2_PT_2010:
            txt = (
                f"Norma: {_backend_display_v59(b)}\n"
                f"Modo: {mode}\n"
                "Motor interno completo: ELU, 2.ª ordem, interacção N-My-Mz, V, T, ELS e pormenorização. "
                "A correcção iterativa gera propostas quando houver falhas bloqueantes."
            )
        elif b == BACKEND_SC_EC2_2004:
            txt = (
                "Norma: Eurocode 2:2004\n"
                "Motor: structuralcodes, em modo strict. Calcula apenas os módulos expostos pela API local: materiais, secções, N-My-Mz e V/ELS quando disponíveis."
            )
        elif b == BACKEND_SC_EC2_2023:
            txt = (
                "Norma: Eurocode 2:2023\n"
                "Motor: structuralcodes, em modo strict. Materiais, secções, N-My-Mz, ELS/fendilhação/deformação quando expostos. Sem fallback interno."
            )
        else:
            txt = (
                "Norma: fib Model Code 2010\n"
                "Motor: structuralcodes, em modo strict. Materiais, secções, V/T/ELS são tentados apenas pela API MC2010 disponível."
            )
        if hasattr(app, "quick_notes_var"):
            app.quick_notes_var.set(txt)
    except Exception:
        pass

def _v59_failure_rows(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    st = df.get("status", pd.Series(index=df.index, dtype=str)).astype(str)
    sev = df.get("failure_severity", pd.Series(index=df.index, dtype=str)).astype(str)
    return df[(st == "Falha") | (sev == "Bloqueante")].copy()

def _v59_failure_summary_text(df: pd.DataFrame, max_rows=12):
    if df is None or df.empty:
        return ""
    lines = []
    for _, r in df.head(max_rows).iterrows():
        pr = str(r.get("prumada", r.get("name", r.get("member", ""))))
        typ = str(r.get("failure_type", ""))
        reason = str(r.get("failure_reason", "")) or str(r.get("failure_summary", ""))
        if len(reason) > 180:
            reason = reason[:177] + "..."
        lines.append(f"• {pr} | membro {r.get('member','')} | caso {r.get('case','')} | {typ}: {reason}")
    if len(df) > max_rows:
        lines.append(f"... + {len(df)-max_rows} ocorrências.")
    return "\n".join(lines)

def _repair_failures_interactive_v59(self):
    """Correcção interactiva controlada: primeiro gera propostas; só aplica após confirmação."""
    if self.df_results is None or self.df_results.empty:
        messagebox.showwarning("Aviso", "Execute primeiro o cálculo.")
        return
    backend = _v59_get_backend(self)
    if _sc_backend_active_v52(backend):
        failures = _v59_failure_rows(self.df_results)
        if failures.empty:
            messagebox.showinfo("Verificação", "Não há falhas bloqueantes a corrigir neste backend.")
        else:
            messagebox.showwarning(
                "Falhas no backend structuralcodes",
                "Foram detectadas falhas no backend seleccionado.\n\n"
                + _v59_failure_summary_text(failures, max_rows=10)
                + "\n\nComo o modo é strict, o ColumnsEC2 não aplica correcções por fórmulas internas. "
                  "Reveja a geometria, armadura candidata, esforços ou use o backend Portugal para propostas internas."
            )
        return

    targets = _v59_failure_rows(self.df_results)
    if targets.empty:
        messagebox.showinfo("Correcção iterativa", "Não foram detectadas falhas bloqueantes.")
        return
    if not messagebox.askyesno(
        "Falhas bloqueantes detectadas",
        "O programa pode tentar gerar propostas de correcção para as falhas abaixo.\n\n"
        + _v59_failure_summary_text(targets, max_rows=10)
        + "\n\nGerar propostas de correcção?"
    ):
        return

    proposals = []
    total = len(targets)
    self.progress_var.set(0.0)
    for k, (idx, row) in enumerate(targets.iterrows(), start=1):
        self.progress_var.set(100.0 * k / max(total, 1))
        self.status_var.set(f"A gerar proposta de correcção... {k}/{total}")
        self.update_idletasks()
        try:
            repaired = _try_repair_result_v44(self, row)
        except Exception as e:
            repaired = dict(row)
            repaired["repair_result"] = "Sem proposta automática"
            repaired["repair_note"] = str(e)
        repaired["_target_index"] = idx
        proposals.append(repaired)
    prop_df = pd.DataFrame(proposals)
    self.df_repair_proposals = prop_df
    msg_ok = int((prop_df.get("status", pd.Series(dtype=str)).astype(str) == "OK").sum()) if not prop_df.empty else 0
    msg_warn = int((prop_df.get("status", pd.Series(dtype=str)).astype(str) == "Aviso").sum()) if not prop_df.empty else 0
    msg_fail = len(prop_df) - msg_ok - msg_warn
    self.status_var.set(f"Propostas geradas: {msg_ok} OK; {msg_warn} avisos; {msg_fail} sem solução automática.")
    if not messagebox.askyesno(
        "Propostas de correcção geradas",
        f"Propostas geradas:\nOK: {msg_ok}\nAvisos: {msg_warn}\nSem solução automática: {msg_fail}\n\n"
        "Aplicar automaticamente apenas as propostas com estado OK?"
    ):
        return

    res = self.df_results.copy()
    applied = 0
    for _, p in prop_df.iterrows():
        idx = p.get("_target_index", None)
        if idx in res.index and str(p.get("status", "")) == "OK":
            for c in prop_df.columns:
                if c != "_target_index":
                    res.at[idx, c] = p.get(c)
            res.at[idx, "auto_repair_applied"] = "Sim"
            applied += 1
    try:
        res = enrich_failures_v43(res)
    except Exception:
        pass
    self.df_results = res
    self.df_summary = self.build_summary_by_member(res) if getattr(self, "var_summary", tk.BooleanVar(value=True)).get() else pd.DataFrame()
    self.df_failures = _v59_failure_rows(res)
    self.df_ok = res[res.get("status", pd.Series(index=res.index, dtype=str)).astype(str).eq("OK")].copy()
    for tree, df in [(self.tree_results, self.df_results), (self.tree_summary, self.df_summary), (self.tree_failures, self.df_failures), (self.tree_shortlists, self.build_shortlists_df())]:
        try:
            self.show_df(tree, df)
        except Exception:
            pass
    self.update_report()
    self.status_var.set(f"Correcção aplicada: {applied} propostas OK.")
    messagebox.showinfo("Correcção aplicada", f"Foram aplicadas {applied} propostas com estado OK.")

ColumnsEC2App.repair_failures_interactive = _repair_failures_interactive_v59

def _run_design_v59(self):
    err = self.validate_inputs()
    if err:
        messagebox.showwarning("Aviso", err)
        return
    backend = _v59_get_backend(self)
    globals()["ACTIVE_CODE_BACKEND_V48"] = backend
    mode_internal = _v53_mode_to_internal(self.var_calc_mode.get())
    try:
        self.var_calc_mode.set(_v53_mode_to_label(mode_internal))
    except Exception:
        pass
    designer = ColumnDesigner(
        cover_mm=safe_float(self.var_cover.get(), 35.0),
        fyk=safe_float(self.var_fyk.get(), 500.0),
        phi_eff=safe_float(getattr(self, "var_phi_eff", tk.StringVar(value="0")).get(), 0.0),
        l0y_factor=safe_float(self.var_l0y.get(), 1.0),
        l0z_factor=safe_float(self.var_l0z.get(), 1.0),
        calc_mode=mode_internal,
        code_backend=backend,
    )
    designer.service_case_override = self.var_service_case.get().strip() if hasattr(self, "var_service_case") else ""
    if getattr(self, "var_reduce_cases", tk.BooleanVar(value=True)).get():
        input_df = _v53_reduced_envelope(self.df_pair, mode=mode_internal)
    else:
        input_df = self.df_pair.copy()
    self.df_calc_input = input_df.copy()
    self.progress_var.set(0.0)
    self.status_var.set(f"A calcular — {_backend_display_v59(backend)}...")

    def progress(done, total):
        pct = 0.0 if total <= 0 else 100.0 * done / total
        self.after(0, lambda p=pct: self.progress_var.set(p))
        self.after(0, lambda d=done, t=total: self.status_var.set(f"A calcular... {d}/{t} casos de envolvente"))

    def worker():
        try:
            results = designer.design_dataframe(input_df, progress_callback=progress)
            try:
                results = apply_service_combination_override_v4(self, results, input_df)
            except Exception:
                pass
            try:
                results = enrich_failures_v43(results)
            except Exception:
                pass
            summary = self.build_summary_by_member(results) if getattr(self, "var_summary", tk.BooleanVar(value=True)).get() else pd.DataFrame()
            failures = _v59_failure_rows(results)
            ok = results[results.get("status", pd.Series(index=results.index, dtype=str)).astype(str).eq("OK")].copy() if not results.empty else pd.DataFrame()
            validation = self.build_data_validation(pre_calc=False)

            def finish():
                self.df_results = results
                self.df_summary = summary
                self.df_failures = failures
                self.df_ok = ok
                self.df_filtered = pd.DataFrame()
                self.df_validation = validation
                self.df_notes = self.build_normative_notes()
                for tree, df in [(self.tree_results, self.df_results), (self.tree_summary, self.df_summary), (self.tree_failures, self.df_failures), (self.tree_shortlists, self.build_shortlists_df()), (self.tree_validation, self.df_validation), (self.tree_notes, self.df_notes)]:
                    try:
                        self.show_df(tree, df)
                    except Exception:
                        pass
                try:
                    self.update_report()
                except Exception:
                    pass
                self.progress_var.set(100.0)
                prumadas = self.df_results.get("prumada", self.df_results.get("name", self.df_results.get("member", pd.Series(dtype=str)))).astype(str).nunique() if not self.df_results.empty else 0
                self.status_var.set(f"Cálculo concluído: {len(results)} casos; {prumadas} prumadas; {len(failures)} falhas bloqueantes.")
                _v59_update_quick_notes(self)
                if not failures.empty:
                    if _sc_backend_active_v52(backend):
                        messagebox.showwarning(
                            "Falhas no backend seleccionado",
                            "Foram detectadas falhas no backend seleccionado.\n\n"
                            + _v59_failure_summary_text(failures, max_rows=10)
                            + "\n\nNão foi aplicada correcção automática porque o modo structuralcodes é strict."
                        )
                    else:
                        if messagebox.askyesno(
                            "Falhas bloqueantes detectadas",
                            "Foram detectadas falhas bloqueantes.\n\n"
                            + _v59_failure_summary_text(failures, max_rows=10)
                            + "\n\nPretende gerar propostas de correcção?"
                        ):
                            self.repair_failures_interactive()
            self.after(0, finish)
        except Exception as err:
            msg = str(err)
            self.after(0, lambda m=msg: messagebox.showerror("Erro", m))
            self.after(0, lambda: self.status_var.set("Falha na análise."))
            self.after(0, lambda: self.progress_var.set(0.0))
    self.analysis_thread = threading.Thread(target=worker, daemon=True)
    self.analysis_thread.start()

ColumnsEC2App.run_design = _run_design_v59

# ---------- Metadados, parâmetros e notas ----------
def _module_description_v59(app=None) -> str:
    b = _v59_get_backend(app)
    if b == BACKEND_EC2_PT_2010:
        return (
            "Módulo Portugal: dimensiona e verifica pilares de betão armado segundo a NP EN 1992-1-1:2010, "
            "com AC:2012 e A1:2019. Inclui selecção de combinações governantes, segunda ordem por curvatura nominal, "
            "interacção N-My-Mz, esforço transverso, torção, ELS simplificado/por combinação indicada, pormenorização "
            "e geração de quadro de pilares/DXF."
        )
    if b == BACKEND_SC_EC2_2004:
        return (
            "Módulo Eurocode 2:2004: usa o pacote structuralcodes em modo strict para materiais, criação da secção armada, "
            "domínio resistente N-My-Mz e verificações EC2:2004 expostas pela API local, nomeadamente esforço transverso "
            "e fendilhação quando disponíveis."
        )
    if b == BACKEND_SC_EC2_2023:
        return (
            "Módulo Eurocode 2:2023: usa o pacote structuralcodes em modo strict para materiais, criação da secção armada, "
            "domínio resistente N-My-Mz e verificações de ELS/fendilhação/deformação quando expostas pela API local."
        )
    return (
        "Módulo fib Model Code 2010: usa o pacote structuralcodes em modo strict para materiais, secções, domínio N-My-Mz "
        "e verificações MC2010 de esforço transverso, torção, ELS, fluência/retracção e aderência quando expostas pela API local."
    )

def _module_limitations_v59(app=None) -> str:
    b = _v59_get_backend(app)
    if b == BACKEND_EC2_PT_2010:
        return (
            "Ferramenta de apoio ao projecto; os pilares críticos devem ser revistos. A resistência N-My-Mz usa discretização "
            "numérica; a pormenorização deve ser confirmada em desenho final. A qualidade dos resultados depende da coerência "
            "dos esforços, sinais locais, comprimentos efectivos e dados geométricos importados."
        )
    if b == BACKEND_SC_EC2_2004:
        return (
            "Modo strict: não usa fórmulas internas como fallback. A cobertura depende da versão instalada do structuralcodes; "
            "verificações não expostas pela API são reportadas como não avaliadas. Pormenorização normativa e 2.ª ordem não são "
            "substituídas pelo motor Portugal."
        )
    if b == BACKEND_SC_EC2_2023:
        return (
            "Modo strict: não usa fórmulas internas como fallback. A cobertura EC2:2023 depende da API structuralcodes instalada; "
            "esforço transverso, torção, 2.ª ordem e pormenorização podem ficar não avaliados se não estiverem expostos."
        )
    return (
        "Modo strict: não usa fórmulas internas como fallback. A cobertura MC2010 depende da API structuralcodes instalada; "
        "qualquer verificação indisponível é assinalada como não avaliada. A geometria de armadura é candidata e deve ser "
        "confirmada em pormenorização de projecto."
    )

def _metadata_df_v59(self) -> pd.DataFrame:
    return pd.DataFrame([
        ["Programa", APP_NAME],
        ["Autor", APP_AUTHOR],
        ["Repositório", GITHUB_URL],
        ["Data de exportação", datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["Ficheiro de origem", self.input_file_path or "-"],
        ["Norma de referência", _backend_reference_v59(self)],
        ["Motor de cálculo", _backend_display_v59(_v59_get_backend(self))],
        ["Modo de cálculo", _v53_mode_to_label(getattr(self, "var_calc_mode", tk.StringVar(value="Dimensionamento")).get())],
        ["Descrição", _module_description_v59(self)],
        ["Limitações", _module_limitations_v59(self)],
    ], columns=["Campo", "Valor"])

ColumnsEC2App._metadata_df = _metadata_df_v59

def _parameters_df_v59(self) -> pd.DataFrame:
    phi, src = _v59_compute_phi_eff_for_app(self)
    try:
        self._phi_eff_source_v53 = src
        self._h0_auto_v54 = _v59_estimate_h0_for_app(self)
    except Exception:
        pass
    return pd.DataFrame([
        ["Recobrimento nominal [mm]", self.var_cover.get()],
        ["Classe de Aço", _v59_steel_class_label(self)],
        ["fyk [MPa]", safe_float(self.var_fyk.get(), 500.0)],
        ["Classes de betão lidas da tabela", _v59_material_classes(self)],
        ["Humidade relativa RH [%]", getattr(self, "var_creep_RH", tk.StringVar(value="70")).get()],
        ["Idade do betão t0 [dias]", getattr(self, "var_creep_t0", tk.StringVar(value="28")).get()],
        ["h0 automático [mm]", f"{_v59_estimate_h0_for_app(self):.0f}"],
        ["φef automático", f"{phi:.3f}" if phi and phi > 0 else "não aplicado"],
        ["Origem / nota φef", src],
        ["l0y/L", self.var_l0y.get()],
        ["l0z/L", self.var_l0z.get()],
        ["Redução para casos governantes", "Sim" if getattr(self, "var_reduce_cases", tk.BooleanVar(value=True)).get() else "Não"],
        ["Combinação ELS indicada", getattr(self, "var_service_case", tk.StringVar(value="")).get().strip() or "não indicada"],
    ], columns=["Parâmetro", "Valor"])

ColumnsEC2App._parameters_df = _parameters_df_v59

def _build_normative_notes_v59(self) -> pd.DataFrame:
    b = _v59_get_backend(self)
    rows = []
    if b == BACKEND_EC2_PT_2010:
        rows = [
            ("Base normativa", "NP EN 1992-1-1:2010 + AC:2012 + A1:2019", "Aplicação com Anexo Nacional português. O módulo é orientado para pilares de edifícios."),
            ("Esforços", "Entrada por combinações", "Os esforços importados devem corresponder às combinações de análise. O programa selecciona casos reais governantes por prumada/member."),
            ("Eixo local", "X longitudinal", "FX é tratado como esforço normal; FY/FZ como esforços transversos; MX como torção; MY/MZ como momentos de flexão."),
            ("2.ª ordem", "Método da curvatura nominal", "São avaliados λ, λlim, M0e, imperfeição geométrica, M2 e momentos finais de cálculo por direcção."),
            ("Fluência", "φef automático", "φef e h0 são calculados automaticamente a partir de RH, t0, geometria e classe de betão; não são editáveis na GUI."),
            ("Resistência", "Interacção N-My-Mz", "A verificação resistente usa catálogo construtivo de armaduras e domínio resistente discreto para o nível de NEd."),
            ("Pormenorização", "Layouts construtivos", "As soluções privilegiam varões nos cantos e varões de face coerentes com a geometria do pilar; avisos construtivos são separados de falhas bloqueantes."),
            ("Esforço transverso", "EC2 6.2", "São avaliados VRd,c, VRd,max e necessidade de armadura transversal; VEd>VRd,c não é falha se VRd,max verificar."),
            ("Torção", "EC2 6.3", "Torção desprezável é classificada como não condicionante; só TEd>TRd,max deve originar falha resistente."),
            ("ELS", "EC2 7", "Se o utilizador indicar combinação ELS, esta é usada; caso contrário o ELS é assinalado como simplificado."),
            ("Relatórios", "PDF/XLSX/DXF", "O PDF resume as decisões de projecto; o XLSX conserva a memória completa; o DXF contém quadro de pilares e secções-tipo."),
        ]
    elif b == BACKEND_SC_EC2_2004:
        rows = [
            ("Base normativa", "Eurocode 2:2004", "Backend calculado por structuralcodes, em modo strict."),
            ("Materiais", "structuralcodes.codes.ec2_2004", "Propriedades de betão e aço obtidas pelas funções expostas pelo pacote."),
            ("Resistência seccional", "structuralcodes.sections", "O domínio N-My-Mz é calculado pelo calculador de secções do pacote, quando disponível."),
            ("Esforço transverso", "EC2 2004 API", "VRdc/VRdmax são avaliados se a instalação local expuser as funções correspondentes."),
            ("ELS/fendilhação", "EC2 2004 API", "Fendilhação/ELS são avaliados se a API local tiver os parâmetros necessários."),
            ("Regra strict", "Sem fallback", "Verificações não expostas pelo pacote são marcadas como não avaliadas, não substituídas pelo motor Portugal."),
        ]
    elif b == BACKEND_SC_EC2_2023:
        rows = [
            ("Base normativa", "Eurocode 2:2023", "Backend calculado por structuralcodes, em modo strict."),
            ("Materiais", "structuralcodes.codes.ec2_2023", "Propriedades de betão e aço obtidas pelas funções expostas pelo pacote."),
            ("Resistência seccional", "structuralcodes.sections", "O domínio N-My-Mz é calculado pelo calculador de secções do pacote, quando disponível."),
            ("ELS/fendilhação/deformação", "EC2 2023 API", "Cálculos de serviço são avaliados apenas quando a API local expõe as funções e parâmetros necessários."),
            ("V/T/2.ª ordem", "Âmbito da API", "Se esforço transverso, torção ou 2.ª ordem não estiverem expostos no pacote, ficam como não avaliados."),
            ("Regra strict", "Sem fallback", "Não são usadas fórmulas internas para completar módulos não disponibilizados pelo structuralcodes."),
        ]
    else:
        rows = [
            ("Base normativa", "fib Model Code 2010", "Backend calculado por structuralcodes, em modo strict."),
            ("Materiais", "structuralcodes.codes.mc2010", "Propriedades de betão e aço obtidas pelas funções expostas pelo pacote."),
            ("Resistência seccional", "structuralcodes.sections", "O domínio N-My-Mz é calculado pelo calculador de secções do pacote, quando disponível."),
            ("Esforço transverso", "MC2010 API", "Verificações de V são tentadas apenas por funções MC2010 expostas no pacote."),
            ("Torção", "MC2010 API", "Verificações de T são tentadas apenas por funções MC2010 expostas no pacote."),
            ("ELS/fluência/retracção/aderência", "MC2010 API", "A cobertura depende da versão instalada do structuralcodes e dos parâmetros disponíveis."),
            ("Regra strict", "Sem fallback", "Módulos não expostos pelo pacote ficam assinalados como não avaliados."),
        ]
    return pd.DataFrame(rows, columns=["Tema", "Referência", "Nota"])

ColumnsEC2App.build_normative_notes = _build_normative_notes_v59

def _backend_coverage_df_v59(app=None) -> pd.DataFrame:
    rows = [
        ["Materiais", BACKEND_DISPLAY_PT_V59, "ColumnsEC2", "Calculado"],
        ["2.ª ordem", BACKEND_DISPLAY_PT_V59, "ColumnsEC2", "Calculado"],
        ["N-My-Mz", BACKEND_DISPLAY_PT_V59, "ColumnsEC2", "Calculado"],
        ["Esforço transverso", BACKEND_DISPLAY_PT_V59, "ColumnsEC2", "Calculado"],
        ["Torção", BACKEND_DISPLAY_PT_V59, "ColumnsEC2", "Calculado"],
        ["ELS", BACKEND_DISPLAY_PT_V59, "ColumnsEC2", "Calculado/simplificado conforme combinação"],
        ["Pormenorização", BACKEND_DISPLAY_PT_V59, "ColumnsEC2", "Calculado"],
        ["Materiais", BACKEND_DISPLAY_EC2004_V59, "structuralcodes.codes.ec2_2004", "Calculado se API disponível"],
        ["N-My-Mz", BACKEND_DISPLAY_EC2004_V59, "structuralcodes.sections", "Calculado se API disponível"],
        ["Esforço transverso", BACKEND_DISPLAY_EC2004_V59, "structuralcodes.codes.ec2_2004", "Calculado se API disponível"],
        ["Torção", BACKEND_DISPLAY_EC2004_V59, "structuralcodes", "Não avaliado se não exposto"],
        ["ELS", BACKEND_DISPLAY_EC2004_V59, "structuralcodes.codes.ec2_2004", "Calculado se API/parâmetros disponíveis"],
        ["Materiais", BACKEND_DISPLAY_EC2023_V59, "structuralcodes.codes.ec2_2023", "Calculado se API disponível"],
        ["N-My-Mz", BACKEND_DISPLAY_EC2023_V59, "structuralcodes.sections", "Calculado se API disponível"],
        ["Esforço transverso", BACKEND_DISPLAY_EC2023_V59, "structuralcodes", "Não avaliado se não exposto"],
        ["Torção", BACKEND_DISPLAY_EC2023_V59, "structuralcodes", "Não avaliado se não exposto"],
        ["ELS/fendilhação/deformação", BACKEND_DISPLAY_EC2023_V59, "structuralcodes.codes.ec2_2023", "Calculado se API/parâmetros disponíveis"],
        ["Materiais", BACKEND_DISPLAY_MC2010_V59, "structuralcodes.codes.mc2010", "Calculado se API disponível"],
        ["N-My-Mz", BACKEND_DISPLAY_MC2010_V59, "structuralcodes.sections", "Calculado se API disponível"],
        ["Esforço transverso", BACKEND_DISPLAY_MC2010_V59, "structuralcodes.codes.mc2010", "Calculado se API disponível"],
        ["Torção", BACKEND_DISPLAY_MC2010_V59, "structuralcodes.codes.mc2010", "Calculado se API disponível"],
        ["ELS/fluência/aderência", BACKEND_DISPLAY_MC2010_V59, "structuralcodes.codes.mc2010", "Calculado se API/parâmetros disponíveis"],
    ]
    return pd.DataFrame(rows, columns=["Verificação", "Norma", "Origem", "Estado"])

def _sc_api_diagnostic_df_v59() -> pd.DataFrame:
    rows = []
    for b in [BACKEND_SC_EC2_2004, BACKEND_SC_EC2_2023, BACKEND_SC_MC2010]:
        sc, err = _sc_import_backend_v52(b)
        disp = _backend_display_v59(b)
        if sc is None:
            rows.append([disp, "structuralcodes", "Indisponível", str(err)])
            continue
        mod = sc["module"]
        for fn in ["fcd","fctm","Ecm","Eci","fyd","Es","VRdc","VRdmax","Asw_s_required","wk","wk_cal","delta_simpl","v_rd","v_rdc","v_rd_max_approx1","t_rd","t_rd_max","phi","creep_coefficient"]:
            rows.append([disp, fn, "Sim" if hasattr(mod, fn) else "Não", ""])
        rows.append([disp, "BeamSection", "Sim" if sc.get("BeamSection") is not None else "Não", "structuralcodes.sections"])
    return pd.DataFrame(rows, columns=["Norma", "Objecto/API", "Disponível", "Nota"])

def _write_excel_v59(self, path: str):
    """Exportação limpa, sem folhas redundantes v5.3/v5.4 e com nomes dinâmicos por backend."""
    self.status_var.set("A exportar XLSX...")
    self.progress_var.set(10.0)
    self.update_idletasks()
    res = self.df_results if self.df_results is not None else pd.DataFrame()
    notes_name = "13_Notas_FIB_MC10" if _v59_get_backend(self) == BACKEND_SC_MC2010 else "13_Notas_EC2"

    def cols(df, names):
        if df is None or df.empty:
            return pd.DataFrame()
        present = [c for c in names if c in df.columns]
        return df[present].copy() if present else pd.DataFrame()

    sheets = {
        "00_Info": self._metadata_df(),
        "01_Parametros": self._parameters_df(),
        "02_Entrada_Dados": self.df_clean,
        "03_Pares_Member_Case": self.df_pair,
        "04_Qualidade_Importacao": getattr(self, "df_import_quality", build_import_quality_cases_v41(self) if "build_import_quality_cases_v41" in globals() else pd.DataFrame()),
        "05_Casos_Calculo": self.df_calc_input,
        "06_Resultados": res,
        "07_Resumo_Membros": self.df_summary,
        "08_Falhas": self.df_failures,
        "09_OK": self.df_ok,
        "10_Shortlists": self.build_shortlists_df(),
        "11_Validacao": self.df_validation,
        "12_Materiais_Assumidos": material_assumptions_df_v41(res) if "material_assumptions_df_v41" in globals() else pd.DataFrame(),
        notes_name: self.df_notes if self.df_notes is not None and not self.df_notes.empty else self.build_normative_notes(),
        "14_ELS": cols(res, ["prumada","member","case","combination_number","limit_state","service_status","service_case_source","service_combination","service_sigma_c_max_MPa","service_sigma_c_min_MPa","service_sigma_s_max_MPa","service_wk_est_mm","service_wk_lim_mm","service_crack_status","service_method","service_note"]),
        "15_Esf_Transverso": cols(res, ["prumada","member","case","v_ed_y_kN","v_rd_c_y_kN","v_rd_max_y_kN","asw_s_y_req_mm2_per_m","shear_status_y","v_ed_z_kN","v_rd_c_z_kN","v_rd_max_z_kN","asw_s_z_req_mm2_per_m","shear_status_z","shear_backend"]),
        "16_Torcao": cols(res, ["prumada","member","case","mx_ed_kNm","torsion_ratio","t_rd_max_kNm","asw_s_t_req_mm2_per_m","asl_t_req_mm2","torsion_status","torsion_backend"]),
        "17_Pormenorizacao": cols(res, ["prumada","member","case","layout_description","solucao","phi_long_mm","n_total","n_bars_y","n_bars_z","phi_st_mm","s_st_mm","detailing_status","detailing_blocking_issues","detailing_warnings","detailing_info","detailing_issues","detailing_min_clear_mm"]),
        "18_Superficie_Resumo": cols(res, ["prumada","member","case","n_ed_kN","my_ed_kNm","mz_ed_kNm","mrd_y_kNm","mrd_z_kNm","utilizacao","biaxial_alpha","biaxial_n_ratio","surface_method","surface_points","nmm_capacity_source"]),
        "19_Superficie_Pontos": surface_points_df_v41(res) if "surface_points_df_v41" in globals() else pd.DataFrame(),
        "20_Memoria_Calculo": cols(res, ["prumada","member","case","material","b_cm","h_cm","length_m","n_ed_kN","my_i_kNm","my_j_kNm","mz_i_kNm","mz_j_kNm","m01_y_ec2_kNm","m02_y_ec2_kNm","rm_y_signed","curvature_y","m01_z_ec2_kNm","m02_z_ec2_kNm","rm_z_signed","curvature_z","lambda_y","lambda_lim_y","lambda_z","lambda_lim_z","m0e_y_kNm","m2_y_kNm","m0e_z_kNm","m2_z_kNm","as_min_mm2","as_req_mm2","as_max_mm2","as_prov_mm2","utilizacao","status","failure_reason"]),
        "22_Quadro_Pilares": build_pillar_schedule_df_v42(res) if "build_pillar_schedule_df_v42" in globals() else self.df_summary,
        "23_Resumo_Prumadas": self.df_summary,
        "24_Gestao_Falhas": cols(res, ["prumada","member","case","status","failure_severity","failure_type","failure_reason","failure_summary","design_decision","review_priority","failure_action","recommendations"]),
        "25_Propostas_Correcoes": getattr(self, "df_repair_proposals", pd.DataFrame()),
        "30_Mapa_Cobertura": _backend_coverage_df_v59(self),
        "31_Diagnostico_API": _sc_api_diagnostic_df_v59(),
        "32_Escopo_Resultados": cols(res, ["prumada","member","case","code_backend","normative_basis","materials_backend","materials_sources","nmm_capacity_source","shear_backend","torsion_backend","service_backend","second_order_status","detailing_status","status","failure_reason"]),
    }

    # remover folhas explicitamente indesejadas se algum nome antigo entrar por engano
    remove_names = {"21_Casos_Tipo_Validacao", "33_Opcoes_v5_3", "34_Cobertura_Backend", "35_Diagnostico_SC", "36_Parametros_v5_4", "13_Notas_FIB_MC10" if notes_name == "13_Notas_EC2" else "13_Notas_EC2"}

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in sheets.items():
            if name in remove_names:
                continue
            if df is None:
                df = pd.DataFrame()
            # Evitar duplicados de sheet name e respeitar limite de 31 caracteres.
            df.to_excel(writer, sheet_name=name[:31], index=False)
        wb = writer.book
        props = wb.properties
        props.title = APP_NAME
        props.subject = APP_SUBJECT
        props.creator = APP_AUTHOR
        props.keywords = APP_KEYWORDS
        props.category = APP_CATEGORY
        props.description = APP_XLSX_DESCRIPTION
        props.lastModifiedBy = APP_AUTHOR
        try:
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            from openpyxl.utils import get_column_letter
            header_fill = PatternFill("solid", fgColor="1F4E5F")
            header_font = Font(color="FFFFFF", bold=True)
            thin = Side(style="thin", color="D9E2E7")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for ws in wb.worksheets:
                ws.sheet_view.showGridLines = False
                ws.freeze_panes = "A2"
                if ws.max_row >= 1:
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        cell.border = border
                for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 5000)):
                    for cell in row:
                        cell.border = border
                        cell.alignment = Alignment(vertical="top", wrap_text=True)
                for col_idx, col in enumerate(ws.columns, start=1):
                    values = [str(c.value) for c in col[:200] if c.value is not None]
                    width = min(max([len(v) for v in values] + [10]) + 2, 55)
                    ws.column_dimensions[get_column_letter(col_idx)].width = width
        except Exception:
            pass
    self.progress_var.set(100.0)
    self.status_var.set(f"XLSX exportado: {path}")

ColumnsEC2App._write_excel = _write_excel_v59

# ---------- PDF profissional ----------
def _write_pdf_v59(self, path: str):
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak

    self.status_var.set("A gerar relatório PDF...")
    self.progress_var.set(10.0)
    self.update_idletasks()

    doc = SimpleDocTemplate(path, pagesize=landscape(A4), rightMargin=12*mm, leftMargin=12*mm, topMargin=12*mm, bottomMargin=12*mm)
    doc.title = APP_NAME
    doc.author = APP_AUTHOR
    doc.subject = APP_SUBJECT
    styles = getSampleStyleSheet()
    if "ReportTitle" not in styles:
        styles.add(ParagraphStyle(name="ReportTitle", parent=styles["Title"], alignment=TA_CENTER, fontName="Courier-Bold", fontSize=14, leading=21, spaceAfter=10))
    if "ReportSubtitle" not in styles:
        styles.add(ParagraphStyle(name="ReportSubtitle", parent=styles["Normal"], alignment=TA_CENTER, fontName="Courier", fontSize=10, leading=15, textColor=colors.darkgrey, spaceAfter=8))
    if "BodyCourier" not in styles:
        styles.add(ParagraphStyle(name="BodyCourier", parent=styles["Normal"], fontName="Courier", fontSize=9, leading=13.5, spaceAfter=5))
    if "Small" not in styles:
        styles.add(ParagraphStyle(name="Small", parent=styles["Normal"], fontName="Courier", fontSize=7.5, leading=10.5))
    if "Cell" not in styles:
        styles.add(ParagraphStyle(name="Cell", parent=styles["Small"], alignment=TA_LEFT, fontName="Courier", fontSize=6.7, leading=9.5))
    if "Section" not in styles:
        styles.add(ParagraphStyle(name="Section", parent=styles["Heading2"], fontName="Courier-Bold", fontSize=11, leading=16, spaceBefore=9, spaceAfter=10))

    def esc(v):
        if v is None or (isinstance(v, float) and not math.isfinite(v)):
            return ""
        s = str(v)
        return s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

    def tbl(data, widths=None, header=False, font_size=6.7):
        pdata = [[Paragraph(esc(c), styles["Cell"]) for c in row] for row in data]
        if widths is None:
            widths = [270*mm/max(1, len(data[0]))] * max(1, len(data[0]))
        t = Table(pdata, colWidths=widths, repeatRows=1 if header else 0)
        cmds = [
            ("GRID", (0,0), (-1,-1), 0.25, colors.lightgrey),
            ("VALIGN", (0,0), (-1,-1), "TOP"),
            ("FONTNAME", (0,0), (-1,-1), "Courier"),
            ("FONTSIZE", (0,0), (-1,-1), font_size),
            ("LEFTPADDING", (0,0), (-1,-1), 3),
            ("RIGHTPADDING", (0,0), (-1,-1), 3),
        ]
        if header:
            cmds += [("BACKGROUND", (0,0), (-1,0), colors.HexColor("#EFEFEF")), ("FONTNAME", (0,0), (-1,0), "Courier-Bold")]
        t.setStyle(TableStyle(cmds))
        return t

    def df_table(df, cols, max_rows=30, widths=None):
        if df is None or df.empty:
            return Paragraph("Sem dados.", styles["Small"])
        present = [c for c in cols if c in df.columns]
        if not present:
            return Paragraph("Sem colunas aplicáveis.", styles["Small"])
        data = [present]
        for _, r in df.head(max_rows).iterrows():
            row = []
            for c in present:
                v = r.get(c, "")
                if isinstance(v, float):
                    row.append("" if not math.isfinite(v) else f"{v:.2f}")
                else:
                    row.append("" if pd.isna(v) else str(v))
            data.append(row)
        return tbl(data, widths=widths, header=True)

    res = self.df_results if self.df_results is not None else pd.DataFrame()
    summary = self.df_summary if self.df_summary is not None and not self.df_summary.empty else res
    failures = self.df_failures if self.df_failures is not None else pd.DataFrame()
    backend = _v59_get_backend(self)
    n_total = len(res)
    n_ok = int((res.get("status", pd.Series(dtype=str)).astype(str) == "OK").sum()) if not res.empty else 0
    n_warn = int((res.get("status", pd.Series(dtype=str)).astype(str) == "Aviso").sum()) if not res.empty else 0
    n_fail = int((res.get("status", pd.Series(dtype=str)).astype(str) == "Falha").sum()) if not res.empty else 0
    prumadas = res.get("prumada", res.get("name", res.get("member", pd.Series(dtype=str)))).astype(str).nunique() if not res.empty else 0

    story = []
    story.append(Paragraph(APP_NAME, styles["ReportTitle"]))
    story.append(Paragraph("Relatório técnico de dimensionamento/verificação de pilares", styles["ReportSubtitle"]))
    meta = [
        ["Norma", _backend_reference_v59(self), "Motor", _backend_display_v59(backend)],
        ["Modo", _v53_mode_to_label(_v59_getvar(self, "var_calc_mode", "Dimensionamento")), "Data", datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["Prumadas", prumadas, "Casos de cálculo", n_total],
        ["OK", n_ok, "Avisos", n_warn],
        ["Falhas", n_fail, "Aço", _v59_steel_class_label(self)],
        ["Betão", _v59_material_classes(self), "Recobrimento", f"{_v59_getvar(self, 'var_cover', '')} mm"],
    ]
    story.append(tbl(meta, widths=[34*mm, 105*mm, 34*mm, 95*mm]))
    story.append(Spacer(1, 4*mm))
    story.append(Paragraph("1. Síntese executiva", styles["Section"]))
    story.append(Paragraph(_module_description_v59(self), styles["BodyCourier"]))
    story.append(Paragraph("<b>Limitações:</b> " + esc(_module_limitations_v59(self)), styles["BodyCourier"]))
    story.append(Spacer(1, 3*mm))

    story.append(Paragraph("2. Parâmetros adoptados", styles["Section"]))
    story.append(df_table(self._parameters_df(), ["Parâmetro", "Valor"], max_rows=18, widths=[70*mm, 190*mm]))

    story.append(Paragraph("3. Resultados governantes por prumada", styles["Section"]))
    cols = ["prumada","member","case","material","b_cm","h_cm","n_ed_kN","my_ed_kNm","mz_ed_kNm","as_prov_mm2","solucao","utilizacao","status"]
    story.append(df_table(summary, cols, max_rows=36))

    if failures is not None and not failures.empty:
        story.append(PageBreak())
        story.append(Paragraph("4. Falhas bloqueantes e acções recomendadas", styles["Section"]))
        story.append(df_table(failures, ["prumada","member","case","failure_type","failure_reason","recommendations","design_decision"], max_rows=45))

    # Verificações complementares
    if res is not None and not res.empty:
        story.append(PageBreak())
        story.append(Paragraph("5. Verificações complementares", styles["Section"]))
        vt_cols = ["prumada","member","case","v_ed_y_kN","v_rd_max_y_kN","shear_status_y","v_ed_z_kN","v_rd_max_z_kN","shear_status_z","mx_ed_kNm","t_rd_max_kNm","torsion_status"]
        story.append(df_table(res, vt_cols, max_rows=34))
        story.append(Spacer(1, 4*mm))
        story.append(Paragraph("6. Estados Limites de Serviço", styles["Section"]))
        els_cols = ["prumada","member","case","service_status","service_case_source","service_sigma_c_max_MPa","service_sigma_s_max_MPa","service_wk_est_mm","service_crack_status"]
        story.append(df_table(res, els_cols, max_rows=34))

    # Cobertura no caso structuralcodes
    if _sc_backend_active_v52(backend):
        story.append(PageBreak())
        story.append(Paragraph("7. Cobertura do backend seleccionado", styles["Section"]))
        cov = _backend_coverage_df_v59(self)
        cov = cov[cov["Norma"].astype(str) == _backend_display_v59(backend)]
        story.append(df_table(cov, ["Verificação","Norma","Origem","Estado"], max_rows=40))

    story.append(Spacer(1, 4*mm))
    story.append(Paragraph("Notas finais: o PDF apresenta a decisão técnica essencial. A memória completa, dados de entrada, shortlists, verificação de serviço, esforço transverso, torção, pormenorização, superfície resistente e diagnóstico de backend permanecem no ficheiro XLSX.", styles["Small"]))

    def footer(canvas, doc_obj):
        canvas.saveState()
        canvas.setAuthor(APP_AUTHOR)
        canvas.setTitle(APP_NAME)
        canvas.setSubject(APP_SUBJECT)
        canvas.setFont("Courier", 7)
        canvas.setFillColor(colors.grey)
        canvas.drawString(12*mm, 7*mm, f"{APP_NAME} {APP_VERSION} | {APP_AUTHOR}")
        canvas.drawRightString(285*mm, 7*mm, f"Página {doc_obj.page}")
        canvas.restoreState()

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    self.progress_var.set(100.0)
    self.status_var.set(f"PDF exportado: {path}")

ColumnsEC2App._write_pdf = _write_pdf_v59

# Export wrappers com barra de estado mantida.
def _export_excel_v59(self):
    if self.df_results is None or self.df_results.empty:
        messagebox.showwarning("Aviso", "Não há resultados para exportar.")
        return
    path = filedialog.asksaveasfilename(title="Exportar resultados", defaultextension=".xlsx", filetypes=[("Excel workbook", "*.xlsx")])
    if not path:
        return
    root, ext = os.path.splitext(path)
    if ext.lower() != ".xlsx":
        path = root + ".xlsx" if ext else path + ".xlsx"
    try:
        self.progress_var.set(5.0)
        self.status_var.set("A preparar XLSX...")
        self.update_idletasks()
        self._write_excel(path)
    except Exception as err:
        self.progress_var.set(0.0)
        messagebox.showerror("Erro", f"Não foi possível exportar XLSX.\n\n{err}")

ColumnsEC2App.export_excel = _export_excel_v59

def _export_pdf_report_v59(self):
    if self.df_results is None or self.df_results.empty:
        messagebox.showwarning("Aviso", "Não há resultados para exportar.")
        return
    path = filedialog.asksaveasfilename(title="Exportar relatório PDF", defaultextension=".pdf", filetypes=[("PDF", "*.pdf")])
    if not path:
        return
    root, ext = os.path.splitext(path)
    if ext.lower() != ".pdf":
        path = root + ".pdf" if ext else path + ".pdf"
    try:
        self.progress_var.set(5.0)
        self.status_var.set("A preparar PDF...")
        self.update_idletasks()
        self._write_pdf(path)
    except Exception as err:
        self.progress_var.set(0.0)
        messagebox.showerror("Erro", f"Não foi possível exportar PDF.\n\n{err}")

ColumnsEC2App.export_pdf_report = _export_pdf_report_v59

# Limpeza da sheet de notas no GUI para FIB
def _build_main_tabs_v59(self, parent):
    _old_build_main_tabs_v4(self, parent) if "_old_build_main_tabs_v4" in globals() else ColumnsEC2App._build_main_tabs(self, parent)
    try:
        # aplicar título dinâmico ao separador de notas, se existir
        pass
    except Exception:
        pass

# Não substituímos _build_main_tabs para evitar duplicar tabs; as folhas XLSX ficam dinâmicas.


# --- v5.9.1 minor robustness: avoid creating Tk variables in non-GUI helper calls ---
def _v59_getvar(obj, name, default=""):
    try:
        v = getattr(obj, name)
        return v.get() if hasattr(v, "get") else v
    except Exception:
        return default

def _v59_get_backend(app=None):
    return _backend_selected_v52(_v59_getvar(app, "var_code_backend", BACKEND_DISPLAY_PT_V59))

def _v59_steel_class_label(app=None) -> str:
    fyk = safe_float(_v59_getvar(app, "var_fyk", "A500"), 500.0)
    return f"A{int(round(fyk))}"

def _v59_compute_phi_eff_for_app(app=None):
    backend = _v59_get_backend(app)
    RH = safe_float(_v59_getvar(app, "var_creep_RH", "70"), 70.0)
    t0 = safe_float(_v59_getvar(app, "var_creep_t0", "28"), 28.0)
    h0 = _v59_estimate_h0_for_app(app)
    fck = 30.0
    try:
        mats = _v59_material_classes(app)
        m = re.search(r"C\s*(\d+)\s*/", mats)
        if m:
            fck = float(m.group(1))
    except Exception:
        pass
    if _sc_backend_active_v52(backend):
        try:
            val, src = _v53_try_structuralcodes_phi(backend, fck, RH, t0, h0)
            if val is not None and math.isfinite(float(val)) and float(val) > 0:
                return float(val), f"{src}; RH={RH:.0f}%; t0={t0:.0f} d; h0={h0:.0f} mm"
        except Exception:
            pass
        if backend in [BACKEND_SC_EC2_2004, BACKEND_SC_EC2_2023]:
            try:
                val = _v53_ec2_creep_phi(fck, RH=RH, t0=t0, h0_mm=h0)
                return float(val), (
                    f"φef estimado por expressão EC2 auxiliar porque a API de fluência não foi exposta; "
                    f"não usado nas verificações strict; RH={RH:.0f}%; t0={t0:.0f} d; h0={h0:.0f} mm"
                )
            except Exception:
                pass
        return 0.0, f"φef não disponível na API structuralcodes local para {_backend_display_v59(backend)}; não aplicado no backend strict; RH={RH:.0f}%; t0={t0:.0f} d; h0={h0:.0f} mm"
    try:
        val = _v53_ec2_creep_phi(fck, RH=RH, t0=t0, h0_mm=h0)
        return float(val), f"NP EN 1992-1-1: φef automático; RH={RH:.0f}%; t0={t0:.0f} d; h0={h0:.0f} mm; fck={fck:.0f} MPa"
    except Exception as e:
        return 2.0, f"φef automático não calculado ({e}); valor adoptado = 2.0"

def _metadata_df_v59(self) -> pd.DataFrame:
    return pd.DataFrame([
        ["Programa", APP_NAME],
        ["Autor", APP_AUTHOR],
        ["Repositório", GITHUB_URL],
        ["Data de exportação", datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["Ficheiro de origem", getattr(self, "input_file_path", "") or "-"],
        ["Norma de referência", _backend_reference_v59(self)],
        ["Motor de cálculo", _backend_display_v59(_v59_get_backend(self))],
        ["Modo de cálculo", _v53_mode_to_label(_v59_getvar(self, "var_calc_mode", "Dimensionamento"))],
        ["Descrição", _module_description_v59(self)],
        ["Limitações", _module_limitations_v59(self)],
    ], columns=["Campo", "Valor"])

ColumnsEC2App._metadata_df = _metadata_df_v59

def _parameters_df_v59(self) -> pd.DataFrame:
    phi, src = _v59_compute_phi_eff_for_app(self)
    try:
        self._phi_eff_source_v53 = src
        self._h0_auto_v54 = _v59_estimate_h0_for_app(self)
    except Exception:
        pass
    return pd.DataFrame([
        ["Recobrimento nominal [mm]", _v59_getvar(self, "var_cover", "")],
        ["Classe de Aço", _v59_steel_class_label(self)],
        ["fyk [MPa]", safe_float(_v59_getvar(self, "var_fyk", "A500"), 500.0)],
        ["Classes de betão lidas da tabela", _v59_material_classes(self)],
        ["Humidade relativa RH [%]", _v59_getvar(self, "var_creep_RH", "70")],
        ["Idade do betão t0 [dias]", _v59_getvar(self, "var_creep_t0", "28")],
        ["h0 automático [mm]", f"{_v59_estimate_h0_for_app(self):.0f}"],
        ["φef automático", f"{phi:.3f}" if phi and phi > 0 else "não aplicado"],
        ["Origem / nota φef", src],
        ["l0y/L", _v59_getvar(self, "var_l0y", "")],
        ["l0z/L", _v59_getvar(self, "var_l0z", "")],
        ["Redução para casos governantes", "Sim" if bool(_v59_getvar(self, "var_reduce_cases", True)) else "Não"],
        ["Combinação ELS indicada", str(_v59_getvar(self, "var_service_case", "")).strip() or "não indicada"],
    ], columns=["Parâmetro", "Valor"])

ColumnsEC2App._parameters_df = _parameters_df_v59

def _build_normative_notes_v59(self) -> pd.DataFrame:
    b = _v59_get_backend(self)
    # reutiliza a função anterior quando possível, mas substitui os textos de backend por rótulos limpos
    rows = []
    if b == BACKEND_EC2_PT_2010:
        rows = [
            ("Base normativa", "NP EN 1992-1-1:2010 + AC:2012 + A1:2019", "Aplicação com Anexo Nacional português. O módulo é orientado para pilares de edifícios."),
            ("Esforços", "Entrada por combinações", "Os esforços importados devem corresponder às combinações de análise. O programa selecciona casos reais governantes por prumada/member."),
            ("Eixo local", "X longitudinal", "FX é tratado como esforço normal; FY/FZ como esforços transversos; MX como torção; MY/MZ como momentos de flexão."),
            ("2.ª ordem", "Método da curvatura nominal", "São avaliados λ, λlim, M0e, imperfeição geométrica, M2 e momentos finais de cálculo por direcção."),
            ("Fluência", "φef automático", "φef e h0 são calculados automaticamente a partir de RH, t0, geometria e classe de betão; não são editáveis na GUI."),
            ("Resistência", "Interacção N-My-Mz", "A verificação resistente usa catálogo construtivo de armaduras e domínio resistente discreto para o nível de NEd."),
            ("Pormenorização", "Layouts construtivos", "As soluções privilegiam varões nos cantos e varões de face coerentes com a geometria do pilar; avisos construtivos são separados de falhas bloqueantes."),
            ("Esforço transverso", "EC2 6.2", "São avaliados VRd,c, VRd,max e necessidade de armadura transversal; VEd>VRd,c não é falha se VRd,max verificar."),
            ("Torção", "EC2 6.3", "Torção desprezável é classificada como não condicionante; só TEd>TRd,max deve originar falha resistente."),
            ("ELS", "EC2 7", "Se o utilizador indicar combinação ELS, esta é usada; caso contrário o ELS é assinalado como simplificado."),
            ("Relatórios", "PDF/XLSX/DXF", "O PDF resume as decisões de projecto; o XLSX conserva a memória completa; o DXF contém quadro de pilares e secções-tipo."),
        ]
    elif b == BACKEND_SC_EC2_2004:
        rows = [
            ("Base normativa", "Eurocode 2:2004", "Backend calculado por structuralcodes, em modo strict."),
            ("Materiais", "structuralcodes.codes.ec2_2004", "Propriedades de betão e aço obtidas pelas funções expostas pelo pacote."),
            ("Resistência seccional", "structuralcodes.sections", "O domínio N-My-Mz é calculado pelo calculador de secções do pacote, quando disponível."),
            ("Esforço transverso", "EC2 2004 API", "VRdc/VRdmax são avaliados se a instalação local expuser as funções correspondentes."),
            ("ELS/fendilhação", "EC2 2004 API", "Fendilhação/ELS são avaliados se a API local tiver os parâmetros necessários."),
            ("Regra strict", "Sem fallback", "Verificações não expostas pelo pacote são marcadas como não avaliadas, não substituídas pelo motor Portugal."),
        ]
    elif b == BACKEND_SC_EC2_2023:
        rows = [
            ("Base normativa", "Eurocode 2:2023", "Backend calculado por structuralcodes, em modo strict."),
            ("Materiais", "structuralcodes.codes.ec2_2023", "Propriedades de betão e aço obtidas pelas funções expostas pelo pacote."),
            ("Resistência seccional", "structuralcodes.sections", "O domínio N-My-Mz é calculado pelo calculador de secções do pacote, quando disponível."),
            ("ELS/fendilhação/deformação", "EC2 2023 API", "Cálculos de serviço são avaliados apenas quando a API local expõe as funções e parâmetros necessários."),
            ("V/T/2.ª ordem", "Âmbito da API", "Se esforço transverso, torção ou 2.ª ordem não estiverem expostos no pacote, ficam como não avaliados."),
            ("Regra strict", "Sem fallback", "Não são usadas fórmulas internas para completar módulos não disponibilizados pelo structuralcodes."),
        ]
    else:
        rows = [
            ("Base normativa", "fib Model Code 2010", "Backend calculado por structuralcodes, em modo strict."),
            ("Materiais", "structuralcodes.codes.mc2010", "Propriedades de betão e aço obtidas pelas funções expostas pelo pacote."),
            ("Resistência seccional", "structuralcodes.sections", "O domínio N-My-Mz é calculado pelo calculador de secções do pacote, quando disponível."),
            ("Esforço transverso", "MC2010 API", "Verificações de V são tentadas apenas por funções MC2010 expostas no pacote."),
            ("Torção", "MC2010 API", "Verificações de T são tentadas apenas por funções MC2010 expostas no pacote."),
            ("ELS/fluência/retracção/aderência", "MC2010 API", "A cobertura depende da versão instalada do structuralcodes e dos parâmetros disponíveis."),
            ("Regra strict", "Sem fallback", "Módulos não expostos pelo pacote ficam assinalados como não avaliados."),
        ]
    return pd.DataFrame(rows, columns=["Tema", "Referência", "Nota"])

ColumnsEC2App.build_normative_notes = _build_normative_notes_v59


# --- v5.9.2 safe XLSX export defaults ---
def _write_excel_v59(self, path: str):
    """Exportação limpa, sem folhas redundantes v5.3/v5.4 e com nomes dinâmicos por backend."""
    try:
        self.status_var.set("A exportar XLSX...")
        self.progress_var.set(10.0)
        self.update_idletasks()
    except Exception:
        pass
    res = self.df_results if getattr(self, "df_results", None) is not None else pd.DataFrame()
    notes_name = "13_Notas_FIB_MC10" if _v59_get_backend(self) == BACKEND_SC_MC2010 else "13_Notas_EC2"

    def cols(df, names):
        if df is None or df.empty:
            return pd.DataFrame()
        present = [c for c in names if c in df.columns]
        return df[present].copy() if present else pd.DataFrame()

    try:
        import_quality = getattr(self, "df_import_quality")
    except Exception:
        try:
            import_quality = build_import_quality_cases_v41(self) if "build_import_quality_cases_v41" in globals() else pd.DataFrame()
        except Exception:
            import_quality = pd.DataFrame()

    try:
        mat_assum = material_assumptions_df_v41(res) if "material_assumptions_df_v41" in globals() else pd.DataFrame()
    except Exception:
        mat_assum = pd.DataFrame()
    try:
        surf_points = surface_points_df_v41(res) if "surface_points_df_v41" in globals() else pd.DataFrame()
    except Exception:
        surf_points = pd.DataFrame()
    try:
        schedule = build_pillar_schedule_df_v42(res) if "build_pillar_schedule_df_v42" in globals() else getattr(self, "df_summary", pd.DataFrame())
    except Exception:
        schedule = getattr(self, "df_summary", pd.DataFrame())

    sheets = {
        "00_Info": self._metadata_df(),
        "01_Parametros": self._parameters_df(),
        "02_Entrada_Dados": getattr(self, "df_clean", pd.DataFrame()),
        "03_Pares_Member_Case": getattr(self, "df_pair", pd.DataFrame()),
        "04_Qualidade_Importacao": import_quality,
        "05_Casos_Calculo": getattr(self, "df_calc_input", pd.DataFrame()),
        "06_Resultados": res,
        "07_Resumo_Membros": getattr(self, "df_summary", pd.DataFrame()),
        "08_Falhas": getattr(self, "df_failures", pd.DataFrame()),
        "09_OK": getattr(self, "df_ok", pd.DataFrame()),
        "10_Shortlists": self.build_shortlists_df() if hasattr(self, "build_shortlists_df") else pd.DataFrame(),
        "11_Validacao": getattr(self, "df_validation", pd.DataFrame()),
        "12_Materiais_Assumidos": mat_assum,
        notes_name: getattr(self, "df_notes", pd.DataFrame()) if getattr(self, "df_notes", pd.DataFrame()) is not None and not getattr(self, "df_notes", pd.DataFrame()).empty else self.build_normative_notes(),
        "14_ELS": cols(res, ["prumada","member","case","combination_number","limit_state","service_status","service_case_source","service_combination","service_sigma_c_max_MPa","service_sigma_c_min_MPa","service_sigma_s_max_MPa","service_wk_est_mm","service_wk_lim_mm","service_crack_status","service_method","service_note"]),
        "15_Esf_Transverso": cols(res, ["prumada","member","case","v_ed_y_kN","v_rd_c_y_kN","v_rd_max_y_kN","asw_s_y_req_mm2_per_m","shear_status_y","v_ed_z_kN","v_rd_c_z_kN","v_rd_max_z_kN","asw_s_z_req_mm2_per_m","shear_status_z","shear_backend"]),
        "16_Torcao": cols(res, ["prumada","member","case","mx_ed_kNm","torsion_ratio","t_rd_max_kNm","asw_s_t_req_mm2_per_m","asl_t_req_mm2","torsion_status","torsion_backend"]),
        "17_Pormenorizacao": cols(res, ["prumada","member","case","layout_description","solucao","phi_long_mm","n_total","n_bars_y","n_bars_z","phi_st_mm","s_st_mm","detailing_status","detailing_blocking_issues","detailing_warnings","detailing_info","detailing_issues","detailing_min_clear_mm"]),
        "18_Superficie_Resumo": cols(res, ["prumada","member","case","n_ed_kN","my_ed_kNm","mz_ed_kNm","mrd_y_kNm","mrd_z_kNm","utilizacao","biaxial_alpha","biaxial_n_ratio","surface_method","surface_points","nmm_capacity_source"]),
        "19_Superficie_Pontos": surf_points,
        "20_Memoria_Calculo": cols(res, ["prumada","member","case","material","b_cm","h_cm","length_m","n_ed_kN","my_i_kNm","my_j_kNm","mz_i_kNm","mz_j_kNm","m01_y_ec2_kNm","m02_y_ec2_kNm","rm_y_signed","curvature_y","m01_z_ec2_kNm","m02_z_ec2_kNm","rm_z_signed","curvature_z","lambda_y","lambda_lim_y","lambda_z","lambda_lim_z","m0e_y_kNm","m2_y_kNm","m0e_z_kNm","m2_z_kNm","as_min_mm2","as_req_mm2","as_max_mm2","as_prov_mm2","utilizacao","status","failure_reason"]),
        "22_Quadro_Pilares": schedule,
        "23_Resumo_Prumadas": getattr(self, "df_summary", pd.DataFrame()),
        "24_Gestao_Falhas": cols(res, ["prumada","member","case","status","failure_severity","failure_type","failure_reason","failure_summary","design_decision","review_priority","failure_action","recommendations"]),
        "25_Propostas_Correcoes": getattr(self, "df_repair_proposals", pd.DataFrame()),
        "30_Mapa_Cobertura": _backend_coverage_df_v59(self),
        "31_Diagnostico_API": _sc_api_diagnostic_df_v59(),
        "32_Escopo_Resultados": cols(res, ["prumada","member","case","code_backend","normative_basis","materials_backend","materials_sources","nmm_capacity_source","shear_backend","torsion_backend","service_backend","second_order_status","detailing_status","status","failure_reason"]),
    }

    remove_names = {"21_Casos_Tipo_Validacao", "33_Opcoes_v5_3", "34_Cobertura_Backend", "35_Diagnostico_SC", "36_Parametros_v5_4"}
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in sheets.items():
            if name in remove_names:
                continue
            if df is None:
                df = pd.DataFrame()
            df.to_excel(writer, sheet_name=name[:31], index=False)
        wb = writer.book
        props = wb.properties
        props.title = APP_NAME
        props.subject = APP_SUBJECT
        props.creator = APP_AUTHOR
        props.keywords = APP_KEYWORDS
        props.category = APP_CATEGORY
        props.description = APP_XLSX_DESCRIPTION
        props.lastModifiedBy = APP_AUTHOR
        try:
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            from openpyxl.utils import get_column_letter
            header_fill = PatternFill("solid", fgColor="1F4E5F")
            header_font = Font(color="FFFFFF", bold=True)
            thin = Side(style="thin", color="D9E2E7")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for ws in wb.worksheets:
                ws.sheet_view.showGridLines = False
                ws.freeze_panes = "A2"
                if ws.max_row >= 1:
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        cell.border = border
                for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 5000)):
                    for cell in row:
                        cell.border = border
                        cell.alignment = Alignment(vertical="top", wrap_text=True)
                for col_idx, col in enumerate(ws.columns, start=1):
                    values = [str(c.value) for c in col[:200] if c.value is not None]
                    width = min(max([len(v) for v in values] + [10]) + 2, 55)
                    ws.column_dimensions[get_column_letter(col_idx)].width = width
        except Exception:
            pass
    try:
        self.progress_var.set(100.0)
        self.status_var.set(f"XLSX exportado: {path}")
    except Exception:
        pass

ColumnsEC2App._write_excel = _write_excel_v59



# ============================================================
# ColumnsEC2 v6.0 — revisão de relatórios, fluência por backend,
# pormenorização construtiva comum e quadro de pilares por prumada.
# ============================================================
APP_VERSION = "v6.2"

APP_XLSX_DESCRIPTION = (
    "Dimensionamento e verificação de pilares de betão armado por flexão composta, "
    "interacção N-My-Mz, verificação de serviço, esforço transverso, torção, "
    "pormenorização construtiva, emissão de memória de cálculo e quadro de pilares."
)

# --------------------------- descrições / limitações por módulo ---------------------------
def _module_description_v60(app=None) -> str:
    b = _v59_get_backend(app)
    if b == BACKEND_EC2_PT_2010:
        return (
            "Módulo NP EN 1992-1-1:2010 PT: dimensionamento de pilares de betão armado "
            "com selecção de casos governantes, verificação de 2.ª ordem, interacção N-My-Mz, "
            "esforço transverso, torção, ELS, pormenorização construtiva e quadro de pilares."
        )
    if b == BACKEND_SC_EC2_2004:
        return (
            "Módulo Eurocode 2:2004: utiliza o pacote structuralcodes para propriedades dos materiais, "
            "fluência/retracção, verificação seccional N-My-Mz, esforço transverso e fendilhação quando "
            "as funções correspondentes se encontram disponíveis na instalação local."
        )
    if b == BACKEND_SC_EC2_2023:
        return (
            "Módulo Eurocode 2:2023: utiliza o pacote structuralcodes para propriedades dos materiais, "
            "fluência/retracção, verificação seccional N-My-Mz, fendilhação e deformação quando "
            "as funções correspondentes se encontram disponíveis na instalação local."
        )
    return (
        "Módulo fib Model Code 2010: utiliza o pacote structuralcodes para propriedades dos materiais, "
        "fluência/retracção, verificação seccional N-My-Mz, esforço transverso, torção, ELS e aderência "
        "quando as funções correspondentes se encontram disponíveis na instalação local."
    )

def _module_limitations_v60(app=None) -> str:
    b = _v59_get_backend(app)
    if b == BACKEND_EC2_PT_2010:
        return (
            "A validade dos resultados depende da coerência dos esforços importados, do sistema de eixos locais, "
            "dos comprimentos efectivos adoptados e da confirmação final da pormenorização em desenho. "
            "Casos críticos devem ser revistos por cálculo independente."
        )
    if b == BACKEND_SC_EC2_2004:
        return (
            "A cobertura depende da API structuralcodes instalada. Verificações não expostas pela API são "
            "assinaladas no relatório. A pormenorização de armaduras é tratada como verificação construtiva "
            "do ColumnsEC2 e deve ser confirmada em desenho."
        )
    if b == BACKEND_SC_EC2_2023:
        return (
            "A cobertura depende da API structuralcodes instalada. A verificação de 2.ª ordem, esforço "
            "transverso ou torção só é executada quando houver função disponível no backend seleccionado. "
            "A pormenorização de armaduras é construtiva e deve ser confirmada em desenho."
        )
    return (
        "A cobertura depende da API structuralcodes instalada e dos parâmetros disponíveis para o fib Model Code 2010. "
        "A pormenorização de armaduras é uma verificação construtiva do ColumnsEC2 e deve ser confirmada "
        "em desenho de projecto."
    )

# manter compatibilidade com chamadas antigas
_module_description_v59 = _module_description_v60
_module_limitations_v59 = _module_limitations_v60

def _v60_short_gui_description() -> str:
    return "Dimensionamento de pilares em betão armado: ELU, ELS, interacção N-My-Mz, pormenorização e relatórios técnicos."

def _v60_quick_notes(app=None) -> str:
    b = _v59_get_backend(app)
    mode = _v53_mode_to_label(_v59_getvar(app, "var_calc_mode", "Dimensionamento"))
    if b == BACKEND_EC2_PT_2010:
        return (
            f"Norma: {_backend_display_v59(b)}\n"
            f"Modo: {mode}\n"
            "Verificação: N-My-Mz, esbelteza, 2.ª ordem, V, T, ELS e pormenorização construtiva. "
            "Os esforços devem corresponder a combinações reais da análise."
        )
    if b == BACKEND_SC_EC2_2004:
        return (
            "Norma: Eurocode 2:2004\n"
            "Motor: structuralcodes. Materiais, fluência/retracção, N-My-Mz, V e fendilhação são avaliados "
            "quando disponíveis na API instalada."
        )
    if b == BACKEND_SC_EC2_2023:
        return (
            "Norma: Eurocode 2:2023\n"
            "Motor: structuralcodes. Materiais, fluência/retracção, N-My-Mz, fendilhação e deformação são "
            "avaliados quando disponíveis na API instalada."
        )
    return (
        "Norma: fib Model Code 2010\n"
        "Motor: structuralcodes. Materiais, fluência/retracção, N-My-Mz, esforço transverso, torção, ELS "
        "e aderência são avaliados quando disponíveis na API instalada."
    )

def _v59_update_quick_notes(app):
    try:
        if hasattr(app, "quick_notes_var"):
            app.quick_notes_var.set(_v60_quick_notes(app))
    except Exception:
        pass

# --------------------------- h0 e φef por backend ---------------------------
def _v60_section_rows_for_creep(app=None):
    """Devolve linhas geométricas válidas para estimar h0/hn/notional size."""
    rows = []
    try:
        df = getattr(app, "df_pair", pd.DataFrame())
        if df is None or df.empty:
            df = getattr(app, "df_clean", pd.DataFrame())
        if df is not None and not df.empty:
            for _, r in df.iterrows():
                b = cm_to_mm(r.get("hy", 0.0))
                h = cm_to_mm(r.get("hz", 0.0))
                if b > 0 and h > 0:
                    rows.append((b, h, str(r.get("material", DEFAULT_CONCRETE_CLASS))))
    except Exception:
        pass
    if not rows:
        rows.append((300.0, 500.0, DEFAULT_CONCRETE_CLASS))
    return rows

def _v60_h0_for_rect(b_mm, h_mm, backend=None):
    """h0/hn = 2Ac/u para perímetro exposto à secagem; usa API structuralcodes se disponível."""
    b = float(b_mm); h = float(h_mm)
    Ac = b * h
    u = 2.0 * (b + h)
    key = _backend_selected_v52(backend)
    if key in [BACKEND_SC_EC2_2004, BACKEND_SC_EC2_2023]:
        try:
            sc, err = _sc_import_backend_v52(key)
            if sc is not None:
                mod = sc.get("module")
                if key == BACKEND_SC_EC2_2004 and hasattr(mod, "h_0"):
                    return float(mod.h_0(Ac, u)), "structuralcodes.codes.ec2_2004.h_0"
                if key == BACKEND_SC_EC2_2023 and hasattr(mod, "hn"):
                    return float(mod.hn(Ac, u)), "structuralcodes.codes.ec2_2023.hn"
        except Exception:
            pass
    # MC2010 usa notional_size = 2A/u.
    return float(2.0 * Ac / max(u, 1e-9)), "2Ac/u"

def _v59_estimate_h0_for_app(app=None) -> float:
    backend = _v59_get_backend(app)
    vals = []
    for b, h, _mat in _v60_section_rows_for_creep(app):
        val, _src = _v60_h0_for_rect(b, h, backend)
        if math.isfinite(val) and val > 0:
            vals.append(val)
    return min(vals) if vals else 200.0

def _v60_structuralcodes_phi(backend: str, fck: float, RH: float, t0: float, h0: float):
    """Calcula φ através das funções structuralcodes documentadas, quando instaladas."""
    key = _backend_selected_v52(backend)
    sc, err = _sc_import_backend_v52(key)
    if sc is None:
        return None, f"structuralcodes indisponível: {err}"
    mod = sc.get("module")
    fck = float(fck)
    fcm = fck + 8.0
    RH = max(40.0, min(100.0, float(RH)))
    t0 = max(1.0, float(t0))
    h0 = max(50.0, min(1000.0, float(h0)))
    t_inf = 36500.0

    try:
        if key == BACKEND_SC_EC2_2004:
            # EN 1992-1-1:2004 Annex B sequence.
            a1 = mod.alpha_1(fcm)
            a2 = mod.alpha_2(fcm)
            a3 = mod.alpha_3(fcm)
            phi_RH = mod.phi_RH(h0, fcm, RH, a1, a2)
            phi0 = mod.phi_0(phi_RH, mod.beta_fcm(fcm), mod.beta_t0(t0))
            betaH = mod.beta_H(h0, fcm, RH, a3)
            betaC = mod.beta_c(t0, t_inf, betaH)
            val = mod.phi(phi0, betaC)
            return float(val), "structuralcodes.codes.ec2_2004: h_0, phi_RH, phi_0, beta_c, phi"

        if key == BACKEND_SC_EC2_2023:
            # EC2:2023 Table 5.2: phi_50y_t0 corrected for fck.
            atm = "humid" if RH >= 70.0 else "dry"
            hn = max(100.0, min(1000.0, h0))
            strength_class = "CN"
            phi50 = mod.phi_50y_t0(t0, atm, hn, strength_class)
            try:
                Aexp = mod.A_phi_correction_exp(hn, atm)
                corr = mod.phi_correction_factor(fck, Aexp)
                val = float(phi50) * float(corr)
                return val, "structuralcodes.codes.ec2_2023: hn, phi_50y_t0, A_phi_correction_exp, phi_correction_factor"
            except Exception:
                return float(phi50), "structuralcodes.codes.ec2_2023: hn, phi_50y_t0"

        if key == BACKEND_SC_MC2010:
            # fib MC2010: basic + drying creep.
            notional_size = max(50.0, float(h0))
            try:
                tT = mod.t_T(20.0, t0)
            except Exception:
                tT = t0
            try:
                t0adj = mod.t0_adj(tT, "42.5 N")
            except Exception:
                t0adj = t0
            phi_bc = mod.phi_bc(mod.beta_bc_fcm(fcm), mod.beta_bc_t(t_inf, t0, t0adj))
            alpha = mod.alpha_fcm(fcm)
            beta_h = mod.beta_h(notional_size, alpha)
            gamma_t0 = mod.gamma_t0(t0adj)
            phi_dc = mod.phi_dc(
                mod.beta_dc_fcm(fcm),
                mod.beta_dc_RH(RH, notional_size),
                mod.beta_dc_t0(t0adj),
                mod.beta_dc_t(t_inf, t0, beta_h, gamma_t0),
            )
            # sigma <= 0.4 fcm para fluência linear.
            val = mod.phi(phi_bc, phi_dc, 0.4 * fcm, fcm)
            return float(val), "structuralcodes.codes.mc2010: phi_bc, phi_dc, phi"

    except Exception as e:
        return None, f"API de fluência encontrada mas não executada: {e}"

    return None, "backend não reconhecido para fluência"

def _v53_try_structuralcodes_phi(backend: str, fck: float, RH: float, t0: float, h0: float):
    return _v60_structuralcodes_phi(backend, fck, RH, t0, h0)

def _v60_compute_phi_eff_for_app(app=None):
    backend = _v59_get_backend(app)
    RH = safe_float(_v59_getvar(app, "var_creep_RH", "70"), 70.0)
    t0 = safe_float(_v59_getvar(app, "var_creep_t0", "28"), 28.0)

    # Calcular phi por material/secção e adoptar valor governante para o motor global.
    candidates = []
    for b, h, mat in _v60_section_rows_for_creep(app):
        h0, h0src = _v60_h0_for_rect(b, h, backend)
        fck = parse_concrete_strength(mat or DEFAULT_CONCRETE_CLASS)
        if _sc_backend_active_v52(backend):
            val, src = _v60_structuralcodes_phi(backend, fck, RH, t0, h0)
            if val is not None and math.isfinite(float(val)) and float(val) > 0:
                candidates.append((float(val), f"{src}; {h0src}; RH={RH:.0f}%; t0={t0:.0f} d; h0={h0:.0f} mm; {mat}"))
        else:
            try:
                val = _v53_ec2_creep_phi(fck, RH=RH, t0=t0, h0_mm=h0)
                candidates.append((float(val), f"NP EN 1992-1-1 Anexo B; {h0src}; RH={RH:.0f}%; t0={t0:.0f} d; h0={h0:.0f} mm; {mat}"))
            except Exception:
                pass

    if candidates:
        val, src = max(candidates, key=lambda x: x[0])
        return val, src + " — valor governante adoptado"
    if _sc_backend_active_v52(backend):
        return 0.0, f"φef não calculado: API de fluência indisponível ou parâmetros fora do domínio de {_backend_display_v59(backend)}."
    return 2.0, "φef não calculado; valor adoptado = 2.0"

_v59_compute_phi_eff_for_app = _v60_compute_phi_eff_for_app

# --------------------------- GUI ---------------------------
_old_build_sidebar_v60_base = ColumnsEC2App._build_sidebar
def _build_sidebar_v60(self, parent):
    _old_build_sidebar_v60_base(self, parent)
    try:
        if not hasattr(self, "quick_notes_var"):
            self.quick_notes_var = tk.StringVar()
        self.quick_notes_var.set(_v60_quick_notes(self))

        def walk(w):
            yield w
            for c in w.winfo_children():
                yield from walk(c)

        # Descrição curta do programa e link invisível no nome.
        for w in walk(parent):
            try:
                txt = str(w.cget("text"))
                if txt == APP_NAME or txt.strip().lower() == "columnsec2":
                    w.configure(cursor="hand2")
                    w.bind("<Button-1>", lambda _e: webbrowser.open_new(GITHUB_URL))
                if "Ferramenta para importação" in txt or "importação de esforços" in txt:
                    w.configure(text=_v60_short_gui_description())
            except Exception:
                pass

        # Notas rápidas com texto dinâmico.
        for lf in walk(parent):
            try:
                if isinstance(lf, ttk.LabelFrame) and "Notas" in str(lf.cget("text")):
                    for child in lf.winfo_children():
                        try:
                            if isinstance(child, ttk.Label):
                                child.configure(textvariable=self.quick_notes_var, text="", wraplength=330, justify="left")
                        except Exception:
                            pass
            except Exception:
                pass

        try:
            self.var_code_backend.trace_add("write", lambda *_: _v59_update_quick_notes(self))
            self.var_calc_mode.trace_add("write", lambda *_: _v59_update_quick_notes(self))
        except Exception:
            pass
    except Exception:
        pass

ColumnsEC2App._build_sidebar = _build_sidebar_v60

# --------------------------- notas normativas ---------------------------
def _build_normative_notes_v60(self) -> pd.DataFrame:
    b = _v59_get_backend(self)
    if b == BACKEND_EC2_PT_2010:
        rows = [
            ("Âmbito", "NP EN 1992-1-1:2010 + AC:2012 + A1:2019", "Verificação de pilares de betão armado em edifícios, com Anexo Nacional português."),
            ("Esforços", "Combinações importadas", "Os esforços de cálculo são lidos da tabela e tratados por casos reais governantes, sem criar combinações artificiais."),
            ("Eixos locais", "X longitudinal", "FX é o esforço normal; FY/FZ são esforços transversos; MX é torção; MY/MZ são momentos flectores."),
            ("Segunda ordem", "EC2 5.8", "São avaliados esbelteza, esbelteza-limite, imperfeição geométrica, momento equivalente e momento de 2.ª ordem."),
            ("Fluência", "Anexo B", "h0 e φef são calculados automaticamente a partir da geometria, classe de betão, RH e idade de carregamento."),
            ("Resistência seccional", "N-My-Mz", "O pilar é verificado por flexão composta desviada, usando catálogo construtivo de armaduras e domínio resistente discreto."),
            ("Esforço transverso", "EC2 6.2", "São avaliados VRd,c, VRd,max e armadura transversal requerida quando VEd excede a resistência sem armadura adicional."),
            ("Torção", "EC2 6.3", "A torção é classificada por relevância; torção desprezável não condiciona a solução de armadura."),
            ("ELS", "EC2 7", "Quando indicada, é usada a combinação ELS fornecida pelo utilizador; caso contrário, o controlo é simplificado/informativo."),
            ("Pormenorização", "EC2 8 e 9.5", "São verificados mínimos, máximos, diâmetros, espaçamentos e coerência construtiva dos varões e estribos."),
            ("Quadro de pilares", "DXF", "As secções são agrupadas por prumada e organizadas por piso/tramo, de baixo para cima."),
        ]
    elif b == BACKEND_SC_EC2_2004:
        rows = [
            ("Âmbito", "Eurocode 2:2004", "Cálculo com structuralcodes para propriedades, fluência/retracção, secções, esforço transverso e fendilhação quando disponíveis."),
            ("Materiais", "structuralcodes.codes.ec2_2004", "Propriedades de betão e aço são obtidas através da API do pacote."),
            ("Fluência", "Creep and shrinkage", "h0, phi_RH, phi_0, beta_c e phi são calculados pelas funções EC2:2004 expostas pelo pacote quando disponíveis."),
            ("Resistência seccional", "structuralcodes.sections", "A interacção N-My-Mz é avaliada por secções armadas criadas no motor structuralcodes."),
            ("Esforço transverso", "EC2 2004 shear API", "VRd,c/VRd,max são avaliados quando as funções estão disponíveis na instalação local."),
            ("Pormenorização", "ColumnsEC2", "A verificação construtiva de armaduras é feita pelo ColumnsEC2 para apoiar o desenho final."),
        ]
    elif b == BACKEND_SC_EC2_2023:
        rows = [
            ("Âmbito", "Eurocode 2:2023", "Cálculo com structuralcodes para materiais, fluência/retracção, secções, fendilhação e deformação quando disponíveis."),
            ("Materiais", "structuralcodes.codes.ec2_2023", "Propriedades de betão e aço são obtidas através da API do pacote."),
            ("Fluência", "EC2:2023 Table 5.2", "hn, phi_50y_t0 e factores de correcção são calculados pelas funções EC2:2023 expostas pelo pacote."),
            ("Resistência seccional", "structuralcodes.sections", "A interacção N-My-Mz é avaliada por secções armadas criadas no motor structuralcodes."),
            ("ELS", "EC2 2023 SLS API", "Fendilhação/deformação são avaliadas quando as funções e parâmetros necessários estão disponíveis."),
            ("Pormenorização", "ColumnsEC2", "A verificação construtiva de armaduras é feita pelo ColumnsEC2 para apoiar o desenho final."),
        ]
    else:
        rows = [
            ("Âmbito", "fib Model Code 2010", "Cálculo com structuralcodes para materiais, fluência/retracção, secções, V, T, ELS e aderência quando disponíveis."),
            ("Materiais", "structuralcodes.codes.mc2010", "Propriedades de betão e aço são obtidas através da API MC2010 do pacote."),
            ("Fluência", "MC2010 5.1", "O coeficiente de fluência é composto por parcelas de fluência básica e de secagem, quando a API local expõe as funções."),
            ("Resistência seccional", "structuralcodes.sections", "A interacção N-My-Mz é avaliada por secções armadas criadas no motor structuralcodes."),
            ("Esforço transverso e torção", "MC2010 API", "V e T são avaliados quando as funções MC2010 estão disponíveis na instalação local."),
            ("Pormenorização", "ColumnsEC2", "A verificação construtiva de armaduras é feita pelo ColumnsEC2 para apoiar o desenho final."),
        ]
    return pd.DataFrame(rows, columns=["Tema", "Referência", "Nota"])

ColumnsEC2App.build_normative_notes = _build_normative_notes_v60

# --------------------------- metadados e parâmetros XLSX ---------------------------
def _metadata_df_v60(self) -> pd.DataFrame:
    return pd.DataFrame([
        ["Programa", APP_NAME],
        ["Autor / Repositório", GITHUB_URL],
        ["Data de emissão", datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["Ficheiro de origem", self.input_file_path or "-"],
        ["Norma de referência", _backend_reference_v59(self)],
        ["Modo de cálculo", _v53_mode_to_label(_v59_getvar(self, "var_calc_mode", "Dimensionamento"))],
        ["Descrição", _module_description_v60(self)],
        ["Limitações", _module_limitations_v60(self)],
    ], columns=["Campo", "Valor"])

ColumnsEC2App._metadata_df = _metadata_df_v60

def _parameters_df_v60(self) -> pd.DataFrame:
    phi, src = _v60_compute_phi_eff_for_app(self)
    h0 = _v59_estimate_h0_for_app(self)
    return pd.DataFrame([
        ["Recobrimento nominal [mm]", _v59_getvar(self, "var_cover", "35")],
        ["Classe de Aço", _v59_steel_class_label(self)],
        ["fyk [MPa]", safe_float(_v59_getvar(self, "var_fyk", "A500"), 500.0)],
        ["Betão", _v59_material_classes(self)],
        ["Humidade relativa RH [%]", _v59_getvar(self, "var_creep_RH", "70")],
        ["Idade do betão t0 [dias]", _v59_getvar(self, "var_creep_t0", "28")],
        ["h0 automático governante [mm]", f"{h0:.0f}"],
        ["φef automático governante", f"{phi:.3f}" if phi and phi > 0 else "não aplicado"],
        ["Método de cálculo de φef", src],
        ["l0y/L", _v59_getvar(self, "var_l0y", "1.0")],
        ["l0z/L", _v59_getvar(self, "var_l0z", "1.0")],
        ["Redução para casos governantes", "Sim" if getattr(self, "var_reduce_cases", tk.BooleanVar(value=True)).get() else "Não"],
        ["Combinação ELS indicada", _v59_getvar(self, "var_service_case", "").strip() or "não indicada"],
    ], columns=["Parâmetro", "Valor"])

ColumnsEC2App._parameters_df = _parameters_df_v60

# --------------------------- cobertura / diagnóstico ---------------------------
def _backend_coverage_df_v60(app=None) -> pd.DataFrame:
    rows = [
        ["Materiais", BACKEND_DISPLAY_PT_V59, "ColumnsEC2", "Calculado"],
        ["h0 / φef", BACKEND_DISPLAY_PT_V59, "ColumnsEC2 / Anexo B", "Calculado automaticamente"],
        ["2.ª ordem", BACKEND_DISPLAY_PT_V59, "ColumnsEC2", "Calculado"],
        ["N-My-Mz", BACKEND_DISPLAY_PT_V59, "ColumnsEC2", "Calculado"],
        ["Esforço transverso", BACKEND_DISPLAY_PT_V59, "ColumnsEC2", "Calculado"],
        ["Torção", BACKEND_DISPLAY_PT_V59, "ColumnsEC2", "Calculado"],
        ["ELS", BACKEND_DISPLAY_PT_V59, "ColumnsEC2", "Calculado/simplificado conforme combinação"],
        ["Pormenorização construtiva", BACKEND_DISPLAY_PT_V59, "ColumnsEC2", "Calculado"],
        ["Materiais", BACKEND_DISPLAY_EC2004_V59, "structuralcodes.codes.ec2_2004", "Calculado se API disponível"],
        ["h0 / φef", BACKEND_DISPLAY_EC2004_V59, "structuralcodes.codes.ec2_2004", "h_0, phi_RH, phi_0, beta_c, phi"],
        ["N-My-Mz", BACKEND_DISPLAY_EC2004_V59, "structuralcodes.sections", "Calculado se API disponível"],
        ["Esforço transverso", BACKEND_DISPLAY_EC2004_V59, "structuralcodes.codes.ec2_2004", "Calculado se API disponível"],
        ["ELS/fendilhação", BACKEND_DISPLAY_EC2004_V59, "structuralcodes.codes.ec2_2004", "Calculado se API/parâmetros disponíveis"],
        ["Pormenorização construtiva", BACKEND_DISPLAY_EC2004_V59, "ColumnsEC2", "Calculado"],
        ["Materiais", BACKEND_DISPLAY_EC2023_V59, "structuralcodes.codes.ec2_2023", "Calculado se API disponível"],
        ["hn / φ", BACKEND_DISPLAY_EC2023_V59, "structuralcodes.codes.ec2_2023", "hn, phi_50y_t0, phi_correction_factor"],
        ["N-My-Mz", BACKEND_DISPLAY_EC2023_V59, "structuralcodes.sections", "Calculado se API disponível"],
        ["ELS/fendilhação/deformação", BACKEND_DISPLAY_EC2023_V59, "structuralcodes.codes.ec2_2023", "Calculado se API/parâmetros disponíveis"],
        ["Pormenorização construtiva", BACKEND_DISPLAY_EC2023_V59, "ColumnsEC2", "Calculado"],
        ["Materiais", BACKEND_DISPLAY_MC2010_V59, "structuralcodes.codes.mc2010", "Calculado se API disponível"],
        ["Fluência/retracção", BACKEND_DISPLAY_MC2010_V59, "structuralcodes.codes.mc2010", "phi_bc, phi_dc, phi"],
        ["N-My-Mz", BACKEND_DISPLAY_MC2010_V59, "structuralcodes.sections", "Calculado se API disponível"],
        ["Esforço transverso", BACKEND_DISPLAY_MC2010_V59, "structuralcodes.codes.mc2010", "Calculado se API disponível"],
        ["Torção", BACKEND_DISPLAY_MC2010_V59, "structuralcodes.codes.mc2010", "Calculado se API disponível"],
        ["ELS/aderência", BACKEND_DISPLAY_MC2010_V59, "structuralcodes.codes.mc2010", "Calculado se API/parâmetros disponíveis"],
        ["Pormenorização construtiva", BACKEND_DISPLAY_MC2010_V59, "ColumnsEC2", "Calculado"],
    ]
    return pd.DataFrame(rows, columns=["Verificação", "Norma", "Origem", "Estado"])

_backend_coverage_df_v59 = _backend_coverage_df_v60

def _sc_api_diagnostic_df_v60() -> pd.DataFrame:
    rows = []
    fn_map = {
        BACKEND_SC_EC2_2004: ["h_0","phi","phi_0","phi_RH","beta_fcm","beta_t0","beta_c","beta_H","VRdc","VRdmax","wk","wk_cal"],
        BACKEND_SC_EC2_2023: ["hn","phi_50y_t0","phi_correction_factor","A_phi_correction_exp","eps_cs_50y","wk","wk_cal","delta_simpl"],
        BACKEND_SC_MC2010: ["phi","phi_bc","phi_dc","beta_bc_fcm","beta_bc_t","beta_dc_RH","beta_h","v_rd","v_rdc","t_rd","t_rd_max"],
    }
    for b, fns in fn_map.items():
        sc, err = _sc_import_backend_v52(b)
        disp = _backend_display_v59(b)
        if sc is None:
            rows.append([disp, "structuralcodes", "Indisponível", str(err)])
            continue
        mod = sc.get("module")
        for fn in ["fcd","fctm","Ecm","Eci","fyd","Es"] + fns:
            rows.append([disp, fn, "Sim" if hasattr(mod, fn) else "Não", ""])
        rows.append([disp, "BeamSection", "Sim" if sc.get("BeamSection") is not None else "Não", "structuralcodes.sections"])
    return pd.DataFrame(rows, columns=["Norma", "Objecto/API", "Disponível", "Nota"])

_sc_api_diagnostic_df_v59 = _sc_api_diagnostic_df_v60

# --------------------------- postprocess pormenorização construtiva comum ---------------------------
def _v60_postprocess_constructive_detailing(results: pd.DataFrame) -> pd.DataFrame:
    if results is None or results.empty:
        return results
    out = results.copy()
    for idx, r in out.iterrows():
        try:
            d = detailing_check_v4(r.to_dict()) if "detailing_check_v4" in globals() else {}
            if d:
                for k, v in d.items():
                    # Não transformar avisos construtivos em falha para backends structuralcodes.
                    if k not in out.columns or pd.isna(out.at[idx, k]) or str(out.at[idx, k]).strip() in ["", "nan", "Não avaliado neste backend"]:
                        out.at[idx, k] = v
                if str(out.get("code_backend", pd.Series(index=out.index)).get(idx, "")).lower().find("structuralcodes") >= 0:
                    if str(out.at[idx, "detailing_status"]).lower().startswith("não conforme"):
                        out.at[idx, "detailing_status"] = "Aviso construtivo"
        except Exception:
            pass
    return out

_old_run_design_v60_base = ColumnsEC2App.run_design
def _run_design_v60(self):
    _old_run_design_v60_base(self)
    # O worker antigo preenche resultados de forma assíncrona; este hook leve reprocessa no finish via after.
    def later():
        try:
            if getattr(self, "df_results", pd.DataFrame()) is not None and not self.df_results.empty:
                self.df_results = _v60_postprocess_constructive_detailing(self.df_results)
                summary_enabled = bool(self.var_summary.get()) if hasattr(self, "var_summary") else True
                self.df_summary = self.build_summary_by_member(self.df_results) if summary_enabled else pd.DataFrame()
                self.df_failures = _v59_failure_rows(self.df_results)
                self.df_ok = self.df_results[self.df_results.get("status", pd.Series(index=self.df_results.index, dtype=str)).astype(str).eq("OK")].copy()
                for tree, df in [(self.tree_results, self.df_results), (self.tree_summary, self.df_summary), (self.tree_failures, self.df_failures)]:
                    try:
                        self.show_df(tree, df)
                    except Exception:
                        pass
        except Exception:
            pass
    try:
        self.after(1500, later)
    except Exception:
        pass

ColumnsEC2App.run_design = _run_design_v60

# --------------------------- propostas de correcção ---------------------------
def _v60_row_for_repair(row: pd.Series, scale=1.0):
    r = row.copy()
    b_cm = _finite(r.get("b_cm", r.get("hy", 0.0)), 0.0) * scale
    h_cm = _finite(r.get("h_cm", r.get("hz", 0.0)), 0.0) * scale
    if b_cm <= 0: b_cm = _finite(r.get("hy", 30.0), 30.0) * scale
    if h_cm <= 0: h_cm = _finite(r.get("hz", 30.0), 30.0) * scale
    r["hy"] = b_cm
    r["hz"] = h_cm
    r["ax"] = b_cm * h_cm
    r["iy"] = b_cm * (h_cm ** 3) / 12.0
    r["iz"] = h_cm * (b_cm ** 3) / 12.0
    n = _finite(row.get("n_ed_kN", row.get("fx", 0.0)), 0.0)
    r["fx_i"] = n
    r["fx_j"] = n
    r["fy_i"] = _finite(row.get("vy_ed_kN", 0.0), 0.0)
    r["fy_j"] = r["fy_i"]
    r["fz_i"] = _finite(row.get("vz_ed_kN", 0.0), 0.0)
    r["fz_j"] = r["fz_i"]
    r["mx_i"] = _finite(row.get("mx_ed_kNm", 0.0), 0.0)
    r["mx_j"] = r["mx_i"]
    r["my_i"] = _finite(row.get("my_i_kNm", row.get("my_ed_kNm", 0.0)), 0.0)
    r["my_j"] = _finite(row.get("my_j_kNm", row.get("my_ed_kNm", 0.0)), 0.0)
    r["mz_i"] = _finite(row.get("mz_i_kNm", row.get("mz_ed_kNm", 0.0)), 0.0)
    r["mz_j"] = _finite(row.get("mz_j_kNm", row.get("mz_ed_kNm", 0.0)), 0.0)
    r["n_nodes_found"] = 2
    r["length"] = _finite(row.get("length_m", row.get("length", 3.0)), 3.0)
    r["material"] = row.get("material", DEFAULT_CONCRETE_CLASS)
    r["member"] = row.get("member", "")
    r["case"] = row.get("case", "")
    r["name"] = row.get("name", row.get("prumada", ""))
    r["prumada"] = row.get("prumada", row.get("name", ""))
    return r

def _v60_make_repair_designer(app):
    return ColumnDesigner(
        cover_mm=safe_float(_v59_getvar(app, "var_cover", "35"), 35.0),
        fyk=safe_float(_v59_getvar(app, "var_fyk", "A500"), 500.0),
        phi_eff=safe_float(_v59_getvar(app, "var_phi_eff", "2.0"), 2.0),
        l0y_factor=safe_float(_v59_getvar(app, "var_l0y", "1.0"), 1.0),
        l0z_factor=safe_float(_v59_getvar(app, "var_l0z", "1.0"), 1.0),
        calc_mode="dimensionamento",
    )

def _v60_generate_repair_proposal(app, row: pd.Series) -> dict:
    designer = _v60_make_repair_designer(app)
    attempts = []
    for scale in [1.00, 1.05, 1.10, 1.15, 1.20, 1.30, 1.40]:
        rr = _v60_row_for_repair(row, scale=scale)
        try:
            out = designer.design_one(rr)
            out["repair_scale_section"] = scale
            out["repair_note"] = "mantida a secção" if abs(scale-1.0) < 1e-9 else f"proposta de aumento de secção x{scale:.2f}"
            out["_target_index"] = row.name
            attempts.append(out)
            if str(out.get("status", "")) in ["OK", "Aviso"]:
                out["repair_result"] = "Proposta aceite" if str(out.get("status")) == "OK" else "Proposta com aviso"
                return out
        except Exception as e:
            attempts.append({"status": "Falha", "repair_scale_section": scale, "repair_note": str(e), "_target_index": row.name})
    best = attempts[-1] if attempts else dict(row)
    best["repair_result"] = "Sem proposta automática"
    best["repair_note"] = "Nenhuma solução verificou no catálogo construtivo testado. Rever esforços, geometria, comprimento efectivo ou adoptar secção superior."
    best["_target_index"] = row.name
    return best

def _repair_failures_interactive_v60(self):
    if self.df_results is None or self.df_results.empty:
        messagebox.showwarning("Aviso", "Execute primeiro o cálculo.")
        return
    backend = _v59_get_backend(self)
    failures = _v59_failure_rows(self.df_results)
    if failures.empty:
        messagebox.showinfo("Correcção iterativa", "Não foram detectadas falhas bloqueantes.")
        return

    if _sc_backend_active_v52(backend):
        # Nestes modos, só são emitidas propostas construtivas; não se substitui o cálculo do backend.
        if not messagebox.askyesno(
            "Falhas detectadas",
            "Foram detectadas falhas/avisos no backend seleccionado.\n\n"
            + _v59_failure_summary_text(failures, max_rows=10)
            + "\n\nGerar apenas propostas construtivas de armadura/secção, mantendo o cálculo normativo do backend seleccionado?"
        ):
            return
    else:
        if not messagebox.askyesno(
            "Falhas bloqueantes detectadas",
            "O programa pode tentar gerar propostas de correcção.\n\n"
            + _v59_failure_summary_text(failures, max_rows=10)
            + "\n\nGerar propostas?"
        ):
            return

    proposals = []
    total = len(failures)
    for k, (idx, row) in enumerate(failures.iterrows(), start=1):
        self.progress_var.set(100.0 * k / max(total, 1))
        self.status_var.set(f"A gerar proposta... {k}/{total}")
        self.update_idletasks()
        prop = _v60_generate_repair_proposal(self, row)
        prop["_target_index"] = idx
        proposals.append(prop)

    self.df_repair_proposals = pd.DataFrame(proposals)
    ok = int((self.df_repair_proposals.get("status", pd.Series(dtype=str)).astype(str) == "OK").sum()) if not self.df_repair_proposals.empty else 0
    warn = int((self.df_repair_proposals.get("status", pd.Series(dtype=str)).astype(str) == "Aviso").sum()) if not self.df_repair_proposals.empty else 0
    fail = len(self.df_repair_proposals) - ok - warn
    self.status_var.set(f"Propostas geradas: {ok} OK; {warn} avisos; {fail} sem solução.")
    if _sc_backend_active_v52(backend):
        messagebox.showinfo("Propostas geradas", f"Propostas construtivas geradas:\nOK: {ok}\nAvisos: {warn}\nSem solução: {fail}\n\nNão foram aplicadas automaticamente por se tratar de backend structuralcodes.")
        return

    if not messagebox.askyesno("Aplicar propostas", f"Propostas geradas:\nOK: {ok}\nAvisos: {warn}\nSem solução: {fail}\n\nAplicar apenas as propostas OK?"):
        return

    res = self.df_results.copy()
    applied = 0
    for _, p in self.df_repair_proposals.iterrows():
        idx = p.get("_target_index", None)
        if idx in res.index and str(p.get("status", "")) == "OK":
            for c in self.df_repair_proposals.columns:
                if c != "_target_index":
                    res.at[idx, c] = p.get(c)
            res.at[idx, "auto_repair_applied"] = "Sim"
            applied += 1
    try:
        res = enrich_failures_v43(res)
    except Exception:
        pass
    self.df_results = _v60_postprocess_constructive_detailing(res)
    self.df_summary = self.build_summary_by_member(self.df_results) if getattr(self, "var_summary", tk.BooleanVar(value=True)).get() else pd.DataFrame()
    self.df_failures = _v59_failure_rows(self.df_results)
    self.df_ok = self.df_results[self.df_results.get("status", pd.Series(index=self.df_results.index, dtype=str)).astype(str).eq("OK")].copy()
    for tree, df in [(self.tree_results, self.df_results), (self.tree_summary, self.df_summary), (self.tree_failures, self.df_failures), (self.tree_shortlists, self.build_shortlists_df())]:
        try:
            self.show_df(tree, df)
        except Exception:
            pass
    self.update_report()
    self.status_var.set(f"Correcção aplicada: {applied} propostas OK.")
    messagebox.showinfo("Correcção aplicada", f"Foram aplicadas {applied} propostas OK.")

ColumnsEC2App.repair_failures_interactive = _repair_failures_interactive_v60

# --------------------------- DXF quadro por prumada e piso de baixo para cima ---------------------------
def _v60_story_label(row):
    for c in ["story", "Story", "piso", "Piso", "floor", "Floor", "level", "Level"]:
        try:
            v = row.get(c, "")
            if str(v).strip() and str(v).strip().lower() not in ["nan", "none"]:
                return str(v).strip()
        except Exception:
            pass
    mem = str(row.get("member", ""))
    return mem or "-"

def _v60_story_key(s):
    txt = str(s)
    nums = re.findall(r"-?\d+", txt)
    return (int(nums[-1]) if nums else 0, txt)

def _v60_dxf_layers():
    names = ["COLUMN_CONCRETE","COLUMN_REBAR","COLUMN_STIRRUP","COLUMN_TEXT","COLUMN_TABLE","COLUMN_AXIS","COLUMN_DIM"]
    s = ["0\nSECTION\n2\nTABLES\n0\nTABLE\n2\nLAYER\n70\n0\n"]
    for nm in names:
        s.append(f"0\nLAYER\n2\n{nm}\n70\n0\n62\n7\n6\nCONTINUOUS\n")
    s.append("0\nENDTAB\n0\nENDSEC\n")
    return "".join(s)

def _v60_draw_section(parts, r, cx, cy, scale=1.0):
    b = _finite(r.get("b_cm")) * 10.0 * scale
    h = _finite(r.get("h_cm")) * 10.0 * scale
    if b <= 0 or h <= 0:
        parts.append(_dxf_line(cx-120, cy, cx+120, cy, "COLUMN_TEXT"))
        return
    left, right, bot, top = cx-b/2, cx+b/2, cy-h/2, cy+h/2
    parts += [
        _dxf_line(left, bot, right, bot, "COLUMN_CONCRETE"),
        _dxf_line(right, bot, right, top, "COLUMN_CONCRETE"),
        _dxf_line(right, top, left, top, "COLUMN_CONCRETE"),
        _dxf_line(left, top, left, bot, "COLUMN_CONCRETE"),
    ]
    cover = _finite(r.get("cover_mm"), 35.0) * scale
    parts += [
        _dxf_line(left+cover, bot+cover, right-cover, bot+cover, "COLUMN_STIRRUP"),
        _dxf_line(right-cover, bot+cover, right-cover, top-cover, "COLUMN_STIRRUP"),
        _dxf_line(right-cover, top-cover, left+cover, top-cover, "COLUMN_STIRRUP"),
        _dxf_line(left+cover, top-cover, left+cover, bot+cover, "COLUMN_STIRRUP"),
    ]
    for y,z,phi in _v56_bar_points_from_result(r):
        parts.append(_dxf_circle(cx + y*scale, cy + z*scale, max(phi*scale/2.0, 2.0), "COLUMN_REBAR"))
    # dimensões e legenda
    parts.append(_dxf_text(left, bot-95, str(r.get("solucao", r.get("layout_description","")))[:80], 20, "COLUMN_TEXT"))
    parts.append(_dxf_text(left, bot-135, f"Est. Ø{int(_finite(r.get('phi_st_mm'),0))}//{_finite(r.get('s_st_mm'))/10:.1f}", 18, "COLUMN_TEXT"))
    parts.append(_dxf_text(left-90, cy-8, f"{_finite(r.get('h_cm'))/100:.2f}", 18, "COLUMN_DIM"))
    parts.append(_dxf_text(cx-35, bot-55, f"{_finite(r.get('b_cm'))/100:.2f}", 18, "COLUMN_DIM"))

def write_columns_dxf_v60(path: str, df: pd.DataFrame):
    parts = ["0\nSECTION\n2\nHEADER\n0\nENDSEC\n", _v60_dxf_layers(), "0\nSECTION\n2\nENTITIES\n"]
    if df is None or df.empty:
        parts.append(_dxf_text(0, 0, "Sem resultados", 50, "COLUMN_TEXT"))
    else:
        work = df.copy()
        if "prumada" not in work.columns:
            work["prumada"] = work.get("name", work.get("member", pd.Series("", index=work.index))).astype(str)
        work["_story_label"] = work.apply(_v60_story_label, axis=1)
        prumadas = sorted(work["prumada"].astype(str).unique(), key=_natural_key_v42 if "_natural_key_v42" in globals() else None)
        stories = sorted(work["_story_label"].astype(str).unique(), key=_v60_story_key)
        prumadas = prumadas[:18]
        stories = stories[:18]

        cw = 1450.0
        rh = 1250.0
        title_h = 260.0
        header_h = 220.0
        row_header_w = 520.0
        x0 = 0.0
        y0 = 0.0
        width = row_header_w + cw*len(prumadas)
        height = title_h + header_h + rh*len(stories)

        left = x0
        bottom = y0
        top = y0 + height
        right = x0 + width

        parts.append(_dxf_text(left + width*0.36, top - 140, "QUADRO DE PILARES", 55, "COLUMN_TEXT"))
        # quadro
        parts += [
            _dxf_line(left, bottom, right, bottom, "COLUMN_TABLE"),
            _dxf_line(right, bottom, right, top, "COLUMN_TABLE"),
            _dxf_line(right, top, left, top, "COLUMN_TABLE"),
            _dxf_line(left, top, left, bottom, "COLUMN_TABLE"),
            _dxf_line(left, top-title_h, right, top-title_h, "COLUMN_TABLE"),
            _dxf_line(left, top-title_h-header_h, right, top-title_h-header_h, "COLUMN_TABLE"),
            _dxf_line(left+row_header_w, top-title_h, left+row_header_w, bottom, "COLUMN_TABLE"),
        ]
        for i in range(len(stories)+1):
            y = bottom + i*rh
            parts.append(_dxf_line(left, y, right, y, "COLUMN_TABLE"))
        for i in range(len(prumadas)+1):
            x = left + row_header_w + i*cw
            parts.append(_dxf_line(x, bottom, x, top-title_h, "COLUMN_TABLE"))

        parts.append(_dxf_text(left+60, top-title_h-140, "Piso", 32, "COLUMN_TEXT"))
        parts.append(_dxf_text(left+240, top-title_h-140, "Pilar", 32, "COLUMN_TEXT"))
        for i,p in enumerate(prumadas):
            x = left+row_header_w+i*cw
            parts.append(_dxf_text(x+cw*0.35, top-title_h-140, p, 36, "COLUMN_TEXT"))

        # desenhar de baixo para cima: stories[0] é linha inferior.
        for r_i, story in enumerate(stories):
            ymid = bottom + r_i*rh + rh/2.0
            parts.append(_dxf_text(left+80, ymid, story, 24, "COLUMN_TEXT"))
            for c_i, p in enumerate(prumadas):
                cell = work[(work["prumada"].astype(str)==str(p)) & (work["_story_label"].astype(str)==str(story))]
                cx = left + row_header_w + c_i*cw + cw/2.0
                cy = ymid + 60
                if cell.empty:
                    parts.append(_dxf_line(cx-150, cy, cx+150, cy, "COLUMN_TEXT"))
                else:
                    rr = cell.sort_values(by=[c for c in ["utilizacao","as_prov_mm2"] if c in cell.columns], ascending=False).iloc[0]
                    _v60_draw_section(parts, rr, cx, cy, scale=0.75)

        # legenda inferior
        ly = bottom - 260
        parts.append(_dxf_text(left, ly, "Notas: quadro organizado por prumada e piso/tramo, de baixo para cima. Confirmar pormenorização final em peça desenhada.", 28, "COLUMN_TEXT"))

    parts.append("0\nENDSEC\n0\nEOF\n")
    Path(path).write_text("".join(parts), encoding="utf-8")

def _export_dxf_v60(self):
    src = self.df_results if getattr(self, "df_results", pd.DataFrame()) is not None and not self.df_results.empty else getattr(self, "df_summary", pd.DataFrame())
    if src is None or src.empty:
        messagebox.showwarning("Aviso", "Não há resultados para exportar em DXF.")
        return
    path = filedialog.asksaveasfilename(title="Exportar quadro de pilares", defaultextension=".dxf", filetypes=[("DXF", "*.dxf")])
    if not path:
        return
    try:
        self.status_var.set("A exportar quadro de pilares DXF...")
        self.progress_var.set(10.0)
        write_columns_dxf_v60(path, src)
        self.progress_var.set(100.0)
        self.status_var.set(f"DXF exportado: {path}")
    except Exception as err:
        messagebox.showerror("Erro", f"Não foi possível exportar DXF.\n\n{err}")

ColumnsEC2App.export_dxf = _export_dxf_v60

# --------------------------- XLSX ---------------------------
def _write_excel_v60(self, path: str):
    try:
        self.status_var.set("A exportar XLSX...")
        self.progress_var.set(10.0)
        self.update_idletasks()
    except Exception:
        pass

    res = self.df_results if getattr(self, "df_results", None) is not None else pd.DataFrame()
    notes_name = "13_Notas_FIB_MC10" if _v59_get_backend(self) == BACKEND_SC_MC2010 else "13_Notas_EC2"

    def cols(df, names):
        if df is None or df.empty:
            return pd.DataFrame()
        present = [c for c in names if c in df.columns]
        return df[present].copy() if present else pd.DataFrame()

    try:
        import_quality = build_import_quality_cases_v41(self) if "build_import_quality_cases_v41" in globals() else getattr(self, "df_validation", pd.DataFrame())
    except Exception:
        import_quality = getattr(self, "df_validation", pd.DataFrame())
    try:
        mat_assum = material_assumptions_df_v41(res) if "material_assumptions_df_v41" in globals() else pd.DataFrame()
    except Exception:
        mat_assum = pd.DataFrame()
    try:
        surf_points = surface_points_df_v41(res) if "surface_points_df_v41" in globals() else pd.DataFrame()
    except Exception:
        surf_points = pd.DataFrame()
    try:
        schedule = build_pillar_schedule_df_v42(res) if "build_pillar_schedule_df_v42" in globals() else getattr(self, "df_summary", pd.DataFrame())
    except Exception:
        schedule = getattr(self, "df_summary", pd.DataFrame())

    sheets = {
        "00_Info": self._metadata_df(),
        "01_Parametros": self._parameters_df(),
        "02_Entrada_Dados": getattr(self, "df_clean", pd.DataFrame()),
        "03_Pares_Member_Case": getattr(self, "df_pair", pd.DataFrame()),
        "04_Qualidade_Importacao": import_quality,
        "05_Casos_Calculo": getattr(self, "df_calc_input", pd.DataFrame()),
        "06_Resultados": res,
        "07_Resumo_Membros": getattr(self, "df_summary", pd.DataFrame()),
        "08_Falhas": getattr(self, "df_failures", pd.DataFrame()),
        "09_OK": getattr(self, "df_ok", pd.DataFrame()),
        "10_Shortlists": self.build_shortlists_df() if hasattr(self, "build_shortlists_df") else pd.DataFrame(),
        "11_Validacao": getattr(self, "df_validation", pd.DataFrame()),
        "12_Materiais_Assumidos": mat_assum,
        notes_name: getattr(self, "df_notes", pd.DataFrame()) if getattr(self, "df_notes", pd.DataFrame()) is not None and not getattr(self, "df_notes", pd.DataFrame()).empty else self.build_normative_notes(),
        "14_ELS": cols(res, ["prumada","member","case","combination_number","limit_state","service_status","service_case_source","service_combination","service_sigma_c_max_MPa","service_sigma_c_min_MPa","service_sigma_s_max_MPa","service_wk_est_mm","service_wk_lim_mm","service_crack_status","service_method","service_note"]),
        "15_Esf_Transverso": cols(res, ["prumada","member","case","v_ed_y_kN","v_rd_c_y_kN","v_rd_max_y_kN","asw_s_y_req_mm2_per_m","shear_status_y","v_ed_z_kN","v_rd_c_z_kN","v_rd_max_z_kN","asw_s_z_req_mm2_per_m","shear_status_z","shear_backend"]),
        "16_Torcao": cols(res, ["prumada","member","case","mx_ed_kNm","torsion_ratio","t_rd_max_kNm","asw_s_t_req_mm2_per_m","asl_t_req_mm2","torsion_status","torsion_backend"]),
        "17_Pormenorizacao": cols(res, ["prumada","member","case","layout_description","solucao","phi_corner_mm","phi_face_mm","n_total","n_face_y_extra","n_face_z_extra","phi_st_mm","s_st_mm","detailing_status","detailing_blocking_issues","detailing_warnings","detailing_info","detailing_issues","detailing_min_clear_mm"]),
        "18_Superficie_Resumo": cols(res, ["prumada","member","case","n_ed_kN","my_ed_kNm","mz_ed_kNm","mrd_y_kNm","mrd_z_kNm","utilizacao","biaxial_alpha","biaxial_n_ratio","surface_method","surface_points","nmm_capacity_source"]),
        "19_Superficie_Pontos": surf_points,
        "20_Memoria_Calculo": cols(res, ["prumada","member","case","material","b_cm","h_cm","length_m","n_ed_kN","my_i_kNm","my_j_kNm","mz_i_kNm","mz_j_kNm","m01_y_ec2_kNm","m02_y_ec2_kNm","rm_y_signed","curvature_y","m01_z_ec2_kNm","m02_z_ec2_kNm","rm_z_signed","curvature_z","lambda_y","lambda_lim_y","lambda_z","lambda_lim_z","m0e_y_kNm","m2_y_kNm","m0e_z_kNm","m2_z_kNm","as_min_mm2","as_req_mm2","as_max_mm2","as_prov_mm2","utilizacao","status","failure_reason"]),
        "22_Quadro_Pilares": schedule,
        "23_Resumo_Prumadas": getattr(self, "df_summary", pd.DataFrame()),
        "24_Gestao_Falhas": cols(res, ["prumada","member","case","status","failure_severity","failure_type","failure_reason","failure_summary","design_decision","review_priority","failure_action","recommendations"]),
        "25_Propostas_Correcoes": getattr(self, "df_repair_proposals", pd.DataFrame()),
        "30_Mapa_Cobertura": _backend_coverage_df_v60(self),
        "31_Diagnostico_API": _sc_api_diagnostic_df_v60(),
        "32_Escopo_Resultados": cols(res, ["prumada","member","case","code_backend","normative_basis","materials_backend","materials_sources","nmm_capacity_source","shear_backend","torsion_backend","service_backend","second_order_status","detailing_status","status","failure_reason"]),
    }
    remove_names = {"21_Casos_Tipo_Validacao", "33_Opcoes_v5_3", "34_Cobertura_Backend", "35_Diagnostico_SC", "36_Parametros_v5_4"}

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in sheets.items():
            if name in remove_names:
                continue
            if df is None:
                df = pd.DataFrame()
            df.to_excel(writer, sheet_name=name[:31], index=False)
        wb = writer.book
        props = wb.properties
        props.title = APP_NAME
        props.subject = APP_SUBJECT
        props.creator = APP_AUTHOR
        props.keywords = APP_KEYWORDS
        props.category = APP_CATEGORY
        props.description = APP_XLSX_DESCRIPTION
        props.lastModifiedBy = APP_AUTHOR
        try:
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            from openpyxl.utils import get_column_letter
            header_fill = PatternFill("solid", fgColor="1F4E5F")
            header_font = Font(color="FFFFFF", bold=True)
            thin = Side(style="thin", color="D9E2E7")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for ws in wb.worksheets:
                ws.sheet_view.showGridLines = False
                ws.freeze_panes = "A2"
                if ws.max_row >= 1:
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        cell.border = border
                for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 5000)):
                    for cell in row:
                        cell.border = border
                        cell.alignment = Alignment(vertical="top", wrap_text=True)
                for col_idx, col in enumerate(ws.columns, start=1):
                    values = [str(c.value) for c in col[:200] if c.value is not None]
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(max([len(v) for v in values] + [10]) + 2, 58)
            # Hyperlink no repositório.
            if "00_Info" in wb.sheetnames:
                ws = wb["00_Info"]
                for row in range(2, ws.max_row + 1):
                    if str(ws.cell(row=row, column=1).value).strip() == "Autor / Repositório":
                        c = ws.cell(row=row, column=2)
                        c.value = GITHUB_URL
                        c.hyperlink = GITHUB_URL
                        c.style = "Hyperlink"
        except Exception:
            pass
    try:
        self.progress_var.set(100.0)
        self.status_var.set(f"XLSX exportado: {path}")
    except Exception:
        pass

ColumnsEC2App._write_excel = _write_excel_v60

# --------------------------- PDF profissional ---------------------------
def _write_pdf_v60(self, path: str):
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak

    try:
        self.status_var.set("A exportar PDF...")
        self.progress_var.set(10.0)
        self.update_idletasks()
    except Exception:
        pass

    res = getattr(self, "df_results", pd.DataFrame())
    summary = getattr(self, "df_summary", pd.DataFrame())
    if summary is None or summary.empty:
        summary = res
    failures = getattr(self, "df_failures", pd.DataFrame())
    backend = _v59_get_backend(self)

    def esc(x):
        return str(x if x is not None else "").replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="ReportTitleCE2", parent=styles["Title"], alignment=TA_CENTER, fontName="Courier-Bold", fontSize=15, leading=22, spaceAfter=8))
    styles.add(ParagraphStyle(name="SubtitleCE2", parent=styles["Normal"], alignment=TA_CENTER, fontName="Courier", fontSize=9.5, leading=14, textColor=colors.darkgrey, spaceAfter=8))
    styles.add(ParagraphStyle(name="BodyCE2", parent=styles["Normal"], fontName="Courier", fontSize=8.5, leading=12.5, spaceAfter=5))
    styles.add(ParagraphStyle(name="SmallCE2", parent=styles["Normal"], fontName="Courier", fontSize=7, leading=10))
    styles.add(ParagraphStyle(name="SectionCE2", parent=styles["Heading2"], fontName="Courier-Bold", fontSize=11, leading=16, spaceBefore=8, spaceAfter=8))
    styles.add(ParagraphStyle(name="CellCE2", parent=styles["SmallCE2"], alignment=TA_LEFT, fontName="Courier", fontSize=6.6, leading=8.3))

    def tbl(data, widths=None, header=False):
        data2 = [[Paragraph(esc(v), styles["CellCE2"]) for v in row] for row in data]
        t = Table(data2, colWidths=widths, repeatRows=1 if header else 0)
        cmds = [
            ("GRID",(0,0),(-1,-1),0.25,colors.lightgrey),
            ("VALIGN",(0,0),(-1,-1),"TOP"),
            ("LEFTPADDING",(0,0),(-1,-1),3),
            ("RIGHTPADDING",(0,0),(-1,-1),3),
        ]
        if header:
            cmds += [("BACKGROUND",(0,0),(-1,0),colors.HexColor("#EFEFEF")),("FONTNAME",(0,0),(-1,0),"Courier-Bold")]
        t.setStyle(TableStyle(cmds))
        return t

    def df_table(df, cols, max_rows=30, widths=None):
        if df is None or df.empty:
            return Paragraph("Sem dados a apresentar.", styles["BodyCE2"])
        present = [c for c in cols if c in df.columns]
        data = [present]
        for _, r in df.head(max_rows).iterrows():
            row = []
            for c in present:
                v = r.get(c, "")
                if isinstance(v, float):
                    row.append("" if not math.isfinite(v) else f"{v:.2f}")
                else:
                    row.append("" if pd.isna(v) else str(v))
            data.append(row)
        return tbl(data, widths=widths, header=True)

    n_total = len(res) if res is not None else 0
    prumadas = res.get("prumada", res.get("name", res.get("member", pd.Series(dtype=str)))).astype(str).nunique() if res is not None and not res.empty else 0
    n_fail = len(failures) if failures is not None else 0
    n_ok = int((res.get("status", pd.Series(dtype=str)).astype(str) == "OK").sum()) if res is not None and not res.empty else 0
    n_warn = int((res.get("status", pd.Series(dtype=str)).astype(str) == "Aviso").sum()) if res is not None and not res.empty else 0

    doc = SimpleDocTemplate(path, pagesize=A4, rightMargin=12*mm, leftMargin=12*mm, topMargin=13*mm, bottomMargin=13*mm)
    doc.title = APP_NAME
    doc.author = APP_AUTHOR
    doc.subject = APP_SUBJECT
    story = []
    story.append(Paragraph("ColumnsEC2", styles["ReportTitleCE2"]))
    story.append(Paragraph("Relatório técnico de dimensionamento/verificação de pilares", styles["SubtitleCE2"]))

    meta = [
        ["Norma", _backend_reference_v59(self), "Data", datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["Modo", _v53_mode_to_label(_v59_getvar(self, "var_calc_mode", "Dimensionamento")), "Prumadas", prumadas],
        ["Casos avaliados", n_total, "OK / Aviso / Falha", f"{n_ok} / {n_warn} / {n_fail}"],
        ["Classe de aço", _v59_steel_class_label(self), "Betão", _v59_material_classes(self)],
    ]
    story.append(tbl(meta, widths=[31*mm, 78*mm, 31*mm, 50*mm]))
    story.append(Spacer(1,4*mm))

    story.append(Paragraph("1. Enquadramento", styles["SectionCE2"]))
    story.append(Paragraph(_module_description_v60(self), styles["BodyCE2"]))
    story.append(Paragraph("<b>Limitações técnicas:</b> " + esc(_module_limitations_v60(self)), styles["BodyCE2"]))

    story.append(Paragraph("2. Parâmetros de cálculo", styles["SectionCE2"]))
    story.append(df_table(self._parameters_df(), ["Parâmetro","Valor"], max_rows=14, widths=[64*mm, 126*mm]))

    story.append(Paragraph("3. Resultados governantes por prumada", styles["SectionCE2"]))
    res_cols = ["prumada","member","case","material","b_cm","h_cm","n_ed_kN","my_ed_kNm","mz_ed_kNm","solucao","utilizacao","status"]
    story.append(df_table(summary, res_cols, max_rows=34))

    story.append(Paragraph("4. Verificações condicionantes", styles["SectionCE2"]))
    cond_cols = ["prumada","member","case","lambda_y","lambda_z","shear_status_y","shear_status_z","torsion_status","service_status","detailing_status"]
    story.append(df_table(summary, cond_cols, max_rows=30))

    if failures is not None and not failures.empty:
        story.append(PageBreak())
        story.append(Paragraph("5. Falhas e acções recomendadas", styles["SectionCE2"]))
        fail_cols = ["prumada","member","case","failure_type","failure_reason","recommendations","design_decision"]
        story.append(df_table(failures, fail_cols, max_rows=45))

    story.append(PageBreak())
    story.append(Paragraph("6. Notas normativas do módulo seleccionado", styles["SectionCE2"]))
    story.append(df_table(self.build_normative_notes(), ["Tema","Referência","Nota"], max_rows=18, widths=[36*mm, 50*mm, 104*mm]))

    if _sc_backend_active_v52(backend):
        story.append(Paragraph("7. Cobertura do backend", styles["SectionCE2"]))
        cov = _backend_coverage_df_v60(self)
        cov = cov[cov["Norma"].astype(str) == _backend_display_v59(backend)]
        story.append(df_table(cov, ["Verificação","Origem","Estado"], max_rows=30, widths=[50*mm, 80*mm, 60*mm]))

    story.append(Spacer(1,4*mm))
    story.append(Paragraph("A memória completa de cálculo, combinações, shortlists, ELS, esforço transverso, torção, pormenorização e diagnóstico do backend constam do ficheiro XLSX.", styles["SmallCE2"]))

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    def footer(canvas, doc_obj):
        canvas.saveState()
        canvas.setAuthor(APP_AUTHOR)
        canvas.setTitle(APP_NAME)
        canvas.setSubject(APP_SUBJECT)
        canvas.setFont("Courier", 7)
        canvas.setFillColor(colors.grey)
        canvas.drawString(12*mm, 7*mm, f"Columns EC2 | {now_str}")
        canvas.drawRightString(198*mm, 7*mm, f"Página {doc_obj.page}")
        canvas.restoreState()

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    try:
        self.progress_var.set(100.0)
        self.status_var.set(f"PDF exportado: {path}")
    except Exception:
        pass

ColumnsEC2App._write_pdf = _write_pdf_v60


# ============================================================
# ColumnsEC2 v6.1 — DXF: quadro de pilares alinhado por prumada
# ============================================================
APP_VERSION = "v6.2"


def _v61_is_blank(value) -> bool:
    s = str(value).strip()
    return (not s) or s.lower() in {"nan", "none", "nat", "-"}


def _v61_prumada_label(row) -> str:
    """Nome da prumada para o quadro: prioridade a prumada > name > member."""
    for c in ["prumada", "pillar_name", "name", "Name"]:
        try:
            v = row.get(c, "")
            if not _v61_is_blank(v):
                return str(v).strip()
        except Exception:
            pass
    return str(row.get("member", "")).strip() or "-"


def _v61_prumada_sort_key(value):
    s = str(value).strip()
    nums = re.findall(r"\d+", s)
    # P1, P2, P10; se houver texto tipo P1=P3=P4, usa o primeiro número para ordenar.
    first = int(nums[0]) if nums else 10**6
    return (first, s.lower())


def _v61_level_label_and_key(row):
    """Linha do quadro, preferindo piso real; se não existir, usa segment_order/tramo.
    A ordenação é sempre de baixo para cima.
    """
    # 1) Piso/nível explícito, se existir na tabela importada ou nos resultados.
    for c in ["story", "Story", "piso", "Piso", "storey", "Storey", "floor", "Floor", "level", "Level", "pavimento", "Pavimento"]:
        try:
            v = row.get(c, "")
            if not _v61_is_blank(v):
                label = str(v).strip()
                low = label.lower()
                if "cobertura" in low or "roof" in low:
                    return label, (900000, label.lower())
                nums = re.findall(r"-?\d+", label)
                if nums:
                    return label, (int(nums[-1]), label.lower())
                return label, (500000, label.lower())
        except Exception:
            pass

    # 2) Ordem de tramo já criada pelo motor de cálculo.
    try:
        so = safe_float(row.get("segment_order", float("nan")), float("nan"))
        if math.isfinite(so):
            n = int(round(so))
            return f"Tramo {n:02d}", (n, f"tramo {n:02d}")
    except Exception:
        pass

    # 3) Fallback: member. Não é ideal, mas preserva alinhamento possível.
    mem = str(row.get("member", "")).strip()
    nums = re.findall(r"-?\d+", mem)
    if nums:
        n = int(nums[-1])
        return f"Member {mem}", (n, f"member {mem}")
    return "Tramo", (0, "tramo")


def _v61_status_rank(row) -> int:
    status = str(row.get("status", "")).lower()
    sev = str(row.get("failure_severity", "")).lower()
    if "falha" in status or "bloqueante" in sev:
        return 4
    if "aviso" in status or "alta" in str(row.get("review_priority", "")).lower():
        return 3
    if "pré" in status or "pre" in status:
        return 1
    if "ok" in status:
        return 2
    return 0


def _v61_select_cell_result(grp: pd.DataFrame) -> pd.Series:
    """Escolhe a linha a desenhar dentro da célula prumada/piso.
    Usa a verificação mais condicionante; em empate usa maior utilização e maior As.
    """
    work = grp.copy()
    work["_rank_dxf"] = work.apply(_v61_status_rank, axis=1)
    if "utilizacao" in work.columns:
        work["_util_dxf"] = work["utilizacao"].map(lambda x: _finite(x, -1.0))
    else:
        work["_util_dxf"] = -1.0
    if "as_prov_mm2" in work.columns:
        work["_as_dxf"] = work["as_prov_mm2"].map(lambda x: _finite(x, -1.0))
    else:
        work["_as_dxf"] = -1.0
    work = work.sort_values(["_rank_dxf", "_util_dxf", "_as_dxf"], ascending=[False, False, False])
    return work.iloc[0]


def _v61_section_label(r: pd.Series) -> str:
    sol = str(r.get("layout_description", "") or r.get("solucao", "") or "").strip()
    # Limpar expressão para ficar mais parecida com quadro de pilares.
    sol = sol.replace(" (cantos)", "").replace(" + estribos", "\nEstr.")
    return sol[:120]


def _v61_draw_missing(parts, cx, cy):
    parts.append(_dxf_line(cx - 115, cy, cx + 115, cy, "COLUMN_TEXT"))


def _v61_draw_section_cell(parts, r, cx, cy, scale=0.80):
    """Desenha a secção armada dentro da célula do quadro."""
    b = _finite(r.get("b_cm")) * 10.0 * scale
    h = _finite(r.get("h_cm")) * 10.0 * scale
    if b <= 0 or h <= 0:
        _v61_draw_missing(parts, cx, cy)
        return

    left, right = cx - b/2, cx + b/2
    bot, top = cy - h/2, cy + h/2
    parts += [
        _dxf_line(left, bot, right, bot, "COLUMN_CONCRETE"),
        _dxf_line(right, bot, right, top, "COLUMN_CONCRETE"),
        _dxf_line(right, top, left, top, "COLUMN_CONCRETE"),
        _dxf_line(left, top, left, bot, "COLUMN_CONCRETE"),
    ]

    cover = max(_finite(r.get("cover_mm"), 30.0) * scale, 15.0)
    if b > 2*cover and h > 2*cover:
        parts += [
            _dxf_line(left + cover, bot + cover, right - cover, bot + cover, "COLUMN_STIRRUP"),
            _dxf_line(right - cover, bot + cover, right - cover, top - cover, "COLUMN_STIRRUP"),
            _dxf_line(right - cover, top - cover, left + cover, top - cover, "COLUMN_STIRRUP"),
            _dxf_line(left + cover, top - cover, left + cover, bot + cover, "COLUMN_STIRRUP"),
        ]

    pts = []
    try:
        pts = _v56_bar_points_from_result(r)
    except Exception:
        pts = []
    if not pts:
        try:
            pts = [(y, z, _finite(r.get("phi_long_mm"), 10.0)) for y, z in _bar_points_for_result(r)]
        except Exception:
            pts = []
    for y, z, phi in pts:
        parts.append(_dxf_circle(cx + y*scale, cy + z*scale, max(phi*scale/2.0, 2.0), "COLUMN_REBAR"))

    # Cotagem simples em m.
    parts.append(_dxf_line(left, bot - 58, right, bot - 58, "COLUMN_DIM"))
    parts.append(_dxf_text(cx - 35, bot - 105, f"{_finite(r.get('b_cm'))/100:.2f}", 17, "COLUMN_DIM"))
    parts.append(_dxf_line(left - 58, bot, left - 58, top, "COLUMN_DIM"))
    parts.append(_dxf_text(left - 112, cy - 8, f"{_finite(r.get('h_cm'))/100:.2f}", 17, "COLUMN_DIM"))

    # Texto da armadura.
    sol = _v61_section_label(r).split("\n")
    if sol:
        parts.append(_dxf_text(cx - 160, bot - 160, sol[0], 19, "COLUMN_TEXT"))
    st = f"Estr. Ø{int(_finite(r.get('phi_st_mm'), 0))}//{_finite(r.get('s_st_mm'))/10:.1f}"
    if _finite(r.get('phi_st_mm'), 0) > 0:
        parts.append(_dxf_text(cx - 160, bot - 205, st, 18, "COLUMN_TEXT"))


def _v62_has_explicit_level(work: pd.DataFrame) -> bool:
    """Indica se existe uma coluna de piso/nível realmente preenchida.

    Quando não há piso explícito, o quadro deve alinhar por posição relativa na
    prumada: Tramo 01, Tramo 02, Tramo 03, etc., e não pelo número global do member.
    """
    cols = ["story", "Story", "piso", "Piso", "storey", "Storey", "floor", "Floor", "level", "Level", "pavimento", "Pavimento"]
    for c in cols:
        if c in work.columns:
            vals = work[c].map(lambda v: not _v61_is_blank(v))
            if bool(vals.any()):
                return True
    return False


def _v62_member_sort_key(row):
    """Ordenação de tramos dentro da mesma prumada, de baixo para cima."""
    # segment_order é criado pelo programa para preservar a ordem física dos tramos.
    try:
        so = safe_float(row.get("segment_order", float("nan")), float("nan"))
        if math.isfinite(so):
            return (float(so), str(row.get("member", "")))
    except Exception:
        pass
    mem = str(row.get("member", ""))
    nums = re.findall(r"-?\d+", mem)
    if nums:
        return (float(int(nums[-1])), mem)
    return (0.0, mem)


def _v62_assign_relative_levels(work: pd.DataFrame) -> pd.DataFrame:
    """Atribui Tramo 01, Tramo 02, ... dentro de cada prumada quando não há Piso.

    Isto é o que permite alinhar todos os P1 numa coluna, todos os P2 noutra,
    com as linhas do quadro a representarem a posição relativa do tramo.
    """
    out = work.copy()
    out["_dxf_level"] = ""
    out["_dxf_level_key"] = None
    for pr, grp in out.groupby("_dxf_prumada", dropna=False):
        # Primeiro reduz para um registo por member/tramo, escolhendo a linha mais condicionante.
        local_rows = []
        for mem, gmem in grp.groupby("member", dropna=False):
            rr = _v61_select_cell_result(gmem).copy()
            rr["_sort_local"] = _v62_member_sort_key(rr)
            local_rows.append(rr)
        loc = pd.DataFrame(local_rows) if local_rows else grp.copy()
        if loc.empty:
            continue
        loc = loc.sort_values("_sort_local", key=lambda s: s.map(lambda x: x if isinstance(x, tuple) else (0, str(x))))
        member_to_level = {}
        for i, (_, rr) in enumerate(loc.iterrows(), start=1):
            member_to_level[str(rr.get("member", ""))] = (f"Tramo {i:02d}", (i, f"tramo {i:02d}"))
        mask = out["_dxf_prumada"].astype(str) == str(pr)
        for idx in out[mask].index:
            key = str(out.at[idx, "member"])
            label, skey = member_to_level.get(key, ("Tramo 01", (1, "tramo 01")))
            out.at[idx, "_dxf_level"] = label
            out.at[idx, "_dxf_level_key"] = skey
    return out


def _v61_build_pillar_schedule(df: pd.DataFrame) -> Tuple[pd.DataFrame, List[str], List[Tuple[str, Tuple]]]:
    """Agenda do quadro: uma linha por prumada/nível, pronta para DXF.

    Regras:
      1. Se existir Piso/Story/Level na tabela, o quadro usa esses pisos reais.
      2. Se não existir piso, o quadro usa níveis relativos por prumada: Tramo 01,
         Tramo 02, Tramo 03, ...; assim todos os P1 ficam numa coluna, P2 noutra,
         e os tramos ficam alinhados por posição relativa.
    """
    work = df.copy()
    work["_dxf_prumada"] = work.apply(_v61_prumada_label, axis=1)

    if _v62_has_explicit_level(work):
        tmp = work.apply(_v61_level_label_and_key, axis=1)
        work["_dxf_level"] = tmp.map(lambda x: x[0])
        work["_dxf_level_key"] = tmp.map(lambda x: x[1])
    else:
        work = _v62_assign_relative_levels(work)

    rows = []
    for (pr, lv), grp in work.groupby(["_dxf_prumada", "_dxf_level"], dropna=False):
        rr = _v61_select_cell_result(grp).copy()
        rr["_dxf_prumada"] = pr
        rr["_dxf_level"] = lv
        rr["_dxf_level_key"] = grp["_dxf_level_key"].iloc[0]
        rows.append(rr)
    sched = pd.DataFrame(rows) if rows else work
    prumadas = sorted(sched["_dxf_prumada"].astype(str).unique(), key=_v61_prumada_sort_key)

    level_map = {}
    for _, r in sched.iterrows():
        label = str(r.get("_dxf_level", ""))
        key = r.get("_dxf_level_key", (0, label.lower()))
        if label not in level_map or key < level_map[label]:
            level_map[label] = key
    levels = sorted(level_map.items(), key=lambda kv: kv[1])
    return sched, prumadas, levels


def write_columns_dxf_v61(path: str, df: pd.DataFrame):
    """Exporta quadro de pilares por prumada.

    Estrutura do DXF:
      - uma coluna por prumada: P1, P2, P3, ...;
      - uma linha por piso/tramo, organizada de baixo para cima;
      - cada célula contém a secção, armadura longitudinal, estribos e dimensões;
      - células sem pilar naquela prumada/piso recebem um traço.
    """
    parts = ["0\nSECTION\n2\nHEADER\n0\nENDSEC\n", _v60_dxf_layers(), "0\nSECTION\n2\nENTITIES\n"]
    if df is None or df.empty:
        parts.append(_dxf_text(0, 0, "Sem resultados", 50, "COLUMN_TEXT"))
    else:
        sched, prumadas, levels = _v61_build_pillar_schedule(df)
        if not prumadas or not levels:
            parts.append(_dxf_text(0, 0, "Sem dados de prumadas", 50, "COLUMN_TEXT"))
        else:
            # Limites para manter o DXF manejável. O XLSX continua com memória completa.
            prumadas = prumadas[:24]
            levels = levels[:24]
            cw = 1550.0
            rh = 1350.0
            title_h = 260.0
            header_h = 230.0
            row_header_w = 520.0
            x0 = 0.0
            y0 = 0.0
            width = row_header_w + cw * len(prumadas)
            height = title_h + header_h + rh * len(levels)
            left, bottom = x0, y0
            right, top = x0 + width, y0 + height

            parts.append(_dxf_text(left + width * 0.38, top - 140, "QUADRO DE PILARES", 55, "COLUMN_TEXT"))
            parts.append(_dxf_text(left + 18, top - title_h - 88, "Piso", 28, "COLUMN_TEXT"))
            parts.append(_dxf_text(left + 220, top - title_h - 88, "Pilar", 28, "COLUMN_TEXT"))

            # Moldura e grelha.
            parts += [
                _dxf_line(left, bottom, right, bottom, "COLUMN_TABLE"),
                _dxf_line(right, bottom, right, top, "COLUMN_TABLE"),
                _dxf_line(right, top, left, top, "COLUMN_TABLE"),
                _dxf_line(left, top, left, bottom, "COLUMN_TABLE"),
                _dxf_line(left, top - title_h, right, top - title_h, "COLUMN_TABLE"),
                _dxf_line(left, top - title_h - header_h, right, top - title_h - header_h, "COLUMN_TABLE"),
                _dxf_line(left + row_header_w, top - title_h, left + row_header_w, bottom, "COLUMN_TABLE"),
            ]
            for i in range(len(levels) + 1):
                y = bottom + i * rh
                parts.append(_dxf_line(left, y, right, y, "COLUMN_TABLE"))
            for i in range(len(prumadas) + 1):
                x = left + row_header_w + i * cw
                parts.append(_dxf_line(x, bottom, x, top - title_h, "COLUMN_TABLE"))

            # Cabeçalho das prumadas.
            for c_i, pr in enumerate(prumadas):
                x = left + row_header_w + c_i * cw
                parts.append(_dxf_text(x + cw * 0.42, top - title_h - 135, str(pr), 38, "COLUMN_TEXT"))

            # Corpo: levels[0] fica em baixo; último fica em cima.
            index = {}
            for _, r in sched.iterrows():
                index[(str(r.get("_dxf_prumada", "")), str(r.get("_dxf_level", "")))] = r

            for r_i, (level_label, _key) in enumerate(levels):
                ymid = bottom + r_i * rh + rh / 2.0
                parts.append(_dxf_text(left + 70, ymid - 12, str(level_label), 24, "COLUMN_TEXT"))
                for c_i, pr in enumerate(prumadas):
                    cx = left + row_header_w + c_i * cw + cw / 2.0
                    cy = ymid + 75
                    rr = index.get((str(pr), str(level_label)))
                    if rr is None:
                        _v61_draw_missing(parts, cx, cy)
                    else:
                        _v61_draw_section_cell(parts, rr, cx, cy, scale=0.82)

            # Nota/legenda.
            note_y = bottom - 270
            parts.append(_dxf_text(left, note_y, "Notas: quadro organizado por prumada; os elementos da mesma prumada ficam alinhados na mesma coluna. As linhas representam pisos/tramos, de baixo para cima.", 26, "COLUMN_TEXT"))
            parts.append(_dxf_text(left, note_y - 70, "Confirmar pormenorização final, amarrações, emendas e grampos em peça desenhada.", 24, "COLUMN_TEXT"))

    parts.append("0\nENDSEC\n0\nEOF\n")
    Path(path).write_text("".join(parts), encoding="utf-8")


def _export_dxf_v61(self):
    # Usar resultados completos para manter todos os tramos/pisos da prumada, não apenas o resumo governante.
    src = self.df_results if getattr(self, "df_results", pd.DataFrame()) is not None and not self.df_results.empty else getattr(self, "df_summary", pd.DataFrame())
    if src is None or src.empty:
        messagebox.showwarning("Aviso", "Não há resultados para exportar em DXF.")
        return
    path = filedialog.asksaveasfilename(title="Exportar quadro de pilares por prumada", defaultextension=".dxf", filetypes=[("DXF", "*.dxf")])
    if not path:
        return
    try:
        self.status_var.set("A exportar quadro de pilares por prumada...")
        self.progress_var.set(10.0)
        self.update_idletasks()
        write_columns_dxf_v61(path, src)
        self.progress_var.set(100.0)
        self.status_var.set(f"DXF exportado por prumada: {path}")
    except Exception as err:
        messagebox.showerror("Erro", f"Não foi possível exportar DXF.\n\n{err}")


# Substitui a rotina DXF activa.
write_columns_dxf_v60 = write_columns_dxf_v61
write_columns_dxf_v4 = write_columns_dxf_v61
ColumnsEC2App.export_dxf = _export_dxf_v61


# ============================================================
# ColumnsEC2 v6.3 — escolha por menor utilização resistente
# ============================================================
APP_VERSION = "v6.7"
APP_XLSX_DESCRIPTION = (
    "Workbook de cálculo de pilares de betão armado com selecção de armadura por menor utilização "
    "N-My-Mz entre os layouts construtivos admissíveis, mantendo memória completa em XLSX."
)



# ============================================================
# ColumnsEC2 v6.4 — GUI profissional, estratégia de armadura,
# eta_NMyMz, casos governantes, PDF/DXF técnicos e correcção iterativa melhorada
# ============================================================
APP_VERSION = "v6.4_1"
APP_XLSX_DESCRIPTION = (
    "Workbook técnico de dimensionamento de pilares de betão armado, com selecção de armadura "
    "por estratégia configurável, verificação N-My-Mz, casos governantes, pormenorização construtiva, "
    "quadro de pilares em DXF e relatórios PDF de gabinete."
)
ACTIVE_REBAR_STRATEGY_V64 = "equilibrada"
REBAR_TARGET_ETA_V64 = 0.80

# --------------------------- inicialização / estratégia ---------------------------
_old_cd_init_v64 = ColumnDesigner.__init__
def _cd_init_v64(self, *args, **kwargs):
    _old_cd_init_v64(self, *args, **kwargs)
    self.design_strategy = str(globals().get("ACTIVE_REBAR_STRATEGY_V64", "equilibrada"))
ColumnDesigner.__init__ = _cd_init_v64


def _v64_strategy_label(value: str) -> str:
    s = str(value or "equilibrada").lower()
    if s.startswith("econ"):
        return "Económica — menor área de aço que verifica"
    if s.startswith("rob"):
        return "Robusta — menor η_NMyMz"
    return "Equilibrada — η_NMyMz alvo 0.70–0.85"


def _v64_strategy_key(value: str) -> str:
    s = str(value or "equilibrada").lower()
    if "econ" in s:
        return "economica"
    if "rob" in s:
        return "robusta"
    return "equilibrada"

# --------------------------- pós-processamento ---------------------------
def _v64_eta_col(results: pd.DataFrame) -> pd.DataFrame:
    if results is None or results.empty:
        return results
    out = results.copy()
    if "utilizacao" in out.columns:
        out["η_NMyMz"] = pd.to_numeric(out["utilizacao"], errors="coerce")
        out["eta_NMyMz"] = out["η_NMyMz"]
    if "estado_global" not in out.columns and "status" in out.columns:
        out["estado_global"] = out["status"]
    if "estado_pormenorizacao" not in out.columns and "detailing_status" in out.columns:
        out["estado_pormenorizacao"] = out["detailing_status"]
    # Grampos / travamento intermédio
    for idx, r in out.iterrows():
        try:
            ny = int(_finite(r.get("n_bars_y", 0), 0))
            nz = int(_finite(r.get("n_bars_z", 0), 0))
            n_clamps = max(0, ny - 4) + max(0, nz - 4)
            if ny > 2 or nz > 2:
                note = "Prever grampos/cintas intermédias para travamento dos varões de face."
            else:
                note = "Estribo fechado nos quatro cantos; grampos intermédios não necessários pela geometria adoptada."
            out.at[idx, "grampos_intermedios"] = int(n_clamps)
            out.at[idx, "pormenorizacao_construtiva"] = note
            # Texto de solução enriquecido
            sol = str(r.get("solucao", ""))
            if note.startswith("Prever") and "grampos" not in sol.lower():
                out.at[idx, "solucao_completa"] = sol + f" + {max(1,n_clamps)} grampo(s) intermédio(s)"
            else:
                out.at[idx, "solucao_completa"] = sol
        except Exception:
            pass
    return out


def _v64_governing_cases(app) -> pd.DataFrame:
    df = getattr(app, "df_pair", pd.DataFrame())
    selected = getattr(app, "df_calc_input", pd.DataFrame())
    if df is None or df.empty:
        return pd.DataFrame(columns=["Prumada","Member","Case","Critério","NEd [kN]","My [kNm]","Mz [kNm]","Vy [kN]","Vz [kN]","T [kNm]"])
    work = df.copy()
    for c in ["fx","fy","fz","mx","my","mz"]:
        if c not in work.columns:
            work[c] = 0.0
        work[f"_abs_{c}"] = pd.to_numeric(work[c], errors="coerce").abs().fillna(0.0)
    work["_prumada"] = work.get("prumada", work.get("name", work.get("member", ""))).astype(str).replace({"":"-","nan":"-"})
    selected_keys = set()
    if selected is not None and not selected.empty:
        for _, r in selected.iterrows():
            selected_keys.add((str(r.get("member","")), str(r.get("case",""))))
    rows = []
    for (pr, mem), grp in work.groupby(["_prumada", "member"], dropna=False):
        criteria = {
            "Nmax": "_abs_fx", "Mymax": "_abs_my", "Mzmax": "_abs_mz",
            "Vymax": "_abs_fy", "Vzmax": "_abs_fz", "Tmax": "_abs_mx",
        }
        grp = grp.copy()
        grp["_nmy_score"] = 0.20*grp["_abs_fx"] + grp["_abs_my"] + grp["_abs_mz"]
        criteria["N+My+Mz"] = "_nmy_score"
        used=set()
        for crit, col in criteria.items():
            if col not in grp.columns or grp.empty:
                continue
            idx = grp[col].idxmax()
            r = grp.loc[idx]
            key=(str(r.get("member","")), str(r.get("case","")), crit)
            if key in used:
                continue
            used.add(key)
            if selected_keys and (str(r.get("member","")), str(r.get("case",""))) not in selected_keys:
                selected_flag="não incluído na envolvente reduzida"
            else:
                selected_flag="incluído"
            rows.append({
                "Prumada": pr, "Member": r.get("member",""), "Case": r.get("case",""), "Critério": crit,
                "NEd [kN]": r.get("fx", None), "My [kNm]": r.get("my", None), "Mz [kNm]": r.get("mz", None),
                "Vy [kN]": r.get("fy", None), "Vz [kN]": r.get("fz", None), "T [kNm]": r.get("mx", None),
                "Estado na redução": selected_flag,
            })
    return pd.DataFrame(rows)

# --------------------------- GUI ---------------------------
_old_init_v64 = ColumnsEC2App.__init__
def _init_v64(self, *args, **kwargs):
    # Não criar StringVar antes de Tk.__init__().
    # A variável é criada em _build_sidebar_v64, quando a janela raiz já existe.
    _old_init_v64(self, *args, **kwargs)
    if not hasattr(self, "var_rebar_strategy"):
        self.var_rebar_strategy = tk.StringVar(master=self, value="Equilibrada")
ColumnsEC2App.__init__ = _init_v64

_old_build_sidebar_v64 = ColumnsEC2App._build_sidebar
def _build_sidebar_v64(self, parent):
    # Nesta fase Tk.__init__ já foi executado, pelo que é seguro criar StringVar.
    if not hasattr(self, "var_rebar_strategy"):
        self.var_rebar_strategy = tk.StringVar(master=self, value="Equilibrada")
    _old_build_sidebar_v64(self, parent)
    # Ajustar textos longos para uma comunicação mais comercial/técnica.
    def walk(w):
        try:
            txt = w.cget("text")
            if isinstance(txt, str):
                if "Dimensionamento de pilares de betão armado" in txt:
                    w.configure(text="Análise e dimensionamento de pilares de betão armado segundo normas seleccionáveis.")
                if "Ferramenta para importação" in txt:
                    w.configure(text="Motor técnico para verificação N-My-Mz, segunda ordem, pormenorização construtiva e relatórios de projecto.")
                if "Aço fyk" in txt:
                    w.configure(text="Classe de Aço")
        except Exception:
            pass
        for ch in w.winfo_children():
            walk(ch)
    walk(parent)
    frame = ttk.LabelFrame(parent, text="Estratégia de armadura")
    frame.pack(fill="x", pady=(0,8))
    ttk.Label(frame, text="Critério de escolha").grid(row=0, column=0, sticky="w", padx=6, pady=4)
    ttk.Combobox(frame, textvariable=self.var_rebar_strategy, values=["Económica", "Equilibrada", "Robusta"], state="readonly", width=18).grid(row=0, column=1, sticky="ew", padx=6, pady=4)
    ttk.Label(frame, text="Económica: menor As; Equilibrada: η≈0.80; Robusta: menor η.", style="Subtle.TLabel", wraplength=320).grid(row=1, column=0, columnspan=2, sticky="w", padx=6, pady=(0,4))
    frame.columnconfigure(1, weight=1)
ColumnsEC2App._build_sidebar = _build_sidebar_v64

# --------------------------- run design hook ---------------------------
_old_run_design_v64 = ColumnsEC2App.run_design
def _run_design_v64(self):
    if not hasattr(self, "var_rebar_strategy"):
        self.var_rebar_strategy = tk.StringVar(master=self, value="Equilibrada")
    globals()["ACTIVE_REBAR_STRATEGY_V64"] = _v64_strategy_key(self.var_rebar_strategy.get())
    _old_run_design_v64(self)
    def later_v64():
        try:
            if getattr(self, "df_results", pd.DataFrame()) is not None and not self.df_results.empty:
                self.df_results = _v64_eta_col(self.df_results)
                summary_enabled = bool(self.var_summary.get()) if hasattr(self, "var_summary") else True
                self.df_summary = self.build_summary_by_member(self.df_results) if summary_enabled else pd.DataFrame()
                self.df_failures = _v59_failure_rows(self.df_results) if "_v59_failure_rows" in globals() else self.df_results[self.df_results.get("status", pd.Series(index=self.df_results.index)).astype(str).eq("Falha")].copy()
                self.df_governing_cases = _v64_governing_cases(self)
                for tree, df in [(self.tree_results,self.df_results),(self.tree_summary,self.df_summary),(self.tree_failures,self.df_failures),(self.tree_shortlists,self.build_shortlists_df())]:
                    try: self.show_df(tree, df)
                    except Exception: pass
                try: self.update_report()
                except Exception: pass
                self.status_var.set(self.status_var.get() + f" | Armadura: {_v64_strategy_label(globals().get('ACTIVE_REBAR_STRATEGY_V64'))}")
        except Exception:
            pass
    try:
        self.after(2200, later_v64)
    except Exception:
        pass
ColumnsEC2App.run_design = _run_design_v64

# --------------------------- resumo por membro: usar η_NMyMz ---------------------------
_old_build_summary_v64 = ColumnsEC2App.build_summary_by_member
def _build_summary_by_member_v64(self, results: pd.DataFrame) -> pd.DataFrame:
    res = _v64_eta_col(results)
    if res is None or res.empty:
        return pd.DataFrame()
    tmp = res.copy()
    # Prioridade: falha > aviso > pré > OK, e dentro da prumada seleccionar caso condicionante por eta/esforços.
    status_rank = {"Falha":0, "Aviso":1, "Pré-dimensionado":2, "OK":3}
    tmp["_status_rank"] = tmp.get("status", pd.Series(index=tmp.index, dtype=str)).map(status_rank).fillna(4)
    eta_col = "η_NMyMz" if "η_NMyMz" in tmp.columns else "utilizacao"
    tmp["_eta_sort"] = pd.to_numeric(tmp.get(eta_col, pd.Series(index=tmp.index)), errors="coerce").fillna(-1.0)
    keys = [c for c in ["prumada", "name", "member"] if c in tmp.columns]
    if "prumada" in tmp.columns:
        group_cols=["prumada"]
    elif "name" in tmp.columns:
        group_cols=["name"]
    else:
        group_cols=["member"]
    tmp = tmp.sort_values(by=group_cols+["_status_rank","_eta_sort"], ascending=[True]*len(group_cols)+[True,False])
    return tmp.groupby(group_cols, as_index=False).first().drop(columns=["_status_rank","_eta_sort"], errors="ignore")
ColumnsEC2App.build_summary_by_member = _build_summary_by_member_v64

# --------------------------- Excel ---------------------------
_old_write_excel_v64 = ColumnsEC2App._write_excel
def _write_excel_v64(self, path: str):
    self.df_results = _v64_eta_col(getattr(self, "df_results", pd.DataFrame()))
    self.df_summary = _v64_eta_col(getattr(self, "df_summary", pd.DataFrame()))
    self.df_governing_cases = _v64_governing_cases(self)
    _old_write_excel_v64(self, path)
    # Acrescentar/substituir folha Casos_Governantes e parâmetros de estratégia.
    try:
        with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            self.df_governing_cases.to_excel(writer, sheet_name="04_Casos_Governantes", index=False)
            extra = pd.DataFrame([
                ["Estratégia de armadura", _v64_strategy_label(globals().get("ACTIVE_REBAR_STRATEGY_V64", "equilibrada"))],
                ["η_NMyMz alvo", "0.80; intervalo recomendado 0.70–0.85 na estratégia equilibrada"],
                ["Unidades DXF", "milímetros"],
            ], columns=["Parâmetro", "Valor"])
            extra.to_excel(writer, sheet_name="01B_Estrategia", index=False)
    except Exception:
        pass
ColumnsEC2App._write_excel = _write_excel_v64

# --------------------------- PDF técnico ---------------------------
def _v64_pdf_style_table(header=True):
    from reportlab.lib import colors
    from reportlab.platypus import TableStyle
    cmds = [
        ("GRID", (0,0), (-1,-1), 0.25, colors.HexColor("#D0D7DE")),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("FONTNAME", (0,0), (-1,-1), "Courier"),
        ("FONTSIZE", (0,0), (-1,-1), 7),
        ("LEFTPADDING", (0,0), (-1,-1), 3),
        ("RIGHTPADDING", (0,0), (-1,-1), 3),
        ("TOPPADDING", (0,0), (-1,-1), 3),
        ("BOTTOMPADDING", (0,0), (-1,-1), 3),
    ]
    if header:
        cmds += [("BACKGROUND", (0,0), (-1,0), colors.HexColor("#F0F2F4")), ("FONTNAME", (0,0), (-1,0), "Courier-Bold")]
    return TableStyle(cmds)


def _v64_pdf_df_table(df, cols, max_rows=30, col_width_total_mm=270):
    from reportlab.platypus import Table, Paragraph
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    p = ParagraphStyle(name="CellV64", fontName="Courier", fontSize=7, leading=9)
    if df is None or df.empty:
        data = [[Paragraph("Sem dados", p)]]
        t = Table(data, colWidths=[col_width_total_mm*mm]); t.setStyle(_v64_pdf_style_table(False)); return t
    present=[c for c in cols if c in df.columns]
    data=[[Paragraph(str(c), p) for c in present]]
    for _, r in df.head(max_rows).iterrows():
        row=[]
        for c in present:
            v=r.get(c,"")
            if isinstance(v, float): txt="" if not math.isfinite(v) else f"{v:.3f}"
            else: txt="" if pd.isna(v) else str(v)
            txt=txt.replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")
            row.append(Paragraph(txt[:420], p))
        data.append(row)
    widths=[(col_width_total_mm*mm)/max(1,len(present))]*max(1,len(present))
    t=Table(data, colWidths=widths, repeatRows=1); t.setStyle(_v64_pdf_style_table(True)); return t


def _write_pdf_v64(self, path: str):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, PageBreak
    self.df_results = _v64_eta_col(getattr(self, "df_results", pd.DataFrame()))
    self.df_summary = _v64_eta_col(getattr(self, "df_summary", pd.DataFrame()))
    styles=getSampleStyleSheet()
    styles.add(ParagraphStyle(name="TitleV64", parent=styles["Title"], fontName="Courier-Bold", fontSize=14, leading=18, alignment=1))
    styles.add(ParagraphStyle(name="SubV64", parent=styles["Normal"], fontName="Courier", fontSize=9, leading=12, alignment=1, textColor=colors.darkgrey))
    styles.add(ParagraphStyle(name="H2V64", parent=styles["Heading2"], fontName="Courier-Bold", fontSize=11, leading=15, spaceBefore=8, spaceAfter=8))
    styles.add(ParagraphStyle(name="BodyV64", parent=styles["Normal"], fontName="Courier", fontSize=8.5, leading=12))
    doc=SimpleDocTemplate(path, pagesize=landscape(A4), leftMargin=12*mm, rightMargin=12*mm, topMargin=12*mm, bottomMargin=12*mm)
    doc.title=APP_NAME; doc.author=APP_AUTHOR; doc.subject=APP_SUBJECT
    res=self.df_results if self.df_results is not None else pd.DataFrame()
    summ=self.df_summary if self.df_summary is not None and not self.df_summary.empty else res
    n_total=len(res); n_ok=int(res.get("status",pd.Series(dtype=str)).astype(str).eq("OK").sum()) if not res.empty else 0
    n_warn=int(res.get("status",pd.Series(dtype=str)).astype(str).eq("Aviso").sum()) if not res.empty else 0
    n_fail=int(res.get("status",pd.Series(dtype=str)).astype(str).eq("Falha").sum()) if not res.empty else 0
    story=[Paragraph("Columns EC2", styles["TitleV64"]), Paragraph("Relatório técnico de dimensionamento de pilares", styles["SubV64"]), Spacer(1,4*mm)]
    meta=[["Norma", _v59_norm_reference(self) if "_v59_norm_reference" in globals() else "Eurocódigo 2", "Estratégia", _v64_strategy_label(globals().get("ACTIVE_REBAR_STRATEGY_V64"))],
          ["Casos", str(n_total), "OK / Aviso / Falha", f"{n_ok} / {n_warn} / {n_fail}"],
          ["Classe de aço", _v59_getvar(self,"var_fyk","A500") if "_v59_getvar" in globals() else self.var_fyk.get(), "Data", datetime.now().strftime("%Y-%m-%d %H:%M")]]
    t=Table(meta, colWidths=[35*mm,100*mm,40*mm,95*mm]); t.setStyle(_v64_pdf_style_table(False)); story.append(t); story.append(Spacer(1,5*mm))
    story.append(Paragraph("1. Síntese executiva", styles["H2V64"]))
    story.append(Paragraph("O presente relatório resume as verificações governantes por prumada. A memória completa, casos de cálculo, shortlists, verificações complementares e metadados encontram-se no ficheiro Excel exportado.", styles["BodyV64"]))
    story.append(Spacer(1,3*mm))
    cols=["prumada","member","case","material","b_cm","h_cm","n_ed_kN","my_ed_kNm","mz_ed_kNm","as_prov_mm2","solucao_completa","η_NMyMz","status"]
    story.append(Paragraph("2. Resultados governantes por prumada", styles["H2V64"]))
    story.append(_v64_pdf_df_table(summ, cols, max_rows=32))
    gov=_v64_governing_cases(self)
    if gov is not None and not gov.empty:
        story.append(PageBreak()); story.append(Paragraph("3. Casos governantes seleccionados", styles["H2V64"]))
        story.append(_v64_pdf_df_table(gov, ["Prumada","Member","Case","Critério","NEd [kN]","My [kNm]","Mz [kNm]","T [kNm]","Estado na redução"], max_rows=45))
    fails = self.df_failures if getattr(self,"df_failures",pd.DataFrame()) is not None else pd.DataFrame()
    if fails is not None and not fails.empty:
        story.append(PageBreak()); story.append(Paragraph("4. Falhas e avisos", styles["H2V64"]))
        story.append(_v64_pdf_df_table(_v64_eta_col(fails), ["prumada","member","case","status","failure_type","failure_reason","failure_warnings","recommendations"], max_rows=45))
    vt_cols=["prumada","member","case","v_ed_y_kN","v_rd_max_y_kN","shear_status_y","v_ed_z_kN","v_rd_max_z_kN","shear_status_z","mx_ed_kNm","torsion_status"]
    if res is not None and not res.empty and any(c in res.columns for c in vt_cols):
        story.append(PageBreak()); story.append(Paragraph("5. Verificações complementares", styles["H2V64"]))
        story.append(_v64_pdf_df_table(res, vt_cols, max_rows=45))
    story.append(Spacer(1,5*mm)); story.append(Paragraph("Observação: confirmar em peça desenhada as amarrações, emendas, grampos, sobreposições e compatibilidade construtiva final.", styles["BodyV64"]))
    def footer(canvas, doc_obj):
        canvas.saveState(); canvas.setFont("Courier",7); canvas.setFillColor(colors.grey)
        canvas.drawString(12*mm,7*mm,f"Columns EC2 | {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        canvas.drawRightString(285*mm,7*mm,f"Página {doc_obj.page}"); canvas.restoreState()
    doc.build(story,onFirstPage=footer,onLaterPages=footer)
ColumnsEC2App._write_pdf = _write_pdf_v64

# --------------------------- DXF refinado por prumada em mm ---------------------------
# Usa a rotina v6.1/v6.2 como base, mas força resultados pós-processados e etiqueta unidades.
_old_write_dxf_v64 = write_columns_dxf_v61 if "write_columns_dxf_v61" in globals() else write_columns_dxf_v4
def write_columns_dxf_v64(path: str, df: pd.DataFrame):
    data=_v64_eta_col(df)
    _old_write_dxf_v64(path, data)
    # Acrescentar uma nota de unidades ao fim do ficheiro antes do EOF.
    try:
        p=Path(path); txt=p.read_text(encoding="utf-8")
        note=_dxf_text(0, -900, "UNIDADES: mm | Quadro organizado por prumada: P1, P2, P3...; tramos de baixo para cima.", 25, "COLUMN_TEXT")
        txt=txt.replace("0\nENDSEC\n0\nEOF\n", note+"0\nENDSEC\n0\nEOF\n")
        p.write_text(txt, encoding="utf-8")
    except Exception:
        pass
write_columns_dxf_v61 = write_columns_dxf_v64
write_columns_dxf_v4 = write_columns_dxf_v64

def _export_dxf_v64(self):
    src = self.df_results if getattr(self,"df_results",pd.DataFrame()) is not None and not self.df_results.empty else getattr(self,"df_summary",pd.DataFrame())
    if src is None or src.empty:
        messagebox.showwarning("Aviso", "Não há resultados para exportar em DXF.")
        return
    path=filedialog.asksaveasfilename(title="Exportar quadro de pilares por prumada [mm]", defaultextension=".dxf", filetypes=[("DXF","*.dxf")])
    if not path: return
    try:
        self.status_var.set("A exportar quadro de pilares por prumada [mm]..."); self.progress_var.set(15); self.update_idletasks()
        write_columns_dxf_v64(path, src)
        self.progress_var.set(100); self.status_var.set(f"DXF exportado: {path}")
    except Exception as err:
        messagebox.showerror("Erro", f"Não foi possível exportar DXF.\n\n{err}")
ColumnsEC2App.export_dxf = _export_dxf_v64

# --------------------------- correcção iterativa melhorada ---------------------------
def _v64_generate_repair_proposal(app, row: pd.Series) -> dict:
    old_strategy=globals().get("ACTIVE_REBAR_STRATEGY_V64","equilibrada")
    globals()["ACTIVE_REBAR_STRATEGY_V64"]="robusta"
    designer = ColumnDesigner(
        cover_mm=safe_float(_v59_getvar(app,"var_cover","35") if "_v59_getvar" in globals() else app.var_cover.get(),35.0),
        fyk=safe_float(_v59_getvar(app,"var_fyk","A500") if "_v59_getvar" in globals() else app.var_fyk.get(),500.0),
        phi_eff=safe_float(_v59_getvar(app,"var_phi_eff","2.0") if "_v59_getvar" in globals() else "2.0",2.0),
        l0y_factor=safe_float(_v59_getvar(app,"var_l0y","1.0") if "_v59_getvar" in globals() else app.var_l0y.get(),1.0),
        l0z_factor=safe_float(_v59_getvar(app,"var_l0z","1.0") if "_v59_getvar" in globals() else app.var_l0z.get(),1.0),
        calc_mode="rigoroso",
    )
    attempts=[]
    try:
        for scale in [1.00,1.05,1.10,1.15,1.20,1.25,1.30,1.40,1.50]:
            rr=_v60_row_for_repair(row, scale=scale) if "_v60_row_for_repair" in globals() else row.copy()
            try:
                out=designer.design_one(rr)
                out=_v64_eta_col(pd.DataFrame([out])).iloc[0].to_dict()
                out["repair_scale_section"]=scale
                out["repair_strategy"]="robusta + modo rigoroso" + ("; aumento de secção" if scale>1.0 else "; secção mantida")
                out["repair_note"]="Proposta gerada por reavaliação do catálogo construtivo admissível."
                out["_target_index"]=row.name
                attempts.append(out)
                if str(out.get("status",""))=="OK":
                    out["repair_result"]="Proposta OK"
                    return out
            except Exception as e:
                attempts.append({"status":"Falha", "repair_scale_section":scale, "repair_note":str(e), "_target_index":row.name})
        best=None
        for a in attempts:
            if str(a.get("status",""))=="Aviso": best=a; break
        if best is None and attempts: best=attempts[-1]
        if best is None: best=dict(row)
        best["repair_result"]="Sem proposta OK" if str(best.get("status",""))!="Aviso" else "Proposta com aviso"
        best["repair_note"]="Não foi encontrada proposta OK; rever esforços, secção, comprimento efectivo ou critérios de pormenorização."
        best["_target_index"]=row.name
        return best
    finally:
        globals()["ACTIVE_REBAR_STRATEGY_V64"]=old_strategy


def _repair_failures_interactive_v64(self):
    if self.df_results is None or self.df_results.empty:
        messagebox.showwarning("Aviso", "Execute primeiro o cálculo."); return
    failures = _v59_failure_rows(self.df_results) if "_v59_failure_rows" in globals() else self.df_results[self.df_results.get("status",pd.Series(index=self.df_results.index)).astype(str).eq("Falha")]
    if failures.empty:
        messagebox.showinfo("Correcção iterativa", "Não foram detectadas falhas bloqueantes."); return
    if not messagebox.askyesno("Correcção iterativa", "Gerar propostas de correcção para as falhas bloqueantes?\n\n" + (_v59_failure_summary_text(failures,10) if "_v59_failure_summary_text" in globals() else str(len(failures)) + " falhas")):
        return
    proposals=[]; total=len(failures)
    for k,(idx,row) in enumerate(failures.iterrows(), start=1):
        self.progress_var.set(100*k/max(total,1)); self.status_var.set(f"A gerar proposta {k}/{total}..."); self.update_idletasks()
        prop=_v64_generate_repair_proposal(self,row); prop["_target_index"]=idx; proposals.append(prop)
    self.df_repair_proposals=pd.DataFrame(proposals)
    ok=int(self.df_repair_proposals.get("status",pd.Series(dtype=str)).astype(str).eq("OK").sum()) if not self.df_repair_proposals.empty else 0
    warn=int(self.df_repair_proposals.get("status",pd.Series(dtype=str)).astype(str).eq("Aviso").sum()) if not self.df_repair_proposals.empty else 0
    fail=len(self.df_repair_proposals)-ok-warn
    if not messagebox.askyesno("Propostas geradas", f"OK: {ok}\nAviso: {warn}\nSem solução: {fail}\n\nAplicar apenas as propostas OK?"):
        self.status_var.set(f"Propostas geradas, não aplicadas: {ok} OK; {warn} avisos; {fail} sem solução."); return
    res=self.df_results.copy(); applied=0
    for _,p in self.df_repair_proposals.iterrows():
        idx=p.get("_target_index",None)
        if idx in res.index and str(p.get("status",""))=="OK":
            for c in self.df_repair_proposals.columns:
                if c != "_target_index": res.at[idx,c]=p.get(c)
            res.at[idx,"auto_repair_applied"]="Sim"; applied+=1
    self.df_results=_v64_eta_col(res)
    self.df_summary=self.build_summary_by_member(self.df_results)
    self.df_failures=_v59_failure_rows(self.df_results) if "_v59_failure_rows" in globals() else self.df_results[self.df_results.get("status",pd.Series(index=self.df_results.index)).astype(str).eq("Falha")]
    for tree,df in [(self.tree_results,self.df_results),(self.tree_summary,self.df_summary),(self.tree_failures,self.df_failures),(self.tree_shortlists,self.build_shortlists_df())]:
        try:self.show_df(tree,df)
        except Exception:pass
    try:self.update_report()
    except Exception:pass
    self.status_var.set(f"Correcção iterativa concluída: {applied} proposta(s) aplicada(s).")
ColumnsEC2App.repair_failures_interactive = _repair_failures_interactive_v64



# ============================================================
# ColumnsEC2 v6.5 — estados separados por módulo
# ============================================================
APP_VERSION = "v6.7"


def _v65_txt(value):
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return str(value or "").strip()


def _v65_norm_state(raw, default="Não avaliado"):
    t = _v65_txt(raw).lower()
    if not t:
        return default
    if any(k in t for k in ["não conforme", "nao conforme", "falha", "vrd,max", "trd,max", "impossível", "impossivel", "bloqueante"]):
        return "Falha"
    if any(k in t for k in ["requer", "verificar", "aviso", "não verificado", "nao verificado", "não avaliado", "nao avaliado", "combinação indicada não encontrada", "combinacao indicada nao encontrada"]):
        return "Aviso"
    if any(k in t for k in ["pré", "pre-"]):
        return "Pré-dimensionado"
    if any(k in t for k in ["informativo", "não conclusivo", "nao conclusivo"]):
        return "Informativo"
    if any(k in t for k in ["ok", "sem aviso", "sem torção", "sem torcao", "desprez", "não condicionante", "nao condicionante"]):
        return "OK"
    return default


def _v65_status_from_row(row):
    old_status = _v65_txt(row.get("status", ""))
    failure_reason = _v65_txt(row.get("failure_reason", ""))
    failure_type = _v65_txt(row.get("failure_type", ""))
    eta = row.get("η_NMyMz", row.get("eta_NMyMz", row.get("utilizacao", None)))

    # 1) Estado resistente N-My-Mz
    resistant = "Não avaliado"
    if old_status == "Pré-dimensionado":
        resistant = "Pré-dimensionado"
    elif "biaxial" in failure_reason.lower() or "resistencia_biaxial" in failure_type.lower() or "resistência biaxial" in failure_reason.lower():
        resistant = "Falha"
    elif _v65_norm_state(old_status, "") == "Falha" and not any(k in failure_reason.lower() for k in ["pormenor", "torção", "torcao", "transverso", "corte", "vr"]) :
        resistant = "Falha"
    elif eta is not None and _v65_txt(eta) != "":
        try:
            resistant = "OK" if float(eta) <= 1.0 + 1e-9 else "Falha"
        except Exception:
            resistant = "OK" if old_status in ["OK", "Aviso"] else _v65_norm_state(old_status, "Não avaliado")
    elif old_status in ["OK", "Aviso"]:
        resistant = "OK"

    # 2) Corte em cada direcção e corte global
    corte_y = _v65_norm_state(row.get("shear_status_y", row.get("estado_corte_y", "")))
    corte_z = _v65_norm_state(row.get("shear_status_z", row.get("estado_corte_z", "")))
    if "Falha" in [corte_y, corte_z]:
        corte = "Falha"
    elif "Aviso" in [corte_y, corte_z]:
        corte = "Aviso"
    elif corte_y == "Não avaliado" and corte_z == "Não avaliado":
        corte = "Não avaliado"
    else:
        corte = "OK"

    # 3) Torção
    tor = _v65_norm_state(row.get("torsion_status", row.get("estado_torcao", "")))
    # Torção nula/desprezável deve ser OK, não aviso.
    tor_txt = _v65_txt(row.get("torsion_status", "")).lower()
    if any(k in tor_txt for k in ["desprez", "sem tor", "não condicionante", "nao condicionante"]):
        tor = "OK"

    # 4) ELS
    els = _v65_norm_state(row.get("service_status", row.get("estado_els", "")), default="Não avaliado")
    if els == "Informativo":
        els = "Aviso"

    # 5) Pormenorização
    det_block = _v65_txt(row.get("detailing_blocking_issues", ""))
    det_warn = _v65_txt(row.get("detailing_warnings", ""))
    det_status_raw = row.get("detailing_status", row.get("estado_pormenorizacao", ""))
    det = _v65_norm_state(det_status_raw, default="OK" if not det_block and not det_warn else "Aviso")
    if det_block and det_block not in ["-", "nan", "None"]:
        det = "Falha"
    elif det_warn and det_warn not in ["-", "nan", "None"] and det != "Falha":
        det = "Aviso"
    # Falha antiga explicitamente de pormenorização deve transferir-se para este módulo.
    if ("pormenor" in failure_reason.lower() or "pormenorizacao" in failure_type.lower()) and old_status == "Falha":
        # Se a questão for apenas aviso construtivo, manter Aviso; se existir blocking, Falha.
        det = "Falha" if det_block and det_block not in ["-", "nan", "None"] else "Aviso"

    # Estado global: só falha por módulos bloqueantes; avisos construtivos/ELS/corte condicionante ficam como Aviso.
    blocking_states = [resistant, corte, tor, det]
    warning_states = [resistant, corte, tor, els, det]
    if "Falha" in blocking_states:
        global_state = "Falha"
    elif "Pré-dimensionado" in warning_states:
        global_state = "Pré-dimensionado"
    elif "Aviso" in warning_states:
        global_state = "Aviso"
    else:
        global_state = "OK"

    # Construir síntese técnica
    notes = []
    if resistant == "Falha":
        notes.append("resistência N-My-Mz não verificada")
    if corte == "Aviso":
        notes.append("corte requer verificação/dimensionamento de armadura transversal")
    if corte == "Falha":
        notes.append("corte excede limite resistente")
    if tor == "Aviso":
        notes.append("torção requer verificação/dimensionamento complementar")
    if tor == "Falha":
        notes.append("torção excede limite resistente")
    if els == "Aviso":
        notes.append("ELS informativo/não conclusivo ou a verificar")
    if det == "Aviso":
        notes.append("pormenorização construtiva a confirmar")
    if det == "Falha":
        notes.append("pormenorização bloqueante")

    return {
        "estado_resistente": resistant,
        "estado_corte_y": corte_y,
        "estado_corte_z": corte_z,
        "estado_corte": corte,
        "estado_torcao": tor,
        "estado_els": els,
        "estado_pormenorizacao": det,
        "estado_global": global_state,
        "decisao_tecnica": "; ".join(notes) if notes else "Sem reservas relevantes no âmbito das verificações efectuadas.",
    }


def _v65_apply_module_statuses(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame() if df is None else df
    out = df.copy()
    # Garantir coluna eta com símbolo.
    try:
        if "_v64_eta_col" in globals():
            out = _v64_eta_col(out)
    except Exception:
        pass
    rows = []
    for _, r in out.iterrows():
        rows.append(_v65_status_from_row(r))
    st = pd.DataFrame(rows, index=out.index)
    for c in st.columns:
        out[c] = st[c]
    # manter compatibilidade: status passa a representar o estado global
    out["status"] = out["estado_global"]
    if "η_NMyMz" not in out.columns and "utilizacao" in out.columns:
        out["η_NMyMz"] = out["utilizacao"]
    if "eta_NMyMz" not in out.columns and "η_NMyMz" in out.columns:
        out["eta_NMyMz"] = out["η_NMyMz"]
    return out


def _v65_failures(df: pd.DataFrame) -> pd.DataFrame:
    d = _v65_apply_module_statuses(df)
    if d.empty or "estado_global" not in d.columns:
        return pd.DataFrame()
    return d[d["estado_global"].astype(str).eq("Falha")].copy()


def _v65_warnings(df: pd.DataFrame) -> pd.DataFrame:
    d = _v65_apply_module_statuses(df)
    if d.empty or "estado_global" not in d.columns:
        return pd.DataFrame()
    return d[d["estado_global"].astype(str).eq("Aviso")].copy()


def _v65_module_status_table(df: pd.DataFrame) -> pd.DataFrame:
    d = _v65_apply_module_statuses(df)
    cols = [
        "prumada", "name", "member", "case", "estado_global", "estado_resistente",
        "estado_corte", "estado_torcao", "estado_els", "estado_pormenorizacao",
        "η_NMyMz", "solucao_completa", "solucao", "decisao_tecnica",
        "failure_reason", "failure_warnings", "detailing_blocking_issues", "detailing_warnings"
    ]
    return d[[c for c in cols if c in d.columns]].copy() if not d.empty else pd.DataFrame()


# --------------------------- run design hook v6.5 ---------------------------
_old_run_design_v65 = ColumnsEC2App.run_design

def _run_design_v65(self):
    _old_run_design_v65(self)
    def finalize_v65():
        try:
            if getattr(self, "df_results", pd.DataFrame()) is not None and not self.df_results.empty:
                self.df_results = _v65_apply_module_statuses(self.df_results)
                self.df_summary = self.build_summary_by_member(self.df_results) if hasattr(self, "build_summary_by_member") else pd.DataFrame()
                self.df_summary = _v65_apply_module_statuses(self.df_summary) if self.df_summary is not None and not self.df_summary.empty else self.df_summary
                self.df_failures = _v65_failures(self.df_results)
                self.df_warnings = _v65_warnings(self.df_results)
                self.df_module_status = _v65_module_status_table(self.df_results)
                for tree, df in [
                    (getattr(self, "tree_results", None), self.df_results),
                    (getattr(self, "tree_summary", None), self.df_summary),
                    (getattr(self, "tree_failures", None), pd.concat([self.df_failures, self.df_warnings], ignore_index=True) if not self.df_warnings.empty else self.df_failures),
                    (getattr(self, "tree_shortlists", None), self.build_shortlists_df() if hasattr(self, "build_shortlists_df") else pd.DataFrame()),
                ]:
                    try:
                        if tree is not None:
                            self.show_df(tree, df)
                    except Exception:
                        pass
                try:
                    self.update_report()
                except Exception:
                    pass
                n_ok = int(self.df_results["estado_global"].astype(str).eq("OK").sum())
                n_warn = int(self.df_results["estado_global"].astype(str).eq("Aviso").sum())
                n_fail = int(self.df_results["estado_global"].astype(str).eq("Falha").sum())
                self.status_var.set(f"Cálculo concluído: {len(self.df_results)} casos | OK: {n_ok} | Avisos: {n_warn} | Falhas: {n_fail}.")
        except Exception as e:
            try:
                self.status_var.set(f"Cálculo concluído; aviso no pós-processamento de estados: {e}")
            except Exception:
                pass
    # O cálculo corre em thread; aplicar o pós-processamento depois de a thread terminar.
    try:
        def wait_and_finalize():
            th = getattr(self, "analysis_thread", None)
            if th is not None and getattr(th, "is_alive", lambda: False)():
                self.after(500, wait_and_finalize)
            else:
                finalize_v65()
        self.after(700, wait_and_finalize)
    except Exception:
        pass

ColumnsEC2App.run_design = _run_design_v65


# --------------------------- summary v6.5 ---------------------------
_old_build_summary_v65 = ColumnsEC2App.build_summary_by_member

def _build_summary_by_member_v65(self, results: pd.DataFrame) -> pd.DataFrame:
    res = _v65_apply_module_statuses(results)
    if res is None or res.empty:
        return pd.DataFrame()
    tmp = res.copy()
    rank = {"Falha": 0, "Aviso": 1, "Pré-dimensionado": 2, "OK": 3, "Informativo": 4, "Não avaliado": 5}
    tmp["_estado_rank"] = tmp.get("estado_global", tmp.get("status", pd.Series(index=tmp.index))).map(rank).fillna(6)
    eta_col = "η_NMyMz" if "η_NMyMz" in tmp.columns else "eta_NMyMz" if "eta_NMyMz" in tmp.columns else "utilizacao"
    tmp["_eta_sort"] = pd.to_numeric(tmp.get(eta_col, pd.Series(index=tmp.index)), errors="coerce").fillna(-1.0)
    if "prumada" in tmp.columns:
        group_cols = ["prumada"]
    elif "name" in tmp.columns:
        group_cols = ["name"]
    else:
        group_cols = ["member"]
    tmp = tmp.sort_values(by=group_cols + ["_estado_rank", "_eta_sort"], ascending=[True]*len(group_cols)+[True, False])
    return tmp.groupby(group_cols, as_index=False).first().drop(columns=["_estado_rank", "_eta_sort"], errors="ignore")

ColumnsEC2App.build_summary_by_member = _build_summary_by_member_v65


# --------------------------- shortlists/report view v6.5 ---------------------------
_old_build_shortlists_v65 = ColumnsEC2App.build_shortlists_df

def _build_shortlists_df_v65(self) -> pd.DataFrame:
    base = _old_build_shortlists_v65(self)
    if getattr(self, "df_results", pd.DataFrame()) is not None and not self.df_results.empty:
        d = _v65_apply_module_statuses(self.df_results)
        cols = ["prumada", "member", "case", "estado_global", "estado_resistente", "estado_pormenorizacao", "η_NMyMz", "shortlist_text", "decisao_tecnica"]
        return d[[c for c in cols if c in d.columns]].copy()
    return base
ColumnsEC2App.build_shortlists_df = _build_shortlists_df_v65


_old_update_report_v65 = ColumnsEC2App.update_report

def _update_report_v65(self):
    try:
        if getattr(self, "df_results", pd.DataFrame()) is None or self.df_results.empty:
            return _old_update_report_v65(self)
        self.df_results = _v65_apply_module_statuses(self.df_results)
        source = self.df_summary if getattr(self, "df_summary", pd.DataFrame()) is not None and not self.df_summary.empty else self.df_results
        source = _v65_apply_module_statuses(source)
        self.report_txt.delete("1.0", "end")
        n_ok = int(self.df_results["estado_global"].astype(str).eq("OK").sum())
        n_warn = int(self.df_results["estado_global"].astype(str).eq("Aviso").sum())
        n_fail = int(self.df_results["estado_global"].astype(str).eq("Falha").sum())
        lines = [f"{APP_NAME} {APP_VERSION}\n", "Relatório interno — estados por módulo\n\n"]
        lines.append(f"Casos: {len(self.df_results)} | OK: {n_ok} | Avisos: {n_warn} | Falhas: {n_fail}\n\n")
        for _, r in source.head(100).iterrows():
            lines.append(f"Prumada {r.get('prumada', r.get('name',''))} | Membro {r.get('member','')} | Caso {r.get('case','')}\n")
            lines.append(f"  NEd={safe_float(r.get('n_ed_kN',0),0):.2f} kN | MyEd={safe_float(r.get('my_ed_kNm',0),0):.2f} kNm | MzEd={safe_float(r.get('mz_ed_kNm',0),0):.2f} kNm | η_NMyMz={safe_float(r.get('η_NMyMz', r.get('eta_NMyMz',0)),0):.3f}\n")
            lines.append(f"  Estados: global={r.get('estado_global','')} | resistente={r.get('estado_resistente','')} | corte={r.get('estado_corte','')} | torção={r.get('estado_torcao','')} | ELS={r.get('estado_els','')} | pormenorização={r.get('estado_pormenorizacao','')}\n")
            sol = r.get('solucao_completa', r.get('solucao',''))
            lines.append(f"  Solução: {sol}\n")
            dec = _v65_txt(r.get('decisao_tecnica',''))
            if dec:
                lines.append(f"  Decisão técnica: {dec}\n")
            lines.append("\n")
        self.report_txt.insert("1.0", "".join(lines))
    except Exception:
        return _old_update_report_v65(self)

ColumnsEC2App.update_report = _update_report_v65


# --------------------------- Excel v6.5 ---------------------------
_old_write_excel_v65 = ColumnsEC2App._write_excel

def _write_excel_v65(self, path: str):
    self.df_results = _v65_apply_module_statuses(getattr(self, "df_results", pd.DataFrame()))
    self.df_summary = _v65_apply_module_statuses(getattr(self, "df_summary", pd.DataFrame())) if getattr(self, "df_summary", pd.DataFrame()) is not None and not self.df_summary.empty else self.df_summary
    self.df_failures = _v65_failures(self.df_results)
    self.df_warnings = _v65_warnings(self.df_results)
    self.df_module_status = _v65_module_status_table(self.df_results)
    _old_write_excel_v65(self, path)
    try:
        with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            self.df_module_status.to_excel(writer, sheet_name="06B_Estados_Modulo", index=False)
            self.df_failures.to_excel(writer, sheet_name="07_Falhas", index=False)
            self.df_warnings.to_excel(writer, sheet_name="07B_Avisos", index=False)
            # Quadro de decisão por prumada, mais sintético.
            dec_cols = ["prumada", "name", "member", "case", "estado_global", "estado_resistente", "estado_corte", "estado_torcao", "estado_els", "estado_pormenorizacao", "η_NMyMz", "solucao_completa", "decisao_tecnica"]
            summary = _v65_apply_module_statuses(getattr(self, "df_summary", pd.DataFrame()))
            summary[[c for c in dec_cols if c in summary.columns]].to_excel(writer, sheet_name="06C_Decisao_Prumadas", index=False)
    except Exception:
        pass

ColumnsEC2App._write_excel = _write_excel_v65


# --------------------------- PDF v6.5 ---------------------------
_old_write_pdf_v65 = ColumnsEC2App._write_pdf

def _write_pdf_v65(self, path: str):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import mm
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, PageBreak
        self.df_results = _v65_apply_module_statuses(getattr(self, "df_results", pd.DataFrame()))
        self.df_summary = _v65_apply_module_statuses(getattr(self, "df_summary", pd.DataFrame())) if getattr(self, "df_summary", pd.DataFrame()) is not None and not self.df_summary.empty else self.df_summary
        res = self.df_results
        summ = self.df_summary if self.df_summary is not None and not self.df_summary.empty else res
        n_ok = int(res["estado_global"].astype(str).eq("OK").sum()) if not res.empty else 0
        n_warn = int(res["estado_global"].astype(str).eq("Aviso").sum()) if not res.empty else 0
        n_fail = int(res["estado_global"].astype(str).eq("Falha").sum()) if not res.empty else 0
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name="TitleV65", parent=styles["Title"], fontName="Courier-Bold", fontSize=14, leading=18, alignment=1))
        styles.add(ParagraphStyle(name="SubV65", parent=styles["Normal"], fontName="Courier", fontSize=9, leading=12, alignment=1, textColor=colors.darkgrey))
        styles.add(ParagraphStyle(name="H2V65", parent=styles["Heading2"], fontName="Courier-Bold", fontSize=11, leading=15, spaceBefore=8, spaceAfter=8))
        styles.add(ParagraphStyle(name="BodyV65", parent=styles["Normal"], fontName="Courier", fontSize=8.5, leading=12))
        doc = SimpleDocTemplate(path, pagesize=landscape(A4), leftMargin=12*mm, rightMargin=12*mm, topMargin=12*mm, bottomMargin=12*mm)
        story = [Paragraph("Columns EC2", styles["TitleV65"]), Paragraph("Relatório técnico de dimensionamento de pilares — estados por módulo", styles["SubV65"]), Spacer(1, 4*mm)]
        meta = [["Norma", _v59_norm_reference(self) if "_v59_norm_reference" in globals() else "Eurocódigo 2", "Data", datetime.now().strftime("%Y-%m-%d %H:%M")],
                ["Casos", str(len(res)), "OK / Aviso / Falha", f"{n_ok} / {n_warn} / {n_fail}"],
                ["Classe de aço", _v59_getvar(self,"var_fyk","A500") if "_v59_getvar" in globals() else self.var_fyk.get(), "Estratégia", _v64_strategy_label(globals().get("ACTIVE_REBAR_STRATEGY_V64")) if "_v64_strategy_label" in globals() else "-"]]
        t = Table(meta, colWidths=[35*mm, 100*mm, 40*mm, 95*mm]); t.setStyle(_v64_pdf_style_table(False) if "_v64_pdf_style_table" in globals() else self._pdf_table_style(False)); story.append(t); story.append(Spacer(1, 5*mm))
        story.append(Paragraph("1. Síntese executiva", styles["H2V65"]))
        story.append(Paragraph("A decisão global resulta da hierarquia dos estados resistentes, corte, torção, ELS e pormenorização. Avisos construtivos não são classificados como falhas resistentes, mas requerem confirmação em projecto de execução.", styles["BodyV65"]))
        story.append(Spacer(1, 3*mm))
        story.append(Paragraph("2. Decisão por prumada", styles["H2V65"]))
        cols = ["prumada", "member", "case", "material", "b_cm", "h_cm", "n_ed_kN", "my_ed_kNm", "mz_ed_kNm", "η_NMyMz", "estado_global", "estado_resistente", "estado_pormenorizacao", "solucao_completa"]
        story.append(_v64_pdf_df_table(summ, cols, max_rows=34) if "_v64_pdf_df_table" in globals() else self._pdf_df_table(summ, cols, max_rows=34))
        module = _v65_module_status_table(res)
        if not module.empty:
            story.append(PageBreak()); story.append(Paragraph("3. Estados por módulo", styles["H2V65"]))
            story.append(_v64_pdf_df_table(module, ["prumada", "member", "case", "estado_global", "estado_resistente", "estado_corte", "estado_torcao", "estado_els", "estado_pormenorizacao", "decisao_tecnica"], max_rows=45) if "_v64_pdf_df_table" in globals() else self._pdf_df_table(module, ["prumada", "member", "case", "estado_global", "estado_resistente", "estado_corte", "estado_torcao", "estado_els", "estado_pormenorizacao", "decisao_tecnica"], max_rows=45))
        fails = _v65_failures(res); warns = _v65_warnings(res)
        if not fails.empty or not warns.empty:
            story.append(PageBreak()); story.append(Paragraph("4. Falhas bloqueantes e avisos", styles["H2V65"]))
            alert = pd.concat([fails, warns], ignore_index=True) if not warns.empty else fails
            story.append(_v64_pdf_df_table(_v65_apply_module_statuses(alert), ["prumada", "member", "case", "estado_global", "failure_reason", "failure_warnings", "detailing_warnings", "recommendations"], max_rows=50) if "_v64_pdf_df_table" in globals() else self._pdf_df_table(alert, ["prumada", "member", "case", "estado_global", "failure_reason", "failure_warnings", "detailing_warnings", "recommendations"], max_rows=50))
        story.append(Spacer(1, 5*mm)); story.append(Paragraph("Observação: o relatório PDF é sintético. A memória completa e os dados por combinação encontram-se no ficheiro Excel.", styles["BodyV65"]))
        def footer(canvas, doc_obj):
            canvas.saveState(); canvas.setFont("Courier", 7); canvas.setFillColor(colors.grey)
            canvas.drawString(12*mm, 7*mm, f"Columns EC2 | {datetime.now().strftime('%Y-%m-%d %H:%M')}")
            canvas.drawRightString(285*mm, 7*mm, f"Página {doc_obj.page}"); canvas.restoreState()
        doc.build(story, onFirstPage=footer, onLaterPages=footer)
    except Exception:
        return _old_write_pdf_v65(self, path)

ColumnsEC2App._write_pdf = _write_pdf_v65


# --------------------------- filtros v6.5 ---------------------------
_old_build_sidebar_v65 = ColumnsEC2App._build_sidebar

def _build_sidebar_v65(self, parent):
    _old_build_sidebar_v65(self, parent)
    # garantir que os combobox de estado reconhecem os novos estados
    def walk(w):
        try:
            if isinstance(w, ttk.Combobox) and str(w.cget("textvariable")) == str(getattr(self, "var_filter_status", "")):
                w.configure(values=["Todos", "OK", "Aviso", "Falha", "Pré-dimensionado"])
        except Exception:
            pass
        for ch in w.winfo_children():
            walk(ch)
    try:
        walk(parent)
    except Exception:
        pass

ColumnsEC2App._build_sidebar = _build_sidebar_v65


# --------------------------- correcção iterativa v6.5 ---------------------------
_old_repair_v65 = getattr(ColumnsEC2App, "repair_failures_interactive", None)

def _repair_failures_interactive_v65(self):
    if getattr(self, "df_results", pd.DataFrame()) is None or self.df_results.empty:
        messagebox.showwarning("Aviso", "Execute primeiro o cálculo.")
        return
    self.df_results = _v65_apply_module_statuses(self.df_results)
    failures = _v65_failures(self.df_results)
    if failures.empty:
        messagebox.showinfo("Correcção iterativa", "Não foram detectadas falhas bloqueantes. Os avisos devem ser tratados por revisão construtiva ou verificações complementares.")
        return
    if _old_repair_v65 is not None:
        return _old_repair_v65(self)
    messagebox.showinfo("Correcção iterativa", f"Foram detectadas {len(failures)} falhas bloqueantes. A rotina de correcção automática não está disponível neste módulo.")

ColumnsEC2App.repair_failures_interactive = _repair_failures_interactive_v65



# ============================================================
# ColumnsEC2 v6.6 — pormenorização construtiva, diagnóstico SC,
# tentativas de correcção, DXF final e superfície N-My-Mz
# ============================================================
APP_VERSION = "v6.7"


def _v66_get_prumada(row):
    for k in ["prumada", "name", "Name", "nome", "member"]:
        try:
            v = row.get(k, "")
            if str(v).strip() and str(v).strip().lower() not in ["nan", "none"]:
                return str(v).strip()
        except Exception:
            pass
    return "-"


def _v66_layout_refinement(row):
    """Gera uma leitura construtiva simples da armadura longitudinal e dos grampos.
    Não substitui a verificação resistente; é uma camada de pormenorização/desenho.
    """
    b = _finite(row.get("b_cm"), 0.0) * 10.0
    h = _finite(row.get("h_cm"), 0.0) * 10.0
    phi = _finite(row.get("phi_long_mm"), 0.0)
    phi_st = _finite(row.get("phi_st_mm"), 8.0)
    s_st = _finite(row.get("s_st_mm"), 0.0)
    ny = int(_finite(row.get("n_bars_y"), 0))
    nz = int(_finite(row.get("n_bars_z"), 0))
    n_total = int(_finite(row.get("n_total"), 0))
    cover = _finite(row.get("cover_mm"), _finite(globals().get("DEFAULT_COVER_MM", 35.0), 35.0))

    warnings = []
    blocking = []
    info = []

    if n_total <= 0 or phi <= 0:
        return {
            "pormenorizacao_construtiva": "sem armadura definida",
            "grampos_intermedios": "não aplicável",
            "numero_grampos_por_nivel": 0,
            "detalhe_grampos": "-",
            "detailing_blocking_issues": "armadura longitudinal não definida",
            "detailing_warnings": "",
            "estado_pormenorizacao": "Falha",
        }

    # descrição construtiva de base
    if n_total == 4:
        desc = f"4Ø{int(phi)} nos cantos"
    else:
        extra = max(0, n_total - 4)
        desc = f"4Ø{int(phi)} nos cantos + {extra}Ø{int(phi)} distribuídos nas faces"

    # espaçamentos livres aproximados por direcção
    edge = cover + phi_st + phi / 2.0
    clear_y = None
    clear_z = None
    if b > 0 and ny > 1:
        clear_y = (b - 2.0 * edge) / max(ny - 1, 1) - phi
    if h > 0 and nz > 1:
        clear_z = (h - 2.0 * edge) / max(nz - 1, 1) - phi

    min_clear_req = max(20.0, phi, 25.0)  # inclui margem expedita para agregado corrente
    if clear_y is not None and clear_y < min_clear_req:
        blocking.append(f"espaçamento livre insuficiente na largura ({clear_y:.0f} mm < {min_clear_req:.0f} mm)")
    if clear_z is not None and clear_z < min_clear_req:
        blocking.append(f"espaçamento livre insuficiente na altura ({clear_z:.0f} mm < {min_clear_req:.0f} mm)")

    # regra prática de travamento: varões intermédios em faces com mais de 2 varões requerem grampos/cintas suplementares.
    links_y = max(0, ny - 2)
    links_z = max(0, nz - 2)
    n_links = links_y + links_z
    if n_links > 0:
        warnings.append("prever grampos/cintas suplementares para travar varões intermédios comprimidos")
        detalhe_grampos = f"{n_links} grampo(s) Ø{int(phi_st)} por nível: {links_y} na largura + {links_z} na altura"
        grampos = f"{n_links}Ø{int(phi_st)} por nível"
    else:
        detalhe_grampos = "não necessários para layout com apenas varões de canto"
        grampos = "sem grampos intermédios"

    # compatibilidade de estribos
    smax = min(15.0 * phi if phi else 300.0, min(b, h) if b and h else 300.0, 300.0)
    if s_st > 0 and s_st > smax + 1e-6:
        blocking.append(f"espaçamento de estribos superior ao limite construtivo ({s_st:.0f} mm > {smax:.0f} mm)")
    if phi_st < max(6.0, phi / 4.0):
        blocking.append("diâmetro de estribo inferior a max(6 mm; Ølong/4)")

    if _finite(row.get("as_prov_mm2"), 0.0) > _finite(row.get("as_max_mm2"), 1e99) + 1e-6:
        blocking.append("As,prov superior a As,max")
    if _finite(row.get("as_prov_mm2"), 0.0) < _finite(row.get("as_min_mm2"), 0.0) - 1e-6:
        blocking.append("As,prov inferior a As,min")

    if _finite(row.get("as_req_mm2"), 0.0) > 0 and _finite(row.get("as_prov_mm2"), 0.0) / max(_finite(row.get("as_req_mm2"), 1.0), 1e-9) > 1.60:
        info.append("As,prov bastante superior a As,req; avaliar estratégia de armadura")

    estado = "Falha" if blocking else ("Aviso" if warnings else "OK")
    if s_st > 0:
        desc_full = f"{desc}; estribos Ø{int(phi_st)}//{int(round(s_st))} mm; {grampos}"
    else:
        desc_full = f"{desc}; estribos Ø{int(phi_st)}; {grampos}"
    return {
        "pormenorizacao_construtiva": desc_full,
        "grampos_intermedios": grampos,
        "numero_grampos_por_nivel": n_links,
        "detalhe_grampos": detalhe_grampos,
        "detailing_blocking_issues": "; ".join(blocking),
        "detailing_warnings": "; ".join(warnings + info),
        "estado_pormenorizacao": estado,
        "detailing_min_clear_y_mm": clear_y,
        "detailing_min_clear_z_mm": clear_z,
        "detailing_smax_ties_v66_mm": smax,
    }


def _v66_apply_constructive_detailing(df):
    if df is None or df.empty:
        return df
    out = df.copy()
    for idx, row in out.iterrows():
        d = _v66_layout_refinement(row)
        for k, v in d.items():
            out.at[idx, k] = v
        # actualizar solução completa sem apagar a solução de cálculo
        if str(out.at[idx, "pormenorizacao_construtiva"]).strip():
            out.at[idx, "solucao_completa"] = out.at[idx, "pormenorizacao_construtiva"]
    # reaplicar estados globais se a função v6.5 existir
    try:
        out = _v65_apply_module_statuses(out)
    except Exception:
        pass
    return out


def _v66_surface_points(df, max_rows=200):
    """Exporta uma secção representativa da superfície N-My-Mz para o nível NEd.
    Quando só existem MRd,y/MRd,z no resultado, gera pontos por interpolação polar da envolvente seccional.
    """
    rows = []
    if df is None or df.empty:
        return pd.DataFrame()
    src = df.copy()
    if "sort_key" in src.columns:
        src = src.sort_values("sort_key")
    for _, r in src.head(max_rows).iterrows():
        mry = _finite(r.get("mrd_y_kNm"), 0.0)
        mrz = _finite(r.get("mrd_z_kNm"), 0.0)
        if mry <= 0 or mrz <= 0:
            continue
        alpha = _finite(r.get("biaxial_alpha"), 2.0)
        alpha = alpha if alpha > 0 else 2.0
        for deg in range(0, 361, 15):
            th = math.radians(deg)
            c = abs(math.cos(th)); s = abs(math.sin(th))
            denom = ((c / max(mry,1e-9)) ** alpha + (s / max(mrz,1e-9)) ** alpha) ** (1.0/alpha)
            R = 0.0 if denom <= 0 else 1.0/denom
            rows.append({
                "prumada": _v66_get_prumada(r),
                "member": r.get("member", ""),
                "case": r.get("case", ""),
                "N_Ed_kN": _finite(r.get("n_ed_kN"), 0.0),
                "angulo_graus": deg,
                "MRd_y_kNm": R * math.cos(th),
                "MRd_z_kNm": R * math.sin(th),
                "MRd_y_base_kNm": mry,
                "MRd_z_base_kNm": mrz,
                "alpha": alpha,
                "metodo": "corte polar representativo da superfície N-My-Mz ao nível NEd",
            })
    return pd.DataFrame(rows)


def _v66_structuralcodes_diagnostics_df():
    rows=[]
    def add(mod_name, funcs):
        try:
            import importlib
            m = importlib.import_module(mod_name)
            ok = "Disponível"
            for f in funcs:
                rows.append({"Módulo": mod_name, "Item": f, "Estado": "Sim" if hasattr(m, f) else "Não", "Nota": ok})
        except Exception as err:
            rows.append({"Módulo": mod_name, "Item": "import", "Estado": "Não", "Nota": str(err)})
    try:
        import importlib.metadata as md
        rows.append({"Módulo":"structuralcodes", "Item":"versão", "Estado": md.version("structuralcodes"), "Nota":"pacote instalado"})
    except Exception as err:
        rows.append({"Módulo":"structuralcodes", "Item":"versão", "Estado":"Indisponível", "Nota":str(err)})
    for pkg in ["numpy", "shapely"]:
        try:
            import importlib.metadata as md
            rows.append({"Módulo":pkg, "Item":"versão", "Estado":md.version(pkg), "Nota":"dependência"})
        except Exception as err:
            rows.append({"Módulo":pkg, "Item":"versão", "Estado":"Indisponível", "Nota":str(err)})
    add("structuralcodes.codes.ec2_2004", ["fcd", "fctm", "Ecm", "h_0", "phi", "phi_0", "beta_c"])
    add("structuralcodes.codes.ec2_2023", ["fcd", "fctm", "Ecm", "hn", "phi_50y_t0", "phi_correction_factor"])
    add("structuralcodes.codes.mc2010", ["fcd", "fctm", "Eci", "phi", "phi_bc", "phi_dc"])
    add("structuralcodes.sections", ["GenericSection", "BeamSection", "SectionCalculator", "calculate_nm_interaction_domain"])
    return pd.DataFrame(rows)


def _v66_show_df_popup(title, df, parent=None, width=980, height=520):
    win = tk.Toplevel(parent) if parent is not None else tk.Toplevel()
    win.title(title)
    win.geometry(f"{width}x{height}")
    win.minsize(720, 360)
    frame = ttk.Frame(win, padding=8)
    frame.pack(fill="both", expand=True)
    tree = ttk.Treeview(frame, show="headings")
    vsb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
    hsb = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
    tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
    tree.grid(row=0, column=0, sticky="nsew")
    vsb.grid(row=0, column=1, sticky="ns")
    hsb.grid(row=1, column=0, sticky="ew")
    frame.rowconfigure(0, weight=1); frame.columnconfigure(0, weight=1)
    df = df if df is not None else pd.DataFrame()
    cols = list(df.columns)
    tree["columns"] = cols
    for c in cols:
        tree.heading(c, text=str(c)); tree.column(c, width=max(120, min(280, len(str(c))*9)), anchor="w")
    for _, r in df.head(2000).iterrows():
        vals=[]
        for c in cols:
            v = r.get(c, "")
            if isinstance(v, float): vals.append("" if not math.isfinite(v) else f"{v:.3f}")
            else: vals.append("" if pd.isna(v) else str(v))
        tree.insert("", "end", values=vals)
    ttk.Button(win, text="Fechar", command=win.destroy).pack(pady=(0,8))
    return win


def _v66_correction_attempt_rows(results):
    rows=[]
    if results is None or results.empty:
        return pd.DataFrame()
    try:
        res = _v65_apply_module_statuses(results.copy())
    except Exception:
        res = results.copy()
    fail_mask = res.get("estado_global", res.get("status", "")).astype(str).eq("Falha") if "estado_global" in res.columns or "status" in res.columns else pd.Series(False, index=res.index)
    for _, r in res[fail_mask].iterrows():
        pr = _v66_get_prumada(r)
        reason = str(r.get("failure_reason", "") or r.get("detailing_blocking_issues", "") or "falha bloqueante")
        candidates = [
            ("Aumentar armadura mantendo secção", "testar estratégia robusta e diâmetros superiores", "não aplicado automaticamente"),
            ("Redistribuir armadura", "privilegiar varões nos cantos e faces mais solicitadas", "não aplicado automaticamente"),
            ("Adicionar grampos/cintas suplementares", "resolver travamento de varões intermédios", "proposta construtiva"),
            ("Aumentar secção", "incrementos de 5 cm na menor/maior dimensão", "exige revisão do modelo")
        ]
        for i,(alt,just,status) in enumerate(candidates, start=1):
            rows.append({"prumada": pr, "member": r.get("member",""), "case": r.get("case",""), "falha": reason, "tentativa": i, "alteracao_proposta": alt, "justificacao": just, "resultado": status})
    return pd.DataFrame(rows)


# --------------------------- XLSX v6.6 ---------------------------
_old_write_excel_v66 = ColumnsEC2App._write_excel

def _write_excel_v66(self, path: str):
    self.df_results = _v66_apply_constructive_detailing(getattr(self, "df_results", pd.DataFrame()))
    if getattr(self, "df_summary", pd.DataFrame()) is not None and not self.df_summary.empty:
        self.df_summary = _v66_apply_constructive_detailing(self.df_summary)
    self.df_correction_attempts = getattr(self, "df_repair_attempts", pd.DataFrame())
    if self.df_correction_attempts is None or self.df_correction_attempts.empty:
        self.df_correction_attempts = _v66_correction_attempt_rows(self.df_results)
    self.df_surface_nmm = _v66_surface_points(self.df_summary if getattr(self,"df_summary", pd.DataFrame()) is not None and not self.df_summary.empty else self.df_results)
    self.df_sc_diagnostics = _v66_structuralcodes_diagnostics_df()
    _old_write_excel_v66(self, path)
    try:
        with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            self.df_correction_attempts.to_excel(writer, sheet_name="26_Tentativas_Correcao", index=False)
            self.df_surface_nmm.to_excel(writer, sheet_name="20_Superficie_NMyMz", index=False)
            self.df_sc_diagnostics.to_excel(writer, sheet_name="30_Diagnostico_SC", index=False)
            # sobrescrever folhas principais com pormenorização v6.6 aplicada
            self.df_results.to_excel(writer, sheet_name="05_Resultados", index=False)
            if getattr(self, "df_summary", pd.DataFrame()) is not None and not self.df_summary.empty:
                self.df_summary.to_excel(writer, sheet_name="06_Resumo_Membros", index=False)
    except Exception as err:
        print("Aviso v6.6: não foi possível acrescentar folhas adicionais ao Excel:", err)

ColumnsEC2App._write_excel = _write_excel_v66


# --------------------------- Balões/popup GUI v6.6 ---------------------------
_old_build_sidebar_v66 = ColumnsEC2App._build_sidebar

def _show_sc_diagnostics_balloon_v66(self):
    df = _v66_structuralcodes_diagnostics_df()
    _v66_show_df_popup("Diagnóstico structuralcodes", df, parent=self)


def _show_correction_attempts_balloon_v66(self):
    df = getattr(self, "df_repair_attempts", pd.DataFrame())
    if df is None or df.empty:
        df = _v66_correction_attempt_rows(getattr(self, "df_results", pd.DataFrame()))
    if df is None or df.empty:
        messagebox.showinfo("Tentativas de correcção", "Não existem falhas bloqueantes ou tentativas registadas.")
        return
    _v66_show_df_popup("Tentativas de correcção interactiva", df, parent=self)


def _build_sidebar_v66(self, parent):
    _old_build_sidebar_v66(self, parent)
    box = ttk.LabelFrame(parent, text="Diagnóstico e auditoria")
    box.pack(fill="x", pady=(0, 8))
    ttk.Button(box, text="Diagnóstico structuralcodes", command=lambda: _show_sc_diagnostics_balloon_v66(self)).pack(fill="x", padx=4, pady=4)
    ttk.Button(box, text="Tentativas de correcção", command=lambda: _show_correction_attempts_balloon_v66(self)).pack(fill="x", padx=4, pady=4)
    ttk.Label(box, text="As tentativas detalhadas são exportadas apenas no ficheiro .xlsx.", style="Subtle.TLabel", wraplength=330, justify="left").pack(fill="x", padx=6, pady=(0,6))

ColumnsEC2App._build_sidebar = _build_sidebar_v66


# --------------------------- reparar correcção interactiva v6.6 ---------------------------
_old_repair_v66 = getattr(ColumnsEC2App, "repair_failures_interactive", None)

def _repair_failures_interactive_v66(self):
    if getattr(self, "df_results", pd.DataFrame()) is None or self.df_results.empty:
        messagebox.showwarning("Aviso", "Execute primeiro o cálculo.")
        return
    attempts = _v66_correction_attempt_rows(self.df_results)
    self.df_repair_attempts = attempts
    if attempts.empty:
        messagebox.showinfo("Correcção interactiva", "Não foram detectadas falhas bloqueantes. Os avisos construtivos devem ser revistos no desenho/pormenorização.")
        return
    # executar rotina anterior se existir, mas mantendo log de tentativas
    if _old_repair_v66 is not None:
        try:
            res = _old_repair_v66(self)
            # se a rotina antiga criar propostas, juntar ao log
            props = getattr(self, "df_repair_proposals", pd.DataFrame())
            if props is not None and not props.empty:
                extra=[]
                for _, r in props.iterrows():
                    extra.append({"prumada": _v66_get_prumada(r), "member": r.get("member", ""), "case": r.get("case", ""), "falha": r.get("failure_reason", ""), "tentativa": r.get("tentativa", ""), "alteracao_proposta": r.get("proposta", r.get("solucao", "")), "justificacao": r.get("nota", "proposta gerada"), "resultado": r.get("status", "")})
                self.df_repair_attempts = pd.concat([attempts, pd.DataFrame(extra)], ignore_index=True)
            return res
        except Exception as err:
            messagebox.showwarning("Correcção interactiva", f"Foi gerado o relatório de tentativas, mas a rotina automática não concluiu.\n\n{err}")
            return
    messagebox.showinfo("Correcção interactiva", "Relatório de tentativas gerado. Exporte o .xlsx para consultar a folha 26_Tentativas_Correcao.")

ColumnsEC2App.repair_failures_interactive = _repair_failures_interactive_v66


# --------------------------- DXF v6.6: cotas, legenda, grampos ---------------------------
def _dxf_layer_table_v66():
    layers = ["COLUMNS_CONCRETE", "COLUMNS_REBAR", "COLUMNS_STIRRUPS", "COLUMNS_LINKS", "COLUMNS_TEXT", "COLUMNS_TABLE", "COLUMNS_DIMENSIONS"]
    out = "0\nSECTION\n2\nTABLES\n0\nTABLE\n2\nLAYER\n70\n{}\n".format(len(layers))
    for i, name in enumerate(layers, start=1):
        out += "0\nLAYER\n2\n{}\n70\n0\n62\n{}\n6\nCONTINUOUS\n".format(name, i)
    out += "0\nENDTAB\n0\nENDSEC\n"
    return out


def _dxf_dim_text_v66(parts, x1, y1, x2, y2, text, off=55):
    parts.append(_dxf_line(x1, y1, x2, y2, "COLUMNS_DIMENSIONS"))
    parts.append(_dxf_text((x1+x2)/2.0, (y1+y2)/2.0 + off, text, 22, "COLUMNS_DIMENSIONS"))


def write_columns_dxf_v66(path: str, df: pd.DataFrame):
    df = _v66_apply_constructive_detailing(df)
    parts = ["0\nSECTION\n2\nHEADER\n9\n$INSUNITS\n70\n4\n0\nENDSEC\n", _dxf_layer_table_v66(), "0\nSECTION\n2\nENTITIES\n"]
    if df is None or df.empty:
        parts.append(_dxf_text(0, 0, "Sem resultados", 50, "COLUMNS_TEXT"))
    else:
        work = df.copy()
        work["_prumada"] = work.apply(_v66_get_prumada, axis=1)
        # tentar ordenar por piso/tramo; caso não exista, ordem por member/case
        def tramo_key(r):
            for k in ["story", "piso", "floor", "level", "tramo", "length_m", "member"]:
                if k in r.index:
                    val = r.get(k)
                    if str(val).strip() and str(val).lower() != "nan":
                        try: return float(re.sub(r"[^0-9.-]", "", str(val)) or 0)
                        except Exception: return str(val)
            return str(r.get("member", ""))
        prumadas = sorted(work["_prumada"].unique(), key=lambda s: [int(t) if t.isdigit() else t for t in re.split(r"(\d+)", str(s))])
        cell_w, cell_h = 1350.0, 1050.0
        margin_x, margin_y = 400.0, 800.0
        # título e legenda
        parts.append(_dxf_text(0, 300, "QUADRO DE PILARES - UNIDADES: mm", 50, "COLUMNS_TEXT"))
        legend_y = 140.0
        parts.append(_dxf_text(0, legend_y, "Legenda: betão=contorno | varões=círculos | estribos=rectângulo interior | grampos=linhas interiores", 25, "COLUMNS_TEXT"))
        for c, pr in enumerate(prumadas):
            x0 = margin_x + c * cell_w
            parts.append(_dxf_text(x0, 0, str(pr), 42, "COLUMNS_TABLE"))
            grp = work[work["_prumada"] == pr].copy()
            grp["_tramo_key"] = grp.apply(tramo_key, axis=1)
            # baixo para cima: chave crescente de baixo para cima, desenhada a subir
            for r_i, (_, r) in enumerate(grp.sort_values("_tramo_key").iterrows()):
                y0 = -margin_y - r_i * cell_h
                b = _finite(r.get("b_cm"))*10.0
                h = _finite(r.get("h_cm"))*10.0
                if b <= 0 or h <= 0:
                    parts.append(_dxf_text(x0, y0, "-", 28, "COLUMNS_TEXT")); continue
                ox = x0 + cell_w/2.0
                oy = y0 - cell_h/2.0 + 130.0
                scale = min(1.0, 620.0/max(b,h,1.0))
                bs, hs = b*scale, h*scale
                left, right = ox-bs/2, ox+bs/2; bot, top = oy-hs/2, oy+hs/2
                parts += [_dxf_line(left,bot,right,bot,"COLUMNS_CONCRETE"), _dxf_line(right,bot,right,top,"COLUMNS_CONCRETE"), _dxf_line(right,top,left,top,"COLUMNS_CONCRETE"), _dxf_line(left,top,left,bot,"COLUMNS_CONCRETE")]
                cov = _finite(r.get("cover_mm"),35.0)*scale
                l2,r2,b2,t2 = left+cov, right-cov, bot+cov, top-cov
                parts += [_dxf_line(l2,b2,r2,b2,"COLUMNS_STIRRUPS"), _dxf_line(r2,b2,r2,t2,"COLUMNS_STIRRUPS"), _dxf_line(r2,t2,l2,t2,"COLUMNS_STIRRUPS"), _dxf_line(l2,t2,l2,b2,"COLUMNS_STIRRUPS")]
                # varões
                phi = _finite(r.get("phi_long_mm"), 10.0)*scale
                for yy, zz in _bar_points_for_result(r):
                    parts.append(_dxf_circle(ox+yy*scale, oy+zz*scale, max(phi/2.0, 3.0), "COLUMNS_REBAR"))
                # grampos intermédios simplificados
                nlinks = int(_finite(r.get("numero_grampos_por_nivel"), 0))
                if nlinks > 0:
                    if _finite(r.get("n_bars_y"),0) > 2:
                        parts.append(_dxf_line(l2, oy, r2, oy, "COLUMNS_LINKS"))
                    if _finite(r.get("n_bars_z"),0) > 2:
                        parts.append(_dxf_line(ox, b2, ox, t2, "COLUMNS_LINKS"))
                _dxf_dim_text_v66(parts, left, bot-45, right, bot-45, f"{b:.0f}", off=-35)
                _dxf_dim_text_v66(parts, right+45, bot, right+45, top, f"{h:.0f}", off=0)
                parts.append(_dxf_text(x0+20, y0-60, f"Tramo {r_i+1}: {r.get('member','')} | {r.get('material','')}", 24, "COLUMNS_TEXT"))
                parts.append(_dxf_text(x0+20, y0-100, str(r.get("solucao_completa", r.get("solucao", "")))[:90], 22, "COLUMNS_TEXT"))
                parts.append(_dxf_text(x0+20, y0-140, f"Estado: {r.get('estado_global', r.get('status',''))} | η={_finite(r.get('η_NMyMz', r.get('eta_NMyMz')), 0.0):.2f}", 22, "COLUMNS_TEXT"))
    parts.append("0\nENDSEC\n0\nEOF\n")
    Path(path).write_text("".join(parts), encoding="utf-8")

# Reencaminhar exportação DXF, mantendo compatibilidade com nomes anteriores.
write_columns_dxf_v66_ref = write_columns_dxf_v66
write_columns_dxf_v61 = write_columns_dxf_v66
write_columns_dxf_v64 = write_columns_dxf_v66
write_columns_dxf_v4 = write_columns_dxf_v66


def _export_dxf_v66(self):
    src = self.df_summary if getattr(self, "df_summary", pd.DataFrame()) is not None and not self.df_summary.empty else getattr(self, "df_results", pd.DataFrame())
    if src is None or src.empty:
        messagebox.showwarning("Aviso", "Não há resultados para exportar em DXF.")
        return
    path = filedialog.asksaveasfilename(title="Exportar quadro de pilares DXF", defaultextension=".dxf", filetypes=[("DXF", "*.dxf")])
    if not path:
        return
    try:
        self.status_var.set("A exportar DXF..."); self.update_idletasks()
        write_columns_dxf_v66(path, src)
        self.status_var.set(f"DXF exportado para: {path}")
    except Exception as err:
        messagebox.showerror("Erro", f"Não foi possível exportar DXF.\n\n{err}")

ColumnsEC2App.export_dxf = _export_dxf_v66


# --------------------------- pós-processamento do cálculo v6.6 ---------------------------
_old_run_design_v66 = ColumnsEC2App.run_design

def _run_design_v66(self):
    # usa rotina existente; em finish posterior, aplicar pormenorização construtiva por polling curto
    _old_run_design_v66(self)
    def _post():
        try:
            if getattr(self, "df_results", pd.DataFrame()) is not None and not self.df_results.empty:
                self.df_results = _v66_apply_constructive_detailing(self.df_results)
                if getattr(self, "df_summary", pd.DataFrame()) is not None and not self.df_summary.empty:
                    self.df_summary = _v66_apply_constructive_detailing(self.df_summary)
                self.df_failures = _v65_failures(self.df_results) if "_v65_failures" in globals() else self.df_results[self.df_results.get("estado_global", self.df_results.get("status", "")).astype(str).eq("Falha")]
                self.df_warnings = _v65_warnings(self.df_results) if "_v65_warnings" in globals() else self.df_results[self.df_results.get("estado_global", self.df_results.get("status", "")).astype(str).eq("Aviso")]
                self.df_repair_attempts = _v66_correction_attempt_rows(self.df_results)
                try:
                    self.show_df(self.tree_results, self.df_results)
                    self.show_df(self.tree_summary, self.df_summary)
                    self.show_df(self.tree_failures, self.df_failures)
                    self.update_report()
                except Exception:
                    pass
            else:
                self.after(700, _post)
        except Exception:
            pass
    self.after(900, _post)

ColumnsEC2App.run_design = _run_design_v66



# ============================================================
# ColumnsEC2 v6.7 — estabilização de performance/exportação
# ============================================================
APP_VERSION = "v6.7"


def _v67_to_safe_df(df):
    """Converte DataFrame para formato exportável, removendo objectos problemáticos."""
    if df is None:
        return pd.DataFrame()
    try:
        out = df.copy()
    except Exception:
        return pd.DataFrame()
    for c in list(out.columns):
        try:
            if out[c].dtype == "object":
                out[c] = out[c].map(lambda v: "" if pd.isna(v) else str(v))
        except Exception:
            out[c] = out[c].astype(str)
    return out


def _v67_sheet_name(name, used):
    base = re.sub(r"[\\/*?:\[\]]", "_", str(name))[:31] or "Sheet"
    candidate = base
    i = 1
    while candidate in used:
        suffix = f"_{i}"
        candidate = (base[:31-len(suffix)] + suffix)
        i += 1
    used.add(candidate)
    return candidate


def _v67_material_classes_from_df(app):
    vals = []
    for df in [getattr(app, "df_clean", pd.DataFrame()), getattr(app, "df_results", pd.DataFrame())]:
        if df is not None and not df.empty and "material" in df.columns:
            vals += [str(x).strip() for x in df["material"].dropna().tolist()]
    mats = []
    for v in vals:
        if re.search(r"C\s*\d+\s*/\s*\d+", v, flags=re.I):
            vv = re.search(r"C\s*\d+\s*/\s*\d+", v, flags=re.I).group(0).replace(" ", "")
            if vv not in mats:
                mats.append(vv)
    return ", ".join(mats) if mats else "não identificado na tabela"


def _v67_info_df(app):
    norm = _v59_norm_reference(app) if "_v59_norm_reference" in globals() else "Eurocódigo 2"
    desc = "Dimensionamento e verificação de pilares de betão armado, com organização por prumada, verificação resistente N-My-Mz, pormenorização construtiva e exportação técnica."
    lim = "Resultados dependentes da qualidade dos esforços importados, da definição dos comprimentos efectivos e das hipóteses de modelação; pilares críticos devem ser revistos por engenheiro responsável."
    return pd.DataFrame([
        ["Programa", APP_NAME],
        ["Norma de referência", norm.replace(" | structuralcodes", "")],
        ["Autor / Repositório", GITHUB_URL],
        ["Data de exportação", datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["Descrição", desc],
        ["Limitações", lim],
    ], columns=["Campo", "Valor"])


def _v67_parameters_df(app):
    steel = _v59_getvar(app, "var_fyk", "500") if "_v59_getvar" in globals() else getattr(app, "var_fyk", tk.StringVar(value="500")).get()
    try:
        strategy = _v64_strategy_label(globals().get("ACTIVE_REBAR_STRATEGY_V64", "equilibrada"))
    except Exception:
        strategy = "Equilibrada"
    rows = [
        ["Classe de Aço", f"A{steel}" if str(steel).isdigit() else str(steel)],
        ["Betão", _v67_material_classes_from_df(app)],
        ["Recobrimento nominal [mm]", _v59_getvar(app, "var_cover", "") if "_v59_getvar" in globals() else getattr(app, "var_cover", tk.StringVar(value="")).get()],
        ["Modo de cálculo", _v59_getvar(app, "var_calc_mode", "") if "_v59_getvar" in globals() else getattr(app, "var_calc_mode", tk.StringVar(value="")).get()],
        ["Estratégia de armadura", strategy],
        ["Redução para casos governantes", "Sim" if getattr(app, "var_reduce_cases", tk.BooleanVar(value=True)).get() else "Não"],
        ["l0y/L", _v59_getvar(app, "var_l0y", "") if "_v59_getvar" in globals() else getattr(app, "var_l0y", tk.StringVar(value="")).get()],
        ["l0z/L", _v59_getvar(app, "var_l0z", "") if "_v59_getvar" in globals() else getattr(app, "var_l0z", tk.StringVar(value="")).get()],
        ["RH [%]", _v59_getvar(app, "var_rh", "70") if "_v59_getvar" in globals() else "70"],
        ["t0 [dias]", _v59_getvar(app, "var_t0", "28") if "_v59_getvar" in globals() else "28"],
        ["h0/hn", "calculado automaticamente pela geometria da secção"],
        ["φef", "calculado automaticamente quando aplicável ao motor seleccionado"],
    ]
    return pd.DataFrame(rows, columns=["Parâmetro", "Valor"])


def _v67_governing_cases_df(app):
    if "_v64_governing_cases" in globals():
        try:
            return _v64_governing_cases(app)
        except Exception:
            pass
    return _v67_to_safe_df(getattr(app, "df_calc_input", pd.DataFrame()))


def _v67_surface_points_fast(app):
    src = getattr(app, "df_summary", pd.DataFrame())
    if src is None or src.empty:
        src = getattr(app, "df_results", pd.DataFrame())
    if src is None or src.empty:
        return pd.DataFrame()
    # mais leve do que v6.6: exporta só por prumada/member resumido, passo angular 30º.
    rows = []
    data = src.copy().head(80)
    for _, r in data.iterrows():
        mry = _finite(r.get("mrd_y_kNm"), 0.0)
        mrz = _finite(r.get("mrd_z_kNm"), 0.0)
        if mry <= 0 or mrz <= 0:
            continue
        alpha = _finite(r.get("biaxial_alpha"), 2.0) or 2.0
        for deg in range(0, 361, 30):
            th = math.radians(deg)
            c = abs(math.cos(th)); s = abs(math.sin(th))
            denom = ((c/max(mry,1e-9))**alpha + (s/max(mrz,1e-9))**alpha) ** (1/alpha)
            R = 0 if denom <= 0 else 1/denom
            rows.append({
                "prumada": _v66_get_prumada(r) if "_v66_get_prumada" in globals() else r.get("prumada", r.get("name", r.get("member", ""))),
                "member": r.get("member", ""),
                "case": r.get("case", ""),
                "N_Ed_kN": _finite(r.get("n_ed_kN"), 0.0),
                "angulo_graus": deg,
                "MRd_y_kNm": R*math.cos(th),
                "MRd_z_kNm": R*math.sin(th),
                "MRd_y_base_kNm": mry,
                "MRd_z_base_kNm": mrz,
                "alpha": alpha,
                "metodo": "corte polar representativo; exportação optimizada v6.7",
            })
    return pd.DataFrame(rows)


def _v67_write_excel_core(app, path):
    res = _v67_to_safe_df(getattr(app, "df_results", pd.DataFrame()))
    summ = _v67_to_safe_df(getattr(app, "df_summary", pd.DataFrame()))
    if res is not None and not res.empty and "_v65_apply_module_statuses" in globals():
        try: res = _v67_to_safe_df(_v65_apply_module_statuses(res))
        except Exception: pass
    if summ is not None and not summ.empty and "_v65_apply_module_statuses" in globals():
        try: summ = _v67_to_safe_df(_v65_apply_module_statuses(summ))
        except Exception: pass
    if res is not None and not res.empty and "_v66_apply_constructive_detailing" in globals():
        try: res = _v67_to_safe_df(_v66_apply_constructive_detailing(res))
        except Exception: pass
    if summ is not None and not summ.empty and "_v66_apply_constructive_detailing" in globals():
        try: summ = _v67_to_safe_df(_v66_apply_constructive_detailing(summ))
        except Exception: pass
    failures = _v67_to_safe_df(_v65_failures(res) if "_v65_failures" in globals() and res is not None and not res.empty else pd.DataFrame())
    warnings = _v67_to_safe_df(_v65_warnings(res) if "_v65_warnings" in globals() and res is not None and not res.empty else pd.DataFrame())
    module = _v67_to_safe_df(_v65_module_status_table(res) if "_v65_module_status_table" in globals() and res is not None and not res.empty else pd.DataFrame())
    attempts = getattr(app, "df_repair_attempts", pd.DataFrame())
    if attempts is None or attempts.empty:
        try: attempts = _v66_correction_attempt_rows(res) if "_v66_correction_attempt_rows" in globals() else pd.DataFrame()
        except Exception: attempts = pd.DataFrame()
    sc_diag = _v66_structuralcodes_diagnostics_df() if "_v66_structuralcodes_diagnostics_df" in globals() else pd.DataFrame()
    sheets = [
        ("00_Info", _v67_info_df(app)),
        ("01_Parametros", _v67_parameters_df(app)),
        ("02_Entrada_Dados", _v67_to_safe_df(getattr(app, "df_clean", pd.DataFrame()))),
        ("03_Pares_Member_Case", _v67_to_safe_df(getattr(app, "df_pair", pd.DataFrame()))),
        ("04_Casos_Governantes", _v67_to_safe_df(_v67_governing_cases_df(app))),
        ("05_Resultados", res),
        ("06_Resumo_Membros", summ),
        ("06B_Estados_Modulo", module),
        ("06C_Decisao_Prumadas", _v67_to_safe_df(summ if summ is not None and not summ.empty else res)),
        ("07_Falhas", failures),
        ("07B_Avisos", warnings),
        ("09_Shortlists", _v67_to_safe_df(app.build_shortlists_df() if hasattr(app, "build_shortlists_df") else pd.DataFrame())),
        ("10_ELS", _v67_to_safe_df(res[[c for c in res.columns if str(c).startswith("service_") or c in ["member","case","prumada"]]] if res is not None and not res.empty else pd.DataFrame())),
        ("11_V_Torcao", _v67_to_safe_df(res[[c for c in res.columns if "shear" in str(c).lower() or "tors" in str(c).lower() or c in ["member","case","prumada","v_ed_y_kN","v_ed_z_kN","mx_ed_kNm"]]] if res is not None and not res.empty else pd.DataFrame())),
        ("13_Notas", _v67_to_safe_df(app.build_normative_notes() if hasattr(app, "build_normative_notes") else pd.DataFrame())),
        ("20_Superficie_NMyMz", _v67_surface_points_fast(app)),
        ("26_Tentativas_Correcao", _v67_to_safe_df(attempts)),
        ("30_Diagnostico_SC", _v67_to_safe_df(sc_diag)),
    ]
    used=set()
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in sheets:
            df = _v67_to_safe_df(df)
            df.to_excel(writer, sheet_name=_v67_sheet_name(name, used), index=False)
        wb = writer.book
        props = wb.properties
        props.title = APP_NAME
        props.subject = _v59_norm_reference(app) if "_v59_norm_reference" in globals() else APP_SUBJECT
        props.creator = APP_AUTHOR
        props.lastModifiedBy = APP_AUTHOR
        props.description = "Columns EC2 — relatório técnico de dimensionamento de pilares."
        try:
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            from openpyxl.utils import get_column_letter
            fill = PatternFill("solid", fgColor="1F4E5F")
            font = Font(color="FFFFFF", bold=True)
            thin = Side(style="thin", color="D9E2E7")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for ws in wb.worksheets:
                ws.sheet_view.showGridLines = False
                ws.freeze_panes = "A2"
                if ws.max_row >= 1:
                    for cell in ws[1]:
                        cell.fill = fill; cell.font = font
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        cell.border = border
                for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 3000)):
                    for cell in row:
                        cell.border = border
                        cell.alignment = Alignment(vertical="top", wrap_text=True)
                for i, col in enumerate(ws.columns, start=1):
                    vals=[str(c.value) for c in col[:200] if c.value is not None]
                    ws.column_dimensions[get_column_letter(i)].width = min(max([len(v) for v in vals]+[10])+2, 48)
        except Exception:
            pass


def _write_excel_v67(self, path: str):
    return _v67_write_excel_core(self, path)

ColumnsEC2App._write_excel = _write_excel_v67


def _export_excel_v67(self):
    if getattr(self, "df_results", pd.DataFrame()) is None or self.df_results.empty:
        messagebox.showwarning("Aviso", "Não há resultados para exportar.")
        return
    path = filedialog.asksaveasfilename(title="Exportar relatório Excel", defaultextension=".xlsx", filetypes=[("Excel workbook", "*.xlsx")])
    if not path:
        return
    if not path.lower().endswith(".xlsx"):
        path += ".xlsx"
    try:
        self.status_var.set("A exportar Excel..."); self.progress_var.set(5); self.update_idletasks()
        tmp = str(Path(path).with_suffix(".tmp.xlsx"))
        if os.path.exists(tmp):
            try: os.remove(tmp)
            except Exception: pass
        self._write_excel(tmp)
        try:
            if os.path.exists(path): os.remove(path)
            os.replace(tmp, path)
        except Exception:
            alt = str(Path(path).with_name(Path(path).stem + "_novo.xlsx"))
            os.replace(tmp, alt); path = alt
        self.progress_var.set(100)
        self.status_var.set(f"Excel exportado: {path}")
    except Exception as err:
        self.progress_var.set(0)
        messagebox.showerror("Erro", f"Não foi possível exportar Excel.\n\n{err}")

ColumnsEC2App.export_excel = _export_excel_v67


def _v67_pdf_para(text, style):
    from reportlab.platypus import Paragraph
    safe = str(text).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")
    return Paragraph(safe, style)


def _v67_pdf_table(df, cols, style, max_rows=40, width_mm=270):
    from reportlab.platypus import Table, Paragraph
    from reportlab.lib.units import mm
    present=[c for c in cols if df is not None and not df.empty and c in df.columns]
    if not present:
        return Table([[Paragraph("Sem dados.", style)]], colWidths=[width_mm*mm])
    data=[[Paragraph(str(c), style) for c in present]]
    for _, r in df.head(max_rows).iterrows():
        row=[]
        for c in present:
            v=r.get(c,"")
            if isinstance(v, float): txt="" if not math.isfinite(v) else f"{v:.2f}"
            else: txt="" if pd.isna(v) else str(v)
            row.append(Paragraph(txt.replace("&","&amp;").replace("<","&lt;").replace(">","&gt;"), style))
        data.append(row)
    widths=[width_mm*mm/max(1,len(present))]*len(present)
    t=Table(data, colWidths=widths, repeatRows=1)
    try:
        t.setStyle(_v64_pdf_style_table(True) if "_v64_pdf_style_table" in globals() else None)
    except Exception:
        pass
    return t


def _write_pdf_v67(self, path: str):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, PageBreak
    res = _v67_to_safe_df(getattr(self, "df_results", pd.DataFrame()))
    summ = _v67_to_safe_df(getattr(self, "df_summary", pd.DataFrame()))
    if summ is None or summ.empty: summ = res
    if "_v65_apply_module_statuses" in globals():
        try: res = _v67_to_safe_df(_v65_apply_module_statuses(res)); summ = _v67_to_safe_df(_v65_apply_module_statuses(summ))
        except Exception: pass
    n_ok=int(res.get("estado_global", res.get("status", pd.Series(dtype=str))).astype(str).eq("OK").sum()) if res is not None and not res.empty else 0
    n_warn=int(res.get("estado_global", res.get("status", pd.Series(dtype=str))).astype(str).eq("Aviso").sum()) if res is not None and not res.empty else 0
    n_fail=int(res.get("estado_global", res.get("status", pd.Series(dtype=str))).astype(str).eq("Falha").sum()) if res is not None and not res.empty else 0
    styles=getSampleStyleSheet()
    styles.add(ParagraphStyle(name="TitleV67", parent=styles["Title"], fontName="Courier-Bold", fontSize=14, leading=18, alignment=1))
    styles.add(ParagraphStyle(name="SubV67", parent=styles["Normal"], fontName="Courier", fontSize=9, leading=12, alignment=1, textColor=colors.darkgrey))
    styles.add(ParagraphStyle(name="H2V67", parent=styles["Heading2"], fontName="Courier-Bold", fontSize=11, leading=15, spaceBefore=8, spaceAfter=8))
    styles.add(ParagraphStyle(name="CellV67", parent=styles["Normal"], fontName="Courier", fontSize=7, leading=9))
    styles.add(ParagraphStyle(name="BodyV67", parent=styles["Normal"], fontName="Courier", fontSize=8.5, leading=12))
    doc=SimpleDocTemplate(path, pagesize=landscape(A4), leftMargin=12*mm, rightMargin=12*mm, topMargin=12*mm, bottomMargin=12*mm)
    story=[]
    story.append(Paragraph("Columns EC2", styles["TitleV67"]))
    story.append(Paragraph("Relatório técnico de dimensionamento de pilares", styles["SubV67"]))
    story.append(Spacer(1, 4*mm))
    meta=[
        ["Norma", _v59_norm_reference(self) if "_v59_norm_reference" in globals() else "Eurocódigo 2", "Data", datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["Casos", str(len(res)), "OK / Aviso / Falha", f"{n_ok} / {n_warn} / {n_fail}"],
        ["Classe de aço", _v59_getvar(self,"var_fyk","500") if "_v59_getvar" in globals() else "500", "Betão", _v67_material_classes_from_df(self)],
    ]
    mt=Table(meta, colWidths=[35*mm,100*mm,40*mm,95*mm])
    try: mt.setStyle(_v64_pdf_style_table(False))
    except Exception: pass
    story.append(mt)
    story.append(Spacer(1, 5*mm))
    story.append(Paragraph("1. Síntese executiva", styles["H2V67"]))
    story.append(Paragraph("A decisão global resulta dos estados resistentes, corte, torção, serviço e pormenorização. O ficheiro Excel contém a memória completa por combinação e as superfícies N-My-Mz exportadas.", styles["BodyV67"]))
    story.append(Paragraph("2. Decisão por prumada", styles["H2V67"]))
    cols=["prumada","member","case","material","b_cm","h_cm","n_ed_kN","my_ed_kNm","mz_ed_kNm","η_NMyMz","estado_global","estado_resistente","estado_pormenorizacao","solucao_completa"]
    story.append(_v67_pdf_table(summ, cols, styles["CellV67"], max_rows=34))
    module = _v65_module_status_table(res) if "_v65_module_status_table" in globals() and res is not None and not res.empty else pd.DataFrame()
    if module is not None and not module.empty:
        story.append(PageBreak()); story.append(Paragraph("3. Estados por módulo", styles["H2V67"]))
        story.append(_v67_pdf_table(module, ["prumada","member","case","estado_global","estado_resistente","estado_corte","estado_torcao","estado_els","estado_pormenorizacao","decisao_tecnica"], styles["CellV67"], max_rows=45))
    fails = _v65_failures(res) if "_v65_failures" in globals() and res is not None and not res.empty else pd.DataFrame()
    warns = _v65_warnings(res) if "_v65_warnings" in globals() and res is not None and not res.empty else pd.DataFrame()
    if (fails is not None and not fails.empty) or (warns is not None and not warns.empty):
        story.append(PageBreak()); story.append(Paragraph("4. Falhas e avisos", styles["H2V67"]))
        alert = pd.concat([fails, warns], ignore_index=True) if warns is not None and not warns.empty else fails
        story.append(_v67_pdf_table(alert, ["prumada","member","case","estado_global","failure_reason","detailing_warnings","decisao_tecnica"], styles["CellV67"], max_rows=50))
    story.append(Spacer(1,4*mm))
    story.append(Paragraph("Nota: os resultados devem ser validados nos pilares críticos, em particular quando existam alterações de geometria, comprimentos efectivos ou hipóteses de contraventamento.", styles["BodyV67"]))
    def footer(canvas, doc_obj):
        canvas.saveState(); canvas.setFont("Courier",7); canvas.setFillColor(colors.grey)
        canvas.drawString(12*mm, 7*mm, f"Columns EC2 | {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        canvas.drawRightString(285*mm, 7*mm, f"Página {doc_obj.page}")
        canvas.restoreState()
    doc.build(story, onFirstPage=footer, onLaterPages=footer)

ColumnsEC2App._write_pdf = _write_pdf_v67


def _export_pdf_report_v67(self):
    if getattr(self, "df_results", pd.DataFrame()) is None or self.df_results.empty:
        messagebox.showwarning("Aviso", "Não há resultados para exportar.")
        return
    path = filedialog.asksaveasfilename(title="Exportar relatório PDF", defaultextension=".pdf", filetypes=[("PDF", "*.pdf")])
    if not path: return
    if not path.lower().endswith(".pdf"): path += ".pdf"
    try:
        self.status_var.set("A exportar PDF..."); self.progress_var.set(5); self.update_idletasks()
        tmp=str(Path(path).with_suffix(".tmp.pdf"))
        if os.path.exists(tmp):
            try: os.remove(tmp)
            except Exception: pass
        self._write_pdf(tmp)
        try:
            if os.path.exists(path): os.remove(path)
            os.replace(tmp,path)
        except Exception:
            alt=str(Path(path).with_name(Path(path).stem + "_novo.pdf"))
            os.replace(tmp,alt); path=alt
        self.progress_var.set(100); self.status_var.set(f"PDF exportado: {path}")
    except Exception as err:
        self.progress_var.set(0)
        messagebox.showerror("Erro", f"Não foi possível exportar PDF.\n\n{err}")

ColumnsEC2App.export_pdf_report = _export_pdf_report_v67

# Remover exportação CSV da interface e neutralizar comando antigo.
def _export_csv_v67(self):
    messagebox.showinfo("Exportação CSV", "A exportação CSV foi removida. Use a exportação .xlsx para auditoria completa.")
ColumnsEC2App.export_csv = _export_csv_v67

_old_build_sidebar_v67 = ColumnsEC2App._build_sidebar

def _build_sidebar_v67(self, parent):
    _old_build_sidebar_v67(self, parent)
    def _walk(w):
        for child in list(w.winfo_children()):
            try:
                txt = child.cget("text")
                if str(txt).strip().lower() in ["exportar .csv", "exportar csv"]:
                    child.destroy(); continue
            except Exception:
                pass
            _walk(child)
    self.after(50, lambda: _walk(parent))
ColumnsEC2App._build_sidebar = _build_sidebar_v67

# Simplificar pós-processamento v6.6: evita repetições pesadas durante cálculo.
_old_run_design_v67 = ColumnsEC2App.run_design

def _run_design_v67(self):
    self.status_var.set("A preparar cálculo optimizado v6.7...")
    _old_run_design_v67(self)

ColumnsEC2App.run_design = _run_design_v67



# ============================================================
# ColumnsEC2 v6.8 — modularização inicial, performance, N-My-Mz,
# critérios parametrizáveis, pormenorização final e PDF por nível
# ============================================================
