# -*- coding: utf-8 -*-
# ColumnsEC2 v0.9 RC11 — Technical EN-UK presentation layer.
# This module does not alter the numerical design/check routines. It rewrites
# the EN-UK GUI/results/report/export presentation in technical engineering
# English and keeps the repository hyperlink on the ColumnsEC2 name.

APP_VERSION = "v0.9 RC11 Modular"

_RC11_REPO_URL = globals().get("GITHUB_URL", "https://github.com/lutondatomalela/ColumnsEC2")

# ---------------------------------------------------------------------------
# Technical English wording
# ---------------------------------------------------------------------------
_RC11_EXACT = {
    # UI sections / controls
    "Relatório PDF": "PDF report",
    "Nível de detalhe": "Report detail level",
    "Resumo executivo": "Executive summary",
    "Relatório técnico": "Technical report",
    "Memória de cálculo": "Detailed calculation note",
    "Memória detalhada": "Detailed calculation note",
    "Estratégia de reinforcement": "Reinforcement strategy",
    "Estratégia de armadura": "Reinforcement strategy",
    "Critério de escolha": "Selection criterion",
    "Critérios da estratégia equilibrada": "Balanced strategy criteria",
    "Diagnóstico and auditoria": "Diagnostics and audit",
    "Diagnóstico e auditoria": "Diagnostics and audit",
    "Diagnóstico structuralcodes": "structuralcodes diagnostics",
    "Tentativas de correcção": "Correction attempts",
    "Tentativas de correção": "Correction attempts",
    "Tentativas de correcção interactiva": "Interactive correction attempts",
    "Tentativas de correção interactiva": "Interactive correction attempts",
    "Classe desenv.": "Exposure class",
    "Classe de exposição": "Exposure class",
    "Classe de aço": "Steel grade",
    "Recobrimento nominal [mm]": "Nominal cover [mm]",
    "Capa nominal [mm]": "Nominal cover [mm]",
    "Betão": "Concrete grade",
    "Aço": "Steel",
    "Modo": "Mode",
    "Dimensionamento": "Design",
    "Pré-dimensionamento": "Preliminary sizing",
    "pre_dimensionamento": "Preliminary sizing",
    "dimensionamento": "Design",
    "Rigoroso": "Detailed numerical check",
    "Económica": "Minimum reinforcement",
    "Equilibrada": "Balanced",
    "Robusta": "Robust",
    "Aplicar": "Apply",
    "Linguagem": "Language",
    "Abrir repositório": "Open repository",
    "Exportar quadro .DXF": "Export column schedule .DXF",
    "Exportar quadro de pilares .DXF": "Export column schedule .DXF",
    "Tabela editável reconhecida": "Recognised editable table",
    "Texto colado": "Pasted text",
    "Interpretar texto": "Parse text",
    "Ler grelha": "Read grid",
    "Adicionar linha": "Add row",
    "Remover linha": "Remove row",
    "Limpar": "Clear",
    "Colar área de transferência": "Paste clipboard",
    "Importar ficheiro": "Import file",
    "Importar .xlsx/.csv": "Import file",
    "Ler caixa de texto": "Parse table",
    "Modelo de tabela": "Table template",
    "Calcular": "Design/check",
    "Design/check": "Design/check",
    "Exportar .xlsx": "Export .xlsx",
    "Relatório .pdf": "PDF report",
    "PDF report": "PDF report",
    "Notas rápidas": "Quick notes",
    "Verificações avançadas": "Advanced checks",
    "Advanced checks": "Advanced checks",
    "Combinação ELS": "SLS combination",
    "SLS combination": "SLS combination",
    "em branco = ELS simplificado por defeito": "blank = simplified SLS check",
    "blank = simplified SLS check": "blank = simplified SLS check",
    "Design standard / calculation engine": "Design standard / calculation engine",
    "Backend": "Calculation engine",
    "Creep": "Creep",
    "Relative humidity RH [%]": "Relative humidity RH [%]",
    "Concrete age t0 [days]": "Concrete age at loading t0 [days]",
    "t_ref / t0 [days]": "Reference age / loading age [days]",
    "9. Design standard / calculation engine": "9. Design standard / calculation engine",
    # Table headings
    "Grupo": "Group",
    "Objeto": "Component",
    "Objecto": "Component",
    "Estado": "Status",
    "Detalhe": "Details",
    "Verificação": "Design check",
    "Origem": "Source",
    "Nota": "Engineering note",
    "Categoria": "Category",
    "Resultado": "Result",
    "Item": "Item",
    "Campo": "Field",
    "Valor": "Value",
    "Prumada": "Column line",
    "Piso": "Storey",
    "Membro": "Member",
    "Caso": "Load case",
    "Solução": "Reinforcement arrangement",
    "Decisão técnica": "Engineering decision",
    "Falha": "Failure",
    "Falhas": "Failures",
    "Aviso": "Warning",
    "Avisos": "Warnings",
    "Recomendações": "Engineering recommendations",
    # Validation/backend values
    "Pacote": "Package",
    "Dependência": "Dependency",
    "Módulo": "Module",
    "API": "API",
    "Disponível": "Available",
    "Indisponível": "Unavailable",
    "desconhecida": "unknown",
    "Materiais": "Material properties",
    "Material": "Concrete grade",
    "2.ª ordem": "Second-order effects",
    "2a ordem": "Second-order effects",
    "N-My-Mz": "N-My-Mz interaction",
    "Esforço transverso": "Shear",
    "Torção": "Torsion",
    "ELS": "SLS",
    "ELS/fendilhação": "SLS crack control",
    "ELS/fendilhação/deformação": "SLS crack control/deflection",
    "Pormenorização": "Detailing",
    "ColumnsEC2 geométrico": "ColumnsEC2 geometry/detailing layer",
    "Calculado": "Calculated",
    "Calculado se API disponível": "Calculated when exposed by the installed API",
    "Não avaliado se API ausente": "Not assessed when the API is unavailable",
    "Informativo": "Informative",
    "Sem fallback interno.": "No internal fallback is used.",
    "Sem fallback interno": "No internal fallback is used.",
    "Motor interno existente; default.": "Existing internal engine; default calculation engine.",
    "Método interno actual.": "Current internal method.",
    "Superfície interna discreta.": "Internal discrete N-My-Mz resistance surface.",
    "Verificação interna actual.": "Current internal check.",
    "Combinação indicada ou simplificada.": "User-defined or simplified SLS combination.",
    "Regras internas de pormenorização.": "Internal detailing rules.",
    "Colunas obrigatórias": "Required input fields",
    "Tabela": "Input table",
    "Dados": "Input data",
    "Pares de nós": "End-node pairs",
    "Consistência entre nós": "End-node data consistency",
    "Unidades": "Units",
    "classe de betão": "concrete grade",
    "linhas importadas": "imported rows",
    "barras distintas": "unique members",
    "member vazio": "blank member identifier",
    "case vazio": "blank load-case identifier",
    "com 2 nós": "with two end-node rows",
    "com 1 nó": "with one end-node row",
    "com mais de 2 nós": "with more than two end-node rows",
    "presente": "present",
    "OK": "OK",
    "Não conforme": "Not compliant",
    "Verificar": "Check required",
    "Pré-dimensionado": "Preliminary sizing",
    "Sim": "Yes",
    "Não": "No",
}

# Phrases are deliberately technical, not literal translations.
_RC11_PHRASES = [
    # UI / long labels
    ("Análise e dimensionamento de pilares de betão armado", "Reinforced concrete column analysis and design"),
    ("Dimensionamento de pilares em betão armado: ELU, ELS, interacção N-My-Mz, pormenorização e relatórios técnicos.", "Reinforced concrete column design and checking: ULS, SLS, N-My-Mz interaction, detailing and technical reporting."),
    ("Dimensionamento de pilares em betão armado: ELU, ELS, interação N-My-Mz, pormenorização e relatórios técnicos.", "Reinforced concrete column design and checking: ULS, SLS, N-My-Mz interaction, detailing and technical reporting."),
    ("Default = EC2 Portugal 2010. EC2:2023 exige", "Default = Portuguese EC2 2010. EC2:2023 requires"),
    ("As tentativas detalhadas são exportadas apenas no ficheiro .xlsx.", "Detailed correction attempts are reported in the Excel workbook only."),
    ("Aplicado apenas quando a estratégia seleccionada é Equilibrada.", "Used only when the Balanced reinforcement strategy is selected."),
    ("Aplicado apenas quando a estratégia selecionada é Equilibrada.", "Used only when the Balanced reinforcement strategy is selected."),
    ("η alvo", "Target η"),
    ("η mínimo", "Minimum η"),
    ("η máximo", "Maximum η"),
    ("Excesso máx. As", "Maximum As surplus"),
    ("lido da coluna Material", "read from the Material column"),
    ("read from table (Material column)", "read from the Material column"),
    ("Betão read from the Material column", "Concrete grade read from the Material column"),
    ("betão read from the Material column", "concrete grade read from the Material column"),
    ("A exportar quadro de pilares por prumada/piso [mm]", "Exporting column schedule by column line/storey [mm]"),
    ("DXF column schedule exported", "DXF column schedule exported"),
    ("Excel exportado", "Excel workbook exported"),
    ("PDF exportado", "PDF report exported"),
    ("Tabela carregada", "Input table loaded"),
    ("tabela editável", "editable table"),
    ("Cálculo completo", "Calculation complete"),
    ("Cálculo concluído", "Calculation complete"),
    ("A calcular", "Calculating"),
    ("casos de envolvente", "governing cases"),
    ("falhas bloqueantes", "blocking design issues"),
    ("Falhas bloqueantes detectadas", "Blocking design issues detected"),
    ("Falhas bloqueantes detetadas", "Blocking design issues detected"),
    ("Foram detectadas falhas bloqueantes.", "Blocking design issues were detected."),
    ("Foram detetadas falhas bloqueantes.", "Blocking design issues were detected."),
    ("Pretende gerar propostas de correcção?", "Generate correction proposals?"),
    ("Pretende gerar propostas de correção?", "Generate correction proposals?"),
    # Validation / backend wording
    ("necessária para ELU/ELS/V/T/DXF", "required for ULS/SLS/shear/torsion and DXF output"),
    ("necessário para identificar combinação", "required to identify the load combination"),
    ("member/node/case deve estar preenchido", "Member/Node/Case must be populated"),
    ("cada member/case deve ter exactamente duas linhas", "each member/case pair should contain exactly two end-node rows"),
    ("cada member/case deve ter exatamente duas linhas", "each member/case pair should contain exactly two end-node rows"),
    ("sem M01/M02 completo", "M01/M02 cannot be reconstructed"),
    ("verificar duplicados ou resultados intermédios", "check for duplicate rows or intermediate station results"),
    ("verificar duplicados ou resultados intermediários", "check for duplicate rows or intermediate station results"),
    ("os dois nós do mesmo member/case devem ter dados geométricos compatíveis", "the two end-node rows of the same member/case pair should have compatible geometric data"),
    ("os dois nós do mesmo member/case devem ter dados", "the two end-node rows of the same member/case pair should have consistent data"),
    ("esperado em cm; verificar exportação", "expected in cm; check the source export"),
    ("esperado em m", "expected in metres"),
    ("pequenas=", "small="),
    ("grandes=", "large="),
    ("a classe deve vir da coluna Material; fallback interno C30/37 quando ausente", "the concrete grade should be provided in the Material column; C30/37 is used only as an internal fallback when the field is blank"),
    ("A geometria candidata é gerada pelo programa; a aceitação normativa depende das APIs disponíveis.", "Candidate geometry is generated by ColumnsEC2; code-level acceptance depends on the checks exposed by the installed structuralcodes API."),
    ("Calculado se API disponível", "Calculated when exposed by the installed API"),
    ("Não avaliado se API ausente", "Not assessed when the API is unavailable"),
    ("Sem fallback interno", "No internal fallback is used"),
    # Results/report wording
    ("Report interno — estados por módulo", "Internal module-status report"),
    ("Report interno", "Internal report"),
    ("estados por módulo", "module status"),
    ("Estados:", "Module status:"),
    ("global=", "overall="),
    ("resistente=", "resistance="),
    ("corte=", "shear="),
    ("torsion=", "torsion="),
    ("torcao=", "torsion="),
    ("torção=", "torsion="),
    ("ELS=", "SLS="),
    ("detailing=", "detailing="),
    ("corte requer verificação/dimensionamento de reinforcement transversal", "shear requires a detailed transverse-reinforcement design check"),
    ("corte requer verificação/dimensionamento de armadura transversal", "shear requires a detailed transverse-reinforcement design check"),
    ("shear requer verificação/dimensionamento de reinforcement transversal", "shear requires a detailed transverse-reinforcement design check"),
    ("torsion requer verificação/dimensionamento complementar", "torsion requires a complementary design check"),
    ("torção requer verificação/dimensionamento complementar", "torsion requires a complementary design check"),
    ("ELS informativo/não conclusivo ou a verificar", "SLS check is informative/non-conclusive; a detailed check is required"),
    ("SLS informativo/não conclusivo ou a verificar", "SLS check is informative/non-conclusive; a detailed check is required"),
    ("detailing construtiva a confirmar", "constructive detailing requires confirmation"),
    ("pormenorização construtiva a confirmar", "constructive detailing requires confirmation"),
    ("verificação/dimensionamento", "detailed design check"),
    ("dimensionamento complementar", "complementary design check"),
    ("a confirmar", "to be confirmed"),
    ("pré-dimensionamento: verificar em modo Dimensionamento antes de adoptar", "preliminary sizing only: run Design mode before adopting the reinforcement"),
    ("pré-dimensionamento: verificar em modo Dimensionamento antes de adotar", "preliminary sizing only: run Design mode before adopting the reinforcement"),
    ("verificar em modo Dimensionamento antes de adoptar", "run Design mode before adopting the reinforcement"),
    ("verificar em modo Dimensionamento antes de adotar", "run Design mode before adopting the reinforcement"),
    ("Falha de resistência biaxial", "N-My-Mz resistance failure"),
    ("Falha de pormenorização", "Detailing failure"),
    ("Falha por insuficiência de armadura", "Insufficient reinforcement"),
    ("Falha em verificações complementares", "Failure in complementary checks"),
    ("Falha de dados", "Input-data failure"),
    ("sem os dois nós necessários", "without the two required end-node rows"),
    ("nenhuma disposição admissível", "no admissible reinforcement layout"),
    ("disposição admissível", "admissible reinforcement layout"),
    ("nenhuma solução", "no solution"),
    ("interação biaxial", "N-My-Mz interaction"),
    ("interacção biaxial", "N-My-Mz interaction"),
    ("interação N-My-Mz", "N-My-Mz interaction"),
    ("interacção N-My-Mz", "N-My-Mz interaction"),
    ("OK sem armadura transversal resistente adicional", "OK without additional shear links"),
    ("Requer armadura de esforço transverso", "shear links required"),
    ("Requer armadura de torção", "torsion reinforcement required"),
    ("Sem aviso relevante", "No relevant warning"),
    ("Sem torção relevante", "No relevant torsion"),
    ("Torção desprezável — não condicionante", "Negligible torsion — not governing"),
    # Reinforcement wording
    ("sem intermediate cross-ties", "no intermediate cross-ties"),
    ("sem cross-ties intermédios", "no intermediate cross-ties"),
    ("sem grampos intermédios", "no intermediate cross-ties"),
    ("sem grampos", "no cross-ties"),
    ("distribuídos nas faces", "face bars"),
    ("distribuídas nas faces", "face bars"),
    ("a meio das faces", "at mid-face"),
    ("nos cantos", "corner bars"),
    ("no canto", "corner bar"),
    ("por nível", "per link level"),
    ("grampo(s)", "intermediate cross-tie(s)"),
    ("grampos intermédios", "intermediate cross-ties"),
    ("grampos intermediários", "intermediate cross-ties"),
    ("grampos", "cross-ties"),
    ("Grampos", "Cross-ties"),
    ("estribos", "links"),
    ("Estribos", "Links"),
    ("varões comprimidos", "compression bars"),
    ("varões", "bars"),
    ("Varões", "Bars"),
    ("armadura transversal", "transverse reinforcement"),
    ("armadura longitudinal", "longitudinal reinforcement"),
    ("armadura", "reinforcement"),
    ("pormenorização construtiva", "constructive detailing"),
    ("pormenorização", "detailing"),
    ("Pormenorização", "Detailing"),
    ("faces longas", "long faces"),
    ("faces curtas", "short faces"),
    # Generic Portuguese engineering terms
    ("fendilhação", "crack control"),
    ("deformação", "deflection"),
    ("fluência", "creep"),
    ("retração", "shrinkage"),
    ("retracção", "shrinkage"),
    ("esforço transverso", "shear"),
    ("esforços transversos", "shear forces"),
    ("corte", "shear"),
    ("Corte", "Shear"),
    ("torção", "torsion"),
    ("Torção", "Torsion"),
    ("prumada", "column line"),
    ("Prumada", "Column line"),
    ("piso", "storey"),
    ("Piso", "Storey"),
    ("tramo", "segment"),
    ("Tramo", "Segment"),
    ("caso", "case"),
    ("Caso", "Load case"),
    ("membro", "member"),
    ("Membro", "Member"),
    ("Bloqueante", "Blocking design issue"),
    ("bloqueante", "blocking design issue"),
    ("Não avaliado", "Not assessed"),
    ("não avaliado", "not assessed"),
    ("não calculado", "not calculated"),
    ("não exposto", "not exposed"),
    ("não exposta", "not exposed"),
    ("não conclusivo", "non-conclusive"),
    ("Informativo", "Informative"),
    ("informativo", "informative"),
    ("Calculado", "Calculated"),
    ("calculado", "calculated"),
    ("Disponível", "Available"),
    ("disponível", "available"),
    ("presente", "present"),
    ("faltante", "missing"),
    ("ausente", "missing"),
    ("Betão", "Concrete"),
    ("betão", "concrete grade"),
    ("combinação", "load combination"),
    ("Combinação", "Load combination"),
    ("envolvente", "governing-case envelope"),
    ("Motor", "Calculation engine"),
    ("Norma", "Standard"),
    ("Método", "Method"),
    ("método", "method"),
    ("Verificação", "Design check"),
    ("verificação", "check"),
    ("Relatório", "Report"),
]

_RC11_COLUMNS = {
    # Validation / diagnostics
    "Grupo": "Group",
    "Objeto": "Component",
    "Objecto": "Component",
    "Estado": "Status",
    "Detalhe": "Details",
    "Verificação": "Design check",
    "Origem": "Source",
    "Nota": "Engineering note",
    "Categoria": "Category",
    "Resultado": "Result",
    "Item": "Item",
    "Campo": "Field",
    "Valor": "Value",
    # Results
    "member": "Member",
    "case": "Load case",
    "name": "Column line",
    "story": "Storey",
    "Story": "Storey",
    "piso": "Storey",
    "Piso": "Storey",
    "Prumada": "Column line",
    "prumada": "Column line",
    "tramo": "Segment",
    "Tramo": "Segment",
    "material": "Concrete grade",
    "status": "Status",
    "Estado": "Status",
    "estado_global": "Overall status",
    "estado_resistente": "Resistance status",
    "estado_corte": "Shear status",
    "estado_torcao": "Torsion status",
    "estado_els": "SLS status",
    "estado_pormenorizacao": "Detailing status",
    "decisao_tecnica": "Engineering decision",
    "failure_reason": "Failure reason",
    "failure_type": "Failure type",
    "recommendations": "Engineering recommendations",
    "shortlist_text": "Candidate layouts",
    "solucao": "Reinforcement arrangement",
    "solucao_completa": "Full reinforcement arrangement",
    "pormenorizacao_construtiva": "Constructive detailing",
    "detalhe_grampos": "Cross-tie arrangement",
    "grampos_intermedios": "Intermediate cross-ties",
    "numero_grampos_por_nivel": "Cross-ties per link level",
    "ramos_estribo_y": "Link legs in y",
    "ramos_estribo_z": "Link legs in z",
    "as_req_mm2": "As,req [mm²]",
    "as_prov_mm2": "As,prov [mm²]",
    "as_min_mm2": "As,min [mm²]",
    "as_max_mm2": "As,max [mm²]",
    "n_ed_kN": "N_Ed [kN]",
    "my_ed_kNm": "My,Ed [kNm]",
    "mz_ed_kNm": "Mz,Ed [kNm]",
    "mrd_y_kNm": "MRd,y [kNm]",
    "mrd_z_kNm": "MRd,z [kNm]",
    "utilizacao": "η_NMyMz",
    "phi_long_mm": "Longitudinal bar diameter [mm]",
    "phi_st_mm": "Link diameter [mm]",
    "s_st_mm": "Link spacing [mm]",
    "s_st_max_mm": "Maximum link spacing [mm]",
    "n_total": "Total number of bars",
    "n_bars_y": "Bars on y-faces",
    "n_bars_z": "Bars on z-faces",
    "Shortlist": "Candidate layouts",
    "Technical decision": "Engineering decision",
}


def _rc11_is_en(app=None):
    try:
        app = app or globals().get("_RC6_ACTIVE_APP", None)
        return str(app.var_language.get()).upper().startswith("EN")
    except Exception:
        return False


def _rc11_is_nan(value):
    try:
        result = pd.isna(value)
        if isinstance(result, (pd.Series, pd.DataFrame, list, tuple)):
            return False
        return bool(result)
    except Exception:
        return False


def _rc11_to_technical_en(value):
    """Rewrite UI/result text in technical EN-UK. Numeric values are untouched."""
    if value is None or _rc11_is_nan(value):
        return value
    if not isinstance(value, str):
        return value
    s = value
    stripped = s.strip()
    if stripped in _RC11_EXACT:
        return _RC11_EXACT[stripped]
    try:
        if "_rc8_translate_technical_en" in globals():
            s = _rc8_translate_technical_en(s)
    except Exception:
        pass
    stripped = s.strip()
    if stripped in _RC11_EXACT:
        return _RC11_EXACT[stripped]
    # Longest phrases first, so full technical wording wins over word fragments.
    for old, new in sorted(_RC11_PHRASES, key=lambda kv: len(kv[0]), reverse=True):
        s = s.replace(old, new)
    # Regex cleanup for remaining isolated Portuguese words and mixed output.
    try:
        import re as _re
        word_map = {
            r"\bGrupo\b": "Group",
            r"\bObjeto\b": "Component",
            r"\bObjecto\b": "Component",
            r"\bEstado\b": "Status",
            r"\bDetalhe\b": "Details",
            r"\bVerificação\b": "Design check",
            r"\bOrigem\b": "Source",
            r"\bNota\b": "Engineering note",
            r"\bCategoria\b": "Category",
            r"\bResultado\b": "Result",
            r"\bFalha\b": "Failure",
            r"\bAviso\b": "Warning",
            r"\bDisponível\b": "Available",
            r"\bIndisponível\b": "Unavailable",
            r"\bCalculado\b": "Calculated",
            r"\bpresente\b": "present",
            r"\bnecessári[ao]s?\b": "required",
        }
        for pattern, repl in word_map.items():
            s = _re.sub(pattern, repl, s)
    except Exception:
        pass
    # Mixed-language artefact cleanup.
    cleanup = {
        "Design/check interna actual": "Current internal check",
        "Current internal design check.": "Current internal check.",
        "Design check interna actual.": "Current internal check.",
        "concrete grade read from the Material column lido da coluna Material": "concrete grade read from the Material column",
        "Concrete concrete grade": "Concrete grade",
        "concrete grade grade": "concrete grade",
        "reinforcement transversal": "transverse reinforcement",
        "shear requires a detailed transverse-reinforcement detailed design check": "shear requires a detailed transverse-reinforcement design check",
        "torsion requires a complementary detailed design check": "torsion requires a complementary design check",
        "SLS check is Informative/non-conclusive": "SLS check is informative/non-conclusive",
        "constructive detailing to be confirmed": "constructive detailing requires confirmation",
        "constructive detailing requires confirmation.": "constructive detailing requires confirmation.",
        "intermediate intermediate cross-ties": "intermediate cross-ties",
        "no cross-ties intermédios": "no intermediate cross-ties",
        "sem intermediate cross-ties": "no intermediate cross-ties",
        "links Ø": "links Ø",
        "1 intermediate cross-tie(s)": "1 intermediate cross-tie",
        "2 intermediate cross-tie(s)": "2 intermediate cross-ties",
        "3 intermediate cross-tie(s)": "3 intermediate cross-ties",
        "4 intermediate cross-tie(s)": "4 intermediate cross-ties",
        "Failure s": "Failures",
        "Warning s": "Warnings",
        "case s": "cases",
        "load case s": "load cases",
        "rows read": "rows read",
    }
    for old, new in cleanup.items():
        s = s.replace(old, new)
    return s


def _rc11_header_en(col):
    s = str(col)
    if s in _RC11_COLUMNS:
        return _RC11_COLUMNS[s]
    try:
        if "_RC8_HEADER_MAP" in globals() and s in _RC8_HEADER_MAP:
            return _RC8_HEADER_MAP[s]
    except Exception:
        pass
    return _rc11_to_technical_en(s)


def _rc11_prepare_display_df(df, lang=None):
    if df is None or getattr(df, "empty", True):
        return pd.DataFrame()
    out = df.copy()
    if lang is None:
        lang = LANG_EN if _rc11_is_en() else LANG_PT
    if lang != LANG_EN:
        return out
    try:
        if "_rc3_deduplicate_columns" in globals():
            out = _rc3_deduplicate_columns(out)
    except Exception:
        pass
    out.columns = [_rc11_header_en(c) for c in out.columns]
    for c in list(out.columns):
        try:
            if out[c].dtype == object or str(out[c].dtype).startswith("string"):
                out[c] = out[c].map(_rc11_to_technical_en)
        except Exception:
            pass
    return out


def _rc11_rewrite_treeview(tree):
    """Translate already-populated Treeview headings and cells in place."""
    try:
        cols = list(tree["columns"])
    except Exception:
        return
    new_cols = []
    for c in cols:
        try:
            current = tree.heading(c).get("text", c)
            tree.heading(c, text=_rc11_header_en(current))
        except Exception:
            pass
        new_cols.append(c)
    try:
        for item in tree.get_children(""):
            values = list(tree.item(item, "values"))
            values = [_rc11_to_technical_en(v) for v in values]
            tree.item(item, values=values)
    except Exception:
        pass


def _rc11_walk_widgets(widget):
    try:
        children = widget.winfo_children()
    except Exception:
        children = []
    for child in children:
        yield child
        yield from _rc11_walk_widgets(child)


def _rc11_set_widget_texts(app):
    for w in _rc11_walk_widgets(app):
        try:
            txt = str(w.cget("text"))
            new = _rc11_to_technical_en(txt)
            if new != txt:
                w.configure(text=new)
        except Exception:
            pass
        try:
            if isinstance(w, ttk.LabelFrame):
                txt = str(w.cget("text"))
                new = _rc11_to_technical_en(txt)
                if new != txt:
                    w.configure(text=new)
        except Exception:
            pass
        try:
            if isinstance(w, ttk.Combobox):
                values = list(w.cget("values") or [])
                new_values = [_rc11_to_technical_en(str(v)) for v in values]
                if new_values != values:
                    w.configure(values=new_values)
                cur = str(w.get())
                new_cur = _rc11_to_technical_en(cur)
                if new_cur != cur:
                    w.set(new_cur)
        except Exception:
            pass
    # Notebook tabs.
    for w in _rc11_walk_widgets(app):
        try:
            if isinstance(w, ttk.Notebook):
                for tab_id in w.tabs():
                    txt = str(w.tab(tab_id, "text"))
                    w.tab(tab_id, text=_rc11_to_technical_en(txt))
        except Exception:
            pass
    # All visible Treeview content.
    for w in _rc11_walk_widgets(app):
        try:
            if isinstance(w, ttk.Treeview):
                _rc11_rewrite_treeview(w)
        except Exception:
            pass
    try:
        app.title("ColumnsEC2 - Reinforced Concrete Column Analysis and Design (EC2)")
    except Exception:
        pass
    try:
        if "_rc7_bind_repository_links" in globals():
            _rc7_bind_repository_links(app)
    except Exception:
        pass


def _rc11_refresh_known_trees(app):
    refresh_pairs = [
        ("tree_input", "df_clean"),
        ("tree_pairs", "df_pair"),
        ("tree_validation", "df_validation"),
        ("tree_results", "df_results"),
        ("tree_summary", "df_summary"),
        ("tree_failures", "df_failures"),
        ("tree_notes", "df_notes"),
        ("tree_backend_coverage", "df_backend_coverage"),
        ("tree_structuralcodes_diag", "df_structuralcodes_diag"),
        ("tree_import_quality", "df_import_quality"),
    ]
    for tree_attr, df_attr in refresh_pairs:
        try:
            tree = getattr(app, tree_attr, None)
            df = getattr(app, df_attr, pd.DataFrame())
            if tree is not None and df is not None and not getattr(df, "empty", True):
                _rc11_show_df(app, tree, df)
            elif tree is not None:
                _rc11_rewrite_treeview(tree)
        except Exception:
            pass
    try:
        if hasattr(app, "tree_shortlists") and hasattr(app, "build_shortlists_df"):
            df = app.build_shortlists_df()
            if df is not None and not getattr(df, "empty", True):
                _rc11_show_df(app, app.tree_shortlists, df)
    except Exception:
        pass


# ---------------------------------------------------------------------------
# Override display/report methods
# ---------------------------------------------------------------------------
_rc11_prev_show_df = getattr(ColumnsEC2App, "show_df", None)

def _rc11_show_df(self, tree, df):
    global _RC6_ACTIVE_APP
    _RC6_ACTIVE_APP = self
    display = df
    if _rc11_is_en(self):
        display = _rc11_prepare_display_df(df, LANG_EN)
    if callable(_rc11_prev_show_df):
        result = _rc11_prev_show_df(self, tree, display)
    else:
        result = None
    if _rc11_is_en(self):
        try:
            _rc11_rewrite_treeview(tree)
        except Exception:
            pass
    return result

ColumnsEC2App.show_df = _rc11_show_df


def _rc11_format_value(v, digits=2):
    try:
        if isinstance(v, float):
            return "" if not math.isfinite(v) else f"{v:.{digits}f}"
        if _rc11_is_nan(v):
            return ""
    except Exception:
        pass
    return _rc11_to_technical_en(str(v))


def _rc11_build_report_en(app):
    res = getattr(app, "df_results", pd.DataFrame())
    if res is None or res.empty:
        return "No calculation results. Import the input table and run the design/check."
    src = getattr(app, "df_summary", pd.DataFrame())
    if src is None or src.empty:
        src = res
    src_d = _rc11_prepare_display_df(src, LANG_EN)
    total = len(res)
    try:
        st = _rc3_col_series(res, ["estado_global", "status", "Estado", "Status"]).astype(str)
    except Exception:
        st = pd.Series(dtype=str)
    n_fail = int((st == "Falha").sum()) if not st.empty else 0
    n_warn = int((st == "Aviso").sum()) if not st.empty else 0
    lines = [
        f"ColumnsEC2 {APP_VERSION}\n",
        "Technical design/check report — reinforced concrete columns\n\n",
        f"Analysed design cases: {total} | Blocking failures: {n_fail} | Warnings: {n_warn}\n\n",
    ]
    for _, r in src_d.head(120).iterrows():
        col_line = r.get("Column line", r.get("Name", ""))
        member = r.get("Member", "")
        case = r.get("Load case", r.get("Case", ""))
        ned = r.get("N_Ed [kN]", r.get("NEd [kN]", ""))
        my = r.get("My,Ed [kNm]", r.get("My_Ed [kNm]", ""))
        mz = r.get("Mz,Ed [kNm]", r.get("Mz_Ed [kNm]", ""))
        eta = r.get("η_NMyMz", "")
        overall = r.get("Overall status", r.get("Status", ""))
        resistance = r.get("Resistance status", "")
        shear = r.get("Shear status", "")
        torsion = r.get("Torsion status", "")
        sls = r.get("SLS status", "")
        detailing = r.get("Detailing status", "")
        sol = r.get("Reinforcement arrangement", r.get("Full reinforcement arrangement", ""))
        decision = r.get("Engineering decision", r.get("Engineering recommendations", ""))
        lines.append(f"Column line {col_line} | Member {member} | Load case {case}\n")
        lines.append(f"N_Ed={_rc11_format_value(ned)} kN | My,Ed={_rc11_format_value(my)} kNm | Mz,Ed={_rc11_format_value(mz)} kNm | η_NMyMz={_rc11_format_value(eta,3)}\n")
        if any(str(x).strip() for x in [overall, resistance, shear, torsion, sls, detailing]):
            lines.append(
                "Module status: "
                f"overall={_rc11_to_technical_en(str(overall))} | "
                f"resistance={_rc11_to_technical_en(str(resistance))} | "
                f"shear={_rc11_to_technical_en(str(shear))} | "
                f"torsion={_rc11_to_technical_en(str(torsion))} | "
                f"SLS={_rc11_to_technical_en(str(sls))} | "
                f"detailing={_rc11_to_technical_en(str(detailing))}\n"
            )
        if str(sol).strip():
            lines.append(f"Reinforcement arrangement: {_rc11_to_technical_en(str(sol))}\n")
        if str(decision).strip():
            lines.append(f"Engineering decision: {_rc11_to_technical_en(str(decision))}\n")
        lines.append("\n")
    return "".join(lines)

_rc11_prev_update_report = getattr(ColumnsEC2App, "update_report", None)

def _rc11_update_report(self):
    if _rc11_is_en(self):
        try:
            self.report_txt.delete("1.0", "end")
            self.report_txt.insert("1.0", _rc11_build_report_en(self))
            return None
        except Exception:
            pass
    return _rc11_prev_update_report(self) if callable(_rc11_prev_update_report) else None

ColumnsEC2App.update_report = _rc11_update_report


# ---------------------------------------------------------------------------
# Language application / callbacks
# ---------------------------------------------------------------------------
_rc11_prev_apply_language = getattr(ColumnsEC2App, "apply_language", None)

def _rc11_apply_language(self):
    global _RC6_ACTIVE_APP
    _RC6_ACTIVE_APP = self
    result = _rc11_prev_apply_language(self) if callable(_rc11_prev_apply_language) else None
    if _rc11_is_en(self):
        try:
            _rc11_set_widget_texts(self)
            _rc11_refresh_known_trees(self)
            self.update_report()
            if hasattr(self, "status_var"):
                self.status_var.set(_rc11_to_technical_en(str(self.status_var.get())))
        except Exception:
            pass
    return result

ColumnsEC2App.apply_language = _rc11_apply_language
try:
    _rc3_apply_language = _rc11_apply_language
except Exception:
    pass


def _rc11_schedule_refresh(app):
    if not _rc11_is_en(app):
        return
    for delay in (100, 400, 1000, 2500):
        try:
            app.after(delay, lambda a=app: (a.apply_language(), _rc11_refresh_known_trees(a)))
        except Exception:
            pass

# Wrap common callbacks that repopulate status/tree/report content.
for _name in ["load_df", "run_design", "export_dxf", "export_excel", "export_pdf_report", "update_report"]:
    try:
        _prev = getattr(ColumnsEC2App, _name)
    except Exception:
        continue
    if _name == "update_report":
        continue
    def _make_wrapper(name, prev):
        def _wrapped(self, *args, **kwargs):
            global _RC6_ACTIVE_APP
            _RC6_ACTIVE_APP = self
            out = prev(self, *args, **kwargs)
            try:
                _rc11_schedule_refresh(self)
            except Exception:
                pass
            return out
        _wrapped.__name__ = f"_rc11_{name}"
        return _wrapped
    setattr(ColumnsEC2App, _name, _make_wrapper(_name, _prev))

# Message boxes.
_rc11_prev_showinfo = messagebox.showinfo
_rc11_prev_showwarning = messagebox.showwarning
_rc11_prev_showerror = messagebox.showerror
_rc11_prev_askyesno = messagebox.askyesno

def _rc11_msg(title, message):
    app = globals().get("_RC6_ACTIVE_APP", None)
    if app is not None and _rc11_is_en(app):
        return _rc11_to_technical_en(str(title)), _rc11_to_technical_en("" if message is None else str(message))
    return title, message

def _rc11_showinfo(title, message=None, *args, **kwargs):
    t, m = _rc11_msg(title, message)
    return _rc11_prev_showinfo(t, m, *args, **kwargs)

def _rc11_showwarning(title, message=None, *args, **kwargs):
    t, m = _rc11_msg(title, message)
    return _rc11_prev_showwarning(t, m, *args, **kwargs)

def _rc11_showerror(title, message=None, *args, **kwargs):
    t, m = _rc11_msg(title, message)
    return _rc11_prev_showerror(t, m, *args, **kwargs)

def _rc11_askyesno(title, message=None, *args, **kwargs):
    t, m = _rc11_msg(title, message)
    return _rc11_prev_askyesno(t, m, *args, **kwargs)

messagebox.showinfo = _rc11_showinfo
messagebox.showwarning = _rc11_showwarning
messagebox.showerror = _rc11_showerror
messagebox.askyesno = _rc11_askyesno

# Excel post-processing.
_rc11_prev_write_excel = getattr(ColumnsEC2App, "_write_excel", None)

def _write_excel_rc11(self, path: str):
    if callable(_rc11_prev_write_excel):
        _rc11_prev_write_excel(self, path)
    if not _rc11_is_en(self):
        return
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path)
        for ws in wb.worksheets:
            ws.title = _rc11_to_technical_en(ws.title)[:31]
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 8000), max_col=min(ws.max_column, 180)):
                for cell in row:
                    if isinstance(cell.value, str):
                        cell.value = _rc11_header_en(cell.value) if cell.row == 1 else _rc11_to_technical_en(cell.value)
        try:
            wb.properties.title = "ColumnsEC2 - Reinforced Concrete Column Analysis and Design"
            wb.properties.subject = "Reinforced concrete column design and checking"
            wb.properties.description = "Technical workbook generated by ColumnsEC2. The programme name contains the repository hyperlink."
        except Exception:
            pass
        try:
            if "_rc7_apply_workbook_links" in globals():
                _rc7_apply_workbook_links(wb)
        except Exception:
            pass
        wb.save(path)
    except Exception:
        pass

ColumnsEC2App._write_excel = _write_excel_rc11

# Initial hook used by CLI/GUI launcher.
def _v092_apply_language_title(app):
    global _RC6_ACTIVE_APP
    _RC6_ACTIVE_APP = app
    try:
        app.apply_language()
    except Exception:
        pass
    try:
        if _rc11_is_en(app):
            _rc11_set_widget_texts(app)
            _rc11_refresh_known_trees(app)
    except Exception:
        pass
    try:
        if "_rc7_bind_repository_links" in globals():
            _rc7_bind_repository_links(app)
    except Exception:
        pass

# Convenience alias for tests and future modules.
technical_en_uk = _rc11_to_technical_en
