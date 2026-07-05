# -*- coding: utf-8 -*-
# Auto-split from ColumnsEC2 v0.9 RC8.
# This module is executed in the shared runtime namespace by columns_ec2.runtime.loader.
# Keep execution order defined in columns_ec2/runtime/manifest.py.

# ============================================================
# ColumnsEC2 v4.0 — extensões modulares: superfície resistente,
# ELS por combinação escolhida, DXF, memória de cálculo detalhada,
# validação robusta, esforço transverso, torção e pormenorização.
# ============================================================

APP_VERSION = "v4.5"


def _finite(x, default=0.0):
    v = safe_float(x, default)
    return v if math.isfinite(v) else default


# --------------------------- validação robusta tabela de cálculo ---------------------------
def _validation_rows_v4(self, pre_calc: bool = False) -> pd.DataFrame:
    rows = []
    df = self.df_clean if self.df_clean is not None else pd.DataFrame()
    pair = self.df_pair if self.df_pair is not None else pd.DataFrame()
    required = ["member_case", "fx", "fy", "fz", "mx", "my", "mz", "length", "material", "hy", "hz", "ax", "iy", "iz", "name"]
    for c in required:
        ok = (not df.empty) and c in df.columns
        rows.append({
            "Categoria": "Colunas obrigatórias",
            "Item": c,
            "Estado": "OK" if ok else "Não conforme",
            "Resultado": "presente" if ok else "em falta",
            "Nota": "necessária para ELU/ELS/V/T/DXF" if ok else "corrigir cabeçalho da folha de importação",
        })

    if df.empty:
        rows.append({"Categoria": "Tabela", "Item": "linhas", "Estado": "Não conforme", "Resultado": 0, "Nota": "sem dados importados"})
        return pd.DataFrame(rows)

    rows.append({"Categoria": "Tabela", "Item": "linhas importadas", "Estado": "OK", "Resultado": len(df), "Nota": "linhas lidas"})
    rows.append({"Categoria": "Tabela", "Item": "members", "Estado": "OK", "Resultado": df["member"].astype(str).nunique() if "member" in df.columns else 0, "Nota": "barras distintas"})

    if "member" in df.columns:
        empty_member = int(df["member"].astype(str).str.strip().isin(["", "nan", "None"]).sum())
        rows.append({"Categoria": "Dados", "Item": "member vazio", "Estado": "OK" if empty_member == 0 else "Não conforme", "Resultado": empty_member, "Nota": "member/node/case deve estar preenchido"})
    if "case" in df.columns:
        empty_case = int(df["case"].astype(str).str.strip().isin(["", "nan", "None"]).sum())
        rows.append({"Categoria": "Dados", "Item": "case vazio", "Estado": "OK" if empty_case == 0 else "Verificar", "Resultado": empty_case, "Nota": "necessário para identificar combinação"})

    # Pares de nós
    if not pair.empty and "n_nodes_found" in pair.columns:
        n1 = int((pair["n_nodes_found"].fillna(0).astype(float) == 1).sum())
        n2 = int((pair["n_nodes_found"].fillna(0).astype(float) == 2).sum())
        n_more = int((pair["n_nodes_found"].fillna(0).astype(float) > 2).sum())
        rows.append({"Categoria": "Pares de nós", "Item": "com 2 nós", "Estado": "OK" if n1 == 0 and n_more == 0 else "Verificar", "Resultado": f"{n2}/{len(pair)}", "Nota": "cada member/case deve ter exactamente duas linhas"})
        rows.append({"Categoria": "Pares de nós", "Item": "com 1 nó", "Estado": "OK" if n1 == 0 else "Não conforme", "Resultado": n1, "Nota": "sem M01/M02 completo"})
        rows.append({"Categoria": "Pares de nós", "Item": "com mais de 2 nós", "Estado": "OK" if n_more == 0 else "Verificar", "Resultado": n_more, "Nota": "verificar duplicados ou resultados intermédios"})

    # Consistência entre nós por group
    geom_cols = ["length", "material", "hy", "hz", "ax", "iy", "iz"]
    inconsistent = {c: 0 for c in geom_cols}
    try:
        for _, grp in df.groupby(["member", "case", "name"], dropna=False):
            if len(grp) < 2:
                continue
            for c in geom_cols:
                if c not in grp.columns:
                    continue
                vals = grp[c].dropna().astype(str).str.strip().unique()
                if len(vals) > 1:
                    # tolerância para numéricos
                    nums = [safe_float(v, float('nan')) for v in vals]
                    nums = [v for v in nums if math.isfinite(v)]
                    if len(nums) >= 2 and max(nums)-min(nums) <= 1e-6:
                        continue
                    inconsistent[c] += 1
    except Exception:
        pass
    for c, count in inconsistent.items():
        rows.append({"Categoria": "Consistência entre nós", "Item": c, "Estado": "OK" if count == 0 else "Verificar", "Resultado": count, "Nota": "os dois nós do mesmo member/case devem ter dados geométricos compatíveis"})

    # Unidades suspeitas
    def _count(cond):
        try:
            return int(cond.sum())
        except Exception:
            return 0
    if "hy" in df.columns and "hz" in df.columns:
        small = _count((df["hy"].astype(float) < 10) | (df["hz"].astype(float) < 10))
        large = _count((df["hy"].astype(float) > 300) | (df["hz"].astype(float) > 300))
        rows.append({"Categoria": "Unidades", "Item": "dimensões HY/HZ", "Estado": "OK" if small == 0 and large == 0 else "Verificar", "Resultado": f"pequenas={small}; grandes={large}", "Nota": "esperado em cm; verificar exportação"})
    if "length" in df.columns:
        long = _count(df["length"].astype(float) > 20)
        zero = _count(df["length"].astype(float) <= 0)
        rows.append({"Categoria": "Unidades", "Item": "Length", "Estado": "OK" if long == 0 and zero == 0 else "Verificar", "Resultado": f">20m={long}; <=0={zero}", "Nota": "esperado em m"})

    # Materiais
    bad_mat = 0
    if "material" in df.columns:
        bad_mat = int(~df["material"].astype(str).str.contains(r"C\s*\d+\s*/\s*\d+", case=False, regex=True, na=False).sum())
        # above gives bool arithmetic badly if inverted scalar; recompute safely
        bad_mat = int((~df["material"].astype(str).str.contains(r"C\s*\d+\s*/\s*\d+", case=False, regex=True, na=False)).sum())
    rows.append({"Categoria": "Materiais", "Item": "classe de betão", "Estado": "OK" if bad_mat == 0 else "Verificar", "Resultado": bad_mat, "Nota": "a classe deve vir da coluna Material; fallback interno C30/37 quando ausente"})

    if not pre_calc and self.df_results is not None and not self.df_results.empty:
        res = self.df_results
        rows.append({"Categoria": "Cálculo", "Item": "casos calculados", "Estado": "OK", "Resultado": len(res), "Nota": "linhas processadas"})
        n_fail = int((res.get("status", pd.Series(dtype=str)) == "Falha").sum())
        rows.append({"Categoria": "Cálculo", "Item": "falhas", "Estado": "OK" if n_fail == 0 else "Verificar", "Resultado": n_fail, "Nota": "consultar separador Falhas"})
        for col in ["shear_status", "torsion_status", "service_status", "detailing_status"]:
            if col in res.columns:
                n_ver = int(res[col].astype(str).str.contains("Verificar|Não conforme|Aviso|Requer", case=False, na=False).sum())
                rows.append({"Categoria": "Cálculo", "Item": col, "Estado": "OK" if n_ver == 0 else "Verificar", "Resultado": n_ver, "Nota": "verificar linhas assinaladas"})
    return pd.DataFrame(rows)


# --------------------------- pormenorização EC2 prática ---------------------------
def detailing_check_v4(result: dict) -> dict:
    b = _finite(result.get("b_cm"), 0.0) * 10.0
    h = _finite(result.get("h_cm"), 0.0) * 10.0
    phi = _finite(result.get("phi_long_mm"), 0.0)
    phi_st = _finite(result.get("phi_st_mm"), 0.0)
    s = _finite(result.get("s_st_mm"), 0.0)
    n_total = int(_finite(result.get("n_total"), 0))
    n_y = int(_finite(result.get("n_bars_y"), 0))
    n_z = int(_finite(result.get("n_bars_z"), 0))
    cover = _finite(result.get("cover_mm", 35.0), 35.0)
    issues = []
    status = "OK"
    min_bars = 6 if abs(b - h) < 1e-6 and "circ" in str(result.get("name", "")).lower() else 4
    if n_total and n_total < min_bars:
        issues.append(f"número mínimo de varões não cumprido ({n_total}<{min_bars})")
    if phi and phi < 10:
        issues.append("diâmetro longitudinal inferior a Ø10")
    if phi and phi_st and phi_st < max(6.0, phi/4.0):
        issues.append("diâmetro dos estribos inferior a max(6 mm; Øl/4)")
    smax = min(12.0*phi if phi else 999, min(b,h) if b and h else 999, 300.0)
    if s and s > smax + 1e-6:
        issues.append(f"espaçamento dos estribos superior ao limite ({s:.0f}>{smax:.0f} mm)")
    edge = cover + phi_st + phi/2.0
    max_ctc = 0.0
    min_clear = 1e9
    if b > 0 and n_y > 1:
        ctc = (b - 2*edge) / max(n_y-1,1)
        max_ctc = max(max_ctc, ctc)
        min_clear = min(min_clear, ctc - phi)
    if h > 0 and n_z > 1:
        ctc = (h - 2*edge) / max(n_z-1,1)
        max_ctc = max(max_ctc, ctc)
        min_clear = min(min_clear, ctc - phi)
    if max_ctc > 300:
        issues.append("espaçamento longitudinal entre varões superior a 300 mm")
    if min_clear < max(20.0, phi, 25.0):
        issues.append("espaçamento livre entre varões insuficiente")
    if n_y > 4 or n_z > 4:
        issues.append("prever estribos/grampos intermédios para travar varões comprimidos")
    if issues:
        status = "Verificar" if all("prever" in x for x in issues) else "Não conforme"
    return {
        "detailing_status": status,
        "detailing_issues": "; ".join(issues) if issues else "-",
        "detailing_min_bars": min_bars,
        "detailing_smax_ties_mm": smax if smax < 998 else None,
        "detailing_max_long_ctc_mm": max_ctc if max_ctc else None,
        "detailing_min_clear_mm": None if min_clear == 1e9 else min_clear,
    }


# --------------------------- esforço transverso e torção EC2 práticos ---------------------------
def shear_check_ec2_v4(v_ed_kN, n_ed_kN, b_mm, h_mm, d_mm, as_long_mm2, fck, fcd, fyd, gamma_c):
    if b_mm <= 0 or h_mm <= 0 or d_mm <= 0:
        return {"VRdc_kN": None, "VRdmax_kN": None, "Asw_s_req_mm2_per_m": None, "status": "Dados insuficientes"}
    bw = b_mm
    k = min(2.0, 1.0 + math.sqrt(200.0 / max(d_mm, 1e-9)))
    rho_l = min(max(as_long_mm2 / max(bw * d_mm, 1e-9), 0.0), 0.02)
    sigma_cp = min(abs(n_ed_kN) * 1e3 / max(b_mm*h_mm,1e-9), 0.2 * fcd)
    crdc = 0.18 / gamma_c
    vrdc_N = (crdc * k * (100.0 * rho_l * fck) ** (1/3) + 0.15 * sigma_cp) * bw * d_mm
    vrdc_kN = vrdc_N / 1e3
    # VRd,max com cot(theta)=2.5, alpha_cw=1.0, z=0.9d, nu1=0.6(1-fck/250)
    cot = 2.5
    z = 0.9 * d_mm
    nu1 = 0.6 * (1.0 - fck/250.0)
    vrdmax_N = bw * z * nu1 * fcd / (cot + 1.0/cot)
    vrdmax_kN = vrdmax_N / 1e3
    ved = abs(v_ed_kN)
    if ved <= vrdc_kN:
        status = "OK sem armadura transversal resistente adicional"
        asw_s = 0.0
    elif ved <= vrdmax_kN:
        status = "Requer armadura de esforço transverso"
        asw_s = (ved*1e3) / max(z * fyd * cot, 1e-9) * 1000.0  # mm²/m
    else:
        status = "Não conforme: VEd > VRd,max"
        asw_s = (ved*1e3) / max(z * fyd * cot, 1e-9) * 1000.0
    return {"VRdc_kN": vrdc_kN, "VRdmax_kN": vrdmax_kN, "Asw_s_req_mm2_per_m": asw_s, "status": status}


def torsion_check_ec2_v4(t_ed_kNm, b_mm, h_mm, cover_mm, fck, fcd, fyd):
    t_ed = abs(t_ed_kNm) * 1e6  # Nmm
    if b_mm <= 0 or h_mm <= 0:
        return {"TRdmax_kNm": None, "Asw_s_t_req_mm2_per_m": None, "Asl_t_req_mm2": None, "torsion_status": "Dados insuficientes"}
    # Parede equivalente aproximada
    tef = max(2.0*cover_mm, min(b_mm,h_mm)/6.0, 50.0)
    bk = max(b_mm - tef, 1.0)
    hk = max(h_mm - tef, 1.0)
    Ak = bk * hk
    uk = 2.0*(bk+hk)
    cot = 1.0
    nu1 = 0.6*(1.0 - fck/250.0)
    trdmax_Nmm = 2.0 * nu1 * fcd * Ak * tef / (cot + 1.0/cot)
    trdmax_kNm = trdmax_Nmm / 1e6
    if t_ed <= 1e-9:
        status = "Sem torção relevante"
        asw_s = 0.0
        asl = 0.0
    elif t_ed <= trdmax_Nmm:
        status = "Requer armadura de torção"
        asw_s = t_ed / max(2.0*Ak*fyd*cot,1e-9) * 1000.0
        asl = t_ed * uk / max(2.0*Ak*fyd,1e-9)
    else:
        status = "Não conforme: TEd > TRd,max"
        asw_s = t_ed / max(2.0*Ak*fyd*cot,1e-9) * 1000.0
        asl = t_ed * uk / max(2.0*Ak*fyd,1e-9)
    return {"TRdmax_kNm": trdmax_kNm, "Asw_s_t_req_mm2_per_m": asw_s, "Asl_t_req_mm2": asl, "torsion_status": status}


# --------------------------- ELS completo simplificado e combinação seleccionada ---------------------------
def elastic_service_check_v4(n_kN, my_kNm, mz_kNm, b_mm, h_mm, iy_mm4, iz_mm4, as_mm2, fck, fyk, ecm, fctm, exposure="XC3"):
    base = elastic_service_check(n_kN, my_kNm, mz_kNm, b_mm, h_mm, iy_mm4, iz_mm4, as_mm2, fck, fyk, ecm, fctm)
    if not base:
        return {"service_status": "Dados insuficientes"}
    # Critério simplificado de fendilhação: se tracção excede fctm, usar tensão no aço e wk estimado.
    sigma_s = _finite(base.get("service_sigma_s_max_MPa"), 0.0)
    sigma_c_min = _finite(base.get("service_sigma_c_min_MPa"), 0.0)
    wk_lim = 0.3 if exposure.upper() in ["XC2", "XC3", "XC4", "XD1", "XS1"] else 0.4
    wk_est = 0.0
    if sigma_c_min < -fctm:
        wk_est = min(1.5, 0.00085 * sigma_s)  # mm, estimativa conservadora expedita
    crack_status = "OK" if wk_est <= wk_lim else "Verificar fendilhação"
    status = "OK" if base.get("service_status") == "OK" and crack_status == "OK" else "Verificar"
    base.update({
        "service_wk_est_mm": wk_est,
        "service_wk_lim_mm": wk_lim,
        "service_crack_status": crack_status,
        "service_status": status,
        "service_method": "ELS elástico simplificado com fendilhação estimada",
    })
    return base


def apply_service_combination_override_v4(app, results: pd.DataFrame, input_df: pd.DataFrame) -> pd.DataFrame:
    selected = getattr(app, "var_service_case", tk.StringVar(value="")).get().strip() if hasattr(app, "var_service_case") else ""
    if results is None or results.empty:
        return results
    out = results.copy()
    if not selected:
        out["service_case_source"] = "automático/simplificado"
        return out
    pairs = app.df_pair.copy() if getattr(app, "df_pair", None) is not None else pd.DataFrame()
    if pairs.empty:
        out["service_case_source"] = "não encontrado"
        return out
    pairs["_case_str"] = pairs["case"].astype(str)
    pairs["_comb_str"] = pairs["case"].map(extract_combination_number).astype(str)
    target = str(selected)
    pair_sel = pairs[(pairs["_case_str"] == target) | (pairs["_comb_str"] == target)]
    if pair_sel.empty:
        out["service_case_source"] = f"combinação ELS {target} não encontrada"
        out["service_status"] = "ELS não verificado — combinação indicada não encontrada"
        return out
    pair_by_member = {str(r.get("member")): r for _, r in pair_sel.iterrows()}
    for idx, r in out.iterrows():
        member = str(r.get("member"))
        p = pair_by_member.get(member)
        if p is None:
            out.at[idx, "service_case_source"] = f"combinação {target} sem linha para o member"
            out.at[idx, "service_status"] = "ELS não verificado — member sem combinação indicada"
            continue
        material = str(r.get("material", DEFAULT_CONCRETE_CLASS) or DEFAULT_CONCRETE_CLASS)
        fck = parse_concrete_strength(material)
        cp = concrete_props(fck)
        b_mm = _finite(r.get("b_cm"))*10.0
        h_mm = _finite(r.get("h_cm"))*10.0
        iy = _finite(p.get("iy"), 0.0)*10000.0
        iz = _finite(p.get("iz"), 0.0)*10000.0
        if iy <= 0: iy = b_mm*h_mm**3/12.0
        if iz <= 0: iz = h_mm*b_mm**3/12.0
        n = max(abs(_finite(p.get("fx_i"))), abs(_finite(p.get("fx_j"))))
        my = max(abs(_finite(p.get("my_i"))), abs(_finite(p.get("my_j"))))
        mz = max(abs(_finite(p.get("mz_i"))), abs(_finite(p.get("mz_j"))))
        els = elastic_service_check_v4(n, my, mz, b_mm, h_mm, iy, iz, _finite(r.get("as_prov_mm2")), fck, _finite(app.var_fyk.get(),500), cp["Ecm"], cp["fctm"])
        for k,v in els.items():
            out.at[idx, k] = v
        out.at[idx, "service_combination"] = target
        out.at[idx, "service_case_source"] = f"combinação indicada pelo utilizador: {target}"
        out.at[idx, "service_n_kN"] = n
        out.at[idx, "service_my_kNm"] = my
        out.at[idx, "service_mz_kNm"] = mz
    return out


# --------------------------- capacidade/superfície resistente melhorada ---------------------------
_old_capacity_for_layout_v31 = ColumnDesigner.capacity_for_layout

def _capacity_for_layout_v4(self, layout: Layout, n_ed_kN: float, fcd: float, fyd: float, Es: float):
    mode = getattr(self, "requested_calc_mode", self.calc_mode)
    key = ("v4", mode, round(layout.b_mm,1), round(layout.h_mm,1), layout.phi_long_mm, layout.phi_st_mm,
           layout.n_bars_y, layout.n_bars_z, round(n_ed_kN,0), round(fcd,3), round(fyd,3))
    if key in self._capacity_cache:
        return self._capacity_cache[key]
    n_ang = 37 if mode == "rigoroso" else 25
    n_c = 120 if mode == "rigoroso" else 80
    angles = [i * math.pi / (2*(n_ang-1)) for i in range(n_ang)]
    capacities = []
    c_max = 2.4 * max(layout.b_mm, layout.h_mm)
    for ang in angles:
        # procura do melhor c por grelha refinada; estável e sem SciPy
        best = None
        for i in range(n_c):
            c_mm = 2.0 + i * (c_max - 2.0) / max(n_c-1,1)
            N, My, Mz = self.section_response(layout, n_ed_kN, ang, c_mm, fcd, fyd, Es)
            diff = abs(N - n_ed_kN)
            if best is None or diff < best[0]:
                best = (diff, My, Mz, c_mm, ang)
        if best is not None:
            capacities.append((best[1], best[2]))
    self._capacity_cache[key] = capacities
    return capacities

ColumnDesigner.capacity_for_layout = _capacity_for_layout_v4


_old_design_one_v31 = ColumnDesigner.design_one

def _design_one_v4(self, row: pd.Series, prebuilt_candidates=None):
    old_mode = self.calc_mode
    self.requested_calc_mode = old_mode
    if old_mode == "rigoroso":
        self.calc_mode = "dimensionamento"
    try:
        out = _old_design_one_v31(self, row, prebuilt_candidates=prebuilt_candidates)
    finally:
        self.calc_mode = old_mode
    if not isinstance(out, dict):
        return out
    out["cover_mm"] = self.cover_mm
    # Se não há armadura escolhida, apenas devolver com campos de auditoria preenchidos.
    if not out.get("phi_long_mm"):
        out.setdefault("detailing_status", "Não verificado")
        out.setdefault("surface_method", "não gerada")
        return out
    material = str(out.get("material", DEFAULT_CONCRETE_CLASS) or DEFAULT_CONCRETE_CLASS)
    fck = parse_concrete_strength(material)
    cp = concrete_props(fck, gamma_c=self.gamma_c)
    fyd = steel_props(self.fyk, self.gamma_s)["fyd"]
    b_mm = _finite(out.get("b_cm"))*10.0
    h_mm = _finite(out.get("h_cm"))*10.0
    asprov = _finite(out.get("as_prov_mm2"))
    d_y = 0.8*b_mm
    d_z = 0.8*h_mm
    vy = _finite(out.get("vy_ed_kN", out.get("v_ed_y_kN", 0.0)))
    vz = _finite(out.get("vz_ed_kN", out.get("v_ed_z_kN", 0.0)))
    n = _finite(out.get("n_ed_kN"))
    phi = _finite(out.get("phi_long_mm"))
    phi_st = _finite(out.get("phi_st_mm"))
    s_st = _finite(out.get("s_st_mm"))
    sh_y = shear_check_ec2_v4(vy, n, h_mm, b_mm, d_y, asprov, fck, cp["fcd"], fyd, self.gamma_c)
    sh_z = shear_check_ec2_v4(vz, n, b_mm, h_mm, d_z, asprov, fck, cp["fcd"], fyd, self.gamma_c)
    tor = torsion_check_ec2_v4(_finite(out.get("mx_ed_kNm")), b_mm, h_mm, self.cover_mm, fck, cp["fcd"], fyd)
    det = detailing_check_v4(out)
    # Campos normalizados
    out.update({
        "v_rd_c_y_kN": sh_y.get("VRdc_kN"), "v_rd_max_y_kN": sh_y.get("VRdmax_kN"), "asw_s_y_req_mm2_per_m": sh_y.get("Asw_s_req_mm2_per_m"), "shear_status_y": sh_y.get("status"),
        "v_rd_c_z_kN": sh_z.get("VRdc_kN"), "v_rd_max_z_kN": sh_z.get("VRdmax_kN"), "asw_s_z_req_mm2_per_m": sh_z.get("Asw_s_req_mm2_per_m"), "shear_status_z": sh_z.get("status"),
        "shear_status": "OK" if str(sh_y.get("status","")).startswith("OK") and str(sh_z.get("status","")).startswith("OK") else "Verificar",
        "t_rd_max_kNm": tor.get("TRdmax_kNm"), "asw_s_t_req_mm2_per_m": tor.get("Asw_s_t_req_mm2_per_m"), "asl_t_req_mm2": tor.get("Asl_t_req_mm2"), "torsion_status": tor.get("torsion_status"),
        **det,
        "surface_method": "superfície discreta N-My-Mz por rotação da linha neutra" if old_mode == "rigoroso" else "critério biaxial EC2 simplificado + pontos discretos",
        "surface_points": 37 if old_mode == "rigoroso" else 25,
    })
    # ELS v4 para a própria combinação, se não houver override de GUI.
    try:
        iy = (b_mm*h_mm**3/12.0)
        iz = (h_mm*b_mm**3/12.0)
        els = elastic_service_check_v4(n, _finite(out.get("my_ed_kNm")), _finite(out.get("mz_ed_kNm")), b_mm, h_mm, iy, iz, asprov, fck, self.fyk, cp["Ecm"], cp["fctm"])
        for k, v in els.items():
            out.setdefault(k, v)
    except Exception:
        pass
    # Se a pormenorização ou VRdmax/Torsão forem não conformes, classificar sem destruir informação de ELU.
    blockers = []
    if det.get("detailing_status") == "Não conforme":
        blockers.append("pormenorização EC2")
    if "Não conforme" in str(sh_y.get("status")) or "Não conforme" in str(sh_z.get("status")):
        blockers.append("esforço transverso")
    if "Não conforme" in str(tor.get("torsion_status")):
        blockers.append("torção")
    if blockers and out.get("status") == "OK":
        out["status"] = "Falha"
        out["failure_reason"] = "Falha em verificações complementares: " + ", ".join(blockers)
        out["failure_type"] = "pormenorizacao" if "pormenorização EC2" in blockers else "outra"
    # Recomendações adicionais
    recs = [str(out.get("recommendations", "") or "").strip()]
    if out.get("shear_status") == "Verificar": recs.append("dimensionar armadura transversal por esforço transverso")
    if "torção" in str(out.get("torsion_status", "")).lower(): recs.append("verificar armadura longitudinal/transversal de torção")
    if det.get("detailing_status") != "OK": recs.append("rever pormenorização de estribos, espaçamentos e varões intermédios")
    out["recommendations"] = "; ".join([x for x in dict.fromkeys([r for r in recs if r])])
    return out

ColumnDesigner.design_one = _design_one_v4


# --------------------------- DXF ---------------------------
def _dxf_pair(code, value):
    return f"{code}\n{value}\n"

def _dxf_line(x1, y1, x2, y2, layer="0"):
    return "0\nLINE\n" + _dxf_pair(8, layer) + _dxf_pair(10, f"{x1:.3f}") + _dxf_pair(20, f"{y1:.3f}") + _dxf_pair(11, f"{x2:.3f}") + _dxf_pair(21, f"{y2:.3f}")

def _dxf_circle(x, y, r, layer="0"):
    return "0\nCIRCLE\n" + _dxf_pair(8, layer) + _dxf_pair(10, f"{x:.3f}") + _dxf_pair(20, f"{y:.3f}") + _dxf_pair(40, f"{r:.3f}")

def _dxf_text(x, y, text, h=35, layer="TEXT"):
    safe = str(text).replace("\n", " ")[:240]
    return "0\nTEXT\n" + _dxf_pair(8, layer) + _dxf_pair(10, f"{x:.3f}") + _dxf_pair(20, f"{y:.3f}") + _dxf_pair(40, f"{h:.3f}") + _dxf_pair(1, safe)

def _bar_points_for_result(r):
    b = _finite(r.get("b_cm"))*10.0
    h = _finite(r.get("h_cm"))*10.0
    cover = _finite(r.get("cover_mm"), 35.0)
    phi = _finite(r.get("phi_long_mm"), 0.0)
    phi_st = _finite(r.get("phi_st_mm"), 8.0)
    ny = int(_finite(r.get("n_bars_y"), 0))
    nz = int(_finite(r.get("n_bars_z"), 0))
    if b <= 0 or h <= 0 or ny < 2 or nz < 2:
        return []
    edge = cover + phi_st + phi/2.0
    y_left, y_right = -b/2+edge, b/2-edge
    z_bot, z_top = -h/2+edge, h/2-edge
    ys = [y_left + i*(y_right-y_left)/(ny-1) for i in range(ny)]
    zs = [z_bot + i*(z_top-z_bot)/(nz-1) for i in range(nz)]
    pts=[]
    for y in ys:
        pts.append((y,z_top)); pts.append((y,z_bot))
    for z in zs[1:-1]:
        pts.append((y_left,z)); pts.append((y_right,z))
    uniq=[]; seen=set()
    for p in pts:
        k=(round(p[0],3), round(p[1],3))
        if k not in seen:
            seen.add(k); uniq.append(p)
    return uniq

def write_columns_dxf_v4(path: str, df: pd.DataFrame):
    parts = ["0\nSECTION\n2\nHEADER\n0\nENDSEC\n0\nSECTION\n2\nTABLES\n0\nENDSEC\n0\nSECTION\n2\nENTITIES\n"]
    if df is None or df.empty:
        parts.append(_dxf_text(0,0,"Sem resultados",50))
    else:
        x0, y0 = 0.0, 0.0
        dx, dy = 1600.0, -1300.0
        for idx, (_, r) in enumerate(df.head(120).iterrows()):
            col = idx % 4
            row = idx // 4
            ox = x0 + col*dx
            oy = y0 + row*dy
            b = _finite(r.get("b_cm"))*10.0
            h = _finite(r.get("h_cm"))*10.0
            if b <= 0 or h <= 0: continue
            # secção exterior
            left, right = ox-b/2, ox+b/2
            bot, top = oy-h/2, oy+h/2
            parts += [_dxf_line(left,bot,right,bot,"CONCRETE"), _dxf_line(right,bot,right,top,"CONCRETE"), _dxf_line(right,top,left,top,"CONCRETE"), _dxf_line(left,top,left,bot,"CONCRETE")]
            # estribo aproximado
            c = _finite(r.get("cover_mm"),35.0)
            l2,r2,b2,t2 = left+c, right-c, bot+c, top-c
            parts += [_dxf_line(l2,b2,r2,b2,"STIRRUP"), _dxf_line(r2,b2,r2,t2,"STIRRUP"), _dxf_line(r2,t2,l2,t2,"STIRRUP"), _dxf_line(l2,t2,l2,b2,"STIRRUP")]
            phi = _finite(r.get("phi_long_mm"), 10.0)
            for y,z in _bar_points_for_result(r):
                parts.append(_dxf_circle(ox+y, oy+z, max(phi/2.0, 3.0), "REBAR"))
            label = f"P{r.get('member','')} {r.get('b_cm','')}x{r.get('h_cm','')}cm {r.get('solucao','')}"
            parts.append(_dxf_text(ox-b/2, oy-h/2-90, label, 28, "TEXT"))
        # quadro de pilares
        tx, ty = 0, y0 + ((len(df.head(120))//4)+2)*dy
        headers = ["MEMBER", "SECÇÃO", "MATERIAL", "ARMADURA", "ESTRIBOS", "ESTADO"]
        widths = [250, 300, 300, 500, 350, 250]
        rowh = 80
        x = tx
        for htxt,w in zip(headers,widths):
            parts.append(_dxf_text(x+5, ty, htxt, 28, "TEXT")); x+=w
        for i, (_, r) in enumerate(df.head(100).iterrows(), start=1):
            y = ty - i*rowh
            vals = [r.get("member",""), f"{_finite(r.get('b_cm')):.0f}x{_finite(r.get('h_cm')):.0f} cm", r.get("material",""), r.get("solucao",""), f"Ø{r.get('phi_st_mm','')}//{_finite(r.get('s_st_mm'))/10:.1f}cm", r.get("status","")]
            x = tx
            for val,w in zip(vals,widths):
                parts.append(_dxf_text(x+5, y, val, 22, "TEXT")); x += w
    parts.append("0\nENDSEC\n0\nEOF\n")
    Path(path).write_text("".join(parts), encoding="utf-8")


# --------------------------- GUI patches ---------------------------
_old_build_sidebar_v31 = ColumnsEC2App._build_sidebar

def _build_sidebar_v4(self, parent):
    if not hasattr(self, "var_service_case"):
        self.var_service_case = tk.StringVar(value="")
    _old_build_sidebar_v31(self, parent)
    # Acrescentar o modo rigoroso no combobox existente de modo de cálculo.
    def _patch_mode_combo(w):
        try:
            if isinstance(w, ttk.Combobox) and str(w.cget("textvariable")) == str(self.var_calc_mode):
                w.configure(values=["pre_dimensionamento", "dimensionamento", "rigoroso"])
        except Exception:
            pass
        for child in w.winfo_children():
            _patch_mode_combo(child)
    _patch_mode_combo(parent)
    v4 = ttk.LabelFrame(parent, text="7. Verificações avançadas v4")
    v4.pack(fill="x", pady=(0,8))
    ttk.Label(v4, text="Combinação ELS").grid(row=0, column=0, sticky="w", padx=6, pady=4)
    ttk.Entry(v4, textvariable=self.var_service_case).grid(row=0, column=1, sticky="ew", padx=6, pady=4)
    ttk.Label(v4, text="em branco = ELS simplificado por defeito", style="Subtle.TLabel").grid(row=1, column=0, columnspan=2, sticky="w", padx=6, pady=(0,4))
    ttk.Button(v4, text="Exportar quadro .DXF", command=self.export_dxf).grid(row=2, column=0, columnspan=2, sticky="ew", padx=4, pady=4)
    v4.columnconfigure(1, weight=1)

ColumnsEC2App._build_sidebar = _build_sidebar_v4

# Patch de modo rigoroso no combobox por substituição visual pós-criação é difícil; validamos no cálculo e aceitamos valor se alterado por código.
_old_validate_inputs_v31 = ColumnsEC2App.validate_inputs

def _validate_inputs_v4(self):
    err = _old_validate_inputs_v31(self)
    if err:
        return err
    if getattr(self, "var_service_case", None) is not None:
        sc = self.var_service_case.get().strip()
        if sc and self.df_pair is not None and not self.df_pair.empty:
            cases = set(self.df_pair.get("case", pd.Series(dtype=str)).astype(str)) | set(self.df_pair.get("case", pd.Series(dtype=str)).map(extract_combination_number).astype(str))
            if sc not in cases:
                # Aviso não bloqueante
                self.status_var.set(f"Aviso: combinação ELS {sc} não encontrada; será assinalado no relatório.")
    return None
ColumnsEC2App.validate_inputs = _validate_inputs_v4

ColumnsEC2App.build_data_validation = _validation_rows_v4

_old_run_design_v31 = ColumnsEC2App.run_design

def _run_design_v4(self):
    # Reimplementa o run_design v3.1 para aplicar ELS por combinação escolhida antes do resumo.
    err = self.validate_inputs()
    if err:
        messagebox.showwarning("Aviso", err)
        return
    designer = ColumnDesigner(
        cover_mm=safe_float(self.var_cover.get(), 35.0),
        fyk=safe_float(self.var_fyk.get(), 500.0),
        phi_eff=safe_float(self.var_phi_eff.get(), 2.0),
        l0y_factor=safe_float(self.var_l0y.get(), 1.0),
        l0z_factor=safe_float(self.var_l0z.get(), 1.0),
        calc_mode=self.var_calc_mode.get(),
    )
    designer.service_case_override = self.var_service_case.get().strip() if hasattr(self, "var_service_case") else ""
    input_df = reduce_to_governing_cases(self.df_pair) if self.var_reduce_cases.get() else self.df_pair.copy()
    self.df_calc_input = input_df.copy()
    self.progress_var.set(0.0)
    self.status_var.set("Análise v4 em curso...")
    def progress(done, total):
        pct = 0.0 if total <= 0 else 100.0 * done / total
        self.after(0, lambda p=pct: self.progress_var.set(p))
        self.after(0, lambda d=done, t=total: self.status_var.set(f"A calcular... {d}/{t} casos member/case"))
    def worker():
        try:
            results = designer.design_dataframe(input_df, progress_callback=progress)
            results = apply_service_combination_override_v4(self, results, input_df)
            summary = self.build_summary_by_member(results) if self.var_summary.get() else pd.DataFrame()
            failures = results[results["status"] == "Falha"].copy() if "status" in results.columns else pd.DataFrame()
            ok = results[results["status"] == "OK"].copy() if "status" in results.columns else pd.DataFrame()
            validation = self.build_data_validation(pre_calc=False)
            def finish():
                self.df_results = results
                self.df_summary = summary
                self.df_failures = failures
                self.df_ok = ok
                self.df_filtered = pd.DataFrame()
                self.df_validation = validation
                self.df_notes = self.build_normative_notes()
                self.show_df(self.tree_results, self.df_results)
                self.show_df(self.tree_summary, self.df_summary)
                self.show_df(self.tree_failures, self.df_failures)
                self.show_df(self.tree_shortlists, self.build_shortlists_df())
                self.show_df(self.tree_validation, self.df_validation)
                self.show_df(self.tree_notes, self.df_notes)
                self.update_report()
                self.progress_var.set(100.0)
                self.status_var.set(f"Cálculo v4 concluído: {len(results)} casos; {len(summary)} membros resumidos; {len(failures)} falhas.")
            self.after(0, finish)
        except Exception as err:
            msg = str(err)
            self.after(0, lambda m=msg: messagebox.showerror("Erro", m))
            self.after(0, lambda: self.status_var.set("Falha na análise."))
            self.after(0, lambda: self.progress_var.set(0.0))
    self.analysis_thread = threading.Thread(target=worker, daemon=True)
    self.analysis_thread.start()

ColumnsEC2App.run_design = _run_design_v4


def _export_dxf_v4(self):
    src = self.df_summary if self.df_summary is not None and not self.df_summary.empty else self.df_results
    if src is None or src.empty:
        messagebox.showwarning("Aviso", "Não há resultados para exportar em DXF.")
        return
    path = filedialog.asksaveasfilename(title="Exportar secções armadas e quadro de pilares", defaultextension=".dxf", filetypes=[("DXF", "*.dxf")])
    if not path:
        return
    try:
        write_columns_dxf_v4(path, src)
        self.status_var.set(f"DXF exportado para: {path}")
    except Exception as err:
        messagebox.showerror("Erro", f"Não foi possível exportar DXF.\n\n{err}")

ColumnsEC2App.export_dxf = _export_dxf_v4

# --------------------------- XLSX/PDF v4 ---------------------------
def _parameters_df_v4(self) -> pd.DataFrame:
    return pd.DataFrame([
        ["Recobrimento [mm]", self.var_cover.get()],
        ["Aço fyk [MPa]", self.var_fyk.get()],
        ["Betão", "lido da coluna Material; fallback interno C30/37 quando ausente"],
        ["φef", self.var_phi_eff.get()],
        ["l0y/L", self.var_l0y.get()],
        ["l0z/L", self.var_l0z.get()],
        ["Modo", self.var_calc_mode.get()],
        ["Combinação ELS indicada", getattr(self, "var_service_case", tk.StringVar(value="")).get().strip() or "não indicada — verificação simplificada"],
        ["Superfície resistente", "discreta N-My-Mz; malha refinada no modo rigoroso"],
        ["Esforço transverso", "VRd,c, VRd,max e Asw/s requerido"],
        ["Torção", "TRd,max, Asw/s e Asl requeridos por modelo tubular simplificado"],
        ["Pormenorização", "mínimos de varões, diâmetros, espaçamentos, travamento e congestionamento"],
    ], columns=["Parâmetro", "Valor"])
ColumnsEC2App._parameters_df = _parameters_df_v4


def _metadata_df_v4(self) -> pd.DataFrame:
    return pd.DataFrame([
        ["Programa", APP_NAME], ["Versão", APP_VERSION], ["Autor", APP_AUTHOR], ["Repositório", GITHUB_URL],
        ["Data de exportação", datetime.now().strftime("%Y-%m-%d %H:%M")], ["Ficheiro de origem", self.input_file_path or "-"],
        ["Norma de referência", "Eurocódigo 2 / NP EN 1992-1-1"],
        ["Âmbito", "Dimensionamento/verificação de pilares de betão armado: ELU, ELS, V, T, pormenorização e DXF"],
        ["Descrição", APP_XLSX_DESCRIPTION],
        ["Limitações", "Ferramenta de apoio ao projecto; resultados críticos devem ser revistos por engenheiro responsável."],
    ], columns=["Campo", "Valor"])
ColumnsEC2App._metadata_df = _metadata_df_v4


def _write_excel_v4(self, path: str):
    res = self.df_results if self.df_results is not None else pd.DataFrame()
    sheets = {
        "00_Info": self._metadata_df(), "01_Parametros": self._parameters_df(), "02_Entrada_Dados": self.df_clean,
        "03_Pares_Member_Case": self.df_pair, "04_Casos_Calculo": self.df_calc_input, "05_Resultados": res,
        "06_Resumo_Membros": self.df_summary, "07_Falhas": self.df_failures, "08_OK": self.df_ok,
        "09_Shortlists": self.build_shortlists_df(), "10_Validacao": self.df_validation, "11_Notas_EC2": self.df_notes,
    }
    # Folhas técnicas filtradas
    tech_cols = {
        "12_ELS": [c for c in res.columns if c.startswith("service") or c in ["member","case","name","status"]],
        "13_Esf_Transverso": [c for c in res.columns if "shear" in c or "v_rd" in c or "asw_s_y" in c or "asw_s_z" in c or c in ["member","case","v_ed_y_kN","v_ed_z_kN"]],
        "14_Torcao": [c for c in res.columns if "torsion" in c or "t_rd" in c or "asl_t" in c or "asw_s_t" in c or c in ["member","case","mx_ed_kNm"]],
        "15_Pormenorizacao": [c for c in res.columns if "detailing" in c or c in ["member","case","solucao","phi_long_mm","phi_st_mm","s_st_mm","n_total","n_bars_y","n_bars_z"]],
        "16_Superficie": [c for c in res.columns if "surface" in c or c in ["member","case","mrd_y_kNm","mrd_z_kNm","utilizacao"]],
        "17_Memoria_Calculo": [c for c in ["member","case","combination_number","limit_state","material","n_ed_kN","my_i_kNm","my_j_kNm","mz_i_kNm","mz_j_kNm","rm_y","rm_z","lambda_y","lambda_lim_y","lambda_z","lambda_lim_z","m_imp_y_kNm","m_imp_z_kNm","m0e_y_kNm","m0e_z_kNm","m2_y_kNm","m2_z_kNm","my_ed_kNm","mz_ed_kNm","as_min_mm2","as_req_mm2","as_prov_mm2","mrd_y_kNm","mrd_z_kNm","utilizacao","status","failure_reason","recommendations"] if c in res.columns],
    }
    for name, cols in tech_cols.items():
        sheets[name] = res[cols].copy() if cols else pd.DataFrame()
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in sheets.items():
            if df is None: df = pd.DataFrame()
            df.to_excel(writer, sheet_name=name[:31], index=False)
        wb = writer.book
        props = wb.properties
        props.title = f"{APP_NAME} {APP_VERSION}"; props.subject = APP_SUBJECT; props.creator = APP_AUTHOR
        props.keywords = APP_KEYWORDS; props.category = APP_CATEGORY; props.description = APP_XLSX_DESCRIPTION; props.lastModifiedBy = APP_AUTHOR
        try:
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            from openpyxl.utils import get_column_letter
            header_fill = PatternFill("solid", fgColor="1F4E5F"); header_font = Font(color="FFFFFF", bold=True)
            thin = Side(style="thin", color="D9E2E7"); border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for ws in wb.worksheets:
                ws.sheet_view.showGridLines = False; ws.freeze_panes = "A2"
                if ws.max_row >= 1:
                    for cell in ws[1]:
                        cell.fill = header_fill; cell.font = header_font; cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); cell.border = border
                for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 5000)):
                    for cell in row:
                        cell.border = border; cell.alignment = Alignment(vertical="top", wrap_text=True)
                for col_idx, col in enumerate(ws.columns, start=1):
                    values=[str(c.value) for c in col[:200] if c.value is not None]
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(max([len(v) for v in values]+[10])+2, 46)
        except Exception:
            pass
ColumnsEC2App._write_excel = _write_excel_v4


def _pdf_table_style_courier(header=True):
    from reportlab.lib import colors
    from reportlab.platypus import TableStyle
    cmds=[("GRID",(0,0),(-1,-1),0.25,colors.HexColor("#D9E2E7")),("VALIGN",(0,0),(-1,-1),"TOP"),("FONTNAME",(0,0),(-1,-1),"Courier"),("FONTSIZE",(0,0),(-1,-1),7),("LEFTPADDING",(0,0),(-1,-1),3),("RIGHTPADDING",(0,0),(-1,-1),3)]
    if header:
        cmds += [("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1F4E5F")),("TEXTCOLOR",(0,0),(-1,0),colors.white),("FONTNAME",(0,0),(-1,0),"Courier-Bold")]
    return TableStyle(cmds)


def _write_pdf_v4(self, path: str):
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, PageBreak
    doc = SimpleDocTemplate(path, pagesize=landscape(A4), rightMargin=12*mm, leftMargin=12*mm, topMargin=12*mm, bottomMargin=12*mm)
    doc.title=f"{APP_NAME} {APP_VERSION}"; doc.author=APP_AUTHOR; doc.subject=APP_SUBJECT
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="ReportTitle", parent=styles["Title"], alignment=TA_CENTER, fontName="Courier-Bold", fontSize=14, leading=21, spaceAfter=10))
    styles.add(ParagraphStyle(name="ReportSubtitle", parent=styles["Normal"], alignment=TA_CENTER, fontName="Courier", fontSize=10, leading=15, textColor=colors.darkgrey, spaceAfter=8))
    styles.add(ParagraphStyle(name="BodyCourier", parent=styles["Normal"], fontName="Courier", fontSize=10, leading=15, spaceAfter=6))
    styles.add(ParagraphStyle(name="Small", parent=styles["Normal"], fontName="Courier", fontSize=8, leading=12))
    styles.add(ParagraphStyle(name="Cell", parent=styles["Small"], alignment=TA_LEFT, fontName="Courier", fontSize=7, leading=10.5))
    styles.add(ParagraphStyle(name="Section", parent=styles["Heading2"], fontName="Courier-Bold", fontSize=12, leading=18, spaceBefore=10, spaceAfter=20))
    story=[]
    story.append(Paragraph(f"{APP_NAME} {APP_VERSION}", styles["ReportTitle"]))
    story.append(Paragraph("Memória de cálculo de pilares de betão armado segundo o Eurocódigo 2", styles["ReportSubtitle"]))
    n_total=len(self.df_results); n_ok=int((self.df_results.get("status",pd.Series(dtype=str))=="OK").sum()); n_fail=int((self.df_results.get("status",pd.Series(dtype=str))=="Falha").sum())
    meta=[["Programa", f"{APP_NAME} {APP_VERSION}", "Autor", APP_AUTHOR], ["Data", datetime.now().strftime("%Y-%m-%d %H:%M"), "Casos", str(n_total)], ["OK",str(n_ok),"Falhas",str(n_fail)], ["ELS indicado", getattr(self,"var_service_case",tk.StringVar(value="")).get().strip() or "simplificado", "Ficheiro", os.path.basename(self.input_file_path or "-")]]
    t=Table(meta, colWidths=[38*mm,90*mm,38*mm,90*mm]); t.setStyle(_pdf_table_style_courier(header=False)); story.append(t); story.append(Spacer(1,5*mm))
    def df_table(df, cols, max_rows=25):
        present=[c for c in cols if c in df.columns]
        data=[[Paragraph(str(c), styles["Cell"]) for c in present]]
        for _,r in df.head(max_rows).iterrows():
            row=[]
            for c in present:
                v=r.get(c,""); txt="" if pd.isna(v) else (f"{v:.2f}" if isinstance(v,float) and math.isfinite(v) else str(v))
                txt=txt.replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")
                row.append(Paragraph(txt, styles["Cell"]))
            data.append(row)
        widths=[270*mm/max(1,len(present))]*max(1,len(present))
        tb=Table(data, colWidths=widths, repeatRows=1); tb.setStyle(_pdf_table_style_courier(header=True)); return tb
    story.append(Paragraph("1. Resumo executivo", styles["Section"]))
    summary = self.df_summary if self.df_summary is not None and not self.df_summary.empty else self.df_results
    story.append(df_table(summary, ["member","case","material","n_ed_kN","my_ed_kNm","mz_ed_kNm","as_req_mm2","as_prov_mm2","solucao","status"], 30))
    story.append(Paragraph("2. Memória de cálculo — pilares críticos", styles["Section"]))
    mem_cols=["member","case","my_i_kNm","my_j_kNm","mz_i_kNm","mz_j_kNm","rm_y","rm_z","lambda_y","lambda_lim_y","lambda_z","lambda_lim_z","m_imp_y_kNm","m2_y_kNm","m_imp_z_kNm","m2_z_kNm","mrd_y_kNm","mrd_z_kNm","utilizacao"]
    story.append(df_table(summary, mem_cols, 25))
    story.append(PageBreak())
    story.append(Paragraph("3. Verificações complementares — ELS, V, T e pormenorização", styles["Section"]))
    comp_cols=["member","case","service_combination","service_status","service_wk_est_mm","shear_status","torsion_status","detailing_status","detailing_issues","recommendations"]
    story.append(df_table(summary, comp_cols, 30))
    if self.df_failures is not None and not self.df_failures.empty:
        story.append(PageBreak()); story.append(Paragraph("4. Falhas e recomendações", styles["Section"]))
        story.append(df_table(self.df_failures, ["member","case","failure_type","failure_reason","recommendations","shortlist_text"], 35))
    story.append(Spacer(1,4*mm)); story.append(Paragraph("Notas: relatório gerado automaticamente. Validar resultados críticos, hipóteses de comprimento efectivo, combinações ELS e pormenorização final em projecto.", styles["Small"]))
    def footer(canvas, doc_obj):
        canvas.saveState(); canvas.setAuthor(APP_AUTHOR); canvas.setTitle(f"{APP_NAME} {APP_VERSION}"); canvas.setSubject(APP_SUBJECT); canvas.setFont("Courier",7); canvas.setFillColor(colors.grey); canvas.drawString(12*mm,7*mm,f"{APP_NAME} {APP_VERSION} | {APP_AUTHOR}"); canvas.drawRightString(285*mm,7*mm,f"Página {doc_obj.page}"); canvas.restoreState()
    doc.build(story, onFirstPage=footer, onLaterPages=footer)
ColumnsEC2App._write_pdf = _write_pdf_v4


def _build_normative_notes_v4(self) -> pd.DataFrame:
    notes = [
        ("Entrada", "Excel/tabela", "Validação robusta de cabeçalhos, pares de nós, materiais, geometria, unidades suspeitas e combinações."),
        ("Betão", "Material", "Classe de betão lida exclusivamente da tabela; C30/37 é apenas fallback interno quando ausente."),
        ("2.ª ordem", "EC2 5.8", "Curvatura nominal, imperfeições geométricas práticas, λlim e razão rm com sinal."),
        ("Biaxial", "EC2 5.8.9", "Modo dimensionamento com critério biaxial; modo rigoroso com malha N-My-Mz mais densa."),
        ("Esforço transverso", "EC2 6.2", "VRd,c, VRd,max e Asw/s requerido em Y e Z."),
        ("Torção", "EC2 6.3", "Modelo tubular simplificado com TRd,max, Asw/s e Asl requeridos."),
        ("ELS", "EC2 7", "Se indicada combinação ELS, usa essa combinação por member; caso contrário usa verificação simplificada da própria combinação."),
        ("Pormenorização", "EC2 8/9", "Mínimos de varões, estribos, espaçamentos, varões intermédios e congestionamento."),
        ("DXF", "Quadro de pilares", "Exporta secções armadas e quadro de pilares em DXF ASCII."),
    ]
    return pd.DataFrame(notes, columns=["Tema", "Referência", "Nota"])
ColumnsEC2App.build_normative_notes = _build_normative_notes_v4




# ============================================================
# Overrides v4.1 — auditabilidade, validação, DXF e exportações
# ============================================================
APP_VERSION = "v4.5"
APP_XLSX_DESCRIPTION = (
    "Workbook de cálculo com dados de entrada, validação robusta, qualidade da importação, "
    "pares member/case, ELU, ELS, esforço transverso, torção, pormenorização, "
    "superfície resistente aproximada, DXF e memória de cálculo."
)


def _case_key_value(v):
    s = str(v or "").strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s


def material_is_valid_ec2_class(value) -> bool:
    return bool(re.search(r"C\s*\d+\s*/\s*\d+", str(value or ""), re.I))


def build_import_quality_cases_v41(app) -> pd.DataFrame:
    """Relatório linha-a-linha por member/case para evitar erros silenciosos da importação de dados."""
    df = getattr(app, "df_clean", pd.DataFrame())
    pair = getattr(app, "df_pair", pd.DataFrame())
    rows = []
    if df is None or df.empty:
        return pd.DataFrame([{
            "member": "", "case": "", "name": "", "estado": "Não conforme",
            "n_nodes_found": 0, "node_i": "", "node_j": "", "avisos": "sem dados importados"
        }])
    if pair is None or pair.empty:
        return pd.DataFrame([{
            "member": "", "case": "", "name": "", "estado": "Não conforme",
            "n_nodes_found": 0, "node_i": "", "node_j": "", "avisos": "não foi possível formar pares member/case"
        }])
    for _, p in pair.iterrows():
        member = str(p.get("member", ""))
        case = str(p.get("case", ""))
        name = str(p.get("name", ""))
        grp = df[(df.get("member", pd.Series(dtype=str)).astype(str) == member) &
                 (df.get("case", pd.Series(dtype=str)).astype(str) == case) &
                 (df.get("name", pd.Series(dtype=str)).astype(str) == name)]
        avisos = []
        estado = "OK"
        n_nodes = int(_finite(p.get("n_nodes_found"), 0))
        if n_nodes != 2:
            estado = "Não conforme" if n_nodes < 2 else "Verificar"
            avisos.append(f"member/case com {n_nodes} nós; esperado 2")
        # consistência geométrica/material por grupo
        for c in ["length", "material", "hy", "hz", "ax", "iy", "iz"]:
            if c in grp.columns and len(grp) >= 2:
                vals = grp[c].dropna().astype(str).str.strip().replace("nan", "").replace("None", "")
                vals = [v for v in vals.tolist() if v != ""]
                if len(vals) >= 2:
                    if c == "material":
                        if len(set(vals)) > 1:
                            estado = "Verificar" if estado == "OK" else estado
                            avisos.append(f"{c} diferente entre nós")
                    else:
                        nums = [safe_float(v, float('nan')) for v in vals]
                        nums = [v for v in nums if math.isfinite(v)]
                        if len(nums) >= 2 and max(nums) - min(nums) > 1e-6:
                            estado = "Verificar" if estado == "OK" else estado
                            avisos.append(f"{c} diferente entre nós")
        mat = str(p.get("material", ""))
        material_assumed = "Não"
        if not material_is_valid_ec2_class(mat):
            material_assumed = "Sim"
            estado = "Verificar" if estado == "OK" else estado
            avisos.append("classe de betão ausente/inválida; fallback interno C30/37")
        # unidades suspeitas
        hy, hz = _finite(p.get("hy"), 0.0), _finite(p.get("hz"), 0.0)
        length = _finite(p.get("length"), 0.0)
        if hy <= 0 or hz <= 0:
            estado = "Não conforme"
            avisos.append("dimensões HY/HZ inválidas")
        elif hy < 10 or hz < 10 or hy > 300 or hz > 300:
            estado = "Verificar" if estado == "OK" else estado
            avisos.append("dimensões HY/HZ suspeitas; esperado em cm")
        if length <= 0 or length > 20:
            estado = "Verificar" if estado == "OK" else estado
            avisos.append("Length suspeito; esperado em m")
        rows.append({
            "member": member,
            "case": case,
            "combination_number": extract_combination_number(case),
            "name": name,
            "estado": estado,
            "n_nodes_found": n_nodes,
            "node_i": p.get("node_i", ""),
            "node_j": p.get("node_j", ""),
            "material": mat,
            "material_assumido": material_assumed,
            "hy_cm": hy,
            "hz_cm": hz,
            "length_m": length,
            "avisos": "; ".join(dict.fromkeys(avisos)) if avisos else "-",
        })
    return pd.DataFrame(rows)


def material_assumptions_df_v41(results: pd.DataFrame) -> pd.DataFrame:
    if results is None or results.empty:
        return pd.DataFrame()
    cols = ["member", "case", "name", "material", "material_assumed", "material_source", "status"]
    present = [c for c in cols if c in results.columns]
    df = results[present].copy() if present else pd.DataFrame()
    if "material_assumed" in df.columns:
        df = df[df["material_assumed"].astype(str).str.lower().isin(["sim", "true", "1"])]
    return df


def surface_points_df_v41(results: pd.DataFrame) -> pd.DataFrame:
    """Exporta pontos representativos da superfície resistente usada no relatório.
    Nota: quando a malha interna completa não fica armazenada, gera-se uma envolvente radial
    coerente com MRd,y/MRd,z e o método indicado, para auditoria e plotagem.
    """
    if results is None or results.empty:
        return pd.DataFrame()
    rows = []
    for _, r in results.iterrows():
        mry = _finite(r.get("mrd_y_kNm"), 0.0)
        mrz = _finite(r.get("mrd_z_kNm"), 0.0)
        if mry <= 0 or mrz <= 0:
            continue
        n = _finite(r.get("n_ed_kN"), 0.0)
        pts = int(_finite(r.get("surface_points"), 25))
        pts = max(9, min(73, pts))
        for i in range(pts):
            th = (math.pi / 2.0) * i / max(pts - 1, 1)
            # elipse de resistência aproximada no plano My-Mz para NEd fixo
            my = mry * math.cos(th)
            mz = mrz * math.sin(th)
            rows.append({
                "member": r.get("member", ""),
                "case": r.get("case", ""),
                "combination_number": r.get("combination_number", ""),
                "N_kN": n,
                "theta_deg": math.degrees(th),
                "MRd_y_kNm": my,
                "MRd_z_kNm": mz,
                "surface_method": r.get("surface_method", ""),
            })
    return pd.DataFrame(rows)


def validation_test_cases_df_v41() -> pd.DataFrame:
    rows = [
        ("TC01", "Compressão centrada", "Pilar curto com My=Mz=0", "NRd aproximado e As,min", "pendente de validação externa"),
        ("TC02", "Flexão uniaxial", "NEd+My, Mz=0", "MRd,y e As escolhida", "pendente de validação externa"),
        ("TC03", "Flexão biaxial", "NEd+My+Mz", "interação biaxial", "pendente de validação externa"),
        ("TC04", "2.ª ordem", "pilar esbelto, dois momentos de extremidade", "λ, λlim, M2", "pendente de validação externa"),
        ("TC05", "Esforço transverso", "FY/FZ elevados", "VRd,c, VRd,max, Asw/s", "pendente de validação externa"),
        ("TC06", "Torção", "MX significativo", "TRd,max e armaduras de torção", "pendente de validação externa"),
        ("TC07", "ELS", "combinação ELS indicada na GUI", "tensões e fendilhação", "pendente de validação externa"),
    ]
    return pd.DataFrame(rows, columns=["ID", "Caso-tipo", "Descrição", "Resultado a validar", "Estado"])


# --------------------------- Design wrapper v4.1 ---------------------------
_old_design_one_v4_bound = ColumnDesigner.design_one

def _design_one_v41(self, row: pd.Series, prebuilt_candidates=None):
    out = _old_design_one_v4_bound(self, row, prebuilt_candidates=prebuilt_candidates)
    if not isinstance(out, dict):
        return out
    raw_mat = str(row.get("material", "") or "").strip()
    valid_mat = material_is_valid_ec2_class(raw_mat)
    out["material_source"] = "tabela" if valid_mat else "fallback interno"
    out["material_assumed"] = "Não" if valid_mat else "Sim"
    if not valid_mat:
        out["material_original"] = raw_mat
        out["material"] = DEFAULT_CONCRETE_CLASS
        note = "classe de betão ausente/inválida na tabela; adoptado C30/37"
        existing = str(out.get("recommendations", "") or "").strip()
        out["recommendations"] = "; ".join([x for x in [existing, note] if x])
    # memória de cálculo adicional
    try:
        out["moment_sign_y"] = "mesmo sinal" if _finite(row.get("my_i"),0)*_finite(row.get("my_j"),0) >= 0 else "sinais opostos"
        out["moment_sign_z"] = "mesmo sinal" if _finite(row.get("mz_i"),0)*_finite(row.get("mz_j"),0) >= 0 else "sinais opostos"
        out["curvature_y"] = "curvatura simples" if out["moment_sign_y"] == "mesmo sinal" else "curvatura dupla"
        out["curvature_z"] = "curvatura simples" if out["moment_sign_z"] == "mesmo sinal" else "curvatura dupla"
        out["NEd_My_Mz_point"] = f"({out.get('n_ed_kN', '')}, {out.get('my_ed_kNm', '')}, {out.get('mz_ed_kNm', '')})"
        out["audit_note"] = "verificar hipóteses de l0, sinais locais da análise estrutural e combinação governante"
    except Exception:
        pass
    return out

ColumnDesigner.design_one = _design_one_v41


# --------------------------- Validação e load/run patches v4.1 ---------------------------
_old_build_main_tabs_v4 = ColumnsEC2App._build_main_tabs

def _build_main_tabs_v41(self, parent):
    _old_build_main_tabs_v4(self, parent)
    try:
        # procurar o Notebook e acrescentar separador de qualidade de importação
        nb = None
        for child in parent.winfo_children():
            if isinstance(child, ttk.Notebook):
                nb = child
                break
        if nb is not None and not hasattr(self, "tab_quality"):
            self.tab_quality = ttk.Frame(nb)
            nb.add(self.tab_quality, text="Qualidade Importação")
            self.tree_quality = self._make_tree(self.tab_quality)
    except Exception:
        pass

ColumnsEC2App._build_main_tabs = _build_main_tabs_v41

_old_load_df_v4 = ColumnsEC2App.load_df

def _load_df_v41(self, df: pd.DataFrame, source: str):
    _old_load_df_v4(self, df, source)
    try:
        self.df_import_quality = build_import_quality_cases_v41(self)
        if hasattr(self, "tree_quality"):
            self.show_df(self.tree_quality, self.df_import_quality)
    except Exception as err:
        self.df_import_quality = pd.DataFrame([{"estado":"Erro", "avisos": str(err)}])

ColumnsEC2App.load_df = _load_df_v41

_old_run_design_v4 = ColumnsEC2App.run_design

def _run_design_v41(self):
    # versão explícita para actualizar qualidade, materiais assumidos e notas após cálculo
    err = self.validate_inputs()
    if err:
        messagebox.showwarning("Aviso", err)
        return
    designer = ColumnDesigner(
        cover_mm=safe_float(self.var_cover.get(), 35.0),
        fyk=safe_float(self.var_fyk.get(), 500.0),
        phi_eff=safe_float(self.var_phi_eff.get(), 2.0),
        l0y_factor=safe_float(self.var_l0y.get(), 1.0),
        l0z_factor=safe_float(self.var_l0z.get(), 1.0),
        calc_mode=self.var_calc_mode.get(),
    )
    input_df = reduce_to_governing_cases(self.df_pair) if self.var_reduce_cases.get() else self.df_pair.copy()
    self.df_calc_input = input_df.copy()
    self.progress_var.set(0.0)
    self.status_var.set("Análise em curso...")

    def progress(done, total):
        pct = 0.0 if total <= 0 else 100.0 * done / total
        self.after(0, lambda p=pct: self.progress_var.set(p))
        self.after(0, lambda d=done, t=total: self.status_var.set(f"A calcular... {d}/{t} casos member/case"))

    def worker():
        try:
            results = designer.design_dataframe(input_df, progress_callback=progress)
            results = apply_service_combination_override_v4(self, results, input_df)
            summary = self.build_summary_by_member(results) if self.var_summary.get() else pd.DataFrame()
            failures = results[results["status"] == "Falha"].copy() if "status" in results.columns else pd.DataFrame()
            ok = results[results["status"] == "OK"].copy() if "status" in results.columns else pd.DataFrame()
            def finish():
                self.df_results = results
                self.df_summary = summary
                self.df_failures = failures
                self.df_ok = ok
                self.df_filtered = pd.DataFrame()
                self.df_import_quality = build_import_quality_cases_v41(self)
                self.df_validation = self.build_data_validation(pre_calc=False)
                self.df_notes = self.build_normative_notes()
                self.show_df(self.tree_results, self.df_results)
                self.show_df(self.tree_summary, self.df_summary)
                self.show_df(self.tree_failures, self.df_failures)
                self.show_df(self.tree_shortlists, self.build_shortlists_df())
                self.show_df(self.tree_validation, self.df_validation)
                self.show_df(self.tree_notes, self.df_notes)
                if hasattr(self, "tree_quality"):
                    self.show_df(self.tree_quality, self.df_import_quality)
                self.update_report()
                self.progress_var.set(100.0)
                n_assumed = int((results.get("material_assumed", pd.Series(dtype=str)).astype(str) == "Sim").sum()) if not results.empty else 0
                mat_note = f"; {n_assumed} materiais assumidos" if n_assumed else ""
                self.status_var.set(f"Cálculo concluído: {len(results)} casos; {len(summary)} membros; {len(failures)} falhas{mat_note}.")
            self.after(0, finish)
        except Exception as err:
            msg = str(err)
            self.after(0, lambda m=msg: messagebox.showerror("Erro", m))
            self.after(0, lambda: self.status_var.set("Falha na análise."))
            self.after(0, lambda: self.progress_var.set(0.0))
    self.analysis_thread = threading.Thread(target=worker, daemon=True)
    self.analysis_thread.start()

ColumnsEC2App.run_design = _run_design_v41


# --------------------------- DXF v4.1 ---------------------------
def _dxf_layer_table_v41():
    layers = [("CONCRETE", 7), ("REBARS", 1), ("STIRRUPS", 3), ("TEXT", 2), ("TABLE", 8)]
    s = "0\nSECTION\n2\nTABLES\n0\nTABLE\n2\nLAYER\n70\n5\n"
    for name, color in layers:
        s += "0\nLAYER\n2\n" + name + "\n70\n0\n62\n" + str(color) + "\n6\nCONTINUOUS\n"
    s += "0\nENDTAB\n0\nENDSEC\n"
    return s


def write_columns_dxf_v41(path: str, df: pd.DataFrame):
    parts = ["0\nSECTION\n2\nHEADER\n0\nENDSEC\n", _dxf_layer_table_v41(), "0\nSECTION\n2\nENTITIES\n"]
    if df is None or df.empty:
        parts.append(_dxf_text(0, 0, "Sem resultados", 50, "TEXT"))
    else:
        # quadro geral
        x_table, y_table = 0.0, 250.0
        headers = ["MEMBER", "SECÇÃO", "BETÃO", "ARM. LONG.", "ESTRIBOS", "ESTADO"]
        widths = [260, 280, 240, 420, 420, 260]
        x = x_table
        for h, w in zip(headers, widths):
            parts.append(_dxf_text(x + 10, y_table, h, 28, "TABLE"))
            parts.append(_dxf_line(x, y_table - 15, x + w, y_table - 15, "TABLE"))
            x += w
        for i, (_, r) in enumerate(df.head(80).iterrows(), start=1):
            y = y_table - i * 70
            vals = [
                str(r.get("member", "")),
                f"{_finite(r.get('b_cm')):.0f}x{_finite(r.get('h_cm')):.0f} cm",
                str(r.get("material", "")),
                str(r.get("solucao", "")).split("+")[0].strip(),
                f"Ø{int(_finite(r.get('phi_st_mm'),0))}//{_finite(r.get('s_st_mm'),0)/10:.1f} cm" if _finite(r.get('phi_st_mm'),0) else "",
                str(r.get("status", "")),
            ]
            x = x_table
            for val, w in zip(vals, widths):
                parts.append(_dxf_text(x + 10, y, val, 22, "TABLE"))
                x += w
        # secções armadas
        x0, y0 = 0.0, -900.0
        dx, dy = 1700.0, -1450.0
        for idx, (_, r) in enumerate(df.head(120).iterrows()):
            col = idx % 4
            row = idx // 4
            ox = x0 + col * dx
            oy = y0 + row * dy
            b = _finite(r.get("b_cm")) * 10.0
            h = _finite(r.get("h_cm")) * 10.0
            if b <= 0 or h <= 0:
                continue
            left, right = ox - b/2, ox + b/2
            bot, top = oy - h/2, oy + h/2
            parts += [_dxf_line(left,bot,right,bot,"CONCRETE"), _dxf_line(right,bot,right,top,"CONCRETE"), _dxf_line(right,top,left,top,"CONCRETE"), _dxf_line(left,top,left,bot,"CONCRETE")]
            cover = _finite(r.get("cover_mm"),35.0) + _finite(r.get("phi_st_mm"),8.0)
            parts += [_dxf_line(left+cover,bot+cover,right-cover,bot+cover,"STIRRUPS"), _dxf_line(right-cover,bot+cover,right-cover,top-cover,"STIRRUPS"), _dxf_line(right-cover,top-cover,left+cover,top-cover,"STIRRUPS"), _dxf_line(left+cover,top-cover,left+cover,bot+cover,"STIRRUPS")]
            phi = max(_finite(r.get("phi_long_mm"),10.0), 6.0)
            for y,z in _bar_points_for_result(r):
                parts.append(_dxf_circle(ox + y, oy + z, phi/2.0, "REBARS"))
            txt = f"P{r.get('member','')} | {b/10:.0f}x{h/10:.0f} | {r.get('solucao','')} | {r.get('status','')}"
            parts.append(_dxf_text(left, top + 90, txt, 28, "TEXT"))
            if str(r.get("detailing_issues", "-")).strip() not in ["", "-"]:
                parts.append(_dxf_text(left, bot - 80, str(r.get("detailing_issues", ""))[:120], 20, "TEXT"))
    parts.append("0\nENDSEC\n0\nEOF\n")
    Path(path).write_text("".join(parts), encoding="utf-8")

# Override global writer and export command uses name lookup at runtime
write_columns_dxf_v4 = write_columns_dxf_v41


# --------------------------- XLSX/PDF v4.1 ---------------------------
def _metadata_df_v41(self) -> pd.DataFrame:
    return pd.DataFrame([
        ["Programa", APP_NAME], ["Versão", APP_VERSION], ["Autor", APP_AUTHOR], ["Repositório", GITHUB_URL],
        ["Data de exportação", datetime.now().strftime("%Y-%m-%d %H:%M")], ["Ficheiro de origem", self.input_file_path or "-"],
        ["Norma de referência", "Eurocódigo 2 / NP EN 1992-1-1"],
        ["Âmbito", "Dimensionamento/verificação de pilares de betão armado: ELU, ELS, V, T, pormenorização, superfície N-My-Mz e DXF"],
        ["Classe de betão", "lida exclusivamente da coluna Material; C30/37 só como fallback interno assinalado"],
        ["Descrição", APP_XLSX_DESCRIPTION],
        ["Limitações", "Ferramenta de apoio ao projecto; validar hipóteses, casos críticos e resultados por engenheiro responsável."],
    ], columns=["Campo", "Valor"])
ColumnsEC2App._metadata_df = _metadata_df_v41


def _write_excel_v41(self, path: str):
    res = self.df_results if self.df_results is not None else pd.DataFrame()
    import_quality = getattr(self, "df_import_quality", pd.DataFrame())
    if import_quality is None or import_quality.empty:
        import_quality = build_import_quality_cases_v41(self)
    surface_df = surface_points_df_v41(res)
    material_df = material_assumptions_df_v41(res)
    test_df = validation_test_cases_df_v41()
    sheets = {
        "00_Info": self._metadata_df(),
        "01_Parametros": self._parameters_df(),
        "02_Entrada_Dados": self.df_clean,
        "03_Pares_Member_Case": self.df_pair,
        "04_Qualidade_Importacao": import_quality,
        "05_Casos_Calculo": self.df_calc_input,
        "06_Resultados": res,
        "07_Resumo_Membros": self.df_summary,
        "08_Falhas": self.df_failures,
        "09_OK": self.df_ok,
        "10_Shortlists": self.build_shortlists_df(),
        "11_Validacao": self.df_validation,
        "12_Materiais_Assumidos": material_df,
        "13_Notas_EC2": self.df_notes,
    }
    tech_cols = {
        "14_ELS": [c for c in res.columns if c.startswith("service") or c in ["member","case","name","status"]],
        "15_Esf_Transverso": [c for c in res.columns if "shear" in c or "v_rd" in c or "asw_s_y" in c or "asw_s_z" in c or c in ["member","case","v_ed_y_kN","v_ed_z_kN"]],
        "16_Torcao": [c for c in res.columns if "torsion" in c or "t_rd" in c or "asl_t" in c or "asw_s_t" in c or c in ["member","case","mx_ed_kNm"]],
        "17_Pormenorizacao": [c for c in res.columns if "detailing" in c or c in ["member","case","solucao","phi_long_mm","phi_st_mm","s_st_mm","n_total","n_bars_y","n_bars_z"]],
        "18_Superficie_Resumo": [c for c in res.columns if "surface" in c or c in ["member","case","mrd_y_kNm","mrd_z_kNm","utilizacao"]],
        "19_Superficie_Pontos": [],
        "20_Memoria_Calculo": [c for c in ["member","case","combination_number","limit_state","material","material_assumed","n_ed_kN","my_i_kNm","my_j_kNm","moment_sign_y","curvature_y","mz_i_kNm","mz_j_kNm","moment_sign_z","curvature_z","rm_y","rm_z","lambda_y","lambda_lim_y","lambda_z","lambda_lim_z","m_imp_y_kNm","m_imp_z_kNm","m0e_y_kNm","m0e_z_kNm","m2_y_kNm","m2_z_kNm","my_ed_kNm","mz_ed_kNm","as_min_mm2","as_req_mm2","as_prov_mm2","mrd_y_kNm","mrd_z_kNm","utilizacao","status","failure_reason","recommendations","audit_note"] if c in res.columns],
        "21_Casos_Tipo_Validacao": [],
    }
    for name, cols in tech_cols.items():
        if name == "19_Superficie_Pontos":
            sheets[name] = surface_df
        elif name == "21_Casos_Tipo_Validacao":
            sheets[name] = test_df
        else:
            sheets[name] = res[cols].copy() if cols else pd.DataFrame()
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in sheets.items():
            if df is None:
                df = pd.DataFrame()
            df.to_excel(writer, sheet_name=name[:31], index=False)
        wb = writer.book
        props = wb.properties
        props.title = f"{APP_NAME} {APP_VERSION}"; props.subject = APP_SUBJECT; props.creator = APP_AUTHOR
        props.keywords = APP_KEYWORDS; props.category = APP_CATEGORY; props.description = APP_XLSX_DESCRIPTION; props.lastModifiedBy = APP_AUTHOR
        try:
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            from openpyxl.utils import get_column_letter
            header_fill = PatternFill("solid", fgColor="1F4E5F"); header_font = Font(color="FFFFFF", bold=True)
            warn_fill = PatternFill("solid", fgColor="FFF2CC"); fail_fill = PatternFill("solid", fgColor="FCE4D6")
            thin = Side(style="thin", color="D9E2E7"); border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for ws in wb.worksheets:
                ws.sheet_view.showGridLines = False; ws.freeze_panes = "A2"
                if ws.max_row >= 1:
                    for cell in ws[1]:
                        cell.fill = header_fill; cell.font = header_font; cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); cell.border = border
                # realçar estados em folhas de validação
                headers = [cell.value for cell in ws[1]] if ws.max_row >= 1 else []
                estado_idx = headers.index("estado") + 1 if "estado" in headers else (headers.index("Estado") + 1 if "Estado" in headers else None)
                for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 5000)):
                    for cell in row:
                        cell.border = border; cell.alignment = Alignment(vertical="top", wrap_text=True)
                    if estado_idx:
                        val = str(row[estado_idx-1].value or "")
                        if "Não conforme" in val or "Falha" in val:
                            for cell in row: cell.fill = fail_fill
                        elif "Verificar" in val:
                            for cell in row: cell.fill = warn_fill
                for col_idx, col in enumerate(ws.columns, start=1):
                    values = [str(c.value) for c in col[:200] if c.value is not None]
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(max([len(v) for v in values] + [10]) + 2, 48)
        except Exception:
            pass
ColumnsEC2App._write_excel = _write_excel_v41


def _write_pdf_v41(self, path: str):
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, PageBreak
    doc = SimpleDocTemplate(path, pagesize=landscape(A4), rightMargin=12*mm, leftMargin=12*mm, topMargin=12*mm, bottomMargin=12*mm)
    doc.title=f"{APP_NAME} {APP_VERSION}"; doc.author=APP_AUTHOR; doc.subject=APP_SUBJECT
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="ReportTitle", parent=styles["Title"], alignment=TA_CENTER, fontName="Courier-Bold", fontSize=14, leading=21, spaceAfter=10))
    styles.add(ParagraphStyle(name="ReportSubtitle", parent=styles["Normal"], alignment=TA_CENTER, fontName="Courier", fontSize=10, leading=15, textColor=colors.darkgrey, spaceAfter=8))
    styles.add(ParagraphStyle(name="BodyCourier", parent=styles["Normal"], fontName="Courier", fontSize=10, leading=15, spaceAfter=6))
    styles.add(ParagraphStyle(name="Small", parent=styles["Normal"], fontName="Courier", fontSize=8, leading=12))
    styles.add(ParagraphStyle(name="Cell", parent=styles["Small"], alignment=TA_LEFT, fontName="Courier", fontSize=7, leading=10.5))
    styles.add(ParagraphStyle(name="Section", parent=styles["Heading2"], fontName="Courier-Bold", fontSize=12, leading=18, spaceBefore=10, spaceAfter=20))
    def df_table(df, cols, max_rows=25):
        if df is None or df.empty:
            df = pd.DataFrame(columns=cols)
        present=[c for c in cols if c in df.columns]
        if not present:
            present = cols[:1]
            df = pd.DataFrame([{present[0]: "sem dados"}])
        data=[[Paragraph(str(c), styles["Cell"]) for c in present]]
        for _,r in df.head(max_rows).iterrows():
            row=[]
            for c in present:
                v=r.get(c,"")
                txt="" if pd.isna(v) else (f"{v:.2f}" if isinstance(v,float) and math.isfinite(v) else str(v))
                txt=txt.replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")
                row.append(Paragraph(txt, styles["Cell"]))
            data.append(row)
        widths=[270*mm/max(1,len(present))]*max(1,len(present))
        tb=Table(data, colWidths=widths, repeatRows=1); tb.setStyle(_pdf_table_style_courier(header=True)); return tb
    story=[]
    story.append(Paragraph(f"{APP_NAME} {APP_VERSION}", styles["ReportTitle"]))
    story.append(Paragraph("Memória de cálculo de pilares de betão armado segundo o Eurocódigo 2", styles["ReportSubtitle"]))
    n_total=len(self.df_results); n_ok=int((self.df_results.get("status",pd.Series(dtype=str))=="OK").sum()); n_fail=int((self.df_results.get("status",pd.Series(dtype=str))=="Falha").sum())
    n_assumed=int((self.df_results.get("material_assumed",pd.Series(dtype=str)).astype(str)=="Sim").sum()) if self.df_results is not None and not self.df_results.empty else 0
    meta=[["Programa", f"{APP_NAME} {APP_VERSION}", "Autor", APP_AUTHOR], ["Data", datetime.now().strftime("%Y-%m-%d %H:%M"), "Casos", str(n_total)], ["OK",str(n_ok),"Falhas",str(n_fail)], ["Materiais assumidos", str(n_assumed), "ELS indicado", getattr(self,"var_service_case",tk.StringVar(value="")).get().strip() or "simplificado"], ["Ficheiro", os.path.basename(self.input_file_path or "-"), "", ""]]
    t=Table(meta, colWidths=[38*mm,90*mm,38*mm,90*mm]); t.setStyle(_pdf_table_style_courier(header=False)); story.append(t); story.append(Spacer(1,5*mm))
    story.append(Paragraph("1. Qualidade da importação", styles["Section"]))
    q = getattr(self, "df_import_quality", pd.DataFrame())
    if q is None or q.empty:
        q = build_import_quality_cases_v41(self)
    story.append(df_table(q, ["member","case","estado","n_nodes_found","material_assumido","avisos"], 30))
    story.append(Paragraph("2. Resumo executivo", styles["Section"]))
    summary = self.df_summary if self.df_summary is not None and not self.df_summary.empty else self.df_results
    story.append(df_table(summary, ["member","case","material","material_assumed","n_ed_kN","my_ed_kNm","mz_ed_kNm","as_req_mm2","as_prov_mm2","solucao","status"], 30))
    story.append(PageBreak())
    story.append(Paragraph("3. Memória de cálculo — pilares críticos", styles["Section"]))
    mem_cols=["member","case","my_i_kNm","my_j_kNm","moment_sign_y","curvature_y","mz_i_kNm","mz_j_kNm","moment_sign_z","curvature_z","lambda_y","lambda_lim_y","lambda_z","lambda_lim_z","m_imp_y_kNm","m2_y_kNm","m_imp_z_kNm","m2_z_kNm","mrd_y_kNm","mrd_z_kNm","utilizacao"]
    story.append(df_table(summary, mem_cols, 25))
    story.append(Paragraph("4. Verificações complementares — ELS, V, T e pormenorização", styles["Section"]))
    comp_cols=["member","case","service_combination","service_status","service_wk_est_mm","shear_status","torsion_status","detailing_status","detailing_issues","recommendations"]
    story.append(df_table(summary, comp_cols, 30))
    if self.df_failures is not None and not self.df_failures.empty:
        story.append(PageBreak()); story.append(Paragraph("5. Falhas e recomendações", styles["Section"]))
        story.append(df_table(self.df_failures, ["member","case","failure_type","failure_reason","recommendations","shortlist_text"], 35))
    story.append(Spacer(1,4*mm)); story.append(Paragraph("Notas: relatório gerado automaticamente. Validar resultados críticos, hipóteses de comprimento efectivo, sinais locais da análise estrutural, combinações ELS, pormenorização final e desenhos DXF em projecto.", styles["Small"]))
    def footer(canvas, doc_obj):
        canvas.saveState(); canvas.setAuthor(APP_AUTHOR); canvas.setTitle(f"{APP_NAME} {APP_VERSION}"); canvas.setSubject(APP_SUBJECT); canvas.setFont("Courier",7); canvas.setFillColor(colors.grey); canvas.drawString(12*mm,7*mm,f"{APP_NAME} {APP_VERSION} | {APP_AUTHOR}"); canvas.drawRightString(285*mm,7*mm,f"Página {doc_obj.page}"); canvas.restoreState()
    doc.build(story, onFirstPage=footer, onLaterPages=footer)
ColumnsEC2App._write_pdf = _write_pdf_v41


def _build_normative_notes_v41(self) -> pd.DataFrame:
    notes = [
        ("Entrada", "Excel/tabela", "Inclui relatório de qualidade por member/case, materiais assumidos e unidades suspeitas."),
        ("Betão", "Material", "Classe de betão lida exclusivamente da tabela; C30/37 é fallback interno assinalado em relatório."),
        ("2.ª ordem", "EC2 5.8", "Curvatura nominal, imperfeições práticas, λlim, rm com sinal e identificação de curvatura simples/dupla."),
        ("Biaxial", "EC2 5.8.9", "Modo dimensionamento com critério biaxial; modo rigoroso com malha N-My-Mz mais densa."),
        ("Superfície", "N-My-Mz", "Exportação de pontos representativos da envolvente para auditoria/plotagem."),
        ("Esforço transverso", "EC2 6.2", "VRd,c, VRd,max e Asw/s requerido em Y e Z."),
        ("Torção", "EC2 6.3", "Modelo tubular simplificado com TRd,max, Asw/s e Asl requeridos."),
        ("ELS", "EC2 7", "Combinação indicada pelo utilizador quando disponível; caso contrário, verificação simplificada."),
        ("Pormenorização", "EC2 8/9", "Mínimos de varões, estribos, espaçamentos, varões intermédios e congestionamento."),
        ("DXF", "Quadro de pilares", "Exporta secções armadas e quadro com layers CONCRETE/REBARS/STIRRUPS/TEXT/TABLE."),
        ("Validação", "Casos-tipo", "Workbook inclui lista de casos-tipo a validar contra cálculo manual/software independente."),
    ]
    return pd.DataFrame(notes, columns=["Tema", "Referência", "Nota"])
ColumnsEC2App.build_normative_notes = _build_normative_notes_v41




# ============================================================
# ColumnsEC2 v4.2 — resumo por prumada/nome e DXF tipo quadro de pilares
# ============================================================
APP_VERSION = "v4.5"
APP_XLSX_DESCRIPTION = (
    "Workbook de cálculo com dados de entrada, validação robusta, pares member/case, "
    "resumo por prumada/nome, quadro de pilares, ELU, ELS, V, T, pormenorização, "
    "superfície resistente, falhas, shortlists e notas EC2."
)

# A coluna Story/Piso passa a ser reconhecida quando vier da tabela do Excel/tabela.
COLUMN_ALIASES.update({
    "story": ["story", "storey", "piso", "level", "floor", "pavimento"],
})


def _natural_key_v42(value):
    s = str(value or "").strip()
    return [int(t) if t.isdigit() else t.lower() for t in re.split(r"(\d+)", s)]


def _is_blank_v42(value) -> bool:
    s = str(value or "").strip()
    return s == "" or s.lower() in ["nan", "none", "null", "-"]


def _pillar_prumada_v42(row) -> str:
    """Identificador da prumada: usa Name quando existe; caso contrário usa Member."""
    name = row.get("name", "") if hasattr(row, "get") else ""
    member = row.get("member", "") if hasattr(row, "get") else ""
    if not _is_blank_v42(name):
        return str(name).strip()
    if not _is_blank_v42(member):
        return f"M{str(member).strip()}"
    return "Sem_nome"


def _combine_member_end_actions_v42(df: pd.DataFrame) -> pd.DataFrame:
    """
    Converte a tabela nó-a-nó da tabela de cálculo numa tabela por member/case,
    preservando também Story/Piso quando existir. A agregação continua por member/case/name,
    mas os resultados passam a transportar a prumada/nome para resumo e DXF.
    """
    if df is None or df.empty:
        return df

    rows = []
    if "story" not in df.columns:
        df = df.copy()
        df["story"] = ""

    for _, grp in df.groupby(["member", "case", "name"], dropna=False):
        grp = grp.sort_values("__row_order")
        r1 = grp.iloc[0]
        r2 = grp.iloc[1] if len(grp) >= 2 else grp.iloc[0]

        story_1 = r1.get("story", "")
        story_2 = r2.get("story", "")
        story = story_1 if not _is_blank_v42(story_1) else story_2

        row = {
            "member": r1.get("member", ""),
            "case": r1.get("case", ""),
            "name": r1.get("name", ""),
            "prumada": _pillar_prumada_v42(r1),
            "story": story,
            "node_i": r1.get("node", ""),
            "node_j": r2.get("node", ""),
            "member_case_i": f"{r1.get('member','')}/{r1.get('node','')}/{r1.get('case','')}",
            "member_case_j": f"{r2.get('member','')}/{r2.get('node','')}/{r2.get('case','')}",
            "fx_i": safe_float(r1.get("fx", 0.0), 0.0),
            "fx_j": safe_float(r2.get("fx", 0.0), 0.0),
            "fy_i": safe_float(r1.get("fy", 0.0), 0.0),
            "fy_j": safe_float(r2.get("fy", 0.0), 0.0),
            "fz_i": safe_float(r1.get("fz", 0.0), 0.0),
            "fz_j": safe_float(r2.get("fz", 0.0), 0.0),
            "mx_i": safe_float(r1.get("mx", 0.0), 0.0),
            "mx_j": safe_float(r2.get("mx", 0.0), 0.0),
            "my_i": safe_float(r1.get("my", 0.0), 0.0),
            "my_j": safe_float(r2.get("my", 0.0), 0.0),
            "mz_i": safe_float(r1.get("mz", 0.0), 0.0),
            "mz_j": safe_float(r2.get("mz", 0.0), 0.0),
            "length": safe_float(r1.get("length", 0.0), 0.0),
            "material": r1.get("material", "") or DEFAULT_CONCRETE_CLASS,
            "hy": safe_float(r1.get("hy", float("nan"))),
            "hz": safe_float(r1.get("hz", float("nan"))),
            "ax": safe_float(r1.get("ax", float("nan"))),
            "iy": safe_float(r1.get("iy", float("nan"))),
            "iz": safe_float(r1.get("iz", float("nan"))),
            "__row_order": safe_float(r1.get("__row_order", 0), 0),
            "n_nodes_found": len(grp),
        }

        row["fx"] = max(abs(row["fx_i"]), abs(row["fx_j"]))
        row["fy"] = max(abs(row["fy_i"]), abs(row["fy_j"]))
        row["fz"] = max(abs(row["fz_i"]), abs(row["fz_j"]))
        row["mx"] = max(abs(row["mx_i"]), abs(row["mx_j"]))
        row["my"] = max(abs(row["my_i"]), abs(row["my_j"]))
        row["mz"] = max(abs(row["mz_i"]), abs(row["mz_j"]))
        rows.append(row)

    return pd.DataFrame(rows).sort_values(["__row_order", "member", "case"]).reset_index(drop=True)


# Substitui a função global usada pela GUI no momento da importação.
combine_member_end_actions = _combine_member_end_actions_v42


_old_design_one_v42_base = ColumnDesigner.design_one

def _design_one_v42(self, row: pd.Series, prebuilt_candidates=None):
    out = _old_design_one_v42_base(self, row, prebuilt_candidates=prebuilt_candidates)
    if isinstance(out, dict):
        out["story"] = row.get("story", "")
        out["prumada"] = _pillar_prumada_v42(out)
        out["pillar_name"] = out["prumada"]
        # chave física do tramo: prumada + member. Útil para quadro de pilares e resumo por piso/tramo.
        out["pillar_segment_key"] = f"{out.get('prumada','')}|{out.get('member','')}"
    return out

ColumnDesigner.design_one = _design_one_v42


def _governing_score_v42(df: pd.DataFrame) -> pd.Series:
    if df is None or df.empty:
        return pd.Series(dtype=float)
    status_fail = df.get("status", pd.Series(index=df.index, dtype=str)).astype(str).eq("Falha").astype(float) * 1e6
    util = pd.to_numeric(df.get("utilizacao", pd.Series(index=df.index, dtype=float)), errors="coerce").fillna(0.0) * 1000.0
    as_req = pd.to_numeric(df.get("as_req_mm2", pd.Series(index=df.index, dtype=float)), errors="coerce").fillna(0.0) / 1000.0
    n_ed = pd.to_numeric(df.get("n_ed_kN", pd.Series(index=df.index, dtype=float)), errors="coerce").fillna(0.0) / 1000.0
    return status_fail + util + as_req + n_ed


def build_pillar_schedule_df_v42(results: pd.DataFrame) -> pd.DataFrame:
    """
    Tabela base do quadro de pilares: uma linha por prumada/member,
    escolhendo o caso governante de cada tramo da prumada.
    """
    if results is None or results.empty:
        return pd.DataFrame()
    work = results.copy()
    if "prumada" not in work.columns:
        work["prumada"] = work.apply(_pillar_prumada_v42, axis=1)
    if "story" not in work.columns:
        work["story"] = ""
    work["_score_v42"] = _governing_score_v42(work)
    work["_story_sort"] = work["story"].astype(str)
    work["_member_sort"] = work["member"].astype(str)
    # escolher o caso governante de cada tramo físico da prumada
    work = work.sort_values(["prumada", "member", "_score_v42"], ascending=[True, True, False])
    group_cols = ["prumada", "member"]
    schedule = work.groupby(group_cols, as_index=False, dropna=False).first()
    schedule["segment_order"] = range(1, len(schedule) + 1)
    schedule = schedule.sort_values(
        by=["prumada", "story", "member"],
        key=lambda col: col.map(lambda x: tuple(_natural_key_v42(x))),
    ).reset_index(drop=True)
    drop_cols = [c for c in ["_score_v42", "_story_sort", "_member_sort"] if c in schedule.columns]
    schedule.drop(columns=drop_cols, inplace=True, errors="ignore")
    return schedule


def _build_summary_by_prumada_v42(self, results: pd.DataFrame) -> pd.DataFrame:
    """
    Resumo principal por prumada/nome, não por linha member/case.
    Todos os pilares com o mesmo Name, por exemplo P1, são tratados como a mesma prumada.
    """
    if results is None or results.empty:
        return pd.DataFrame()
    work = results.copy()
    if "prumada" not in work.columns:
        work["prumada"] = work.apply(_pillar_prumada_v42, axis=1)
    if "story" not in work.columns:
        work["story"] = ""
    work["_score_v42"] = _governing_score_v42(work)
    summaries = []
    for prumada, grp in work.groupby("prumada", dropna=False):
        grp_sorted = grp.sort_values("_score_v42", ascending=False)
        gov = grp_sorted.iloc[0].copy()
        members = sorted([str(v) for v in grp["member"].dropna().unique()], key=_natural_key_v42) if "member" in grp.columns else []
        cases = sorted([str(v) for v in grp["case"].dropna().unique()], key=_natural_key_v42) if "case" in grp.columns else []
        stories = sorted([str(v) for v in grp["story"].dropna().unique() if not _is_blank_v42(v)], key=_natural_key_v42) if "story" in grp.columns else []
        gov["prumada"] = prumada
        gov["pillar_name"] = prumada
        gov["members_in_prumada"] = ", ".join(members)
        gov["cases_in_prumada"] = ", ".join(cases[:12]) + (" ..." if len(cases) > 12 else "")
        gov["stories_in_prumada"] = ", ".join(stories)
        gov["n_members_prumada"] = len(members)
        gov["n_cases_prumada"] = len(grp)
        gov["status_prumada"] = "Falha" if (grp.get("status", pd.Series(dtype=str)).astype(str) == "Falha").any() else gov.get("status", "")
        # Soluções distintas da prumada, útil para detectar alterações de armadura ao longo dos pisos.
        sols = [str(v) for v in grp.get("solucao", pd.Series(dtype=str)).dropna().unique() if not _is_blank_v42(v)]
        gov["solucoes_prumada"] = " | ".join(sols[:6]) + (" ..." if len(sols) > 6 else "")
        summaries.append(gov)
    out = pd.DataFrame(summaries)
    if not out.empty:
        out = out.sort_values("prumada", key=lambda col: col.map(lambda x: tuple(_natural_key_v42(x)))).reset_index(drop=True)
        out.drop(columns=["_score_v42"], inplace=True, errors="ignore")
    return out

ColumnsEC2App.build_summary_by_member = _build_summary_by_prumada_v42


def _format_dim_m_v42(mm_value: float) -> str:
    try:
        return f"{mm_value/1000.0:.2f}"
    except Exception:
        return ""


def _long_label_v42(r) -> str:
    sol = str(r.get("solucao", "") or "")
    if "+" in sol:
        return sol.split("+")[0].strip()
    if not _is_blank_v42(sol):
        return sol.strip()
    n = int(_finite(r.get("n_total"), 0))
    phi = int(_finite(r.get("phi_long_mm"), 0))
    return f"{n}Ø{phi}" if n and phi else ""


def _stirrup_label_v42(r) -> str:
    phi = int(_finite(r.get("phi_st_mm"), 0))
    s = _finite(r.get("s_st_mm"), 0.0)
    if not phi or not s:
        return ""
    return f"Estr. Ø{phi}//{s/1000.0:.2f}"


def _draw_section_in_cell_v42(parts: list, r, cx: float, cy: float, cw: float, ch: float):
    b = _finite(r.get("b_cm"), 0.0) * 10.0
    h = _finite(r.get("h_cm"), 0.0) * 10.0
    if b <= 0 or h <= 0:
        parts.append(_dxf_line(cx - 120, cy, cx + 120, cy, "TEXT"))
        return
    scale = min((cw * 0.38) / max(b, 1.0), (ch * 0.32) / max(h, 1.0), 2.2)
    bs, hs = b * scale, h * scale
    left, right = cx - bs/2.0, cx + bs/2.0
    bot, top = cy - hs/2.0, cy + hs/2.0

    # Betão e estribo
    parts += [
        _dxf_line(left, bot, right, bot, "CONCRETE"),
        _dxf_line(right, bot, right, top, "CONCRETE"),
        _dxf_line(right, top, left, top, "CONCRETE"),
        _dxf_line(left, top, left, bot, "CONCRETE"),
    ]
    cover = (_finite(r.get("cover_mm"), 35.0) + _finite(r.get("phi_st_mm"), 8.0)) * scale
    if cover * 2 < min(bs, hs):
        parts += [
            _dxf_line(left + cover, bot + cover, right - cover, bot + cover, "STIRRUPS"),
            _dxf_line(right - cover, bot + cover, right - cover, top - cover, "STIRRUPS"),
            _dxf_line(right - cover, top - cover, left + cover, top - cover, "STIRRUPS"),
            _dxf_line(left + cover, top - cover, left + cover, bot + cover, "STIRRUPS"),
        ]
    phi = max(_finite(r.get("phi_long_mm"), 10.0) * scale, 7.0)
    for y, z in _bar_points_for_result(r):
        parts.append(_dxf_circle(cx + y * scale, cy + z * scale, phi/2.0, "REBARS"))

    # Dimensões simples tipo quadro de pilares
    dim_off = 95.0
    parts.append(_dxf_line(left, bot - dim_off, right, bot - dim_off, "TEXT"))
    parts.append(_dxf_line(left, bot - dim_off - 25, left, bot - dim_off + 25, "TEXT"))
    parts.append(_dxf_line(right, bot - dim_off - 25, right, bot - dim_off + 25, "TEXT"))
    parts.append(_dxf_text(cx - 55, bot - dim_off - 55, _format_dim_m_v42(b), 24, "TEXT"))

    parts.append(_dxf_line(left - dim_off, bot, left - dim_off, top, "TEXT"))
    parts.append(_dxf_line(left - dim_off - 25, bot, left - dim_off + 25, bot, "TEXT"))
    parts.append(_dxf_line(left - dim_off - 25, top, left - dim_off + 25, top, "TEXT"))
    parts.append(_dxf_text(left - dim_off - 90, cy - 15, _format_dim_m_v42(h), 24, "TEXT"))

    # Identificação da armadura
    parts.append(_dxf_text(cx - 120, bot - 190, _long_label_v42(r), 25, "TEXT"))
    parts.append(_dxf_text(cx - 170, bot - 235, _stirrup_label_v42(r), 23, "TEXT"))
    story = str(r.get("story", "") or "").strip()
    if story:
        parts.append(_dxf_text(left, top + 55, story, 20, "TEXT"))


def write_columns_dxf_v42(path: str, df: pd.DataFrame):
    """DXF tipo quadro de pilares: uma coluna por prumada/nome e linhas por tramo/member."""
    schedule = build_pillar_schedule_df_v42(df)
    parts = ["0\nSECTION\n2\nHEADER\n0\nENDSEC\n", _dxf_layer_table_v41(), "0\nSECTION\n2\nENTITIES\n"]
    if schedule is None or schedule.empty:
        parts.append(_dxf_text(0, 0, "Sem resultados", 50, "TEXT"))
        parts.append("0\nENDSEC\n0\nEOF\n")
        Path(path).write_text("".join(parts), encoding="utf-8")
        return

    prumadas = sorted(schedule["prumada"].astype(str).unique(), key=_natural_key_v42)
    groups = {p: schedule[schedule["prumada"].astype(str) == p].copy() for p in prumadas}
    max_rows = max(len(g) for g in groups.values())
    # Limitar a dimensão para DXF manejável; continua tudo no XLSX.
    prumadas = prumadas[:18]
    max_rows = min(max_rows, 12)

    cw, ch = 1450.0, 1280.0
    title_h, header_h = 260.0, 240.0
    x0, y0 = 0.0, 0.0
    total_w = cw * len(prumadas)
    total_h = title_h + header_h + ch * max_rows

    title = "QUADRO DE PILARES (por prumada/nome)"
    parts.append(_dxf_text(x0 + total_w * 0.34, y0 + 110, title, 55, "TEXT"))

    # Grelha principal
    top = y0
    bottom = y0 - total_h
    left = x0
    right = x0 + total_w
    # contorno e linhas horizontais
    parts += [
        _dxf_line(left, top, right, top, "TABLE"),
        _dxf_line(right, top, right, bottom, "TABLE"),
        _dxf_line(right, bottom, left, bottom, "TABLE"),
        _dxf_line(left, bottom, left, top, "TABLE"),
        _dxf_line(left, top - title_h, right, top - title_h, "TABLE"),
        _dxf_line(left, top - title_h - header_h, right, top - title_h - header_h, "TABLE"),
    ]
    for r_i in range(1, max_rows + 1):
        y = top - title_h - header_h - r_i * ch
        parts.append(_dxf_line(left, y, right, y, "TABLE"))
    # linhas verticais e cabeçalhos
    for c_i, p in enumerate(prumadas):
        x = left + c_i * cw
        parts.append(_dxf_line(x, top - title_h, x, bottom, "TABLE"))
        parts.append(_dxf_text(x + cw * 0.42, top - title_h - 145, str(p), 40, "TEXT"))
    parts.append(_dxf_line(right, top - title_h, right, bottom, "TABLE"))

    # Conteúdo das células
    for c_i, p in enumerate(prumadas):
        grp = groups[p].sort_values(
            by=["story", "member"],
            key=lambda col: col.map(lambda x: tuple(_natural_key_v42(x))),
        ).reset_index(drop=True)
        for r_i in range(max_rows):
            cell_left = left + c_i * cw
            cell_top = top - title_h - header_h - r_i * ch
            cx = cell_left + cw / 2.0
            cy = cell_top - ch * 0.47
            if r_i < len(grp):
                _draw_section_in_cell_v42(parts, grp.iloc[r_i], cx, cy, cw, ch)
            else:
                parts.append(_dxf_line(cx - 120, cy, cx + 120, cy, "TEXT"))

    # Quadro-resumo textual abaixo para auditoria rápida
    y_table = bottom - 420.0
    parts.append(_dxf_text(left, y_table + 130, "Resumo governante por prumada", 36, "TABLE"))
    headers = ["PRUMADA", "MEMBERS", "SECÇÃO", "ARM. LONG.", "ESTRIBOS", "ESTADO"]
    widths = [260, 420, 260, 360, 360, 220]
    x = left
    for h, w in zip(headers, widths):
        parts.append(_dxf_text(x + 8, y_table, h, 24, "TABLE")); x += w
    # Uma linha por prumada com a solução governante.
    for i, p in enumerate(prumadas[:40], start=1):
        g = groups[p]
        g = g.copy()
        g["_score_v42"] = _governing_score_v42(g)
        r = g.sort_values("_score_v42", ascending=False).iloc[0]
        members = ",".join(sorted([str(v) for v in g["member"].dropna().unique()], key=_natural_key_v42)[:6])
        vals = [
            str(p), members,
            f"{_finite(r.get('b_cm')):.0f}x{_finite(r.get('h_cm')):.0f}",
            _long_label_v42(r), _stirrup_label_v42(r), str(r.get("status", "")),
        ]
        x = left; y = y_table - i * 55
        for val, w in zip(vals, widths):
            parts.append(_dxf_text(x + 8, y, val, 20, "TABLE")); x += w

    parts.append("0\nENDSEC\n0\nEOF\n")
    Path(path).write_text("".join(parts), encoding="utf-8")


# O exportador de DXF passa a usar os resultados completos, para manter os tramos da prumada.
def _export_dxf_v42(self):
    src = self.df_results if self.df_results is not None and not self.df_results.empty else self.df_summary
    if src is None or src.empty:
        messagebox.showwarning("Aviso", "Não há resultados para exportar em DXF.")
        return
    path = filedialog.asksaveasfilename(
        title="Exportar quadro de pilares por prumada/nome",
        defaultextension=".dxf",
        filetypes=[("DXF", "*.dxf")],
    )
    if not path:
        return
    try:
        write_columns_dxf_v42(path, src)
        self.status_var.set(f"DXF exportado por prumada/nome para: {path}")
    except Exception as err:
        messagebox.showerror("Erro", f"Não foi possível exportar DXF.\n\n{err}")

write_columns_dxf_v4 = write_columns_dxf_v42
ColumnsEC2App.export_dxf = _export_dxf_v42


# Acrescenta folhas específicas ao Excel sem reescrever todo o exportador v4.1.
_old_write_excel_v42_base = ColumnsEC2App._write_excel

def _write_excel_v42(self, path: str):
    _old_write_excel_v42_base(self, path)
    try:
        schedule = build_pillar_schedule_df_v42(self.df_results)
        summary_prumadas = self.build_summary_by_member(self.df_results)
        with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            schedule.to_excel(writer, sheet_name="22_Quadro_Pilares", index=False)
            summary_prumadas.to_excel(writer, sheet_name="23_Resumo_Prumadas", index=False)
            wb = writer.book
            props = wb.properties
            props.title = f"{APP_NAME} {APP_VERSION}"
            props.description = APP_XLSX_DESCRIPTION
            try:
                from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
                from openpyxl.utils import get_column_letter
                header_fill = PatternFill("solid", fgColor="1F4E5F")
                header_font = Font(color="FFFFFF", bold=True)
                thin = Side(style="thin", color="D9E2E7")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                for wsname in ["22_Quadro_Pilares", "23_Resumo_Prumadas"]:
                    ws = wb[wsname]
                    ws.sheet_view.showGridLines = False
                    ws.freeze_panes = "A2"
                    if ws.max_row >= 1:
                        for cell in ws[1]:
                            cell.fill = header_fill; cell.font = header_font
                            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                            cell.border = border
                    for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 5000)):
                        for cell in row:
                            cell.border = border; cell.alignment = Alignment(vertical="top", wrap_text=True)
                    for col_idx, col in enumerate(ws.columns, start=1):
                        values = [str(c.value) for c in col[:200] if c.value is not None]
                        ws.column_dimensions[get_column_letter(col_idx)].width = min(max([len(v) for v in values] + [10]) + 2, 48)
            except Exception:
                pass
    except Exception:
        # A exportação principal não deve falhar se a folha de quadro tiver algum problema.
        pass

ColumnsEC2App._write_excel = _write_excel_v42


# Actualiza notas normativas para reflectir resumo por prumada/nome.
def _build_normative_notes_v42(self) -> pd.DataFrame:
    notes = _build_normative_notes_v41(self)
    extra = pd.DataFrame([
        ("Resumo", "Prumada/Nome", "Os resultados resumidos são agregados por Name; se Name estiver vazio, é usado Member como fallback."),
        ("DXF", "Quadro de pilares", "O DXF organiza uma coluna por prumada/nome e uma linha por tramo/member governante, em formato próximo de quadro de pilares de projecto."),
        ("Excel", "Quadro_Pilares", "O workbook inclui folhas específicas 22_Quadro_Pilares e 23_Resumo_Prumadas."),
    ], columns=["Tema", "Referência", "Nota"])
    return pd.concat([notes, extra], ignore_index=True)

ColumnsEC2App.build_normative_notes = _build_normative_notes_v42





# ============================================================
# ColumnsEC2 v4.3 — PDF sintético, sem menções a software específico
# e gestão prática das falhas
# ============================================================
APP_VERSION = "v4.5"
APP_XLSX_DESCRIPTION = (
    "Workbook de cálculo com entrada de dados, validação robusta, pares member/case, "
    "resumo por prumada/nome, quadro de pilares, ELU, ELS, V, T, pormenorização, "
    "superfície resistente, falhas, shortlists e notas EC2. O PDF é uma síntese executiva; "
    "o XLSX conserva a memória completa."
)


def _failure_policy_v43(row) -> Dict[str, str]:
    """Classifica a falha em termos práticos de decisão de projecto."""
    status = str(row.get("status", "") or "")
    failure_type = str(row.get("failure_type", "") or "")
    reason = str(row.get("failure_reason", "") or "")
    service_status = str(row.get("service_status", "") or "")
    shear_status = str(row.get("shear_status", "") or "")
    torsion_status = str(row.get("torsion_status", "") or "")
    detailing_status = str(row.get("detailing_status", "") or "")

    if status == "OK":
        warnings = []
        if "Verificar" in service_status or "Falha" in service_status:
            warnings.append("ELS")
        if "Falha" in shear_status:
            warnings.append("esforço transverso")
        if "Falha" in torsion_status:
            warnings.append("torção")
        if "Falha" in detailing_status:
            warnings.append("pormenorização")
        if warnings:
            return {
                "failure_severity": "Aviso",
                "design_decision": "Aceitável apenas após revisão das verificações complementares",
                "review_priority": "Média",
                "failure_action": "Rever: " + ", ".join(warnings),
            }
        return {
            "failure_severity": "OK",
            "design_decision": "Aceitável no âmbito das hipóteses adoptadas",
            "review_priority": "Normal",
            "failure_action": "Sem acção correctiva obrigatória",
        }

    # Estado diferente de OK deve ser bloqueante para projecto final.
    if failure_type == "dados_incompletos":
        return {
            "failure_severity": "Bloqueante",
            "design_decision": "Não dimensionar até corrigir os dados de entrada",
            "review_priority": "Alta",
            "failure_action": "Corrigir a tabela: garantir dois nós por member/case, material, secção, comprimento e esforços completos.",
        }
    if failure_type == "armadura_insuficiente":
        return {
            "failure_severity": "Bloqueante",
            "design_decision": "Solução não aceite",
            "review_priority": "Alta",
            "failure_action": "Aumentar secção, permitir mais armadura longitudinal ou rever esforços governantes.",
        }
    if failure_type == "pormenorizacao":
        return {
            "failure_severity": "Bloqueante",
            "design_decision": "Solução não aceite sem alteração construtiva",
            "review_priority": "Alta",
            "failure_action": "Reduzir congestionamento, aumentar secção, rever recobrimento, varões por face, estribos e espaçamentos.",
        }
    if failure_type == "resistencia_biaxial" or "biaxial" in reason.lower():
        return {
            "failure_severity": "Bloqueante",
            "design_decision": "Solução resistente não verificada",
            "review_priority": "Alta",
            "failure_action": "Aumentar secção/armadura, executar modo rigoroso, rever combinação crítica e validar superfície N-My-Mz.",
        }
    if failure_type == "esbelteza":
        return {
            "failure_severity": "Bloqueante",
            "design_decision": "Solução não aceite sem revisão da estabilidade",
            "review_priority": "Alta",
            "failure_action": "Reduzir comprimento efectivo, aumentar rigidez, rever contraventamento e segunda ordem.",
        }
    return {
        "failure_severity": "Bloqueante",
        "design_decision": "Requer revisão técnica antes de aceitação",
        "review_priority": "Alta",
        "failure_action": str(row.get("recommendations", "") or "Rever dados, esforços, secção e armaduras."),
    }


def enrich_failures_v43(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    out = df.copy()
    policies = out.apply(lambda r: pd.Series(_failure_policy_v43(r)), axis=1)
    for c in policies.columns:
        out[c] = policies[c]
    # Campo curto para relatórios.
    out["failure_summary"] = out.apply(
        lambda r: (
            f"{r.get('failure_severity','')} | {r.get('failure_type','')} | "
            f"{r.get('failure_action','')}"
        ).strip(),
        axis=1,
    )
    return out


_old_design_dataframe_v43 = ColumnDesigner.design_dataframe

def _design_dataframe_v43(self, df: pd.DataFrame, progress_callback=None):
    res = _old_design_dataframe_v43(self, df, progress_callback=progress_callback)
    return enrich_failures_v43(res)

ColumnDesigner.design_dataframe = _design_dataframe_v43


# Mantém o XLSX completo, mas actualiza nomes de folhas e adiciona a gestão de falhas.
_old_write_excel_v43_base = ColumnsEC2App._write_excel

def _write_excel_v43(self, path: str):
    # Primeiro usa o exportador completo existente.
    _old_write_excel_v43_base(self, path)
    try:
        res = enrich_failures_v43(self.df_results) if self.df_results is not None else pd.DataFrame()
        failures = res[res.get("status", pd.Series(dtype=str)).astype(str) == "Falha"].copy() if not res.empty else pd.DataFrame()
        failure_policy = res[[c for c in [
            "prumada", "member", "case", "name", "status", "failure_type", "failure_severity",
            "design_decision", "review_priority", "failure_reason", "failure_action", "recommendations"
        ] if c in res.columns]].copy() if not res.empty else pd.DataFrame()
        with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            # Folhas adicionais com nomes neutros e gestão prática das falhas.
            res.to_excel(writer, sheet_name="06_Resultados", index=False)
            failures.to_excel(writer, sheet_name="07_Falhas", index=False)
            failure_policy.to_excel(writer, sheet_name="24_Gestao_Falhas", index=False)
            wb = writer.book
            props = wb.properties
            props.title = f"{APP_NAME} {APP_VERSION}"
            props.description = APP_XLSX_DESCRIPTION
            try:
                from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
                from openpyxl.utils import get_column_letter
                header_fill = PatternFill("solid", fgColor="1F4E5F")
                header_font = Font(color="FFFFFF", bold=True)
                fail_fill = PatternFill("solid", fgColor="FCE4D6")
                warn_fill = PatternFill("solid", fgColor="FFF2CC")
                ok_fill = PatternFill("solid", fgColor="E2F0D9")
                thin = Side(style="thin", color="D9E2E7")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                for wsname in ["06_Resultados", "07_Falhas", "24_Gestao_Falhas"]:
                    if wsname not in wb.sheetnames:
                        continue
                    ws = wb[wsname]
                    ws.sheet_view.showGridLines = False
                    ws.freeze_panes = "A2"
                    headers = [cell.value for cell in ws[1]] if ws.max_row >= 1 else []
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        cell.border = border
                    sev_idx = headers.index("failure_severity") + 1 if "failure_severity" in headers else None
                    for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 5000)):
                        for cell in row:
                            cell.border = border
                            cell.alignment = Alignment(vertical="top", wrap_text=True)
                        if sev_idx:
                            sev = str(row[sev_idx - 1].value or "")
                            fill = fail_fill if sev == "Bloqueante" else warn_fill if sev == "Aviso" else ok_fill if sev == "OK" else None
                            if fill:
                                for cell in row:
                                    cell.fill = fill
                    for col_idx, col in enumerate(ws.columns, start=1):
                        values = [str(c.value) for c in col[:200] if c.value is not None]
                        ws.column_dimensions[get_column_letter(col_idx)].width = min(max([len(v) for v in values] + [10]) + 2, 55)
            except Exception:
                pass
    except Exception:
        pass

ColumnsEC2App._write_excel = _write_excel_v43


# Relatório PDF v4.3: síntese executiva apenas; XLSX mantém a memória completa.
def _write_pdf_v43(self, path: str):
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, PageBreak

    res = enrich_failures_v43(self.df_results) if self.df_results is not None else pd.DataFrame()
    summary = self.df_summary if self.df_summary is not None and not self.df_summary.empty else res
    summary = enrich_failures_v43(summary) if summary is not None and not summary.empty else pd.DataFrame()
    failures = res[res.get("status", pd.Series(dtype=str)).astype(str) == "Falha"].copy() if not res.empty else pd.DataFrame()

    doc = SimpleDocTemplate(
        path,
        pagesize=landscape(A4),
        rightMargin=12 * mm,
        leftMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )
    doc.title = f"{APP_NAME} {APP_VERSION}"
    doc.author = APP_AUTHOR
    doc.subject = APP_SUBJECT
    doc.creator = APP_NAME

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="ReportTitle", parent=styles["Title"], alignment=TA_CENTER, fontName="Courier-Bold", fontSize=14, leading=21, spaceAfter=10))
    styles.add(ParagraphStyle(name="ReportSubtitle", parent=styles["Normal"], alignment=TA_CENTER, fontName="Courier", fontSize=10, leading=15, textColor=colors.darkgrey, spaceAfter=8))
    styles.add(ParagraphStyle(name="BodyCourier", parent=styles["Normal"], fontName="Courier", fontSize=10, leading=15, spaceAfter=6))
    styles.add(ParagraphStyle(name="Small", parent=styles["Normal"], fontName="Courier", fontSize=8, leading=12))
    styles.add(ParagraphStyle(name="Cell", parent=styles["Small"], alignment=TA_LEFT, fontName="Courier", fontSize=7, leading=10.5))
    styles.add(ParagraphStyle(name="Section", parent=styles["Heading2"], fontName="Courier-Bold", fontSize=12, leading=18, spaceBefore=10, spaceAfter=12))

    def _clean_txt(x):
        if x is None or (isinstance(x, float) and not math.isfinite(x)) or pd.isna(x):
            return ""
        if isinstance(x, float):
            return f"{x:.2f}"
        return str(x).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

    def df_table(df, cols, max_rows=25, widths=None):
        if df is None or df.empty:
            df = pd.DataFrame(columns=cols)
        present = [c for c in cols if c in df.columns]
        if not present:
            present = ["nota"]
            df = pd.DataFrame([{"nota": "sem dados"}])
        data = [[Paragraph(str(c), styles["Cell"]) for c in present]]
        for _, r in df.head(max_rows).iterrows():
            data.append([Paragraph(_clean_txt(r.get(c, "")), styles["Cell"]) for c in present])
        if widths is None:
            widths = [270 * mm / max(1, len(present))] * len(present)
        tb = Table(data, colWidths=widths, repeatRows=1)
        tb.setStyle(_pdf_table_style_courier(header=True))
        return tb

    n_total = len(res)
    n_ok = int((res.get("status", pd.Series(dtype=str)) == "OK").sum()) if not res.empty else 0
    n_fail = int((res.get("status", pd.Series(dtype=str)) == "Falha").sum()) if not res.empty else 0
    n_warn = int((res.get("failure_severity", pd.Series(dtype=str)) == "Aviso").sum()) if not res.empty else 0
    n_block = int((res.get("failure_severity", pd.Series(dtype=str)) == "Bloqueante").sum()) if not res.empty else 0
    n_assumed = int((res.get("material_assumed", pd.Series(dtype=str)).astype(str) == "Sim").sum()) if not res.empty and "material_assumed" in res.columns else 0

    story = []
    story.append(Paragraph(f"{APP_NAME} {APP_VERSION}", styles["ReportTitle"]))
    story.append(Paragraph("Síntese de dimensionamento de pilares segundo o Eurocódigo 2", styles["ReportSubtitle"]))

    meta = [
        ["Programa", f"{APP_NAME} {APP_VERSION}", "Autor", APP_AUTHOR],
        ["Data", datetime.now().strftime("%Y-%m-%d %H:%M"), "Casos analisados", str(n_total)],
        ["OK", str(n_ok), "Falhas bloqueantes", str(n_block)],
        ["Avisos", str(n_warn), "Materiais assumidos", str(n_assumed)],
        ["ELS", getattr(self, "var_service_case", tk.StringVar(value="")).get().strip() or "simplificado", "Ficheiro", os.path.basename(self.input_file_path or "-")],
    ]
    t = Table(meta, colWidths=[38 * mm, 90 * mm, 38 * mm, 90 * mm])
    t.setStyle(_pdf_table_style_courier(header=False))
    story.append(t)
    story.append(Spacer(1, 5 * mm))

    story.append(Paragraph("1. Resumo executivo por prumada/nome", styles["Section"]))
    exec_cols = ["prumada", "member", "material", "b_cm", "h_cm", "n_ed_kN", "as_req_mm2", "as_prov_mm2", "solucao", "status", "failure_severity"]
    story.append(df_table(summary, exec_cols, max_rows=32))

    story.append(Paragraph("2. Verificações condicionantes", styles["Section"]))
    cond_cols = ["prumada", "member", "case", "utilizacao", "service_status", "shear_status", "torsion_status", "detailing_status", "design_decision"]
    story.append(df_table(summary, cond_cols, max_rows=28))

    if failures is not None and not failures.empty:
        story.append(PageBreak())
        story.append(Paragraph("3. Falhas — decisão e acção recomendada", styles["Section"]))
        fail_cols = ["prumada", "member", "case", "failure_type", "failure_severity", "design_decision", "failure_action"]
        story.append(df_table(failures, fail_cols, max_rows=45, widths=[25*mm, 22*mm, 18*mm, 34*mm, 25*mm, 58*mm, 88*mm]))
        story.append(Spacer(1, 3 * mm))
        story.append(Paragraph(
            "Critério de utilização: qualquer falha classificada como Bloqueante não deve ser aceite em projecto sem correcção dos dados, revisão da solução ou validação independente do pilar crítico.",
            styles["BodyCourier"],
        ))

    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph(
        "Notas: este PDF é uma síntese executiva. A memória completa, shortlists, combinações, ELS, esforço transverso, torção, pormenorização, superfície resistente e metadados permanecem no ficheiro XLSX exportado.",
        styles["Small"],
    ))

    def footer(canvas, doc_obj):
        canvas.saveState()
        canvas.setAuthor(APP_AUTHOR)
        canvas.setTitle(f"{APP_NAME} {APP_VERSION}")
        canvas.setSubject(APP_SUBJECT)
        canvas.setFont("Courier", 7)
        canvas.setFillColor(colors.grey)
        canvas.drawString(12 * mm, 7 * mm, f"{APP_NAME} {APP_VERSION} | {APP_AUTHOR}")
        canvas.drawRightString(285 * mm, 7 * mm, f"Página {doc_obj.page}")
        canvas.restoreState()

    doc.build(story, onFirstPage=footer, onLaterPages=footer)

ColumnsEC2App._write_pdf = _write_pdf_v43


# Relatório interno da GUI também passa a ser sintético e orientado à decisão.
def _update_report_v43(self):
    self.report_txt.delete("1.0", "end")
    if self.df_results is None or self.df_results.empty:
        self.report_txt.insert("1.0", "Sem resultados. Importe a tabela e execute o cálculo.")
        return
    res = enrich_failures_v43(self.df_results)
    n_total = len(res)
    n_ok = int((res.get("status", pd.Series(dtype=str)) == "OK").sum())
    n_block = int((res.get("failure_severity", pd.Series(dtype=str)) == "Bloqueante").sum())
    n_warn = int((res.get("failure_severity", pd.Series(dtype=str)) == "Aviso").sum())
    lines = []
    lines.append(f"{APP_NAME} {APP_VERSION}\n")
    lines.append("Síntese de dimensionamento de pilares\n\n")
    lines.append(f"Casos analisados: {n_total} | OK: {n_ok} | Falhas bloqueantes: {n_block} | Avisos: {n_warn}\n\n")
    failures = res[res.get("status", pd.Series(dtype=str)).astype(str) == "Falha"].copy()
    if failures.empty:
        lines.append("Não foram identificadas falhas bloqueantes nos casos calculados.\n")
    else:
        lines.append("Falhas principais e acção recomendada:\n")
        for _, r in failures.head(30).iterrows():
            lines.append(
                f"- Prumada {r.get('prumada', r.get('name',''))} | Member {r.get('member','')} | Caso {r.get('case','')} | "
                f"{r.get('failure_type','')} | {r.get('failure_action','')}\n"
            )
    lines.append("\nNota: o PDF é uma síntese; exporte o XLSX para memória completa e auditoria técnica.\n")
    self.report_txt.insert("1.0", "".join(lines))

ColumnsEC2App.update_report = _update_report_v43


# Ajustes finais de textos visíveis que possam ter ficado por actualizar.
_old_build_normative_notes_v43_base = ColumnsEC2App.build_normative_notes

def _build_normative_notes_v43(self) -> pd.DataFrame:
    notes = _old_build_normative_notes_v43_base(self).copy()
    for col in notes.columns:
        notes[col] = notes[col].astype(str)
    extra = pd.DataFrame([
        ("Relatório PDF", "Síntese executiva", "O PDF contém apenas resultados essenciais, falhas bloqueantes e acções recomendadas."),
        ("Relatório XLSX", "Memória completa", "O XLSX mantém a auditoria completa: entrada, validação, resultados, shortlists, ELS, V, T, superfície e metadados."),
        ("Falhas", "Gestão de projecto", "Falhas bloqueantes não devem ser aceites sem correcção, revisão da solução ou validação independente."),
    ], columns=["Tema", "Referência", "Nota"])
    return pd.concat([notes, extra], ignore_index=True)

ColumnsEC2App.build_normative_notes = _build_normative_notes_v43


# Garante que a barra lateral e instruções não mencionam software específico.
# A alteração é feita por substituição textual nos métodos já existentes onde possível via notas/exportação;
# os nomes de botões e folhas activos passam a usar entrada de dados / tabela de cálculo.




# ============================================================
# ColumnsEC2 v4.4 — correcção interactiva de falhas
# ============================================================
APP_VERSION = "v4.5"
APP_XLSX_DESCRIPTION = (
    "Workbook de cálculo com entrada de dados, validação, pares member/case, resumo por prumada/nome, "
    "quadro de pilares, ELU, ELS, V, T, pormenorização, superfície resistente, gestão e correcção interactiva "
    "de falhas, shortlists e notas EC2. O PDF é uma síntese executiva; o XLSX conserva a memória completa."
)


def _v44_bool_var(app, name: str, default: bool):
    if not hasattr(app, name):
        setattr(app, name, tk.BooleanVar(value=default))
    return getattr(app, name)


def _v44_str_var(app, name: str, default: str):
    if not hasattr(app, name):
        setattr(app, name, tk.StringVar(value=default))
    return getattr(app, name)


def _as_bool(value) -> bool:
    try:
        return bool(value.get())
    except Exception:
        return bool(value)


def _status_is_unresolved_v44(row) -> bool:
    status = str(row.get("status", "") or "")
    sev = str(row.get("failure_severity", "") or "")
    ftype = str(row.get("failure_type", "") or "")
    if status == "Falha":
        return True
    if sev == "Bloqueante":
        return True
    if ftype in {"armadura_insuficiente", "pormenorizacao", "resistencia_biaxial", "esbelteza"} and status != "OK":
        return True
    return False


def _repair_policy_v44(row) -> str:
    ftype = str(row.get("failure_type", "") or "")
    reason = str(row.get("failure_reason", "") or "").lower()
    if ftype == "dados_incompletos" or "dados" in reason:
        return "dados_nao_corrigiveis_automaticamente"
    if ftype == "armadura_insuficiente":
        return "tentar_mais_armadura"
    if ftype == "pormenorizacao":
        return "tentar_rearranjo_armadura"
    if ftype == "resistencia_biaxial" or "biaxial" in reason:
        return "tentar_modo_rigoroso_e_mais_armadura"
    if ftype == "esbelteza" or "esbelteza" in reason:
        return "tentar_secao_maior_ou_rever_l0"
    return "tentar_recalculo_rigoroso"


def _repair_note_failure_v44(row, attempts: list) -> str:
    base = str(row.get("failure_reason", "") or "Falha não resolvida pelo processo interactivo.")
    att = "; ".join([str(a) for a in attempts if str(a).strip()])
    if att:
        return f"{base} Processo interactivo executado sem solução admissível. Tentativas: {att}."
    return f"{base} Processo interactivo executado sem solução admissível."


def _match_input_row_v44(app, result_row) -> Optional[pd.Series]:
    src = getattr(app, "df_calc_input", None)
    if src is None or src.empty:
        src = getattr(app, "df_pair", None)
    if src is None or src.empty:
        return None
    work = src.copy()
    def s(v): return str(v or "").strip()
    mask = pd.Series(True, index=work.index)
    for c in ["member", "case", "name"]:
        if c in work.columns and c in result_row.index:
            rv = s(result_row.get(c, ""))
            if rv:
                mask &= work[c].astype(str).str.strip().eq(rv)
    cand = work[mask]
    if cand.empty and "member" in work.columns:
        cand = work[work["member"].astype(str).str.strip().eq(s(result_row.get("member", "")))]
    if cand.empty:
        return None
    return cand.iloc[0].copy()


def _make_repair_designer_v44(app, mode: str = "rigoroso") -> ColumnDesigner:
    designer = ColumnDesigner(
        cover_mm=safe_float(app.var_cover.get(), 35.0),
        fyk=safe_float(app.var_fyk.get(), 500.0),
        phi_eff=safe_float(app.var_phi_eff.get(), 2.0),
        l0y_factor=safe_float(app.var_l0y.get(), 1.0),
        l0z_factor=safe_float(app.var_l0z.get(), 1.0),
        calc_mode=mode,
    )
    max_phi = safe_float(getattr(app, "var_repair_max_phi", tk.StringVar(value="32")).get(), 32.0)
    catalogue = [10.0, 12.0, 16.0, 20.0, 25.0, 32.0, 40.0]
    designer.long_diams = [p for p in catalogue if p <= max_phi + 1e-9]
    if not designer.long_diams:
        designer.long_diams = [10.0, 12.0, 16.0, 20.0, 25.0, 32.0]
    designer.stirrup_diams = [8.0, 10.0, 12.0]
    min_s_cm = safe_float(getattr(app, "var_repair_min_st_spacing_cm", tk.StringVar(value="7.5")).get(), 7.5)
    base_spacing = [50.0, 75.0, 100.0, 125.0, 150.0, 200.0, 250.0, 300.0]
    designer.spacing_candidates_mm = [s for s in base_spacing if s >= min_s_cm * 10.0 - 1e-9]
    if not designer.spacing_candidates_mm:
        designer.spacing_candidates_mm = [max(min_s_cm * 10.0, 50.0)]
    extra_bars = int(max(0, safe_float(getattr(app, "var_repair_extra_bars", tk.StringVar(value="0")).get(), 0)))
    base_method = designer.max_bars_per_face
    import types
    def max_bars_per_face_repair(self_obj, b_mm: float, h_mm: float, is_circular: bool = False):
        y, z = base_method(b_mm, h_mm, is_circular=is_circular)
        return max(2, y + extra_bars), max(2, z + extra_bars)
    designer.max_bars_per_face = types.MethodType(max_bars_per_face_repair, designer)
    designer.service_case_override = getattr(app, "var_service_case", tk.StringVar(value="")).get().strip() if hasattr(app, "var_service_case") else ""
    return designer


def _scale_section_row_v44(row: pd.Series, factor: float) -> pd.Series:
    r = row.copy()
    hy = safe_float(r.get("hy", 0.0), 0.0)
    hz = safe_float(r.get("hz", 0.0), 0.0)
    if hy <= 0 or hz <= 0:
        return r
    hy2 = hy * factor
    hz2 = hz * factor
    r["hy"] = hy2
    r["hz"] = hz2
    r["ax"] = hy2 * hz2
    r["iy"] = hy2 * (hz2 ** 3) / 12.0
    r["iz"] = hz2 * (hy2 ** 3) / 12.0
    return r


def _accept_repaired_result_v44(result: dict, original: pd.Series, strategy: str, note: str, section_changed: bool = False) -> dict:
    out = dict(result)
    original_status = str(original.get("status", "") or "")
    original_reason = str(original.get("failure_reason", "") or "")
    out["original_status"] = original_status
    out["original_failure_type"] = str(original.get("failure_type", "") or "")
    out["original_failure_reason"] = original_reason
    out["auto_repair_applied"] = "Sim"
    out["repair_strategy"] = strategy
    out["repair_note"] = note
    out["repair_result"] = "Corrigido" if not section_changed else "Proposta com alteração de secção"
    if section_changed:
        out["status"] = "Aviso"
        out["failure_severity"] = "Aviso"
        out["design_decision"] = "Solução proposta requer alteração da secção no modelo/projecto"
        out["review_priority"] = "Alta"
        out["failure_type"] = "alteracao_secao_proposta"
        out["failure_reason"] = "A solução resistente só foi encontrada com aumento de secção; validar e actualizar a geometria de projecto antes de aceitar."
        out["failure_action"] = "Actualizar a secção no modelo e repetir a análise global; depois recalcular o pilar."
        out["recommendations"] = out["failure_action"]
    else:
        out["status"] = "OK"
        out["failure_severity"] = "OK"
        out["design_decision"] = "Falha corrigida automaticamente por reajuste de armadura/pormenorização dentro dos limites permitidos"
        out["review_priority"] = "Normal"
        out["failure_type"] = "corrigido_interativamente"
        out["failure_reason"] = ""
        out["failure_action"] = "Sem acção correctiva obrigatória; rever a solução proposta antes de emissão."
        out["recommendations"] = "Rever desenho de armaduras e confirmar compatibilidade com o projecto."
    out["failure_summary"] = f"{out.get('failure_severity','')} | {out.get('repair_result','')} | {out.get('repair_strategy','')}"
    return out


def _unresolved_after_repair_v44(original: pd.Series, attempts: list) -> dict:
    out = dict(original)
    out["auto_repair_applied"] = "Não"
    out["repair_strategy"] = _repair_policy_v44(original)
    out["repair_result"] = "Sem solução automática"
    out["repair_note"] = _repair_note_failure_v44(original, attempts)
    if str(out.get("status", "") or "") != "Falha":
        out["status"] = "Aviso"
        out["failure_severity"] = "Aviso"
    else:
        out["failure_severity"] = "Bloqueante"
    out["design_decision"] = out.get("design_decision", "Requer revisão técnica") or "Requer revisão técnica"
    out["failure_action"] = out.get("failure_action", "Rever manualmente a solução") or "Rever manualmente a solução"
    out["failure_summary"] = f"{out.get('failure_severity','')} | {out.get('failure_type','')} | {out.get('repair_note','')}"
    return out


def _try_repair_result_v44(app, result_row: pd.Series) -> dict:
    input_row = _match_input_row_v44(app, result_row)
    attempts = []
    if input_row is None:
        attempts.append("sem correspondência com a linha de cálculo")
        return _unresolved_after_repair_v44(result_row, attempts)
    if _repair_policy_v44(result_row) == "dados_nao_corrigiveis_automaticamente":
        attempts.append("falha de dados não pode ser corrigida por armadura")
        return _unresolved_after_repair_v44(result_row, attempts)

    # Tentativa 1: recalcular com catálogo alargado, mais pontos e modo rigoroso.
    try:
        designer = _make_repair_designer_v44(app, mode="rigoroso")
        repaired = designer.design_one(input_row)
        attempts.append("catálogo alargado + modo rigoroso")
        if str(repaired.get("status", "")) == "OK":
            return _accept_repaired_result_v44(
                repaired, result_row, "catálogo alargado de armaduras/estribos", 
                "Solução encontrada sem alterar a secção.", section_changed=False
            )
    except Exception as err:
        attempts.append(f"catálogo alargado falhou: {err}")

    # Tentativa 2: caso permitido, propor aumento progressivo da secção.
    allow_section = _as_bool(getattr(app, "var_repair_allow_section", False))
    if allow_section:
        max_inc = safe_float(getattr(app, "var_repair_max_section_inc_pct", tk.StringVar(value="15")).get(), 15.0)
        max_factor = 1.0 + max(0.0, max_inc) / 100.0
        factors = []
        f = 1.05
        while f <= max_factor + 1e-9:
            factors.append(round(f, 3)); f += 0.05
        for fac in factors:
            try:
                row2 = _scale_section_row_v44(input_row, fac)
                designer2 = _make_repair_designer_v44(app, mode="rigoroso")
                repaired2 = designer2.design_one(row2)
                attempts.append(f"aumento de secção {int((fac-1)*100)}%")
                if str(repaired2.get("status", "")) == "OK":
                    return _accept_repaired_result_v44(
                        repaired2, result_row,
                        f"proposta de aumento de secção {int((fac-1)*100)}%",
                        "Solução encontrada com alteração geométrica; requer actualização do projecto.",
                        section_changed=True,
                    )
            except Exception as err:
                attempts.append(f"aumento secção {fac:.2f} falhou: {err}")
    else:
        attempts.append("aumento de secção não autorizado")

    return _unresolved_after_repair_v44(result_row, attempts)


# Ajusta a política de falhas para reconhecer Avisos e correcções interactivas.
_old_enrich_failures_v44_base = enrich_failures_v43

def enrich_failures_v43(df: pd.DataFrame) -> pd.DataFrame:
    out = _old_enrich_failures_v44_base(df)
    if out is None or out.empty:
        return out
    if "auto_repair_applied" not in out.columns:
        out["auto_repair_applied"] = ""
    if "repair_result" not in out.columns:
        out["repair_result"] = ""
    mask_auto_ok = out.get("auto_repair_applied", pd.Series(index=out.index, dtype=str)).astype(str).eq("Sim") & out.get("status", pd.Series(index=out.index, dtype=str)).astype(str).eq("OK")
    out.loc[mask_auto_ok, "failure_severity"] = "OK"
    out.loc[mask_auto_ok, "design_decision"] = "Falha corrigida automaticamente por reajuste de armadura/pormenorização"
    out.loc[mask_auto_ok, "review_priority"] = "Normal"
    out.loc[mask_auto_ok, "failure_action"] = "Rever solução proposta antes de emissão."
    mask_warn = out.get("status", pd.Series(index=out.index, dtype=str)).astype(str).eq("Aviso")
    out.loc[mask_warn, "failure_severity"] = "Aviso"
    out.loc[mask_warn, "design_decision"] = out.loc[mask_warn, "design_decision"].replace("", "Aceitável apenas após revisão/actualização do projecto") if "design_decision" in out.columns else "Aceitável apenas após revisão/actualização do projecto"
    out["failure_summary"] = out.apply(
        lambda r: (f"{r.get('failure_severity','')} | {r.get('failure_type','')} | {r.get('failure_action','')}").strip(), axis=1
    )
    return out

# Actualiza a referência global usada pelos métodos já existentes.
globals()["enrich_failures_v43"] = enrich_failures_v43


# Sidebar: acrescenta opções de correcção interactiva e inclui o estado Aviso nos filtros.
_old_build_sidebar_v44_base = ColumnsEC2App._build_sidebar

def _build_sidebar_v44(self, parent):
    _old_build_sidebar_v44_base(self, parent)
    _v44_bool_var(self, "var_auto_repair_on_run", False)
    _v44_bool_var(self, "var_repair_allow_section", False)
    _v44_str_var(self, "var_repair_max_phi", "32")
    _v44_str_var(self, "var_repair_extra_bars", "0")
    _v44_str_var(self, "var_repair_min_st_spacing_cm", "7.5")
    _v44_str_var(self, "var_repair_max_section_inc_pct", "15")

    # Actualiza combobox de estados, se existir.
    def patch_combo(w):
        try:
            if isinstance(w, ttk.Combobox) and str(w.cget("textvariable")) == str(self.var_filter_status):
                w.configure(values=["Todos", "OK", "Aviso", "Falha", "Pré-dimensionado"])
        except Exception:
            pass
        for child in w.winfo_children():
            patch_combo(child)
    patch_combo(parent)

    frame = ttk.LabelFrame(parent, text="8. Correcção interactiva de falhas")
    frame.pack(fill="x", pady=(0, 8))
    ttk.Checkbutton(frame, text="Aplicar automaticamente após cálculo", variable=self.var_auto_repair_on_run).grid(row=0, column=0, columnspan=2, sticky="w", padx=6, pady=3)
    ttk.Label(frame, text="Ø máximo long. [mm]").grid(row=1, column=0, sticky="w", padx=6, pady=3)
    ttk.Entry(frame, textvariable=self.var_repair_max_phi, width=8).grid(row=1, column=1, sticky="ew", padx=6, pady=3)
    ttk.Label(frame, text="Varões extra/face").grid(row=2, column=0, sticky="w", padx=6, pady=3)
    ttk.Entry(frame, textvariable=self.var_repair_extra_bars, width=8).grid(row=2, column=1, sticky="ew", padx=6, pady=3)
    ttk.Label(frame, text="s mín. estribos [cm]").grid(row=3, column=0, sticky="w", padx=6, pady=3)
    ttk.Entry(frame, textvariable=self.var_repair_min_st_spacing_cm, width=8).grid(row=3, column=1, sticky="ew", padx=6, pady=3)
    ttk.Checkbutton(frame, text="Permitir proposta de aumento de secção", variable=self.var_repair_allow_section).grid(row=4, column=0, columnspan=2, sticky="w", padx=6, pady=3)
    ttk.Label(frame, text="Aumento máx. [%]").grid(row=5, column=0, sticky="w", padx=6, pady=3)
    ttk.Entry(frame, textvariable=self.var_repair_max_section_inc_pct, width=8).grid(row=5, column=1, sticky="ew", padx=6, pady=3)
    ttk.Button(frame, text="Corrigir falhas agora", command=self.repair_failures_interactive).grid(row=6, column=0, columnspan=2, sticky="ew", padx=6, pady=(5, 3))
    ttk.Label(frame, text="Se não houver solução, mantém Falha/Aviso com nota justificativa.", style="Subtle.TLabel").grid(row=7, column=0, columnspan=2, sticky="w", padx=6, pady=(0,4))
    frame.columnconfigure(1, weight=1)

ColumnsEC2App._build_sidebar = _build_sidebar_v44


def _refresh_after_repair_v44(self):
    self.df_results = enrich_failures_v43(self.df_results)
    self.df_summary = self.build_summary_by_member(self.df_results) if getattr(self, "var_summary", tk.BooleanVar(value=True)).get() else pd.DataFrame()
    self.df_failures = self.df_results[self.df_results.get("status", pd.Series(dtype=str)).astype(str).eq("Falha")].copy() if self.df_results is not None and not self.df_results.empty else pd.DataFrame()
    self.df_ok = self.df_results[self.df_results.get("status", pd.Series(dtype=str)).astype(str).eq("OK")].copy() if self.df_results is not None and not self.df_results.empty else pd.DataFrame()
    self.df_repair_log = self.build_repair_log_df()
    try:
        self.show_df(self.tree_results, self.df_results)
        self.show_df(self.tree_summary, self.df_summary)
        self.show_df(self.tree_failures, self.df_failures)
        self.show_df(self.tree_shortlists, self.build_shortlists_df())
        self.update_report()
    except Exception:
        pass


def _repair_failures_interactive_v44(self):
    if self.df_results is None or self.df_results.empty:
        messagebox.showwarning("Aviso", "Execute o cálculo antes de corrigir falhas.")
        return
    if getattr(self, "df_calc_input", None) is None or self.df_calc_input.empty:
        messagebox.showwarning("Aviso", "Não existe tabela de cálculo associada aos resultados actuais.")
        return
    res = self.df_results.copy()
    targets = res[res.apply(_status_is_unresolved_v44, axis=1)].copy()
    if targets.empty:
        messagebox.showinfo("Correcção interactiva", "Não há falhas bloqueantes a corrigir.")
        return
    self.status_var.set(f"Correcção interactiva em curso... {len(targets)} casos")
    corrected = []
    corrected_count = 0
    warning_count = 0
    unresolved_count = 0
    for idx, row in res.iterrows():
        if idx not in targets.index:
            r = dict(row)
            if "auto_repair_applied" not in r:
                r["auto_repair_applied"] = ""
            corrected.append(r)
            continue
        repaired = _try_repair_result_v44(self, row)
        repaired["_original_index"] = idx
        if str(repaired.get("auto_repair_applied", "")) == "Sim" and str(repaired.get("status", "")) == "OK":
            corrected_count += 1
        elif str(repaired.get("status", "")) == "Aviso":
            warning_count += 1
        else:
            unresolved_count += 1
        corrected.append(repaired)
    self.df_results = pd.DataFrame(corrected)
    _refresh_after_repair_v44(self)
    self.status_var.set(
        f"Correcção concluída: {corrected_count} corrigidas; {warning_count} avisos/propostas; {unresolved_count} sem solução automática."
    )
    messagebox.showinfo(
        "Correcção interactiva concluída",
        f"Corrigidas: {corrected_count}\nAvisos/propostas: {warning_count}\nSem solução automática: {unresolved_count}",
    )

ColumnsEC2App.repair_failures_interactive = _repair_failures_interactive_v44
ColumnsEC2App._refresh_after_repair = _refresh_after_repair_v44


# Run design: aplica correcção automática quando a opção estiver activa.
_old_run_design_v44_base = ColumnsEC2App.run_design

def _run_design_v44(self):
    _old_run_design_v44_base(self)
    # A execução é assíncrona; a correcção automática é feita no fim através de um watcher simples.
    if not _as_bool(getattr(self, "var_auto_repair_on_run", False)):
        return
    def watcher(tries=0):
        try:
            if getattr(self, "analysis_thread", None) is not None and self.analysis_thread.is_alive() and tries < 600:
                self.after(250, lambda: watcher(tries + 1))
                return
            if self.df_results is not None and not self.df_results.empty:
                self.repair_failures_interactive()
        except Exception as err:
            self.status_var.set(f"Aviso: correcção automática não concluída: {err}")
    self.after(500, watcher)

ColumnsEC2App.run_design = _run_design_v44


def _build_repair_log_df_v44(self) -> pd.DataFrame:
    if self.df_results is None or self.df_results.empty:
        return pd.DataFrame()
    cols = [c for c in [
        "prumada", "member", "case", "name", "status", "original_status", "original_failure_type",
        "auto_repair_applied", "repair_result", "repair_strategy", "repair_note", "solucao",
        "b_cm", "h_cm", "as_req_mm2", "as_prov_mm2", "failure_severity", "design_decision", "failure_action"
    ] if c in self.df_results.columns]
    if not cols:
        return pd.DataFrame()
    return self.df_results[cols].copy()

ColumnsEC2App.build_repair_log_df = _build_repair_log_df_v44


# Exportação XLSX: mantém memória completa e acrescenta a folha de correcções.
_old_write_excel_v44_base = ColumnsEC2App._write_excel

def _write_excel_v44(self, path: str):
    _old_write_excel_v44_base(self, path)
    try:
        repair_log = self.build_repair_log_df()
        with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            repair_log.to_excel(writer, sheet_name="25_Correcoes_Interativas", index=False)
            wb = writer.book
            try:
                props = wb.properties
                props.title = f"{APP_NAME} {APP_VERSION}"
                props.description = APP_XLSX_DESCRIPTION
            except Exception:
                pass
            try:
                from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
                from openpyxl.utils import get_column_letter
                if "25_Correcoes_Interativas" in wb.sheetnames:
                    ws = wb["25_Correcoes_Interativas"]
                    ws.sheet_view.showGridLines = False
                    ws.freeze_panes = "A2"
                    header_fill = PatternFill("solid", fgColor="1F4E5F")
                    header_font = Font(color="FFFFFF", bold=True)
                    thin = Side(style="thin", color="D9E2E7")
                    border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    for cell in ws[1]:
                        cell.fill = header_fill; cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        cell.border = border
                    for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 5000)):
                        for cell in row:
                            cell.border = border; cell.alignment = Alignment(vertical="top", wrap_text=True)
                    for col_idx, col in enumerate(ws.columns, start=1):
                        values = [str(c.value) for c in col[:300] if c.value is not None]
                        ws.column_dimensions[get_column_letter(col_idx)].width = min(max([len(v) for v in values] + [10]) + 2, 60)
            except Exception:
                pass
    except Exception:
        pass

ColumnsEC2App._write_excel = _write_excel_v44


# PDF sintético: acrescenta resumo de correcções sem poluir a memória completa.
_old_write_pdf_v44_base = ColumnsEC2App._write_pdf

def _write_pdf_v44(self, path: str):
    # Usa o PDF sintético existente; o detalhe das correcções fica no XLSX.
    _old_write_pdf_v44_base(self, path)

ColumnsEC2App._write_pdf = _write_pdf_v44


# Relatório interno: inclui estado da correcção interactiva.
_old_update_report_v44_base = ColumnsEC2App.update_report

def _update_report_v44(self):
    _old_update_report_v44_base(self)
    if self.df_results is None or self.df_results.empty:
        return
    try:
        txt = self.report_txt.get("1.0", "end").rstrip() + "\n\n"
        log = self.build_repair_log_df()
        if not log.empty and "repair_result" in log.columns:
            corrected = int((log.get("auto_repair_applied", pd.Series(dtype=str)).astype(str) == "Sim").sum())
            proposed = int((log.get("repair_result", pd.Series(dtype=str)).astype(str) == "Proposta com alteração de secção").sum())
            unresolved = int((log.get("repair_result", pd.Series(dtype=str)).astype(str) == "Sem solução automática").sum())
            txt += "Correcção interactiva de falhas\n"
            txt += f"Corrigidas automaticamente: {corrected}\n"
            txt += f"Propostas com alteração de secção: {proposed}\n"
            txt += f"Sem solução automática: {unresolved}\n"
            txt += "Detalhe completo na folha 25_Correcoes_Interativas do XLSX.\n"
        self.report_txt.delete("1.0", "end")
        self.report_txt.insert("1.0", txt)
    except Exception:
        pass

ColumnsEC2App.update_report = _update_report_v44


# Notas normativas e de uso.
_old_build_normative_notes_v44_base = ColumnsEC2App.build_normative_notes

def _build_normative_notes_v44(self) -> pd.DataFrame:
    notes = _old_build_normative_notes_v44_base(self).copy()
    extra = pd.DataFrame([
        ("Correcção interactiva", "Armadura", "O programa tenta resolver falhas por armadura/pormenorização aumentando o catálogo permitido, reduzindo espaçamentos e usando modo rigoroso."),
        ("Correcção interactiva", "Secção", "Aumento de secção só é proposto quando autorizado; nesse caso o resultado é Aviso e exige actualização do projecto/modelo."),
        ("Correcção interactiva", "Limitação", "Falhas de dados não são corrigidas automaticamente; a tabela de entrada deve ser corrigida."),
        ("Relatório", "PDF/XLSX", "O PDF mantém síntese executiva; a folha 25_Correcoes_Interativas do XLSX contém o detalhe das tentativas de correcção."),
    ], columns=["Tema", "Referência", "Nota"])
    return pd.concat([notes, extra], ignore_index=True)

ColumnsEC2App.build_normative_notes = _build_normative_notes_v44



# ============================================================
# ColumnsEC2 v4.5 — materiais obrigatórios, Es=210 GPa,
# envolvente ELU, shortlist inteligente e correcção com progresso
# ============================================================
