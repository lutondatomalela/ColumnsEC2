# -*- coding: utf-8 -*-
# Auto-split from ColumnsEC2 v0.9 RC8.
# This module is executed in the shared runtime namespace by columns_ec2.runtime.loader.
# Keep execution order defined in columns_ec2/runtime/manifest.py.

# ============================================================
# ColumnsEC2 v0.9 RC7 — technical EN-UK result localisation
# and repository hyperlinks embedded in the programme name
# ============================================================
APP_VERSION = "v0.9 RC7"

_RC7_REPO_URL = globals().get("GITHUB_URL", "https://github.com/lutondatomalela/ColumnsEC2")

_RC7_HEADER_MAP = {
    "member": "Member",
    "case": "Case",
    "name": "Column line",
    "story": "Storey",
    "Story": "Storey",
    "Piso": "Storey",
    "piso": "Storey",
    "Prumada": "Column line",
    "prumada": "Column line",
    "Tramo": "Segment",
    "tramo": "Segment",
    "Secção [cm]": "Section [cm]",
    "section_cm": "Section [cm]",
    "material": "Concrete grade",
    "status": "Status",
    "Estado": "Status",
    "estado_global": "Overall status",
    "estado_resistente": "Resistance status",
    "estado_corte": "Shear status",
    "estado_torcao": "Torsion status",
    "estado_els": "SLS status",
    "estado_pormenorizacao": "Detailing status",
    "decisao_tecnica": "Technical decision",
    "failure_reason": "Failure reason",
    "failure_type": "Failure type",
    "recommendations": "Recommendations",
    "shortlist_text": "Shortlist",
    "solucao": "Reinforcement solution",
    "solucao_completa": "Full reinforcement solution",
    "pormenorizacao_construtiva": "Constructive detailing",
    "detalhe_grampos": "Cross-tie detailing",
    "grampos_intermedios": "Intermediate cross-ties",
    "numero_grampos_por_nivel": "Cross-ties per level",
    "ramos_estribo_y": "Link legs in y",
    "ramos_estribo_z": "Link legs in z",
    "as_req_mm2": "As,req [mm²]",
    "as_prov_mm2": "As,prov [mm²]",
    "as_min_mm2": "As,min [mm²]",
    "as_max_mm2": "As,max [mm²]",
    "n_ed_kN": "N_Ed [kN]",
    "my_ed_kNm": "My,Ed [kNm]",
    "mz_ed_kNm": "Mz,Ed [kNm]",
    "mrd_y_kNm": "MRd,y [kNm]",
    "mrd_z_kNm": "MRd,z [kNm]",
    "utilizacao": "η_NMyMz",
    "phi_long_mm": "Longitudinal bar diameter [mm]",
    "phi_st_mm": "Link diameter [mm]",
    "s_st_mm": "Link spacing [mm]",
    "s_st_max_mm": "Maximum link spacing [mm]",
    "n_total": "Total number of bars",
    "n_bars_y": "Bars on y-faces",
    "n_bars_z": "Bars on z-faces",
}

_RC7_EXACT_VALUE_MAP = {
    "OK": "OK",
    "Aviso": "Warning",
    "Falha": "Failure",
    "Verificar": "Check required",
    "Não conforme": "Not compliant",
    "Pré-dimensionado": "Preliminary sizing",
    "Pré-dimensionamento": "Preliminary sizing",
    "Dimensionamento": "Design",
    "Dispensada": "Waived",
    "Sim": "Yes",
    "Não": "No",
    "Bloqueante": "Blocking design issue",
    "Não bloqueante": "Non-blocking item",
    "Económica": "Minimum reinforcement",
    "Equilibrada": "Balanced",
    "Robusta": "Robust",
    "Relatório técnico": "Technical report",
    "Resumo executivo": "Executive summary",
    "Memória de cálculo": "Detailed calculation note",
    "Sem aviso relevante": "No relevant warning",
    "Sem torção relevante": "No relevant torsion",
    "Torção desprezável — não condicionante": "Negligible torsion — not governing",
    "OK sem armadura transversal resistente adicional": "OK without additional shear links",
    "Requer armadura de esforço transverso": "Shear links required",
    "Requer armadura de torção": "Torsion reinforcement required",
    "Não avaliado neste backend structuralcodes pelo ColumnsEC2; sem fallback interno.": "Not assessed in this structuralcodes backend by ColumnsEC2; no internal fallback has been used.",
}

_RC7_TECHNICAL_REPLACEMENTS = [
    ("Falhas bloqueantes detectadas", "Blocking design issues detected"),
    ("Foram detectadas falhas bloqueantes.", "Blocking design issues were detected."),
    ("Pretende gerar propostas de correcção?", "Generate correction proposals?"),
    ("pré-dimensionamento: verificar em modo Dimensionamento antes de adoptar", "preliminary sizing only: run Design mode before adopting the reinforcement"),
    ("pre-dimensionamento: verificar em modo Dimensionamento antes de adoptar", "preliminary sizing only: run Design mode before adopting the reinforcement"),
    ("verificar em modo Dimensionamento antes de adoptar", "run Design mode before adopting the reinforcement"),
    ("verificar em modo Design antes de adoptar", "run Design mode before adopting the reinforcement"),
    ("antes de adoptar", "before adopting the reinforcement"),
    ("Falha de resistência biaxial", "N-My-Mz resistance failure"),
    ("Falha de pormenorização", "Detailing failure"),
    ("Falha por insuficiência de armadura", "Insufficient reinforcement"),
    ("Falha em verificações complementares", "Failure in complementary checks"),
    ("Falha de dados", "Input-data failure"),
    ("sem os dois nós necessários", "without the two required end nodes"),
    ("nenhuma solução", "no solution"),
    ("nenhuma disposição admissível", "no admissible reinforcement layout"),
    ("disposição admissível", "admissible reinforcement layout"),
    ("interação biaxial", "N-My-Mz interaction"),
    ("interacção biaxial", "N-My-Mz interaction"),
    ("interação N-My-Mz", "N-My-Mz interaction"),
    ("interacção N-My-Mz", "N-My-Mz interaction"),
    ("pormenorização construtiva", "constructive detailing"),
    ("pormenorização", "detailing"),
    ("esforço transverso", "shear"),
    ("esforços transversos", "shear forces"),
    ("torção", "torsion"),
    ("fendilhação", "crack control"),
    ("fluência", "creep"),
    ("retracção", "shrinkage"),
    ("retração", "shrinkage"),
    ("armadura longitudinal", "longitudinal reinforcement"),
    ("armadura transversal", "transverse reinforcement"),
    ("armadura de esforço transverso", "shear links"),
    ("armadura de torção", "torsion reinforcement"),
    ("estribos", "links"),
    ("Estribos", "Links"),
    ("grampos intermédios", "intermediate cross-ties"),
    ("grampos", "cross-ties"),
    ("Grampos", "Cross-ties"),
    ("varões comprimidos", "compression bars"),
    ("varões", "bars"),
    ("Varões", "Bars"),
    ("nos cantos", "at the corners"),
    ("cantos", "corners"),
    ("faces longas", "long faces"),
    ("faces curtas", "short faces"),
    ("a meio das faces", "at mid-face"),
    ("por nível", "per level"),
    ("Ramos", "Legs"),
    ("ramos", "legs"),
    ("Solução", "Reinforcement solution"),
    ("Estado resistente", "Resistance status"),
    ("Estado global", "Overall status"),
    ("Decisão técnica", "Technical decision"),
    ("Prumada", "Column line"),
    ("prumada", "column line"),
    ("Piso", "Storey"),
    ("piso", "storey"),
    ("Tramo", "Segment"),
    ("tramo", "segment"),
    ("caso", "case"),
    ("Caso", "Case"),
    ("Bloqueante", "Blocking design issue"),
    ("bloqueante", "blocking design issue"),
    ("Aviso", "Warning"),
    ("avisos", "warnings"),
    ("Falha", "Failure"),
    ("falhas", "failures"),
    ("Não avaliado", "Not assessed"),
    ("não avaliado", "not assessed"),
    ("não calculado", "not calculated"),
    ("não exposto", "not exposed"),
    ("não exposta", "not exposed"),
    ("não gerada", "not generated"),
    ("sem fallback interno", "no internal fallback"),
    ("método", "method"),
    ("Método", "Method"),
    ("concluído", "complete"),
    ("Concluído", "Complete"),
    ("A calcular", "Calculating"),
    ("a calcular", "calculating"),
    ("casos de envolvente", "governing cases"),
    ("member/case sem dois nós", "member/case pair without two end nodes"),
    ("Betão lido da coluna Material", "Concrete grade read from the Material column"),
    ("betão read from the Material column", "concrete grade read from the Material column"),
    ("betão", "concrete grade"),
    ("Betão", "Concrete"),
]


def _rc7_lang(app):
    try:
        return LANG_EN if str(app.var_language.get()).upper().startswith("EN") else LANG_PT
    except Exception:
        return LANG_PT


def _rc7_is_en(app):
    return _rc7_lang(app) == LANG_EN


def _rc7_translate_technical_en(value):
    try:
        if _rc3_is_scalar_nan(value):
            return value
    except Exception:
        pass
    if not isinstance(value, str):
        return value
    s = value
    stripped = s.strip()
    if stripped in _RC7_EXACT_VALUE_MAP:
        return _RC7_EXACT_VALUE_MAP[stripped]
    try:
        s = _rc6_technical_en_text(s)
    except Exception:
        try:
            s = _v69_status_text(s, LANG_EN)
        except Exception:
            pass
    if s.strip() in _RC7_EXACT_VALUE_MAP:
        return _RC7_EXACT_VALUE_MAP[s.strip()]
    for old, new in sorted(_RC7_TECHNICAL_REPLACEMENTS, key=lambda kv: len(kv[0]), reverse=True):
        s = s.replace(old, new)
    # Clean mixed-language artefacts produced by older fragment translators.
    s = s.replace("Designamento", "Design")
    s = s.replace("reinforcement strategy equilibrada", "balanced reinforcement strategy")
    s = s.replace("Estratégia de reinforcement", "Reinforcement strategy")
    s = s.replace("Diagnóstico and auditoria", "Diagnostics and audit")
    s = s.replace("Instruções de utilização and tabela tipo", "User instructions and input table format")
    s = s.replace("concrete grade read from the Material column lido da coluna Material", "concrete grade read from the Material column")
    s = s.replace("case s", "cases")
    s = s.replace("Failure s", "Failures")
    return s


def _rc7_translate_header(col):
    s = str(col)
    if s in _RC7_HEADER_MAP:
        return _RC7_HEADER_MAP[s]
    try:
        if s in _V69_HEADER_MAP_PT_EN:
            return _V69_HEADER_MAP_PT_EN[s]
    except Exception:
        pass
    return _rc7_translate_technical_en(s)


def _rc7_prepare_display_df(df, lang=LANG_EN):
    if df is None or getattr(df, "empty", True):
        return pd.DataFrame()
    out = df.copy()
    if lang != LANG_EN:
        return out
    try:
        out = _rc3_deduplicate_columns(out) if "_rc3_deduplicate_columns" in globals() else out
    except Exception:
        pass
    out.columns = [_rc7_translate_header(c) for c in out.columns]
    for c in list(out.columns):
        try:
            if out[c].dtype == object or str(out[c].dtype).startswith("string"):
                out[c] = out[c].map(_rc7_translate_technical_en)
        except Exception:
            pass
    # Status-like and decision columns often hold PT text even when dtype is object-like.
    for c in list(out.columns):
        cname = str(c).lower()
        if any(k in cname for k in ["status", "decision", "reason", "recommend", "solution", "detailing", "shortlist", "scope", "note", "criterion", "failure"]):
            try:
                out[c] = out[c].map(_rc7_translate_technical_en)
            except Exception:
                pass
    return out


# Improve the treeview display only; internal dataframes remain unchanged for calculation.
_rc7_prev_show_df = getattr(ColumnsEC2App, "show_df", None)
def _rc7_show_df(self, tree, df):
    global _RC6_ACTIVE_APP
    _RC6_ACTIVE_APP = self
    display_df = df
    try:
        if _rc7_is_en(self):
            display_df = _rc7_prepare_display_df(df, LANG_EN)
    except Exception:
        display_df = df
    if callable(_rc7_prev_show_df):
        out = _rc7_prev_show_df(self, tree, display_df)
    else:
        out = None
    try:
        if _rc7_is_en(self):
            _rc6_translate_tree_headings(self)
    except Exception:
        pass
    return out
ColumnsEC2App.show_df = _rc7_show_df


# Re-translate report tab content after update_report.
_rc7_prev_update_report = getattr(ColumnsEC2App, "update_report", None)
def _rc7_update_report(self):
    out = _rc7_prev_update_report(self) if callable(_rc7_prev_update_report) else None
    try:
        if _rc7_is_en(self) and hasattr(self, "report_txt"):
            txt = self.report_txt.get("1.0", "end")
            new = _rc7_translate_technical_en(txt)
            if new != txt:
                self.report_txt.delete("1.0", "end")
                self.report_txt.insert("1.0", new)
    except Exception:
        pass
    return out
ColumnsEC2App.update_report = _rc7_update_report


def _rc7_bind_repository_links(app):
    """Make the programme name clickable wherever it appears in the GUI."""
    try:
        widgets = list(_rc5_walk_widgets(app)) if "_rc5_walk_widgets" in globals() else list(_rc3_walk_widgets(app))
    except Exception:
        widgets = []
    for w in widgets:
        try:
            txt = str(w.cget("text"))
        except Exception:
            continue
        if "ColumnsEC2" in txt or "Columns EC2" in txt:
            try:
                w.configure(cursor="hand2")
            except Exception:
                pass
            try:
                w.bind("<Button-1>", lambda _e, url=_RC7_REPO_URL: webbrowser.open_new(url), add="+")
            except Exception:
                pass


# Excel: hyperlink the programme name and translate remaining PT values in EN-UK workbooks.
_rc7_prev_write_excel = getattr(ColumnsEC2App, "_write_excel", None)
def _write_excel_rc7(self, path: str):
    if callable(_rc7_prev_write_excel):
        _rc7_prev_write_excel(self, path)
    if not _rc7_is_en(self):
        # Still embed the repository link in PT workbooks.
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path)
            _rc7_apply_workbook_links(wb)
            wb.save(path)
        except Exception:
            pass
        return
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path)
        for ws in wb.worksheets:
            max_rows = min(ws.max_row, 5000)
            max_cols = min(ws.max_column, 120)
            for row in ws.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols):
                for cell in row:
                    if isinstance(cell.value, str):
                        # Header first; value translation afterwards.
                        original = cell.value
                        translated = _rc7_translate_header(original) if cell.row == 1 else _rc7_translate_technical_en(original)
                        cell.value = translated
        _rc7_apply_workbook_links(wb)
        try:
            wb.properties.title = "ColumnsEC2 - Reinforced Concrete Column Design"
            wb.properties.subject = "Reinforced concrete column design and checking according to EC2"
            wb.properties.description = "Technical workbook generated by ColumnsEC2. The programme name contains the repository hyperlink."
        except Exception:
            pass
        wb.save(path)
    except Exception:
        pass
ColumnsEC2App._write_excel = _write_excel_rc7


def _rc7_apply_workbook_links(wb):
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 80), max_col=min(ws.max_column, 20)):
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.strip() in ["ColumnsEC2", "Columns EC2"]:
                        cell.hyperlink = _RC7_REPO_URL
                        try:
                            cell.style = "Hyperlink"
                        except Exception:
                            pass
    except Exception:
        pass


# PDF report: technical EN-UK display frames and repository hyperlink embedded in the programme name.
def _write_pdf_rc7(self, path: str):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak
    import tempfile, shutil
    lang = _rc7_lang(self)
    styles = _v68_pdf_styles() if "_v68_pdf_styles" in globals() else _pdf_styles_v3()
    level = _rc3_get_var(self, "var_pdf_level", "Relatório técnico" if lang == LANG_PT else "Technical report")
    if lang == LANG_EN:
        level = _rc6_pdf_level_display(level, LANG_EN) if "_rc6_pdf_level_display" in globals() else level
    else:
        level = _rc6_pdf_level_display(level, LANG_PT) if "_rc6_pdf_level_display" in globals() else level

    res = getattr(self, "df_results", pd.DataFrame())
    if res is None:
        res = pd.DataFrame()
    try:
        if "_v68_apply_constructive_detailing" in globals() and not res.empty:
            res = _v68_apply_constructive_detailing(res)
    except Exception:
        pass

    summ_src = getattr(self, "df_summary", pd.DataFrame())
    if summ_src is None or summ_src.empty:
        summ_src = res
    try:
        if "_v683_build_tramo_schedule" in globals():
            summ = _v683_build_tramo_schedule(summ_src)
        elif "_v681_decision_by_prumada" in globals():
            summ = _v681_decision_by_prumada(summ_src)
        else:
            summ = summ_src
    except Exception:
        summ = summ_src
    try:
        inter = _v68_interaction_summary(summ_src) if "_v68_interaction_summary" in globals() else pd.DataFrame()
    except Exception:
        inter = pd.DataFrame()
    try:
        module = _v65_module_status_table(res) if "_v65_module_status_table" in globals() and not res.empty else pd.DataFrame()
    except Exception:
        module = pd.DataFrame()
    try:
        gov = _v681_governing_cases_for_pdf(self) if "_v681_governing_cases_for_pdf" in globals() else getattr(self, "df_calc_input", pd.DataFrame())
    except Exception:
        gov = getattr(self, "df_calc_input", pd.DataFrame())
    try:
        perf = _v68_performance_df(self) if "_v68_performance_df" in globals() else pd.DataFrame()
    except Exception:
        perf = pd.DataFrame()

    summ_d = _rc7_prepare_display_df(summ, lang)
    inter_d = _rc7_prepare_display_df(inter, lang)
    module_d = _rc7_prepare_display_df(module, lang)
    gov_d = _rc7_prepare_display_df(gov, lang)
    perf_d = _rc7_prepare_display_df(perf, lang)

    n_total = len(res)
    st_series = _rc3_col_series(res, ["estado_global", "status", "Estado", "Status"]).astype(str) if "_rc3_col_series" in globals() else pd.Series(dtype=str)
    n_fail = int((st_series == "Falha").sum()) if not st_series.empty else 0
    n_warn = int((st_series == "Aviso").sum()) if not st_series.empty else 0
    pr_series = _rc3_col_series(summ_d, ["Column line", "Prumada", "prumada", "name", "Name"]).astype(str) if "_rc3_col_series" in globals() else pd.Series(dtype=str)
    n_pr = int(pr_series.nunique()) if not pr_series.empty else 0

    if lang == LANG_EN:
        subtitle = "Technical report for reinforced concrete column design and checking"
        section_decision = "1. Design decision by column line and segment"
        section_inter = "2. N-My-Mz resistance interaction"
        section_module = "3. Design-module status"
        section_gov = "4. Governing design cases"
        section_perf = "5. Calculation performance"
        meta = pd.DataFrame([
            {"Field": "Reference standard / calculation engine", "Value": _v59_norm_reference(self) if "_v59_norm_reference" in globals() else "Eurocode 2"},
            {"Field": "Report level", "Value": _rc7_translate_technical_en(level)},
            {"Field": "Reinforcement strategy", "Value": _rc7_translate_technical_en(_rc3_get_var(self, "var_rebar_strategy", "Balanced"))},
            {"Field": "Column lines", "Value": n_pr},
            {"Field": "Analysed cases", "Value": n_total},
            {"Field": "Failures / Warnings", "Value": f"{n_fail} / {n_warn}"},
        ])
        meta_cols = ["Field", "Value"]
        decision_cols = ["Column line", "Storey", "Section [cm]", "Concrete grade", "N_Ed [kN]", "My,Ed [kNm]", "Mz,Ed [kNm]", "η_NMyMz", "Reinforcement solution", "Status"]
        inter_cols = ["Column line", "Case", "N_Ed [kN]", "My_Ed [kNm]", "Mz_Ed [kNm]", "MRd_y [kNm]", "MRd_z [kNm]", "η_NMyMz", "Resistance status"]
        mod_cols = ["Column line", "case", "Overall status", "Resistance status", "Shear status", "Torsion status", "SLS status", "Detailing status", "Technical decision"]
        gov_cols = ["Column line", "Case", "Criterion", "NEd [kN]", "My [kNm]", "Mz [kNm]", "Vy [kN]", "Vz [kN]", "T [kNm]"]
        perf_cols = ["Item", "Value", "Note"]
        page_word = "Page"
    else:
        subtitle = "Relatório técnico de dimensionamento/verificação de pilares de betão armado"
        section_decision = "1. Decisão por prumada / tramo"
        section_inter = "2. Interacção N-My-Mz"
        section_module = "3. Estados por módulo"
        section_gov = "4. Casos governantes"
        section_perf = "5. Performance"
        meta = pd.DataFrame([
            {"Campo": "Norma/Motor", "Valor": _v59_norm_reference(self) if "_v59_norm_reference" in globals() else "Eurocódigo 2"},
            {"Campo": "Nível do relatório", "Valor": level},
            {"Campo": "Estratégia de armadura", "Valor": _rc3_get_var(self, "var_rebar_strategy", "Equilibrada")},
            {"Campo": "Prumadas", "Valor": n_pr},
            {"Campo": "Casos analisados", "Valor": n_total},
            {"Campo": "Falhas / Avisos", "Valor": f"{n_fail} / {n_warn}"},
        ])
        meta_cols = ["Campo", "Valor"]
        decision_cols = ["Prumada", "Piso", "Secção [cm]", "material", "N_Ed_kN", "My_Ed_kNm", "Mz_Ed_kNm", "η_NMyMz", "Solução", "Estado"]
        inter_cols = ["Prumada", "Case", "N_Ed [kN]", "My_Ed [kNm]", "Mz_Ed [kNm]", "MRd_y [kNm]", "MRd_z [kNm]", "η_NMyMz", "Estado resistente"]
        mod_cols = ["Prumada", "case", "estado_global", "estado_resistente", "estado_corte", "estado_torcao", "estado_els", "estado_pormenorizacao", "decisao_tecnica"]
        gov_cols = ["Prumada", "Case", "Critério", "NEd [kN]", "My [kNm]", "Mz [kNm]", "Vy [kN]", "Vz [kN]", "T [kNm]"]
        perf_cols = ["Item", "Valor", "Nota"]
        page_word = "Página"

    tmp_path = path
    try:
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        tmp_path = tmp.name
        tmp.close()
    except Exception:
        pass
    doc = SimpleDocTemplate(tmp_path, pagesize=landscape(A4), rightMargin=12*mm, leftMargin=12*mm, topMargin=12*mm, bottomMargin=12*mm)
    doc.title = "ColumnsEC2"
    doc.author = APP_AUTHOR
    doc.subject = "Reinforced concrete column design" if lang == LANG_EN else APP_SUBJECT
    linked_title = f'<link href="{_RC7_REPO_URL}">ColumnsEC2</link>'
    story = [Paragraph(linked_title, styles["T68"]), Paragraph(subtitle, styles["B68"]), Spacer(1, 4*mm)]
    story.append(_rc3_pdf_table(meta, meta_cols, styles["C68"], max_rows=10, lang=lang))
    story.append(Paragraph(section_decision, styles["H68"]))
    story.append(_rc3_pdf_table(summ_d, decision_cols, styles["C68"], max_rows=100, lang=lang))
    if level in ["Relatório técnico", "Memória de cálculo", "Technical report", "Detailed calculation note"]:
        story.append(PageBreak())
        story.append(Paragraph(section_inter, styles["H68"]))
        story.append(_rc3_pdf_table(inter_d, inter_cols, styles["C68"], max_rows=100, lang=lang))
        story.append(Paragraph(section_module, styles["H68"]))
        story.append(_rc3_pdf_table(module_d, mod_cols, styles["C68"], max_rows=100, lang=lang))
    if level in ["Memória de cálculo", "Detailed calculation note"]:
        story.append(PageBreak())
        story.append(Paragraph(section_gov, styles["H68"]))
        story.append(_rc3_pdf_table(gov_d, gov_cols, styles["C68"], max_rows=120, lang=lang))
        story.append(PageBreak())
        story.append(Paragraph(section_perf, styles["H68"]))
        story.append(_rc3_pdf_table(perf_d, perf_cols, styles["C68"], max_rows=80, lang=lang))

    def footer(canvas, doc_obj):
        canvas.saveState()
        canvas.setFont("Courier", 7)
        canvas.setFillColor(colors.grey)
        x0, y0 = 12*mm, 7*mm
        canvas.drawString(x0, y0, f"ColumnsEC2 | {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        try:
            canvas.linkURL(_RC7_REPO_URL, (x0, y0 - 1*mm, x0 + 32*mm, y0 + 4*mm), relative=0, thickness=0)
        except Exception:
            pass
        canvas.drawRightString(285*mm, 7*mm, f"{page_word} {doc_obj.page}")
        canvas.restoreState()

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    if tmp_path != path:
        try:
            shutil.move(tmp_path, path)
        except Exception:
            root, ext = os.path.splitext(path)
            alt = root + ("_new" if lang == LANG_EN else "_novo") + ext
            shutil.move(tmp_path, alt)

ColumnsEC2App._write_pdf = _write_pdf_rc7


# Message boxes/status: route through RC7 technical wording.
_rc7_prev_showinfo = messagebox.showinfo
_rc7_prev_showwarning = messagebox.showwarning
_rc7_prev_showerror = messagebox.showerror
_rc7_prev_askyesno = messagebox.askyesno

def _rc7_msg(title, message):
    app = globals().get("_RC6_ACTIVE_APP", None)
    if app is not None and _rc7_is_en(app):
        return _rc7_translate_technical_en(title), _rc7_translate_technical_en("" if message is None else str(message))
    return title, message

def _rc7_showinfo(title, message=None, *args, **kwargs):
    t, m = _rc7_msg(str(title), message)
    return _rc7_prev_showinfo(t, m, *args, **kwargs)

def _rc7_showwarning(title, message=None, *args, **kwargs):
    t, m = _rc7_msg(str(title), message)
    return _rc7_prev_showwarning(t, m, *args, **kwargs)

def _rc7_showerror(title, message=None, *args, **kwargs):
    t, m = _rc7_msg(str(title), message)
    return _rc7_prev_showerror(t, m, *args, **kwargs)

def _rc7_askyesno(title, message=None, *args, **kwargs):
    t, m = _rc7_msg(str(title), message)
    return _rc7_prev_askyesno(t, m, *args, **kwargs)

messagebox.showinfo = _rc7_showinfo
messagebox.showwarning = _rc7_showwarning
messagebox.showerror = _rc7_showerror
messagebox.askyesno = _rc7_askyesno


# Apply language: bind repository links and force technical EN-UK updates.
_rc7_prev_apply_language = getattr(ColumnsEC2App, "apply_language", None)
def _rc7_apply_language(self):
    global _RC6_ACTIVE_APP
    _RC6_ACTIVE_APP = self
    out = _rc7_prev_apply_language(self) if callable(_rc7_prev_apply_language) else None
    try:
        _rc7_bind_repository_links(self)
    except Exception:
        pass
    try:
        if _rc7_is_en(self):
            # Force result tables already visible to use technical English.
            for attr, df_attr in [
                ("tree_results", "df_results"), ("tree_summary", "df_summary"), ("tree_failures", "df_failures"),
                ("tree_shortlists", None), ("tree_validation", "df_validation"), ("tree_notes", "df_notes")
            ]:
                tree = getattr(self, attr, None)
                if tree is None:
                    continue
                if df_attr is None:
                    df = self.build_shortlists_df() if hasattr(self, "build_shortlists_df") else pd.DataFrame()
                else:
                    df = getattr(self, df_attr, pd.DataFrame())
                if df is not None and not getattr(df, "empty", True):
                    _rc7_show_df(self, tree, df)
            try:
                old = self.status_var.get()
                self.status_var.set(_rc7_translate_technical_en(old))
            except Exception:
                pass
        _rc7_bind_repository_links(self)
    except Exception:
        pass
    return out
ColumnsEC2App.apply_language = _rc7_apply_language
try:
    _rc3_apply_language = _rc7_apply_language
except Exception:
    pass


# Keep active context and translate visible results when calculation finishes.
_rc7_prev_run_design = getattr(ColumnsEC2App, "run_design", None)
def _rc7_run_design(self):
    global _RC6_ACTIVE_APP
    _RC6_ACTIVE_APP = self
    result = _rc7_prev_run_design(self) if callable(_rc7_prev_run_design) else None
    try:
        self.after(300, lambda: self.apply_language() if _rc7_is_en(self) else None)
        self.after(1200, lambda: self.apply_language() if _rc7_is_en(self) else None)
    except Exception:
        pass
    return result
ColumnsEC2App.run_design = _rc7_run_design


# Initial language/title hook.
def _v091_apply_language_title(app):
    global _RC6_ACTIVE_APP
    _RC6_ACTIVE_APP = app
    try:
        _rc5_install_live_localisation(app)
    except Exception:
        pass
    try:
        app.apply_language()
    except Exception:
        pass
    try:
        _rc7_bind_repository_links(app)
    except Exception:
        pass
    try:
        app.title("ColumnsEC2 - Reinforced Concrete Column Design (EC2)" if _rc7_is_en(app) else APP_TITLE)
    except Exception:
        pass


# ============================================================
# ColumnsEC2 v0.9 RC8 — full EN-UK technical presentation layer
# ============================================================
APP_VERSION = "v0.9 RC8"

_RC8_REPO_URL = globals().get("GITHUB_URL", "https://github.com/lutondatomalela/ColumnsEC2")

_RC8_HEADER_MAP = {
    # generic table headings
    "Grupo": "Group",
    "Objeto": "Object",
    "Objecto": "Object",
    "Estado": "Status",
    "Detalhe": "Details",
    "Verificação": "Design check",
    "Origem": "Source",
    "Nota": "Engineering note",
    "Categoria": "Category",
    "Resultado": "Result",
    "Item": "Item",
    "Backend": "Calculation engine",
    "Campo": "Field",
    "Valor": "Value",
    "Critério": "Criterion",
    "Prumada": "Column line",
    "Piso": "Storey",
    "Membro": "Member",
    "Caso": "Load case",
    "Solução": "Reinforcement arrangement",
    "Decisão técnica": "Engineering decision",
    "Falha": "Design issue",
    "Falhas": "Design issues",
    "Avisos": "Warnings",
    "Recomendações": "Engineering recommendations",
    "shortlist_text": "Candidate layouts",
    "Shortlist": "Candidate layouts",
    "Technical decision": "Engineering decision",
    # common result aliases from older versions
    "estado_global": "Overall status",
    "estado_resistente": "Resistance status",
    "estado_corte": "Shear status",
    "estado_torcao": "Torsion status",
    "estado_els": "SLS status",
    "estado_pormenorizacao": "Detailing status",
    "decisao_tecnica": "Engineering decision",
    "solucao": "Reinforcement arrangement",
    "pormenorizacao_construtiva": "Detailing arrangement",
    "detalhe_grampos": "Cross-tie arrangement",
}

_RC8_EXACT_VALUE_MAP = {
    "Pacote": "Package",
    "Dependência": "Dependency",
    "Módulo": "Module",
    "Disponível": "Available",
    "Indisponível": "Unavailable",
    "desconhecida": "unknown",
    "Materiais": "Materials",
    "Material": "Concrete grade",
    "2.ª ordem": "Second-order effects",
    "2a ordem": "Second-order effects",
    "N-My-Mz": "N-My-Mz interaction",
    "Esforço transverso": "Shear",
    "Torção": "Torsion",
    "ELS": "SLS",
    "ELS/fendilhação": "SLS crack control",
    "ELS/fendilhação/deformação": "SLS crack control/deflection",
    "Pormenorização": "Detailing",
    "ColumnsEC2 geométrico": "ColumnsEC2 geometry layer",
    "Calculado": "Calculated",
    "Calculado se API disponível": "Calculated where the API is available",
    "Não avaliado se API ausente": "Not assessed when the API is unavailable",
    "Informativo": "Informative",
    "Sem fallback interno.": "No internal fallback is used.",
    "Motor interno existente; default.": "Existing internal engine; default calculation engine.",
    "Método interno actual.": "Current internal method.",
    "Superfície interna discreta.": "Internal discrete N-My-Mz resistance surface.",
    "Verificação interna actual.": "Current internal check.",
    "Combinação indicada ou simplificada.": "User-defined or simplified SLS combination.",
    "Regras internas de pormenorização.": "Internal detailing rules.",
    "Colunas obrigatórias": "Required input fields",
    "Tabela": "Input table",
    "Dados": "Input data",
    "Pares de nós": "End-node pairs",
    "Consistência entre nós": "End-node data consistency",
    "Unidades": "Units",
    "classe de betão": "concrete grade",
    "linhas importadas": "imported rows",
    "barras distintas": "unique members",
    "member vazio": "blank member field",
    "case vazio": "blank load-case field",
    "com 2 nós": "with two end nodes",
    "com 1 nó": "with one end node",
    "com mais de 2 nós": "with more than two end nodes",
    "presente": "present",
    "OK": "OK",
    "Aviso": "Warning",
    "Falha": "Failure",
    "Pré-dimensionamento": "Preliminary sizing",
    "Pré-dimensionado": "Preliminary sizing",
    "Dimensionamento": "Design",
    "Económica": "Minimum reinforcement",
    "Equilibrada": "Balanced",
    "Robusta": "Robust",
    "Relatório técnico": "Technical report",
    "Resumo executivo": "Executive summary",
    "Memória de cálculo": "Detailed calculation note",
    "Memória detalhada": "Detailed calculation note",
    "Sim": "Yes",
    "Não": "No",
}

_RC8_TECHNICAL_REPLACEMENTS = [
    # GUI panels / labels
    ("Classe desenv.", "Exposure class"),
    ("Classe de exposição", "Exposure class"),
    ("Estratégia de reinforcement", "Reinforcement strategy"),
    ("Estratégia de armadura", "Reinforcement strategy"),
    ("Critério de escolha", "Selection criterion"),
    ("Nível de detalhe", "Report detail level"),
    ("Relatório PDF", "PDF report"),
    ("Diagnóstico and auditoria", "Diagnostics and audit"),
    ("Diagnóstico e auditoria", "Diagnostics and audit"),
    ("Diagnóstico structuralcodes", "structuralcodes diagnostics"),
    ("Tentativas de correcção interactiva", "Interactive correction attempts"),
    ("Tentativas de correcção", "Correction attempts"),
    ("As tentativas detalhadas são exportadas apenas no ficheiro .xlsx.", "Detailed correction attempts are reported in the Excel workbook only."),
    ("Instruções de utilização and tabela tipo", "User instructions and input table format"),
    ("Instruções de utilização", "User instructions"),
    ("tabela tipo", "input table template"),
    ("η alvo", "Target η"),
    ("η mínimo", "Minimum η"),
    ("η máximo", "Maximum η"),
    ("Excesso máx. As", "Maximum As surplus"),
    ("Aplicado apenas quando a estratégia seleccionada é Equilibrada.", "Used only when the Balanced reinforcement strategy is selected."),
    ("Default = EC2 Portugal 2010. EC2:2023 exige", "Default = Portuguese EC2 2010. EC2:2023 requires"),
    ("python -m pip install structuralcodes", "python -m pip install structuralcodes"),
    # validation / backend tables
    ("necessária para ELU/ELS/V/T/DXF", "required for ULS/SLS/shear/torsion/DXF output"),
    ("necessário para identificar combinação", "required to identify the load combination"),
    ("member/node/case deve estar preenchido", "Member/Node/Case must be populated"),
    ("cada member/case deve ter exactamente duas linhas", "each member/case pair should contain exactly two end-node rows"),
    ("sem M01/M02 completo", "M01/M02 cannot be established"),
    ("verificar duplicados ou resultados intermédios", "check for duplicate rows or intermediate station results"),
    ("os dois nós do mesmo member/case devem ter dados geométricos compatíveis", "the two end nodes of the same member/case pair should have compatible geometry data"),
    ("os dois nós do mesmo member/case devem ter dados", "the two end nodes of the same member/case pair should have consistent data"),
    ("esperado em cm; verificar exportação", "expected in cm; check the source export"),
    ("esperado em m", "expected in metres"),
    ("pequenas=", "small="),
    ("grandes=", "large="),
    ("a classe deve vir da coluna Material; fallback interno C30/37 quando ausente", "the concrete grade should be provided in the Material column; the internal C30/37 fallback is used only when the field is blank"),
    ("A geometria candidata é gerada pelo programa; a aceitação normativa depende das APIs disponíveis.", "Candidate geometry is generated by ColumnsEC2; code acceptance depends on the checks exposed by the installed structuralcodes API."),
    ("Calculado se API disponível", "Calculated where the API is available"),
    ("Não avaliado se API ausente", "Not assessed when the API is unavailable"),
    ("Sem fallback interno", "No internal fallback is used"),
    # report/result wording: rewrite as technical English, not literal fragments
    ("Report interno — estados por módulo", "Internal module-status report"),
    ("Report interno", "Internal report"),
    ("estados por módulo", "module status"),
    ("Estados:", "Module status:"),
    ("global=", "overall="),
    ("resistente=", "resistance="),
    ("corte=", "shear="),
    ("torsion=", "torsion="),
    ("ELS=", "SLS="),
    ("detailing=", "detailing="),
    ("corte requer verificação/dimensionamento de reinforcement transversal", "shear requires a detailed transverse-reinforcement design check"),
    ("shear requer verificação/dimensionamento de reinforcement transversal", "shear requires a detailed transverse-reinforcement design check"),
    ("torsion requer verificação/dimensionamento complementar", "torsion requires a complementary design check"),
    ("torção requer verificação/dimensionamento complementar", "torsion requires a complementary design check"),
    ("ELS informativo/não conclusivo ou a verificar", "SLS check is informative/non-conclusive; detailed check required"),
    ("SLS informativo/não conclusivo ou a verificar", "SLS check is informative/non-conclusive; detailed check required"),
    ("detailing construtiva a confirmar", "constructive detailing to be confirmed"),
    ("detailing constructive a confirmar", "constructive detailing to be confirmed"),
    ("verificação/dimensionamento", "detailed design check"),
    ("dimensionamento complementar", "complementary design check"),
    ("a confirmar", "to be confirmed"),
    ("sem intermediate cross-ties", "no intermediate cross-ties"),
    ("sem cross-ties intermédios", "no intermediate cross-ties"),
    ("sem grampos intermédios", "no intermediate cross-ties"),
    ("distribuídos nas faces", "distributed along the faces"),
    ("distribuídas nas faces", "distributed along the faces"),
    ("a meio das faces", "at mid-face"),
    ("nos cantos", "at the corners"),
    ("no canto", "at the corner"),
    ("por nível", "per level"),
    ("grampo(s)", "cross-tie(s)"),
    ("grampos", "cross-ties"),
    ("Grampos", "Cross-ties"),
    ("estribos", "links"),
    ("Estribos", "Links"),
    ("varões", "bars"),
    ("Varões", "Bars"),
    ("armadura transversal", "transverse reinforcement"),
    ("armadura longitudinal", "longitudinal reinforcement"),
    ("reinforcement transversal", "transverse reinforcement"),
    ("pormenorização construtiva", "constructive detailing"),
    ("pormenorização", "detailing"),
    ("Pormenorização", "Detailing"),
    ("fendilhação", "crack control"),
    ("deformação", "deflection"),
    ("fluência", "creep"),
    ("retração", "shrinkage"),
    ("retracção", "shrinkage"),
    ("prumada", "column line"),
    ("Prumada", "Column line"),
    ("piso", "storey"),
    ("Piso", "Storey"),
    ("tramo", "segment"),
    ("Tramo", "Segment"),
    ("caso", "case"),
    ("Caso", "Case"),
    ("Bloqueante", "Blocking design issue"),
    ("bloqueante", "blocking design issue"),
    ("Aviso", "Warning"),
    ("Falha", "Failure"),
    ("falhas", "failures"),
    ("avisos", "warnings"),
    ("Não avaliado", "Not assessed"),
    ("não avaliado", "not assessed"),
    ("não calculado", "not calculated"),
    ("não exposto", "not exposed"),
    ("não exposta", "not exposed"),
    ("não conclusivo", "non-conclusive"),
    ("Informativo", "Informative"),
    ("informativo", "informative"),
    ("Calculado", "Calculated"),
    ("calculado", "calculated"),
    ("Disponível", "Available"),
    ("disponível", "available"),
    ("presente", "present"),
    ("Betão lido da coluna Material", "Concrete grade read from the Material column"),
    ("betão read from the Material column", "concrete grade read from the Material column"),
    ("betão", "concrete grade"),
    ("Betão", "Concrete"),
    ("tabela editável", "editable table"),
    ("envolvente", "governing-case envelope"),
    ("A calcular", "Calculating"),
    ("a calcular", "calculating"),
    ("Cálculo concluído", "Calculation complete"),
    ("concluído", "complete"),
    ("Concluído", "Complete"),
    ("Motor", "Calculation engine"),
    ("Norma", "Standard"),
    ("Verificação", "Design check"),
    ("verificação", "check"),
]

# Regex-level cleanup for common mixed-language artefacts produced by older callbacks.
def _rc8_clean_technical_english(s: str) -> str:
    try:
        import re as _re
        s = _re.sub(r"\bMembro\b", "Member", s)
        s = _re.sub(r"\bEstado\b", "Status", s)
        s = _re.sub(r"\bObjecto\b", "Object", s)
        s = _re.sub(r"\bObjeto\b", "Object", s)
        s = _re.sub(r"\bOrigem\b", "Source", s)
        s = _re.sub(r"\bDetalhe\b", "Details", s)
        s = _re.sub(r"\bCategoria\b", "Category", s)
        s = _re.sub(r"\bResultado\b", "Result", s)
        s = _re.sub(r"\bColunas obrigatórias\b", "Required input fields", s)
        s = _re.sub(r"\bPares de nós\b", "End-node pairs", s)
        s = _re.sub(r"\bConsistência entre nós\b", "End-node data consistency", s)
        s = _re.sub(r"\bUnidades\b", "Units", s)
        s = _re.sub(r"\bMateriais\b", "Materials", s)
        # Avoid doubled words after layered replacements.
        s = s.replace("SLS/SLS", "SLS")
        s = s.replace("shear shear", "shear")
        s = s.replace("torsion torsion", "torsion")
        s = s.replace("Concrete concrete grade", "Concrete grade")
        s = s.replace("concrete grade grade", "concrete grade")
        s = s.replace("cross-tie(s) Ø", "cross-tie(s) Ø")
        s = s.replace("cross-ties intermédios", "intermediate cross-ties")
        s = s.replace("intermediate intermediate cross-ties", "intermediate cross-ties")
        s = s.replace("at the corners +", "at the corners +")
        s = s.replace("Solução", "Reinforcement arrangement")
    except Exception:
        pass
    return s


def _rc8_translate_header(col):
    s = str(col)
    if s in _RC8_HEADER_MAP:
        return _RC8_HEADER_MAP[s]
    try:
        if s in _RC7_HEADER_MAP:
            return _RC7_HEADER_MAP[s]
    except Exception:
        pass
    try:
        return _rc8_translate_technical_en(s)
    except Exception:
        return s


def _rc8_translate_technical_en(value):
    try:
        if _rc3_is_scalar_nan(value):
            return value
    except Exception:
        pass
    if not isinstance(value, str):
        return value
    s = value
    stripped = s.strip()
    if stripped in _RC8_EXACT_VALUE_MAP:
        return _RC8_EXACT_VALUE_MAP[stripped]
    try:
        s = _rc7_translate_technical_en(s)
    except Exception:
        pass
    stripped = s.strip()
    if stripped in _RC8_EXACT_VALUE_MAP:
        return _RC8_EXACT_VALUE_MAP[stripped]
    # Apply longest first to avoid partial Portuguese fragments surviving.
    for old, new in sorted(_RC8_TECHNICAL_REPLACEMENTS, key=lambda kv: len(kv[0]), reverse=True):
        s = s.replace(old, new)
    s = _rc8_clean_technical_english(s)
    return s


def _rc8_lang(app):
    try:
        return LANG_EN if str(app.var_language.get()).upper().startswith("EN") else LANG_PT
    except Exception:
        return LANG_PT


def _rc8_is_en(app):
    return _rc8_lang(app) == LANG_EN


def _rc8_prepare_display_df(df, lang=LANG_EN):
    if df is None or getattr(df, "empty", True):
        return pd.DataFrame()
    out = df.copy()
    if lang != LANG_EN:
        return out
    try:
        out = _rc3_deduplicate_columns(out) if "_rc3_deduplicate_columns" in globals() else out
    except Exception:
        pass
    out.columns = [_rc8_translate_header(c) for c in out.columns]
    for c in list(out.columns):
        try:
            # Translate all text-like display cells. The source dataframes are left untouched.
            if out[c].dtype == object or str(out[c].dtype).startswith("string"):
                out[c] = out[c].map(_rc8_translate_technical_en)
        except Exception:
            pass
    return out


# More explicit technical report for the on-screen Report tab in EN-UK.
def _rc8_build_report_en(app):
    res = getattr(app, "df_results", pd.DataFrame())
    if res is None or res.empty:
        return "No calculation results. Import the input table and run the design/check."
    try:
        src = getattr(app, "df_summary", pd.DataFrame())
        if src is None or src.empty:
            src = res
    except Exception:
        src = res
    src_d = _rc8_prepare_display_df(src, LANG_EN)
    total = len(res)
    try:
        st = _rc3_col_series(res, ["estado_global", "status", "Estado", "Status"]).astype(str)
    except Exception:
        st = pd.Series(dtype=str)
    n_fail = int((st == "Falha").sum()) if not st.empty else 0
    n_warn = int((st == "Aviso").sum()) if not st.empty else 0
    lines = []
    lines.append(f"ColumnsEC2 {APP_VERSION}\n")
    lines.append("Technical design/check report — module status summary\n\n")
    lines.append(f"Analysed cases: {total} | Failures: {n_fail} | Warnings: {n_warn}\n\n")
    # Get columns after display translation.
    for _, r in src_d.head(80).iterrows():
        column_line = r.get("Column line", r.get("name", r.get("Name", "")))
        member = r.get("Member", r.get("member", ""))
        case = r.get("Load case", r.get("Case", r.get("case", "")))
        ned = r.get("N_Ed [kN]", r.get("NEd [kN]", r.get("n_ed_kN", "")))
        my = r.get("My,Ed [kNm]", r.get("My_Ed [kNm]", r.get("my_ed_kNm", "")))
        mz = r.get("Mz,Ed [kNm]", r.get("Mz_Ed [kNm]", r.get("mz_ed_kNm", "")))
        eta = r.get("η_NMyMz", r.get("utilizacao", ""))
        overall = r.get("Overall status", r.get("Status", ""))
        resistance = r.get("Resistance status", "")
        shear = r.get("Shear status", "")
        torsion = r.get("Torsion status", "")
        sls = r.get("SLS status", "")
        detailing = r.get("Detailing status", "")
        sol = r.get("Reinforcement arrangement", r.get("Full reinforcement solution", ""))
        dec = r.get("Engineering decision", r.get("Recommendations", ""))
        lines.append(f"Column line {column_line} | Member {member} | Load case {case}\n")
        lines.append(f"N_Ed={ned} kN | My,Ed={my} kNm | Mz,Ed={mz} kNm | η_NMyMz={eta}\n")
        if any(str(x).strip() for x in [overall, resistance, shear, torsion, sls, detailing]):
            lines.append(f"Module status: overall={overall} | resistance={resistance} | shear={shear} | torsion={torsion} | SLS={sls} | detailing={detailing}\n")
        if str(sol).strip():
            lines.append(f"Reinforcement arrangement: {_rc8_translate_technical_en(str(sol))}\n")
        if str(dec).strip():
            lines.append(f"Engineering decision: {_rc8_translate_technical_en(str(dec))}\n")
        lines.append("\n")
    return "".join(lines)


_rc8_prev_show_df = getattr(ColumnsEC2App, "show_df", None)
def _rc8_show_df(self, tree, df):
    global _RC6_ACTIVE_APP
    _RC6_ACTIVE_APP = self
    display_df = df
    try:
        if _rc8_is_en(self):
            display_df = _rc8_prepare_display_df(df, LANG_EN)
    except Exception:
        display_df = df
    if callable(_rc8_prev_show_df):
        out = _rc8_prev_show_df(self, tree, display_df)
    else:
        out = None
    try:
        if _rc8_is_en(self):
            _rc8_translate_widget_tree(self)
    except Exception:
        pass
    return out
ColumnsEC2App.show_df = _rc8_show_df


_rc8_prev_update_report = getattr(ColumnsEC2App, "update_report", None)
def _rc8_update_report(self):
    if _rc8_is_en(self):
        try:
            self.report_txt.delete("1.0", "end")
            self.report_txt.insert("1.0", _rc8_build_report_en(self))
            return None
        except Exception:
            pass
    out = _rc8_prev_update_report(self) if callable(_rc8_prev_update_report) else None
    return out
ColumnsEC2App.update_report = _rc8_update_report


def _rc8_translate_widget_tree(app):
    try:
        widgets = list(_rc5_walk_widgets(app)) if "_rc5_walk_widgets" in globals() else list(_rc3_walk_widgets(app))
    except Exception:
        widgets = []
    for w in widgets:
        # Static text on labels, buttons, label frames, checkbuttons, notebook tabs via cget('text').
        try:
            txt = str(w.cget("text"))
            new = _rc8_translate_technical_en(txt)
            if new != txt:
                w.configure(text=new)
        except Exception:
            pass
        # Combobox display values.
        try:
            if isinstance(w, ttk.Combobox):
                vals = list(w.cget("values") or [])
                new_vals = [_rc8_translate_technical_en(str(v)) for v in vals]
                if new_vals != vals:
                    w.configure(values=new_vals)
                cur = str(w.get())
                new_cur = _rc8_translate_technical_en(cur)
                if new_cur != cur:
                    w.set(new_cur)
        except Exception:
            pass
    # Specific variables whose displayed value is stored in StringVars.
    try:
        if hasattr(app, "var_pdf_level"):
            v = str(app.var_pdf_level.get())
            app.var_pdf_level.set(_rc8_translate_technical_en(v))
    except Exception:
        pass
    try:
        if hasattr(app, "var_rebar_strategy"):
            v = str(app.var_rebar_strategy.get())
            app.var_rebar_strategy.set(_rc8_translate_technical_en(v))
    except Exception:
        pass
    try:
        if hasattr(app, "var_calc_mode"):
            v = str(app.var_calc_mode.get())
            app.var_calc_mode.set(_rc8_translate_technical_en(v))
    except Exception:
        pass
    try:
        if hasattr(app, "status_var"):
            app.status_var.set(_rc8_translate_technical_en(str(app.status_var.get())))
    except Exception:
        pass
    try:
        app.title("ColumnsEC2 - Reinforced Concrete Column Design (EC2)")
    except Exception:
        pass
    try:
        _rc7_bind_repository_links(app) if "_rc7_bind_repository_links" in globals() else None
    except Exception:
        pass


_rc8_prev_apply_language = getattr(ColumnsEC2App, "apply_language", None)
def _rc8_apply_language(self):
    global _RC6_ACTIVE_APP
    _RC6_ACTIVE_APP = self
    out = _rc8_prev_apply_language(self) if callable(_rc8_prev_apply_language) else None
    if _rc8_is_en(self):
        try:
            _rc8_translate_widget_tree(self)
        except Exception:
            pass
        # Force all known trees to refresh using technical EN-UK.
        try:
            refresh_pairs = [
                ("tree_input", "df_clean"),
                ("tree_pairs", "df_pair"),
                ("tree_validation", "df_validation"),
                ("tree_results", "df_results"),
                ("tree_summary", "df_summary"),
                ("tree_failures", "df_failures"),
                ("tree_notes", "df_notes"),
                ("tree_backend_coverage", "df_backend_coverage"),
                ("tree_structuralcodes_diag", "df_structuralcodes_diag"),
            ]
            for tree_attr, df_attr in refresh_pairs:
                tree = getattr(self, tree_attr, None)
                df = getattr(self, df_attr, pd.DataFrame())
                if tree is not None and df is not None and not getattr(df, "empty", True):
                    _rc8_show_df(self, tree, df)
            # shortlists are generated on demand
            if hasattr(self, "tree_shortlists") and hasattr(self, "build_shortlists_df"):
                df = self.build_shortlists_df()
                if df is not None and not df.empty:
                    _rc8_show_df(self, self.tree_shortlists, df)
            self.update_report()
        except Exception:
            pass
    return out
ColumnsEC2App.apply_language = _rc8_apply_language
try:
    _rc3_apply_language = _rc8_apply_language
except Exception:
    pass


# Translate message boxes with the stronger technical dictionary.
_rc8_prev_showinfo = messagebox.showinfo
_rc8_prev_showwarning = messagebox.showwarning
_rc8_prev_showerror = messagebox.showerror
_rc8_prev_askyesno = messagebox.askyesno

def _rc8_msg(title, message):
    app = globals().get("_RC6_ACTIVE_APP", None)
    if app is not None and _rc8_is_en(app):
        return _rc8_translate_technical_en(str(title)), _rc8_translate_technical_en("" if message is None else str(message))
    return title, message

def _rc8_showinfo(title, message=None, *args, **kwargs):
    t, m = _rc8_msg(title, message)
    return _rc8_prev_showinfo(t, m, *args, **kwargs)

def _rc8_showwarning(title, message=None, *args, **kwargs):
    t, m = _rc8_msg(title, message)
    return _rc8_prev_showwarning(t, m, *args, **kwargs)

def _rc8_showerror(title, message=None, *args, **kwargs):
    t, m = _rc8_msg(title, message)
    return _rc8_prev_showerror(t, m, *args, **kwargs)

def _rc8_askyesno(title, message=None, *args, **kwargs):
    t, m = _rc8_msg(title, message)
    return _rc8_prev_askyesno(t, m, *args, **kwargs)

messagebox.showinfo = _rc8_showinfo
messagebox.showwarning = _rc8_showwarning
messagebox.showerror = _rc8_showerror
messagebox.askyesno = _rc8_askyesno


# Re-translate visible content after calculation/export callbacks update the GUI.
_rc8_prev_run_design = getattr(ColumnsEC2App, "run_design", None)
def _rc8_run_design(self):
    global _RC6_ACTIVE_APP
    _RC6_ACTIVE_APP = self
    result = _rc8_prev_run_design(self) if callable(_rc8_prev_run_design) else None
    try:
        self.after(400, lambda: self.apply_language() if _rc8_is_en(self) else None)
        self.after(1500, lambda: self.apply_language() if _rc8_is_en(self) else None)
        self.after(3500, lambda: self.apply_language() if _rc8_is_en(self) else None)
    except Exception:
        pass
    return result
ColumnsEC2App.run_design = _rc8_run_design


# Excel: one more technical-English pass over workbook text and repository links.
_rc8_prev_write_excel = getattr(ColumnsEC2App, "_write_excel", None)
def _write_excel_rc8(self, path: str):
    if callable(_rc8_prev_write_excel):
        _rc8_prev_write_excel(self, path)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path)
        if _rc8_is_en(self):
            for ws in wb.worksheets:
                for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 5000), max_col=min(ws.max_column, 150)):
                    for cell in row:
                        if isinstance(cell.value, str):
                            cell.value = _rc8_translate_header(cell.value) if cell.row == 1 else _rc8_translate_technical_en(cell.value)
            try:
                wb.properties.title = "ColumnsEC2 - Reinforced Concrete Column Design"
                wb.properties.subject = "Reinforced concrete column design and checking"
                wb.properties.description = "Technical workbook generated by ColumnsEC2. The programme name contains the repository hyperlink."
            except Exception:
                pass
        try:
            _rc7_apply_workbook_links(wb) if "_rc7_apply_workbook_links" in globals() else None
        except Exception:
            pass
        wb.save(path)
    except Exception:
        pass
ColumnsEC2App._write_excel = _write_excel_rc8


# PDF tables: ensure RC8 terminology is used even when the underlying RC7 PDF writer is called.
_rc8_prev_write_pdf = getattr(ColumnsEC2App, "_write_pdf", None)
def _write_pdf_rc8(self, path: str):
    return _rc8_prev_write_pdf(self, path) if callable(_rc8_prev_write_pdf) else None
ColumnsEC2App._write_pdf = _write_pdf_rc8


# Initial hook used by the __main__ block.
def _v092_apply_language_title(app):
    global _RC6_ACTIVE_APP
    _RC6_ACTIVE_APP = app
    try:
        app.apply_language()
    except Exception:
        pass
    try:
        _rc8_translate_widget_tree(app) if _rc8_is_en(app) else None
    except Exception:
        pass
    try:
        _rc7_bind_repository_links(app) if "_rc7_bind_repository_links" in globals() else None
    except Exception:
        pass


# RC8 hotfix: make header-only words translate when they appear as widget text or cell values.
_RC8_DIRECT_TEXT_MAP = {
    "Grupo": "Group", "Objeto": "Object", "Objecto": "Object", "Estado": "Status",
    "Detalhe": "Details", "Verificação": "Design check", "Origem": "Source",
    "Nota": "Engineering note", "Categoria": "Category", "Resultado": "Result",
    "Relatório PDF": "PDF report", "Nível de detalhe": "Report detail level",
}
_rc8_translate_technical_en_base = _rc8_translate_technical_en

def _rc8_translate_technical_en(value):
    try:
        if _rc3_is_scalar_nan(value):
            return value
    except Exception:
        pass
    if not isinstance(value, str):
        return value
    stripped = value.strip()
    if stripped in _RC8_DIRECT_TEXT_MAP:
        return _RC8_DIRECT_TEXT_MAP[stripped]
    if stripped in _RC8_HEADER_MAP:
        return _RC8_HEADER_MAP[stripped]
    if stripped in _RC8_EXACT_VALUE_MAP:
        return _RC8_EXACT_VALUE_MAP[stripped]
    s = _rc8_translate_technical_en_base(value)
    # Upgrade a few older literal translations to technical UK wording.
    upgrades = {
        "Check": "Design check",
        "Note": "Engineering note",
        "Calculation engine scope": "Calculation-engine scope",
    }
    if str(s).strip() in upgrades:
        return upgrades[str(s).strip()]
    return _rc8_clean_technical_english(str(s)) if isinstance(s, str) else s


