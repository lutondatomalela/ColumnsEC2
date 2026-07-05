# -*- coding: utf-8 -*-
# Auto-split from ColumnsEC2 v0.9 RC8.
# This module is executed in the shared runtime namespace by columns_ec2.runtime.loader.
# Keep execution order defined in columns_ec2/runtime/manifest.py.

# -*- coding: utf-8 -*-
"""
Created on Tue Jan 20 17:01:33 2026

@author: Lutonda Tomalela
"""

import io
import math
import re
import threading
import os
import webbrowser
from pathlib import Path
from datetime import datetime
from dataclasses import dataclass
from typing import Dict, List, Tuple, Optional

import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

APP_TITLE = "ColumnsEC2 - Dimensionamento de Pilares (EC2)"
APP_NAME = "ColumnsEC2"
APP_VERSION = "v5.3"
APP_AUTHOR = "Eng.º Lutonda Tomalela"
APP_SUBJECT = "Dimensionamento e verificação de pilares de betão armado segundo o Eurocódigo 2"
APP_KEYWORDS = "ColumnsEC2, Eurocódigo 2, EC2, NP EN 1992-1-1, pilares, betão armado, armaduras, segunda ordem, flexão composta, interação biaxial"
APP_CATEGORY = "Structural Engineering / Reinforced Concrete Design"
APP_XLSX_DESCRIPTION = "Workbook de cálculo com dados de entrada, validação robusta, pares member/case, ELU, ELS, esforço transverso, torção, pormenorização, superfície resistente, resumo, falhas, shortlists e notas EC2."
GITHUB_URL = "https://github.com/lutondatomalela/ColumnsEC2"
MAX_PREVIEW_ROWS = 20000
DEFAULT_CONCRETE_CLASS = "C30/37"


# ============================================================
# Utilidades
# ============================================================
def normalize_text(s: str) -> str:
    s = str(s).strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def safe_float(value, default=float("nan")):
    try:
        if pd.isna(value):
            return default
        if isinstance(value, (int, float)):
            return float(value)
        s = str(value).strip()
        if not s:
            return default
        s = s.replace("\u00a0", " ").replace(" ", "")
        if re.fullmatch(r"-?\d{1,3}(\.\d{3})+,\d+", s):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", ".")
        return float(s)
    except Exception:
        return default


def parse_concrete_strength(material: str) -> float:
    s = str(material)
    m = re.search(r"C\s*(\d+(?:[.,]\d+)?)\s*/\s*(\d+(?:[.,]\d+)?)", s, re.I)
    if m:
        return float(m.group(1).replace(",", "."))
    return parse_concrete_strength(DEFAULT_CONCRETE_CLASS)


def concrete_props(fck: float, alpha_cc: float = 1.0, gamma_c: float = 1.5) -> Dict[str, float]:
    fcm = fck + 8.0
    fcd = alpha_cc * fck / gamma_c
    fctm = 0.30 * fck ** (2.0 / 3.0) if fck <= 50 else 2.12 * math.log(1 + fcm / 10.0)
    ecm = 22.0 * (fcm / 10.0) ** 0.3 * 1000.0
    return {"fck": fck, "fcm": fcm, "fcd": fcd, "fctm": fctm, "Ecm": ecm}


def steel_props(fyk: float = 500.0, gamma_s: float = 1.15) -> Dict[str, float]:
    return {"fyd": fyk / gamma_s, "Es": 210000.0}


def bar_area_mm2(phi_mm: float) -> float:
    return math.pi * phi_mm * phi_mm / 4.0


def cm_to_mm(x) -> float:
    return safe_float(x, 0.0) * 10.0


def m_to_mm(x) -> float:
    return safe_float(x, 0.0) * 1000.0


def split_member_case(text: str) -> Tuple[str, str, str]:
    s = str(text)
    parts = [p.strip() for p in s.split("/")]
    member = parts[0] if len(parts) > 0 else ""
    node = parts[1] if len(parts) > 1 else ""
    case = parts[2] if len(parts) > 2 else ""
    return member, node, case


# ============================================================
# Leitura / limpeza de tabelas
# ============================================================
def parse_pasted_table(text: str) -> pd.DataFrame:
    text = text.strip()
    if not text:
        return pd.DataFrame()

    for sep in ("\t", ";"):
        try:
            df = pd.read_csv(io.StringIO(text), sep=sep, engine="python", dtype=str)
            if len(df.columns) > 1:
                return df
        except Exception:
            pass

    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    if not lines:
        return pd.DataFrame()

    rows = [re.split(r"\s{2,}", line) for line in lines]
    if len(rows) < 2:
        return pd.DataFrame()

    header = rows[0]
    body = rows[1:]
    width = len(header)
    body = [r[:width] + [""] * max(0, width - len(r)) for r in body]
    return pd.DataFrame(body, columns=header)


COLUMN_ALIASES = {
    "member_case": ["member/n", "member/node/case", "member", "member/node"],
    "fx": ["fx (kn)", "fx"],
    "fy": ["fy (kn)", "fy"],
    "fz": ["fz (kn)", "fz"],
    "mx": ["mx (knm)", "mx"],
    "my": ["my (knm)", "my"],
    "mz": ["mz (knm)", "mz"],
    "length": ["length (m)", "length(m)", "length"],
    "material": ["material"],
    "hy": ["hy (cm)", "hy"],
    "hz": ["hz (cm)", "hz"],
    "vy": ["vy (cm)", "vy"],
    "vz": ["vz (cm)", "vz"],
    "vpy": ["vpy (cm)", "vpy"],
    "vpz": ["vpz (cm)", "vpz"],
    "ax": ["ax (cm2)", "ax"],
    "ay": ["ay (cm2)", "ay"],
    "az": ["az (cm2)", "az"],
    "ix": ["ix (cm4)", "ix"],
    "iy": ["iy (cm4)", "iy"],
    "iz": ["iz (cm4)", "iz"],
    "name": ["name", "nome"],
}


def rename_known_columns(df: pd.DataFrame) -> pd.DataFrame:
    norm_to_original = {normalize_text(c): c for c in df.columns}
    rename_map = {}
    for target, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in norm_to_original:
                rename_map[norm_to_original[alias]] = target
                break
    return df.rename(columns=rename_map).copy()


def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    df = rename_known_columns(df)
    df["__row_order"] = range(len(df))

    numeric_cols = [
        "fx", "fy", "fz", "mx", "my", "mz", "length", "hy", "hz",
        "vy", "vz", "vpy", "vpz", "ax", "ay", "az", "ix", "iy", "iz"
    ]
    for c in numeric_cols:
        if c in df.columns:
            df[c] = df[c].map(safe_float)

    if "member_case" in df.columns:
        split_vals = df["member_case"].map(split_member_case)
        df["member"] = split_vals.map(lambda x: x[0])
        df["node"] = split_vals.map(lambda x: x[1])
        df["case"] = split_vals.map(lambda x: x[2])
    else:
        for c in ["member", "node", "case"]:
            if c not in df.columns:
                df[c] = ""

    if "name" not in df.columns:
        df["name"] = ""

    return df


def combine_member_end_actions(df: pd.DataFrame) -> pd.DataFrame:
    """
    Converte a tabela nó-a-nó da tabela de cálculo numa tabela por member/case,
    preservando a ordem dos nós como aparece no ficheiro.
    Também conserva FY, FZ e MX para verificações complementares.
    """
    if df is None or df.empty:
        return df

    rows = []
    for _, grp in df.groupby(["member", "case", "name"], dropna=False):
        grp = grp.sort_values("__row_order")
        r1 = grp.iloc[0]
        r2 = grp.iloc[1] if len(grp) >= 2 else grp.iloc[0]

        row = {
            "member": r1.get("member", ""),
            "case": r1.get("case", ""),
            "name": r1.get("name", ""),
            "node_i": r1.get("node", ""),
            "node_j": r2.get("node", ""),
            "member_case_i": f"{r1.get('member','')}/{r1.get('node','')}/{r1.get('case','')}",
            "member_case_j": f"{r2.get('member','')}/{r2.get('node','')}/{r2.get('case','')}",
            "fx_i": safe_float(r1.get("fx", 0.0), 0.0),
            "fx_j": safe_float(r2.get("fx", 0.0), 0.0),
            "fy_i": safe_float(r1.get("fy", 0.0), 0.0),
            "fy_j": safe_float(r2.get("fy", 0.0), 0.0),
            "fz_i": safe_float(r1.get("fz", 0.0), 0.0),
            "fz_j": safe_float(r2.get("fz", 0.0), 0.0),
            "mx_i": safe_float(r1.get("mx", 0.0), 0.0),
            "mx_j": safe_float(r2.get("mx", 0.0), 0.0),
            "my_i": safe_float(r1.get("my", 0.0), 0.0),
            "my_j": safe_float(r2.get("my", 0.0), 0.0),
            "mz_i": safe_float(r1.get("mz", 0.0), 0.0),
            "mz_j": safe_float(r2.get("mz", 0.0), 0.0),
            "length": safe_float(r1.get("length", 0.0), 0.0),
            "material": r1.get("material", "") or DEFAULT_CONCRETE_CLASS,
            "hy": safe_float(r1.get("hy", float("nan"))),
            "hz": safe_float(r1.get("hz", float("nan"))),
            "ax": safe_float(r1.get("ax", float("nan"))),
            "iy": safe_float(r1.get("iy", float("nan"))),
            "iz": safe_float(r1.get("iz", float("nan"))),
            "__row_order": safe_float(r1.get("__row_order", 0), 0),
            "n_nodes_found": len(grp),
        }

        row["fx"] = max(abs(row["fx_i"]), abs(row["fx_j"]))
        row["fy"] = max(abs(row["fy_i"]), abs(row["fy_j"]))
        row["fz"] = max(abs(row["fz_i"]), abs(row["fz_j"]))
        row["mx"] = max(abs(row["mx_i"]), abs(row["mx_j"]))
        row["my"] = max(abs(row["my_i"]), abs(row["my_j"]))
        row["mz"] = max(abs(row["mz_i"]), abs(row["mz_j"]))
        rows.append(row)

    return pd.DataFrame(rows).sort_values(["__row_order", "member", "case"]).reset_index(drop=True)



def reduce_to_governing_cases(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df

    work = df.copy()
    work["_abs_fx"] = work["fx"].abs().fillna(0.0)
    work["_abs_my"] = work["my"].abs().fillna(0.0)
    work["_abs_mz"] = work["mz"].abs().fillna(0.0)
    work["_score"] = work["_abs_fx"] * 0.20 + work["_abs_my"] + work["_abs_mz"]

    selected_idx = set()
    for _, grp in work.groupby(["member", "name"], dropna=False):
        if grp.empty:
            continue
        selected_idx.add(grp["_score"].idxmax())
        selected_idx.add(grp["_abs_fx"].idxmax())
        selected_idx.add(grp["_abs_my"].idxmax())
        selected_idx.add(grp["_abs_mz"].idxmax())

        grp2 = grp.assign(
            _my_over_n=grp["_abs_my"] / grp["_abs_fx"].replace(0.0, 1e-9),
            _mz_over_n=grp["_abs_mz"] / grp["_abs_fx"].replace(0.0, 1e-9),
        )
        selected_idx.add(grp2["_my_over_n"].idxmax())
        selected_idx.add(grp2["_mz_over_n"].idxmax())

    reduced = work.loc[sorted(selected_idx)].copy().sort_values(["member", "name", "case"]).reset_index(drop=True)
    reduced.drop(columns=["_abs_fx", "_abs_my", "_abs_mz", "_score"], inplace=True, errors="ignore")
    return reduced


# ============================================================
# Modelos de armadura
# ============================================================
@dataclass
class Layout:
    phi_long_mm: float
    phi_st_mm: float
    n_bars_y: int
    n_bars_z: int
    b_mm: float
    h_mm: float
    cover_mm: float

    @property
    def n_total(self) -> int:
        return 2 * self.n_bars_y + 2 * self.n_bars_z - 4

    @property
    def as_prov_mm2(self) -> float:
        return self.n_total * bar_area_mm2(self.phi_long_mm)

    def clear_spacing_ok(self, agg_mm: float = 20.0, min_clear_mm: float = 20.0) -> bool:
        req = max(min_clear_mm, self.phi_long_mm, agg_mm + 5.0)
        edge = self.cover_mm + self.phi_st_mm + self.phi_long_mm / 2.0

        ok_y = True
        ok_z = True
        if self.n_bars_y > 1:
            span = self.b_mm - 2.0 * edge
            ctc = span / (self.n_bars_y - 1)
            ok_y = (ctc - self.phi_long_mm) >= req
        if self.n_bars_z > 1:
            span = self.h_mm - 2.0 * edge
            ctc = span / (self.n_bars_z - 1)
            ok_z = (ctc - self.phi_long_mm) >= req

        return ok_y and ok_z


def classify_failure_reason(reason: str) -> str:
    txt = str(reason or "").lower()
    if not txt:
        return ""
    if "pormenorização" in txt:
        return "pormenorizacao"
    if "insuficiência de armadura" in txt:
        return "armadura_insuficiente"
    if "esbelteza" in txt:
        return "esbelteza"
    if "biaxial" in txt:
        return "resistencia_biaxial"
    if "dados" in txt:
        return "dados_incompletos"
    return "outra"


def recommend_actions(row: pd.Series) -> str:
    recs = []
    fail_type = str(row.get("failure_type", "") or "")
    b = safe_float(row.get("b_cm", 0.0), 0.0)
    h = safe_float(row.get("h_cm", 0.0), 0.0)
    lam_y = safe_float(row.get("lambda_y", 0.0), 0.0)
    lam_z = safe_float(row.get("lambda_z", 0.0), 0.0)
    lam_lim_y = safe_float(row.get("lambda_lim_y", 0.0), 0.0)
    lam_lim_z = safe_float(row.get("lambda_lim_z", 0.0), 0.0)

    if fail_type == "armadura_insuficiente":
        recs += ["aumentar a secção", "permitir mais varões por face", "aumentar o diâmetro dos varões"]
    elif fail_type == "pormenorizacao":
        recs += ["reduzir congestionamento de armaduras", "aumentar a secção", "rever recobrimento e arranjo dos varões"]
    elif fail_type == "resistencia_biaxial":
        recs += ["aumentar a secção", "aumentar a armadura longitudinal", "rever os esforços de cálculo governantes"]
    elif fail_type == "esbelteza":
        recs += ["reduzir o comprimento efectivo", "aumentar rigidez da secção", "melhorar contraventamento"]
    elif fail_type == "dados_incompletos":
        recs += ["verificar exportação da tabela de cálculo", "confirmar os dois nós por member/case"]

    if lam_lim_y > 0 and lam_y > lam_lim_y:
        recs.append("reduzir esbelteza em Y")
    if lam_lim_z > 0 and lam_z > lam_lim_z:
        recs.append("reduzir esbelteza em Z")
    if b and h and min(b, h) < 25:
        recs.append("avaliar aumento da menor dimensão da secção")

    recs = list(dict.fromkeys(recs))
    return "; ".join(recs[:4])


def serialize_shortlist(shortlist_rows: List[Dict]) -> str:
    parts = []
    for i, item in enumerate(shortlist_rows, start=1):
        txt = (
            f"{i}) {item.get('solucao','')} | "
            f"As={item.get('as_prov_mm2',0):.0f} mm² | "
            f"util={item.get('utilizacao','')}"
        )
        if item.get("status_short") == "OK":
            txt += " | OK"
        else:
            txt += f" | {item.get('failure_short','')}"
        parts.append(txt)
    return " || ".join(parts)



class ColumnDesigner:
    def __init__(
        self,
        cover_mm=35.0,
        fyk=500.0,
        gamma_c=1.5,
        gamma_s=1.15,
        phi_eff=2.0,
        l0y_factor=1.0,
        l0z_factor=1.0,
        calc_mode="dimensionamento",
    ):
        self.cover_mm = cover_mm
        self.fyk = fyk
        self.gamma_c = gamma_c
        self.gamma_s = gamma_s
        self.phi_eff = phi_eff
        self.l0y_factor = l0y_factor
        self.l0z_factor = l0z_factor
        self.calc_mode = calc_mode

        self.long_diams = [10.0, 12.0, 16.0, 20.0, 25.0, 32.0]
        self.stirrup_diams = [8.0, 10.0]
        self.spacing_candidates_mm = [75.0, 100.0, 150.0, 200.0, 250.0, 300.0]

        self._layout_cache = {}
        self._capacity_cache = {}

    def min_longitudinal_as(self, n_ed_kN: float, ac_mm2: float, fyd: float) -> float:
        return max(0.10 * n_ed_kN * 1e3 / fyd, 0.002 * ac_mm2)

    def max_longitudinal_as(self, ac_mm2: float) -> float:
        return 0.04 * ac_mm2

    def tie_spacing_max(self, b_mm: float, h_mm: float, phi_long_mm: float) -> float:
        return min(12.0 * phi_long_mm, min(b_mm, h_mm), 300.0)

    def choose_stirrup(self, phi_long_mm: float) -> float:
        req = max(6.0, phi_long_mm / 4.0)
        for phi in self.stirrup_diams:
            if phi >= req:
                return phi
        return self.stirrup_diams[-1]

    def choose_spacing(self, smax_mm: float) -> float:
        valid = [s for s in self.spacing_candidates_mm if s <= smax_mm + 1e-9]
        return max(valid) if valid else smax_mm

    def second_order_nominal_curvature(
        self,
        n_ed_kN: float,
        m01_kNm: float,
        m02_kNm: float,
        l0_mm: float,
        ac_mm2: float,
        i_mm4: float,
        h_mm: float,
        as_total_mm2: float,
        fcd: float,
        fck: float,
        fyd: float,
        es: float,
        phi_eff: Optional[float] = None,
    ):
        if phi_eff is None:
            phi_eff = self.phi_eff

        if ac_mm2 <= 0 or i_mm4 <= 0 or h_mm <= 0 or fcd <= 0 or es <= 0:
            m0ed = max(abs(m01_kNm), abs(m02_kNm))
            return m0ed, 0.0, 0.0, 1.0, False, m0ed, 0.0

        i_radius = math.sqrt(i_mm4 / ac_mm2)
        slenderness = l0_mm / i_radius if i_radius > 0 else 0.0

        m01 = abs(m01_kNm)
        m02 = abs(m02_kNm)
        if m02 < m01:
            m01, m02 = m02, m01

        m0e = max(0.6 * m02 + 0.4 * m01, 0.4 * m02)
        m0ed = max(m02, m0e)

        omega = (as_total_mm2 * fyd) / (ac_mm2 * fcd) if ac_mm2 * fcd > 0 else 0.0
        n_red = (n_ed_kN * 1e3) / (ac_mm2 * fcd) if ac_mm2 * fcd > 0 else 0.0
        nu = 1.0 + omega
        nbal = 0.4
        denom = max(nu - nbal, 1e-9)
        kr = min(max((nu - n_red) / denom, 0.0), 1.0)

        beta = 0.35 + fck / 200.0 - slenderness / 150.0
        kphi = max(1.0 + beta * max(phi_eff, 0.0), 1.0)

        eps_yd = fyd / es
        d_eff = 0.8 * h_mm
        inv_r0 = eps_yd / (0.45 * d_eff) if d_eff > 0 else 0.0
        inv_r = kr * kphi * inv_r0

        c_coeff = 8.0
        e2_mm = inv_r * (l0_mm ** 2) / c_coeff
        m2_kNm = n_ed_kN * e2_mm / 1000.0
        m_ed = m0ed + m2_kNm

        second_order = bool(m2_kNm > 1e-9)
        return m_ed, slenderness, inv_r, kphi, second_order, m0ed, m2_kNm

    def infer_is_circular(self, row: pd.Series, b_mm: float, h_mm: float) -> bool:
        name = str(row.get("name", "")).lower()
        return ("circ" in name or "circular" in name or "ø" in name or "phi" in name) and abs(b_mm - h_mm) <= 1e-6

    def max_bars_per_face(self, b_mm: float, h_mm: float, is_circular: bool = False) -> Tuple[int, int]:
        if is_circular:
            perimeter_m = math.pi * (max(b_mm, h_mm) / 1000.0)
            nmax = max(2, math.ceil(perimeter_m / 0.1))
            return nmax, nmax

        max_y = max(2, math.ceil((b_mm / 1000.0) / 0.1))
        max_z = max(2, math.ceil((h_mm / 1000.0) / 0.1))
        return max_y, max_z

    def build_candidate_layouts(self, b_mm, h_mm, is_circular=False):
        key = (round(b_mm, 1), round(h_mm, 1), round(self.cover_mm, 1), bool(is_circular))
        if key in self._layout_cache:
            return self._layout_cache[key]

        layouts = []
        max_face_y, max_face_z = self.max_bars_per_face(b_mm, h_mm, is_circular=is_circular)

        for phi_long in self.long_diams:
            phi_st = self.choose_stirrup(phi_long)
            for n_bars_y in range(2, max_face_y + 1):
                for n_bars_z in range(2, max_face_z + 1):
                    layout = Layout(phi_long, phi_st, n_bars_y, n_bars_z, b_mm, h_mm, self.cover_mm)
                    if layout.clear_spacing_ok():
                        layouts.append(layout)

        layouts.sort(key=lambda x: (x.as_prov_mm2, x.phi_long_mm, x.n_total))
        self._layout_cache[key] = layouts
        return layouts

    def layout_coordinates(self, layout: Layout) -> List[Tuple[float, float]]:
        edge = layout.cover_mm + layout.phi_st_mm + layout.phi_long_mm / 2.0
        y_left = -layout.b_mm / 2.0 + edge
        y_right = layout.b_mm / 2.0 - edge
        z_bot = -layout.h_mm / 2.0 + edge
        z_top = layout.h_mm / 2.0 - edge

        ys = [y_left] if layout.n_bars_y == 1 else [
            y_left + i * (y_right - y_left) / (layout.n_bars_y - 1)
            for i in range(layout.n_bars_y)
        ]
        zs = [z_bot] if layout.n_bars_z == 1 else [
            z_bot + i * (z_top - z_bot) / (layout.n_bars_z - 1)
            for i in range(layout.n_bars_z)
        ]

        pts = []
        for y in ys:
            pts.append((float(y), float(z_top)))
            pts.append((float(y), float(z_bot)))

        for z in zs[1:-1]:
            pts.append((float(y_left), float(z)))
            pts.append((float(y_right), float(z)))

        unique = []
        seen = set()
        for p in pts:
            key = (round(p[0], 6), round(p[1], 6))
            if key not in seen:
                seen.add(key)
                unique.append(p)
        return unique

    def section_response(
        self,
        layout: Layout,
        n_ed_kN: float,
        angle_rad: float,
        c_mm: float,
        fcd: float,
        fyd: float,
        Es: float,
    ):
        eps_cu = 0.0035
        pts = self.layout_coordinates(layout)
        a_bar = bar_area_mm2(layout.phi_long_mm)
        ca = math.cos(angle_rad)
        sa = math.sin(angle_rad)

        def ucoord(y, z):
            return y * ca + z * sa

        corners = [
            (-layout.b_mm / 2.0, -layout.h_mm / 2.0),
            (-layout.b_mm / 2.0,  layout.h_mm / 2.0),
            ( layout.b_mm / 2.0, -layout.h_mm / 2.0),
            ( layout.b_mm / 2.0,  layout.h_mm / 2.0),
        ]
        us = [ucoord(y, z) for y, z in corners]
        u_max = max(us)
        u_na = u_max - c_mm

        ny = max(10, int(layout.b_mm / 35.0))
        nz = max(10, int(layout.h_mm / 35.0))
        dy = layout.b_mm / ny
        dz = layout.h_mm / nz
        dA = dy * dz

        N = 0.0
        My = 0.0
        Mz = 0.0

        for iy in range(ny):
            y = -layout.b_mm / 2.0 + (iy + 0.5) * dy
            for iz in range(nz):
                z = -layout.h_mm / 2.0 + (iz + 0.5) * dz
                u = ucoord(y, z)
                if u <= u_na:
                    continue
                eps = eps_cu * (u - u_na) / max(c_mm, 1e-9)
                sig = 0.0 if eps <= 0.0 else min(fcd, fcd * eps / 0.002)
                Fc = sig * dA
                N += Fc
                My += Fc * z
                Mz += Fc * y

        for y, z in pts:
            u = ucoord(y, z)
            eps_s = eps_cu * (u - u_na) / max(c_mm, 1e-9)
            sig_s = max(-fyd, min(fyd, Es * eps_s))
            Fs = sig_s * a_bar
            N += Fs
            My += Fs * z
            Mz += Fs * y

        return N / 1000.0, abs(My) / 1e6, abs(Mz) / 1e6

    def capacity_for_layout(self, layout: Layout, n_ed_kN: float, fcd: float, fyd: float, Es: float):
        key = (
            round(layout.b_mm, 1),
            round(layout.h_mm, 1),
            layout.phi_long_mm,
            layout.phi_st_mm,
            layout.n_bars_y,
            layout.n_bars_z,
            round(n_ed_kN, 0),
            round(fcd, 3),
            round(fyd, 3),
        )
        if key in self._capacity_cache:
            return self._capacity_cache[key]

        angles = [i * math.pi / 36.0 for i in range(19)]
        capacities = []
        c_max = 2.2 * max(layout.b_mm, layout.h_mm)
        c_values = [5.0 + i * (c_max - 5.0) / 79.0 for i in range(80)]

        for ang in angles:
            best = None
            for c_mm in c_values:
                N, My, Mz = self.section_response(layout, n_ed_kN, ang, c_mm, fcd, fyd, Es)
                diff = abs(N - n_ed_kN)
                if best is None or diff < best[0]:
                    best = (diff, My, Mz)
            if best is not None:
                capacities.append((best[1], best[2]))

        self._capacity_cache[key] = capacities
        return capacities

    def biaxial_ok(self, my_ed_kNm: float, mz_ed_kNm: float, capacities: List[Tuple[float, float]]):
        if not capacities:
            return False, None, None, None

        my_req = abs(my_ed_kNm)
        mz_req = abs(mz_ed_kNm)

        best_util = None
        best_my = None
        best_mz = None

        for my_cap, mz_cap in capacities:
            if my_cap <= 1e-9 or mz_cap <= 1e-9:
                continue

            util_lin = my_req / my_cap + mz_req / mz_cap
            util_ell = math.sqrt((my_req / my_cap) ** 2 + (mz_req / mz_cap) ** 2)
            util = min(util_lin, util_ell)

            if best_util is None or util < best_util:
                best_util = util
                best_my = my_cap
                best_mz = mz_cap

        if best_util is None:
            return False, None, None, None

        return best_util <= 1.0, best_util, best_my, best_mz



    def approx_required_as(self, n_ed_kN, my_ed_kNm, mz_ed_kNm, b_mm, h_mm, fyd, as_min_mm2):
        z_y = 0.80 * h_mm
        z_z = 0.80 * b_mm
        as_n = 0.10 * n_ed_kN * 1e3 / fyd
        as_my = abs(my_ed_kNm) * 1e6 / (0.87 * fyd * z_y) if z_y > 0 else 0.0
        as_mz = abs(mz_ed_kNm) * 1e6 / (0.87 * fyd * z_z) if z_z > 0 else 0.0
        return max(as_min_mm2, as_n + as_my + as_mz)

    def design_one(self, row: pd.Series, prebuilt_candidates=None):
        material = str(row.get("material", "C30/37"))
        fck = parse_concrete_strength(material)
        cp = concrete_props(fck, gamma_c=self.gamma_c)
        sp = steel_props(self.fyk, gamma_s=self.gamma_s)
        fyd = sp["fyd"]
        Es = sp["Es"]

        b_mm = cm_to_mm(row.get("hy", 0.0))
        h_mm = cm_to_mm(row.get("hz", 0.0))
        ac_mm2 = safe_float(row.get("ax", float("nan"))) * 100.0
        iy_mm4 = safe_float(row.get("iy", float("nan"))) * 10000.0
        iz_mm4 = safe_float(row.get("iz", float("nan"))) * 10000.0

        if b_mm <= 0 and ac_mm2 > 0:
            b_mm = math.sqrt(ac_mm2)
        if h_mm <= 0:
            h_mm = b_mm
        if not math.isfinite(ac_mm2) or ac_mm2 <= 0:
            ac_mm2 = b_mm * h_mm
        if not math.isfinite(iy_mm4) or iy_mm4 <= 0:
            iy_mm4 = b_mm * h_mm ** 3 / 12.0
        if not math.isfinite(iz_mm4) or iz_mm4 <= 0:
            iz_mm4 = h_mm * b_mm ** 3 / 12.0

        ly_mm = m_to_mm(row.get("length", 0.0)) * self.l0y_factor
        lz_mm = m_to_mm(row.get("length", 0.0)) * self.l0z_factor

        n_ed_kN = max(
            abs(safe_float(row.get("fx_i", 0.0), 0.0)),
            abs(safe_float(row.get("fx_j", 0.0), 0.0)),
        )
        my1_kNm = safe_float(row.get("my_i", 0.0), 0.0)
        my2_kNm = safe_float(row.get("my_j", 0.0), 0.0)
        mz1_kNm = safe_float(row.get("mz_i", 0.0), 0.0)
        mz2_kNm = safe_float(row.get("mz_j", 0.0), 0.0)

        if int(safe_float(row.get("n_nodes_found", 0), 0)) < 2:
            reason = "Falha de dados: member/case sem os dois nós necessários"
            out = {
                "member": row.get("member", ""),
                "case": row.get("case", ""),
                "name": row.get("name", ""),
                "node_i": row.get("node_i", ""),
                "node_j": row.get("node_j", ""),
                "member_case_i": row.get("member_case_i", ""),
                "member_case_j": row.get("member_case_j", ""),
                "material": material,
                "b_cm": b_mm / 10.0 if b_mm else None,
                "h_cm": h_mm / 10.0 if h_mm else None,
                "length_m": safe_float(row.get("length", 0.0), 0.0),
                "n_nodes_found": int(safe_float(row.get("n_nodes_found", 0), 0)),
                "n_ed_kN": n_ed_kN,
                "my_i_kNm": my1_kNm,
                "my_j_kNm": my2_kNm,
                "mz_i_kNm": mz1_kNm,
                "mz_j_kNm": mz2_kNm,
                "my_ed_kNm": None,
                "mz_ed_kNm": None,
                "lambda_y": None,
                "lambda_z": None,
                "lambda_lim_y": None,
                "lambda_lim_z": None,
                "lambda_check_y": "",
                "lambda_check_z": "",
                "m0e_y_kNm": None,
                "m0e_z_kNm": None,
                "m2_y_kNm": None,
                "m2_z_kNm": None,
                "second_order_y": "",
                "second_order_z": "",
                "as_min_mm2": None,
                "as_req_mm2": None,
                "as_max_mm2": None,
                "max_bars_face_y": None,
                "max_bars_face_z": None,
                "phi_long_mm": None,
                "n_total": None,
                "n_bars_y": None,
                "n_bars_z": None,
                "as_prov_mm2": None,
                "phi_st_mm": None,
                "s_st_mm": None,
                "s_st_max_mm": None,
                "mrd_y_kNm": None,
                "mrd_z_kNm": None,
                "status": "Falha",
                "utilizacao": None,
                "solucao": "",
                "failure_reason": reason,
                "failure_type": classify_failure_reason(reason),
                "recommendations": "verificar exportação da tabela de cálculo; confirmar dois nós por member/case",
                "shortlist_text": "",
            }
            return out

        as_min_mm2 = self.min_longitudinal_as(n_ed_kN, ac_mm2, fyd)
        as_max_mm2 = self.max_longitudinal_as(ac_mm2)

        is_circular = self.infer_is_circular(row, b_mm, h_mm)
        max_face_y, max_face_z = self.max_bars_per_face(b_mm, h_mm, is_circular=is_circular)

        as_seed_mm2 = max(as_min_mm2, 0.002 * ac_mm2)

        my_ed_kNm, lambda_y, invr_y, kphi_y, second_order_y, my0eq_kNm, my2nd_kNm = \
            self.second_order_nominal_curvature(
                n_ed_kN, my1_kNm, my2_kNm, ly_mm, ac_mm2, iy_mm4, h_mm,
                as_seed_mm2, cp["fcd"], fck, fyd, Es
            )

        mz_ed_kNm, lambda_z, invr_z, kphi_z, second_order_z, mz0eq_kNm, mz2nd_kNm = \
            self.second_order_nominal_curvature(
                n_ed_kN, mz1_kNm, mz2_kNm, lz_mm, ac_mm2, iz_mm4, b_mm,
                as_seed_mm2, cp["fcd"], fck, fyd, Es
            )

        n_red = (n_ed_kN * 1e3) / (ac_mm2 * cp["fcd"]) if ac_mm2 * cp["fcd"] > 0 else 0.0
        n_red = max(n_red, 1e-6)
        omega_seed = (as_seed_mm2 * fyd) / (ac_mm2 * cp["fcd"]) if ac_mm2 * cp["fcd"] > 0 else 0.0
        A = 1.0 / (1.0 + 0.2 * max(self.phi_eff, 0.0))
        B = math.sqrt(max(1.0 + 2.0 * omega_seed, 1.0))

        def lambda_lim(m1, m2):
            m1a = abs(m1)
            m2a = abs(m2)
            if m2a < 1e-9:
                rm = 1.0
            else:
                rm = min(m1a, m2a) / max(m1a, m2a)
            C = max(0.7, 1.7 - rm)
            return 20.0 * A * B * C / math.sqrt(n_red)

        lambda_lim_y = lambda_lim(my1_kNm, my2_kNm)
        lambda_lim_z = lambda_lim(mz1_kNm, mz2_kNm)

        lambda_check_y = "Dispensa 2a ordem" if lambda_y <= lambda_lim_y else "Considerar 2a ordem"
        lambda_check_z = "Dispensa 2a ordem" if lambda_z <= lambda_lim_z else "Considerar 2a ordem"

        as_req_mm2 = self.approx_required_as(
            n_ed_kN, my_ed_kNm, mz_ed_kNm, b_mm, h_mm, fyd, as_min_mm2
        )

        base = {
            "member": row.get("member", ""),
            "case": row.get("case", ""),
            "name": row.get("name", ""),
            "node_i": row.get("node_i", ""),
            "node_j": row.get("node_j", ""),
            "member_case_i": row.get("member_case_i", ""),
            "member_case_j": row.get("member_case_j", ""),
            "material": material,
            "b_cm": b_mm / 10.0,
            "h_cm": h_mm / 10.0,
            "length_m": safe_float(row.get("length", 0.0), 0.0),
            "n_nodes_found": int(safe_float(row.get("n_nodes_found", 0), 0)),
            "n_ed_kN": n_ed_kN,
            "my_i_kNm": my1_kNm,
            "my_j_kNm": my2_kNm,
            "mz_i_kNm": mz1_kNm,
            "mz_j_kNm": mz2_kNm,
            "my_ed_kNm": my_ed_kNm,
            "mz_ed_kNm": mz_ed_kNm,
            "lambda_y": lambda_y,
            "lambda_z": lambda_z,
            "lambda_lim_y": lambda_lim_y,
            "lambda_lim_z": lambda_lim_z,
            "lambda_check_y": lambda_check_y,
            "lambda_check_z": lambda_check_z,
            "m0e_y_kNm": my0eq_kNm,
            "m0e_z_kNm": mz0eq_kNm,
            "m2_y_kNm": my2nd_kNm,
            "m2_z_kNm": mz2nd_kNm,
            "second_order_y": "Sim" if second_order_y else "Não",
            "second_order_z": "Sim" if second_order_z else "Não",
            "as_min_mm2": as_min_mm2,
            "as_req_mm2": as_req_mm2,
            "as_max_mm2": as_max_mm2,
            "max_bars_face_y": max_face_y,
            "max_bars_face_z": max_face_z,
            "failure_reason": "",
            "failure_type": "",
            "recommendations": "",
            "shortlist_text": "",
        }

        candidates = prebuilt_candidates if prebuilt_candidates is not None else self.build_candidate_layouts(
            b_mm, h_mm, is_circular=is_circular
        )
        candidates = [
            ly for ly in candidates
            if ly.as_prov_mm2 <= as_max_mm2 and ly.n_bars_y <= max_face_y and ly.n_bars_z <= max_face_z
        ]

        if not candidates:
            reason = "Falha de pormenorização: não existe disposição admissível dentro do limite máximo de varões por face e/ou As,max"
            out = {
                **base,
                "phi_long_mm": None,
                "n_total": None,
                "n_bars_y": None,
                "n_bars_z": None,
                "as_prov_mm2": None,
                "phi_st_mm": None,
                "s_st_mm": None,
                "s_st_max_mm": None,
                "mrd_y_kNm": None,
                "mrd_z_kNm": None,
                "status": "Falha",
                "utilizacao": None,
                "solucao": "",
                "failure_reason": reason,
            }
            out["failure_type"] = classify_failure_reason(reason)
            out["recommendations"] = recommend_actions(pd.Series(out))
            return out

        candidates_as_ok = [ly for ly in candidates if ly.as_prov_mm2 >= as_req_mm2]

        if not candidates_as_ok:
            reason = "Falha por insuficiência de armadura: nenhuma disposição admissível atingiu As,req dentro das restrições geométricas"
            out = {
                **base,
                "phi_long_mm": None,
                "n_total": None,
                "n_bars_y": None,
                "n_bars_z": None,
                "as_prov_mm2": None,
                "phi_st_mm": None,
                "s_st_mm": None,
                "s_st_max_mm": None,
                "mrd_y_kNm": None,
                "mrd_z_kNm": None,
                "status": "Falha",
                "utilizacao": None,
                "solucao": "",
                "failure_reason": reason,
            }
            out["failure_type"] = classify_failure_reason(reason)
            out["recommendations"] = recommend_actions(pd.Series(out))
            return out

        candidates_as_ok.sort(key=lambda ly: (abs(ly.as_prov_mm2 - as_req_mm2), ly.as_prov_mm2))
        shortlist = smart_shortlist_v45(candidates_as_ok, as_req_mm2, max_n=30)

        shortlist_rows = []
        chosen = None
        chosen_util = None
        chosen_caps = (None, None)
        status = "Pré-dimensionado"
        failure_reason = ""

        if self.calc_mode == "dimensionamento":
            for layout in shortlist:
                capacities = self.capacity_for_layout(layout, n_ed_kN, cp["fcd"], fyd, Es)
                ok, util, my_cap, mz_cap = self.biaxial_ok(my_ed_kNm, mz_ed_kNm, capacities)

                shortlist_rows.append({
                    "solucao": f"{layout.n_total}Ø{int(layout.phi_long_mm)}",
                    "as_prov_mm2": layout.as_prov_mm2,
                    "utilizacao": "" if util is None else f"{util:.3f}",
                    "status_short": "OK" if ok else "Falha",
                    "failure_short": "" if ok else "biaxial",
                })

                if ok:
                    chosen = layout
                    chosen_util = util
                    chosen_caps = (my_cap, mz_cap)
                    status = "OK"
                    break

            if chosen is None:
                chosen = min(shortlist, key=lambda ly: ly.as_prov_mm2)
                chosen_util = as_req_mm2 / chosen.as_prov_mm2 if chosen.as_prov_mm2 > 0 else None
                status = "Falha"
                failure_reason = "Falha de resistência biaxial: embora As,prov >= As,req, nenhuma solução da shortlist verificou a interação biaxial no modo dimensionamento"
        else:
            for layout in shortlist:
                shortlist_rows.append({
                    "solucao": f"{layout.n_total}Ø{int(layout.phi_long_mm)}",
                    "as_prov_mm2": layout.as_prov_mm2,
                    "utilizacao": "",
                    "status_short": "Pré",
                    "failure_short": "",
                })
            chosen = min(shortlist, key=lambda ly: ly.as_prov_mm2)
            chosen_util = as_req_mm2 / chosen.as_prov_mm2 if chosen.as_prov_mm2 > 0 else None

        smax = self.tie_spacing_max(b_mm, h_mm, chosen.phi_long_mm)
        sprov = self.choose_spacing(smax)

        out = {
            **base,
            "phi_long_mm": chosen.phi_long_mm,
            "n_total": chosen.n_total,
            "n_bars_y": chosen.n_bars_y,
            "n_bars_z": chosen.n_bars_z,
            "as_prov_mm2": chosen.as_prov_mm2,
            "phi_st_mm": chosen.phi_st_mm,
            "s_st_mm": sprov,
            "s_st_max_mm": smax,
            "mrd_y_kNm": chosen_caps[0],
            "mrd_z_kNm": chosen_caps[1],
            "status": status,
            "utilizacao": chosen_util,
            "solucao": f"{chosen.n_total}Ø{int(chosen.phi_long_mm)} + estribos Ø{int(chosen.phi_st_mm)}//{sprov / 10:.1f} cm",
            "failure_reason": failure_reason,
            "shortlist_text": serialize_shortlist(shortlist_rows),
        }

        if status == "Falha" and not failure_reason:
            out["failure_reason"] = "Falha por motivo não classificado"

        out["failure_type"] = classify_failure_reason(out["failure_reason"])
        out["recommendations"] = recommend_actions(pd.Series(out))
        return out

    def design_dataframe(self, df: pd.DataFrame, progress_callback=None):
        results = []
        total = len(df)
        grouped_candidates = {}

        for _, row in df.iterrows():
            b_mm = cm_to_mm(row.get("hy", 0.0))
            h_mm = cm_to_mm(row.get("hz", 0.0))
            ac_mm2 = safe_float(row.get("ax", float("nan"))) * 100.0
            if b_mm <= 0 and ac_mm2 > 0:
                b_mm = math.sqrt(ac_mm2)
            if h_mm <= 0:
                h_mm = b_mm
            is_circular = self.infer_is_circular(row, b_mm, h_mm)
            sec_key = (round(b_mm, 1), round(h_mm, 1), bool(is_circular))
            if sec_key not in grouped_candidates:
                grouped_candidates[sec_key] = self.build_candidate_layouts(b_mm, h_mm, is_circular=is_circular)

        for i, row in enumerate((r for _, r in df.iterrows()), start=1):
            b_mm = cm_to_mm(row.get("hy", 0.0))
            h_mm = cm_to_mm(row.get("hz", 0.0))
            ac_mm2 = safe_float(row.get("ax", float("nan"))) * 100.0
            if b_mm <= 0 and ac_mm2 > 0:
                b_mm = math.sqrt(ac_mm2)
            if h_mm <= 0:
                h_mm = b_mm
            is_circular = self.infer_is_circular(row, b_mm, h_mm)
            sec_key = (round(b_mm, 1), round(h_mm, 1), bool(is_circular))
            results.append(self.design_one(row, prebuilt_candidates=grouped_candidates[sec_key]))

            if progress_callback and (i == total or i % 10 == 0):
                progress_callback(i, total)

        out = pd.DataFrame(results)
        if not out.empty and "utilizacao" in out.columns:
            out["sort_key"] = out["utilizacao"].fillna(999.0)
        return out







# ============================================================
# Cálculos EC2 v3 — ELS, esforço transverso, torção e verificações auxiliares
# ============================================================
def extract_combination_number(case_value) -> str:
    s = str(case_value or "").strip()
    m = re.search(r"\b(\d+)\b", s)
    return m.group(1) if m else s


def classify_limit_state_from_case(case_value) -> str:
    s = str(case_value or "").upper()
    if any(k in s for k in ["ELS", "SLS", "SERV", "SERVICE", "RARA", "FREQ", "QUASE", "QP", "Q.P.", "(S)"]):
        return "ELS"
    if any(k in s for k in ["ELU", "ULS", "STR", "EQU", "GEO", "(U)", "(C)"]):
        return "ELU"
    return "ELU"


def lambda_lim_ec2_practical(m01, m02, n_red, omega, phi_eff):
    """Forma prática de λlim com razão de momentos com sinal."""
    n_red = max(float(n_red), 1e-6)
    A = 1.0 / (1.0 + 0.2 * max(phi_eff, 0.0))
    B = math.sqrt(max(1.0 + 2.0 * max(omega, 0.0), 1.0))
    a = float(m01)
    b = float(m02)
    if abs(a) > abs(b):
        a, b = b, a
    rm = 1.0 if abs(b) < 1e-9 else max(-1.0, min(1.0, a / b))
    C = max(0.7, min(2.7, 1.7 - rm))
    return 20.0 * A * B * C / math.sqrt(n_red), rm, A, B, C


def elastic_service_check(n_kN, my_kNm, mz_kNm, b_mm, h_mm, iy_mm4, iz_mm4, as_mm2, fck, fyk, ecm, fctm):
    """Verificação ELS elástica simplificada: tensões no betão e aço por secção bruta."""
    if b_mm <= 0 or h_mm <= 0 or iy_mm4 <= 0 or iz_mm4 <= 0:
        return {}
    A = b_mm * h_mm
    N = abs(n_kN) * 1e3
    My = abs(my_kNm) * 1e6
    Mz = abs(mz_kNm) * 1e6
    corners = [(-b_mm/2, -h_mm/2), (-b_mm/2, h_mm/2), (b_mm/2, -h_mm/2), (b_mm/2, h_mm/2)]
    sigmas = []
    for y,z in corners:
        sigmas.append(N/A + My*z/iy_mm4 + Mz*y/iz_mm4)
    sigma_c_max = max(sigmas)
    sigma_c_min = min(sigmas)
    alpha_e = 200000.0 / max(ecm, 1e-9)
    sigma_s_max = abs(alpha_e * sigma_c_min) if sigma_c_min < 0 else abs(alpha_e * max(abs(v) for v in sigmas))
    lim_c = 0.60 * fck
    lim_s = 0.80 * fyk
    cracking = "Potencial fendilhação" if sigma_c_min < -fctm else "Sem tração relevante"
    status = "OK" if (sigma_c_max <= lim_c and sigma_s_max <= lim_s) else "Verificar"
    return {
        "service_sigma_c_max_MPa": sigma_c_max,
        "service_sigma_c_min_MPa": sigma_c_min,
        "service_sigma_s_max_MPa": sigma_s_max,
        "service_sigma_c_lim_MPa": lim_c,
        "service_sigma_s_lim_MPa": lim_s,
        "service_crack_warning": cracking,
        "service_status": status,
    }


def shear_check_ec2_practical(v_ed_kN, n_ed_kN, b_mm, h_mm, d_mm, as_long_mm2, fck, fcd, gamma_c):
    """VRd,c simplificado para pilares; não dimensiona estribos por V, apenas classifica."""
    if b_mm <= 0 or h_mm <= 0 or d_mm <= 0:
        return {"VRdc_kN": None, "status": "Dados insuficientes"}
    bw = b_mm
    k = min(2.0, 1.0 + math.sqrt(200.0 / max(d_mm, 1e-9)))
    rho_l = min(max(as_long_mm2 / max(bw * d_mm, 1e-9), 0.0), 0.02)
    sigma_cp = min(abs(n_ed_kN) * 1e3 / max(b_mm*h_mm,1e-9), 0.2 * fcd)
    crdc = 0.18 / gamma_c
    vrdc_N = (crdc * k * (100.0 * rho_l * fck) ** (1/3) + 0.15 * sigma_cp) * bw * d_mm
    vrdc_kN = vrdc_N / 1e3
    status = "OK" if abs(v_ed_kN) <= vrdc_kN else "Requer verificação/armadura de esforço transverso"
    return {"VRdc_kN": vrdc_kN, "status": status}


def _design_one_v3(self, row: pd.Series, prebuilt_candidates=None):
    material = str(row.get("material", "") or DEFAULT_CONCRETE_CLASS)
    if material.strip().lower() in ["", "nan", "none"]:
        material = DEFAULT_CONCRETE_CLASS
    fck = parse_concrete_strength(material)
    cp = concrete_props(fck, gamma_c=self.gamma_c)
    sp = steel_props(self.fyk, gamma_s=self.gamma_s)
    fyd = sp["fyd"]
    Es = sp["Es"]

    b_mm = cm_to_mm(row.get("hy", 0.0))
    h_mm = cm_to_mm(row.get("hz", 0.0))
    ac_mm2 = safe_float(row.get("ax", float("nan"))) * 100.0
    iy_mm4 = safe_float(row.get("iy", float("nan"))) * 10000.0
    iz_mm4 = safe_float(row.get("iz", float("nan"))) * 10000.0
    if b_mm <= 0 and ac_mm2 > 0:
        b_mm = math.sqrt(ac_mm2)
    if h_mm <= 0:
        h_mm = b_mm
    if not math.isfinite(ac_mm2) or ac_mm2 <= 0:
        ac_mm2 = b_mm * h_mm
    if not math.isfinite(iy_mm4) or iy_mm4 <= 0:
        iy_mm4 = b_mm * h_mm ** 3 / 12.0
    if not math.isfinite(iz_mm4) or iz_mm4 <= 0:
        iz_mm4 = h_mm * b_mm ** 3 / 12.0

    ly_mm = m_to_mm(row.get("length", 0.0)) * self.l0y_factor
    lz_mm = m_to_mm(row.get("length", 0.0)) * self.l0z_factor
    n_ed_kN = max(abs(safe_float(row.get("fx_i", 0.0), 0.0)), abs(safe_float(row.get("fx_j", 0.0), 0.0)))
    vy_ed_kN = max(abs(safe_float(row.get("fy_i", row.get("fy", 0.0)), 0.0)), abs(safe_float(row.get("fy_j", row.get("fy", 0.0)), 0.0)))
    vz_ed_kN = max(abs(safe_float(row.get("fz_i", row.get("fz", 0.0)), 0.0)), abs(safe_float(row.get("fz_j", row.get("fz", 0.0)), 0.0)))
    mx_ed_kNm = max(abs(safe_float(row.get("mx_i", row.get("mx", 0.0)), 0.0)), abs(safe_float(row.get("mx_j", row.get("mx", 0.0)), 0.0)))
    my1_kNm = safe_float(row.get("my_i", 0.0), 0.0)
    my2_kNm = safe_float(row.get("my_j", 0.0), 0.0)
    mz1_kNm = safe_float(row.get("mz_i", 0.0), 0.0)
    mz2_kNm = safe_float(row.get("mz_j", 0.0), 0.0)
    case_value = row.get("case", "")
    limit_state = classify_limit_state_from_case(case_value)
    combination_number = extract_combination_number(case_value)

    if int(safe_float(row.get("n_nodes_found", 0), 0)) < 2:
        reason = "Falha de dados: member/case sem os dois nós necessários"
        out = {
            "member": row.get("member", ""), "case": case_value, "combination_number": combination_number,
            "limit_state": limit_state, "name": row.get("name", ""), "node_i": row.get("node_i", ""), "node_j": row.get("node_j", ""),
            "material": material, "b_cm": b_mm/10 if b_mm else None, "h_cm": h_mm/10 if h_mm else None,
            "length_m": safe_float(row.get("length", 0.0), 0.0), "n_nodes_found": int(safe_float(row.get("n_nodes_found", 0), 0)),
            "n_ed_kN": n_ed_kN, "vy_ed_kN": vy_ed_kN, "vz_ed_kN": vz_ed_kN, "mx_ed_kNm": mx_ed_kNm,
            "my_i_kNm": my1_kNm, "my_j_kNm": my2_kNm, "mz_i_kNm": mz1_kNm, "mz_j_kNm": mz2_kNm,
            "status": "Falha", "failure_reason": reason, "failure_type": classify_failure_reason(reason),
            "recommendations": "verificar exportação da tabela de cálculo; confirmar dois nós por member/case", "shortlist_text": "",
        }
        return out

    as_min_mm2 = self.min_longitudinal_as(n_ed_kN, ac_mm2, fyd)
    as_max_mm2 = self.max_longitudinal_as(ac_mm2)
    is_circular = self.infer_is_circular(row, b_mm, h_mm)
    max_face_y, max_face_z = self.max_bars_per_face(b_mm, h_mm, is_circular=is_circular)
    as_seed_mm2 = max(as_min_mm2, 0.002 * ac_mm2)

    # Imperfeições geométricas práticas: theta_i = theta0 * alpha_h; e_i = theta_i*l0/2.
    theta0 = 1.0/200.0
    L_m = max(safe_float(row.get("length", 0.0), 0.0), 1e-6)
    alpha_h = max(2.0/3.0, min(1.0, 2.0/math.sqrt(max(L_m, 1e-9))))
    theta_i = theta0 * alpha_h
    e_imp_y_mm = theta_i * ly_mm / 2.0
    e_imp_z_mm = theta_i * lz_mm / 2.0
    m_imp_y_kNm = n_ed_kN * e_imp_y_mm / 1000.0
    m_imp_z_kNm = n_ed_kN * e_imp_z_mm / 1000.0

    my_ed_kNm, lambda_y, invr_y, kphi_y, second_order_y, my0eq_kNm, my2nd_kNm = self.second_order_nominal_curvature(
        n_ed_kN, my1_kNm, my2_kNm, ly_mm, ac_mm2, iy_mm4, h_mm, as_seed_mm2, cp["fcd"], fck, fyd, Es
    )
    mz_ed_kNm, lambda_z, invr_z, kphi_z, second_order_z, mz0eq_kNm, mz2nd_kNm = self.second_order_nominal_curvature(
        n_ed_kN, mz1_kNm, mz2_kNm, lz_mm, ac_mm2, iz_mm4, b_mm, as_seed_mm2, cp["fcd"], fck, fyd, Es
    )

    n_red = max((n_ed_kN * 1e3) / max(ac_mm2 * cp["fcd"], 1e-9), 1e-6)
    omega_seed = (as_seed_mm2 * fyd) / max(ac_mm2 * cp["fcd"], 1e-9)
    lambda_lim_y, rm_y, A_lam, B_lam, C_y = lambda_lim_ec2_practical(my1_kNm, my2_kNm, n_red, omega_seed, self.phi_eff)
    lambda_lim_z, rm_z, _A, _B, C_z = lambda_lim_ec2_practical(mz1_kNm, mz2_kNm, n_red, omega_seed, self.phi_eff)
    needs_2y = lambda_y > lambda_lim_y
    needs_2z = lambda_z > lambda_lim_z
    my0eq_kNm = max(my0eq_kNm, m_imp_y_kNm)
    mz0eq_kNm = max(mz0eq_kNm, m_imp_z_kNm)
    if not needs_2y:
        my2nd_kNm = 0.0
    if not needs_2z:
        mz2nd_kNm = 0.0
    my_ed_kNm = my0eq_kNm + my2nd_kNm
    mz_ed_kNm = mz0eq_kNm + mz2nd_kNm
    lambda_check_y = "Dispensa 2.ª ordem" if not needs_2y else "Considerar 2.ª ordem"
    lambda_check_z = "Dispensa 2.ª ordem" if not needs_2z else "Considerar 2.ª ordem"

    as_req_mm2 = self.approx_required_as(n_ed_kN, my_ed_kNm, mz_ed_kNm, b_mm, h_mm, fyd, as_min_mm2)
    base = {
        "member": row.get("member", ""), "case": case_value, "combination_number": combination_number, "limit_state": limit_state,
        "name": row.get("name", ""), "node_i": row.get("node_i", ""), "node_j": row.get("node_j", ""),
        "member_case_i": row.get("member_case_i", ""), "member_case_j": row.get("member_case_j", ""),
        "material": material, "fck_MPa": fck, "b_cm": b_mm/10.0, "h_cm": h_mm/10.0, "length_m": safe_float(row.get("length", 0.0), 0.0),
        "n_nodes_found": int(safe_float(row.get("n_nodes_found", 0), 0)),
        "n_ed_kN": n_ed_kN, "vy_ed_kN": vy_ed_kN, "vz_ed_kN": vz_ed_kN, "mx_ed_kNm": mx_ed_kNm,
        "my_i_kNm": my1_kNm, "my_j_kNm": my2_kNm, "mz_i_kNm": mz1_kNm, "mz_j_kNm": mz2_kNm,
        "my_ed_kNm": my_ed_kNm, "mz_ed_kNm": mz_ed_kNm,
        "lambda_y": lambda_y, "lambda_z": lambda_z, "lambda_lim_y": lambda_lim_y, "lambda_lim_z": lambda_lim_z,
        "lambda_check_y": lambda_check_y, "lambda_check_z": lambda_check_z, "rm_y": rm_y, "rm_z": rm_z,
        "m0e_y_kNm": my0eq_kNm, "m0e_z_kNm": mz0eq_kNm, "m2_y_kNm": my2nd_kNm, "m2_z_kNm": mz2nd_kNm,
        "m_imp_y_kNm": m_imp_y_kNm, "m_imp_z_kNm": m_imp_z_kNm,
        "second_order_y": "Sim" if needs_2y else "Dispensada", "second_order_z": "Sim" if needs_2z else "Dispensada",
        "as_min_mm2": as_min_mm2, "as_req_mm2": as_req_mm2, "as_max_mm2": as_max_mm2,
        "max_bars_face_y": max_face_y, "max_bars_face_z": max_face_z,
        "failure_reason": "", "failure_type": "", "recommendations": "", "shortlist_text": "",
    }
    candidates = prebuilt_candidates if prebuilt_candidates is not None else self.build_candidate_layouts(b_mm, h_mm, is_circular=is_circular)
    candidates = [ly for ly in candidates if ly.as_prov_mm2 <= as_max_mm2 and ly.n_bars_y <= max_face_y and ly.n_bars_z <= max_face_z]
    if not candidates:
        reason = "Falha de pormenorização: não existe disposição admissível dentro do limite máximo de varões por face e/ou As,max"
        out = {**base, "phi_long_mm": None, "n_total": None, "n_bars_y": None, "n_bars_z": None, "as_prov_mm2": None,
               "phi_st_mm": None, "s_st_mm": None, "s_st_max_mm": None, "mrd_y_kNm": None, "mrd_z_kNm": None,
               "status": "Falha", "utilizacao": None, "solucao": "", "failure_reason": reason}
        out["failure_type"] = classify_failure_reason(reason); out["recommendations"] = recommend_actions(pd.Series(out)); return out
    candidates_as_ok = [ly for ly in candidates if ly.as_prov_mm2 >= as_req_mm2]
    if not candidates_as_ok:
        reason = "Falha por insuficiência de armadura: nenhuma disposição admissível atingiu As,req dentro das restrições geométricas"
        out = {**base, "phi_long_mm": None, "n_total": None, "n_bars_y": None, "n_bars_z": None, "as_prov_mm2": None,
               "phi_st_mm": None, "s_st_mm": None, "s_st_max_mm": None, "mrd_y_kNm": None, "mrd_z_kNm": None,
               "status": "Falha", "utilizacao": None, "solucao": "", "failure_reason": reason}
        out["failure_type"] = classify_failure_reason(reason); out["recommendations"] = recommend_actions(pd.Series(out)); return out
    candidates_as_ok.sort(key=lambda ly: (abs(ly.as_prov_mm2 - as_req_mm2), ly.as_prov_mm2))
    shortlist = candidates_as_ok[:10]
    shortlist_rows=[]; chosen=None; chosen_util=None; chosen_caps=(None,None); status="Pré-dimensionado"; failure_reason=""
    if self.calc_mode == "dimensionamento":
        for layout in shortlist:
            capacities = self.capacity_for_layout(layout, n_ed_kN, cp["fcd"], fyd, Es)
            ok, util, my_cap, mz_cap = self.biaxial_ok(my_ed_kNm, mz_ed_kNm, capacities)
            shortlist_rows.append({"solucao": f"{layout.n_total}Ø{int(layout.phi_long_mm)}", "as_prov_mm2": layout.as_prov_mm2,
                                   "utilizacao": "" if util is None else f"{util:.3f}", "status_short": "OK" if ok else "Falha", "failure_short": "" if ok else "biaxial"})
            if ok:
                chosen=layout; chosen_util=util; chosen_caps=(my_cap,mz_cap); status="OK"; break
        if chosen is None:
            chosen=min(shortlist,key=lambda ly:ly.as_prov_mm2); chosen_util=as_req_mm2/chosen.as_prov_mm2 if chosen.as_prov_mm2>0 else None
            status="Falha"; failure_reason="Falha de resistência biaxial: embora As,prov >= As,req, nenhuma solução da shortlist verificou a interação biaxial"
    else:
        for layout in shortlist:
            shortlist_rows.append({"solucao": f"{layout.n_total}Ø{int(layout.phi_long_mm)}", "as_prov_mm2": layout.as_prov_mm2, "utilizacao":"", "status_short":"Pré", "failure_short":""})
        chosen=min(shortlist,key=lambda ly: ly.as_prov_mm2); chosen_util=as_req_mm2/chosen.as_prov_mm2 if chosen.as_prov_mm2>0 else None
    smax = self.tie_spacing_max(b_mm,h_mm,chosen.phi_long_mm); sprov=self.choose_spacing(smax)
    shear_y = shear_check_ec2_practical(vy_ed_kN, n_ed_kN, h_mm, b_mm, 0.8*b_mm, chosen.as_prov_mm2, fck, cp["fcd"], self.gamma_c)
    shear_z = shear_check_ec2_practical(vz_ed_kN, n_ed_kN, b_mm, h_mm, 0.8*h_mm, chosen.as_prov_mm2, fck, cp["fcd"], self.gamma_c)
    torsion_ratio = mx_ed_kNm / max(abs(my_ed_kNm), abs(mz_ed_kNm), 1.0)
    torsion_status = "Aviso: torção relevante não dimensionada" if torsion_ratio > 0.10 else "Sem aviso relevante"
    els = elastic_service_check(n_ed_kN, my_ed_kNm, mz_ed_kNm, b_mm, h_mm, iy_mm4, iz_mm4, chosen.as_prov_mm2, fck, self.fyk, cp["Ecm"], cp["fctm"])
    if limit_state != "ELS":
        els["service_status"] = "Informativo — caso não identificado como ELS"
    out = {**base, "phi_long_mm": chosen.phi_long_mm, "n_total": chosen.n_total, "n_bars_y": chosen.n_bars_y, "n_bars_z": chosen.n_bars_z,
           "as_prov_mm2": chosen.as_prov_mm2, "phi_st_mm": chosen.phi_st_mm, "s_st_mm": sprov, "s_st_max_mm": smax,
           "mrd_y_kNm": chosen_caps[0], "mrd_z_kNm": chosen_caps[1], "status": status, "utilizacao": chosen_util,
           "solucao": f"{chosen.n_total}Ø{int(chosen.phi_long_mm)} + estribos Ø{int(chosen.phi_st_mm)}//{sprov/10:.1f} cm",
           "failure_reason": failure_reason, "shortlist_text": serialize_shortlist(shortlist_rows),
           "v_ed_y_kN": vy_ed_kN, "v_ed_z_kN": vz_ed_kN, "v_rd_c_y_kN": shear_y.get("VRdc_kN"), "v_rd_c_z_kN": shear_z.get("VRdc_kN"),
           "shear_status_y": shear_y.get("status"), "shear_status_z": shear_z.get("status"),
           "torsion_ratio": torsion_ratio, "torsion_status": torsion_status,
           "service_combination": combination_number, **els}
    if status == "Falha" and not failure_reason:
        out["failure_reason"] = "Falha por motivo não classificado"
    out["failure_type"] = classify_failure_reason(out["failure_reason"])
    out["recommendations"] = recommend_actions(pd.Series(out))
    if out.get("shear_status_y","").startswith("Requer") or out.get("shear_status_z","").startswith("Requer"):
        out["recommendations"] = (out.get("recommendations","") + "; verificar esforço transverso segundo EC2 6.2").strip("; ")
    if "Aviso" in out.get("torsion_status",""):
        out["recommendations"] = (out.get("recommendations","") + "; verificar torção segundo EC2 6.3").strip("; ")
    return out

ColumnDesigner.design_one = _design_one_v3

# ============================================================
# GUI profissional + exportação XLSX/PDF
# ============================================================
class ColumnsEC2App(tk.Tk):
    TEMPLATE_COLUMNS = [
        "Member/Node/Case", "FX (kN)", "FY (kN)", "FZ (kN)", "MX (kNm)", "MY (kNm)", "MZ (kNm)",
        "Length (m)", "Material", "HY (cm)", "HZ (cm)", "VY (cm)", "VZ (cm)", "VPY (cm)", "VPZ (cm)",
        "AX (cm2)", "AY (cm2)", "AZ (cm2)", "IX (cm4)", "IY (cm4)", "IZ (cm4)", "Name"
    ]

    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1260x760")
        self.minsize(960, 620)

        self.df_raw = pd.DataFrame()
        self.df_clean = pd.DataFrame()
        self.df_pair = pd.DataFrame()
        self.df_calc_input = pd.DataFrame()
        self.df_results = pd.DataFrame()
        self.df_summary = pd.DataFrame()
        self.df_failures = pd.DataFrame()
        self.df_ok = pd.DataFrame()
        self.df_validation = pd.DataFrame()
        self.df_notes = pd.DataFrame()
        self.df_filtered = pd.DataFrame()
        self.input_file_path = ""
        self.analysis_thread = None

        self.var_cover = tk.StringVar(value="35")
        self.var_fyk = tk.StringVar(value="500")
        self.var_concrete_default = tk.StringVar(value=DEFAULT_CONCRETE_CLASS)  # fallback interno se a tabela não trouxer material
        self.var_phi_eff = tk.StringVar(value="2.0")
        self.var_l0y = tk.StringVar(value="1.0")
        self.var_l0z = tk.StringVar(value="1.0")
        self.var_summary = tk.BooleanVar(value=True)
        self.var_reduce_cases = tk.BooleanVar(value=True)
        self.var_calc_mode = tk.StringVar(value="dimensionamento")
        self.var_filter_status = tk.StringVar(value="Todos")
        self.var_filter_fail = tk.StringVar(value="Todos")
        self.var_filter_member = tk.StringVar(value="")
        self.var_pdf_scope = tk.StringVar(value="Resumo + falhas")

        self.status_var = tk.StringVar(value="Cole ou importe a tabela de esforços.")
        self.progress_var = tk.DoubleVar(value=0.0)
        self.progress_text_var = tk.StringVar(value="0%")
        self.progress_var.trace_add("write", lambda *args: self.progress_text_var.set(f"{self.progress_var.get():.0f}%"))

        self._build_ui()

    # --------------------------- UI ---------------------------
    def _build_ui(self):
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except Exception:
            pass

        bg = style.lookup("TFrame", "background") or "#f3f5f7"
        self.configure(background=bg)
        style.configure("TLabelframe", padding=8)
        style.configure("TLabelframe.Label", font=("Segoe UI", 9, "bold"))
        style.configure("TButton", padding=(8, 6), font=("Segoe UI", 9))
        style.configure("Primary.TButton", padding=(10, 8), font=("Segoe UI", 9, "bold"))
        style.configure("Treeview", rowheight=24, font=("Segoe UI", 9))
        style.configure("Treeview.Heading", font=("Segoe UI", 9, "bold"))
        style.configure("TNotebook.Tab", padding=(10, 6), font=("Segoe UI", 9))
        style.configure("Header.TLabel", font=("Segoe UI Semibold", 11))
        style.configure("Subtle.TLabel", font=("Segoe UI", 8), foreground="#5f6b7a")
        style.configure("Status.TLabel", font=("Segoe UI", 9))

        self.rowconfigure(0, weight=1)
        self.rowconfigure(1, weight=0)
        self.columnconfigure(0, weight=1)

        root = ttk.Frame(self, padding=8)
        root.grid(row=0, column=0, sticky="nsew")
        root.rowconfigure(0, weight=1)
        root.columnconfigure(0, weight=1)

        paned = ttk.Panedwindow(root, orient="horizontal")
        paned.grid(row=0, column=0, sticky="nsew")

        sidebar_host = ttk.Frame(paned, width=370)
        sidebar_host.pack_propagate(False)
        sidebar_host.rowconfigure(0, weight=1)
        sidebar_host.columnconfigure(0, weight=1)

        self.sidebar_canvas = tk.Canvas(sidebar_host, highlightthickness=0, borderwidth=0, background=bg)
        sb_y = ttk.Scrollbar(sidebar_host, orient="vertical", command=self.sidebar_canvas.yview)
        self.sidebar_canvas.configure(yscrollcommand=sb_y.set)
        self.sidebar_canvas.grid(row=0, column=0, sticky="nsew")
        sb_y.grid(row=0, column=1, sticky="ns")

        sidebar_inner = ttk.Frame(self.sidebar_canvas, padding=(0, 0, 6, 0))
        sidebar_window = self.sidebar_canvas.create_window((0, 0), window=sidebar_inner, anchor="nw")

        def _sync_sidebar(_event=None):
            self.sidebar_canvas.configure(scrollregion=self.sidebar_canvas.bbox("all"))
            self.sidebar_canvas.itemconfigure(sidebar_window, width=self.sidebar_canvas.winfo_width())

        sidebar_inner.bind("<Configure>", _sync_sidebar)
        self.sidebar_canvas.bind("<Configure>", _sync_sidebar)

        def _on_mousewheel(event):
            try:
                self.sidebar_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            except Exception:
                pass

        self.sidebar_canvas.bind("<Enter>", lambda e: self.sidebar_canvas.bind_all("<MouseWheel>", _on_mousewheel))
        self.sidebar_canvas.bind("<Leave>", lambda e: self.sidebar_canvas.unbind_all("<MouseWheel>"))

        right = ttk.Frame(paned)
        right.rowconfigure(0, weight=1)
        right.columnconfigure(0, weight=1)
        paned.add(sidebar_host, weight=0)
        paned.add(right, weight=1)

        bottom = ttk.Frame(self, padding=(8, 4, 8, 8))
        bottom.grid(row=1, column=0, sticky="ew")
        bottom.columnconfigure(1, weight=1)
        ttk.Label(bottom, text="Estado:", style="Header.TLabel").grid(row=0, column=0, sticky="w", padx=(0, 8))
        ttk.Label(bottom, textvariable=self.status_var, style="Status.TLabel").grid(row=0, column=1, sticky="ew")
        ttk.Progressbar(bottom, variable=self.progress_var, maximum=100, length=260).grid(row=0, column=2, sticky="e", padx=(12, 6))
        ttk.Label(bottom, textvariable=self.progress_text_var, width=6, anchor="e").grid(row=0, column=3, sticky="e")

        self._build_sidebar(sidebar_inner)
        self._build_main_tabs(right)
        self.after(120, lambda: paned.sashpos(0, 385))

    def _build_sidebar(self, parent):
        hero = ttk.LabelFrame(parent, text="ColumnsEC2")
        hero.pack(fill="x", pady=(0, 8))
        program_link = ttk.Label(hero, text="ColumnsEC2", style="Header.TLabel", cursor="hand2")
        program_link.pack(anchor="w")
        program_link.bind("<Button-1>", lambda _e: webbrowser.open_new(GITHUB_URL))
        ttk.Label(hero, text="Dimensionamento de pilares de betão armado (EC2)", style="Header.TLabel").pack(anchor="w", pady=(2, 0))
        ttk.Label(
            hero,
            text="Ferramenta para importação de esforços do Excel/tabela, verificação ELU, 2.ª ordem, interação biaxial, pormenorização e relatórios .xlsx/.pdf.",
            style="Subtle.TLabel",
            wraplength=330,
            justify="left",
        ).pack(anchor="w", pady=(2, 0))

        data = ttk.LabelFrame(parent, text="1. Entrada")
        data.pack(fill="x", pady=(0, 8))
        ttk.Button(data, text="Colar área de transferência", command=self.paste_clipboard).grid(row=0, column=0, sticky="ew", padx=4, pady=4)
        ttk.Button(data, text="Importar .xlsx/.csv", command=self.import_file).grid(row=0, column=1, sticky="ew", padx=4, pady=4)
        ttk.Button(data, text="Ler caixa de texto", command=self.load_from_textbox).grid(row=1, column=0, sticky="ew", padx=4, pady=4)
        ttk.Button(data, text="Modelo de tabela", command=self.export_template).grid(row=1, column=1, sticky="ew", padx=4, pady=4)
        data.columnconfigure(0, weight=1)
        data.columnconfigure(1, weight=1)

        params = ttk.LabelFrame(parent, text="2. Parâmetros EC2")
        params.pack(fill="x", pady=(0, 8))
        self._add_label_entry(params, "Recobrimento [mm]", self.var_cover, 0)
        ttk.Label(params, text="Aço fyk [MPa]").grid(row=1, column=0, sticky="w", padx=6, pady=4)
        ttk.Combobox(params, textvariable=self.var_fyk, values=["400", "500"], state="readonly", width=14).grid(row=1, column=1, sticky="ew", padx=6, pady=4)
        ttk.Label(params, text="Betão").grid(row=2, column=0, sticky="w", padx=6, pady=4)
        ttk.Label(params, text="lido da tabela (coluna Material)", style="Subtle.TLabel", anchor="w").grid(row=2, column=1, sticky="ew", padx=6, pady=4)
        self._add_label_entry(params, "φef", self.var_phi_eff, 3)
        self._add_label_entry(params, "l0y/L", self.var_l0y, 4)
        self._add_label_entry(params, "l0z/L", self.var_l0z, 5)
        ttk.Label(params, text="Modo").grid(row=6, column=0, sticky="w", padx=6, pady=4)
        ttk.Combobox(params, textvariable=self.var_calc_mode, values=["pre_dimensionamento", "dimensionamento"], state="readonly", width=18).grid(row=6, column=1, sticky="ew", padx=6, pady=4)
        ttk.Checkbutton(params, text="Reduzir para casos governantes", variable=self.var_reduce_cases).grid(row=7, column=0, columnspan=2, sticky="w", padx=6, pady=3)
        ttk.Checkbutton(params, text="Resumo por membro", variable=self.var_summary).grid(row=8, column=0, columnspan=2, sticky="w", padx=6, pady=3)
        params.columnconfigure(1, weight=1)

        filters = ttk.LabelFrame(parent, text="3. Filtros")
        filters.pack(fill="x", pady=(0, 8))
        ttk.Label(filters, text="Member").grid(row=0, column=0, sticky="w", padx=6, pady=4)
        ttk.Entry(filters, textvariable=self.var_filter_member).grid(row=0, column=1, sticky="ew", padx=6, pady=4)
        ttk.Label(filters, text="Estado").grid(row=1, column=0, sticky="w", padx=6, pady=4)
        ttk.Combobox(filters, textvariable=self.var_filter_status, values=["Todos", "OK", "Falha", "Pré-dimensionado"], state="readonly").grid(row=1, column=1, sticky="ew", padx=6, pady=4)
        ttk.Label(filters, text="Falha").grid(row=2, column=0, sticky="w", padx=6, pady=4)
        ttk.Combobox(filters, textvariable=self.var_filter_fail, values=["Todos", "pormenorizacao", "armadura_insuficiente", "esbelteza", "resistencia_biaxial", "dados_incompletos", "outra"], state="readonly").grid(row=2, column=1, sticky="ew", padx=6, pady=4)
        ttk.Button(filters, text="Aplicar", command=self.apply_filters).grid(row=3, column=0, sticky="ew", padx=4, pady=4)
        ttk.Button(filters, text="Limpar", command=self.clear_filters).grid(row=3, column=1, sticky="ew", padx=4, pady=4)
        filters.columnconfigure(1, weight=1)

        actions = ttk.LabelFrame(parent, text="4. Cálculo e exportação")
        actions.pack(fill="x", pady=(0, 8))
        ttk.Button(actions, text="Calcular", command=self.run_design, style="Primary.TButton").grid(row=0, column=0, columnspan=2, sticky="ew", padx=4, pady=4)
        ttk.Button(actions, text="Exportar .xlsx", command=self.export_excel).grid(row=1, column=0, sticky="ew", padx=4, pady=4)
        ttk.Button(actions, text="Relatório .pdf", command=self.export_pdf_report).grid(row=1, column=1, sticky="ew", padx=4, pady=4)
        ttk.Button(actions, text="Exportar .csv", command=self.export_csv).grid(row=2, column=0, sticky="ew", padx=4, pady=4)
        ttk.Button(actions, text="Abrir repositório", command=lambda: webbrowser.open_new(GITHUB_URL)).grid(row=2, column=1, sticky="ew", padx=4, pady=4)
        actions.columnconfigure(0, weight=1)
        actions.columnconfigure(1, weight=1)

        progress_box = ttk.LabelFrame(parent, text="5. Estado")
        progress_box.pack(fill="x", pady=(0, 8))
        ttk.Label(progress_box, textvariable=self.status_var, wraplength=330, justify="left").pack(fill="x", padx=6, pady=(4, 2))
        ttk.Progressbar(progress_box, variable=self.progress_var, maximum=100).pack(fill="x", padx=6, pady=(2, 2))
        ttk.Label(progress_box, textvariable=self.progress_text_var, anchor="e").pack(fill="x", padx=6, pady=(0, 4))

        quick = ttk.LabelFrame(parent, text="6. Notas rápidas")
        quick.pack(fill="x", pady=(0, 8))
        quick_text = (
            "• Entrada tipo: Member/Node/Case + esforços + geometria.\n"
            "• Cada member/case deve ter 2 nós para cálculo completo.\n"
            "• Eixo do pilar: X local; flexões em MY e MZ.\n"
            "• FX é tratado como esforço axial de cálculo.\n"
            "• MX/torção e FY/FZ ainda não são verificados como dimensionamento final."
        )
        ttk.Label(quick, text=quick_text, wraplength=330, justify="left").pack(fill="x", padx=6, pady=6)

    def _build_main_tabs(self, parent):
        nb = ttk.Notebook(parent)
        nb.grid(row=0, column=0, sticky="nsew")

        self.tab_instructions = ttk.Frame(nb)
        self.tab_paste = ttk.Frame(nb)
        self.tab_input = ttk.Frame(nb)
        self.tab_pairs = ttk.Frame(nb)
        self.tab_validation = ttk.Frame(nb)
        self.tab_results = ttk.Frame(nb)
        self.tab_summary = ttk.Frame(nb)
        self.tab_failures = ttk.Frame(nb)
        self.tab_shortlists = ttk.Frame(nb)
        self.tab_report = ttk.Frame(nb)
        self.tab_notes = ttk.Frame(nb)

        for frame, title in [
            (self.tab_instructions, "Instruções"),
            (self.tab_paste, "Colar"),
            (self.tab_input, "Tabela"),
            (self.tab_pairs, "Pares"),
            (self.tab_validation, "Validação"),
            (self.tab_results, "Resultados"),
            (self.tab_summary, "Resumo"),
            (self.tab_failures, "Falhas"),
            (self.tab_shortlists, "Shortlists"),
            (self.tab_report, "Relatório"),
            (self.tab_notes, "Notas EC2"),
        ]:
            nb.add(frame, text=title)

        self._build_instructions_tab(self.tab_instructions)
        self._build_paste_tab(self.tab_paste)
        self.tree_input = self._make_tree(self.tab_input)
        self.tree_pairs = self._make_tree(self.tab_pairs)
        self.tree_validation = self._make_tree(self.tab_validation)
        self.tree_results = self._make_tree(self.tab_results)
        self.tree_summary = self._make_tree(self.tab_summary)
        self.tree_failures = self._make_tree(self.tab_failures)
        self.tree_shortlists = self._make_tree(self.tab_shortlists)
        self.report_txt = self._make_text_view(self.tab_report)
        self.tree_notes = self._make_tree(self.tab_notes)

    def _add_label_entry(self, parent, label, var, row):
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", padx=6, pady=4)
        ttk.Entry(parent, textvariable=var, width=14).grid(row=row, column=1, sticky="ew", padx=6, pady=4)

    def _build_instructions_tab(self, parent):
        outer = ttk.Frame(parent, padding=10)
        outer.pack(fill="both", expand=True)
        outer.rowconfigure(1, weight=1)
        outer.columnconfigure(0, weight=1)
        ttk.Label(outer, text="Instruções de importação da tabela de pilares", style="Header.TLabel").grid(row=0, column=0, sticky="w", pady=(0, 8))
        text_host = ttk.Frame(outer)
        text_host.grid(row=1, column=0, sticky="nsew")
        txt = self._make_text_view(text_host)
        instructions = (
            "OBJECTIVO\n"
            "Dimensionar/verificar pilares de betão armado a partir de esforços exportados do software de análise estrutural ou de uma folha Excel.\n\n"
            "COLUNAS DA FOLHA DE IMPORTAÇÃO TIPO\n"
            + " | ".join(self.TEMPLATE_COLUMNS) + "\n\n"
            "UNIDADES ESPERADAS\n"
            "FX, FY, FZ em kN; MX, MY, MZ em kNm; Length em m; dimensões HY/HZ em cm; áreas em cm²; inércias em cm⁴.\n\n"
            "CONVENÇÃO ADOPTADA\n"
            "O eixo longitudinal do pilar é o eixo local X. A flexão de dimensionamento é feita com MY e MZ. "
            "A coluna Member/Node/Case deve permitir identificar as duas extremidades de cada barra.\n\n"
            "EXEMPLO DE LINHA\n"
            "119/24/101 (C)    832,57    5,20    -83,34    -3,35    113,21    9,37    3,30    C40/50    50    50    ...    PF13\n\n"
            "NOTAS\n"
            "• Para cálculo completo, cada member/case deve aparecer com dois nós.\n"
            "• Caso só exista um nó, o programa assinala falha de dados.\n"
            "• A redução para casos governantes acelera o cálculo em tabelas grandes.\n"
            "• O relatório PDF é resumido; o Excel mantém a auditoria completa.\n"
        )
        txt.insert("1.0", instructions)
        txt.config(state="disabled")

    def _build_paste_tab(self, parent):
        top = ttk.Frame(parent, padding=6)
        top.pack(fill="x")
        ttk.Label(top, text="Cole aqui a tabela de esforços e clique em 'Ler caixa de texto'.").pack(side="left")
        ttk.Button(top, text="Ler caixa de texto", command=self.load_from_textbox).pack(side="right")
        ttk.Button(top, text="Limpar", command=lambda: self.txt_paste.delete("1.0", "end")).pack(side="right", padx=(0, 6))
        frame = ttk.Frame(parent, padding=(6, 0, 6, 6))
        frame.pack(fill="both", expand=True)
        self.txt_paste = tk.Text(frame, wrap="none", undo=True)
        vsb = ttk.Scrollbar(frame, orient="vertical", command=self.txt_paste.yview)
        hsb = ttk.Scrollbar(frame, orient="horizontal", command=self.txt_paste.xview)
        self.txt_paste.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.txt_paste.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)

    def _make_text_view(self, parent):
        frame = ttk.Frame(parent)
        # Avoid mixing Tk geometry managers in the same parent.
        # If the parent already uses grid, attach this container with grid; otherwise use pack.
        try:
            grid_slaves = parent.grid_slaves()
            pack_slaves = parent.pack_slaves()
        except Exception:
            grid_slaves = []
            pack_slaves = []

        if grid_slaves and not pack_slaves:
            next_row = max([int(w.grid_info().get("row", 0)) for w in grid_slaves] + [-1]) + 1
            parent.rowconfigure(next_row, weight=1)
            parent.columnconfigure(0, weight=1)
            frame.grid(row=next_row, column=0, sticky="nsew")
        else:
            frame.pack(fill="both", expand=True)

        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)
        txt = tk.Text(frame, wrap="word", undo=False, font=("Segoe UI", 9))
        vsb = ttk.Scrollbar(frame, orient="vertical", command=txt.yview)
        txt.configure(yscrollcommand=vsb.set)
        txt.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        return txt

    def _make_tree(self, parent):
        frame = ttk.Frame(parent)
        frame.pack(fill="both", expand=True)
        tree = ttk.Treeview(frame, show="headings")
        vsb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)
        return tree

    def show_df(self, tree: ttk.Treeview, df: pd.DataFrame):
        tree.delete(*tree.get_children())
        if df is None or df.empty:
            tree["columns"] = []
            return
        cols = list(df.columns)
        tree["columns"] = cols
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=max(85, min(220, len(str(c)) * 9)), anchor="center")
        for _, row in df.head(MAX_PREVIEW_ROWS).iterrows():
            vals = []
            for c in cols:
                v = row[c]
                if isinstance(v, float):
                    vals.append("" if not math.isfinite(v) else f"{v:.3f}")
                else:
                    vals.append("" if pd.isna(v) else str(v))
            tree.insert("", "end", values=vals)

    # --------------------------- importação ---------------------------
    def load_from_textbox(self):
        text = self.txt_paste.get("1.0", "end").strip()
        if not text:
            messagebox.showwarning("Aviso", "Cole primeiro a tabela na caixa de texto.")
            return
        df = parse_pasted_table(text)
        if df.empty:
            messagebox.showwarning("Aviso", "A tabela colada não foi reconhecida.")
            return
        self.load_df(df, source="tabela colada")

    def paste_clipboard(self):
        try:
            text = self.clipboard_get()
        except Exception:
            messagebox.showwarning("Aviso", "Não foi possível ler a área de transferência.")
            return
        df = parse_pasted_table(text)
        if df.empty:
            messagebox.showwarning("Aviso", "A tabela colada não foi reconhecida.")
            return
        self.txt_paste.delete("1.0", "end")
        self.txt_paste.insert("1.0", text)
        self.load_df(df, source="área de transferência")

    def import_file(self):
        path = filedialog.askopenfilename(title="Importar tabela", filetypes=[("Excel", "*.xlsx *.xls"), ("CSV", "*.csv"), ("Todos", "*.*")])
        if not path:
            return
        try:
            if path.lower().endswith((".xlsx", ".xls")):
                df = pd.read_excel(path, dtype=str)
            else:
                try:
                    df = pd.read_csv(path, dtype=str)
                except Exception:
                    df = pd.read_csv(path, sep=";", dtype=str)
            self.input_file_path = path
            self.load_df(df, source=os.path.basename(path))
        except Exception as err:
            messagebox.showerror("Erro", f"Não foi possível importar o ficheiro.\n\n{err}")

    def load_df(self, df: pd.DataFrame, source: str = ""):
        self.df_raw = df.copy()
        self.df_clean = clean_dataframe(df)
        if "material" in self.df_clean.columns:
            mask = self.df_clean["material"].astype(str).str.strip().isin(["", "nan", "None"])
            self.df_clean.loc[mask, "material"] = DEFAULT_CONCRETE_CLASS
        self.df_pair = combine_member_end_actions(self.df_clean)
        self.df_calc_input = pd.DataFrame()
        self.df_results = pd.DataFrame()
        self.df_summary = pd.DataFrame()
        self.df_failures = pd.DataFrame()
        self.df_ok = pd.DataFrame()
        self.df_filtered = pd.DataFrame()
        self.df_validation = self.build_data_validation(pre_calc=True)
        self.df_notes = self.build_normative_notes()
        self.show_df(self.tree_input, self.df_clean)
        self.show_df(self.tree_pairs, self.df_pair)
        self.show_df(self.tree_validation, self.df_validation)
        self.show_df(self.tree_results, self.df_results)
        self.show_df(self.tree_summary, self.df_summary)
        self.show_df(self.tree_failures, self.df_failures)
        self.show_df(self.tree_shortlists, pd.DataFrame())
        self.show_df(self.tree_notes, self.df_notes)
        self.update_report()
        self.progress_var.set(0.0)
        bad_pairs = int((self.df_pair.get("n_nodes_found", pd.Series(dtype=float)).fillna(0).astype(float) < 2).sum()) if not self.df_pair.empty else 0
        warn = f"; {bad_pairs} member/case sem dois nós" if bad_pairs else ""
        self.status_var.set(f"Tabela carregada ({source}): {len(self.df_clean)} linhas; {len(self.df_pair)} pares member/case{warn}.")

    def validate_inputs(self) -> Optional[str]:
        if self.df_pair is None or self.df_pair.empty:
            return "Cole ou importe uma tabela de esforços primeiro."
        c = safe_float(self.var_cover.get(), 0.0)
        if c <= 0:
            return "Recobrimento inválido."
        if safe_float(self.var_fyk.get(), 0.0) <= 0:
            return "Valor de fyk inválido."
        return None

    # --------------------------- validação e notas ---------------------------
    def build_data_validation(self, pre_calc: bool = False) -> pd.DataFrame:
        rows = []
        required = ["member_case", "fx", "my", "mz", "length", "material", "hy", "hz", "ax", "iy", "iz", "name"]
        df = self.df_clean
        for c in required:
            ok = df is not None and not df.empty and c in df.columns
            rows.append({"Categoria": "Colunas", "Item": c, "Estado": "OK" if ok else "Não conforme", "Resultado": "presente" if ok else "em falta", "Nota": "coluna reconhecida" if ok else "corrigir cabeçalho da folha de importação"})
        if df is not None and not df.empty:
            rows.append({"Categoria": "Tabela", "Item": "linhas", "Estado": "OK", "Resultado": len(df), "Nota": "linhas importadas"})
            rows.append({"Categoria": "Tabela", "Item": "members", "Estado": "OK", "Resultado": df["member"].astype(str).nunique() if "member" in df.columns else 0, "Nota": "barras distintas detectadas"})
        if self.df_pair is not None and not self.df_pair.empty:
            n_bad = int((self.df_pair["n_nodes_found"].fillna(0).astype(float) < 2).sum())
            rows.append({"Categoria": "Pares de nós", "Item": "member/case com 2 nós", "Estado": "OK" if n_bad == 0 else "Verificar", "Resultado": f"{len(self.df_pair)-n_bad}/{len(self.df_pair)}", "Nota": "cada member/case deve ter duas linhas, uma por nó"})
        if not pre_calc and self.df_results is not None and not self.df_results.empty:
            rows.append({"Categoria": "Cálculo", "Item": "resultados", "Estado": "OK", "Resultado": len(self.df_results), "Nota": "linhas calculadas"})
            n_fail = int((self.df_results["status"] == "Falha").sum()) if "status" in self.df_results.columns else 0
            rows.append({"Categoria": "Cálculo", "Item": "falhas", "Estado": "OK" if n_fail == 0 else "Verificar", "Resultado": n_fail, "Nota": "ver separador Falhas"})
        return pd.DataFrame(rows)

    def build_normative_notes(self) -> pd.DataFrame:
        notes = [
            ("Âmbito", "Pilares de betão armado", "Ferramenta orientada para ELU: compressão composta, segunda ordem simplificada e interação biaxial."),
            ("Eixos", "X local = eixo do pilar", "MY e MZ são tratados como momentos de flexão nos eixos principais locais da secção."),
            ("Entrada de dados", "Member/Node/Case", "Cada member/case deve conter duas linhas para permitir reconstruir M01/M02."),
            ("2.ª ordem", "Método da curvatura nominal", "Implementação prática baseada no EC2 5.8; confirmar comprimentos efectivos e fluência efectiva."),
            ("Esbelteza", "λlim explícito", "O programa reporta λ e λlim em Y e Z, indicando se a segunda ordem é dispensável ou deve ser considerada."),
            ("Biaxial", "Verificação numérica aproximada", "A verificação biaxial usa uma pesquisa da secção e critério de utilização; não substitui uma superfície N-My-Mz refinada."),
            ("Pormenorização", "Armadura longitudinal e estribos", "Controla As,min, As,max, espaçamento entre varões e espaçamento máximo de estribos."),
            ("Limitação", "FY/FZ/MX", "Nesta versão, esforço transverso e torção não são dimensionados de forma final; recomenda-se verificação complementar quando relevantes."),
        ]
        return pd.DataFrame(notes, columns=["Tema", "Critério adoptado", "Nota"])

    def build_summary_by_member(self, results: pd.DataFrame) -> pd.DataFrame:
        if results is None or results.empty:
            return pd.DataFrame()
        tmp = results.copy()
        tmp["utilizacao_sort"] = tmp["utilizacao"].fillna(999.0)
        tmp["as_req_sort"] = tmp["as_req_mm2"].fillna(-1.0)
        tmp = tmp.sort_values(by=["member", "name", "utilizacao_sort", "as_req_sort"], ascending=[True, True, False, False])
        return tmp.groupby(["member", "name"], as_index=False).first()

    def build_shortlists_df(self) -> pd.DataFrame:
        if self.df_results is None or self.df_results.empty:
            return pd.DataFrame()
        cols = ["member", "case", "name", "status", "failure_type", "shortlist_text", "recommendations"]
        return self.df_results[[c for c in cols if c in self.df_results.columns]].copy()

    # --------------------------- cálculo ---------------------------
    def run_design(self):
        err = self.validate_inputs()
        if err:
            messagebox.showwarning("Aviso", err)
            return

        designer = ColumnDesigner(
            cover_mm=safe_float(self.var_cover.get(), 35.0),
            fyk=safe_float(self.var_fyk.get(), 500.0),
            phi_eff=safe_float(self.var_phi_eff.get(), 2.0),
            l0y_factor=safe_float(self.var_l0y.get(), 1.0),
            l0z_factor=safe_float(self.var_l0z.get(), 1.0),
            calc_mode=self.var_calc_mode.get(),
        )

        input_df = reduce_to_governing_cases(self.df_pair) if self.var_reduce_cases.get() else self.df_pair.copy()
        self.df_calc_input = input_df.copy()
        self.progress_var.set(0.0)
        self.status_var.set("Análise em curso...")

        def progress(done, total):
            pct = 0.0 if total <= 0 else 100.0 * done / total
            self.after(0, lambda p=pct: self.progress_var.set(p))
            self.after(0, lambda d=done, t=total: self.status_var.set(f"A calcular... {d}/{t} casos member/case"))

        def worker():
            try:
                results = designer.design_dataframe(input_df, progress_callback=progress)
                summary = self.build_summary_by_member(results) if self.var_summary.get() else pd.DataFrame()
                failures = results[results["status"] == "Falha"].copy() if "status" in results.columns else pd.DataFrame()
                ok = results[results["status"] == "OK"].copy() if "status" in results.columns else pd.DataFrame()
                validation = self.build_data_validation(pre_calc=False)

                def finish():
                    self.df_results = results
                    self.df_summary = summary
                    self.df_failures = failures
                    self.df_ok = ok
                    self.df_filtered = pd.DataFrame()
                    self.df_validation = validation
                    self.df_notes = self.build_normative_notes()
                    self.show_df(self.tree_results, self.df_results)
                    self.show_df(self.tree_summary, self.df_summary)
                    self.show_df(self.tree_failures, self.df_failures)
                    self.show_df(self.tree_shortlists, self.build_shortlists_df())
                    self.show_df(self.tree_validation, self.df_validation)
                    self.show_df(self.tree_notes, self.df_notes)
                    self.update_report()
                    self.progress_var.set(100.0)
                    self.status_var.set(f"Cálculo concluído: {len(results)} casos; {len(summary)} membros resumidos; {len(failures)} falhas.")
                self.after(0, finish)
            except Exception as err:
                msg = str(err)
                self.after(0, lambda m=msg: messagebox.showerror("Erro", m))
                self.after(0, lambda: self.status_var.set("Falha na análise."))
                self.after(0, lambda: self.progress_var.set(0.0))

        self.analysis_thread = threading.Thread(target=worker, daemon=True)
        self.analysis_thread.start()

    # --------------------------- filtros / relatório ---------------------------
    def apply_filters(self):
        if self.df_results is None or self.df_results.empty:
            return
        df = self.df_results.copy()
        member = self.var_filter_member.get().strip()
        st = self.var_filter_status.get()
        ft = self.var_filter_fail.get()
        if member:
            df = df[df["member"].astype(str).str.contains(member, case=False, na=False)]
        if st != "Todos":
            df = df[df["status"] == st]
        if ft != "Todos" and "failure_type" in df.columns:
            df = df[df["failure_type"] == ft]
        self.df_filtered = df
        self.show_df(self.tree_results, df)
        self.status_var.set(f"Filtros aplicados: {len(df)} linhas visíveis.")

    def clear_filters(self):
        self.var_filter_member.set("")
        self.var_filter_status.set("Todos")
        self.var_filter_fail.set("Todos")
        self.df_filtered = pd.DataFrame()
        self.show_df(self.tree_results, self.df_results)
        self.status_var.set("Filtros removidos.")

    def update_report(self):
        self.report_txt.delete("1.0", "end")
        if self.df_results is None or self.df_results.empty:
            self.report_txt.insert("1.0", "Sem resultados. Importe a tabela e execute o cálculo.")
            return
        source = self.df_summary if self.df_summary is not None and not self.df_summary.empty else self.df_results
        n_total = len(self.df_results)
        n_ok = int((self.df_results["status"] == "OK").sum()) if "status" in self.df_results.columns else 0
        n_fail = int((self.df_results["status"] == "Falha").sum()) if "status" in self.df_results.columns else 0
        lines = []
        lines.append(f"{APP_NAME} {APP_VERSION}\n")
        lines.append("Relatório resumido de dimensionamento de pilares\n\n")
        lines.append(f"Casos analisados: {n_total} | OK: {n_ok} | Falhas: {n_fail}\n")
        lines.append(f"Recobrimento: {self.var_cover.get()} mm | fyk: {self.var_fyk.get()} MPa | φef: {self.var_phi_eff.get()}\n\n")
        for _, r in source.head(80).iterrows():
            lines.append(f"Membro {r.get('member','')} | Caso {r.get('case','')} | Nós {r.get('node_i','')} -> {r.get('node_j','')}\n")
            lines.append(f"  NEd = {safe_float(r.get('n_ed_kN',0),0):.2f} kN | My,Ed = {safe_float(r.get('my_ed_kNm',0),0):.2f} kNm | Mz,Ed = {safe_float(r.get('mz_ed_kNm',0),0):.2f} kNm\n")
            lines.append(f"  λy = {safe_float(r.get('lambda_y',0),0):.2f} / λlim,y = {safe_float(r.get('lambda_lim_y',0),0):.2f}; λz = {safe_float(r.get('lambda_z',0),0):.2f} / λlim,z = {safe_float(r.get('lambda_lim_z',0),0):.2f}\n")
            as_prov = r.get("as_prov_mm2", None)
            as_prov_str = "" if pd.isna(as_prov) else f"{float(as_prov):.0f}"
            lines.append(f"  As,min = {safe_float(r.get('as_min_mm2',0),0):.0f} mm² | As,req = {safe_float(r.get('as_req_mm2',0),0):.0f} mm² | As,prov = {as_prov_str} mm²\n")
            lines.append(f"  Solução: {r.get('solucao','')} | Estado: {r.get('status','')}\n")
            motivo = str(r.get("failure_reason", "") or "").strip()
            if motivo:
                lines.append(f"  Tipo de falha: {r.get('failure_type','')}\n")
                lines.append(f"  Motivo: {motivo}\n")
            rec = str(r.get("recommendations", "") or "").strip()
            if rec:
                lines.append(f"  Recomendações: {rec}\n")
            short = str(r.get("shortlist_text", "") or "").strip()
            if short:
                lines.append(f"  Shortlist: {short}\n")
            lines.append("\n")
        self.report_txt.insert("1.0", "".join(lines))

    # --------------------------- exportação ---------------------------
    def _metadata_df(self) -> pd.DataFrame:
        return pd.DataFrame([
            ["Programa", APP_NAME],
            ["Versão", APP_VERSION],
            ["Autor", APP_AUTHOR],
            ["Repositório", GITHUB_URL],
            ["Data de exportação", datetime.now().strftime("%Y-%m-%d %H:%M")],
            ["Ficheiro de origem", self.input_file_path or "-"] ,
            ["Norma de referência", "Eurocódigo 2 / NP EN 1992-1-1"],
            ["Âmbito", "Dimensionamento/verificação ELU de pilares de betão armado"],
            ["Descrição", APP_XLSX_DESCRIPTION],
            ["Limitações", "Torção, esforço transverso e ELS não estão verificados como dimensionamento final nesta versão."],
        ], columns=["Campo", "Valor"])

    def _parameters_df(self) -> pd.DataFrame:
        return pd.DataFrame([
            ["Recobrimento [mm]", self.var_cover.get()],
            ["Aço fyk [MPa]", self.var_fyk.get()],
            ["Betão", "lido da coluna Material; fallback interno C30/37 se ausente"],
            ["φef", self.var_phi_eff.get()],
            ["l0y/L", self.var_l0y.get()],
            ["l0z/L", self.var_l0z.get()],
            ["Modo de cálculo", self.var_calc_mode.get()],
            ["Redução para casos governantes", "Sim" if self.var_reduce_cases.get() else "Não"],
        ], columns=["Parâmetro", "Valor"])

    def export_excel(self):
        if self.df_results is None or self.df_results.empty:
            messagebox.showwarning("Aviso", "Não há resultados para exportar.")
            return
        path = filedialog.asksaveasfilename(title="Exportar resultados", defaultextension=".xlsx", filetypes=[("Excel workbook", "*.xlsx")])
        if not path:
            return
        root, ext = os.path.splitext(path)
        if ext.lower() != ".xlsx":
            path = root + ".xlsx" if ext else path + ".xlsx"
        try:
            self._write_excel(path)
            self.status_var.set(f"Resultados exportados para: {path}")
        except Exception as err:
            messagebox.showerror("Erro", f"Não foi possível exportar.\n\n{err}")

    def _write_excel(self, path: str):
        sheets = {
            "00_Info": self._metadata_df(),
            "01_Parametros": self._parameters_df(),
            "02_Entrada_Dados": self.df_clean,
            "03_Pares_Member_Case": self.df_pair,
            "04_Casos_Calculo": self.df_calc_input,
            "05_Resultados": self.df_results,
            "06_Resumo_Membros": self.df_summary,
            "07_Falhas": self.df_failures,
            "08_OK": self.df_ok,
            "09_Shortlists": self.build_shortlists_df(),
            "10_Validacao": self.df_validation,
            "11_Notas_EC2": self.df_notes,
        }
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            for name, df in sheets.items():
                if df is None:
                    df = pd.DataFrame()
                df.to_excel(writer, sheet_name=name[:31], index=False)
            wb = writer.book
            props = wb.properties
            props.title = APP_NAME
            props.subject = APP_SUBJECT
            props.creator = APP_AUTHOR
            props.keywords = APP_KEYWORDS
            props.category = APP_CATEGORY
            props.description = APP_XLSX_DESCRIPTION
            props.lastModifiedBy = APP_AUTHOR
            try:
                from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
                from openpyxl.utils import get_column_letter
                header_fill = PatternFill("solid", fgColor="1F4E5F")
                header_font = Font(color="FFFFFF", bold=True)
                light_fill = PatternFill("solid", fgColor="EAF2F5")
                thin = Side(style="thin", color="D9E2E7")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                for ws in wb.worksheets:
                    ws.sheet_view.showGridLines = False
                    ws.freeze_panes = "A2"
                    if ws.max_row >= 1:
                        for cell in ws[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                            cell.border = border
                    for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 5000)):
                        for cell in row:
                            cell.border = border
                            cell.alignment = Alignment(vertical="top", wrap_text=True)
                    for col_idx, col in enumerate(ws.columns, start=1):
                        values = [str(c.value) for c in col[:200] if c.value is not None]
                        width = min(max([len(v) for v in values] + [10]) + 2, 42)
                        ws.column_dimensions[get_column_letter(col_idx)].width = width
                    if ws.title == "00_Info":
                        ws["A1"].fill = light_fill
                        ws["B1"].fill = light_fill
            except Exception:
                pass

    def export_pdf_report(self):
        if self.df_results is None or self.df_results.empty:
            messagebox.showwarning("Aviso", "Não há resultados para exportar.")
            return
        path = filedialog.asksaveasfilename(title="Exportar relatório PDF", defaultextension=".pdf", filetypes=[("PDF", "*.pdf")])
        if not path:
            return
        root, ext = os.path.splitext(path)
        if ext.lower() != ".pdf":
            path = root + ".pdf" if ext else path + ".pdf"
        try:
            self._write_pdf(path)
            self.status_var.set(f"Relatório PDF exportado para: {path}")
        except Exception as err:
            messagebox.showerror("Erro", f"Não foi possível exportar PDF.\n\n{err}")

    def _write_pdf(self, path: str):
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak

        doc = SimpleDocTemplate(path, pagesize=landscape(A4), rightMargin=12*mm, leftMargin=12*mm, topMargin=12*mm, bottomMargin=12*mm)
        doc.title = APP_NAME
        doc.author = APP_AUTHOR
        doc.subject = APP_SUBJECT
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name="TitleCenter", parent=styles["Title"], alignment=TA_CENTER, fontName="Helvetica-Bold", fontSize=16, leading=20, textColor=colors.HexColor("#1F4E5F")))
        styles.add(ParagraphStyle(name="Small", parent=styles["BodyText"], fontSize=7, leading=9, alignment=TA_LEFT))
        styles.add(ParagraphStyle(name="H2Custom", parent=styles["Heading2"], fontSize=11, leading=14, textColor=colors.HexColor("#1F4E5F")))
        story = []
        story.append(Paragraph(APP_NAME, styles["TitleCenter"]))
        story.append(Paragraph("Relatório resumido de dimensionamento de pilares segundo o Eurocódigo 2", styles["BodyText"]))
        story.append(Spacer(1, 5*mm))
        n_total = len(self.df_results)
        n_ok = int((self.df_results["status"] == "OK").sum()) if "status" in self.df_results.columns else 0
        n_fail = int((self.df_results["status"] == "Falha").sum()) if "status" in self.df_results.columns else 0
        meta = [
            ["Programa", f"{APP_NAME} {APP_VERSION}", "Autor", APP_AUTHOR],
            ["Data", datetime.now().strftime("%Y-%m-%d %H:%M"), "Casos analisados", str(n_total)],
            ["OK", str(n_ok), "Falhas", str(n_fail)],
            ["Recobrimento", f"{self.var_cover.get()} mm", "fyk", f"{self.var_fyk.get()} MPa"],
        ]
        t = Table(meta, colWidths=[38*mm, 90*mm, 38*mm, 90*mm])
        t.setStyle(self._pdf_table_style(header=False))
        story.append(t)
        story.append(Spacer(1, 6*mm))
        story.append(Paragraph("Resumo por membro", styles["H2Custom"]))
        summary = self.df_summary if self.df_summary is not None and not self.df_summary.empty else self.df_results
        cols = ["member", "case", "name", "n_ed_kN", "my_ed_kNm", "mz_ed_kNm", "as_req_mm2", "as_prov_mm2", "solucao", "status"]
        story.append(self._pdf_df_table(summary, cols, max_rows=28))
        if self.df_failures is not None and not self.df_failures.empty:
            story.append(PageBreak())
            story.append(Paragraph("Falhas e recomendações", styles["H2Custom"]))
            fcols = ["member", "case", "name", "failure_type", "failure_reason", "recommendations"]
            story.append(self._pdf_df_table(self.df_failures, fcols, max_rows=40, small=True))
        story.append(Spacer(1, 5*mm))
        story.append(Paragraph("Notas: o relatório PDF é uma síntese. O ficheiro Excel exportado contém a tabela completa, validação, pares de nós, shortlists e notas EC2.", styles["Small"]))

        def footer(canvas, doc_obj):
            canvas.saveState()
            canvas.setAuthor(APP_AUTHOR)
            canvas.setTitle(APP_NAME)
            canvas.setSubject(APP_SUBJECT)
            canvas.setFont("Helvetica", 7)
            canvas.setFillColor(colors.grey)
            canvas.drawString(12*mm, 7*mm, f"{APP_NAME} {APP_VERSION} | {APP_AUTHOR}")
            canvas.drawRightString(285*mm, 7*mm, f"Página {doc_obj.page}")
            canvas.restoreState()
        doc.build(story, onFirstPage=footer, onLaterPages=footer)

    def _pdf_table_style(self, header=True):
        from reportlab.lib import colors
        from reportlab.platypus import TableStyle
        cmds = [
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D9E2E7")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ]
        if header:
            cmds += [("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E5F")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold")]
        return TableStyle(cmds)

    def _pdf_df_table(self, df: pd.DataFrame, cols: List[str], max_rows: int = 30, small: bool = False):
        from reportlab.platypus import Table, Paragraph
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        styles = getSampleStyleSheet()
        pstyle = ParagraphStyle(name="Cell", parent=styles["BodyText"], fontSize=6 if small else 7, leading=8 if small else 9)
        present = [c for c in cols if c in df.columns]
        data = [[Paragraph(str(c), pstyle) for c in present]]
        for _, r in df.head(max_rows).iterrows():
            row = []
            for c in present:
                v = r.get(c, "")
                if isinstance(v, float):
                    txt = "" if not math.isfinite(v) else f"{v:.2f}"
                else:
                    txt = "" if pd.isna(v) else str(v)
                txt = txt.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                row.append(Paragraph(txt, pstyle))
            data.append(row)
        total_width = 270 * mm
        widths = [total_width / max(1, len(present))] * max(1, len(present))
        t = Table(data, colWidths=widths, repeatRows=1)
        t.setStyle(self._pdf_table_style(header=True))
        return t

    def export_csv(self):
        if self.df_results is None or self.df_results.empty:
            messagebox.showwarning("Aviso", "Não há resultados para exportar.")
            return
        path = filedialog.asksaveasfilename(title="Exportar CSV", defaultextension=".csv", filetypes=[("CSV", "*.csv")])
        if not path:
            return
        self.df_results.to_csv(path, index=False, encoding="utf-8-sig")
        self.status_var.set(f"CSV exportado para: {path}")

    def export_template(self):
        path = filedialog.asksaveasfilename(title="Guardar modelo de importação", defaultextension=".xlsx", filetypes=[("Excel workbook", "*.xlsx")])
        if not path:
            return
        root, ext = os.path.splitext(path)
        if ext.lower() != ".xlsx":
            path = root + ".xlsx" if ext else path + ".xlsx"
        sample = pd.DataFrame([{
            "Member/Node/Case": "119/24/101 (C)", "FX (kN)": "832,57", "FY (kN)": "5,20", "FZ (kN)": "-83,34", "MX (kNm)": "-3,35", "MY (kNm)": "113,21", "MZ (kNm)": "9,37",
            "Length (m)": "3,30", "Material": "C40/50", "HY (cm)": "50", "HZ (cm)": "50", "VY (cm)": "25", "VZ (cm)": "25", "VPY (cm)": "25", "VPZ (cm)": "25",
            "AX (cm2)": "2500", "AY (cm2)": "2083,33", "AZ (cm2)": "2083,33", "IX (cm4)": "878644,3", "IY (cm4)": "520833,33", "IZ (cm4)": "520833,33", "Name": "PF13"
        }])
        try:
            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                sample.to_excel(writer, sheet_name="FOLHA_IMPORTACAO_TIPO", index=False)
                ws = writer.book["FOLHA_IMPORTACAO_TIPO"]
                writer.book.properties.title = f"{APP_NAME} - folha de importação tipo"
                writer.book.properties.creator = APP_AUTHOR
                try:
                    from openpyxl.styles import Font, PatternFill, Alignment
                    fill = PatternFill("solid", fgColor="1F4E5F")
                    for cell in ws[1]:
                        cell.fill = fill
                        cell.font = Font(color="FFFFFF", bold=True)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    for col in ws.columns:
                        ws.column_dimensions[col[0].column_letter].width = max(12, min(22, len(str(col[0].value)) + 2))
                except Exception:
                    pass
            self.status_var.set(f"Modelo de tabela guardado: {path}")
        except Exception as err:
            messagebox.showerror("Erro", f"Não foi possível guardar o modelo.\n\n{err}")


# ============================================================
# Overrides v3 — relatório PDF Courier, XLSX enriquecido e notas EC2
# ============================================================
def _metadata_df_v3(self) -> pd.DataFrame:
    return pd.DataFrame([
        ["Programa", APP_NAME], ["Versão", APP_VERSION], ["Autor", APP_AUTHOR], ["Repositório", GITHUB_URL],
        ["Data de exportação", datetime.now().strftime("%Y-%m-%d %H:%M")], ["Ficheiro de origem", self.input_file_path or "-"],
        ["Norma de referência", "Eurocódigo 2 / NP EN 1992-1-1"],
        ["Âmbito", "ELU, verificações informativas ELS, esforço transverso e torção em pilares de betão armado"],
        ["Descrição", APP_XLSX_DESCRIPTION],
        ["Limitações", "ELS, esforço transverso e torção são verificações práticas/informativas; validar casos críticos com cálculo dedicado."],
    ], columns=["Campo", "Valor"])


def _parameters_df_v3(self) -> pd.DataFrame:
    return pd.DataFrame([
        ["Recobrimento [mm]", self.var_cover.get()], ["Aço fyk [MPa]", self.var_fyk.get()],
        ["Betão", "lido da coluna Material; fallback interno C30/37 quando ausente"],
        ["φef", self.var_phi_eff.get()], ["l0y/L", self.var_l0y.get()], ["l0z/L", self.var_l0z.get()],
        ["Modo de cálculo", self.var_calc_mode.get()], ["Redução para casos governantes", "Sim" if self.var_reduce_cases.get() else "Não"],
    ], columns=["Parâmetro", "Valor"])


def _build_normative_notes_v3(self) -> pd.DataFrame:
    notes = [
        ("Âmbito", "Pilares de betão armado", "Ferramenta orientada para ELU, com verificações complementares ELS, V e torção."),
        ("Material", "Classe da tabela", "A classe de betão é lida da coluna Material; C30/37 é apenas fallback interno se estiver ausente."),
        ("Entrada de dados", "Member/Node/Case", "Cada member/case deve conter duas linhas para reconstruir M01/M02."),
        ("2.ª ordem", "Curvatura nominal", "Inclui esbelteza-limite, imperfeição geométrica prática e fluência efectiva introduzida pelo utilizador."),
        ("Esbelteza", "λ vs λlim", "A segunda ordem é dispensada quando λ <= λlim no eixo considerado."),
        ("Biaxial", "Pesquisa numérica aproximada", "A verificação usa discretização da secção e critério de utilização; para casos críticos recomenda-se superfície N-My-Mz refinada."),
        ("ELS", "Tensões elásticas", "Verifica tensões em serviço de forma simplificada e identifica o número da combinação relevante."),
        ("Esforço transverso", "VRd,c prático", "Verificação informativa segundo forma simplificada; quando falha, dimensionar armadura transversal por EC2 6.2."),
        ("Torção", "Aviso por MX", "A torção é assinalada quando relevante; dimensionamento dedicado por EC2 6.3 deve ser feito separadamente."),
        ("Pormenorização", "Armadura longitudinal e estribos", "Controla As,min, As,max, espaçamentos e limite prático de varões por face."),
    ]
    return pd.DataFrame(notes, columns=["Tema", "Critério adoptado", "Nota"])


def _write_excel_v3(self, path: str):
    els_cols = [c for c in ["member","case","combination_number","limit_state","service_status","service_sigma_c_max_MPa","service_sigma_c_min_MPa","service_sigma_s_max_MPa","service_sigma_c_lim_MPa","service_sigma_s_lim_MPa","service_crack_warning"] if c in self.df_results.columns]
    vt_cols = [c for c in ["member","case","v_ed_y_kN","v_rd_c_y_kN","shear_status_y","v_ed_z_kN","v_rd_c_z_kN","shear_status_z","mx_ed_kNm","torsion_ratio","torsion_status"] if c in self.df_results.columns]
    sheets = {
        "00_Info": self._metadata_df(), "01_Parametros": self._parameters_df(), "02_Entrada_Dados": self.df_clean,
        "03_Pares_Member_Case": self.df_pair, "04_Casos_Calculo": self.df_calc_input, "05_Resultados": self.df_results,
        "06_Resumo_Membros": self.df_summary, "07_Falhas": self.df_failures, "08_OK": self.df_ok,
        "09_Shortlists": self.build_shortlists_df(), "10_ELS": self.df_results[els_cols].copy() if els_cols else pd.DataFrame(),
        "11_V_Torcao": self.df_results[vt_cols].copy() if vt_cols else pd.DataFrame(),
        "12_Validacao": self.df_validation, "13_Notas_EC2": self.df_notes,
    }
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in sheets.items():
            if df is None: df = pd.DataFrame()
            df.to_excel(writer, sheet_name=name[:31], index=False)
        wb = writer.book
        props = wb.properties
        props.title = APP_NAME; props.subject = APP_SUBJECT; props.creator = APP_AUTHOR; props.keywords = APP_KEYWORDS
        props.category = APP_CATEGORY; props.description = APP_XLSX_DESCRIPTION; props.lastModifiedBy = APP_AUTHOR
        try:
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            from openpyxl.utils import get_column_letter
            header_fill = PatternFill("solid", fgColor="1F4E5F"); header_font = Font(color="FFFFFF", bold=True)
            light_fill = PatternFill("solid", fgColor="EAF2F5"); thin = Side(style="thin", color="D9E2E7")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for ws in wb.worksheets:
                ws.sheet_view.showGridLines = False; ws.freeze_panes = "A2"
                if ws.max_row >= 1:
                    for cell in ws[1]:
                        cell.fill = header_fill; cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); cell.border = border
                for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 5000)):
                    for cell in row:
                        cell.border = border; cell.alignment = Alignment(vertical="top", wrap_text=True)
                for col_idx, col in enumerate(ws.columns, start=1):
                    values = [str(c.value) for c in col[:200] if c.value is not None]
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(max([len(v) for v in values] + [10]) + 2, 46)
                if ws.title == "00_Info":
                    ws["A1"].fill = light_fill; ws["B1"].fill = light_fill
        except Exception:
            pass


def _pdf_styles_v3():
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="ReportTitle", parent=styles["Title"], alignment=TA_CENTER, fontName="Courier-Bold", fontSize=14, leading=21, spaceAfter=10))
    styles.add(ParagraphStyle(name="ReportSubtitle", parent=styles["Normal"], alignment=TA_CENTER, fontName="Courier", fontSize=10, leading=15, textColor=colors.darkgrey, spaceAfter=8))
    styles.add(ParagraphStyle(name="BodyCourier", parent=styles["Normal"], fontName="Courier", fontSize=10, leading=15, spaceAfter=6))
    styles.add(ParagraphStyle(name="Small", parent=styles["Normal"], fontName="Courier", fontSize=8, leading=12))
    styles.add(ParagraphStyle(name="Cell", parent=styles["Small"], alignment=TA_LEFT, fontName="Courier", fontSize=7, leading=10.5))
    styles.add(ParagraphStyle(name="Section", parent=styles["Heading2"], fontName="Courier-Bold", fontSize=12, leading=18, spaceBefore=10, spaceAfter=20))
    return styles


def _pdf_table_style_v3(self, header=True):
    from reportlab.lib import colors
    from reportlab.platypus import TableStyle
    cmds = [("GRID",(0,0),(-1,-1),0.25,colors.lightgrey),("VALIGN",(0,0),(-1,-1),"TOP"),("FONTNAME",(0,0),(-1,-1),"Courier"),("FONTSIZE",(0,0),(-1,-1),7),("LEFTPADDING",(0,0),(-1,-1),3),("RIGHTPADDING",(0,0),(-1,-1),3)]
    if header:
        cmds += [("BACKGROUND",(0,0),(-1,0),colors.HexColor("#EFEFEF")), ("FONTNAME",(0,0),(-1,0),"Courier-Bold")]
    return TableStyle(cmds)


def _pdf_df_table_v3(self, df: pd.DataFrame, cols: List[str], max_rows: int = 30, small: bool = False):
    from reportlab.platypus import Table, Paragraph
    from reportlab.lib.units import mm
    styles = _pdf_styles_v3(); pstyle = styles["Cell"]
    present = [c for c in cols if c in df.columns]
    data = [[Paragraph(str(c), pstyle) for c in present]]
    for _, r in df.head(max_rows).iterrows():
        row=[]
        for c in present:
            v=r.get(c,"")
            if isinstance(v,float): txt="" if not math.isfinite(v) else f"{v:.2f}"
            else: txt="" if pd.isna(v) else str(v)
            txt=txt.replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")
            row.append(Paragraph(txt,pstyle))
        data.append(row)
    total_width=270*mm; widths=[total_width/max(1,len(present))]*max(1,len(present))
    t=Table(data,colWidths=widths,repeatRows=1); t.setStyle(self._pdf_table_style(header=True)); return t


def _write_pdf_v3(self, path: str):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, PageBreak
    styles = _pdf_styles_v3()
    doc = SimpleDocTemplate(path, pagesize=landscape(A4), rightMargin=12*mm, leftMargin=12*mm, topMargin=12*mm, bottomMargin=12*mm)
    doc.title=APP_NAME; doc.author=APP_AUTHOR; doc.subject=APP_SUBJECT
    story=[]
    story.append(Paragraph(APP_NAME, styles["ReportTitle"]))
    story.append(Paragraph("Dimensionamento e verificação de pilares segundo o Eurocódigo 2", styles["ReportSubtitle"]))
    n_total=len(self.df_results); n_ok=int((self.df_results["status"]=="OK").sum()) if "status" in self.df_results.columns else 0; n_fail=int((self.df_results["status"]=="Falha").sum()) if "status" in self.df_results.columns else 0
    meta=[["Programa",f"{APP_NAME} {APP_VERSION}","Autor",APP_AUTHOR],["Data",datetime.now().strftime("%Y-%m-%d %H:%M"),"Casos",str(n_total)],["OK",str(n_ok),"Falhas",str(n_fail)],["Betão","lido da tabela","fyk",f"{self.var_fyk.get()} MPa"]]
    t=Table(meta,colWidths=[38*mm,90*mm,38*mm,90*mm]); t.setStyle(self._pdf_table_style(header=False)); story.append(t); story.append(Spacer(1,5*mm))
    story.append(Paragraph("Resumo por membro", styles["Section"]))
    summary=self.df_summary if self.df_summary is not None and not self.df_summary.empty else self.df_results
    cols=["member","case","combination_number","limit_state","n_ed_kN","my_ed_kNm","mz_ed_kNm","as_req_mm2","as_prov_mm2","solucao","status"]
    story.append(self._pdf_df_table(summary,cols,max_rows=30))
    if self.df_failures is not None and not self.df_failures.empty:
        story.append(PageBreak()); story.append(Paragraph("Falhas e recomendações", styles["Section"])); story.append(self._pdf_df_table(self.df_failures,["member","case","failure_type","failure_reason","recommendations"],max_rows=45,small=True))
    els_cols=[c for c in ["member","case","combination_number","service_status","service_sigma_c_max_MPa","service_sigma_s_max_MPa","service_crack_warning"] if c in self.df_results.columns]
    if els_cols:
        story.append(PageBreak()); story.append(Paragraph("Verificações em serviço", styles["Section"])); story.append(self._pdf_df_table(self.df_results,els_cols,max_rows=45,small=True))
    vt_cols=[c for c in ["member","case","v_ed_y_kN","v_rd_c_y_kN","shear_status_y","v_ed_z_kN","v_rd_c_z_kN","shear_status_z","mx_ed_kNm","torsion_status"] if c in self.df_results.columns]
    if vt_cols:
        story.append(PageBreak()); story.append(Paragraph("Esforço transverso e torção", styles["Section"])); story.append(self._pdf_df_table(self.df_results,vt_cols,max_rows=45,small=True))
    story.append(Spacer(1,5*mm)); story.append(Paragraph("Notas: relatório sintético. O ficheiro Excel contém resultados completos, shortlists, ELS, esforço transverso, torção e metadados.", styles["Small"]))
    def footer(canvas, doc_obj):
        canvas.saveState(); canvas.setAuthor(APP_AUTHOR); canvas.setTitle(APP_NAME); canvas.setSubject(APP_SUBJECT); canvas.setFont("Courier",7); canvas.setFillColor(colors.grey); canvas.drawString(12*mm,7*mm,f"{APP_NAME} {APP_VERSION} | {APP_AUTHOR}"); canvas.drawRightString(285*mm,7*mm,f"Página {doc_obj.page}"); canvas.restoreState()
    doc.build(story,onFirstPage=footer,onLaterPages=footer)

ColumnsEC2App._metadata_df = _metadata_df_v3
ColumnsEC2App._parameters_df = _parameters_df_v3
ColumnsEC2App.build_normative_notes = _build_normative_notes_v3
ColumnsEC2App._write_excel = _write_excel_v3
ColumnsEC2App._pdf_table_style = _pdf_table_style_v3
ColumnsEC2App._pdf_df_table = _pdf_df_table_v3
ColumnsEC2App._write_pdf = _write_pdf_v3




def _reduce_to_governing_cases_v3(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    work = df.copy()
    for c in ["fx", "fy", "fz", "mx", "my", "mz"]:
        if c not in work.columns:
            work[c] = 0.0
    work["_abs_fx"] = work["fx"].abs().fillna(0.0)
    work["_abs_fy"] = work["fy"].abs().fillna(0.0)
    work["_abs_fz"] = work["fz"].abs().fillna(0.0)
    work["_abs_mx"] = work["mx"].abs().fillna(0.0)
    work["_abs_my"] = work["my"].abs().fillna(0.0)
    work["_abs_mz"] = work["mz"].abs().fillna(0.0)
    work["_score"] = work["_abs_fx"] * 0.20 + work["_abs_my"] + work["_abs_mz"] + 0.35 * work["_abs_mx"] + 0.10 * (work["_abs_fy"] + work["_abs_fz"])
    selected_idx = set()
    for _, grp in work.groupby(["member", "name"], dropna=False):
        if grp.empty:
            continue
        for col in ["_score", "_abs_fx", "_abs_fy", "_abs_fz", "_abs_mx", "_abs_my", "_abs_mz"]:
            selected_idx.add(grp[col].idxmax())
        grp2 = grp.assign(
            _my_over_n=grp["_abs_my"] / grp["_abs_fx"].replace(0.0, 1e-9),
            _mz_over_n=grp["_abs_mz"] / grp["_abs_fx"].replace(0.0, 1e-9),
            _v_over_n=(grp["_abs_fy"] + grp["_abs_fz"]) / grp["_abs_fx"].replace(0.0, 1e-9),
        )
        for col in ["_my_over_n", "_mz_over_n", "_v_over_n"]:
            selected_idx.add(grp2[col].idxmax())
        els_mask = grp["case"].astype(str).str.upper().str.contains("ELS|SLS|SERV|SERVICE|RARA|FREQ|QUASE|QP|\\(S\\)", regex=True, na=False)
        if els_mask.any():
            els_grp = grp[els_mask]
            selected_idx.add(els_grp["_score"].idxmax())
    reduced = work.loc[sorted(selected_idx)].copy().sort_values(["member", "name", "case"]).reset_index(drop=True)
    reduced.drop(columns=[c for c in reduced.columns if c.startswith("_")], inplace=True, errors="ignore")
    return reduced

reduce_to_governing_cases = _reduce_to_governing_cases_v3



# ============================================================
# v3.1 - Melhorias EC2 de cálculo e auditoria
# ============================================================
def order_end_moments_ec2(m_i, m_j):
    """Devolve (M01, M02, rm, tipo_curvatura) com M02 positivo e |M02| >= |M01|.
    M01 preserva o sinal relativo face a M02, para distinguir curvatura simples/dupla.
    """
    a = safe_float(m_i, 0.0)
    b = safe_float(m_j, 0.0)
    if abs(a) > abs(b):
        m02_raw, m01_raw = a, b
    else:
        m02_raw, m01_raw = b, a
    sign = 1.0 if m02_raw >= 0 else -1.0
    m02 = abs(m02_raw)
    m01 = m01_raw * sign
    rm = 1.0 if m02 < 1e-9 else max(-1.0, min(1.0, m01 / m02))
    curvature = "dupla" if rm < 0 else "simples"
    return m01, m02, rm, curvature


def biaxial_alpha_ec2_practical(n_ed_kN, ac_mm2, fcd, as_mm2, fyd):
    """Expoente prático para flexão desviada EC2 5.8.9.
    Usa o nível axial relativo para variar entre 1.0 e 2.0.
    """
    nrd = max(ac_mm2 * fcd + as_mm2 * fyd, 1e-9) / 1000.0
    nu = max(0.0, min(1.0, abs(n_ed_kN) / nrd))
    if nu <= 0.10:
        alpha = 1.00
    elif nu >= 0.70:
        alpha = 2.00
    else:
        alpha = 1.00 + (nu - 0.10) / 0.60
    return alpha, nu


def second_order_nominal_curvature_v31(
    self,
    n_ed_kN: float,
    m01_kNm: float,
    m02_kNm: float,
    l0_mm: float,
    ac_mm2: float,
    i_mm4: float,
    h_mm: float,
    as_total_mm2: float,
    fcd: float,
    fck: float,
    fyd: float,
    es: float,
    phi_eff: Optional[float] = None,
):
    if phi_eff is None:
        phi_eff = self.phi_eff
    if ac_mm2 <= 0 or i_mm4 <= 0 or h_mm <= 0 or fcd <= 0 or es <= 0:
        m0ed = max(abs(m01_kNm), abs(m02_kNm))
        return m0ed, 0.0, 0.0, 1.0, False, m0ed, 0.0

    i_radius = math.sqrt(i_mm4 / ac_mm2)
    slenderness = l0_mm / i_radius if i_radius > 0 else 0.0
    m01s, m02s, rm, curvature = order_end_moments_ec2(m01_kNm, m02_kNm)

    m0e_signed = 0.6 * m02s + 0.4 * m01s
    m0e = max(abs(m0e_signed), 0.4 * abs(m02s))
    m0ed = max(abs(m02s), m0e)

    omega = (as_total_mm2 * fyd) / max(ac_mm2 * fcd, 1e-9)
    n_red = (n_ed_kN * 1e3) / max(ac_mm2 * fcd, 1e-9)
    nu = 1.0 + omega
    nbal = 0.4
    denom = max(nu - nbal, 1e-9)
    kr = min(max((nu - n_red) / denom, 0.0), 1.0)

    beta = 0.35 + fck / 200.0 - slenderness / 150.0
    kphi = max(1.0 + beta * max(phi_eff, 0.0), 1.0)

    eps_yd = fyd / es
    d_eff = 0.8 * h_mm
    inv_r0 = eps_yd / max(0.45 * d_eff, 1e-9)
    inv_r = kr * kphi * inv_r0

    # c=10 é a hipótese geral do método da curvatura nominal; c=8 é conservador para momento constante.
    c_coeff = 8.0 if rm > 0.85 else 10.0
    e2_mm = inv_r * (l0_mm ** 2) / c_coeff
    m2_kNm = n_ed_kN * e2_mm / 1000.0
    m_ed = m0ed + m2_kNm
    return m_ed, slenderness, inv_r, kphi, bool(m2_kNm > 1e-9), m0ed, m2_kNm


ColumnDesigner.second_order_nominal_curvature = second_order_nominal_curvature_v31


def capacity_for_layout_v31(self, layout: Layout, n_ed_kN: float, fcd: float, fyd: float, Es: float):
    """Superfície resistente N-My-Mz por varrimento angular e resolução de N≈NEd por bissecção.
    Continua simplificada, mas substitui a grelha bruta por uma procura mais estável da linha neutra.
    """
    key = (
        "v31", round(layout.b_mm, 1), round(layout.h_mm, 1), layout.phi_long_mm, layout.phi_st_mm,
        layout.n_bars_y, layout.n_bars_z, round(n_ed_kN, 0), round(fcd, 3), round(fyd, 3)
    )
    if key in self._capacity_cache:
        return self._capacity_cache[key]

    capacities = []
    angles = [i * math.pi / 72.0 for i in range(37)]  # 0º a 90º, passo 2.5º
    c_min = 2.0
    c_max = 3.0 * max(layout.b_mm, layout.h_mm)

    for ang in angles:
        def n_at(c):
            N, My, Mz = self.section_response(layout, n_ed_kN, ang, c, fcd, fyd, Es)
            return N, My, Mz

        lo, hi = c_min, c_max
        n_lo, _, _ = n_at(lo)
        n_hi, _, _ = n_at(hi)
        best = None

        if (n_lo - n_ed_kN) * (n_hi - n_ed_kN) <= 0:
            for _ in range(48):
                mid = 0.5 * (lo + hi)
                N, My, Mz = n_at(mid)
                diff = N - n_ed_kN
                if best is None or abs(diff) < best[0]:
                    best = (abs(diff), My, Mz)
                if abs(diff) < 1e-3:
                    break
                if (n_lo - n_ed_kN) * diff <= 0:
                    hi = mid
                    n_hi = N
                else:
                    lo = mid
                    n_lo = N
        else:
            # fallback robusto se não houver mudança de sinal
            for c in [c_min + i * (c_max - c_min) / 120.0 for i in range(121)]:
                N, My, Mz = n_at(c)
                diff = abs(N - n_ed_kN)
                if best is None or diff < best[0]:
                    best = (diff, My, Mz)

        if best is not None:
            capacities.append((best[1], best[2]))

    self._capacity_cache[key] = capacities
    return capacities


ColumnDesigner.capacity_for_layout = capacity_for_layout_v31


def biaxial_ok_v31(self, my_ed_kNm: float, mz_ed_kNm: float, capacities: List[Tuple[float, float]]):
    """Flexão desviada: usa superfície discreta N-My-Mz e critério tipo EC2 5.8.9.
    O expoente alpha é definido em design_one e fica em self._biaxial_alpha.
    """
    if not capacities:
        return False, None, None, None
    my_req = abs(my_ed_kNm)
    mz_req = abs(mz_ed_kNm)
    alpha = float(getattr(self, "_biaxial_alpha", 2.0) or 2.0)
    best_util = None
    best_my = None
    best_mz = None
    for my_cap, mz_cap in capacities:
        if my_cap <= 1e-9 or mz_cap <= 1e-9:
            continue
        util = (my_req / my_cap) ** alpha + (mz_req / mz_cap) ** alpha
        util = util ** (1.0 / alpha)
        if best_util is None or util < best_util:
            best_util = util
            best_my = my_cap
            best_mz = mz_cap
    if best_util is None:
        return False, None, None, None
    return best_util <= 1.0, best_util, best_my, best_mz


ColumnDesigner.biaxial_ok = biaxial_ok_v31


def shear_check_ec2_v31(v_ed_kN, n_ed_kN, b_mm, h_mm, d_mm, as_long_mm2, fck, fcd, gamma_c):
    """Verificação prática EC2 6.2 com VRd,c, VRd,max e Asw/s requerido.
    Não substitui um dimensionamento completo, mas já separa resistência do betão, limite máximo e necessidade de estribos.
    """
    if b_mm <= 0 or h_mm <= 0 or d_mm <= 0:
        return {"VRdc_kN": None, "VRdmax_kN": None, "Asw_s_req_mm2_per_mm": None, "status": "Dados insuficientes"}
    v_ed = abs(v_ed_kN)
    bw = b_mm
    z = 0.9 * d_mm
    k = min(2.0, 1.0 + math.sqrt(200.0 / max(d_mm, 1e-9)))
    rho_l = min(max(as_long_mm2 / max(bw * d_mm, 1e-9), 0.0), 0.02)
    sigma_cp = min(abs(n_ed_kN) * 1e3 / max(b_mm * h_mm, 1e-9), 0.2 * fcd)
    crdc = 0.18 / gamma_c
    vrdc_N = (crdc * k * (100.0 * rho_l * fck) ** (1.0 / 3.0) + 0.15 * sigma_cp) * bw * d_mm
    vmin = 0.035 * k ** 1.5 * math.sqrt(fck)
    vrdc_min_N = (vmin + 0.15 * sigma_cp) * bw * d_mm
    vrdc_kN = max(vrdc_N, vrdc_min_N) / 1e3

    # VRd,max com cot(theta)=2.0 e nu1=0.6(1-fck/250)
    cot_theta = 2.0
    tan_theta = 1.0 / cot_theta
    nu1 = 0.6 * (1.0 - fck / 250.0)
    alpha_cw = 1.0
    vrdmax_N = alpha_cw * bw * z * nu1 * fcd / (cot_theta + tan_theta)
    vrdmax_kN = vrdmax_N / 1e3

    fyk_assumed = 500.0
    fyd_assumed = fyk_assumed / 1.15
    asw_s_req = 0.0
    if v_ed > vrdc_kN:
        asw_s_req = (v_ed * 1e3) / max(z * fyd_assumed * cot_theta, 1e-9)

    if v_ed <= vrdc_kN:
        status = "OK sem armadura adicional por V"
    elif v_ed <= vrdmax_kN:
        status = "Requer armadura de esforço transverso"
    else:
        status = "Falha VRd,max"
    return {"VRdc_kN": vrdc_kN, "VRdmax_kN": vrdmax_kN, "Asw_s_req_mm2_per_mm": asw_s_req, "status": status}


# design_one resolve este nome global em tempo de execução
shear_check_ec2_practical = shear_check_ec2_v31


def elastic_service_check_v31(n_kN, my_kNm, mz_kNm, b_mm, h_mm, iy_mm4, iz_mm4, as_mm2, fck, fyk, ecm, fctm):
    """ELS elástico melhorado: tensões por secção bruta e estimativa transformada no aço.
    Reporta explicitamente que é um controlo simplificado para pilares.
    """
    if b_mm <= 0 or h_mm <= 0 or iy_mm4 <= 0 or iz_mm4 <= 0:
        return {}
    A = b_mm * h_mm
    N = abs(n_kN) * 1e3
    My = abs(my_kNm) * 1e6
    Mz = abs(mz_kNm) * 1e6
    corners = [(-b_mm/2, -h_mm/2), (-b_mm/2, h_mm/2), (b_mm/2, -h_mm/2), (b_mm/2, h_mm/2)]
    sigmas = [N/A + My*z/iy_mm4 + Mz*y/iz_mm4 for y, z in corners]
    sigma_c_max = max(sigmas)
    sigma_c_min = min(sigmas)
    alpha_e = 200000.0 / max(ecm, 1e-9)
    sigma_s_max = alpha_e * max(abs(sigma_c_min), abs(sigma_c_max))
    lim_c = 0.60 * fck
    lim_s = 0.80 * fyk
    cracking = "Potencial fendilhação" if sigma_c_min < -fctm else "Sem tração relevante"
    status = "OK" if (sigma_c_max <= lim_c and sigma_s_max <= lim_s and sigma_c_min >= -fctm) else "Verificar"
    return {
        "service_sigma_c_max_MPa": sigma_c_max,
        "service_sigma_c_min_MPa": sigma_c_min,
        "service_sigma_s_max_MPa": sigma_s_max,
        "service_sigma_c_lim_MPa": lim_c,
        "service_sigma_s_lim_MPa": lim_s,
        "service_crack_warning": cracking,
        "service_status": status,
        "service_method": "ELS elástico simplificado; tensões por secção bruta/transformada aproximada",
    }


elastic_service_check = elastic_service_check_v31


# Patch cirúrgico do design_one v3: acrescenta campos de auditoria sem reescrever toda a rotina.
_original_design_one_v31_base = ColumnDesigner.design_one

def design_one_v31(self, row: pd.Series, prebuilt_candidates=None):
    # calcular alpha antes do dimensionamento e expor para biaxial_ok
    material_raw = str(row.get("material", "") or "").strip()
    material = material_raw if material_raw and material_raw.lower() not in ["nan", "none"] else DEFAULT_CONCRETE_CLASS
    fck = parse_concrete_strength(material)
    cp = concrete_props(fck, gamma_c=self.gamma_c)
    fyd = steel_props(self.fyk, gamma_s=self.gamma_s)["fyd"]
    b_mm = cm_to_mm(row.get("hy", 0.0))
    h_mm = cm_to_mm(row.get("hz", 0.0))
    ac_mm2 = safe_float(row.get("ax", float("nan"))) * 100.0
    if b_mm <= 0 and ac_mm2 > 0:
        b_mm = math.sqrt(ac_mm2)
    if h_mm <= 0:
        h_mm = b_mm
    if not math.isfinite(ac_mm2) or ac_mm2 <= 0:
        ac_mm2 = b_mm * h_mm
    n_ed_kN = max(abs(safe_float(row.get("fx_i", 0.0), 0.0)), abs(safe_float(row.get("fx_j", 0.0), 0.0)))
    as_seed = max(self.min_longitudinal_as(n_ed_kN, ac_mm2, fyd), 0.002 * ac_mm2)
    alpha, nu_bi = biaxial_alpha_ec2_practical(n_ed_kN, ac_mm2, cp["fcd"], as_seed, fyd)
    self._biaxial_alpha = alpha

    out = _original_design_one_v31_base(self, row, prebuilt_candidates=prebuilt_candidates)
    out["concrete_source"] = "tabela" if material_raw and material_raw.lower() not in ["nan", "none"] else "fallback interno C30/37"
    out["biaxial_method"] = "EC2 5.8.9 simplificado + superfície discreta N-My-Mz"
    out["biaxial_alpha"] = alpha
    out["biaxial_n_ratio"] = nu_bi
    out["m01_y_ec2_kNm"], out["m02_y_ec2_kNm"], out["rm_y_signed"], out["curvature_y"] = order_end_moments_ec2(row.get("my_i", 0.0), row.get("my_j", 0.0))
    out["m01_z_ec2_kNm"], out["m02_z_ec2_kNm"], out["rm_z_signed"], out["curvature_z"] = order_end_moments_ec2(row.get("mz_i", 0.0), row.get("mz_j", 0.0))
    if "v_rd_max_y_kN" not in out:
        # Se a rotina base não exportou os novos campos de shear, calcular aqui para o layout escolhido.
        try:
            as_prov = safe_float(out.get("as_prov_mm2", as_seed), as_seed)
            shear_y = shear_check_ec2_v31(out.get("vy_ed_kN", 0.0), n_ed_kN, h_mm, b_mm, 0.8*b_mm, as_prov, fck, cp["fcd"], self.gamma_c)
            shear_z = shear_check_ec2_v31(out.get("vz_ed_kN", 0.0), n_ed_kN, b_mm, h_mm, 0.8*h_mm, as_prov, fck, cp["fcd"], self.gamma_c)
            out["v_rd_max_y_kN"] = shear_y.get("VRdmax_kN")
            out["v_rd_max_z_kN"] = shear_z.get("VRdmax_kN")
            out["asw_s_req_y_mm2_per_mm"] = shear_y.get("Asw_s_req_mm2_per_mm")
            out["asw_s_req_z_mm2_per_mm"] = shear_z.get("Asw_s_req_mm2_per_mm")
            out["shear_status_y"] = shear_y.get("status", out.get("shear_status_y"))
            out["shear_status_z"] = shear_z.get("status", out.get("shear_status_z"))
        except Exception:
            pass
    return out


ColumnDesigner.design_one = design_one_v31


# Importação: o betão é sempre lido da tabela; se estiver em falta, usa fallback interno fixo.
def load_df_v31(self, df: pd.DataFrame, source: str = ""):
    self.df_raw = df.copy()
    self.df_clean = clean_dataframe(df)
    if "material" in self.df_clean.columns:
        mask = self.df_clean["material"].astype(str).str.strip().str.lower().isin(["", "nan", "none"])
        self.df_clean.loc[mask, "material"] = DEFAULT_CONCRETE_CLASS
    else:
        self.df_clean["material"] = DEFAULT_CONCRETE_CLASS
    self.df_pair = combine_member_end_actions(self.df_clean)
    self.df_calc_input = pd.DataFrame()
    self.df_results = pd.DataFrame()
    self.df_summary = pd.DataFrame()
    self.df_failures = pd.DataFrame()
    self.df_ok = pd.DataFrame()
    self.df_filtered = pd.DataFrame()
    self.df_validation = self.build_data_validation(pre_calc=True)
    self.df_notes = self.build_normative_notes()
    self.show_df(self.tree_input, self.df_clean)
    self.show_df(self.tree_pairs, self.df_pair)
    self.show_df(self.tree_validation, self.df_validation)
    self.show_df(self.tree_results, self.df_results)
    self.show_df(self.tree_summary, self.df_summary)
    self.show_df(self.tree_failures, self.df_failures)
    self.show_df(self.tree_shortlists, pd.DataFrame())
    self.show_df(self.tree_notes, self.df_notes)
    self.update_report()
    self.progress_var.set(0.0)
    bad_pairs = int((self.df_pair.get("n_nodes_found", pd.Series(dtype=float)).fillna(0).astype(float) < 2).sum()) if not self.df_pair.empty else 0
    warn = f"; {bad_pairs} member/case sem dois nós" if bad_pairs else ""
    self.status_var.set(f"Tabela carregada ({source}): {len(self.df_clean)} linhas; {len(self.df_pair)} pares member/case{warn}. Betão lido da coluna Material.")


ColumnsEC2App.load_df = load_df_v31


# Parâmetros e notas v3.1
def _parameters_df_v31(self) -> pd.DataFrame:
    return pd.DataFrame([
        ["Recobrimento [mm]", self.var_cover.get()],
        ["Aço fyk [MPa]", self.var_fyk.get()],
        ["Betão", "lido da coluna Material; fallback interno C30/37 quando ausente"],
        ["φef", self.var_phi_eff.get()],
        ["l0y/L", self.var_l0y.get()],
        ["l0z/L", self.var_l0z.get()],
        ["Modo", self.var_calc_mode.get()],
        ["Método biaxial", "EC2 5.8.9 simplificado + superfície discreta N-My-Mz"],
        ["Esforço transverso", "VRd,c, VRd,max e Asw/s requerido informativo"],
        ["ELS", "tensões elásticas simplificadas e número da combinação relevante"],
    ], columns=["Parâmetro", "Valor"])


ColumnsEC2App._parameters_df = _parameters_df_v31


def _build_normative_notes_v31(self) -> pd.DataFrame:
    notes = [
        ("Materiais", "Betão", "Classe de betão lida da coluna Material da tabela importada; fallback interno C30/37 se o campo vier vazio."),
        ("2.ª ordem", "EC2 5.8", "M01/M02 tratados com sinal relativo para distinguir curvatura simples e dupla; relatório inclui rm assinado."),
        ("Esbelteza", "λlim", "λlim calculado com razão de momentos com sinal, coeficientes A, B e C práticos."),
        ("Biaxial", "EC2 5.8.9", "Verificação por critério com expoente alpha dependente do nível axial e superfície discreta N-My-Mz."),
        ("Esforço transverso", "EC2 6.2", "Cálculo prático de VRd,c, VRd,max e Asw/s requerido; não substitui desenho detalhado dos estribos."),
        ("Torção", "EC2 6.3", "MX é importado e classificado por aviso; dimensionamento completo de torção ainda não é efectuado."),
        ("ELS", "EC2 7", "Verificação simplificada de tensões em serviço e alerta de fendilhação; usar combinações ELS identificadas."),
        ("Auditoria", "Relatório", "Exporta combinação relevante, sinais dos momentos, rm, alpha biaxial, ELS, V e torção."),
    ]
    return pd.DataFrame(notes, columns=["Tema", "Referência", "Nota"])


ColumnsEC2App.build_normative_notes = _build_normative_notes_v31


# Actualizar PDF: remover betão por defeito e incluir método biaxial/ELS.
_old_write_pdf = ColumnsEC2App._write_pdf

def _write_pdf_v31(self, path: str):
    # reutiliza a rotina v3, mas os metadados/parametrização já não expõem betão na GUI.
    return _old_write_pdf(self, path)

ColumnsEC2App._write_pdf = _write_pdf_v31



