# -*- coding: utf-8 -*-
# Auto-split from ColumnsEC2 v0.9 RC8.
# This module is executed in the shared runtime namespace by columns_ec2.runtime.loader.
# Keep execution order defined in columns_ec2/runtime/manifest.py.

APP_VERSION = "v6.8"
REBAR_ETA_MIN_V68 = 0.70
REBAR_ETA_MAX_V68 = 0.90
REBAR_AS_EXCESS_MAX_V68 = 0.40

# --------------------------- módulo lógico: performance ---------------------------
def _v68_now():
    try:
        import time
        return time.perf_counter()
    except Exception:
        return 0.0


def _v68_performance_df(app):
    rows = []
    def add(item, valor, nota=""):
        rows.append({"Item": item, "Valor": valor, "Nota": nota})
    df_clean = getattr(app, "df_clean", pd.DataFrame())
    df_pair = getattr(app, "df_pair", pd.DataFrame())
    df_calc = getattr(app, "df_calc_input", pd.DataFrame())
    df_res = getattr(app, "df_results", pd.DataFrame())
    add("Linhas importadas", 0 if df_clean is None else len(df_clean), "linhas lidas da tabela")
    add("Pares member/case", 0 if df_pair is None else len(df_pair), "pares de extremidades reconstruídos")
    add("Casos governantes calculados", 0 if df_calc is None else len(df_calc), "casos reais seleccionados para cálculo")
    add("Resultados", 0 if df_res is None else len(df_res), "linhas de cálculo produzidas")
    add("Estratégia de armadura", _v64_strategy_label(globals().get("ACTIVE_REBAR_STRATEGY_V64", "equilibrada")) if "_v64_strategy_label" in globals() else globals().get("ACTIVE_REBAR_STRATEGY_V64", "equilibrada"), "critério de escolha da solução")
    add("η_NMyMz alvo", globals().get("REBAR_TARGET_ETA_V64", 0.80), "usado na estratégia equilibrada")
    add("η_NMyMz mínimo", globals().get("REBAR_ETA_MIN_V68", 0.70), "janela de aceitação preferencial")
    add("η_NMyMz máximo", globals().get("REBAR_ETA_MAX_V68", 0.90), "janela de aceitação preferencial")
    add("Excesso máximo de As", globals().get("REBAR_AS_EXCESS_MAX_V68", 0.40), "penalização na estratégia equilibrada")
    perf = getattr(app, "_v68_perf", {}) or {}
    for k, v in perf.items():
        add(k, v, "medido nesta sessão")
    return pd.DataFrame(rows)


# --------------------------- módulo lógico: interacção N-My-Mz ---------------------------
def _v68_interaction_summary(df):
    if df is None or df.empty:
        return pd.DataFrame(columns=["Prumada", "Member", "Case", "N_Ed [kN]", "My_Ed [kNm]", "Mz_Ed [kNm]", "MRd_y [kNm]", "MRd_z [kNm]", "η_NMyMz", "Estado resistente", "Método"])
    out = df.copy()
    if "η_NMyMz" not in out.columns and "utilizacao" in out.columns:
        out["η_NMyMz"] = pd.to_numeric(out["utilizacao"], errors="coerce")
    rows=[]
    for _, r in out.iterrows():
        pr = r.get("prumada", r.get("name", r.get("member", "")))
        rows.append({
            "Prumada": pr,
            "Member": r.get("member", ""),
            "Case": r.get("case", ""),
            "N_Ed [kN]": _finite(r.get("n_ed_kN"), 0.0),
            "My_Ed [kNm]": _finite(r.get("my_ed_kNm"), 0.0),
            "Mz_Ed [kNm]": _finite(r.get("mz_ed_kNm"), 0.0),
            "MRd_y [kNm]": _finite(r.get("mrd_y_kNm"), 0.0),
            "MRd_z [kNm]": _finite(r.get("mrd_z_kNm"), 0.0),
            "η_NMyMz": _finite(r.get("η_NMyMz", r.get("utilizacao")), float("nan")),
            "Estado resistente": r.get("estado_resistente", r.get("status", "")),
            "Método": r.get("biaxial_method", r.get("surface_method", "N-My-Mz")),
        })
    return pd.DataFrame(rows)


# --------------------------- módulo lógico: pormenorização final ---------------------------
def _v68_layout_refinement(row):
    base = _v66_layout_refinement(row) if "_v66_layout_refinement" in globals() else {}
    b = _finite(row.get("b_cm"), 0.0) * 10.0
    h = _finite(row.get("h_cm"), 0.0) * 10.0
    phi = _finite(row.get("phi_long_mm"), 0.0)
    phi_st = _finite(row.get("phi_st_mm"), 8.0)
    s_st = _finite(row.get("s_st_mm"), 0.0)
    ny = int(_finite(row.get("n_bars_y"), 0))
    nz = int(_finite(row.get("n_bars_z"), 0))
    # número de ramos de estribo: 2 exteriores + grampos para varões interiores por direcção
    links_y = max(0, ny - 2)
    links_z = max(0, nz - 2)
    n_grampos = links_y + links_z
    ramos_y = 2 + links_y
    ramos_z = 2 + links_z
    faces = []
    if ny > 2:
        faces.append(f"{ny-2} varão(ões) intermédio(s) nas faces horizontais")
    if nz > 2:
        faces.append(f"{nz-2} varão(ões) intermédio(s) nas faces verticais")
    if not faces:
        faces_txt = "apenas varões de canto"
    else:
        faces_txt = "; ".join(faces)
    grampos_txt = "sem grampos intermédios" if n_grampos == 0 else f"{n_grampos} grampo(s) Ø{int(phi_st)} por nível"
    estribos_txt = f"estribos Ø{int(phi_st)}//{int(round(s_st))} mm" if s_st else f"estribos Ø{int(phi_st)}"
    desc = str(base.get("pormenorizacao_construtiva", row.get("solucao", "")))
    if desc and "; estribos" in desc:
        # substituir por texto final normalizado
        desc0 = desc.split("; estribos")[0]
    else:
        desc0 = str(row.get("layout_description", row.get("solucao", "")))
    base.update({
        "ramos_estribo_y": ramos_y,
        "ramos_estribo_z": ramos_z,
        "numero_grampos_por_nivel": n_grampos,
        "grampos_intermedios": grampos_txt,
        "detalhe_grampos": f"{grampos_txt}; {links_y} na largura + {links_z} na altura; ramos de travamento: {ramos_y} em Y e {ramos_z} em Z",
        "faces_com_varoes_intermedios": faces_txt,
        "pormenorizacao_construtiva": f"{desc0}; {estribos_txt}; {grampos_txt}",
        "solucao_completa": f"{desc0}; {estribos_txt}; {grampos_txt}",
    })
    # Aviso de construtibilidade quando a menor dimensão é reduzida para muitos ramos.
    warnings = str(base.get("detailing_warnings", "") or "")
    if min(b, h) > 0 and n_grampos >= 3:
        extra = "avaliar congestionamento devido ao número de grampos por nível"
        warnings = (warnings + "; " + extra).strip("; ") if warnings else extra
    base["detailing_warnings"] = warnings
    return base


def _v68_apply_constructive_detailing(df):
    if df is None or df.empty:
        return df
    out = df.copy()
    for idx, row in out.iterrows():
        try:
            d = _v68_layout_refinement(row)
            for k, v in d.items():
                out.at[idx, k] = v
        except Exception:
            pass
    try:
        out = _v65_apply_module_statuses(out)
    except Exception:
        pass
    return out

# Substituir a função global usada pela exportação v6.7
_v66_apply_constructive_detailing = _v68_apply_constructive_detailing


# --------------------------- GUI: critérios e nível do PDF ---------------------------
_old_build_sidebar_v68 = ColumnsEC2App._build_sidebar

def _build_sidebar_v68(self, parent):
    if not hasattr(self, "var_eta_target"):
        self.var_eta_target = tk.StringVar(master=self, value=str(globals().get("REBAR_TARGET_ETA_V64", 0.80)))
    if not hasattr(self, "var_eta_min"):
        self.var_eta_min = tk.StringVar(master=self, value=str(globals().get("REBAR_ETA_MIN_V68", 0.70)))
    if not hasattr(self, "var_eta_max"):
        self.var_eta_max = tk.StringVar(master=self, value=str(globals().get("REBAR_ETA_MAX_V68", 0.90)))
    if not hasattr(self, "var_as_excess"):
        self.var_as_excess = tk.StringVar(master=self, value=str(globals().get("REBAR_AS_EXCESS_MAX_V68", 0.40)))
    if not hasattr(self, "var_pdf_level"):
        self.var_pdf_level = tk.StringVar(master=self, value="Relatório técnico")
    _old_build_sidebar_v68(self, parent)
    box = ttk.LabelFrame(parent, text="Critérios da estratégia equilibrada")
    box.pack(fill="x", pady=(0,8))
    for i, (lab, var) in enumerate([
        ("η alvo", self.var_eta_target),
        ("η mínimo", self.var_eta_min),
        ("η máximo", self.var_eta_max),
        ("Excesso máx. As", self.var_as_excess),
    ]):
        ttk.Label(box, text=lab).grid(row=i, column=0, sticky="w", padx=6, pady=3)
        ttk.Entry(box, textvariable=var, width=10).grid(row=i, column=1, sticky="ew", padx=6, pady=3)
    ttk.Label(box, text="Aplicado apenas quando a estratégia seleccionada é Equilibrada.", style="Subtle.TLabel", wraplength=320).grid(row=4, column=0, columnspan=2, sticky="w", padx=6, pady=(2,4))
    box.columnconfigure(1, weight=1)
    pdf = ttk.LabelFrame(parent, text="Relatório PDF")
    pdf.pack(fill="x", pady=(0,8))
    ttk.Label(pdf, text="Nível de detalhe").grid(row=0, column=0, sticky="w", padx=6, pady=4)
    ttk.Combobox(pdf, textvariable=self.var_pdf_level, values=["Resumo executivo", "Relatório técnico", "Memória de cálculo"], state="readonly", width=20).grid(row=0, column=1, sticky="ew", padx=6, pady=4)
    pdf.columnconfigure(1, weight=1)

ColumnsEC2App._build_sidebar = _build_sidebar_v68


# --------------------------- run hook: critérios e performance ---------------------------
_old_run_design_v68 = ColumnsEC2App.run_design

def _run_design_v68(self):
    try:
        globals()["REBAR_TARGET_ETA_V64"] = float(str(self.var_eta_target.get()).replace(",", "."))
        globals()["REBAR_ETA_MIN_V68"] = float(str(self.var_eta_min.get()).replace(",", "."))
        globals()["REBAR_ETA_MAX_V68"] = float(str(self.var_eta_max.get()).replace(",", "."))
        globals()["REBAR_AS_EXCESS_MAX_V68"] = float(str(self.var_as_excess.get()).replace(",", "."))
    except Exception:
        pass
    self._v68_t0 = _v68_now()
    self._v68_perf = {"Início do cálculo": datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
    _old_run_design_v68(self)
    def poll():
        try:
            if getattr(self, "df_results", pd.DataFrame()) is not None and not self.df_results.empty and float(self.progress_var.get()) >= 99:
                elapsed = _v68_now() - getattr(self, "_v68_t0", _v68_now())
                self._v68_perf["Tempo total de cálculo [s]"] = round(elapsed, 3)
                self._v68_perf["Fim do cálculo"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                # aplicar pormenorização final e actualizar tabelas visíveis
                try:
                    self.df_results = _v68_apply_constructive_detailing(self.df_results)
                    self.df_summary = self.build_summary_by_member(self.df_results) if getattr(self, "var_summary", tk.BooleanVar(value=True)).get() else pd.DataFrame()
                    self.show_df(self.tree_results, self.df_results)
                    self.show_df(self.tree_summary, self.df_summary)
                except Exception:
                    pass
                return
        except Exception:
            return
        try: self.after(700, poll)
        except Exception: pass
    try: self.after(900, poll)
    except Exception: pass

ColumnsEC2App.run_design = _run_design_v68


# --------------------------- Excel: folhas novas ---------------------------
_old_write_excel_v68 = ColumnsEC2App._write_excel

def _write_excel_v68(self, path: str):
    t0 = _v68_now()
    try:
        if getattr(self, "df_results", pd.DataFrame()) is not None and not self.df_results.empty:
            self.df_results = _v68_apply_constructive_detailing(self.df_results)
            self.df_summary = _v68_apply_constructive_detailing(getattr(self, "df_summary", pd.DataFrame()))
    except Exception:
        pass
    _old_write_excel_v68(self, path)
    elapsed = _v68_now() - t0
    try:
        self._v68_perf = getattr(self, "_v68_perf", {}) or {}
        self._v68_perf["Tempo de exportação XLSX [s]"] = round(elapsed, 3)
        with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            _v68_performance_df(self).to_excel(writer, sheet_name="12_Performance", index=False)
            _v68_interaction_summary(getattr(self, "df_summary", pd.DataFrame()) if getattr(self, "df_summary", pd.DataFrame()) is not None and not self.df_summary.empty else getattr(self, "df_results", pd.DataFrame())).to_excel(writer, sheet_name="14_Interacao_NMyMz", index=False)
            pd.DataFrame([
                ["η_NMyMz alvo", globals().get("REBAR_TARGET_ETA_V64", 0.80)],
                ["η_NMyMz mínimo", globals().get("REBAR_ETA_MIN_V68", 0.70)],
                ["η_NMyMz máximo", globals().get("REBAR_ETA_MAX_V68", 0.90)],
                ["Excesso máximo de As", globals().get("REBAR_AS_EXCESS_MAX_V68", 0.40)],
                ["Nível PDF", getattr(self, "var_pdf_level", tk.StringVar(value="Relatório técnico")).get()],
            ], columns=["Critério", "Valor"]).to_excel(writer, sheet_name="01C_Criterios", index=False)
    except Exception:
        pass

ColumnsEC2App._write_excel = _write_excel_v68


# --------------------------- PDF: níveis de detalhe ---------------------------
def _v68_pdf_styles():
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="T68", parent=styles["Title"], alignment=TA_CENTER, fontName="Courier-Bold", fontSize=14, leading=18, spaceAfter=8))
    styles.add(ParagraphStyle(name="H68", parent=styles["Heading2"], fontName="Courier-Bold", fontSize=11, leading=14, spaceBefore=8, spaceAfter=6))
    styles.add(ParagraphStyle(name="B68", parent=styles["Normal"], fontName="Courier", fontSize=8.5, leading=12, spaceAfter=4))
    styles.add(ParagraphStyle(name="C68", parent=styles["Normal"], fontName="Courier", fontSize=6.5, leading=8, alignment=TA_LEFT))
    return styles


def _v68_table(df, cols, style, max_rows=40):
    from reportlab.platypus import Table, Paragraph
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import TableStyle
    if df is None or df.empty:
        return Paragraph("Sem dados.", style)
    present=[c for c in cols if c in df.columns]
    if not present:
        return Paragraph("Sem colunas aplicáveis.", style)
    data=[[Paragraph(str(c), style) for c in present]]
    for _, r in df.head(max_rows).iterrows():
        row=[]
        for c in present:
            v=r.get(c,"")
            if isinstance(v, float): txt="" if not math.isfinite(v) else f"{v:.2f}"
            else: txt="" if pd.isna(v) else str(v)
            txt=txt.replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")
            row.append(Paragraph(txt, style))
        data.append(row)
    widths=[270*mm/max(1,len(present))]*len(present)
    tbl=Table(data, colWidths=widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("GRID",(0,0),(-1,-1),0.25,colors.lightgrey),
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#EFEFEF")),
        ("FONTNAME",(0,0),(-1,0),"Courier-Bold"),
        ("VALIGN",(0,0),(-1,-1),"TOP"),
        ("LEFTPADDING",(0,0),(-1,-1),3), ("RIGHTPADDING",(0,0),(-1,-1),3),
    ]))
    return tbl


def _write_pdf_v68(self, path: str):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, PageBreak
    from reportlab.platypus import TableStyle
    t0=_v68_now()
    styles=_v68_pdf_styles()
    level = getattr(self, "var_pdf_level", tk.StringVar(value="Relatório técnico")).get()
    res = _v68_apply_constructive_detailing(getattr(self, "df_results", pd.DataFrame()))
    summ = _v68_apply_constructive_detailing(getattr(self, "df_summary", pd.DataFrame()))
    if summ is None or summ.empty: summ=res
    module = _v65_module_status_table(res) if "_v65_module_status_table" in globals() and res is not None and not res.empty else pd.DataFrame()
    inter = _v68_interaction_summary(summ)
    perf = _v68_performance_df(self)
    doc=SimpleDocTemplate(path, pagesize=landscape(A4), rightMargin=12*mm, leftMargin=12*mm, topMargin=12*mm, bottomMargin=12*mm)
    story=[]
    story.append(Paragraph("Columns EC2", styles["T68"]))
    story.append(Paragraph("Relatório técnico de dimensionamento/verificação de pilares de betão armado", styles["B68"]))
    story.append(Spacer(1,4*mm))
    n_total=len(res) if res is not None else 0
    n_fail=int((res.get("estado_global", res.get("status", pd.Series(dtype=str))).astype(str)=="Falha").sum()) if res is not None and not res.empty else 0
    n_warn=int((res.get("estado_global", res.get("status", pd.Series(dtype=str))).astype(str)=="Aviso").sum()) if res is not None and not res.empty else 0
    meta=pd.DataFrame([
        {"Campo":"Norma/Motor", "Valor": _v59_norm_reference(self) if "_v59_norm_reference" in globals() else "Eurocódigo 2"},
        {"Campo":"Estratégia de armadura", "Valor": _v64_strategy_label(globals().get("ACTIVE_REBAR_STRATEGY_V64","equilibrada")) if "_v64_strategy_label" in globals() else "-"},
        {"Campo":"Casos analisados", "Valor": n_total},
        {"Campo":"Falhas / Avisos", "Valor": f"{n_fail} / {n_warn}"},
    ])
    story.append(_v68_table(meta, ["Campo","Valor"], styles["C68"], max_rows=10))
    story.append(Paragraph("1. Decisão por prumada", styles["H68"]))
    story.append(_v68_table(summ, ["prumada","member","case","b_cm","h_cm","material","n_ed_kN","my_ed_kNm","mz_ed_kNm","η_NMyMz","solucao_completa","estado_global"], styles["C68"], max_rows=45))
    if level in ["Relatório técnico", "Memória de cálculo"]:
        story.append(PageBreak()); story.append(Paragraph("2. Interacção N-My-Mz", styles["H68"]))
        story.append(_v68_table(inter, ["Prumada","Member","Case","N_Ed [kN]","My_Ed [kNm]","Mz_Ed [kNm]","MRd_y [kNm]","MRd_z [kNm]","η_NMyMz","Estado resistente"], styles["C68"], max_rows=60))
        story.append(Paragraph("3. Estados por módulo", styles["H68"]))
        story.append(_v68_table(module, ["prumada","member","case","estado_global","estado_resistente","estado_corte","estado_torcao","estado_els","estado_pormenorizacao","decisao_tecnica"], styles["C68"], max_rows=60))
    if level == "Memória de cálculo":
        story.append(PageBreak()); story.append(Paragraph("4. Casos governantes", styles["H68"]))
        gov = _v67_governing_cases_df(self) if "_v67_governing_cases_df" in globals() else pd.DataFrame()
        story.append(_v68_table(gov, ["Prumada","Member","Case","Critério","NEd [kN]","My [kNm]","Mz [kNm]","Vy [kN]","Vz [kN]","T [kNm]"], styles["C68"], max_rows=80))
        story.append(PageBreak()); story.append(Paragraph("5. Performance", styles["H68"]))
        story.append(_v68_table(perf, ["Item","Valor","Nota"], styles["C68"], max_rows=80))
    def footer(canvas, doc_obj):
        canvas.saveState(); canvas.setFont("Courier",7); canvas.setFillColor(colors.grey)
        canvas.drawString(12*mm,7*mm,f"Columns EC2 | {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        canvas.drawRightString(285*mm,7*mm,f"Página {doc_obj.page}")
        canvas.restoreState()
    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    try:
        self._v68_perf = getattr(self, "_v68_perf", {}) or {}
        self._v68_perf["Tempo de exportação PDF [s]"] = round(_v68_now()-t0,3)
    except Exception:
        pass

ColumnsEC2App._write_pdf = _write_pdf_v68


# --------------------------- modularização inicial ---------------------------
def _v68_write_module_skeleton():
    """Cria uma pasta técnica mínima de módulos auxiliares junto ao ficheiro exportado.
    Não é necessária para correr a GUI, mas inicia a separação real de responsabilidades.
    """
    try:
        base = Path(__file__).with_name("columnsec2_core")
        base.mkdir(exist_ok=True)
        files = {
            "__init__.py": "# ColumnsEC2 core package — início de modularização v6.8\n",
            "performance.py": "# Métricas de performance e auditoria de cálculo.\n",
            "interaction_nmm.py": "# Resumos e exportação da interacção N-My-Mz.\n",
            "detailing.py": "# Pormenorização construtiva de pilares, ramos de estribos e grampos.\n",
            "reports.py": "# Exportações PDF/XLSX/DXF.\n",
        }
        for name, content in files.items():
            p = base / name
            if not p.exists():
                p.write_text(content, encoding="utf-8")
    except Exception:
        pass

# RC11 modular package: the real package structure already exists.
# The legacy v6.8 skeleton writer is intentionally not executed.
# _v68_write_module_skeleton()



# ============================================================
# v6.8.1 — PDF organizado por prumada/nome e sem nota sintética
# ============================================================
APP_VERSION = "v6.8.1"


def _v681_norm_prumada_series(df: pd.DataFrame):
    if df is None or df.empty:
        return pd.Series(dtype=str)
    idx = df.index
    for col in ["prumada", "name", "Name", "nome", "Pilar", "pilar"]:
        if col in df.columns:
            s = df[col].astype(str).str.strip()
            s = s.replace({"": pd.NA, "nan": pd.NA, "None": pd.NA})
            if s.notna().any():
                return s.fillna(df.get("member", pd.Series("-", index=idx)).astype(str))
    return df.get("member", pd.Series("-", index=idx)).astype(str)


def _v681_status_rank(value):
    v = str(value or "").strip().lower()
    if "falha" in v:
        return 3
    if "aviso" in v or "verificar" in v:
        return 2
    if "pré" in v or "pre" in v:
        return 1
    if "ok" in v:
        return 0
    return 1


def _v681_prepare_pdf_df(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    out = df.copy()
    out["Prumada"] = _v681_norm_prumada_series(out)
    if "η_NMyMz" not in out.columns and "utilizacao" in out.columns:
        out["η_NMyMz"] = out["utilizacao"]
    if "Estado" not in out.columns:
        out["Estado"] = out.get("estado_global", out.get("status", ""))
    if "Secção [cm]" not in out.columns:
        out["Secção [cm]"] = out.apply(lambda r: f"{safe_float(r.get('b_cm',0),0):.0f}x{safe_float(r.get('h_cm',0),0):.0f}", axis=1)
    if "Solução" not in out.columns:
        out["Solução"] = out.get("solucao_completa", out.get("solucao", ""))
    sort_cols = ["Prumada"]
    if "story_order" in out.columns:
        sort_cols.append("story_order")
    elif "level_order" in out.columns:
        sort_cols.append("level_order")
    elif "member" in out.columns:
        sort_cols.append("member")
    if "case" in out.columns:
        sort_cols.append("case")
    try:
        out = out.sort_values(sort_cols, kind="mergesort")
    except Exception:
        out = out.sort_values(["Prumada"], kind="mergesort")
    return out


def _v681_decision_by_prumada(df: pd.DataFrame) -> pd.DataFrame:
    work = _v681_prepare_pdf_df(df)
    if work.empty:
        return work
    rows = []
    for pr, grp in work.groupby("Prumada", dropna=False):
        g = grp.copy()
        g["_rank"] = g.get("Estado", pd.Series("", index=g.index)).map(_v681_status_rank)
        if "η_NMyMz" in g.columns:
            g["_eta"] = g["η_NMyMz"].map(lambda x: safe_float(x, -1.0))
        else:
            g["_eta"] = -1.0
        g = g.sort_values(["_rank", "_eta"], ascending=[False, False])
        r = g.iloc[0].copy()
        r["N.º tramos/casos"] = len(grp)
        r["Prumada"] = pr
        rows.append(r.drop(labels=[c for c in ["_rank", "_eta"] if c in r.index]))
    out = pd.DataFrame(rows)
    try:
        out = out.sort_values("Prumada", key=lambda s: s.astype(str).str.extract(r'(\d+)')[0].astype(float).fillna(1e9).astype(str) + '_' + s.astype(str))
    except Exception:
        out = out.sort_values("Prumada")
    return out


def _v681_governing_cases_for_pdf(self) -> pd.DataFrame:
    gov = _v67_governing_cases_df(self) if "_v67_governing_cases_df" in globals() else pd.DataFrame()
    if gov is None or gov.empty:
        return pd.DataFrame()
    out = gov.copy()
    if "Prumada" not in out.columns:
        out["Prumada"] = out.get("name", out.get("member", ""))
    # Para o PDF, a organização é por prumada/nome; o member fica apenas no Excel.
    return out.sort_values(["Prumada", "Case"] if "Case" in out.columns else ["Prumada"])


def _write_pdf_v681(self, path: str):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak
    t0 = _v68_now() if "_v68_now" in globals() else 0.0
    styles = _v68_pdf_styles() if "_v68_pdf_styles" in globals() else _pdf_styles_v3()
    level = getattr(self, "var_pdf_level", tk.StringVar(value="Relatório técnico")).get()

    res = _v68_apply_constructive_detailing(getattr(self, "df_results", pd.DataFrame())) if "_v68_apply_constructive_detailing" in globals() else getattr(self, "df_results", pd.DataFrame())
    summ_src = _v68_apply_constructive_detailing(getattr(self, "df_summary", pd.DataFrame())) if "_v68_apply_constructive_detailing" in globals() else getattr(self, "df_summary", pd.DataFrame())
    if summ_src is None or summ_src.empty:
        summ_src = res
    summ = _v681_decision_by_prumada(summ_src)
    res_pdf = _v681_prepare_pdf_df(res)
    module = _v65_module_status_table(res) if "_v65_module_status_table" in globals() and res is not None and not res.empty else pd.DataFrame()
    module = _v681_prepare_pdf_df(module)
    inter = _v68_interaction_summary(summ_src) if "_v68_interaction_summary" in globals() else pd.DataFrame()
    if inter is not None and not inter.empty:
        inter = inter.rename(columns={"Member": "Elemento"})
        # manter apenas prumada no PDF; Member/Elemento fica dispensável para leitura de gabinete
        if "Elemento" in inter.columns:
            inter = inter.drop(columns=["Elemento"], errors="ignore")
    perf = _v68_performance_df(self) if "_v68_performance_df" in globals() else pd.DataFrame()

    doc = SimpleDocTemplate(path, pagesize=landscape(A4), rightMargin=12*mm, leftMargin=12*mm, topMargin=12*mm, bottomMargin=12*mm)
    story = []
    story.append(Paragraph("Columns EC2", styles["T68"]))
    story.append(Paragraph("Relatório técnico de dimensionamento/verificação de pilares de betão armado", styles["B68"]))
    story.append(Spacer(1, 4*mm))

    n_total = len(res) if res is not None else 0
    st_series = res.get("estado_global", res.get("status", pd.Series(dtype=str))).astype(str) if res is not None and not res.empty else pd.Series(dtype=str)
    n_fail = int((st_series == "Falha").sum()) if not st_series.empty else 0
    n_warn = int((st_series == "Aviso").sum()) if not st_series.empty else 0
    n_pr = int(summ["Prumada"].astype(str).nunique()) if summ is not None and not summ.empty and "Prumada" in summ.columns else 0
    meta = pd.DataFrame([
        {"Campo": "Norma/Motor", "Valor": _v59_norm_reference(self) if "_v59_norm_reference" in globals() else "Eurocódigo 2"},
        {"Campo": "Estratégia de armadura", "Valor": _v64_strategy_label(globals().get("ACTIVE_REBAR_STRATEGY_V64", "equilibrada")) if "_v64_strategy_label" in globals() else "Equilibrada"},
        {"Campo": "Prumadas", "Valor": n_pr},
        {"Campo": "Casos analisados", "Valor": n_total},
        {"Campo": "Falhas / Avisos", "Valor": f"{n_fail} / {n_warn}"},
    ])
    story.append(_v68_table(meta, ["Campo", "Valor"], styles["C68"], max_rows=10))

    story.append(Paragraph("1. Decisão por prumada", styles["H68"]))
    story.append(_v68_table(
        summ,
        ["Prumada", "N.º tramos/casos", "case", "Secção [cm]", "material", "n_ed_kN", "my_ed_kNm", "mz_ed_kNm", "η_NMyMz", "Solução", "Estado"],
        styles["C68"],
        max_rows=80,
    ))

    if level in ["Relatório técnico", "Memória de cálculo"]:
        story.append(PageBreak())
        story.append(Paragraph("2. Interacção N-My-Mz por prumada", styles["H68"]))
        story.append(_v68_table(
            inter,
            ["Prumada", "Case", "N_Ed [kN]", "My_Ed [kNm]", "Mz_Ed [kNm]", "MRd_y [kNm]", "MRd_z [kNm]", "η_NMyMz", "Estado resistente"],
            styles["C68"],
            max_rows=80,
        ))
        story.append(Paragraph("3. Estados por módulo", styles["H68"]))
        story.append(_v68_table(
            module,
            ["Prumada", "case", "estado_global", "estado_resistente", "estado_corte", "estado_torcao", "estado_els", "estado_pormenorizacao", "decisao_tecnica"],
            styles["C68"],
            max_rows=80,
        ))

    if level == "Memória de cálculo":
        story.append(PageBreak())
        story.append(Paragraph("4. Casos governantes", styles["H68"]))
        gov = _v681_governing_cases_for_pdf(self)
        story.append(_v68_table(
            gov,
            ["Prumada", "Case", "Critério", "NEd [kN]", "My [kNm]", "Mz [kNm]", "Vy [kN]", "Vz [kN]", "T [kNm]"],
            styles["C68"],
            max_rows=100,
        ))
        story.append(PageBreak())
        story.append(Paragraph("5. Performance", styles["H68"]))
        story.append(_v68_table(perf, ["Item", "Valor", "Nota"], styles["C68"], max_rows=80))

    def footer(canvas, doc_obj):
        canvas.saveState()
        canvas.setFont("Courier", 7)
        canvas.setFillColor(colors.grey)
        canvas.drawString(12*mm, 7*mm, f"Columns EC2 | {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        canvas.drawRightString(285*mm, 7*mm, f"Página {doc_obj.page}")
        canvas.restoreState()
    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    try:
        self._v68_perf = getattr(self, "_v68_perf", {}) or {}
        self._v68_perf["Tempo de exportação PDF [s]"] = round((_v68_now() if "_v68_now" in globals() else 0.0)-t0, 3) if t0 else "-"
    except Exception:
        pass

ColumnsEC2App._write_pdf = _write_pdf_v681



# ============================================================
# ColumnsEC2 v6.8.2 — secções variáveis na mesma prumada
# - cálculo/resumo por tramo físico e não por secção dominante
# - PDF e DXF por prumada + tramo, preservando alterações de secção/armadura
# ============================================================
APP_VERSION = "v6.8.2"


def _v682_natural_key(value):
    s = str(value if value is not None else "")
    parts = re.split(r"(\d+)", s)
    out = []
    for p in parts:
        if p.isdigit():
            out.append((0, int(p)))
        else:
            out.append((1, p.lower()))
    return tuple(out)


def _v682_is_blank(v):
    return str(v if v is not None else "").strip().lower() in ["", "nan", "none", "-", "<na>"]


def _v682_prumada_from_row(row):
    # A prumada é o nome do pilar; member é apenas o tramo/barra dentro dessa prumada.
    for k in ["prumada", "name", "Name", "nome", "Pilar", "pilar", "pillar_name"]:
        try:
            v = row.get(k, "")
        except Exception:
            v = ""
        if not _v682_is_blank(v):
            return str(v).strip()
    try:
        return str(row.get("member", "")).strip() or "-"
    except Exception:
        return "-"


def _v682_story_from_row(row):
    for k in ["story", "Story", "storey", "Storey", "piso", "Piso", "floor", "Floor", "level", "Level", "pavimento", "Pavimento"]:
        try:
            v = row.get(k, "")
        except Exception:
            v = ""
        if not _v682_is_blank(v):
            return str(v).strip()
    return ""


def _v682_section_signature(row):
    b = round(_finite(row.get("b_cm", row.get("hy", 0.0)), 0.0), 3)
    h = round(_finite(row.get("h_cm", row.get("hz", 0.0)), 0.0), 3)
    mat = str(row.get("material", "") or "").strip()
    # A solução também entra no identificador para que mudanças de armadura no mesmo member não sejam perdidas.
    sol = str(row.get("solucao_completa", row.get("solucao", "")) or "").strip()
    return f"{b:g}x{h:g}|{mat}|{sol}"


def _v682_segment_label(row, fallback_index=None):
    st = _v682_story_from_row(row)
    if st:
        return st
    mem = str(row.get("member", "") or "").strip()
    if mem:
        return f"Tramo {mem}"
    if fallback_index is not None:
        return f"Tramo {int(fallback_index):02d}"
    return "Tramo"


def _v682_segment_sort_tuple(row):
    """Ordenação de baixo para cima.
    Se existir piso/cota, usa-a; caso contrário usa a ordem original/member.
    """
    st = _v682_story_from_row(row)
    if st:
        nums = re.findall(r"-?\d+(?:[\.,]\d+)?", st)
        if nums:
            try:
                return (0, float(nums[-1].replace(",", ".")), _v682_natural_key(st))
            except Exception:
                pass
        return (1, 0.0, _v682_natural_key(st))
    # usar ordem original quando disponível; é geralmente a ordem vertical da exportação/modelo
    ro = _finite(row.get("__row_order", row.get("row_order", row.get("segment_order", 0))), 0.0)
    mem = str(row.get("member", "") or "")
    return (2, ro, _v682_natural_key(mem))


def _v682_governing_score_df(df: pd.DataFrame) -> pd.Series:
    if df is None or df.empty:
        return pd.Series(dtype=float)
    idx = df.index
    status = df.get("estado_global", df.get("status", pd.Series("", index=idx))).astype(str)
    rank = status.map({"Falha": 1_000_000.0, "Aviso": 500_000.0, "Pré-dimensionado": 250_000.0, "OK": 0.0}).fillna(0.0)
    eta_name = "η_NMyMz" if "η_NMyMz" in df.columns else "eta_NMyMz" if "eta_NMyMz" in df.columns else "utilizacao"
    eta = pd.to_numeric(df.get(eta_name, pd.Series(0.0, index=idx)), errors="coerce").fillna(0.0) * 10_000.0
    n = pd.to_numeric(df.get("n_ed_kN", pd.Series(0.0, index=idx)), errors="coerce").abs().fillna(0.0)
    my = pd.to_numeric(df.get("my_ed_kNm", pd.Series(0.0, index=idx)), errors="coerce").abs().fillna(0.0)
    mz = pd.to_numeric(df.get("mz_ed_kNm", pd.Series(0.0, index=idx)), errors="coerce").abs().fillna(0.0)
    return rank + eta + 0.02*n + my + mz


def _v682_build_tramo_schedule(results: pd.DataFrame) -> pd.DataFrame:
    """Cria uma tabela de decisão por prumada e por tramo físico.

    Ao contrário do resumo antigo por prumada, esta função NÃO escolhe a secção
    que aparece mais vezes nem colapsa todos os P1 numa única linha. Cada tramo
    físico/member é mantido, preservando mudanças de secção, material ou armadura
    ao longo da prumada.
    """
    if results is None or results.empty:
        return pd.DataFrame()
    work = results.copy()
    try:
        if "_v65_apply_module_statuses" in globals():
            work = _v65_apply_module_statuses(work)
    except Exception:
        pass
    try:
        if "_v68_apply_constructive_detailing" in globals():
            work = _v68_apply_constructive_detailing(work)
    except Exception:
        pass
    work["Prumada"] = work.apply(_v682_prumada_from_row, axis=1)
    work["_story_label"] = work.apply(_v682_story_from_row, axis=1)
    work["_section_signature"] = work.apply(_v682_section_signature, axis=1)
    work["_segment_sort"] = work.apply(_v682_segment_sort_tuple, axis=1)
    work["_gov_score_v682"] = _v682_governing_score_df(work)

    # Chave de tramo físico. O member é obrigatório para não fundir secções diferentes da mesma prumada.
    # A assinatura da secção/armadura protege contra alterações de secção dentro da mesma prumada.
    group_cols = ["Prumada"]
    if "member" in work.columns:
        group_cols.append("member")
    if "_story_label" in work.columns:
        group_cols.append("_story_label")
    group_cols.append("_section_signature")

    rows = []
    for _, grp in work.groupby(group_cols, dropna=False, sort=False):
        g = grp.sort_values("_gov_score_v682", ascending=False)
        r = g.iloc[0].copy()
        r["Prumada"] = _v682_prumada_from_row(r)
        r["Tramo"] = _v682_segment_label(r)
        r["N.º combinações/tramo"] = len(grp)
        r["Secção [cm]"] = f"{_finite(r.get('b_cm', r.get('hy',0)),0):.0f}x{_finite(r.get('h_cm', r.get('hz',0)),0):.0f}"
        r["Solução"] = r.get("solucao_completa", r.get("solucao", ""))
        r["Estado"] = r.get("estado_global", r.get("status", ""))
        r["_segment_sort"] = grp["_segment_sort"].iloc[0]
        rows.append(r)
    out = pd.DataFrame(rows)
    if out.empty:
        return out
    try:
        out = out.sort_values(["Prumada", "_segment_sort"], key=lambda s: s.map(lambda x: str(x)), kind="mergesort")
    except Exception:
        out = out.sort_values(["Prumada"], kind="mergesort")

    # Ordem relativa dentro de cada prumada: baixo -> cima.
    out["Ordem na prumada"] = out.groupby("Prumada").cumcount() + 1
    out["Tramo"] = out.apply(lambda r: r.get("Tramo") if not _v682_is_blank(r.get("Tramo")) else f"Tramo {int(r.get('Ordem na prumada',1)):02d}", axis=1)
    return out.drop(columns=[c for c in ["_gov_score_v682", "_section_signature"] if c in out.columns], errors="ignore")


# Resumo da GUI/Excel: passa a ser por prumada + tramo, não uma única linha por prumada.
def _build_summary_by_member_v682(self, results: pd.DataFrame) -> pd.DataFrame:
    return _v682_build_tramo_schedule(results)

ColumnsEC2App.build_summary_by_member = _build_summary_by_member_v682


def _v682_pdf_prepare(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    out = df.copy()
    if "Prumada" not in out.columns:
        out["Prumada"] = out.apply(_v682_prumada_from_row, axis=1)
    if "Tramo" not in out.columns:
        out["Tramo"] = out.apply(_v682_segment_label, axis=1)
    if "Secção [cm]" not in out.columns:
        out["Secção [cm]"] = out.apply(lambda r: f"{_finite(r.get('b_cm', r.get('hy',0)),0):.0f}x{_finite(r.get('h_cm', r.get('hz',0)),0):.0f}", axis=1)
    if "Solução" not in out.columns:
        out["Solução"] = out.get("solucao_completa", out.get("solucao", ""))
    if "Estado" not in out.columns:
        out["Estado"] = out.get("estado_global", out.get("status", ""))
    if "Case" not in out.columns and "case" in out.columns:
        out["Case"] = out["case"]
    return out


# PDF revisto: decisões por prumada/tramo, preservando mudanças de secção na mesma prumada.
def _write_pdf_v682(self, path: str):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak
    t0 = _v68_now() if "_v68_now" in globals() else 0.0
    styles = _v68_pdf_styles() if "_v68_pdf_styles" in globals() else _pdf_styles_v3()
    level = getattr(self, "var_pdf_level", tk.StringVar(value="Relatório técnico")).get()

    res = getattr(self, "df_results", pd.DataFrame())
    schedule = _v682_build_tramo_schedule(res)
    schedule_pdf = _v682_pdf_prepare(schedule)
    n_pr = int(schedule_pdf["Prumada"].astype(str).nunique()) if not schedule_pdf.empty and "Prumada" in schedule_pdf.columns else 0
    n_tramos = len(schedule_pdf)
    st_series = schedule_pdf.get("Estado", pd.Series(dtype=str)).astype(str) if not schedule_pdf.empty else pd.Series(dtype=str)
    n_fail = int((st_series == "Falha").sum()) if not st_series.empty else 0
    n_warn = int((st_series == "Aviso").sum()) if not st_series.empty else 0

    doc = SimpleDocTemplate(path, pagesize=landscape(A4), rightMargin=12*mm, leftMargin=12*mm, topMargin=12*mm, bottomMargin=12*mm)
    story = []
    story.append(Paragraph("Columns EC2", styles["T68"]))
    story.append(Paragraph("Relatório técnico de dimensionamento/verificação de pilares de betão armado", styles["B68"]))
    story.append(Spacer(1, 4*mm))
    meta = pd.DataFrame([
        {"Campo": "Norma/Motor", "Valor": _v59_norm_reference(self) if "_v59_norm_reference" in globals() else "Eurocódigo 2"},
        {"Campo": "Estratégia de armadura", "Valor": _v64_strategy_label(globals().get("ACTIVE_REBAR_STRATEGY_V64", "equilibrada")) if "_v64_strategy_label" in globals() else "Equilibrada"},
        {"Campo": "Prumadas", "Valor": n_pr},
        {"Campo": "Tramos/secções verificados", "Valor": n_tramos},
        {"Campo": "Falhas / Avisos", "Valor": f"{n_fail} / {n_warn}"},
    ])
    story.append(_v68_table(meta, ["Campo", "Valor"], styles["C68"], max_rows=10))

    story.append(Paragraph("1. Decisão por prumada e tramo", styles["H68"]))
    story.append(_v68_table(
        schedule_pdf,
        ["Prumada", "Ordem na prumada", "Tramo", "Case", "Secção [cm]", "material", "n_ed_kN", "my_ed_kNm", "mz_ed_kNm", "η_NMyMz", "Solução", "Estado"],
        styles["C68"],
        max_rows=120,
    ))

    if level in ["Relatório técnico", "Memória de cálculo"]:
        story.append(PageBreak())
        story.append(Paragraph("2. Interacção N-My-Mz por tramo", styles["H68"]))
        inter_src = schedule if not schedule.empty else res
        inter = _v68_interaction_summary(inter_src) if "_v68_interaction_summary" in globals() else pd.DataFrame()
        if inter is not None and not inter.empty:
            # manter prumada + member/tramo para distinguir mudanças de secção
            if "Member" in inter.columns:
                inter = inter.rename(columns={"Member": "Tramo/Member"})
        story.append(_v68_table(
            inter,
            ["Prumada", "Tramo/Member", "Case", "N_Ed [kN]", "My_Ed [kNm]", "Mz_Ed [kNm]", "MRd_y [kNm]", "MRd_z [kNm]", "η_NMyMz", "Estado resistente"],
            styles["C68"],
            max_rows=120,
        ))
        story.append(Paragraph("3. Estados por módulo", styles["H68"]))
        module = _v65_module_status_table(res) if "_v65_module_status_table" in globals() and res is not None and not res.empty else pd.DataFrame()
        module_sched = _v682_build_tramo_schedule(module) if module is not None and not module.empty else module
        module_pdf = _v682_pdf_prepare(module_sched)
        story.append(_v68_table(
            module_pdf,
            ["Prumada", "Tramo", "Case", "estado_global", "estado_resistente", "estado_corte", "estado_torcao", "estado_els", "estado_pormenorizacao", "decisao_tecnica"],
            styles["C68"],
            max_rows=120,
        ))

    if level == "Memória de cálculo":
        story.append(PageBreak())
        story.append(Paragraph("4. Casos governantes", styles["H68"]))
        gov = _v681_governing_cases_for_pdf(self) if "_v681_governing_cases_for_pdf" in globals() else pd.DataFrame()
        story.append(_v68_table(
            gov,
            ["Prumada", "Case", "Critério", "NEd [kN]", "My [kNm]", "Mz [kNm]", "Vy [kN]", "Vz [kN]", "T [kNm]"],
            styles["C68"],
            max_rows=140,
        ))
        story.append(PageBreak())
        story.append(Paragraph("5. Performance", styles["H68"]))
        perf = _v68_performance_df(self) if "_v68_performance_df" in globals() else pd.DataFrame()
        story.append(_v68_table(perf, ["Item", "Valor", "Nota"], styles["C68"], max_rows=80))

    def footer(canvas, doc_obj):
        canvas.saveState()
        canvas.setFont("Courier", 7)
        canvas.setFillColor(colors.grey)
        canvas.drawString(12*mm, 7*mm, f"Columns EC2 | {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        canvas.drawRightString(285*mm, 7*mm, f"Página {doc_obj.page}")
        canvas.restoreState()
    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    try:
        self._v68_perf = getattr(self, "_v68_perf", {}) or {}
        self._v68_perf["Tempo de exportação PDF [s]"] = round((_v68_now() if "_v68_now" in globals() else 0.0)-t0, 3) if t0 else "-"
    except Exception:
        pass

ColumnsEC2App._write_pdf = _write_pdf_v682


# Excel: acrescentar quadro por prumada/tramo e usar df_summary já corrigido por tramos.
_old_write_excel_v682_base = ColumnsEC2App._write_excel

def _write_excel_v682(self, path: str):
    try:
        if getattr(self, "df_results", pd.DataFrame()) is not None and not self.df_results.empty:
            self.df_summary = _v682_build_tramo_schedule(self.df_results)
    except Exception:
        pass
    _old_write_excel_v682_base(self, path)
    try:
        sched = _v682_build_tramo_schedule(getattr(self, "df_results", pd.DataFrame()))
        with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            sched.to_excel(writer, sheet_name="06D_Quadro_Prumada_Tramo", index=False)
    except Exception:
        pass

ColumnsEC2App._write_excel = _write_excel_v682


# DXF: usar SEMPRE o quadro por tramo gerado a partir de df_results, não df_summary colapsado.
def write_columns_dxf_v682(path: str, df: pd.DataFrame):
    sched = _v682_build_tramo_schedule(df)
    if sched is None or sched.empty:
        sched = df.copy() if df is not None else pd.DataFrame()
    try:
        sched = _v66_apply_constructive_detailing(sched) if "_v66_apply_constructive_detailing" in globals() else sched
    except Exception:
        pass
    parts = ["0\nSECTION\n2\nHEADER\n9\n$INSUNITS\n70\n4\n0\nENDSEC\n", _dxf_layer_table_v66() if "_dxf_layer_table_v66" in globals() else "0\nSECTION\n2\nTABLES\n0\nENDSEC\n", "0\nSECTION\n2\nENTITIES\n"]
    if sched is None or sched.empty:
        parts.append(_dxf_text(0, 0, "Sem resultados", 50, "COLUMNS_TEXT"))
    else:
        work = sched.copy()
        if "Prumada" not in work.columns:
            work["Prumada"] = work.apply(_v682_prumada_from_row, axis=1)
        if "Tramo" not in work.columns:
            work["Tramo"] = work.apply(_v682_segment_label, axis=1)
        prumadas = sorted(work["Prumada"].astype(str).unique(), key=_v682_natural_key)
        cell_w, cell_h = 1450.0, 1180.0
        margin_x, margin_y = 420.0, 950.0
        parts.append(_dxf_text(0, 330, "QUADRO DE PILARES - UNIDADES: mm", 52, "COLUMNS_TEXT"))
        parts.append(_dxf_text(0, 175, "Organização por prumada: cada coluna corresponde ao nome do pilar; alterações de secção/armadura são mantidas por tramo.", 25, "COLUMNS_TEXT"))
        parts.append(_dxf_text(0, 115, "Legenda: betão=contorno | varões=círculos | estribos=rectângulo interior | grampos=linhas interiores | cotas em mm", 24, "COLUMNS_TEXT"))
        for c, pr in enumerate(prumadas[:32]):
            x0 = margin_x + c * cell_w
            parts.append(_dxf_text(x0 + 15, 0, str(pr), 44, "COLUMNS_TABLE"))
            grp = work[work["Prumada"].astype(str) == str(pr)].copy()
            try:
                grp = grp.sort_values(["Ordem na prumada"], kind="mergesort")
            except Exception:
                pass
            for r_i, (_, r) in enumerate(grp.iterrows()):
                y0 = -margin_y - r_i * cell_h
                b = _finite(r.get("b_cm", r.get("hy", 0.0)), 0.0) * 10.0
                h = _finite(r.get("h_cm", r.get("hz", 0.0)), 0.0) * 10.0
                # moldura da célula
                parts += [
                    _dxf_line(x0, y0, x0 + cell_w - 80, y0, "COLUMNS_TABLE"),
                    _dxf_line(x0 + cell_w - 80, y0, x0 + cell_w - 80, y0 - cell_h + 60, "COLUMNS_TABLE"),
                    _dxf_line(x0 + cell_w - 80, y0 - cell_h + 60, x0, y0 - cell_h + 60, "COLUMNS_TABLE"),
                    _dxf_line(x0, y0 - cell_h + 60, x0, y0, "COLUMNS_TABLE"),
                ]
                if b <= 0 or h <= 0:
                    parts.append(_dxf_text(x0 + 25, y0 - 120, "Sem geometria", 26, "COLUMNS_TEXT")); continue
                ox = x0 + cell_w/2.0
                oy = y0 - cell_h/2.0 + 120.0
                scale = min(1.0, 650.0/max(b, h, 1.0))
                bs, hs = b*scale, h*scale
                left, right = ox-bs/2, ox+bs/2
                bot, top = oy-hs/2, oy+hs/2
                # betão e estribo
                parts += [_dxf_line(left,bot,right,bot,"COLUMNS_CONCRETE"), _dxf_line(right,bot,right,top,"COLUMNS_CONCRETE"), _dxf_line(right,top,left,top,"COLUMNS_CONCRETE"), _dxf_line(left,top,left,bot,"COLUMNS_CONCRETE")]
                cov = _finite(r.get("cover_mm"), 35.0)*scale
                l2, r2, b2, t2 = left+cov, right-cov, bot+cov, top-cov
                parts += [_dxf_line(l2,b2,r2,b2,"COLUMNS_STIRRUPS"), _dxf_line(r2,b2,r2,t2,"COLUMNS_STIRRUPS"), _dxf_line(r2,t2,l2,t2,"COLUMNS_STIRRUPS"), _dxf_line(l2,t2,l2,b2,"COLUMNS_STIRRUPS")]
                phi = _finite(r.get("phi_long_mm"), 10.0) * scale
                try:
                    pts = _bar_points_for_result(r)
                except Exception:
                    pts = []
                for yy, zz in pts:
                    parts.append(_dxf_circle(ox+yy*scale, oy+zz*scale, max(phi/2.0, 3.0), "COLUMNS_REBAR"))
                nlinks = int(_finite(r.get("numero_grampos_por_nivel", r.get("grampos_intermedios", 0)), 0))
                if nlinks > 0:
                    if _finite(r.get("n_bars_y"), 0) > 2:
                        parts.append(_dxf_line(l2, oy, r2, oy, "COLUMNS_LINKS"))
                    if _finite(r.get("n_bars_z"), 0) > 2:
                        parts.append(_dxf_line(ox, b2, ox, t2, "COLUMNS_LINKS"))
                # cotas
                _dxf_dim_text_v66(parts, left, bot-45, right, bot-45, f"{b:.0f}", off=-35)
                _dxf_dim_text_v66(parts, right+45, bot, right+45, top, f"{h:.0f}", off=0)
                tramo = str(r.get("Tramo", _v682_segment_label(r)))
                member = str(r.get("member", ""))
                mat = str(r.get("material", ""))
                sec = f"{b:.0f}x{h:.0f} mm"
                sol = str(r.get("solucao_completa", r.get("Solução", r.get("solucao", ""))))[:95]
                estado = str(r.get("estado_global", r.get("Estado", r.get("status", ""))))
                eta = _finite(r.get("η_NMyMz", r.get("eta_NMyMz", r.get("utilizacao", 0.0))), 0.0)
                parts.append(_dxf_text(x0+25, y0-70, f"{tramo} | Member {member} | {mat}", 23, "COLUMNS_TEXT"))
                parts.append(_dxf_text(x0+25, y0-105, f"Secção: {sec}", 22, "COLUMNS_TEXT"))
                parts.append(_dxf_text(x0+25, y0-140, sol, 21, "COLUMNS_TEXT"))
                parts.append(_dxf_text(x0+25, y0-175, f"Estado: {estado} | eta={eta:.2f}", 21, "COLUMNS_TEXT"))
                if nlinks > 0:
                    parts.append(_dxf_text(x0+25, y0-210, f"Grampos: {nlinks} por nível", 20, "COLUMNS_TEXT"))
    parts.append("0\nENDSEC\n0\nEOF\n")
    Path(path).write_text("".join(parts), encoding="utf-8")

write_columns_dxf_v66 = write_columns_dxf_v682
write_columns_dxf_v61 = write_columns_dxf_v682
write_columns_dxf_v64 = write_columns_dxf_v682
write_columns_dxf_v4 = write_columns_dxf_v682


def _export_dxf_v682(self):
    src = getattr(self, "df_results", pd.DataFrame())
    if src is None or src.empty:
        src = getattr(self, "df_summary", pd.DataFrame())
    if src is None or src.empty:
        messagebox.showwarning("Aviso", "Não há resultados para exportar em DXF.")
        return
    path = filedialog.asksaveasfilename(title="Exportar quadro de pilares por prumada [mm]", defaultextension=".dxf", filetypes=[("DXF", "*.dxf")])
    if not path:
        return
    try:
        self.status_var.set("A exportar quadro de pilares por prumada/tramo [mm]...")
        self.update_idletasks()
        write_columns_dxf_v682(path, src)
        self.status_var.set(f"DXF exportado: {path}")
    except Exception as err:
        messagebox.showerror("Erro", f"Não foi possível exportar DXF.\n\n{err}")

ColumnsEC2App.export_dxf = _export_dxf_v682


# ============================================================
# ColumnsEC2 v6.8.3 — ordenação por piso/story/andar e DXF alinhado por prumada
# - reconhece Story/Piso/Andar/Andares
# - cria índice de piso para ordenação ascendente: inferior -> superior
# - melhora 06D_Quadro_Prumada_Tramo e DXF com uma linha global por piso
# ============================================================
APP_VERSION = "v6.8.3"

# Reforço dos aliases de importação para o campo de piso.
try:
    COLUMN_ALIASES.update({
        "story": [
            "story", "storey", "piso", "piso/andar", "andar", "andares", "andare", "andars",
            "level", "floor", "floor level", "pavimento", "cota", "nível", "nivel", "storey name",
            "nome do piso", "nome piso", "piso estrutural",
        ],
    })
except Exception:
    pass


def _v683_story_sort_tuple_from_value(value, fallback=0):
    """Converte labels como PISO 1, Piso 07, Cave -1, RC, Cobertura em chave ordenável.

    A ordenação é ascendente: pisos inferiores primeiro, pisos superiores depois.
    """
    s_raw = str(value if value is not None else "").strip()
    s = s_raw.lower()
    s_norm = (
        s.replace("rés", "res")
         .replace("r/c", "rc")
         .replace("r.c.", "rc")
         .replace("r-c", "rc")
         .replace("_", " ")
    )
    if not s_raw or s_norm in ["nan", "none", "null", "-", "<na>"]:
        return (50_000, float(fallback), _v682_natural_key(s_raw))

    # Cobertura/topo sempre depois dos pisos correntes.
    if any(k in s_norm for k in ["cobertura", "roof", "top", "terraço", "terraco", "cover"]):
        return (10_000, 0.0, _v682_natural_key(s_raw))

    # Rés-do-chão / ground floor.
    if re.search(r"(^|\b)(rc|res do chao|res-do-chao|ground|terreo|térreo)(\b|$)", s_norm):
        return (0, 0.0, _v682_natural_key(s_raw))

    # Pisos negativos: cave/basement/subsolo.
    nums = re.findall(r"-?\d+(?:[\.,]\d+)?", s_norm)
    if any(k in s_norm for k in ["cave", "basement", "subsolo", "sub-solo", "parking", "garagem"]):
        if nums:
            try:
                n = abs(float(nums[-1].replace(",", ".")))
            except Exception:
                n = 1.0
        else:
            n = 1.0
        return (-int(n), -n, _v682_natural_key(s_raw))

    # Pisos numéricos correntes: PISO 1, Piso 07, Floor 12.
    if nums:
        try:
            n = float(nums[-1].replace(",", "."))
        except Exception:
            n = 0.0
        return (int(n), n, _v682_natural_key(s_raw))

    # Qualquer outro texto fica depois dos numéricos, mas com ordem natural.
    return (20_000, 0.0, _v682_natural_key(s_raw))


def _v683_story_label_from_row(row):
    label = _v682_story_from_row(row) if "_v682_story_from_row" in globals() else ""
    if not _v682_is_blank(label):
        return str(label).strip()
    # Sem Story, usar tramo/member como fallback visual.
    mem = str(row.get("member", "") if hasattr(row, "get") else "").strip()
    if mem:
        return f"Tramo {mem}"
    return "Tramo"


def _v683_story_sort_tuple(row):
    story = _v682_story_from_row(row) if "_v682_story_from_row" in globals() else ""
    fallback = _finite(row.get("__row_order", row.get("segment_order", 0)), 0.0) if hasattr(row, "get") else 0.0
    if not _v682_is_blank(story):
        return _v683_story_sort_tuple_from_value(story, fallback=fallback)
    mem = str(row.get("member", "") if hasattr(row, "get") else "")
    # Sem piso: preserva a ordem física/exportação.
    return (50_000, float(fallback), _v682_natural_key(mem))


def _v683_group_level_key(label):
    return _v683_story_sort_tuple_from_value(label)


# Limpeza/combinação de dados: garantir que a coluna story é preservada mesmo quando vem como Andar/Andares.
_old_clean_dataframe_v683 = clean_dataframe

def clean_dataframe_v683(df: pd.DataFrame) -> pd.DataFrame:
    out = _old_clean_dataframe_v683(df)
    if out is None:
        return out
    if "story" not in out.columns:
        out["story"] = ""
    # fallback: procurar colunas originais que possam ter escapado ao rename.
    if df is not None and not df.empty:
        candidate_cols = []
        for c in df.columns:
            cn = normalize_text(c)
            if cn in ["story", "storey", "piso", "andar", "andares", "level", "floor", "pavimento", "nivel", "nível"]:
                candidate_cols.append(c)
        if candidate_cols:
            src_col = candidate_cols[0]
            try:
                mask = out["story"].astype(str).str.strip().isin(["", "nan", "None", "none"])
                vals = df[src_col].astype(str).reset_index(drop=True)
                out.loc[mask, "story"] = vals.loc[mask].values
            except Exception:
                pass
    return out

clean_dataframe = clean_dataframe_v683


_old_design_one_v683_base = ColumnDesigner.design_one

def _design_one_v683(self, row: pd.Series, prebuilt_candidates=None):
    out = _old_design_one_v683_base(self, row, prebuilt_candidates=prebuilt_candidates)
    if isinstance(out, dict):
        story = _v682_story_from_row(row)
        out["story"] = story
        out["Piso"] = story
        out["story_sort_key"] = str(_v683_story_sort_tuple(row))
        out["Prumada"] = _v682_prumada_from_row(out)
    return out

ColumnDesigner.design_one = _design_one_v683


def _v683_build_tramo_schedule(results: pd.DataFrame) -> pd.DataFrame:
    """Quadro de decisão por prumada e por piso/tramo.

    Mantém secções diferentes na mesma prumada e ordena os tramos em ordem ascendente
    de piso: inferior -> superior.
    """
    if results is None or results.empty:
        return pd.DataFrame()
    work = results.copy()
    try:
        if "_v65_apply_module_statuses" in globals():
            work = _v65_apply_module_statuses(work)
    except Exception:
        pass
    try:
        if "_v68_apply_constructive_detailing" in globals():
            work = _v68_apply_constructive_detailing(work)
    except Exception:
        pass
    work["Prumada"] = work.apply(_v682_prumada_from_row, axis=1)
    work["Piso"] = work.apply(_v683_story_label_from_row, axis=1)
    work["_story_sort_tuple"] = work.apply(_v683_story_sort_tuple, axis=1)
    work["_story_rank"] = work["_story_sort_tuple"].map(lambda x: x[0] if isinstance(x, tuple) and x else 0)
    work["_story_rank_float"] = work["_story_sort_tuple"].map(lambda x: x[1] if isinstance(x, tuple) and len(x) > 1 else 0.0)
    work["_section_signature"] = work.apply(_v682_section_signature, axis=1)
    work["_gov_score_v683"] = _v682_governing_score_df(work) if "_v682_governing_score_df" in globals() else pd.Series(0.0, index=work.index)

    group_cols = ["Prumada"]
    if "member" in work.columns:
        group_cols.append("member")
    group_cols.append("Piso")
    group_cols.append("_section_signature")

    rows = []
    for _, grp in work.groupby(group_cols, dropna=False, sort=False):
        g = grp.sort_values("_gov_score_v683", ascending=False)
        r = g.iloc[0].copy()
        r["Prumada"] = _v682_prumada_from_row(r)
        r["Piso"] = _v683_story_label_from_row(r)
        r["Tramo"] = r["Piso"] if not _v682_is_blank(r["Piso"]) else _v682_segment_label(r)
        r["N.º combinações/tramo"] = len(grp)
        r["Secção [cm]"] = f"{_finite(r.get('b_cm', r.get('hy',0)),0):.0f}x{_finite(r.get('h_cm', r.get('hz',0)),0):.0f}"
        r["Solução"] = r.get("solucao_completa", r.get("solucao", ""))
        r["Estado"] = r.get("estado_global", r.get("status", ""))
        r["_story_sort_tuple"] = grp["_story_sort_tuple"].iloc[0]
        r["_story_rank"] = grp["_story_rank"].iloc[0]
        r["_story_rank_float"] = grp["_story_rank_float"].iloc[0]
        rows.append(r)
    out = pd.DataFrame(rows)
    if out.empty:
        return out

    # Ordenação final: Prumada natural; dentro de cada prumada, piso inferior -> superior.
    out["_prumada_sort"] = out["Prumada"].map(_v682_natural_key)
    out = out.sort_values(
        ["_prumada_sort", "_story_rank", "_story_rank_float", "member"],
        kind="mergesort",
    ).reset_index(drop=True)
    out["Ordem na prumada"] = out.groupby("Prumada").cumcount() + 1
    return out.drop(columns=[c for c in ["_gov_score_v683", "_section_signature", "_prumada_sort"] if c in out.columns], errors="ignore")

# Override global usado no Excel/PDF/DXF.
_v682_build_tramo_schedule = _v683_build_tramo_schedule


def _build_summary_by_member_v683(self, results: pd.DataFrame) -> pd.DataFrame:
    return _v683_build_tramo_schedule(results)

ColumnsEC2App.build_summary_by_member = _build_summary_by_member_v683


# Excel: substituir/actualizar a folha de quadro com Piso e ordem de piso.
_old_write_excel_v683_base = ColumnsEC2App._write_excel

def _write_excel_v683(self, path: str):
    try:
        if getattr(self, "df_results", pd.DataFrame()) is not None and not self.df_results.empty:
            self.df_summary = _v683_build_tramo_schedule(self.df_results)
    except Exception:
        pass
    _old_write_excel_v683_base(self, path)
    try:
        sched = _v683_build_tramo_schedule(getattr(self, "df_results", pd.DataFrame()))
        with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            sched.to_excel(writer, sheet_name="06D_Quadro_Prumada_Tramo", index=False)
            wb = writer.book
            try:
                from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
                from openpyxl.utils import get_column_letter
                ws = wb["06D_Quadro_Prumada_Tramo"]
                header_fill = PatternFill("solid", fgColor="1F4E5F")
                header_font = Font(color="FFFFFF", bold=True)
                thin = Side(style="thin", color="D9E2E7")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                ws.sheet_view.showGridLines = False
                ws.freeze_panes = "A2"
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = border
                for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 5000)):
                    for cell in row:
                        cell.border = border
                        cell.alignment = Alignment(vertical="top", wrap_text=True)
                for col_idx, col in enumerate(ws.columns, start=1):
                    values = [str(c.value) for c in col[:200] if c.value is not None]
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(max([len(v) for v in values] + [10]) + 2, 48)
            except Exception:
                pass
    except Exception:
        pass

ColumnsEC2App._write_excel = _write_excel_v683


# DXF v6.8.3: grelha global por piso, alinhando todos os P1/P2/P3... na mesma linha de piso.
def write_columns_dxf_v683(path: str, df: pd.DataFrame):
    sched = _v683_build_tramo_schedule(df)
    try:
        sched = _v66_apply_constructive_detailing(sched) if "_v66_apply_constructive_detailing" in globals() else sched
    except Exception:
        pass
    parts = ["0\nSECTION\n2\nHEADER\n9\n$INSUNITS\n70\n4\n0\nENDSEC\n", _dxf_layer_table_v66() if "_dxf_layer_table_v66" in globals() else "0\nSECTION\n2\nTABLES\n0\nENDSEC\n", "0\nSECTION\n2\nENTITIES\n"]
    if sched is None or sched.empty:
        parts.append(_dxf_text(0, 0, "Sem resultados", 50, "COLUMNS_TEXT"))
        parts.append("0\nENDSEC\n0\nEOF\n")
        Path(path).write_text("".join(parts), encoding="utf-8")
        return

    work = sched.copy()
    if "Prumada" not in work.columns:
        work["Prumada"] = work.apply(_v682_prumada_from_row, axis=1)
    if "Piso" not in work.columns:
        work["Piso"] = work.apply(_v683_story_label_from_row, axis=1)
    if "_story_sort_tuple" not in work.columns:
        work["_story_sort_tuple"] = work.apply(_v683_story_sort_tuple, axis=1)
    work["_story_rank"] = work["_story_sort_tuple"].map(lambda x: x[0] if isinstance(x, tuple) and x else 0)
    work["_story_rank_float"] = work["_story_sort_tuple"].map(lambda x: x[1] if isinstance(x, tuple) and len(x) > 1 else 0.0)

    prumadas = sorted(work["Prumada"].astype(str).unique(), key=_v682_natural_key)[:32]
    # Níveis globais alinhados por piso. Se não houver Story, usa os Tramos de cada linha.
    levels_df = work[["Piso", "_story_rank", "_story_rank_float", "_story_sort_tuple"]].drop_duplicates()
    levels_df = levels_df.sort_values(["_story_rank", "_story_rank_float"], kind="mergesort")
    levels = list(levels_df["Piso"].astype(str))[:24]

    cell_w, cell_h = 1450.0, 1180.0
    level_w = 420.0
    margin_x, base_y = 480.0, 0.0
    title_y = base_y + len(levels)*cell_h + 780.0
    header_y = base_y + len(levels)*cell_h + 420.0

    parts.append(_dxf_text(margin_x, title_y, "QUADRO DE PILARES - UNIDADES: mm", 52, "COLUMNS_TEXT"))
    parts.append(_dxf_text(margin_x, title_y-130, "Organização por prumada e piso: colunas = P1, P2, P3...; linhas = pisos em ordem ascendente, do inferior para o superior.", 25, "COLUMNS_TEXT"))
    parts.append(_dxf_text(margin_x, title_y-195, "Legenda: betão=contorno | varões=círculos | estribos=rectângulo interior | grampos=linhas interiores | cotas em mm", 24, "COLUMNS_TEXT"))

    # Cabeçalhos
    parts.append(_dxf_text(0, header_y, "Piso", 34, "COLUMNS_TABLE"))
    for c, pr in enumerate(prumadas):
        x0 = margin_x + c * cell_w
        parts.append(_dxf_text(x0 + cell_w*0.38, header_y, str(pr), 42, "COLUMNS_TABLE"))

    # Grelha exterior alinhada por piso
    total_w = level_w + len(prumadas)*cell_w
    total_h = len(levels)*cell_h
    left = 0.0
    bottom = base_y
    top = base_y + total_h
    right = left + total_w
    parts += [
        _dxf_line(left, bottom, right, bottom, "COLUMNS_TABLE"),
        _dxf_line(right, bottom, right, top, "COLUMNS_TABLE"),
        _dxf_line(right, top, left, top, "COLUMNS_TABLE"),
        _dxf_line(left, top, left, bottom, "COLUMNS_TABLE"),
        _dxf_line(level_w, bottom, level_w, top, "COLUMNS_TABLE"),
    ]
    for c in range(len(prumadas)+1):
        x = level_w + c*cell_w
        parts.append(_dxf_line(x, bottom, x, top, "COLUMNS_TABLE"))
    for r in range(len(levels)+1):
        y = base_y + r*cell_h
        parts.append(_dxf_line(left, y, right, y, "COLUMNS_TABLE"))

    # Índice por prumada/piso
    lookup = {}
    for _, r in work.iterrows():
        key = (str(r.get("Prumada", "")), str(r.get("Piso", "")))
        # se houver mais do que um resultado no mesmo piso/prumada, manter o mais governante
        if key not in lookup:
            lookup[key] = r
        else:
            old = pd.DataFrame([lookup[key]])
            new = pd.DataFrame([r])
            try:
                if float(_v682_governing_score_df(new).iloc[0]) > float(_v682_governing_score_df(old).iloc[0]):
                    lookup[key] = r
            except Exception:
                pass

    # Conteúdo: o primeiro nível fica na linha inferior; níveis seguintes sobem.
    for r_i, level in enumerate(levels):
        y0 = base_y + r_i*cell_h
        parts.append(_dxf_text(25, y0 + cell_h*0.47, str(level), 27, "COLUMNS_TABLE"))
        for c, pr in enumerate(prumadas):
            x0 = level_w + c*cell_w
            row = lookup.get((str(pr), str(level)))
            if row is None:
                # célula vazia
                cx = x0 + cell_w/2.0
                cy = y0 + cell_h/2.0
                parts.append(_dxf_line(cx-120, cy, cx+120, cy, "COLUMNS_TEXT"))
                continue

            b = _finite(row.get("b_cm", row.get("hy", 0.0)), 0.0) * 10.0
            h = _finite(row.get("h_cm", row.get("hz", 0.0)), 0.0) * 10.0
            if b <= 0 or h <= 0:
                parts.append(_dxf_text(x0 + 25, y0 + cell_h - 120, "Sem geometria", 26, "COLUMNS_TEXT"))
                continue
            ox = x0 + cell_w/2.0
            oy = y0 + cell_h/2.0 + 70.0
            scale = min(1.0, 620.0/max(b, h, 1.0))
            bs, hs = b*scale, h*scale
            lft, rgt = ox-bs/2, ox+bs/2
            bot, tp = oy-hs/2, oy+hs/2
            parts += [_dxf_line(lft,bot,rgt,bot,"COLUMNS_CONCRETE"), _dxf_line(rgt,bot,rgt,tp,"COLUMNS_CONCRETE"), _dxf_line(rgt,tp,lft,tp,"COLUMNS_CONCRETE"), _dxf_line(lft,tp,lft,bot,"COLUMNS_CONCRETE")]
            cov = _finite(row.get("cover_mm"), 35.0)*scale
            l2, r2, b2, t2 = lft+cov, rgt-cov, bot+cov, tp-cov
            parts += [_dxf_line(l2,b2,r2,b2,"COLUMNS_STIRRUPS"), _dxf_line(r2,b2,r2,t2,"COLUMNS_STIRRUPS"), _dxf_line(r2,t2,l2,t2,"COLUMNS_STIRRUPS"), _dxf_line(l2,t2,l2,b2,"COLUMNS_STIRRUPS")]
            phi = _finite(row.get("phi_long_mm"), 10.0)*scale
            try:
                pts = _bar_points_for_result(row)
            except Exception:
                pts = []
            for yy, zz in pts:
                parts.append(_dxf_circle(ox+yy*scale, oy+zz*scale, max(phi/2.0, 3.0), "COLUMNS_REBAR"))
            nlinks = int(_finite(row.get("numero_grampos_por_nivel", row.get("grampos_intermedios", 0)), 0))
            if nlinks > 0:
                if _finite(row.get("n_bars_y"), 0) > 2:
                    parts.append(_dxf_line(l2, oy, r2, oy, "COLUMNS_LINKS"))
                if _finite(row.get("n_bars_z"), 0) > 2:
                    parts.append(_dxf_line(ox, b2, ox, t2, "COLUMNS_LINKS"))
            # cotas em mm
            try:
                _dxf_dim_text_v66(parts, lft, bot-45, rgt, bot-45, f"{b:.0f}", off=-35)
                _dxf_dim_text_v66(parts, rgt+45, bot, rgt+45, tp, f"{h:.0f}", off=0)
            except Exception:
                pass
            member = str(row.get("member", ""))
            mat = str(row.get("material", ""))
            sec = f"{b:.0f}x{h:.0f} mm"
            sol = str(row.get("solucao_completa", row.get("Solução", row.get("solucao", ""))))[:95]
            estado = str(row.get("estado_global", row.get("Estado", row.get("status", ""))))
            eta = _finite(row.get("η_NMyMz", row.get("eta_NMyMz", row.get("utilizacao", 0.0))), 0.0)
            parts.append(_dxf_text(x0+25, y0+cell_h-70, f"{level} | Member {member} | {mat}", 22, "COLUMNS_TEXT"))
            parts.append(_dxf_text(x0+25, y0+cell_h-105, f"Secção: {sec}", 21, "COLUMNS_TEXT"))
            parts.append(_dxf_text(x0+25, y0+cell_h-140, sol, 20, "COLUMNS_TEXT"))
            parts.append(_dxf_text(x0+25, y0+cell_h-175, f"Estado: {estado} | eta={eta:.2f}", 20, "COLUMNS_TEXT"))
            if nlinks > 0:
                parts.append(_dxf_text(x0+25, y0+cell_h-210, f"Grampos: {nlinks} por nível", 19, "COLUMNS_TEXT"))

    parts.append("0\nENDSEC\n0\nEOF\n")
    Path(path).write_text("".join(parts), encoding="utf-8")

# Override dos exportadores DXF activos.
write_columns_dxf_v682 = write_columns_dxf_v683
write_columns_dxf_v66 = write_columns_dxf_v683
write_columns_dxf_v61 = write_columns_dxf_v683
write_columns_dxf_v64 = write_columns_dxf_v683
write_columns_dxf_v4 = write_columns_dxf_v683


def _export_dxf_v683(self):
    src = getattr(self, "df_results", pd.DataFrame())
    if src is None or src.empty:
        src = getattr(self, "df_summary", pd.DataFrame())
    if src is None or src.empty:
        messagebox.showwarning("Aviso", "Não há resultados para exportar em DXF.")
        return
    path = filedialog.asksaveasfilename(title="Exportar quadro de pilares por prumada e piso [mm]", defaultextension=".dxf", filetypes=[("DXF", "*.dxf")])
    if not path:
        return
    try:
        self.status_var.set("A exportar quadro de pilares por prumada/piso [mm]...")
        self.update_idletasks()
        write_columns_dxf_v683(path, src)
        self.status_var.set(f"DXF exportado por prumada e piso: {path}")
    except Exception as err:
        messagebox.showerror("Erro", f"Não foi possível exportar DXF.\n\n{err}")

ColumnsEC2App.export_dxf = _export_dxf_v683


# ============================================================
# ColumnsEC2 v6.8.4 — instruções profissionais e tabela-tipo com Story/Piso
# - actualiza o modelo de tabela para incluir Story/Piso;
# - melhora a descrição técnica da GUI;
# - substitui as instruções por um guia de utilização claro para novos utilizadores;
# - harmoniza notas rápidas e mensagens de entrada com a organização por prumada/piso.
# ============================================================
APP_VERSION = "v6.8.4"

try:
    ColumnsEC2App.TEMPLATE_COLUMNS = [
        "Member/Node/Case", "FX (kN)", "FY (kN)", "FZ (kN)", "MX (kNm)", "MY (kNm)", "MZ (kNm)",
        "Length (m)", "Material", "HY (cm)", "HZ (cm)", "VY (cm)", "VZ (cm)", "VPY (cm)", "VPZ (cm)",
        "AX (cm2)", "AY (cm2)", "AZ (cm2)", "IX (cm4)", "IY (cm4)", "IZ (cm4)", "Name", "Story"
    ]
except Exception:
    pass


def _v684_user_instructions_text(app=None):
    cols = getattr(app, "TEMPLATE_COLUMNS", getattr(ColumnsEC2App, "TEMPLATE_COLUMNS", []))
    return (
        "OBJECTIVO DO PROGRAMA\n"
        "ColumnsEC2 efectua o dimensionamento e a verificação de pilares de betão armado a partir de esforços nodais e propriedades geométricas da barra. "
        "O cálculo inclui selecção de casos governantes, verificação ELU, efeitos de 2.ª ordem, interacção N-My-Mz, verificações complementares, pormenorização construtiva, relatório técnico, ficheiro Excel e quadro de pilares em DXF.\n\n"

        "FLUXO DE UTILIZAÇÃO\n"
        "1. Prepare a tabela de entrada com as colunas indicadas no modelo.\n"
        "2. Cole a tabela ou importe o ficheiro Excel.\n"
        "3. Confirme a leitura das colunas, materiais, prumadas e pisos.\n"
        "4. Escolha o motor de cálculo, o modo de cálculo e a estratégia de armadura.\n"
        "5. Execute o cálculo.\n"
        "6. Consulte o resumo por prumada/piso, os estados por módulo e os avisos.\n"
        "7. Exporte o relatório PDF, o ficheiro Excel e o quadro de pilares em DXF.\n\n"

        "FORMATO DA TABELA TIPO\n"
        + " | ".join(cols) + "\n\n"

        "DESCRIÇÃO DAS COLUNAS PRINCIPAIS\n"
        "Member/Node/Case  — identificação da barra, do nó e da combinação/caso. Exemplo: 72/28/101 (C).\n"
        "FX, FY, FZ        — esforços nodais em kN no sistema local da barra.\n"
        "MX, MY, MZ        — momentos nodais em kNm no sistema local da barra.\n"
        "Length            — comprimento da barra em metros.\n"
        "Material          — classe de betão da barra, por exemplo C30/37 ou C40/50.\n"
        "HY, HZ            — dimensões principais da secção transversal em centímetros.\n"
        "AX, IY, IZ        — área e inércias geométricas da secção nas unidades indicadas.\n"
        "Name              — nome da prumada do pilar. Todos os elementos com o mesmo Name são agrupados na mesma prumada; por exemplo P1, P2, P17.\n"
        "Story             — piso/tramo associado ao elemento. Também são reconhecidos nomes equivalentes como Piso, Andar, Andares, Floor, Level ou Pavimento.\n\n"

        "CONVENÇÃO DE EIXOS E ESFORÇOS\n"
        "O eixo local X é assumido como o eixo longitudinal do pilar. FX é tratado como esforço axial; FY e FZ como esforços transversos; MX como torção; MY e MZ como momentos flectores. "
        "A reconstrução dos momentos de extremidade exige, idealmente, duas linhas por Member/Case, uma por nó extremo da barra.\n\n"

        "ORGANIZAÇÃO POR PRUMADA E PISO\n"
        "A coluna Name define a prumada. A coluna Story/Piso define a posição vertical do tramo. O programa ordena os tramos do piso inferior para o superior e preserva alterações de secção ou armadura dentro da mesma prumada. "
        "No DXF, cada prumada é apresentada numa coluna própria e cada piso/tramo numa linha.\n\n"

        "UNIDADES ESPERADAS\n"
        "FX, FY, FZ: kN; MX, MY, MZ: kNm; Length: m; HY/HZ/VY/VZ/VPY/VPZ: cm; AX/AY/AZ: cm²; IX/IY/IZ: cm⁴.\n\n"

        "VALIDAÇÕES AUTOMÁTICAS\n"
        "O programa verifica a existência das colunas necessárias, a consistência dos pares Member/Case, materiais em falta, dimensões suspeitas, número de nós por caso e coerência dos dados geométricos. "
        "As linhas com dados insuficientes são assinaladas para revisão.\n\n"

        "EXEMPLO DE TABELA\n"
        "Member/Node/Case    FX (kN)   FY (kN)   FZ (kN)   MX (kNm)   MY (kNm)   MZ (kNm)   Length (m)   Material   HY (cm)   HZ (cm)   AX (cm2)   IY (cm4)   IZ (cm4)   Name   Story\n"
        "72/28/101 (C)       420,50    3,20     -12,40    0,15       48,60      -7,80      3,67         C30/37     30,0      30,0      900,00     67500,00   67500,00   P17    PISO 1\n"
        "72/31/101 (C)       418,10    2,95     -11,90    0,12       45,20      -8,10      3,67         C30/37     30,0      30,0      900,00     67500,00   67500,00   P17    PISO 1\n\n"

        "RECOMENDAÇÕES DE UTILIZAÇÃO\n"
        "Para cálculo final, confirme sempre os pilares condicionantes, os comprimentos efectivos, a coerência dos sinais de MY/MZ, a classe de betão por tramo e a pormenorização construtiva proposta. "
        "O ficheiro Excel mantém a memória completa por caso, e o PDF apresenta a síntese técnica por prumada/tramo."
    )


def _build_instructions_tab_v684(self, parent):
    outer = ttk.Frame(parent, padding=10)
    outer.pack(fill="both", expand=True)
    outer.rowconfigure(1, weight=1)
    outer.columnconfigure(0, weight=1)
    ttk.Label(
        outer,
        text="Instruções de utilização e tabela tipo",
        style="Header.TLabel"
    ).grid(row=0, column=0, sticky="w", pady=(0, 8))
    text_host = ttk.Frame(outer)
    text_host.grid(row=1, column=0, sticky="nsew")
    txt = self._make_text_view(text_host)
    txt.insert("1.0", _v684_user_instructions_text(self))
    txt.config(state="disabled")

ColumnsEC2App._build_instructions_tab = _build_instructions_tab_v684


_old_build_sidebar_v684 = ColumnsEC2App._build_sidebar

def _build_sidebar_v684(self, parent):
    _old_build_sidebar_v684(self, parent)

    technical_description = (
        "Dimensionamento e verificação técnica de pilares de betão armado com interacção N-My-Mz, "
        "2.ª ordem, pormenorização construtiva e relatórios profissionais."
    )
    quick_text = (
        "• Name define a prumada; Story/Piso define a ordem vertical dos tramos.\n"
        "• A tabela deve conter esforços nodais, geometria, material, prumada e piso.\n"
        "• O eixo local X é o eixo do pilar; MY e MZ são os momentos de flexão.\n"
        "• Alterações de secção na mesma prumada são preservadas por piso/tramo.\n"
        "• O DXF organiza uma coluna por prumada e uma linha por piso/tramo."
    )

    def walk(widget):
        for child in widget.winfo_children():
            yield child
            yield from walk(child)

    for w in walk(parent):
        try:
            txt = str(w.cget("text"))
        except Exception:
            continue
        if "Ferramenta para importação de esforços" in txt:
            try:
                w.configure(text=technical_description)
            except Exception:
                pass
        elif txt.startswith("• Entrada tipo:") or "MX/torção" in txt:
            try:
                w.configure(text=quick_text)
            except Exception:
                pass
        elif txt == "Importar .xlsx/.csv":
            try:
                w.configure(text="Importar ficheiro")
            except Exception:
                pass
        elif txt == "Modelo de tabela":
            try:
                w.configure(text="Modelo de tabela tipo")
            except Exception:
                pass
        elif txt == "Ler caixa de texto":
            try:
                w.configure(text="Interpretar tabela")
            except Exception:
                pass

ColumnsEC2App._build_sidebar = _build_sidebar_v684


_old_build_paste_tab_v684 = ColumnsEC2App._build_paste_tab

def _build_paste_tab_v684(self, parent):
    _old_build_paste_tab_v684(self, parent)
    def walk(widget):
        for child in widget.winfo_children():
            yield child
            yield from walk(child)
    for w in walk(parent):
        try:
            txt = str(w.cget("text"))
        except Exception:
            continue
        if "Cole aqui a tabela" in txt:
            try:
                w.configure(text="Cole a tabela de esforços, geometria, prumada e piso; depois clique em 'Interpretar tabela'.")
            except Exception:
                pass
        elif txt == "Ler caixa de texto":
            try:
                w.configure(text="Interpretar tabela")
            except Exception:
                pass

ColumnsEC2App._build_paste_tab = _build_paste_tab_v684


def _export_template_v684(self):
    path = filedialog.asksaveasfilename(
        title="Guardar modelo de tabela tipo",
        defaultextension=".xlsx",
        filetypes=[("Excel workbook", "*.xlsx")]
    )
    if not path:
        return
    root, ext = os.path.splitext(path)
    if ext.lower() != ".xlsx":
        path = root + ".xlsx" if ext else path + ".xlsx"

    sample = pd.DataFrame([
        {
            "Member/Node/Case": "72/28/101 (C)", "FX (kN)": "420,50", "FY (kN)": "3,20", "FZ (kN)": "-12,40",
            "MX (kNm)": "0,15", "MY (kNm)": "48,60", "MZ (kNm)": "-7,80",
            "Length (m)": "3,67", "Material": "C30/37", "HY (cm)": "30,0", "HZ (cm)": "30,0",
            "VY (cm)": "15,0", "VZ (cm)": "15,0", "VPY (cm)": "15,0", "VPZ (cm)": "15,0",
            "AX (cm2)": "900,00", "AY (cm2)": "750,00", "AZ (cm2)": "750,00",
            "IX (cm4)": "113872,30", "IY (cm4)": "67500,00", "IZ (cm4)": "67500,00",
            "Name": "P17", "Story": "PISO 1"
        },
        {
            "Member/Node/Case": "72/31/101 (C)", "FX (kN)": "418,10", "FY (kN)": "2,95", "FZ (kN)": "-11,90",
            "MX (kNm)": "0,12", "MY (kNm)": "45,20", "MZ (kNm)": "-8,10",
            "Length (m)": "3,67", "Material": "C30/37", "HY (cm)": "30,0", "HZ (cm)": "30,0",
            "VY (cm)": "15,0", "VZ (cm)": "15,0", "VPY (cm)": "15,0", "VPZ (cm)": "15,0",
            "AX (cm2)": "900,00", "AY (cm2)": "750,00", "AZ (cm2)": "750,00",
            "IX (cm4)": "113872,30", "IY (cm4)": "67500,00", "IZ (cm4)": "67500,00",
            "Name": "P17", "Story": "PISO 1"
        },
        {
            "Member/Node/Case": "73/40/201 (C)", "FX (kN)": "365,80", "FY (kN)": "2,10", "FZ (kN)": "-9,60",
            "MX (kNm)": "0,08", "MY (kNm)": "36,40", "MZ (kNm)": "-5,20",
            "Length (m)": "3,20", "Material": "C30/37", "HY (cm)": "25,0", "HZ (cm)": "40,0",
            "VY (cm)": "12,5", "VZ (cm)": "20,0", "VPY (cm)": "12,5", "VPZ (cm)": "20,0",
            "AX (cm2)": "1000,00", "AY (cm2)": "833,33", "AZ (cm2)": "833,33",
            "IX (cm4)": "170833,33", "IY (cm4)": "133333,33", "IZ (cm4)": "52083,33",
            "Name": "P17", "Story": "PISO 2"
        },
        {
            "Member/Node/Case": "73/44/201 (C)", "FX (kN)": "363,70", "FY (kN)": "1,95", "FZ (kN)": "-9,10",
            "MX (kNm)": "0,06", "MY (kNm)": "34,90", "MZ (kNm)": "-5,60",
            "Length (m)": "3,20", "Material": "C30/37", "HY (cm)": "25,0", "HZ (cm)": "40,0",
            "VY (cm)": "12,5", "VZ (cm)": "20,0", "VPY (cm)": "12,5", "VPZ (cm)": "20,0",
            "AX (cm2)": "1000,00", "AY (cm2)": "833,33", "AZ (cm2)": "833,33",
            "IX (cm4)": "170833,33", "IY (cm4)": "133333,33", "IZ (cm4)": "52083,33",
            "Name": "P17", "Story": "PISO 2"
        },
    ], columns=ColumnsEC2App.TEMPLATE_COLUMNS)

    notes = pd.DataFrame([
        ["Name", "Identifica a prumada. Todos os elementos com o mesmo Name são agrupados no mesmo alinhamento vertical."],
        ["Story", "Identifica o piso/tramo. É usado para ordenar os elementos do piso inferior para o superior e para alinhar o quadro DXF."],
        ["Member/Node/Case", "Cada barra/caso deve ter, preferencialmente, duas linhas correspondentes aos dois nós extremos."],
        ["Material", "Classe de betão por linha/tramo, por exemplo C30/37. Podem existir várias classes no mesmo ficheiro."],
        ["Unidades", "Esforços em kN/kNm; comprimento em m; dimensões em cm; áreas em cm²; inércias em cm⁴."],
    ], columns=["Campo", "Descrição"])

    try:
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            sample.to_excel(writer, sheet_name="Tabela_Tipo", index=False)
            notes.to_excel(writer, sheet_name="Notas", index=False)
            wb = writer.book
            wb.properties.title = f"{APP_NAME} - tabela tipo"
            wb.properties.creator = APP_AUTHOR
            try:
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils import get_column_letter
                fill = PatternFill("solid", fgColor="1F4E5F")
                font = Font(color="FFFFFF", bold=True)
                thin = Side(style="thin", color="D9E2E7")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                for ws in wb.worksheets:
                    ws.sheet_view.showGridLines = False
                    ws.freeze_panes = "A2"
                    for cell in ws[1]:
                        cell.fill = fill
                        cell.font = font
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        cell.border = border
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                        for cell in row:
                            cell.border = border
                            cell.alignment = Alignment(vertical="top", wrap_text=True)
                    for col_idx, col in enumerate(ws.columns, start=1):
                        vals = [str(c.value) for c in col[:100] if c.value is not None]
                        ws.column_dimensions[get_column_letter(col_idx)].width = min(max([len(v) for v in vals] + [10]) + 2, 32)
            except Exception:
                pass
        self.status_var.set(f"Modelo de tabela tipo guardado: {path}")
    except Exception as err:
        messagebox.showerror("Erro", f"Não foi possível guardar o modelo de tabela tipo.\n\n{err}")

ColumnsEC2App.export_template = _export_template_v684


# Actualizar descrição técnica no workbook, sem expor detalhes internos da GUI.
APP_XLSX_DESCRIPTION = (
    "Dimensionamento e verificação técnica de pilares de betão armado por prumada e piso, "
    "incluindo selecção de casos governantes, interacção N-My-Mz, efeitos de 2.ª ordem, "
    "pormenorização construtiva, relatórios técnicos e quadro de pilares em DXF."
)



# ============================================================
# ColumnsEC2 v6.9 — Interface bilingue PT/EN-UK
# - selector PT/EN na GUI;
# - instruções, notas rápidas, mensagens principais e botões traduzidos;
# - PDF técnico bilingue;
# - XLSX com metadados/parâmetros/notas/headers/sheets traduzidos quando EN;
# - DXF com legenda/títulos em inglês quando EN.
# ============================================================
APP_VERSION = "v6.9"

LANG_PT = "PT"
LANG_EN = "EN"

_V69_UI_MAP_PT_EN = {
    "ColumnsEC2": "ColumnsEC2",
    "Dimensionamento de pilares de betão armado (EC2)": "Design of reinforced concrete columns (EC2)",
    "Dimensionamento e verificação técnica de pilares de betão armado com interacção N-My-Mz, 2.ª ordem, pormenorização construtiva e relatórios profissionais.": "Technical design and checking of reinforced concrete columns with N-My-Mz interaction, second-order effects, detailing and professional reports.",
    "1. Entrada": "1. Input",
    "Colar área de transferência": "Paste clipboard",
    "Importar ficheiro": "Import file",
    "Interpretar tabela": "Parse table",
    "Modelo de tabela tipo": "Table template",
    "2. Parâmetros EC2": "2. Design parameters",
    "Recobrimento [mm]": "Nominal cover [mm]",
    "Classe de Aço": "Steel grade",
    "Aço fyk [MPa]": "Steel grade",
    "Betão": "Concrete",
    "lido da tabela (coluna Material)": "read from table (Material column)",
    "Modo": "Mode",
    "Reduzir para casos governantes": "Reduce to governing cases",
    "Resumo por membro": "Member summary",
    "3. Filtros": "3. Filters",
    "Member": "Member",
    "Estado": "Status",
    "Falha": "Failure",
    "Todos": "All",
    "Aplicar": "Apply",
    "Limpar": "Clear",
    "4. Cálculo e exportação": "4. Design and export",
    "Calcular": "Design/check",
    "Exportar .xlsx": "Export .xlsx",
    "Relatório .pdf": "PDF report",
    "Exportar quadro .DXF": "Export column schedule .DXF",
    "Abrir repositório": "Open repository",
    "5. Estado": "5. Status",
    "6. Notas rápidas": "6. Quick notes",
    "7. Verificações avançadas v4": "7. Advanced checks",
    "Combinação ELS": "SLS combination",
    "em branco = ELS simplificado por defeito": "blank = simplified SLS by default",
    "Instruções": "Instructions",
    "Colar": "Paste",
    "Tabela": "Table",
    "Pares": "Pairs",
    "Validação": "Validation",
    "Resultados": "Results",
    "Resumo": "Summary",
    "Falhas": "Failures",
    "Shortlists": "Shortlists",
    "Relatório": "Report",
    "Notas EC2": "EC2 Notes",
    "Qualidade Importação": "Input quality",
    "Cobertura Backend": "Backend scope",
    "Diagnóstico SC": "SC diagnostics",
    "Interacção N-My-Mz": "N-My-Mz interaction",
    "Performance": "Performance",
    "Estratégia armadura": "Reinforcement strategy",
    "Económica": "Economic",
    "Equilibrada": "Balanced",
    "Robusta": "Robust",
    "Resumo executivo": "Executive summary",
    "Relatório técnico": "Technical report",
    "Memória de cálculo": "Detailed calculation note",
    "Linguagem": "Language",
    "Português": "Portuguese",
    "Inglês (UK)": "English (UK)",
}
_V69_UI_MAP_EN_PT = {v: k for k, v in _V69_UI_MAP_PT_EN.items()}

_V69_CELL_MAP_PT_EN = {
    "Programa": "Program",
    "Autor / Repositório": "Author / Repository",
    "Autor": "Author",
    "Repositório": "Repository",
    "Data de exportação": "Export date",
    "Ficheiro de origem": "Source file",
    "Norma de referência": "Reference standard",
    "Âmbito": "Scope",
    "Descrição": "Description",
    "Limitações": "Technical limitations",
    "Parâmetro": "Parameter",
    "Valor": "Value",
    "Recobrimento [mm]": "Nominal cover [mm]",
    "Classe de Aço": "Steel grade",
    "Betão": "Concrete",
    "Classes de betão lidas da coluna Material": "Concrete grades read from the Material column",
    "Modo de cálculo": "Design mode",
    "Redução para casos governantes": "Reduction to governing cases",
    "Estratégia de armadura": "Reinforcement strategy",
    "Tema": "Topic",
    "Referência": "Reference",
    "Nota": "Note",
    "Prumada": "Column line",
    "Piso": "Storey",
    "Tramo": "Segment",
    "Solução": "Reinforcement solution",
    "Estado": "Status",
    "Falha": "Failure",
    "Aviso": "Warning",
    "OK": "OK",
    "Secção [cm]": "Section [cm]",
    "Secção": "Section",
    "Caso": "Case",
    "Critério": "Criterion",
    "Estado resistente": "Resistance status",
    "Estado global": "Overall status",
    "Decisão técnica": "Technical decision",
    "Pormenorização": "Detailing",
    "Corte": "Shear",
    "Torção": "Torsion",
    "ELS": "SLS",
    "Grampos": "Cross-ties",
    "Ramos": "Tie legs",
    "A calcular": "Calculating",
    "Concluído": "Completed",
    "Sim": "Yes",
    "Não": "No",
}

_V69_SHEET_MAP_PT_EN = {
    "00_Info": "00_Info",
    "01_Parametros": "01_Parameters",
    "01B_Estrategia": "01B_Strategy",
    "01C_Criterios": "01C_Criteria",
    "02_Entrada_Dados": "02_Input_Data",
    "03_Pares_Member_Case": "03_Member_Case_Pairs",
    "04_Casos_Calculo": "04_Design_Cases",
    "04_Casos_Governantes": "04_Governing_Cases",
    "05_Resultados": "05_Results",
    "06_Resumo_Membros": "06_Member_Summary",
    "06B_Estados_Modulo": "06B_Module_Status",
    "06C_Decisao_Prumadas": "06C_Column_Line_Decision",
    "06D_Quadro_Prumada_Tramo": "06D_Column_Line_Schedule",
    "07_Falhas": "07_Failures",
    "07B_Avisos": "07B_Warnings",
    "08_OK": "08_OK",
    "09_Shortlists": "09_Shortlists",
    "10_ELS": "10_SLS",
    "11_V_Torcao": "11_Shear_Torsion",
    "12_Performance": "12_Performance",
    "13_Notas_EC2": "13_EC2_Notes",
    "13_Notas_FIB_MC10": "13_FIB_MC10_Notes",
    "14_Interacao_NMyMz": "14_NMyMz_Interaction",
    "20_Superficie_NMyMz": "20_NMyMz_Surface",
    "26_Tentativas_Correcao": "26_Correction_Attempts",
    "30_Diagnostico_SC": "30_SC_Diagnostics",
}

_V69_HEADER_MAP_PT_EN = {
    "Prumada": "Column line",
    "Piso": "Storey",
    "Tramo": "Segment",
    "N.º tramos/casos": "No. of segments/cases",
    "Solução": "Reinforcement solution",
    "Estado": "Status",
    "Estado global": "Overall status",
    "estado_global": "overall_status",
    "estado_resistente": "resistance_status",
    "estado_corte": "shear_status",
    "estado_torcao": "torsion_status",
    "estado_els": "sls_status",
    "estado_pormenorizacao": "detailing_status",
    "decisao_tecnica": "technical_decision",
    "η_NMyMz": "η_NMyMz",
    "eta_NMyMz": "η_NMyMz",
    "solucao": "reinforcement_solution",
    "solucao_completa": "full_reinforcement_solution",
    "pormenorizacao_construtiva": "constructive_detailing",
    "grampos_intermedios": "intermediate_cross_ties",
    "numero_grampos_por_nivel": "cross_ties_per_level",
    "ramos_estribo_y": "tie_legs_y",
    "ramos_estribo_z": "tie_legs_z",
    "member": "member",
    "case": "case",
    "name": "name",
    "material": "material",
    "failure_reason": "failure_reason",
    "failure_type": "failure_type",
    "recommendations": "recommendations",
    "as_req_mm2": "As_req_mm2",
    "as_prov_mm2": "As_prov_mm2",
    "n_ed_kN": "N_Ed_kN",
    "my_ed_kNm": "My_Ed_kNm",
    "mz_ed_kNm": "Mz_Ed_kNm",
    "mrd_y_kNm": "MRd_y_kNm",
    "mrd_z_kNm": "MRd_z_kNm",
}


def _v69_lang(app):
    try:
        return LANG_EN if str(app.var_language.get()).upper().startswith("EN") else LANG_PT
    except Exception:
        return LANG_PT


def _v69_is_en(app):
    return _v69_lang(app) == LANG_EN


def _v69_tr_text(text, target_lang):
    s = str(text)
    if target_lang == LANG_EN:
        return _V69_UI_MAP_PT_EN.get(s, _V69_CELL_MAP_PT_EN.get(s, s))
    return _V69_UI_MAP_EN_PT.get(s, s)


def _v69_status_text(value, lang):
    if lang != LANG_EN:
        return value
    s = str(value)
    return {
        "Aviso": "Warning",
        "Falha": "Failure",
        "Pré-dimensionado": "Preliminary",
        "Não avaliado": "Not assessed",
        "Verificar": "Check",
        "Não conforme": "Not compliant",
        "Dispensada": "Waived",
        "Sem aviso relevante": "No relevant warning",
        "Torção desprezável — não condicionante": "Negligible torsion — not governing",
    }.get(s, s)


def _v69_technical_description(lang):
    if lang == LANG_EN:
        return (
            "Technical design and checking of reinforced concrete columns by column line and storey, including governing-case selection, "
            "N-My-Mz interaction, second-order effects, constructive detailing, technical reports and DXF column schedules."
        )
    return (
        "Dimensionamento e verificação técnica de pilares de betão armado por prumada e piso, incluindo selecção de casos governantes, "
        "interacção N-My-Mz, efeitos de 2.ª ordem, pormenorização construtiva, relatórios técnicos e quadro de pilares em DXF."
    )


def _v69_module_description(app):
    lang = _v69_lang(app)
    try:
        backend = _v59_backend_key(app)
    except Exception:
        backend = "pt2010"
    if lang == LANG_EN:
        desc = {
            "pt2010": "Internal Portuguese EC2 2010 design engine for reinforced concrete columns, including second-order effects, N-My-Mz interaction, shear/torsion checks, constructive detailing, reports and DXF output.",
            "ec2_2004": "Eurocode 2:2004 backend using structuralcodes where available for material, creep/shrinkage and section-related calculations; ColumnsEC2 manages case selection, detailing, reporting and drawing output.",
            "ec2_2023": "Eurocode 2:2023 backend using structuralcodes where available for material, creep/shrinkage, serviceability and section-related calculations; unavailable package functions are reported as not assessed.",
            "fib2010": "fib Model Code 2010 backend using structuralcodes where available for material, time-dependent effects and resistance checks; ColumnsEC2 manages detailing, reporting and drawing output.",
        }
        return desc.get(backend, desc["pt2010"])
    desc = {
        "pt2010": "Motor interno português EC2 2010 para dimensionamento de pilares de betão armado, incluindo 2.ª ordem, interacção N-My-Mz, corte/torção, pormenorização construtiva, relatórios e DXF.",
        "ec2_2004": "Motor Eurocode 2:2004 via structuralcodes quando disponível para materiais, fluência/retracção e cálculo seccional; o ColumnsEC2 gere casos, pormenorização, relatórios e desenho.",
        "ec2_2023": "Motor Eurocode 2:2023 via structuralcodes quando disponível para materiais, fluência/retracção, serviço e cálculo seccional; funções não expostas pelo pacote são assinaladas como não avaliadas.",
        "fib2010": "Motor fib Model Code 2010 via structuralcodes quando disponível para materiais, efeitos diferidos e verificações resistentes; o ColumnsEC2 gere pormenorização, relatórios e desenho.",
    }
    return desc.get(backend, desc["pt2010"])


def _v69_module_limitations(app):
    lang = _v69_lang(app)
    try:
        backend = _v59_backend_key(app)
    except Exception:
        backend = "pt2010"
    if lang == LANG_EN:
        lim = {
            "pt2010": "Results depend on the imported design actions, effective-length assumptions, material data and adopted detailing constraints. Critical columns should be independently reviewed.",
            "ec2_2004": "Strict package-based checks are limited by the functions exposed by the installed structuralcodes version. Detailing and drawing logic are handled by ColumnsEC2.",
            "ec2_2023": "Strict package-based checks are limited by the functions exposed by the installed structuralcodes version. Modules not exposed by the package are reported as not assessed.",
            "fib2010": "Strict package-based checks are limited by the functions exposed by the installed structuralcodes version. Detailing and drawing logic are handled by ColumnsEC2.",
        }
        return lim.get(backend, lim["pt2010"])
    lim = {
        "pt2010": "Os resultados dependem dos esforços importados, hipóteses de comprimento efectivo, materiais e critérios de pormenorização adoptados. Pilares críticos devem ser revistos independentemente.",
        "ec2_2004": "As verificações estritas dependem das funções expostas pela versão instalada do structuralcodes. A pormenorização e o desenho são geridos pelo ColumnsEC2.",
        "ec2_2023": "As verificações estritas dependem das funções expostas pela versão instalada do structuralcodes. Módulos não expostos pelo pacote são reportados como não avaliados.",
        "fib2010": "As verificações estritas dependem das funções expostas pela versão instalada do structuralcodes. A pormenorização e o desenho são geridos pelo ColumnsEC2.",
    }
    return lim.get(backend, lim["pt2010"])


def _v69_instructions_text(app=None):
    lang = _v69_lang(app) if app is not None else LANG_PT
    cols = getattr(app, "TEMPLATE_COLUMNS", getattr(ColumnsEC2App, "TEMPLATE_COLUMNS", []))
    if lang == LANG_EN:
        return (
            "PROGRAMME PURPOSE\n"
            "ColumnsEC2 designs and checks reinforced concrete columns using nodal actions and member geometric properties. The workflow includes governing-case selection, ULS verification, second-order effects, N-My-Mz interaction, complementary checks, constructive detailing, technical PDF reporting, Excel output and a DXF column schedule.\n\n"
            "WORKFLOW\n"
            "1. Prepare the input table using the template columns.\n"
            "2. Paste the table or import the Excel file.\n"
            "3. Check the recognised columns, materials, column lines and storeys.\n"
            "4. Select the design engine, design mode and reinforcement strategy.\n"
            "5. Run the calculation.\n"
            "6. Review the summary by column line/storey, module status and warnings.\n"
            "7. Export the PDF report, Excel workbook and DXF column schedule.\n\n"
            "INPUT TABLE FORMAT\n" + " | ".join(cols) + "\n\n"
            "MAIN COLUMNS\n"
            "Member/Node/Case  — member, node and load case/combination identifier. Example: 72/28/101 (C).\n"
            "FX, FY, FZ        — nodal forces in kN in the member local coordinate system.\n"
            "MX, MY, MZ        — nodal moments in kNm in the member local coordinate system.\n"
            "Length            — member length in metres.\n"
            "Material          — concrete grade for the member, for example C30/37 or C40/50.\n"
            "HY, HZ            — main cross-section dimensions in centimetres.\n"
            "AX, IY, IZ        — area and second moments of area in the stated units.\n"
            "Name              — column line name. Elements with the same Name are grouped vertically, for example P1, P2, P17.\n"
            "Story             — storey/segment associated with the element. Equivalent headings such as Piso, Andar, Floor, Level or Pavimento are also recognised.\n\n"
            "AXES AND ACTION CONVENTION\n"
            "The local X axis is assumed to be the column longitudinal axis. FX is treated as axial force; FY and FZ as shear forces; MX as torsion; MY and MZ as bending moments. Ideally, each Member/Case has two rows, one for each end node.\n\n"
            "COLUMN LINE AND STOREY ORGANISATION\n"
            "The Name column defines the column line. The Story/Storey column defines the vertical position of the segment. The programme sorts segments from lower to upper storeys and preserves cross-section or reinforcement changes within the same column line.\n\n"
            "EXPECTED UNITS\n"
            "FX, FY, FZ: kN; MX, MY, MZ: kNm; Length: m; HY/HZ/VY/VZ/VPY/VPZ: cm; AX/AY/AZ: cm²; IX/IY/IZ: cm⁴.\n\n"
            "AUTOMATIC VALIDATION\n"
            "The programme checks required columns, Member/Case pairs, missing materials, suspicious dimensions, number of nodes per case and consistency of geometric data. Lines with insufficient data are flagged for review.\n\n"
            "RECOMMENDATIONS\n"
            "For final design, check the governing columns, effective lengths, MY/MZ signs, concrete grade by segment and the proposed detailing. The Excel workbook keeps the full calculation record by case; the PDF presents the technical summary by column line/segment."
        )
    return _v684_user_instructions_text(app)


def _v69_quick_notes(lang):
    if lang == LANG_EN:
        return (
            "• Name defines the column line; Story/Storey defines the vertical order of segments.\n"
            "• The input table must include nodal actions, geometry, material, column line and storey.\n"
            "• The local X axis is the column axis; MY and MZ are bending moments.\n"
            "• Cross-section changes within the same column line are preserved by storey/segment.\n"
            "• The DXF uses one column per column line and one row per storey/segment."
        )
    return (
        "• Name define a prumada; Story/Piso define a ordem vertical dos tramos.\n"
        "• A tabela deve conter esforços nodais, geometria, material, prumada e piso.\n"
        "• O eixo local X é o eixo do pilar; MY e MZ são os momentos de flexão.\n"
        "• Alterações de secção na mesma prumada são preservadas por piso/tramo.\n"
        "• O DXF organiza uma coluna por prumada e uma linha por piso/tramo."
    )


def _v69_refresh_instructions(self):
    try:
        txt_widgets = []
        def walk(w):
            for ch in w.winfo_children():
                if isinstance(ch, tk.Text):
                    txt_widgets.append(ch)
                walk(ch)
        walk(getattr(self, "tab_instructions", self))
        for t in txt_widgets:
            t.config(state="normal")
            t.delete("1.0", "end")
            t.insert("1.0", _v69_instructions_text(self))
            t.config(state="disabled")
    except Exception:
        pass


def _v69_apply_language(self):
    lang = _v69_lang(self)
    target = lang
    # Widget text
    def walk(widget):
        for child in widget.winfo_children():
            yield child
            yield from walk(child)
    for w in walk(self):
        try:
            txt = str(w.cget("text"))
        except Exception:
            continue
        new = _v69_tr_text(txt, target)
        if new != txt:
            try:
                w.configure(text=new)
            except Exception:
                pass
    # Notebook tabs
    for nb in [w for w in walk(self) if isinstance(w, ttk.Notebook)]:
        for tab_id in nb.tabs():
            try:
                txt = nb.tab(tab_id, "text")
                nb.tab(tab_id, text=_v69_tr_text(txt, target))
            except Exception:
                pass
    # Main title and instructions
    try:
        self.title("ColumnsEC2 - Reinforced Concrete Column Design (EC2)" if lang == LANG_EN else APP_TITLE)
    except Exception:
        pass
    _v69_refresh_instructions(self)
    # Quick notes / description text replacement by content pattern
    try:
        for w in walk(self):
            try:
                txt = str(w.cget("text"))
            except Exception:
                continue
            if "Technical design and checking" in txt or "Dimensionamento e verificação técnica" in txt:
                w.configure(text=_v69_technical_description(lang))
            if ("Name defines the column line" in txt) or ("Name define a prumada" in txt):
                w.configure(text=_v69_quick_notes(lang))
    except Exception:
        pass
    try:
        self.status_var.set("Language set to English (UK)." if lang == LANG_EN else "Idioma definido para Português.")
    except Exception:
        pass


_old_build_sidebar_v69 = ColumnsEC2App._build_sidebar

def _build_sidebar_v69(self, parent):
    if not hasattr(self, "var_language"):
        self.var_language = tk.StringVar(value=LANG_PT)
    _old_build_sidebar_v69(self, parent)
    lang_box = ttk.LabelFrame(parent, text="Linguagem")
    lang_box.pack(fill="x", pady=(0, 8))
    ttk.Combobox(lang_box, textvariable=self.var_language, values=["PT", "EN-UK"], state="readonly", width=14).grid(row=0, column=0, sticky="ew", padx=6, pady=4)
    ttk.Button(lang_box, text="Aplicar", command=lambda: _v69_apply_language(self)).grid(row=0, column=1, sticky="ew", padx=6, pady=4)
    lang_box.columnconfigure(0, weight=1)

ColumnsEC2App._build_sidebar = _build_sidebar_v69


_old_init_v69 = ColumnsEC2App.__init__
def _init_v69(self, *args, **kwargs):
    _old_init_v69(self, *args, **kwargs)
    try:
        if not hasattr(self, "var_language"):
            self.var_language = tk.StringVar(value=LANG_PT)
    except Exception:
        pass
ColumnsEC2App.__init__ = _init_v69


_old_build_instructions_v69 = ColumnsEC2App._build_instructions_tab

def _build_instructions_tab_v69(self, parent):
    # Criar a aba directamente para permitir PT/EN.
    outer = ttk.Frame(parent, padding=10)
    outer.pack(fill="both", expand=True)
    outer.rowconfigure(1, weight=1)
    outer.columnconfigure(0, weight=1)
    title = "User guide and input table" if _v69_is_en(self) else "Instruções de utilização e tabela tipo"
    ttk.Label(outer, text=title, style="Header.TLabel").grid(row=0, column=0, sticky="w", pady=(0, 8))
    text_host = ttk.Frame(outer)
    text_host.grid(row=1, column=0, sticky="nsew")
    txt = self._make_text_view(text_host)
    txt.insert("1.0", _v69_instructions_text(self))
    txt.config(state="disabled")
ColumnsEC2App._build_instructions_tab = _build_instructions_tab_v69


# Metadata and notes in selected language.
def _metadata_df_v69(self) -> pd.DataFrame:
    lang = _v69_lang(self)
    if lang == LANG_EN:
        return pd.DataFrame([
            ["Program", APP_NAME],
            ["Author / Repository", GITHUB_URL],
            ["Export date", datetime.now().strftime("%Y-%m-%d %H:%M")],
            ["Source file", self.input_file_path or "-"],
            ["Reference standard", _v59_norm_reference(self) if "_v59_norm_reference" in globals() else "Eurocode 2"],
            ["Scope", "Design and technical checking of reinforced concrete columns by column line and storey."],
            ["Description", _v69_module_description(self)],
            ["Technical limitations", _v69_module_limitations(self)],
        ], columns=["Field", "Value"])
    return pd.DataFrame([
        ["Programa", APP_NAME],
        ["Autor / Repositório", GITHUB_URL],
        ["Data de exportação", datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["Ficheiro de origem", self.input_file_path or "-"],
        ["Norma de referência", _v59_norm_reference(self) if "_v59_norm_reference" in globals() else "Eurocódigo 2"],
        ["Âmbito", "Dimensionamento e verificação técnica de pilares de betão armado por prumada e piso."],
        ["Descrição", _v69_module_description(self)],
        ["Limitações", _v69_module_limitations(self)],
    ], columns=["Campo", "Valor"])

ColumnsEC2App._metadata_df = _metadata_df_v69


def _v69_concrete_classes_from_results(app):
    vals = []
    for df in [getattr(app, "df_clean", pd.DataFrame()), getattr(app, "df_results", pd.DataFrame())]:
        try:
            if df is not None and not df.empty and "material" in df.columns:
                vals += [str(x).strip() for x in df["material"].dropna().tolist() if str(x).strip()]
        except Exception:
            pass
    vals = sorted(set(vals), key=str)
    return ", ".join(vals) if vals else "-"


def _parameters_df_v69(self) -> pd.DataFrame:
    lang = _v69_lang(self)
    concrete = _v69_concrete_classes_from_results(self)
    try:
        strategy = self.var_rebar_strategy.get()
    except Exception:
        strategy = globals().get("ACTIVE_REBAR_STRATEGY_V64", "equilibrada")
    if lang == LANG_EN:
        return pd.DataFrame([
            ["Nominal cover [mm]", self.var_cover.get()],
            ["Steel grade", f"B{int(float(self.var_fyk.get()))}" if str(self.var_fyk.get()).replace('.', '', 1).isdigit() else self.var_fyk.get()],
            ["Concrete", concrete],
            ["Design mode", self.var_calc_mode.get()],
            ["Reinforcement strategy", strategy],
            ["Reduction to governing cases", "Yes" if self.var_reduce_cases.get() else "No"],
            ["Relative humidity RH [%]", getattr(self, "var_rh", tk.StringVar(value="70")).get() if hasattr(self, "var_rh") else "70"],
            ["Concrete age at loading t0 [days]", getattr(self, "var_t0", tk.StringVar(value="28")).get() if hasattr(self, "var_t0") else "28"],
            ["h0 / hn", "calculated automatically"],
            ["φef", "calculated automatically"],
        ], columns=["Parameter", "Value"])
    return pd.DataFrame([
        ["Recobrimento [mm]", self.var_cover.get()],
        ["Classe de Aço", f"B{int(float(self.var_fyk.get()))}" if str(self.var_fyk.get()).replace('.', '', 1).isdigit() else self.var_fyk.get()],
        ["Betão", concrete],
        ["Modo de cálculo", self.var_calc_mode.get()],
        ["Estratégia de armadura", strategy],
        ["Redução para casos governantes", "Sim" if self.var_reduce_cases.get() else "Não"],
        ["Humidade relativa RH [%]", getattr(self, "var_rh", tk.StringVar(value="70")).get() if hasattr(self, "var_rh") else "70"],
        ["Idade do betão no carregamento t0 [dias]", getattr(self, "var_t0", tk.StringVar(value="28")).get() if hasattr(self, "var_t0") else "28"],
        ["h0 / hn", "calculado automaticamente"],
        ["φef", "calculado automaticamente"],
    ], columns=["Parâmetro", "Valor"])

ColumnsEC2App._parameters_df = _parameters_df_v69


def _build_normative_notes_v69(self) -> pd.DataFrame:
    lang = _v69_lang(self)
    try:
        backend = _v59_backend_key(self)
    except Exception:
        backend = "pt2010"
    if lang == LANG_EN:
        if backend == "fib2010":
            rows = [
                ("Scope", "fib Model Code 2010", "Package-based checks are used where exposed by structuralcodes. ColumnsEC2 handles column-line organisation, detailing and reporting."),
                ("Section resistance", "N-My-Mz", "The column resistance is assessed by the selected backend where available; the utilisation index is reported as η_NMyMz."),
                ("Detailing", "Constructive detailing", "Longitudinal bars, closed links and intermediate cross-ties are proposed by ColumnsEC2 using constructive rules."),
                ("Limitations", "Backend scope", "Checks not exposed by the installed structuralcodes version are reported as not assessed."),
            ]
            return pd.DataFrame(rows, columns=["Topic", "Reference", "Note"])
        rows = [
            ("Materials", "Concrete and reinforcement", "Concrete grades are read from the Material column. Reinforcement grade is selected in the GUI."),
            ("Actions", "Imported design cases", "The programme does not generate load combinations; it checks the actions imported in the table."),
            ("Second-order effects", "EC2 column design", "Effective lengths and second-order moments are considered according to the selected design engine and assumptions."),
            ("N-My-Mz interaction", "Biaxial bending with axial force", "η_NMyMz is the resistance utilisation index for the adopted reinforcement layout."),
            ("Detailing", "Constructive reinforcement", "The proposed reinforcement includes corner bars, face bars, links and intermediate cross-ties where required."),
            ("Reports", "PDF/XLSX/DXF", "The PDF gives the technical summary; Excel stores the calculation record; DXF gives the column schedule."),
        ]
        return pd.DataFrame(rows, columns=["Topic", "Reference", "Note"])
    if backend == "fib2010":
        rows = [
            ("Âmbito", "fib Model Code 2010", "São usadas verificações do structuralcodes quando expostas pelo pacote. O ColumnsEC2 gere prumadas, pormenorização e relatórios."),
            ("Resistência seccional", "N-My-Mz", "A resistência do pilar é avaliada pelo backend seleccionado quando disponível; o índice é reportado como η_NMyMz."),
            ("Pormenorização", "Detalhe construtivo", "Varões longitudinais, estribos fechados e grampos intermédios são propostos pelo ColumnsEC2 com critérios construtivos."),
            ("Limitações", "Cobertura do backend", "Verificações não expostas pela versão instalada do structuralcodes são reportadas como não avaliadas."),
        ]
        return pd.DataFrame(rows, columns=["Tema", "Referência", "Nota"])
    rows = [
        ("Materiais", "Betão e armadura", "As classes de betão são lidas da coluna Material. A classe de aço é seleccionada na GUI."),
        ("Esforços", "Casos importados", "O programa não gera combinações de acções; verifica os esforços importados na tabela."),
        ("2.ª ordem", "Dimensionamento de pilares", "Os comprimentos efectivos e momentos de 2.ª ordem são considerados conforme o motor de cálculo seleccionado e hipóteses adoptadas."),
        ("Interacção N-My-Mz", "Flexão desviada com esforço normal", "η_NMyMz é o índice de utilização resistente da solução de armadura adoptada."),
        ("Pormenorização", "Armadura construtiva", "A solução proposta inclui varões de canto, varões de face, estribos e grampos intermédios quando necessários."),
        ("Relatórios", "PDF/XLSX/DXF", "O PDF apresenta a síntese técnica; o Excel guarda a memória de cálculo; o DXF apresenta o quadro de pilares."),
    ]
    return pd.DataFrame(rows, columns=["Tema", "Referência", "Nota"])

ColumnsEC2App.build_normative_notes = _build_normative_notes_v69


# PDF bilingue.
def _v69_prepare_display_df(df, lang):
    if df is None or df.empty:
        return pd.DataFrame()
    out = df.copy()
    if lang == LANG_EN:
        out = out.rename(columns=_V69_HEADER_MAP_PT_EN)
        for c in out.columns:
            if str(c).lower().endswith("status") or "estado" in str(c).lower() or c in ["Status", "Overall status", "resistance_status", "sls_status", "detailing_status"]:
                try:
                    out[c] = out[c].map(lambda x: _v69_status_text(x, lang))
                except Exception:
                    pass
    return out


def _write_pdf_v69(self, path: str):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak
    import tempfile, shutil
    lang = _v69_lang(self)
    styles = _v68_pdf_styles() if "_v68_pdf_styles" in globals() else _pdf_styles_v3()
    level = _rc3_get_var(self, "var_pdf_level", "Relatório técnico" if lang == LANG_PT else "Technical report")
    level_en = {"Resumo executivo": "Executive summary", "Relatório técnico": "Technical report", "Memória de cálculo": "Detailed calculation note"}.get(level, level)

    res = getattr(self, "df_results", pd.DataFrame())
    res = _v68_apply_constructive_detailing(res) if "_v68_apply_constructive_detailing" in globals() and res is not None and not res.empty else res
    summ_src = getattr(self, "df_summary", pd.DataFrame())
    if summ_src is None or summ_src.empty:
        summ_src = res
    if "_v683_build_tramo_schedule" in globals():
        try:
            summ = _v683_build_tramo_schedule(summ_src)
        except Exception:
            summ = _v681_decision_by_prumada(summ_src) if "_v681_decision_by_prumada" in globals() else summ_src
    else:
        summ = _v681_decision_by_prumada(summ_src) if "_v681_decision_by_prumada" in globals() else summ_src
    inter = _v68_interaction_summary(summ_src) if "_v68_interaction_summary" in globals() else pd.DataFrame()
    module = _v65_module_status_table(res) if "_v65_module_status_table" in globals() and res is not None and not res.empty else pd.DataFrame()
    gov = _v681_governing_cases_for_pdf(self) if "_v681_governing_cases_for_pdf" in globals() else getattr(self, "df_calc_input", pd.DataFrame())
    perf = _v68_performance_df(self) if "_v68_performance_df" in globals() else pd.DataFrame()

    # Normalise/translate display frames.
    summ_d = _v69_prepare_display_df(summ, lang)
    inter_d = _v69_prepare_display_df(inter, lang)
    module_d = _v69_prepare_display_df(module, lang)
    gov_d = _v69_prepare_display_df(gov, lang)
    perf_d = _v69_prepare_display_df(perf, lang)

    title = "Columns EC2"
    subtitle = "Technical report for reinforced concrete column design/checking" if lang == LANG_EN else "Relatório técnico de dimensionamento/verificação de pilares de betão armado"
    s_decision = "1. Design decision by column line / segment" if lang == LANG_EN else "1. Decisão por prumada / tramo"
    s_inter = "2. N-My-Mz interaction" if lang == LANG_EN else "2. Interacção N-My-Mz"
    s_module = "3. Module status" if lang == LANG_EN else "3. Estados por módulo"
    s_gov = "4. Governing cases" if lang == LANG_EN else "4. Casos governantes"
    s_perf = "5. Performance" if lang == LANG_EN else "5. Performance"

    n_total = len(res) if res is not None else 0
    st_series = res.get("estado_global", res.get("status", pd.Series(dtype=str))).astype(str) if res is not None and not res.empty else pd.Series(dtype=str)
    n_fail = int((st_series == "Falha").sum()) if not st_series.empty else 0
    n_warn = int((st_series == "Aviso").sum()) if not st_series.empty else 0
    n_pr = int(summ.get("Prumada", summ.get("Column line", pd.Series(dtype=str))).astype(str).nunique()) if summ is not None and not summ.empty else 0
    if lang == LANG_EN:
        meta = pd.DataFrame([
            {"Field": "Reference standard", "Value": _v59_norm_reference(self) if "_v59_norm_reference" in globals() else "Eurocode 2"},
            {"Field": "Report level", "Value": level_en},
            {"Field": "Reinforcement strategy", "Value": _rc3_get_var(self, "var_rebar_strategy", "Balanced")},
            {"Field": "Column lines", "Value": n_pr},
            {"Field": "Analysed cases", "Value": n_total},
            {"Field": "Failures / Warnings", "Value": f"{n_fail} / {n_warn}"},
        ])
        meta_cols = ["Field", "Value"]
    else:
        meta = pd.DataFrame([
            {"Campo": "Norma/Motor", "Valor": _v59_norm_reference(self) if "_v59_norm_reference" in globals() else "Eurocódigo 2"},
            {"Campo": "Nível do relatório", "Valor": level},
            {"Campo": "Estratégia de armadura", "Valor": _rc3_get_var(self, "var_rebar_strategy", "Equilibrada")},
            {"Campo": "Prumadas", "Valor": n_pr},
            {"Campo": "Casos analisados", "Valor": n_total},
            {"Campo": "Falhas / Avisos", "Valor": f"{n_fail} / {n_warn}"},
        ])
        meta_cols = ["Campo", "Valor"]

    tmp_path = path
    try:
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        tmp_path = tmp.name
        tmp.close()
    except Exception:
        pass
    doc = SimpleDocTemplate(tmp_path, pagesize=landscape(A4), rightMargin=12*mm, leftMargin=12*mm, topMargin=12*mm, bottomMargin=12*mm)
    story = [Paragraph(title, styles["T68"]), Paragraph(subtitle, styles["B68"]), Spacer(1, 4*mm)]
    story.append(_v68_table(meta, meta_cols, styles["C68"], max_rows=10))
    story.append(Paragraph(s_decision, styles["H68"]))
    decision_cols = ["Column line", "Storey", "Section [cm]", "material", "N_Ed_kN", "My_Ed_kNm", "Mz_Ed_kNm", "η_NMyMz", "Reinforcement solution", "Status"] if lang == LANG_EN else ["Prumada", "Piso", "Secção [cm]", "material", "n_ed_kN", "my_ed_kNm", "mz_ed_kNm", "η_NMyMz", "Solução", "Estado"]
    story.append(_v68_table(summ_d, decision_cols, styles["C68"], max_rows=100))
    if level in ["Relatório técnico", "Memória de cálculo", "Technical report", "Detailed calculation note"]:
        story.append(PageBreak())
        story.append(Paragraph(s_inter, styles["H68"]))
        inter_cols = ["Column line", "Case", "N_Ed [kN]", "My_Ed [kNm]", "Mz_Ed [kNm]", "MRd_y [kNm]", "MRd_z [kNm]", "η_NMyMz", "Resistance status"] if lang == LANG_EN else ["Prumada", "Case", "N_Ed [kN]", "My_Ed [kNm]", "Mz_Ed [kNm]", "MRd_y [kNm]", "MRd_z [kNm]", "η_NMyMz", "Estado resistente"]
        story.append(_v68_table(inter_d, inter_cols, styles["C68"], max_rows=100))
        story.append(Paragraph(s_module, styles["H68"]))
        mod_cols = ["Column line", "case", "overall_status", "resistance_status", "shear_status", "torsion_status", "sls_status", "detailing_status", "technical_decision"] if lang == LANG_EN else ["Prumada", "case", "estado_global", "estado_resistente", "estado_corte", "estado_torcao", "estado_els", "estado_pormenorizacao", "decisao_tecnica"]
        story.append(_v68_table(module_d, mod_cols, styles["C68"], max_rows=100))
    if level in ["Memória de cálculo", "Detailed calculation note"]:
        story.append(PageBreak())
        story.append(Paragraph(s_gov, styles["H68"]))
        gov_cols = ["Column line", "Case", "Criterion", "NEd [kN]", "My [kNm]", "Mz [kNm]", "Vy [kN]", "Vz [kN]", "T [kNm]"] if lang == LANG_EN else ["Prumada", "Case", "Critério", "NEd [kN]", "My [kNm]", "Mz [kNm]", "Vy [kN]", "Vz [kN]", "T [kNm]"]
        story.append(_v68_table(gov_d, gov_cols, styles["C68"], max_rows=120))
        story.append(PageBreak())
        story.append(Paragraph(s_perf, styles["H68"]))
        story.append(_v68_table(perf_d, ["Item", "Value", "Note"] if lang == LANG_EN else ["Item", "Valor", "Nota"], styles["C68"], max_rows=80))
    def footer(canvas, doc_obj):
        canvas.saveState()
        canvas.setFont("Courier", 7)
        canvas.setFillColor(colors.grey)
        canvas.drawString(12*mm, 7*mm, f"Columns EC2 | {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        page_word = "Page" if lang == LANG_EN else "Página"
        canvas.drawRightString(285*mm, 7*mm, f"{page_word} {doc_obj.page}")
        canvas.restoreState()
    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    if tmp_path != path:
        try:
            shutil.move(tmp_path, path)
        except Exception:
            root, ext = os.path.splitext(path)
            alt = root + "_new" + ext
            shutil.move(tmp_path, alt)

ColumnsEC2App._write_pdf = _write_pdf_v69


# Excel wrapper: translate sheet names, headers and fixed labels when EN is active.
_old_write_excel_v69_base = ColumnsEC2App._write_excel

def _write_excel_v69(self, path: str):
    _old_write_excel_v69_base(self, path)
    if not _v69_is_en(self):
        return
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path)
        # Rename sheets first.
        for ws in list(wb.worksheets):
            if ws.title in _V69_SHEET_MAP_PT_EN:
                new_title = _V69_SHEET_MAP_PT_EN[ws.title][:31]
                if new_title != ws.title:
                    # avoid conflicts
                    base_title = new_title
                    i = 1
                    while new_title in wb.sheetnames and new_title != ws.title:
                        suffix = f"_{i}"
                        new_title = (base_title[:31-len(suffix)] + suffix)
                        i += 1
                    ws.title = new_title
        # Translate headers and common labels.
        cell_map = {**_V69_CELL_MAP_PT_EN, **_V69_HEADER_MAP_PT_EN}
        for ws in wb.worksheets:
            max_rows = min(ws.max_row, 1000)
            max_cols = min(ws.max_column, 80)
            for row in ws.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols):
                for cell in row:
                    if isinstance(cell.value, str):
                        v = cell.value.strip()
                        if v in cell_map:
                            cell.value = cell_map[v]
                        elif v in ["Aviso", "Falha", "Pré-dimensionado", "Não avaliado", "Verificar"]:
                            cell.value = _v69_status_text(v, LANG_EN)
        wb.properties.title = "ColumnsEC2"
        wb.properties.subject = "Reinforced concrete column design/checking"
        wb.properties.creator = APP_AUTHOR
        wb.properties.description = "Technical design and checking of reinforced concrete columns by column line and storey."
        wb.save(path)
    except Exception:
        pass

ColumnsEC2App._write_excel = _write_excel_v69


# DXF wrapper: export using the active writer and translate drawing labels when EN is selected.
_old_export_dxf_v69_base = ColumnsEC2App.export_dxf

def _export_dxf_v69(self):
    _old_export_dxf_v69_base(self)
    # The base function owns the file dialog and path is local. Translation is therefore handled inside a safer
    # writer patch by translating common DXF text strings globally when a path is later available in custom writer.
ColumnsEC2App.export_dxf = _old_export_dxf_v69_base

# Patch DXF writer text content by wrapping the v6.8.3 writer if present.
try:
    _old_write_columns_dxf_v683 = write_columns_dxf_v683
    def write_columns_dxf_v69(path: str, df: pd.DataFrame, lang: str = LANG_PT):
        _old_write_columns_dxf_v683(path, df)
        if lang == LANG_EN:
            try:
                p = Path(path)
                txt = p.read_text(encoding="utf-8")
                repl = {
                    "QUADRO DE PILARES - UNIDADES: mm": "COLUMN SCHEDULE - UNITS: mm",
                    "Organização por prumada e piso: colunas = P1, P2, P3...; linhas = pisos em ordem ascendente, do inferior para o superior.": "Column-line and storey layout: columns = P1, P2, P3...; rows = storeys in ascending order from lower to upper.",
                    "Legenda: betão=contorno | varões=círculos | estribos=rectângulo interior | grampos=linhas interiores | cotas em mm": "Legend: concrete=outline | bars=circles | links=inner rectangle | cross-ties=inner lines | dimensions in mm",
                    "Piso": "Storey",
                    "Secção:": "Section:",
                    "Estado:": "Status:",
                    "Grampos:": "Cross-ties:",
                    "por nível": "per level",
                    "Sem resultados": "No results",
                    "Member": "Member",
                }
                for a, b in repl.items():
                    txt = txt.replace(a, b)
                p.write_text(txt, encoding="utf-8")
            except Exception:
                pass
    def _export_dxf_v69_final(self):
        src = getattr(self, "df_results", pd.DataFrame())
        if src is None or src.empty:
            src = getattr(self, "df_summary", pd.DataFrame())
        if src is None or src.empty:
            messagebox.showwarning("Warning" if _v69_is_en(self) else "Aviso", "No results to export to DXF." if _v69_is_en(self) else "Não há resultados para exportar em DXF.")
            return
        title = "Export column schedule by column line and storey [mm]" if _v69_is_en(self) else "Exportar quadro de pilares por prumada e piso [mm]"
        path = filedialog.asksaveasfilename(title=title, defaultextension=".dxf", filetypes=[("DXF", "*.dxf")])
        if not path:
            return
        try:
            self.status_var.set("Exporting column schedule [mm]..." if _v69_is_en(self) else "A exportar quadro de pilares [mm]...")
            self.update_idletasks()
            write_columns_dxf_v69(path, src, _v69_lang(self))
            self.status_var.set(("DXF column schedule exported: " if _v69_is_en(self) else "DXF exportado: ") + str(path))
        except Exception as err:
            messagebox.showerror("Error" if _v69_is_en(self) else "Erro", ("Could not export DXF.\n\n" if _v69_is_en(self) else "Não foi possível exportar DXF.\n\n") + str(err))
    ColumnsEC2App.export_dxf = _export_dxf_v69_final
except Exception:
    pass


# Template export in selected language.
_old_export_template_v69_base = ColumnsEC2App.export_template

def _export_template_v69(self):
    if not _v69_is_en(self):
        return _old_export_template_v69_base(self)
    path = filedialog.asksaveasfilename(
        title="Save input table template",
        defaultextension=".xlsx",
        filetypes=[("Excel workbook", "*.xlsx")]
    )
    if not path:
        return
    root, ext = os.path.splitext(path)
    if ext.lower() != ".xlsx":
        path = root + ".xlsx" if ext else path + ".xlsx"
    cols = list(getattr(self, "TEMPLATE_COLUMNS", ColumnsEC2App.TEMPLATE_COLUMNS))
    sample = pd.DataFrame([
        {
            "Member/Node/Case": "72/28/101 (C)", "FX (kN)": "420.50", "FY (kN)": "3.20", "FZ (kN)": "-12.40",
            "MX (kNm)": "0.15", "MY (kNm)": "48.60", "MZ (kNm)": "-7.80", "Length (m)": "3.67",
            "Material": "C30/37", "HY (cm)": "30.0", "HZ (cm)": "30.0", "VY (cm)": "15.0", "VZ (cm)": "15.0",
            "VPY (cm)": "15.0", "VPZ (cm)": "15.0", "AX (cm2)": "900.00", "AY (cm2)": "750.00", "AZ (cm2)": "750.00",
            "IX (cm4)": "113872.30", "IY (cm4)": "67500.00", "IZ (cm4)": "67500.00", "Name": "P17", "Story": "STOREY 1"
        }
    ], columns=cols)
    notes = pd.DataFrame([
        ["Name", "Column line identifier. Elements with the same Name are grouped vertically."],
        ["Story", "Storey/segment identifier used to sort elements from lower to upper levels and align the DXF schedule."],
        ["Member/Node/Case", "Each member/case should preferably have two rows, one per end node."],
        ["Material", "Concrete grade by line/segment, for example C30/37. Multiple grades may exist in the same workbook."],
        ["Units", "Forces in kN/kNm; length in m; dimensions in cm; areas in cm²; second moments of area in cm⁴."],
    ], columns=["Field", "Description"])
    try:
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            sample.to_excel(writer, sheet_name="Input_Template", index=False)
            notes.to_excel(writer, sheet_name="Notes", index=False)
        self.status_var.set(f"Input table template saved: {path}")
    except Exception as err:
        messagebox.showerror("Error", f"Could not save the input table template.\n\n{err}")

ColumnsEC2App.export_template = _export_template_v69


APP_XLSX_DESCRIPTION = (
    "Technical design and checking of reinforced concrete columns by column line and storey, including governing-case selection, "
    "N-My-Mz interaction, second-order effects, constructive detailing, technical reports and DXF column schedules."
)


# ============================================================
# ColumnsEC2 v0.9 RC1 — backends structuralcodes com pesquisa
# adaptativa de layouts + interface bilingue PT/EN-UK mantida.
# ============================================================
