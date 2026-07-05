# -*- coding: utf-8 -*-
"""RC25 export stability and report sanitisation patch.

Scope:
- keep RC24 calculation performance;
- remove literal 'nan;' artefacts from exported frames;
- replace the PDF/DXF callbacks at the end of the runtime chain with direct,
  robust exporters that do not depend on older wrapper order;
- use df_summary / 06D-style schedule for the DXF so the full column schedule is
  exported when available.
"""
from __future__ import annotations

import math
import os
import re
import tempfile
from datetime import datetime
from pathlib import Path

import pandas as pd

APP_VERSION = "v0.9 RC25 Modular"

_RC25_REPO_URL = "https://github.com/lutondatomalela/ColumnsEC2"


def _rc25_is_nan_scalar(x) -> bool:
    try:
        if x is None:
            return True
        if isinstance(x, float) and not math.isfinite(x):
            return True
        return bool(pd.isna(x))
    except Exception:
        return False


def _rc25_clean_text(x, dash_for_empty: bool = False) -> str:
    if _rc25_is_nan_scalar(x):
        return "-" if dash_for_empty else ""
    s = str(x)
    # Clean repeated artefacts introduced by pandas string conversion.
    s = s.replace("\x00", "")
    s = re.sub(r"(?i)^\s*nan\s*;\s*", "", s)
    s = re.sub(r"(?i)\s*;\s*nan\s*(;|$)", lambda m: ";" if m.group(1) == ";" else "", s)
    s = re.sub(r"(?i)^\s*(nan|none|null)\s*$", "", s)
    s = re.sub(r"\s*;\s*;\s*", "; ", s)
    s = re.sub(r"\s+", " ", s).strip()
    if not s or s == ";" or s == "-;":
        return "-" if dash_for_empty else ""
    return s


def _rc25_sanitize_df(df: pd.DataFrame | None) -> pd.DataFrame:
    if df is None:
        return pd.DataFrame()
    if not isinstance(df, pd.DataFrame):
        try:
            df = pd.DataFrame(df)
        except Exception:
            return pd.DataFrame()
    out = df.copy()
    for c in out.columns:
        if out[c].dtype == object:
            out[c] = out[c].map(lambda v: _rc25_clean_text(v, dash_for_empty=False))
    for c in ["warning_reason", "failure_warnings", "failure_info", "failure_action", "recommendations"]:
        if c in out.columns:
            out[c] = out[c].map(lambda v: _rc25_clean_text(v, dash_for_empty=True))
    # Avoid presenting a failed/empty design as an adoptable solution.
    for c in ["solucao", "solucao_completa", "Solução", "Solução adoptada", "Solução local"]:
        if c in out.columns:
            out[c] = out[c].map(lambda v: "" if str(v).strip().lower() in {"nan", "none", "null"} else _rc25_clean_text(v))
    return out


def _rc25_getvar(self, name: str, default=""):
    try:
        v = getattr(self, name, None)
        return v.get() if hasattr(v, "get") else (v if v is not None else default)
    except Exception:
        return default


def _rc25_safe_filename(path: str, suffix: str) -> str:
    root, ext = os.path.splitext(str(path))
    if ext.lower() != suffix.lower():
        return root + suffix if ext else str(path) + suffix
    return str(path)


def _rc25_schedule_from_app(self) -> pd.DataFrame:
    """Return the best available full schedule: summary first, then rebuilt results."""
    summary = getattr(self, "df_summary", pd.DataFrame())
    if summary is not None and not getattr(summary, "empty", True):
        return _rc25_sanitize_df(summary)
    res = getattr(self, "df_results", pd.DataFrame())
    if res is None or getattr(res, "empty", True):
        return pd.DataFrame()
    try:
        if "_rc22_build_tramo_schedule" in globals():
            return _rc25_sanitize_df(_rc22_build_tramo_schedule(res))
    except Exception:
        pass
    try:
        if hasattr(self, "build_summary_by_member"):
            return _rc25_sanitize_df(self.build_summary_by_member(res))
    except Exception:
        pass
    return _rc25_sanitize_df(res)


def _rc25_prepare_frames(self):
    for attr in ["df_results", "df_summary", "df_failures", "df_warnings", "df_validation", "df_notes", "df_calc_input", "df_pair"]:
        try:
            df = getattr(self, attr, None)
            if isinstance(df, pd.DataFrame) and not df.empty:
                setattr(self, attr, _rc25_sanitize_df(df))
        except Exception:
            pass
    try:
        self.df_summary = _rc25_schedule_from_app(self)
    except Exception:
        pass


def _rc25_write_excel_direct(self, path: str):
    _rc25_prepare_frames(self)
    res = _rc25_sanitize_df(getattr(self, "df_results", pd.DataFrame()))
    els_cols = [c for c in res.columns if str(c).startswith("service") or c in ["member", "case", "name", "story", "status"]]
    frames = {
        "00_Resumo": _rc25_sanitize_df(getattr(self, "df_summary", pd.DataFrame())),
        "01_Resultados": res,
        "02_ELS": res[els_cols].copy() if els_cols else pd.DataFrame(),
        "03_Calculo": _rc25_sanitize_df(getattr(self, "df_calc_input", pd.DataFrame())),
        "04_Pares": _rc25_sanitize_df(getattr(self, "df_pair", pd.DataFrame())),
        "05_Validacao": _rc25_sanitize_df(getattr(self, "df_validation", pd.DataFrame())),
        "06_Notas": _rc25_sanitize_df(getattr(self, "df_notes", pd.DataFrame())),
    }
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet, df in frames.items():
            if df is None or getattr(df, "empty", True):
                df = pd.DataFrame([{"nota": "Sem dados"}])
            df = _rc25_sanitize_df(df).replace({pd.NA: "", None: ""})
            df.to_excel(writer, sheet_name=sheet[:31], index=False)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path)
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            if ws.max_row and ws.max_column:
                ws.auto_filter.ref = ws.dimensions
            for col in ws.columns:
                try:
                    width = min(42, max(10, max(len(str(cell.value or "")) for cell in col[:80]) + 2))
                    ws.column_dimensions[col[0].column_letter].width = width
                except Exception:
                    pass
        wb.save(path)
    except Exception:
        pass


# Sanitise immediately after the calculation finishes.
_rc25_prev_run_design = getattr(ColumnsEC2App, "run_design", None)

def _run_design_rc25(self):
    out = _rc25_prev_run_design(self) if callable(_rc25_prev_run_design) else None
    try:
        _rc25_prepare_frames(self)
    except Exception:
        pass
    return out

try:
    ColumnsEC2App.run_design = _run_design_rc25
except Exception:
    pass


# Sanitise before Excel export. Keep the previous writer, but clean the data frames.
_rc25_prev_write_excel = getattr(ColumnsEC2App, "_write_excel", None)

def _write_excel_rc25(self, path: str):
    _rc25_prepare_frames(self)
    if callable(_rc25_prev_write_excel):
        try:
            return _rc25_prev_write_excel(self, path)
        except RuntimeError as err:
            if "default root window" not in str(err) and "Too early to create variable" not in str(err):
                raise
            return _rc25_write_excel_direct(self, path)
    raise RuntimeError("Excel writer não disponível.")

try:
    ColumnsEC2App._write_excel = _write_excel_rc25
except Exception:
    pass


def _rc25_pdf_safe(x) -> str:
    s = _rc25_clean_text(x, dash_for_empty=False)
    # Built-in PDF fonts may render Ø as a plain zero in some viewers.
    # Use explicit engineering text in the PDF; XLSX/DXF keep the original symbol.
    s = s.replace("Ø", "phi")
    return s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _rc25_pdf_table(df: pd.DataFrame, cols, style, table_style, max_rows=35, widths=None):
    from reportlab.platypus import Paragraph, Table
    from reportlab.lib.units import mm
    if df is None or df.empty:
        df = pd.DataFrame(columns=cols)
    present = [c for c in cols if c in df.columns]
    if not present:
        present = ["nota"]
        df = pd.DataFrame([{"nota": "Sem dados."}])
    data = [[Paragraph(_rc25_pdf_safe(c), style) for c in present]]
    for _, r in df.head(max_rows).iterrows():
        data.append([Paragraph(_rc25_pdf_safe(r.get(c, "")), style) for c in present])
    if widths is None:
        widths = [270 * mm / max(1, len(present))] * len(present)
    t = Table(data, colWidths=widths, repeatRows=1)
    t.setStyle(table_style)
    return t


def _rc25_pdf_escape(text: str) -> str:
    return str(text).replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def _write_pdf_minimal_rc25(self, path: str):
    _rc25_prepare_frames(self)
    summary = _rc25_schedule_from_app(self)
    res = _rc25_sanitize_df(getattr(self, "df_results", pd.DataFrame()))
    lines = [
        f"ColumnsEC2 {APP_VERSION}",
        "Relatorio tecnico sintetico",
        f"Tramos: {len(summary) if summary is not None else 0}",
        f"Prumadas: {int(summary.get('name', summary.get('Prumada', pd.Series(dtype=str))).astype(str).nunique()) if summary is not None and not summary.empty else 0}",
        "",
        "Prumada | Piso | member | case | Estado | Solucao",
    ]
    src = summary if summary is not None and not summary.empty else res
    for _, r in src.head(65).iterrows():
        vals = [
            r.get("name", r.get("Prumada", "")),
            r.get("story", r.get("Piso", "")),
            r.get("member", ""),
            r.get("case", ""),
            r.get("estado_global", r.get("Estado", r.get("status", ""))),
            r.get("solucao", r.get("Solução", "")),
        ]
        lines.append(" | ".join(_rc25_clean_text(v, dash_for_empty=True) for v in vals)[:150])
    text_ops = ["BT", "/F1 9 Tf", "40 560 Td"]
    for i, line in enumerate(lines[:72]):
        if i:
            text_ops.append("0 -12 Td")
        text_ops.append(f"({_rc25_pdf_escape(line)}) Tj")
    text_ops.append("ET")
    stream = "\n".join(text_ops).encode("latin-1", "replace")
    objects = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
        b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 842 595] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >>",
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
        b"<< /Length " + str(len(stream)).encode("ascii") + b" >>\nstream\n" + stream + b"\nendstream",
    ]
    data = bytearray(b"%PDF-1.4\n")
    offsets = [0]
    for idx, obj in enumerate(objects, start=1):
        offsets.append(len(data))
        data.extend(f"{idx} 0 obj\n".encode("ascii"))
        data.extend(obj)
        data.extend(b"\nendobj\n")
    xref = len(data)
    data.extend(f"xref\n0 {len(objects)+1}\n".encode("ascii"))
    data.extend(b"0000000000 65535 f \n")
    for off in offsets[1:]:
        data.extend(f"{off:010d} 00000 n \n".encode("ascii"))
    data.extend(f"trailer << /Size {len(objects)+1} /Root 1 0 R >>\nstartxref\n{xref}\n%%EOF\n".encode("ascii"))
    Path(path).write_bytes(bytes(data))


def _write_pdf_rc25(self, path: str):
    """Small but robust PDF report independent of previous wrapper order."""
    _rc25_prepare_frames(self)
    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    except ModuleNotFoundError:
        return _write_pdf_minimal_rc25(self, path)

    res = _rc25_sanitize_df(getattr(self, "df_results", pd.DataFrame()))
    summary = _rc25_schedule_from_app(self)
    def _status_series(df):
        if df is None or df.empty:
            return pd.Series([], dtype=str)
        if "estado_global" in df.columns:
            return df["estado_global"].astype(str)
        if "status" in df.columns:
            return df["status"].astype(str)
        if "Estado" in df.columns:
            return df["Estado"].astype(str)
        return pd.Series([""] * len(df), index=df.index, dtype=str)

    failures = summary[_status_series(summary).str.contains("Falha", case=False, na=False)].copy() if not summary.empty else pd.DataFrame()
    if failures.empty and not res.empty:
        failures = res[_status_series(res).str.contains("Falha", case=False, na=False)].copy()
    warnings = summary[_status_series(summary).str.contains("Aviso", case=False, na=False)].copy() if not summary.empty else pd.DataFrame()

    doc = SimpleDocTemplate(str(path), pagesize=landscape(A4), rightMargin=10*mm, leftMargin=10*mm, topMargin=10*mm, bottomMargin=10*mm)
    try:
        doc.title = f"{APP_NAME} {APP_VERSION}"
        doc.author = _RC25_REPO_URL
        doc.subject = "Dimensionamento e verificação de pilares de betão armado"
        doc.creator = APP_NAME
    except Exception:
        pass

    styles = getSampleStyleSheet()
    if "RC25_Title" not in styles.byName:
        styles.add(ParagraphStyle(name="RC25_Title", parent=styles["Title"], alignment=TA_CENTER, fontName="Courier-Bold", fontSize=14, leading=18, spaceAfter=8))
        styles.add(ParagraphStyle(name="RC25_Subtitle", parent=styles["Normal"], alignment=TA_CENTER, fontName="Courier", fontSize=9, leading=12, textColor=colors.darkgrey, spaceAfter=8))
        styles.add(ParagraphStyle(name="RC25_Section", parent=styles["Heading2"], fontName="Courier-Bold", fontSize=11, leading=14, spaceBefore=8, spaceAfter=6))
        styles.add(ParagraphStyle(name="RC25_Cell", parent=styles["Normal"], fontName="Courier", fontSize=6.4, leading=8.2, alignment=TA_LEFT))
        styles.add(ParagraphStyle(name="RC25_Small", parent=styles["Normal"], fontName="Courier", fontSize=7.2, leading=9.5))
    table_style = TableStyle([
        ("GRID", (0,0), (-1,-1), 0.25, colors.HexColor("#D8DEE9")),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1F4E5F")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTNAME", (0,0), (-1,0), "Courier-Bold"),
        ("FONTSIZE", (0,0), (-1,-1), 6.4),
        ("LEFTPADDING", (0,0), (-1,-1), 2.0),
        ("RIGHTPADDING", (0,0), (-1,-1), 2.0),
        ("TOPPADDING", (0,0), (-1,-1), 2.0),
        ("BOTTOMPADDING", (0,0), (-1,-1), 2.0),
    ])
    meta_style = TableStyle([
        ("GRID", (0,0), (-1,-1), 0.25, colors.HexColor("#D8DEE9")),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("FONTNAME", (0,0), (-1,-1), "Courier"),
        ("FONTSIZE", (0,0), (-1,-1), 7.0),
        ("BACKGROUND", (0,0), (0,-1), colors.HexColor("#F2F5F7")),
        ("BACKGROUND", (2,0), (2,-1), colors.HexColor("#F2F5F7")),
    ])

    def count_status(df, word):
        if df is None or df.empty:
            return 0
        s = _status_series(df)
        return int(s.str.contains(word, case=False, na=False).sum())

    n_tramos = len(summary)
    n_prumadas = int(summary.get("name", summary.get("Prumada", pd.Series(dtype=str))).astype(str).nunique()) if not summary.empty else 0
    n_ok = count_status(summary, "OK")
    n_warn = count_status(summary, "Aviso")
    n_fail = count_status(summary, "Falha")
    perf = getattr(self, "_calculation_perf", {})

    story = []
    story.append(Paragraph(f'<link href="{_RC25_REPO_URL}">ColumnsEC2 {APP_VERSION}</link>', styles["RC25_Title"]))
    story.append(Paragraph("Relatório técnico sintético - NP EN 1992-1-1:2010 + AC:2012 + A1:2019", styles["RC25_Subtitle"]))
    meta = [
        ["Data", datetime.now().strftime("%Y-%m-%d %H:%M"), "Repositório", _RC25_REPO_URL],
        ["Tramos", str(n_tramos), "Prumadas", str(n_prumadas)],
        ["OK / Aviso / Falha", f"{n_ok} / {n_warn} / {n_fail}", "ELS", str(_rc25_getvar(self, "var_service_case", "") or "nao indicada")],
        ["Estratégia", str(_rc25_getvar(self, "var_rebar_strategy", "Equilibrada")), "Nível", str(_rc25_getvar(self, "var_pdf_level", "Relatório técnico"))],
    ]
    mt = Table(meta, colWidths=[36*mm, 74*mm, 36*mm, 124*mm]); mt.setStyle(meta_style); story.append(mt); story.append(Spacer(1,4*mm))

    story.append(Paragraph("1. Quadro de decisão por prumada e tramo", styles["RC25_Section"]))
    cols = ["name", "story", "member", "case", "material", "b_cm", "h_cm", "η_NMyMz", "solucao", "estado_global"]
    # Prefer PT schedule columns if present.
    if "Prumada" in summary.columns:
        cols = ["Prumada", "Piso", "member", "case", "material", "Secção [cm]", "η_NMyMz", "Solução", "Estado"]
    story.append(_rc25_pdf_table(summary, cols, styles["RC25_Cell"], table_style, max_rows=70))

    if not failures.empty:
        story.append(PageBreak())
        story.append(Paragraph("2. Falhas a rever", styles["RC25_Section"]))
        fcols = ["name", "story", "member", "case", "b_cm", "h_cm", "η_NMyMz", "failure_reason", "failure_action"]
        if "Prumada" in failures.columns:
            fcols = ["Prumada", "Piso", "member", "case", "Secção [cm]", "η_NMyMz", "Estado", "decisao_tecnica"]
        story.append(_rc25_pdf_table(failures, fcols, styles["RC25_Cell"], table_style, max_rows=50))

    if not warnings.empty:
        story.append(Paragraph("3. Avisos principais", styles["RC25_Section"]))
        wcols = ["name", "story", "member", "case", "estado_corte", "estado_torcao", "estado_els", "warning_reason"]
        if "Prumada" in warnings.columns:
            wcols = ["Prumada", "Piso", "member", "case", "estado_corte", "estado_torcao", "estado_els", "decisao_tecnica"]
        story.append(_rc25_pdf_table(warnings, wcols, styles["RC25_Cell"], table_style, max_rows=50))

    # Lightweight performance/notes block.
    story.append(Spacer(1, 4*mm))
    story.append(Paragraph("Nota: a memória completa, incluindo dados importados, casos governantes, ELS, esforço transverso, torção e superfície N-My-Mz, permanece no ficheiro XLSX.", styles["RC25_Small"]))

    def footer(canvas, doc_obj):
        canvas.saveState()
        canvas.setFont("Courier", 7)
        canvas.setFillColor(colors.grey)
        canvas.drawString(10*mm, 7*mm, f"ColumnsEC2 {APP_VERSION} | {_RC25_REPO_URL}")
        try:
            canvas.linkURL(_RC25_REPO_URL, (10*mm, 6*mm, 95*mm, 10*mm), relative=0, thickness=0)
        except Exception:
            pass
        canvas.drawRightString(287*mm, 7*mm, f"Página {doc_obj.page}")
        canvas.restoreState()

    doc.build(story, onFirstPage=footer, onLaterPages=footer)


try:
    ColumnsEC2App._write_pdf = _write_pdf_rc25
except Exception:
    pass


def _export_pdf_report_rc25(self):
    src = getattr(self, "df_results", pd.DataFrame())
    summ = getattr(self, "df_summary", pd.DataFrame())
    if (src is None or getattr(src, "empty", True)) and (summ is None or getattr(summ, "empty", True)):
        messagebox.showwarning("Aviso", "Não há resultados para exportar em PDF.")
        return
    path = filedialog.asksaveasfilename(title="Exportar relatório PDF", defaultextension=".pdf", filetypes=[("PDF", "*.pdf")])
    if not path:
        return
    path = _rc25_safe_filename(path, ".pdf")
    out_dir = os.path.dirname(os.path.abspath(path)) or os.getcwd()
    os.makedirs(out_dir, exist_ok=True)
    fd, tmp = tempfile.mkstemp(prefix=".__columns_ec2_rc25_", suffix=".pdf", dir=out_dir)
    os.close(fd)
    try:
        try:
            self.progress_var.set(5.0)
        except Exception:
            pass
        try:
            self.status_var.set("A preparar PDF...")
            self.update_idletasks()
        except Exception:
            pass
        self._write_pdf(tmp)
        final_path = path
        try:
            os.replace(tmp, path)
        except Exception:
            root, ext = os.path.splitext(path)
            final_path = root + "_novo" + ext
            os.replace(tmp, final_path)
        try:
            self.progress_var.set(100.0)
            self.status_var.set(f"PDF exportado: {final_path}")
        except Exception:
            pass
        messagebox.showinfo("Exportação concluída", f"PDF exportado:\n{final_path}")
    except Exception as err:
        try:
            if os.path.exists(tmp):
                os.remove(tmp)
        except Exception:
            pass
        try:
            self.progress_var.set(0.0)
        except Exception:
            pass
        messagebox.showerror("Erro", f"Não foi possível exportar PDF.\n\n{err}")

try:
    ColumnsEC2App.export_pdf_report = _export_pdf_report_rc25
except Exception:
    pass


_RC25_PREV_DXF_WRITERS = []
for _n in ["write_columns_dxf_v22", "write_columns_dxf_v16", "write_columns_dxf_v683", "write_columns_dxf_v4"]:
    _f = globals().get(_n)
    if callable(_f):
        _RC25_PREV_DXF_WRITERS.append((_n, _f))


def _write_columns_dxf_rc25(path: str, df: pd.DataFrame, lang=None):
    df = _rc25_sanitize_df(df)
    if df is None or df.empty:
        Path(path).write_text("0\nSECTION\n2\nENTITIES\n0\nTEXT\n8\nTEXT\n10\n0\n20\n0\n40\n50\n1\nSem resultados\n0\nENDSEC\n0\nEOF\n", encoding="utf-8")
        return
    errors = []
    # Prefer the RC22 full schedule writer captured before RC25 aliases are installed.
    for fn_name, fn in _RC25_PREV_DXF_WRITERS:
        if callable(fn) and fn is not _write_columns_dxf_rc25:
            try:
                if "lang" in getattr(fn, "__code__", object()).co_varnames:
                    return fn(path, df, lang=lang or globals().get("LANG_PT", "pt"))
                return fn(path, df)
            except Exception as err:
                errors.append(f"{fn_name}: {err}")
    # Minimal fallback: textual schedule only, but valid DXF.
    parts = ["0\nSECTION\n2\nHEADER\n9\n$INSUNITS\n70\n4\n0\nENDSEC\n0\nSECTION\n2\nENTITIES\n"]
    parts += ["0\nTEXT\n8\nTEXT\n10\n0\n20\n0\n40\n60\n1\nQUADRO DE PILARES - FALLBACK RC25\n"]
    y = -100.0
    cols = [c for c in ["name", "story", "member", "case", "material", "b_cm", "h_cm", "solucao", "estado_global"] if c in df.columns]
    if not cols:
        cols = list(df.columns[:8])
    for _, r in df.head(500).iterrows():
        txt = " | ".join(_rc25_clean_text(r.get(c, "")) for c in cols)[:240]
        txt = txt.replace("\n", " ")
        parts += [f"0\nTEXT\n8\nTEXT\n10\n0\n20\n{y:.3f}\n40\n22\n1\n{txt}\n"]
        y -= 45.0
    parts += ["0\nENDSEC\n0\nEOF\n"]
    Path(path).write_text("".join(parts), encoding="utf-8")

# Keep named writer aliases for downstream code.
write_columns_dxf_rc25 = _write_columns_dxf_rc25
write_columns_dxf_v22 = _write_columns_dxf_rc25
write_columns_dxf_v16 = _write_columns_dxf_rc25
write_columns_dxf_v4 = _write_columns_dxf_rc25


def _export_dxf_rc25(self):
    src = _rc25_schedule_from_app(self)
    if src is None or src.empty:
        messagebox.showwarning("Aviso", "Não há resultados para exportar em DXF.")
        return
    path = filedialog.asksaveasfilename(title="Exportar quadro completo de pilares [DXF]", defaultextension=".dxf", filetypes=[("DXF", "*.dxf")])
    if not path:
        return
    path = _rc25_safe_filename(path, ".dxf")
    try:
        try:
            self.status_var.set("A exportar quadro completo de pilares [DXF]...")
            self.update_idletasks()
        except Exception:
            pass
        _write_columns_dxf_rc25(path, src, lang=globals().get("LANG_PT", "pt"))
        try:
            self.status_var.set(f"DXF exportado: {path}")
        except Exception:
            pass
        messagebox.showinfo("Exportação concluída", f"DXF exportado:\n{path}")
    except Exception as err:
        messagebox.showerror("Erro", f"Não foi possível exportar o quadro de pilares em DXF.\n\n{err}")

try:
    ColumnsEC2App.export_dxf = _export_dxf_rc25
except Exception:
    pass

# Final visible version marker.
try:
    globals()["APP_VERSION"] = "v0.9 RC25 Modular"
except Exception:
    pass
